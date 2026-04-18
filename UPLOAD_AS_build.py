#!/usr/bin/env python3
"""
Pinellas Ice Co \u2014 Prospect Tool Builder
Complete, self-contained. Drop in any folder with your CSV files and run.
"""

import sys, os, json, re, warnings, csv
from pathlib import Path
from datetime import date, timedelta
from math import radians, cos, sin, sqrt, atan2
from collections import Counter

warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
TARGET_COUNTIES  = ['Pinellas', 'Hillsborough', 'Pasco',
                    'Citrus', 'Hernando', 'Polk', 'Sumter']
MIN_SCORE        = 5   # Very low - include everything, filter in browser
TODAY            = date.today()
OUTPUT_FILE      = Path(__file__).parent / 'prospecting_tool.html'

# ── CHAIN CLASSIFICATION ──────────────────────────────────────────────────────
# Corporate chains: national vendor contracts, skip entirely
CORPORATE_CHAINS = [
    'mcdonald','burger king','wendy','chick-fil-a','starbucks','chipotle',
    'ihop','waffle house','olive garden','red lobster','outback steakhouse',
    'buffalo wild wings','panera bread','cracker barrel','texas roadhouse',
    'golden corral','panda express','jack in the box','raising cane',
    'shake shack','wingstop','habit burger',
]

# Franchise brands: independently owned, CAN be prospects
FRANCHISE_BRANDS = [
    'subway','taco bell','dunkin','pizza hut','domino','papa john',
    'little caesar','sonic','dairy queen','five guys','popeyes','kfc',
    'chilis','applebee','denny','hooters','longhorn','buffalo wild wing',
    'jersey mike','jimmy john','firehouse subs','mcalister',
    'tijuana flats','miller ale house',
]

def classify_business(name):
    """Returns ('corporate','franchise','independent') + is_chain bool."""
    n = name.lower()
    for c in CORPORATE_CHAINS:
        if c in n: return 'corporate', True
    for f in FRANCHISE_BRANDS:
        if f in n: return 'franchise', True
    return 'independent', False

# ── MACHINE ESTIMATION ────────────────────────────────────────────────────────
def est_machines(seats, is_full_bar, rank_code):
    """Estimate ice machine count from seat count + license type."""
    if rank_code in ('NOST','CNOSEAT','MFDV'): return 1
    if seats <= 0:    base = 1
    elif seats < 40:  base = 1
    elif seats < 80:  base = 1
    elif seats < 150: base = 2
    elif seats < 250: base = 2
    elif seats < 400: base = 3
    else:             base = 4
    bonus = 1 if is_full_bar else 0
    return min(base + bonus, 6)

def est_monthly(machines):
    """Tiered: $149 first, $89 second (+$238), $69 each additional."""
    if machines <= 1: return 149
    if machines == 2: return 238
    return 238 + (machines - 2) * 69

def est_onetime(machines):
    """One-time deep clean price."""
    if machines <= 1: return 249
    if machines == 2: return 378
    return 378 + (machines - 2) * 99

def account_tier(seats, rank_code, machines, chronic, confirmed):
    """Account quality tier."""
    if rank_code in ('NOST','CNOSEAT') and not confirmed and not chronic: return 'COLD'
    if seats > 0 and seats < 15 and not confirmed and not chronic: return 'COLD'
    if machines >= 3 and (chronic or confirmed): return 'PLATINUM'
    if machines >= 2 and (chronic or confirmed): return 'GOLD'
    if machines >= 2:                            return 'SILVER'
    if confirmed and chronic:                    return 'GOLD'
    if confirmed:                                return 'SILVER'
    if chronic:                                  return 'SILVER'
    if seats >= 60:                              return 'BRONZE'
    return 'COLD'

def confidence_score(n_inspections, n_years_data, has_ice_history, days_since):
    """
    Confidence in the scoring/prediction (0-100).
    Higher with more inspection history, recent data, ice violation history.
    """
    base = 30
    # More inspections = more confident
    if n_inspections >= 6:   base += 25
    elif n_inspections >= 3: base += 15
    elif n_inspections >= 2: base += 8
    # More years of data
    if n_years_data >= 4:    base += 20
    elif n_years_data >= 2:  base += 10
    # Ice history is a strong signal
    if has_ice_history:      base += 15
    # Recent inspection = more reliable
    if days_since <= 30:     base += 10
    elif days_since <= 90:   base += 5
    return min(100, base)

def seat_score_bonus(machines, rank_code):
    if rank_code in ('NOST','CNOSEAT'): return 0
    if machines >= 3: return 15
    if machines >= 2: return 10
    return 0


# ──────────────────────────────────────────────────────────────────────────────
# STEP 0: CHECK DEPENDENCIES
# ──────────────────────────────────────────────────────────────────────────────
def check_deps():
    missing = []
    for pkg, install in [('pandas','pandas'), ('sklearn','scikit-learn'), 
                          ('numpy','numpy'), ('openpyxl','openpyxl')]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(install)
    if missing:
        print("\n" + "="*55)
        print("  Missing packages. Run this once in terminal:")
        print(f"  pip install {' '.join(missing)}")
        print("="*55 + "\n")
        sys.exit(1)

check_deps()

import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN NORMALIZATION
# ──────────────────────────────────────────────────────────────────────────────
COLUMN_MAP = {
    'Inspection Date': ['InspectionDate','INSPECTION DATE','inspection_date','Date','Insp Date'],
    'License ID': ['LicenseID','LICENSE ID','license_id','License Number','LicenseNumber','Lic ID','License_ID'],
    'Business (DBA-Does Business As) Name': [
        'Business Name','DBA Name','DBA','Business','BUSINESS NAME',
        'DBA-Does Business As','Restaurant Name','RestaurantName','Establishment Name'],
    'County Name': ['County','COUNTY','county_name','CountyName'],
    'Location Address': ['Address','LOCATION ADDRESS','Street Address','StreetAddress','Addr'],
    'Location City': ['City','CITY','city_name'],
    'Location Zip Code': ['Zip','ZIP','Zip Code','ZipCode','Postal Code','PostalCode','zip_code'],
    'Inspection Type': ['InspectionType','INSPECTION TYPE','Type','Insp Type'],
    'Inspection Disposition': ['Disposition','DISPOSITION','InspectionDisposition','Result'],
    'Number of High Priority Violations': [
        'High Priority Violations','HighPriorityViolations','High Violations','HP Violations','HV'],
    'Number of Total Violations': ['Total Violations','TotalViolations','Total','TV'],
    'Number of Intermediate Violations': ['Intermediate Violations','IntermediateViolations','IV'],
    'Number of Basic Violations': ['Basic Violations','BasicViolations','BV'],
    'Visit Number': ['VisitNumber','Visit','VISIT NUMBER','Visit_Number'],
}

DISP_RISK = {
    'Inspection Completed - No Further Action': 0, 'Warning Issued': 1,
    'Call Back - Complied': 2, 'Call Back - Extension given, pending': 3,
    'Admin. Complaint Callback Complied': 3, 'Call Back - Admin. complaint recommended': 4,
    'Administrative complaint recommended': 4, 'Administrative determination recommended': 4,
    'Emergency Order Callback Complied': 3, 'Emergency Order Callback Time Extension': 4,
    'Emergency Order Callback Not Complied': 5, 'Emergency order recommended': 5,
}

ICE_CODES = {14: 100, 22: 80, 50: 70, 23: 60, 37: 50, 51: 30, 36: 20, 55: 15}

ZIP_COORDS = {
    '33511':(27.924,-82.312),'33547':(27.861,-82.212),'33570':(27.714,-82.398),
    '33572':(27.752,-82.378),'33573':(27.710,-82.354),'33578':(27.855,-82.324),
    '33579':(27.793,-82.278),'33584':(27.992,-82.269),'33592':(28.096,-82.280),
    '33594':(27.910,-82.232),'33596':(27.875,-82.258),'33598':(27.660,-82.349),
    '33602':(27.949,-82.459),'33603':(27.980,-82.468),'33604':(28.001,-82.468),
    '33605':(27.956,-82.427),'33606':(27.931,-82.474),'33607':(27.966,-82.489),
    '33609':(27.945,-82.495),'33610':(27.978,-82.406),'33611':(27.900,-82.508),
    '33612':(28.022,-82.457),'33613':(28.055,-82.448),'33614':(27.997,-82.509),
    '33615':(27.989,-82.563),'33616':(27.894,-82.532),'33617':(28.012,-82.407),
    '33618':(28.053,-82.499),'33619':(27.950,-82.389),'33620':(28.062,-82.415),
    '33621':(27.860,-82.520),'33624':(28.069,-82.535),'33625':(28.080,-82.556),
    '33626':(28.082,-82.585),'33629':(27.924,-82.522),'33634':(27.993,-82.567),
    '33635':(28.010,-82.609),'33637':(28.056,-82.380),'33647':(28.119,-82.378),
    '33701':(27.770,-82.634),'33702':(27.811,-82.647),'33703':(27.817,-82.633),
    '33704':(27.798,-82.640),'33705':(27.750,-82.638),'33706':(27.747,-82.747),
    '33707':(27.761,-82.717),'33708':(27.810,-82.800),'33709':(27.806,-82.754),
    '33710':(27.785,-82.734),'33711':(27.745,-82.682),'33712':(27.726,-82.663),
    '33713':(27.779,-82.669),'33714':(27.803,-82.685),'33715':(27.703,-82.717),
    '33716':(27.855,-82.652),'33755':(27.966,-82.800),'33756':(27.953,-82.801),
    '33759':(27.974,-82.750),'33760':(27.900,-82.687),'33761':(28.014,-82.740),
    '33762':(27.870,-82.676),'33763':(28.005,-82.748),'33764':(27.934,-82.740),
    '33765':(27.984,-82.777),'33767':(27.978,-82.828),'33770':(27.910,-82.789),
    '33771':(27.912,-82.762),'33772':(27.878,-82.795),'33773':(27.880,-82.720),
    '33774':(27.869,-82.833),'33776':(27.839,-82.815),'33777':(27.854,-82.754),
    '33778':(27.873,-82.740),'33781':(27.841,-82.717),'33782':(27.863,-82.697),
    '33785':(27.740,-82.781),'33786':(27.924,-82.826),'34677':(28.053,-82.669),
    '34681':(28.011,-82.688),'34683':(28.080,-82.765),'34684':(28.102,-82.769),
    '34685':(28.117,-82.731),'34688':(28.128,-82.704),'34689':(28.146,-82.756),
    '34695':(27.986,-82.695),'34698':(28.019,-82.775),'34652':(28.240,-82.727),
    '34653':(28.262,-82.724),'34654':(28.279,-82.666),'34655':(28.218,-82.673),
    '34667':(28.338,-82.652),'34668':(28.280,-82.697),'34669':(28.226,-82.706),
    '34690':(28.243,-82.675),'34691':(28.195,-82.754),'34638':(28.190,-82.712),
    '34637':(28.214,-82.666),'34639':(28.234,-82.628),
    '34429':(28.896,-82.580),'34442':(28.767,-82.472),'34448':(28.956,-82.582),
    '34450':(28.836,-82.330),'34452':(28.814,-82.335),'34465':(28.947,-82.469),
    '34601':(28.552,-82.388),'34604':(28.538,-82.466),'34606':(28.468,-82.549),
    '34608':(28.506,-82.540),'34609':(28.441,-82.480),'34610':(28.369,-82.547),
    '34613':(28.527,-82.517),'33801':(28.041,-81.975),'33803':(28.020,-81.970),
    '33810':(28.059,-82.036),'33811':(27.945,-82.028),'33813':(27.969,-82.001),
    '33823':(28.071,-81.897),'33830':(27.899,-81.840),'33880':(28.023,-81.731),
    '33897':(28.291,-81.643),'32162':(28.950,-81.966),'34785':(28.839,-82.051),
}

HIGH_ICE = ['bar','pub','tavern','lounge','grill','grille','sports','club',
            'seafood','sushi','hibachi','oyster','marina','brewery','taproom',
            'resort','hotel','inn','cantina','steakhouse']
LOW_ICE  = ['bakery','donut','coffee','smoothie','juice','dessert','cupcake']

# Median days until callback inspection per disposition type.
# Derived from 25k+ FL DBPR inspection interval records — data-driven, not guesses.
CALLBACK_DAYS = {
    'Emergency order recommended':               1,
    'Emergency Order Callback Not Complied':     1,
    'Administrative complaint recommended':      4,
    'Administrative determination recommended':  4,
    'Call Back - Admin. complaint recommended':  22,
    'Admin. Complaint Callback Complied':        45,
    'Warning Issued':                            12,
    'Call Back - Extension given, pending':      50,
    'Emergency Order Callback Time Extension':   62,
    'Emergency Order Callback Complied':         64,
    'Call Back - Complied':                      118,
    'Inspection Completed - No Further Action':  130,
}
STATIC_PHONES = {
    "9784166": "+1 813-993-3924",
    "2183411": "+1 813-752-2236",
    "9681679": "+1 813-657-5600",
    "9468965": "+1 813-374-5363",
    "9473519": "+1 727-314-4004",
    "9405890": "+1 813-909-6354",
    "9428645": "+1 813-265-2111",
    "9435298": "+1 813-295-8450",
    "9317331": "+1 727-242-8400",
    "9394343": "+1 813-344-3311",
    "9283918": "+1 813-815-7662",
    "9296161": "+1 727-381-0088",
    "4115014": "+1 727-844-3125",
    "3754544": "+1 727-924-2000",
    "3774587": "+1 727-586-5797",
    "3243192": "+1 727-443-4900",
    "3425752": "+1 813-654-4262",
    "3523975": "+1 813-825-1373",
    "9256830": "+1 727-270-6655",
    "9155471": "+1 813-994-9666",
    "4421753": "+1 813-961-4092",
    "7089467": "+1 727-726-4608",
    "2976346": "+1 813-374-2739",
    "6462895": "+1 727-934-4047",
    "2238952": "+1 727-327-8090",
    "5654752": "+1 727-863-0965",
    "5787421": "+1 727-614-0574",
    "7432888": "+1 727-954-7369"
}

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def normalize_columns(df):
    col_lower = {c.lower().strip(): c for c in df.columns}
    renames = {}
    for standard, variants in COLUMN_MAP.items():
        if standard in df.columns:
            continue
        for v in variants:
            if v.lower().strip() in col_lower:
                renames[col_lower[v.lower().strip()]] = standard
                break
    # Violation column variants: V14, Viol14, violation_14 → Violation 14
    viol_pat = re.compile(r'^[Vv](?:iol(?:ation)?[\s_-]?)?0?(\d+)$')
    for col in list(df.columns):
        m = viol_pat.match(col.strip())
        if m:
            new = f'Violation {int(m.group(1)):02d}'
            if new != col and col not in renames:
                renames[col] = new
    if renames:
        df = df.rename(columns=renames)
    return df

def load_csvs(paths):
    frames, ref_cols = [], None
    for p in paths:
        path = Path(p)
        if not path.exists():
            print(f"  WARNING: {path.name} not found, skipping")
            continue
        try:
            ext = path.suffix.lower()
            if ext in ('.xlsx', '.xlsm', '.xls'):
                # Stream XLSX row-by-row to handle large statewide files
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                headers = [str(c).strip() if c is not None else '' for c in rows[0]]
                data = [['' if v is None else str(v) for v in row] for row in rows[1:]]
                df = pd.DataFrame(data, columns=headers)
                print(f"  {path.name}: {len(df):,} rows (xlsx)")
            else:
                df = pd.read_csv(path, low_memory=False, encoding='utf-8',
                                 encoding_errors='replace')
                print(f"  {path.name}: {len(df):,} rows")

            df.columns = df.columns.str.strip()
            df = normalize_columns(df)
            if ref_cols is None:
                ref_cols = list(df.columns)
            frames.append(df)

        except Exception as e:
            try:
                if ref_cols and path.suffix.lower() == '.csv':
                    df = pd.read_csv(path, header=None, names=ref_cols,
                                     low_memory=False, encoding_errors='replace')
                    df = normalize_columns(df)
                    frames.append(df)
                    print(f"  {path.name}: {len(df):,} rows (headerless)")
                else:
                    print(f"  SKIP {path.name}: {e}")
            except Exception as e2:
                print(f"  SKIP {path.name}: {e2}")

    if not frames:
        raise RuntimeError("No valid data files loaded.")
    return pd.concat(frames, ignore_index=True)

def load_license_extract(data_dir=None):
    """
    Load FL DBPR active license extract.
    Returns dict keyed by License Number with phone, seats, rank, license_type.
    Tries data/ folder first, then current folder.
    """
    import pandas as pd
    candidates = []
    if data_dir:
        candidates.append(Path(data_dir) / 'hrfood3_licenses.csv')
    candidates += [
        Path(__file__).parent / 'data' / 'hrfood3_licenses.csv',
        Path(__file__).parent / 'hrfood3_licenses.csv',
    ]
    for path in candidates:
        if path.exists():
            try:
                df = pd.read_csv(path, low_memory=False, encoding_errors='replace')
                df.columns = df.columns.str.strip()
                # Normalize license number -- strip prefix letters for join
                # DBPR license numbers in extract: "SEA6213532"
                # In inspection data: numeric "6213532" or full "SEA6213532"
                result = {}
                for _, r in df.iterrows():
                    lic_raw = str(r.get('License Number', '')).strip()
                    # Store under both full and numeric forms
                    phone = str(r.get('Primary Phone Number', '') or
                                r.get('Secondary Phone Number', '') or '').strip()
                    phone = re.sub(r'[^0-9]', '', phone)
                    if len(phone) == 10:
                        phone = '+1 ' + phone[:3] + '-' + phone[3:6] + '-' + phone[6:]
                    elif len(phone) == 11 and phone.startswith('1'):
                        phone = '+1 ' + phone[1:4] + '-' + phone[4:7] + '-' + phone[7:]
                    else:
                        phone = ''
                    seats_raw = r.get('Number of Seats or Rental Units', 0)
                    try:
                        seats = int(float(str(seats_raw))) if seats_raw and str(seats_raw).strip() not in ('', 'nan') else 0
                    except:
                        seats = 0
                    entry = {
                        'phone':        phone,
                        'seats':        seats,
                        'rank':         str(r.get('Rank Code', '') or '').strip(),
                        'license_type': str(r.get('License Type Code', '') or '').strip(),
                        'risk_level':   str(r.get('Base Risk Level', '') or '').strip(),
                    }
                    result[lic_raw] = entry
                    # Also index by numeric portion for flexible joins
                    numeric = re.sub(r'[^0-9]', '', lic_raw)
                    if numeric and numeric not in result:
                        result[numeric] = entry
                n_phones = sum(1 for v in result.values() if v.get('phone'))
                print(f"  License extract: {len(df):,} records, {n_phones:,} with phones")
                return result
            except Exception as e:
                print(f"  WARNING: license extract load failed -- {e}")
    print("  License extract not found -- phones/seats from extract unavailable")
    return {}

def match_license(license_id, extract):
    """Flexible license ID matching -- try multiple formats."""
    lid = str(license_id).strip()
    if lid in extract:
        return extract[lid]
    numeric = re.sub(r'[^0-9]', '', lid)
    if numeric in extract:
        return extract[numeric]
    # Try common prefixes
    for prefix in ('SEA', 'NOS', 'MFD'):
        key = prefix + numeric
        if key in extract:
            return extract[key]
    return None

def ice_usage_label(name):
    n = name.lower()
    if any(w in n for w in HIGH_ICE): return 'high'
    if any(w in n for w in LOW_ICE):  return 'low'
    return 'medium'

def ice_insight_text(codes):
    if 'V14' in codes: return 'Food contact surfaces (V14) \u2014 inspectors directly cited food contact equipment. Ice machines are the primary target of this violation code.'
    if 'V22' in codes: return 'Non-PHF food contact surfaces (V22) \u2014 ice machines fall directly into this category.'
    if 'V50' in codes: return 'Food contact surfaces not clean (V50) \u2014 same root issue as V14/V22, cited at basic level.'
    if 'V37' in codes: return 'Equipment not in good repair (V37) \u2014 ice machine may be broken or malfunctioning.'
    if 'V23' in codes: return 'Utensil/container sanitation (V23) \u2014 includes ice scoops and bin components.'
    return ''

def _callback_fields(disposition: str, last_insp_date) -> dict:
    """Compute estimated callback inspection date from disposition type."""
    cb_days = CALLBACK_DAYS.get(disposition, 30)
    est_callback = last_insp_date + timedelta(days=cb_days)
    days_to_callback = (TODAY - est_callback).days * -1
    urgency = ('overdue'   if days_to_callback < 0  else
               'imminent'  if days_to_callback <= 7  else
               'soon'      if days_to_callback <= 21 else 'upcoming')
    return {
        'est_callback_date': est_callback.isoformat(),
        'days_to_callback':  days_to_callback,
        'callback_urgency':  urgency,
    }

def load_emergency_closures(data_dir=None):
    """Load recent DBPR emergency closure reports. Returns set of license IDs."""
    import openpyxl
    closed_ids = set()
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        if not d.exists(): continue
        for f in sorted(d.glob('EOS_Weekly_Extract_*.xlsx'), reverse=True)[:4]:
            try:
                wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows: continue
                headers = [str(h).strip() if h else '' for h in rows[0]]
                lic_col = next((i for i,h in enumerate(headers)
                                if 'license' in h.lower()), None)
                if lic_col is None: continue
                for row in rows[1:]:
                    val = row[lic_col]
                    if val:
                        num = re.sub(r'[^0-9]', '', str(val))
                        if num: closed_ids.add(num)
            except Exception:
                pass
    if closed_ids:
        print(f"  Emergency closures: {len(closed_ids)} businesses flagged")
    return closed_ids

def nominatim_geocode(address, city, state='FL', cache={}):
    """Geocode a single address using Nominatim (free, no key required).
    Uses in-memory cache to avoid duplicate requests."""
    import time, urllib.request, urllib.parse
    key = f"{address},{city},{state}"
    if key in cache:
        return cache[key]
    try:
        query = urllib.parse.urlencode({'q': f"{address}, {city}, {state}",
                                        'format': 'json', 'limit': 1})
        url = f"https://nominatim.openstreetmap.org/search?{query}"
        req = urllib.request.Request(url,
            headers={'User-Agent': 'PinellasIceCo/1.0 ice machine inspection tool'})
        with urllib.request.urlopen(req, timeout=5) as r:
            results = json.loads(r.read())
        if results:
            lat = float(results[0]['lat'])
            lon = float(results[0]['lon'])
            cache[key] = (lat, lon)
            time.sleep(1.1)  # Nominatim rate limit: 1 req/sec
            return (lat, lon)
    except Exception:
        pass
    cache[key] = None
    return None

def load_osm_cache(data_dir):
    """Load OSM phone/hours cache from download_data.py output."""
    cache_path = Path(data_dir) / 'osm_phones.json' if data_dir else None
    if not cache_path or not cache_path.exists():
        # Try default data/ folder
        alt = Path(__file__).parent / 'data' / 'osm_phones.json'
        if not alt.exists():
            return {}
        cache_path = alt
    try:
        import re as _re
        raw = json.loads(cache_path.read_text(encoding='utf-8'))
        osm = {}
        for county_data in raw.values():
            for biz in county_data.get('businesses', []):
                name = biz.get('name', '')
                if name and biz.get('phone'):
                    key = _re.sub(r'[^a-z0-9]', '', name.lower())
                    if key not in osm:
                        osm[key] = biz
        print(f"  OSM cache: {len(osm):,} businesses with phones/hours")
        return osm
    except Exception as e:
        print(f"  OSM cache load failed: {e}")
        return {}

def osm_match(name, city, osm_cache):
    """Fuzzy match a business name to OSM cache entry."""
    import re as _re
    if not osm_cache:
        return None
    key = _re.sub(r'[^a-z0-9]', '', name.lower())
    if key in osm_cache:
        return osm_cache[key]
    words = set(_re.sub(r'[^a-z0-9 ]', ' ', name.lower()).split()) - {
        'llc','inc','the','and','bar','grill','restaurant','cafe','pub'}
    if len(words) < 2:
        return None
    best_score, best_match = 0.7, None
    for cached_biz in osm_cache.values():
        cw = set(_re.sub(r'[^a-z0-9 ]', ' ', cached_biz.get('name','').lower()).split())
        overlap = len(words & cw) / max(len(words), len(cw), 1)
        if overlap > best_score:
            cc = (cached_biz.get('city') or '').lower()
            if not cc or city.lower()[:5] in cc or cc[:5] in city.lower():
                best_score = overlap
                best_match = cached_biz
    return best_match

def load_geo_cache(data_dir=None):
    """Load persistent geocoding cache from disk."""
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        p = d / 'geocache.json'
        if p.exists():
            try:
                data = json.loads(p.read_text())
                print(f"  Geocache: {len(data)} cached addresses")
                return data
            except Exception:
                pass
    return {}

def save_geo_cache(cache, data_dir=None):
    """Save geocoding cache to disk for next run."""
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        try:
            d.mkdir(exist_ok=True)
            (d / 'geocache.json').write_text(json.dumps(cache))
            return
        except Exception:
            pass

# ──────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ──────────────────────────────────────────────────────────────────────────────
def run(csv_paths):
    print(f"\n{'='*55}")
    print(f"  Pinellas Ice Co \u2014 Building Prospect Tool")
    print(f"  {TODAY}")
    print(f"{'='*55}\n")

    # LOAD
    print("Loading CSV files...")
    df = load_csvs(csv_paths)
    print(f"  Total rows loaded: {len(df):,}\n")

    # LOAD LICENSE EXTRACT (phones + seats + rank)
    print("Loading license extract...")
    data_dir = Path(csv_paths[0]).parent if csv_paths else None
    license_data = load_license_extract(data_dir)

    # LOAD EMERGENCY CLOSURES
    print("Loading emergency closures...")
    emergency_ids = load_emergency_closures(data_dir)

    # LOAD GEO CACHE
    geo_cache = load_geo_cache(data_dir)
    osm_cache = load_osm_cache(data_dir)
    print()

    # COERCE
    df['inspection_date'] = pd.to_datetime(df.get('Inspection Date'), errors='coerce')
    df = df.dropna(subset=['inspection_date'])
    df['county']    = df.get('County Name',          pd.Series('', index=df.index)).astype(str).str.strip()
    df['biz_name']  = df.get('Business (DBA-Does Business As) Name',
                              pd.Series('', index=df.index)).astype(str).str.strip()
    df['address']   = df.get('Location Address',     pd.Series('', index=df.index)).astype(str).str.strip()
    df['city']      = df.get('Location City',        pd.Series('', index=df.index)).astype(str).str.strip()
    df['zip']       = df.get('Location Zip Code',    pd.Series('', index=df.index)).astype(str).str.strip().str[:5]
    df['disp_raw']  = df.get('Inspection Disposition',pd.Series('',index=df.index)).astype(str).str.strip()
    df['insp_type'] = df.get('Inspection Type',      pd.Series('', index=df.index)).astype(str).str.strip()
    df['license_id']= df.get('License ID',           pd.Series(range(len(df)), index=df.index))

    for col, src_name in [('hv','Number of High Priority Violations'),
                           ('tv','Number of Total Violations'),
                           ('iv','Number of Intermediate Violations'),
                           ('bv','Number of Basic Violations'),
                           ('vn','Visit Number')]:
        df[col] = pd.to_numeric(df.get(src_name, pd.Series(0, index=df.index)),
                                errors='coerce').fillna(0)

    # Violation codes
    for code in ICE_CODES:
        col = f'Violation {code:02d}'
        df[col] = pd.to_numeric(df.get(col, pd.Series(0, index=df.index)),
                                errors='coerce').fillna(0)

    df['ice_rel']    = sum(df[f'Violation {c:02d}'].clip(upper=1) * s for c,s in ICE_CODES.items())
    df['direct_ice'] = (df['Violation 14'].clip(upper=1) +
                        df['Violation 22'].clip(upper=1) +
                        df['Violation 50'].clip(upper=1)) > 0
    df['dr'] = df['disp_raw'].map(DISP_RISK).fillna(2)

    # Filter
    df_t = df[df['county'].isin(TARGET_COUNTIES)].copy()
    print(f"Rows in target counties: {len(df_t):,}")

    # Routine inspections
    routine = df_t[df_t['insp_type'] == 'Routine - Food'].copy()
    routine = routine.sort_values(['license_id', 'inspection_date'])
    routine['next_date'] = routine.groupby('license_id')['inspection_date'].shift(-1)
    routine['interval']  = (routine['next_date'] - routine['inspection_date']).dt.days
    rv = routine[(routine['interval'] > 0) & (routine['interval'] < 730)].copy()
    MED = float(rv['interval'].median()) if len(rv) else 125.0
    print(f"Training on {len(rv):,} inspection intervals (median={MED:.0f}d)")

    # Train model
    le = LabelEncoder()
    rv['county_enc'] = le.fit_transform(rv['county'].fillna('Unknown'))
    rv['prev']  = rv.groupby('license_id')['interval'].transform(
        lambda x: x.expanding().mean().shift(1)).fillna(MED)
    rv['month'] = rv['inspection_date'].dt.month
    FEATS = ['hv','iv','bv','tv','dr','vn','month','county_enc','prev']
    rf = RandomForestRegressor(n_estimators=120, max_depth=8, random_state=42, n_jobs=-1)
    rf.fit(rv[FEATS].fillna(0), rv['interval'])
    print("Prediction model trained.\n")

    # Per-establishment history
    ice_hist = routine.groupby('license_id').agg(
        avg_interval  =('interval','mean'),
        n_insp        =('interval','count'),
        avg_hv        =('hv','mean'),
        avg_ice       =('ice_rel','mean'),
        direct_count  =('direct_ice','sum'),
    ).reset_index()
    ice_hist['chronic']   = ice_hist['direct_count'] >= 2
    ice_hist['confirmed'] = ice_hist['direct_count'] >= 1

    # ── NEW SIGNALS ────────────────────────────────────────────────────────────

    # 1. Callback inspection count per business
    #    Businesses with callbacks are actively being monitored — highest urgency
    callback_disps = ['Call Back - Admin. complaint recommended',
                      'Administrative complaint recommended',
                      'Admin. Complaint Callback Complied',
                      'Emergency Order Callback Not Complied',
                      'Emergency order recommended']
    routine['is_callback'] = routine['disp_raw'].isin(callback_disps)
    callback_hist = routine.groupby('license_id')['is_callback'].sum().reset_index()
    callback_hist.columns = ['license_id','n_callbacks']
    ice_hist = ice_hist.merge(callback_hist, on='license_id', how='left')
    ice_hist['n_callbacks'] = ice_hist['n_callbacks'].fillna(0).astype(int)

    # 2. Disposition escalation — did risk level increase over time?
    #    Compare avg disp_risk in first half of history vs second half
    def escalation_score(g):
        if len(g) < 4: return 0
        g = g.sort_values('inspection_date')
        mid = len(g)//2
        early = g.iloc[:mid]['dr'].mean()
        late  = g.iloc[mid:]['dr'].mean()
        return max(0, late - early)  # positive = getting worse
    esc = routine.groupby('license_id').apply(escalation_score)
    esc_df = esc.reset_index()
    esc_df.columns = ['license_id','escalation']
    ice_hist = ice_hist.merge(esc_df, on='license_id', how='left')
    ice_hist['escalation'] = ice_hist['escalation'].fillna(0)

    # 3. Ice violation recency — was the last ice violation recent?
    #    Recent = within 18 months, old = 3+ years ago
    ice_rows = routine[routine['direct_ice'] == True].copy()
    if len(ice_rows):
        last_ice = ice_rows.groupby('license_id')['inspection_date'].max().reset_index()
        last_ice.columns = ['license_id','last_ice_date']
        last_ice['days_since_ice'] = (pd.Timestamp(TODAY) - last_ice['last_ice_date']).dt.days
        ice_hist = ice_hist.merge(last_ice[['license_id','days_since_ice']], on='license_id', how='left')
    else:
        ice_hist['days_since_ice'] = 999
    ice_hist['days_since_ice'] = ice_hist['days_since_ice'].fillna(999).astype(int)
    ice_hist['ice_recent']  = ice_hist['days_since_ice'] <= 365   # within 1 year
    ice_hist['ice_fresh']   = ice_hist['days_since_ice'] <= 180   # within 6 months

    # 4. Violation code diversity — how many different violation types?
    def code_diversity(g):
        codes_seen = set()
        for col in [f'Violation {c:02d}' for c in ICE_CODES if f'Violation {c:02d}' in g.columns]:
            if (g[col]>0).any():
                codes_seen.add(col)
        return len(codes_seen)
    div = routine.groupby('license_id').apply(code_diversity)
    div_df = div.reset_index()
    div_df.columns = ['license_id','code_diversity']
    ice_hist = ice_hist.merge(div_df, on='license_id', how='left')
    ice_hist['code_diversity'] = ice_hist['code_diversity'].fillna(0).astype(int)

    # 5. Failed-first-visit rate — visit_number > 1 means they failed initial
    if 'vn' in routine.columns:
        visit_hist = routine.groupby('license_id')['vn'].agg(
            avg_visit='mean', max_visit='max').reset_index()
        ice_hist = ice_hist.merge(visit_hist, on='license_id', how='left')
        ice_hist['avg_visit'] = ice_hist['avg_visit'].fillna(1)
        ice_hist['has_callbacks_visit'] = ice_hist['avg_visit'] > 1.1
    else:
        ice_hist['avg_visit'] = 1.0
        ice_hist['has_callbacks_visit'] = False

    print(f"  New signals computed:")
    print(f"    Businesses with callbacks: {(ice_hist['n_callbacks']>0).sum():,}")
    print(f"    Businesses with escalation: {(ice_hist['escalation']>0.5).sum():,}")
    print(f"    Ice violation within 1yr: {ice_hist['ice_recent'].sum():,}")
    print(f"    Ice violation within 6mo: {ice_hist['ice_fresh'].sum():,}")
    print(f"    Multi-code violators: {(ice_hist['code_diversity']>=2).sum():,}")
    def trend_slope(g):
        if len(g) < 3: return 0.0
        x = np.arange(len(g), dtype=float)
        return float(np.polyfit(x, g['hv'].values.astype(float), 1)[0])
    trends = routine.groupby('license_id').apply(trend_slope).to_dict()

    # Fired ice codes per establishment
    def get_fired(g):
        codes = []
        for code in ICE_CODES:
            col = f'Violation {code:02d}'
            if col in g.columns and (pd.to_numeric(g[col], errors='coerce').fillna(0) > 0).any():
                codes.append(f'V{code}')
        return codes[:4]
    fired = routine.groupby('license_id').apply(get_fired).to_dict()

    # Predict next inspection date
    latest = routine.sort_values('inspection_date').groupby('license_id').last().reset_index()
    latest = latest.merge(ice_hist, on='license_id', how='left')
    for c in ['avg_interval','avg_hv','avg_ice','direct_count']:
        latest[c] = latest.get(c, 0).fillna(0)
    for c in ['chronic','confirmed']:
        latest[c] = latest.get(c, False).fillna(False).astype(bool)

    cs = latest['county'].where(latest['county'].isin(le.classes_), 'Unknown').fillna('Unknown')
    latest['county_enc'] = le.transform(cs)
    latest['prev']  = latest['avg_interval'].fillna(MED)
    latest['month'] = latest['inspection_date'].dt.month
    Xp = pd.DataFrame({f: latest.get(f, 0) for f in FEATS}).fillna(0)
    latest['pred_days']  = np.clip(rf.predict(Xp).round().astype(int), 7, 730)
    latest['pred_next']  = latest['inspection_date'] + pd.to_timedelta(latest['pred_days'], unit='d')
    latest['days_until'] = (latest['pred_next'] - pd.Timestamp(TODAY)).dt.days.fillna(999).astype(int)

    # Score
    def pitch_type(dr, du, ice, chronic, confirmed, days_since=0):
        if dr >= 4 and days_since <= 120: return 'callback'   # admin complaint/emergency = always callback if recent
        if dr >= 4 and -30 <= du < 75: return 'callback'
        if du < 0 and dr >= 2:         return 'overdue_urgent'
        if du < 0:                     return 'overdue'
        if du <= 21:                   return 'pre_hot'
        if du <= 45:                   return 'pre_warm'
        if (chronic or ice >= 80) and du <= 90: return 'high_risk'
        return 'routine'

    def ice_score(pt, dr, hv, ni, ice, chronic, confirmed,
                  n_callbacks=0, escalation=0, ice_recent=False,
                  ice_fresh=False, code_diversity=0, avg_visit=1.0):
        # Base by pitch type - recalibrated for better spread
        b = {'callback':55,'overdue_urgent':48,'overdue':38,'pre_hot':45,
             'pre_warm':35,'high_risk':30,'routine':15}.get(pt, 15)
        # Ice history - primary differentiator
        if ice_fresh:      b += 20  # ice violation within 6 months = urgent
        elif ice_recent:   b += 12  # ice violation within 1 year
        if chronic:        b += 15  # 2+ inspections with ice violations
        elif confirmed:    b += 8   # 1 ice violation on record
        # Callback signals - inspector already watching this business
        if n_callbacks >= 3: b += 18
        elif n_callbacks == 2: b += 12
        elif n_callbacks == 1: b += 7
        # Escalation - getting systematically worse
        if escalation > 1.5: b += 12
        elif escalation > 0.8: b += 7
        elif escalation > 0.3: b += 3
        # Code diversity - breadth of violation types
        if code_diversity >= 4: b += 8
        elif code_diversity >= 3: b += 5
        elif code_diversity >= 2: b += 2
        # Failed first visits
        if avg_visit > 1.5: b += 6
        elif avg_visit > 1.2: b += 3
        # Violation severity
        if hv >= 6: b += 8
        elif hv >= 4: b += 5
        elif hv >= 2: b += 2
        if (ni or 1) >= 6: b += 5
        elif (ni or 1) >= 3: b += 2
        if dr >= 5: b += 8
        elif dr >= 4: b += 4
        if ice >= 200: b += 6
        elif ice >= 100: b += 3
        return min(100, max(5, b))

    def priority(pt):
        return {'callback':'CALLBACK','overdue_urgent':'CALLBACK',
                'pre_hot':'HOT','pre_warm':'WARM','overdue':'WARM',
                'high_risk':'WATCH'}.get(pt, 'LATER')

    # Score ALL businesses - no caps, no sampling
    # Browser does the filtering
    recent = latest[latest['inspection_date'] >= '2023-07-01'].copy()
    recent['_days_since'] = (pd.Timestamp(TODAY) - recent['inspection_date']).dt.days.fillna(999).astype(int)

    print("Scoring all businesses...")
    pts, scs, prs = [], [], []
    for _, r in recent.iterrows():
        pt = pitch_type(r['dr'], r['days_until'], r['avg_ice'],
                        bool(r['chronic']), bool(r['confirmed']), int(r['_days_since']))
        sc = ice_score(
            pt, r['dr'],
            r['avg_hv'] if pd.notna(r['avg_hv']) else r['hv'],
            r.get('n_insp',1) or 1,
            r['avg_ice'],
            bool(r['chronic']),
            bool(r['confirmed']),
            n_callbacks   = int(r.get('n_callbacks', 0) or 0),
            escalation    = float(r.get('escalation', 0) or 0),
            ice_recent    = bool(r.get('ice_recent', False)),
            ice_fresh     = bool(r.get('ice_fresh', False)),
            code_diversity= int(r.get('code_diversity', 0) or 0),
            avg_visit     = float(r.get('avg_visit', 1.0) or 1.0),
        )
        pts.append(pt); scs.append(sc); prs.append(priority(pt))

    recent['_pt'] = pts
    recent['_sc'] = scs
    recent['_pr'] = prs
    recent['_pinellas'] = (recent['county'] == 'Pinellas').astype(int)
    recent = recent[recent['_sc'] >= MIN_SCORE]
    recent = recent.sort_values(['_sc','_pinellas'], ascending=[False, False])

    result = recent.drop_duplicates('license_id').reset_index(drop=True)
    print(f"  {len(result):,} businesses scored")

    # Rename underscore columns -- itertuples strips leading underscores
    result = result.rename(columns={
        '_sc': 'sc', '_pt': 'pt', '_pr': 'pr',
        '_pinellas': 'pinellas', '_days_since': 'days_since_col'
    })

    print(f"Building records (vectorized)...")
    # Pre-compute all classification in bulk -- avoids slow iterrows
    result['biz_type'] = result['biz_name'].apply(
        lambda n: classify_business(str(n)[:50])[0])
    result = result[result['biz_type'] != 'corporate'].copy()
    print(f"  {len(result):,} non-corporate businesses")

    # Vectorized lat/lon from ZIP
    result['z5'] = result['zip'].astype(str).str[:5]
    result['lat'] = result['z5'].map(lambda z: ZIP_COORDS.get(z, (None,None))[0])
    result['lon'] = result['z5'].map(lambda z: ZIP_COORDS.get(z, (None,None))[1])

    # Vectorized bar detection
    bar_words = ['bar','pub','tavern','lounge','brewery','taproom','cantina',
                 'saloon','sports bar','grill & bar','grille & bar','raw bar']
    result['is_bar'] = result['biz_name'].str.lower().apply(
        lambda n: any(w in str(n) for w in bar_words))

    # Vectorized emergency flag
    result['is_emergency'] = result['license_id'].astype(str).isin(emergency_ids)

    # Vectorized days_since
    result['days_since'] = (pd.Timestamp(TODAY) - result['inspection_date']).dt.days.fillna(999).astype(int)

    # Fill missing columns safely
    for col, default in [('direct_count',0),('avg_ice',0),('n_insp',1),('disp_raw',''),
                          ('chronic',False),('confirmed',False)]:
        if col not in result.columns: result[col] = default
    result['direct_count'] = result['direct_count'].fillna(0).astype(int)
    result['avg_ice']  = result['avg_ice'].fillna(0)
    result['n_insp']   = result['n_insp'].fillna(1).astype(int)
    result['disp_raw'] = result['disp_raw'].fillna('').astype(str)

    # Build records using itertuples (fast)
    records = []
    for row in result.itertuples(index=False):
        try:
            lid   = int(row.license_id)
            name  = str(row.biz_name)[:50]
            codes = fired.get(lid, [])
            z5    = str(row.zip)[:5]

            lic_info  = match_license(lid, license_data) or {}
            phone_raw = lic_info.get('phone','') or STATIC_PHONES.get(str(lid),'')
            hours_raw = ''
            rating_raw = 0
            # OSM enrichment - phone + hours for unmatched records
            if not phone_raw or not hours_raw:
                osm_hit = osm_match(name, str(row.city), osm_cache)
                if osm_hit:
                    if not phone_raw:
                        phone_raw = osm_hit.get('phone','')
                    if not hours_raw:
                        hours_raw = osm_hit.get('hours','')[:60] if osm_hit.get('hours') else ''
            seats     = int(lic_info.get('seats', 0) or 0)
            rank      = lic_info.get('rank', 'SEAT')

            machines    = est_machines(seats, bool(row.is_bar), rank)
            monthly_val = est_monthly(machines)
            onetime_val = est_onetime(machines)
            intro_val   = 99  # flat intro offer regardless of machine count
            confirmed   = bool(row.confirmed)
            chronic     = bool(row.chronic)
            tier        = account_tier(seats, rank, machines, chronic, confirmed)
            days_since  = int(row.days_since)

            lat = row.lat if row.lat and row.lat == row.lat else None
            lon = row.lon if row.lon and row.lon == row.lon else None
            if lat is None:
                cached = geo_cache.get(f"{row.address},{row.city},FL")
                if cached: lat, lon = cached[0], cached[1]

            n_insp      = int(row.n_insp or 1)
            avg_ice     = float(row.avg_ice or 0)
            disp_raw    = str(row.disp_raw or '')
            final_score = min(100, int(row.sc) + seat_score_bonus(machines, rank))
            conf        = confidence_score(n_insp, 1, confirmed, days_since)

            records.append({
                'id':          lid,
                'name':        name,
                'county':      str(row.county),
                'city':        str(row.city),
                'address':     str(row.address),
                'zip':         z5,
                'lat':         lat,
                'lon':         lon,
                'last_insp':   row.inspection_date.strftime('%Y-%m-%d'),
                'last_disp':   disp_raw,
                'disp_risk':   int(row.dr),
                'high_viol':   int(row.hv),
                'total_viol':  int(row.tv),
                'n_insp':      n_insp,
                'pred_next':   row.pred_next.strftime('%Y-%m-%d'),
                'days_until':  int(row.days_until),
                'days_since':  days_since,
                'pitch_type':  str(row.pt),
                'score':       final_score,
                'confidence':  conf,
                'priority':    str(row.pr),
                'ice_rel':     round(avg_ice, 0),
                'confirmed':   confirmed,
                'chronic':     chronic,
                'ice_count':   int(row.direct_count),
                'codes':       codes,
                'trending':    float(trends.get(lid, 0)) > 0.5,
                'is_emergency':bool(row.is_emergency),
                'biz_type':    str(row.biz_type),
                'is_bar':      bool(row.is_bar),
                'seats':       seats,
                'machines':    machines,
                'monthly':     monthly_val,
                'onetime':     onetime_val,
                'intro':       intro_val,
                'tier':        tier,
                'phone':       phone_raw,
                'status':      'prospect',
                'rating':      0,
                'hours':       hours_raw,
                # New intelligence signals
                'n_callbacks':    int(row.n_callbacks) if hasattr(row,'n_callbacks') else 0,
                'escalation':     round(float(row.escalation),2) if hasattr(row,'escalation') else 0,
                'ice_recent':     bool(row.ice_recent) if hasattr(row,'ice_recent') else False,
                'ice_fresh':      bool(row.ice_fresh) if hasattr(row,'ice_fresh') else False,
                'days_since_ice': int(row.days_since_ice) if hasattr(row,'days_since_ice') else 999,
                'code_diversity': int(row.code_diversity) if hasattr(row,'code_diversity') else 0,
                'avg_visit':      round(float(row.avg_visit),1) if hasattr(row,'avg_visit') else 1.0,
            })
        except Exception:
            pass

    PO = {'CALLBACK':0,'HOT':1,'WARM':2,'WATCH':3,'LATER':4}
    records.sort(key=lambda x: (PO.get(x['priority'], 5), -x['score']))

    # Save geocoding cache for future runs
    save_geo_cache(geo_cache, data_dir)

    # Preserve phones from previous run
    if OUTPUT_FILE.exists():
        try:
            content = OUTPUT_FILE.read_text(encoding='utf-8')
            m = re.search(r'const PHONES=(\{.*?\});', content, re.DOTALL)
            if m:
                saved_phones = json.loads(m.group(1))
                n_restored = 0
                for rec in records:
                    ph = saved_phones.get(str(rec['id']), {})
                    if ph.get('phone'):
                        rec['phone']  = ph['phone']
                        rec['rating'] = ph.get('rating', 0)
                        rec['hours']  = ph.get('hours', '')
                        n_restored += 1
                if n_restored:
                    print(f"  Phone data restored for {n_restored} businesses")
        except Exception:
            pass

    pc = dict(Counter(r['priority'] for r in records))
    print(f"\nResults:")
    print(f"  Total prospects:       {len(records)}")
    print(f"  Callback urgent:       {pc.get('CALLBACK',0)}")
    print(f"  Hot (≤21 days):        {pc.get('HOT',0)}")
    print(f"  Warm (22-45 days):     {pc.get('WARM',0)}")
    print(f"  Watch:                 {pc.get('WATCH',0)}")
    print(f"  Chronic ice offenders: {sum(1 for r in records if r.get('chronic'))}")
    print(f"  With map coordinates:  {sum(1 for r in records if r['lat'])}")
    print(f"  With phone number:     {sum(1 for r in records if r['phone'])}")

    return records

# ──────────────────────────────────────────────────────────────────────────────
# HTML
# ──────────────────────────────────────────────────────────────────────────────
def build_html(records):
    data_js   = json.dumps(records, separators=(',',':'))
    phones_js = json.dumps(
        {str(r['id']): {'phone':r['phone'],'rating':r['rating'],'hours':r['hours']}
         for r in records if r['phone']},
        separators=(',',':')
    )
    zip_js = json.dumps({z: list(c) for z,c in ZIP_COORDS.items()}, separators=(',',':'))

    n_cb    = sum(1 for r in records if r['priority']=='CALLBACK')
    n_hot   = sum(1 for r in records if r['priority']=='HOT')
    n_warm  = sum(1 for r in records if r['priority']=='WARM')
    n_phone = sum(1 for r in records if r['phone'])
    n_chron = sum(1 for r in records if r.get('chronic'))
    n_geo   = sum(1 for r in records if r['lat'])

    # Read the HTML template embedded below
    return HTML_TEMPLATE.replace('%%DATA%%', data_js)\
                        .replace('%%PHONES%%', phones_js)\
                        .replace('%%ZIPS%%', zip_js)\
                        .replace('%%DATE%%', str(TODAY))\
                        .replace('%%TOTAL%%', str(len(records)))\
                        .replace('%%NCB%%', str(n_cb))\
                        .replace('%%NHOT%%', str(n_hot))\
                        .replace('%%NWARM%%', str(n_warm))\
                        .replace('%%NPHONE%%', str(n_phone))\
                        .replace('%%NCHRON%%', str(n_chron))\
                        .replace('%%NGEO%%', str(n_geo))\
                        .replace('%%NCONF%%', str(sum(1 for r in records if r['confirmed'])))

# ──────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE  (everything between the triple-quotes)
# ──────────────────────────────────────────────────────────────────────────────

def build_html(records):
    data_js   = json.dumps(records, separators=(',',':'))
    phones_js = json.dumps(
        {str(r['id']): {'phone':r['phone'],'rating':r['rating'],'hours':r['hours']}
         for r in records if r['phone']},
        separators=(',',':')
    )
    zip_js = json.dumps({z: list(c) for z,c in ZIP_COORDS.items()}, separators=(',',':'))

    n_cb    = sum(1 for r in records if r['priority']=='CALLBACK')
    n_hot   = sum(1 for r in records if r['priority']=='HOT')
    n_warm  = sum(1 for r in records if r['priority']=='WARM')
    n_phone = sum(1 for r in records if r['phone'])
    n_chron = sum(1 for r in records if r.get('chronic'))
    n_geo   = sum(1 for r in records if r['lat'])

    # Read the HTML template embedded below
    return HTML_TEMPLATE.replace('%%DATA%%', data_js)\
                        .replace('%%PHONES%%', phones_js)\
                        .replace('%%ZIPS%%', zip_js)\
                        .replace('%%DATE%%', str(TODAY))\
                        .replace('%%TOTAL%%', str(len(records)))\
                        .replace('%%NCB%%', str(n_cb))\
                        .replace('%%NHOT%%', str(n_hot))\
                        .replace('%%NWARM%%', str(n_warm))\
                        .replace('%%NPHONE%%', str(n_phone))\
                        .replace('%%NCHRON%%', str(n_chron))\
                        .replace('%%NGEO%%', str(n_geo))\
                        .replace('%%NCONF%%', str(sum(1 for r in records if r['confirmed'])))

# ──────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE  (everything between the triple-quotes)
# ──────────────────────────────────────────────────────────────────────────────
HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Pinellas Ice Co &#xB7; Prospects</title>
<!-- Leaflet loaded on-demand when Route tab opens -->
<link rel="manifest" href="manifest.json">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="PIC Prospects">
<link rel="apple-touch-icon" href="apple-touch-icon.png">
<meta name="theme-color" content="#2d3e50">
<link href="https://fonts.googleapis.com/css2?family=Lexend:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
:root{
  --bg:#f5f8fa;--surf:#ffffff;--brd:#dfe3eb;--brd2:#eaf0f6;
  --txt:#33475b;--sub:#7c98b6;--sub2:#516f90;
  --cb:#f2545b;--hot:#ff7a59;--warm:#f5c26b;--watch:#00bda5;
  --grn:#00bda5;--blu:#0091ae;--ora:#ff7a59;--navy:#2d3e50;
  --shadow:0 1px 3px rgba(45,62,80,.1),0 1px 8px rgba(45,62,80,.06);
  --shadow-md:0 3px 12px rgba(45,62,80,.12),0 1px 4px rgba(45,62,80,.08);
}
html,body{height:100%;background:var(--bg);color:var(--txt);
  font-family:'Lexend',sans-serif;
  font-size:13px;overflow:hidden;-webkit-font-smoothing:antialiased}
#app{display:flex;flex-direction:column;height:100vh}
header{background:var(--navy);
  padding:0 16px;display:flex;align-items:center;gap:10px;flex-shrink:0;flex-wrap:wrap;height:52px}
.logo{display:flex;align-items:center;gap:8px;flex-shrink:0}
.logo-icon{width:30px;height:30px;border-radius:8px;
  background:var(--ora);
  display:flex;align-items:center;justify-content:center;font-size:16px}
.logo-name{font-weight:700;font-size:14px;color:#fff;letter-spacing:-.02em}
.hchips{display:flex;gap:5px;margin-left:auto;flex-wrap:wrap}
.hc{font-size:10px;padding:3px 10px;border-radius:20px;font-weight:600;cursor:pointer;user-select:none;transition:.15s}
.hc.cb{background:rgba(242,84,91,.2);color:#ffb3b6;border:1px solid rgba(242,84,91,.3)}
.hc.ht{background:rgba(255,122,89,.2);color:#ffcab8;border:1px solid rgba(255,122,89,.3)}
.hc.wm{background:rgba(245,194,107,.2);color:#ffe5a8;border:1px solid rgba(245,194,107,.3)}
.hc.ic{background:rgba(0,189,165,.2);color:#7eeee3;border:1px solid rgba(0,189,165,.3)}
.hc.rt{background:rgba(0,145,174,.2);color:#7dd4e8;border:1px solid rgba(0,145,174,.3)}
.srch{position:relative}
.srch input{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);border-radius:8px;padding:6px 10px 6px 28px;color:#fff;font-size:12px;outline:none;width:160px;font-family:inherit}
.srch input::placeholder{color:rgba(255,255,255,.65)}
.srch input:focus{background:rgba(255,255,255,.18);border-color:rgba(255,255,255,.4)}
.si{position:absolute;left:9px;top:50%;transform:translateY(-50%);color:rgba(255,255,255,.5);font-size:11px;pointer-events:none}
.tabs{display:flex;background:var(--surf);border-bottom:2px solid var(--brd2);flex-shrink:0;
  box-shadow:0 1px 3px rgba(45,62,80,.06)}
.tab{flex:1;padding:10px 4px;text-align:center;font-size:11px;font-weight:600;
  color:var(--sub);cursor:pointer;border-bottom:2px solid transparent;
  margin-bottom:-2px;transition:.15s;user-select:none}
.tab.on{color:var(--ora);border-color:var(--ora)}
.panel{flex:1;overflow-y:auto;padding:14px 16px;display:none;background:var(--bg)}
.panel.on{display:block}
.fbar{display:flex;gap:6px;margin-bottom:12px;flex-wrap:wrap;align-items:center}
.fbar select{background:var(--surf);border:1px solid var(--brd);border-radius:6px;
  padding:5px 9px;color:var(--txt);font-size:11px;outline:none;cursor:pointer;
  font-family:inherit;font-weight:500;box-shadow:var(--shadow)}
.fcnt{font-size:10px;color:var(--sub);margin-left:auto;font-weight:500}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(276px,1fr));gap:10px}
.card{background:var(--surf);border:1px solid var(--brd);border-radius:10px;
  padding:13px 14px;cursor:pointer;transition:box-shadow .15s,transform .12s,border-color .15s;
  position:relative;overflow:hidden;box-shadow:var(--shadow)}
.card:hover{box-shadow:var(--shadow-md);transform:translateY(-1px);border-color:#c5d1de}
.card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;border-radius:10px 0 0 10px}
.card.CALLBACK::before{background:var(--cb)}
.card.HOT::before{background:var(--hot)}
.card.WARM::before{background:var(--warm)}
.card.WATCH::before{background:var(--watch)}
.card.done{opacity:.45}
.ctop{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:3px}
.cname{font-weight:700;font-size:12.5px;flex:1;padding-right:6px;line-height:1.3;color:var(--navy)}
.pbadge{font-size:9px;font-weight:700;padding:2px 8px;border-radius:20px;letter-spacing:.04em;white-space:nowrap}
.pbadge.CALLBACK{background:#fdeaea;color:#c0392b;border:1px solid #f5c6c6}
.pbadge.HOT{background:#fff2ee;color:#c9542b;border:1px solid #fdd5c6}
.pbadge.WARM{background:#fef9ee;color:#9e6b1a;border:1px solid #fbe8b5}
.pbadge.WATCH{background:#e8faf8;color:#007a6e;border:1px solid #b3ece6}
.cloc{font-size:10px;color:var(--sub);margin-bottom:6px;font-weight:500}
.phrow{display:flex;align-items:center;gap:6px;padding:7px 9px;
  background:#f5f8fa;border-radius:7px;border:1px solid var(--brd2);margin-bottom:6px}
.phnum{font-size:12px;font-weight:600;color:var(--blu);flex:1}
.phnum.none{color:var(--sub);font-size:10px;font-style:italic;font-weight:400}
.abtn{border:none;border-radius:5px;padding:3px 9px;font-size:10px;font-weight:700;
  cursor:pointer;text-decoration:none;display:inline-block;white-space:nowrap;font-family:inherit}
.call-a{background:var(--grn);color:#fff}
.find-a{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7}
.icebadge{font-size:9px;font-weight:700;padding:3px 9px;border-radius:20px;
  margin-bottom:5px;display:inline-block}
.icebadge.chronic{background:#e8faf8;color:#007a6e;border:1px solid #b3ece6}
.icebadge.confirmed{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7}
.cmeta{display:flex;gap:8px;margin-bottom:6px;flex-wrap:wrap}
.mi{display:flex;flex-direction:column;gap:1px}
.ml{font-size:8px;color:var(--sub);letter-spacing:.05em;text-transform:uppercase;font-weight:600}
.mv{font-weight:700;font-size:11px;color:var(--txt)}
.mv.u{color:var(--cb)}.mv.h{color:var(--hot)}.mv.w{color:#9e6b1a}.mv.bad{color:var(--cb)}
.insight{font-size:9px;color:var(--sub2);background:#f5f8fa;border-radius:5px;border:1px solid var(--brd2);
  padding:4px 8px;margin-bottom:5px;line-height:1.5}
.lastc{font-size:10px;color:var(--sub);margin-bottom:6px}
.lastc.hc{color:#007a6e;font-weight:500}
.cacts{display:flex;gap:5px}
.btn{border:none;border-radius:6px;padding:5px 11px;font-size:10px;font-weight:600;cursor:pointer;font-family:inherit}
.blog{background:var(--ora);color:#fff}.blog:hover{background:#e86744}
.bskip{background:transparent;color:var(--sub);border:1px solid var(--brd)}
.brt{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7;font-size:9px}
.ptbl{width:100%;border-collapse:collapse}
.ptbl th{padding:7px 10px;text-align:left;font-size:9px;color:var(--sub2);font-weight:700;
  letter-spacing:.06em;text-transform:uppercase;border-bottom:2px solid var(--brd);
  white-space:nowrap;position:sticky;top:0;background:var(--surf);z-index:1}
.ptbl td{padding:8px 10px;border-bottom:1px solid var(--brd2);vertical-align:middle}
.ptbl tr:hover td{background:#f5f8fa;cursor:pointer}
.ptbl tr.done td{opacity:.4}
.tn{font-weight:600;font-size:11px;max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--navy)}
.ts{font-size:9px;color:var(--sub)}
.tph{font-size:11px;color:var(--blu);font-weight:600;text-decoration:none;white-space:nowrap}
.tph:hover{color:#007a9e}
.tph.none{color:#b0c4d8;font-size:9px}
.td{font-weight:700;font-size:12px}
.td.u{color:var(--cb)}.td.h{color:var(--hot)}.td.w{color:#9e6b1a}
/* ROUTE */
.route-wrap{display:grid;grid-template-columns:280px 1fr;gap:12px;height:calc(100vh - 120px)}
@media(max-width:600px){.route-wrap{grid-template-columns:1fr;grid-template-rows:auto 1fr}}
.rsidebar{display:flex;flex-direction:column;gap:8px;overflow-y:auto}
.rcontrols{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:12px;box-shadow:var(--shadow)}
.rtitle{font-weight:700;font-size:12px;margin-bottom:8px;color:var(--navy)}
.rlist{display:flex;flex-direction:column;gap:4px;overflow-y:auto;max-height:320px}
.rcard{background:var(--surf);border:1px solid var(--brd);border-radius:7px;
  padding:8px 10px;cursor:pointer;transition:.12s;display:flex;align-items:center;gap:7px;
  box-shadow:var(--shadow)}
.rcard:hover{border-color:#c5d1de;box-shadow:var(--shadow-md)}
.rcard.sel{border-color:var(--ora);background:#fff7f5}
.rdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.rname{font-weight:600;font-size:11px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--navy)}
.rdist{font-size:9px;color:var(--sub);flex-shrink:0}
.map-area{background:#e8edf2;border:1px solid var(--brd);border-radius:10px;
  display:flex;align-items:center;justify-content:center;overflow:hidden;position:relative;min-height:300px;
  box-shadow:var(--shadow)}
.map-empty{color:var(--sub);font-size:12px;text-align:center;padding:20px}
.day-route{background:var(--surf);border:1px solid var(--brd);border-radius:10px;
  padding:11px;margin-top:8px;box-shadow:var(--shadow)}
.day-stop{display:flex;align-items:center;gap:7px;padding:6px 0;
  border-bottom:1px solid var(--brd2);font-size:11px}
.day-stop:last-child{border-bottom:none}
.stopnum{width:20px;height:20px;border-radius:50%;background:var(--ora);
  color:#fff;font-weight:700;font-size:9px;display:flex;align-items:center;
  justify-content:center;flex-shrink:0}
.route-btn{background:var(--ora);color:#fff;border:none;border-radius:8px;
  padding:9px;font-size:11px;font-weight:700;cursor:pointer;width:100%;margin-top:8px;font-family:inherit}
.route-btn:hover{background:#e86744}
.route-btn.sec{background:#f5f8fa;color:var(--sub2);border:1px solid var(--brd);margin-top:5px}
/* MODAL */
#mbg{position:fixed;inset:0;background:rgba(45,62,80,.6);z-index:100;
  display:none;backdrop-filter:blur(4px)}
#mbg.on{display:flex;align-items:flex-end;justify-content:center}
@media(min-width:600px){#mbg.on{align-items:center}}
#modal{background:#fff;border:1px solid var(--brd);border-radius:16px 16px 0 0;
  width:100%;max-width:540px;padding:18px;max-height:90vh;overflow-y:auto;
  box-shadow:0 -4px 24px rgba(45,62,80,.15)}
@media(min-width:600px){#modal{border-radius:16px;box-shadow:0 8px 40px rgba(45,62,80,.18)}}
.mh{width:36px;height:4px;background:var(--brd);border-radius:2px;margin:0 auto 14px}
.mname{font-size:17px;font-weight:700;margin-bottom:2px;color:var(--navy)}
.mloc{font-size:10px;color:var(--sub);margin-bottom:12px;font-weight:500}
.mphsec{background:#f5f8fa;border:1px solid var(--brd2);border-radius:10px;padding:12px;margin-bottom:11px}
.mphl{font-size:8px;color:var(--sub);letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px;font-weight:700}
.mphnum{font-size:20px;font-weight:700;color:var(--blu);margin-bottom:8px}
.mphnum.none{font-size:12px;color:var(--sub);font-weight:400;font-style:italic}
.mphacts{display:flex;gap:6px;flex-wrap:wrap}
.mcall{background:var(--grn);color:#fff;border:none;border-radius:8px;padding:8px 16px;
  font-size:11px;font-weight:700;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;gap:5px;font-family:inherit}
.msms{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7;border-radius:8px;
  padding:8px 12px;font-size:11px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:4px}
.mgoog{background:#f5f8fa;color:var(--sub2);border:1px solid var(--brd);border-radius:8px;
  padding:8px 12px;font-size:11px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:4px}
.mhours{font-size:9px;color:var(--sub);margin-top:5px}
.micebox{background:#e8faf8;border:1px solid #b3ece6;border-radius:9px;padding:11px;margin-bottom:11px}
.msect{font-size:9px;color:var(--sub2);letter-spacing:.07em;text-transform:uppercase;
  margin-bottom:6px;padding-bottom:5px;border-bottom:1px solid var(--brd2);font-weight:700}
.pitch{background:#f5f8fa;border:1px solid var(--brd2);border-radius:8px;padding:10px;
  font-size:11px;color:var(--sub2);line-height:1.7;margin-bottom:9px}
.pitch b{color:var(--navy)}
.ogrid{display:grid;grid-template-columns:repeat(3,1fr);gap:5px;margin-bottom:8px}
.obtn{padding:7px 4px;border-radius:7px;border:1px solid var(--brd);
  background:var(--surf);color:var(--sub2);font-size:10px;font-weight:600;
  cursor:pointer;text-align:center;transition:.12s;font-family:inherit}
.obtn:hover{border-color:#c5d1de;background:#f5f8fa}
.obtn.on{border-color:var(--ora);background:#fff7f5;color:var(--ora)}
.ntxt{width:100%;background:#f5f8fa;border:1px solid var(--brd);border-radius:8px;
  padding:8px;color:var(--txt);font-size:11px;resize:none;outline:none;
  line-height:1.5;font-family:inherit}
.ntxt:focus{border-color:var(--blu)}
.mfacts{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.fact{background:#f5f8fa;border:1px solid var(--brd2);border-radius:7px;padding:8px 9px}
.fl{font-size:8px;color:var(--sub);letter-spacing:.05em;text-transform:uppercase;margin-bottom:2px;font-weight:600}
.fv{font-weight:700;font-size:11px;color:var(--navy)}
.fv.r{color:var(--cb)}.fv.o{color:var(--hot)}.fv.g{color:var(--grn)}.fv.b{color:var(--blu)}
.mhist{background:#f5f8fa;border-radius:8px;padding:9px}
.hi{padding:4px 0;border-bottom:1px solid var(--brd2);font-size:10px;color:var(--txt)}
.hi:last-child{border-bottom:none}
.psect{margin-bottom:16px}
.pst{font-size:10px;font-weight:700;letter-spacing:.05em;text-transform:uppercase;
  margin-bottom:7px;display:flex;align-items:center;gap:6px;color:var(--navy)}
.pct{background:var(--brd2);color:var(--sub2);border-radius:20px;padding:1px 8px;font-size:9px;font-weight:700}
.pitem{background:var(--surf);border:1px solid var(--brd);border-radius:8px;
  padding:9px 12px;margin-bottom:5px;display:flex;align-items:center;gap:8px;cursor:pointer;
  box-shadow:var(--shadow);transition:.12s}
.pitem:hover{box-shadow:var(--shadow-md);border-color:#c5d1de}
.pdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.piname{font-weight:600;font-size:11px;flex:1;color:var(--navy)}
.piph{font-size:10px;color:var(--blu);font-weight:600}
.dc{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:14px;margin-bottom:10px;box-shadow:var(--shadow)}
.dct{font-weight:700;font-size:13px;margin-bottom:8px;color:var(--navy)}
.ds{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid var(--brd2);font-size:11px}
.ds:last-child{border-bottom:none}
.dsv{font-weight:700;color:var(--blu)}
.ibox{background:#f0f9fb;border:1px solid #b3dce7;border-radius:8px;
  padding:11px;font-size:11px;color:var(--sub2);line-height:1.8}
.ibox code{background:#dff0f5;padding:2px 6px;border-radius:4px;
  font-family:monospace;font-size:10px;color:var(--blu)}
.ibox b{color:var(--navy)}
.phinput{background:#f5f8fa;border:1px solid var(--brd);border-radius:7px;
  padding:7px 10px;color:var(--txt);font-size:11px;outline:none;width:100%;margin-bottom:7px;font-family:inherit}
.phinput:focus{border-color:var(--blu)}
.xbtn{background:var(--ora);color:#fff;border:none;
  border-radius:8px;padding:9px 12px;font-size:11px;font-weight:700;cursor:pointer;width:100%;margin-top:6px;font-family:inherit}
.xbtn:hover{background:#e86744}
.dbtn{background:#fff;color:var(--cb);border:1px solid #f5c6c6;
  border-radius:8px;padding:9px 12px;font-size:11px;font-weight:600;cursor:pointer;width:100%;margin-top:5px;font-family:inherit}
.empty{text-align:center;padding:36px 20px;color:var(--sub)}
.ei{font-size:32px;margin-bottom:10px}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.2}}.pulse{animation:pulse 2s ease-in-out infinite}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--brd);border-radius:2px}
#toast{position:fixed;bottom:20px;left:50%;transform:translateX(-50%) translateY(60px);
  background:var(--navy);border-radius:10px;
  padding:9px 16px;font-size:11px;color:#fff;transition:transform .2s;z-index:200;white-space:nowrap;
  box-shadow:0 4px 16px rgba(45,62,80,.25);font-weight:500}
.tsect-hdr{display:flex;flex-direction:column;gap:2px;margin-bottom:10px;padding-bottom:8px;border-bottom:2px solid var(--brd2)}
.tsect-hdr span:first-child{font-weight:700;font-size:14px;color:var(--navy)}
.tsect-sub{font-size:10px;color:var(--sub);font-weight:400}
.tempty{text-align:center;padding:16px;color:var(--sub);font-size:11px;background:var(--surf);border-radius:8px;border:1px dashed var(--brd)}
.emerg-badge{display:inline-block;font-size:8px;font-weight:700;padding:2px 6px;border-radius:20px;background:#fef2f2;color:#dc2626;border:1px solid #fecaca;margin-left:4px}
.preset-btn{flex-shrink:0;padding:6px 12px;border-radius:20px;border:1px solid var(--brd);background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;white-space:nowrap;font-family:inherit;transition:.15s}
.preset-btn.on{background:var(--navy);color:#fff;border-color:var(--navy)}
.svc-tab{flex:1;padding:8px 4px;border:none;background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;font-family:inherit;border-right:1px solid var(--brd);transition:.15s}
.svc-tab:last-child{border-right:none}
.svc-tab.on{background:var(--navy);color:#fff}
.svc-card{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:12px;margin-bottom:8px;display:flex;flex-direction:column;gap:7px}
.svc-card.overdue{border-left:4px solid #dc2626;background:#fef2f2}
.svc-card.due-soon{border-left:4px solid #d97706;background:#fef9ee}
.svc-card.on-track{border-left:4px solid #059669;background:#f0fdf4}
.svc-card.no-date{border-left:4px solid var(--sub);opacity:.7}

/* ── MOBILE RESPONSIVE ─────────────────────────────── */
@media(max-width:480px){
  /* Bottom nav for phone */
  .tabs{
    position:fixed;bottom:0;left:0;right:0;
    background:var(--surf);
    border-top:1px solid var(--brd);
    border-bottom:none;
    display:flex;overflow-x:auto;
    z-index:90;padding-bottom:env(safe-area-inset-bottom);
    box-shadow:0 -2px 8px rgba(0,0,0,.08);
  }
  .tab{
    flex:1;flex-shrink:0;
    padding:6px 4px 8px;
    font-size:8px;font-weight:700;
    display:flex;flex-direction:column;align-items:center;gap:2px;
    min-width:48px;
  }
  /* Add bottom padding so content clears bottom nav */
  .panels{padding-bottom:80px;}
  /* Header more compact */
  .hdr{padding:8px 12px 6px;}
  .hdr-title{font-size:13px;}
  /* Bigger tap targets */
  button,.btn,.cta-btn{min-height:44px;}
  .tab{min-height:52px;}
  /* Cards full width */
  .grid{grid-template-columns:1fr !important;}
  /* Briefing compact */
  #brief-stats{grid-template-columns:repeat(2,1fr) !important;}
  /* Service modal full screen on phone */
  #svc-log-bg > div{
    max-height:100vh !important;
    border-radius:0 !important;
    padding-bottom:env(safe-area-inset-bottom);
  }
  /* Offline banner */
  #offline-banner{display:flex !important;}
}

/* Offline indicator */
#offline-banner{
  display:none;
  position:fixed;top:0;left:0;right:0;
  background:#dc2626;color:#fff;
  padding:6px 12px;
  font-size:11px;font-weight:700;
  text-align:center;z-index:200;
  align-items:center;justify-content:center;gap:6px;
}
#offline-banner.active{display:flex;}

/* Quick action fab - phone only */
.fab{
  display:none !important;
}
@media(max-width:480px){
  .fab{display:flex !important;}
  #offline-banner{display:none;}
  #offline-banner.active{display:flex !important;}
}

/* Larger modal inputs on mobile */
@media(max-width:480px){
  #svc-atp{font-size:18px !important;padding:12px !important;}
  #svc-notes{font-size:14px !important;min-height:80px;}
  .svc-card{padding:14px;}
  .svc-card button{min-height:44px;font-size:12px !important;}
}
.flt-sel{background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:5px 7px;color:var(--txt);font-size:10px;outline:none;cursor:pointer;font-family:inherit;max-width:130px}

/* ── NEW OUTCOME BUTTON STYLES ─────────────────────── */
.obtn-green  { background:#ecfdf5 !important; border-color:#6ee7b7 !important; color:#059669 !important; }
.obtn-blue   { background:#ecfeff !important; border-color:#67e8f9 !important; color:#0891b2 !important; }
.obtn-yellow { background:#fffbeb !important; border-color:#fcd34d !important; color:#d97706 !important; }
.obtn-gray   { background:#f8fafc !important; border-color:#cbd5e1 !important; color:#475569 !important; }
.obtn-orange { background:#fff7ed !important; border-color:#fdba74 !important; color:#c2410c !important; }
.obtn-red    { background:#fef2f2 !important; border-color:#fca5a5 !important; color:#dc2626 !important; }
.obtn-dark   { background:#f1f5f9 !important; border-color:#94a3b8 !important; color:#374151 !important; }
.obtn-teal   { background:#f0fdfa !important; border-color:#5eead4 !important; color:#0d9488 !important; }
.obtn.on.obtn-green  { background:#059669 !important; color:#fff !important; border-color:#059669 !important; }
.obtn.on.obtn-blue   { background:#0891b2 !important; color:#fff !important; border-color:#0891b2 !important; }
.obtn.on.obtn-yellow { background:#d97706 !important; color:#fff !important; border-color:#d97706 !important; }
.obtn.on.obtn-gray   { background:#475569 !important; color:#fff !important; border-color:#475569 !important; }
.obtn.on.obtn-orange { background:#c2410c !important; color:#fff !important; border-color:#c2410c !important; }
.obtn.on.obtn-red    { background:#dc2626 !important; color:#fff !important; border-color:#dc2626 !important; }
.obtn.on.obtn-dark   { background:#374151 !important; color:#fff !important; border-color:#374151 !important; }
.obtn.on.obtn-teal   { background:#0d9488 !important; color:#fff !important; border-color:#0d9488 !important; }

/* ── TYPE BUTTON SELECTED ──────────────────────────── */
.mtype-on { background:var(--navy) !important; color:#fff !important; border-color:var(--navy) !important; }

/* ── REASON CHIP ───────────────────────────────────── */
.reason-chip { font-size:10px; padding:4px 8px; border-radius:12px; border:1px solid var(--brd2); background:var(--surf); color:var(--sub); cursor:pointer; font-family:inherit; transition:.15s; font-weight:500; }
.reason-chip.on { background:var(--navy); color:#fff; border-color:var(--navy); }

/* ── QUEUE MODE ────────────────────────────────────── */
#queue-bg { display:none; }
#queue-bg.on { display:flex !important; }
.qbtn { padding:10px 6px; border:none; border-radius:8px; font-size:11px; font-weight:700; cursor:pointer; font-family:inherit; }
.qbtn-green  { background:#ecfdf5; color:#059669; }
.qbtn-yellow { background:#fffbeb; color:#d97706; }
.qbtn-blue   { background:#ecfeff; color:#0891b2; }
.qbtn-red    { background:#fef2f2; color:#dc2626; }
.qbtn-gray   { background:#f8fafc; color:#475569; }

/* ── FUNNEL STAGE ──────────────────────────────────── */
.funnel-stage { display:flex; align-items:center; gap:8px; }
.funnel-bar-wrap { flex:1; height:6px; background:var(--brd2); border-radius:3px; }
.funnel-bar { height:6px; border-radius:3px; transition:width .4s; }
.funnel-label { font-size:11px; color:var(--txt); font-weight:600; white-space:nowrap; min-width:80px; }
.funnel-val { font-size:11px; color:var(--sub); white-space:nowrap; min-width:32px; text-align:right; }

/* ── GOAL PACING ───────────────────────────────────── */
.pace-on-track  { color:#059669; font-weight:700; }
.pace-behind    { color:#dc2626; font-weight:700; }
.pace-ahead     { color:#0891b2; font-weight:700; }

</style>
</head>
<body>
<div id="app">
  <header>
    <div class="logo">
      <div class="logo-icon">&#x1F9CA;</div>
      <div><div class="logo-name">Pinellas Ice Co</div>
        <div style="font-size:8px;color:var(--sub);letter-spacing:.04em">PROSPECT TOOL &bull; %%DATE%%</div>
      </div>
    </div>
    <div class="hchips">
      <span class="hc cb" onclick="setF('CALLBACK')">&#x25CF; %%NCB%% Callback</span>
      <span class="hc ht" onclick="setF('HOT')">&#x25CF; %%NHOT%% Hot</span>
      <span class="hc wm" onclick="setF('WARM')">&#x25CF; %%NWARM%% Warm</span>
      <span class="hc ic">&#x1F9CA; %%NCHRON%% Chronic</span>
      <span class="hc rt" onclick="sw('route')">&#x1F5FA; Route</span>
    </div>
    <div class="srch"><span class="si">&#x1F50D;</span>
      <input type="text" id="si" placeholder="Search&#x2026;" oninput="onS()" autocomplete="off" autocorrect="off" autocapitalize="off" spellcheck="false">
    </div>
  </header>
  <nav class="tabs" id="main-nav">
    <div class="tab on"  onclick="sw('today')"     ><span class="tab-icon">&#x1F4CA;</span><span class="tab-lbl">Home</span></div>
    <div class="tab"     onclick="sw('all')"        ><span class="tab-icon">&#x1F4CD;</span><span class="tab-lbl">Prospects</span></div>
    <div class="tab"     onclick="sw('route')"      ><span class="tab-icon">&#x1F5FA;</span><span class="tab-lbl">Route</span></div>
    <div class="tab"     onclick="sw('customers')"  ><span class="tab-icon">&#x1F91D;</span><span class="tab-lbl">Clients</span></div>
    <div class="tab"     onclick="sw('service')"    ><span class="tab-icon">&#x1F9FC;</span><span class="tab-lbl">Service</span></div>
  </nav>

  <!-- Offline banner -->
  <div id="offline-banner">&#x26A1; Offline &mdash; data saved locally, all features available</div>

  <!-- FAB: Service tab shortcut (phone only) -->
  <button id="svc-fab" onclick="sw('service')">&#x1F9FC;</button>

  <!-- TODAY -->
  <div class="panel on" id="p-today">

    <!-- ── KPI ROW ─────────────────────────────────── -->
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px" id="kpi-row">
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">MRR</div>
        <div style="font-size:20px;font-weight:800;color:var(--grn)" id="kpi-mrr">$0</div>
      </div>
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">Clients</div>
        <div style="font-size:20px;font-weight:800;color:var(--navy)" id="kpi-clients">0</div>
      </div>
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">Pipeline</div>
        <div style="font-size:20px;font-weight:800;color:var(--blu)" id="kpi-pipe">0</div>
      </div>
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">This Wk</div>
        <div style="font-size:20px;font-weight:800;color:var(--ora)" id="kpi-week">0</div>
      </div>
    </div>

    <!-- ── FUNNEL ────────────────────────────────────── -->
    <div class="dc" style="margin-bottom:12px;padding:12px 14px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
        <div style="font-weight:700;font-size:12px;color:var(--navy)">&#x1F4CA; Weekly Funnel</div>
        <div style="font-size:9px;color:var(--sub)" id="funnel-week-label"></div>
      </div>
      <div id="funnel-stages" style="display:flex;flex-direction:column;gap:6px"></div>
    </div>

    <!-- ── GOAL PACING ───────────────────────────────── -->
    <div class="dc" style="margin-bottom:12px;padding:12px 14px" id="goal-pacing-card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:700;font-size:12px;color:var(--navy)">&#x1F3AF; Goal Pacing</div>
        <button onclick="sw('data')" style="font-size:9px;padding:3px 8px;border:1px solid var(--brd);border-radius:6px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">Edit Goals</button>
      </div>
      <div id="goal-pacing-content"></div>
    </div>

    <!-- ── LOSS REASON BREAKDOWN ─────────────────────── -->
    <div class="dc" style="margin-bottom:12px;padding:12px 14px" id="loss-breakdown-card">
      <div style="font-weight:700;font-size:12px;color:var(--navy);margin-bottom:8px">&#x274C; Why You're Losing</div>
      <div id="loss-breakdown" style="display:flex;flex-direction:column;gap:4px"></div>
      <div id="loss-empty" style="font-size:11px;color:var(--sub);display:none">Log some walk-ins to see loss patterns.</div>
    </div>

    <!-- ── ACTIVE NURTURE (In Play due) ─────────────── -->
    <div style="margin-bottom:18px" id="nurture-section">
      <div class="tsect-hdr">
        <span>&#x1F7E1; In Play — Follow Up Due</span>
        <span class="tsect-sub">Warm prospects with a date you set. Do these first.</span>
      </div>
      <div class="grid" id="nurture-grid"></div>
      <div class="tempty" id="nurture-empty" style="display:none">No follow-ups due. &#x2713;</div>
    </div>

    <!-- ── BEST MOVE (cold targets) ─────────────────── -->
    <div id="t-actnow">
      <div class="tsect-hdr">
        <span>&#x1F534; Best Cold Targets Today</span>
        <span class="tsect-sub">Top 5 uncontacted. High score + callback priority.</span>
      </div>
      <div class="grid" id="tgrid-actnow"></div>
      <div class="tempty" id="empty-actnow" style="display:none">No urgent prospects.</div>
    </div>

  </div>

  <div class="panel" id="p-all">

    <!-- Preset filters -->
    <div style="display:flex;gap:6px;overflow-x:auto;padding-bottom:8px;margin-bottom:10px;-webkit-overflow-scrolling:touch">
      <button class="preset-btn on" onclick="setPreset('all')"      id="pre-all">All</button>
      <button class="preset-btn"    onclick="setPreset('actnow')"   id="pre-actnow">&#x1F534; Act Now</button>
      <button class="preset-btn"    onclick="setPreset('callback')" id="pre-callback">Callbacks</button>
      <button class="preset-btn"    onclick="setPreset('phone')"    id="pre-phone">Has Phone</button>
      <button class="preset-btn"    onclick="setPreset('chronic')"  id="pre-chronic">Chronic Ice</button>
      <button class="preset-btn"    onclick="setPreset('notyet')"   id="pre-notyet">Not Contacted</button>
      <button class="preset-btn"    onclick="setPreset('inplay')"   id="pre-inplay">&#x1F7E1; In Play</button>
      <button class="preset-btn"    onclick="setPreset('freshice')" id="pre-freshice">&#x1F525; Ice Viol.</button>
    </div>

    <!-- Filters row -->
    <div style="display:flex;gap:5px;flex-wrap:wrap;margin-bottom:8px;align-items:center">
      <select id="ac"     onchange="populateCityFilter();rA()" class="flt-sel"><option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option><option>Citrus</option><option>Hernando</option><option>Polk</option><option>Sumter</option></select>
      <select id="ac-city" onchange="rA()" class="flt-sel"><option value="">All Cities</option></select>
      <select id="ap"     onchange="rA()" class="flt-sel"><option value="">All Priorities</option><option>CALLBACK</option><option>HOT</option><option>WARM</option><option>WATCH</option></select>
      <select id="as_"    onchange="rA()" class="flt-sel">
        <option value="">All Status</option>
        <option value="not_contacted">Not Contacted</option>
        <option value="in_play">In Play</option>
        <option value="intro_set">Intro Set</option>
        <option value="not_now">Not Now</option>
        <option value="signed">Signed</option>
        <option value="dead">Dead</option>
      </select>
      <button onclick="enterQueueMode()" style="font-size:10px;padding:5px 10px;border:1px solid var(--blu);border-radius:6px;background:#0a84ff22;color:var(--blu);cursor:pointer;font-family:inherit;font-weight:700;flex-shrink:0">&#x25B6; Queue</button>
      <button onclick="clearFilters()" style="font-size:10px;padding:5px 8px;border:1px solid var(--brd);border-radius:6px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit;flex-shrink:0">Clear</button>
      <span class="fcnt" id="acnt"></span>
    </div>

    <div class="grid" id="agrid"></div>
    <div class="tempty" id="a-empty" style="display:none">
      <div class="ei">&#x1F50D;</div>
      <div>No prospects match these filters</div>
    </div>
  </div>

  <!-- QUEUE MODE OVERLAY -->
  <div id="queue-bg" style="display:none;position:fixed;inset:0;background:var(--bg);z-index:90;flex-direction:column">
    <div style="background:var(--navy);padding:12px 16px;display:flex;align-items:center;gap:10px;flex-shrink:0">
      <button onclick="exitQueueMode()" style="border:none;background:rgba(255,255,255,.15);color:#fff;border-radius:8px;padding:6px 12px;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">&#x2715; Exit</button>
      <div style="flex:1;text-align:center;font-size:11px;font-weight:700;color:rgba(255,255,255,.8)" id="queue-progress"></div>
      <button onclick="queueNext()" style="border:none;background:var(--ora);color:#fff;border-radius:8px;padding:6px 12px;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">Skip &#x2192;</button>
    </div>
    <div style="flex:1;overflow-y:auto;padding:14px 16px">
      <div id="queue-card-wrap"></div>
    </div>
    <div style="padding:12px 16px;background:var(--surf);border-top:1px solid var(--brd2);display:grid;grid-template-columns:repeat(3,1fr);gap:8px;flex-shrink:0" id="queue-actions">
      <button class="qbtn qbtn-green"  onclick="queueLog('intro_set')"  >&#x1F4C5; Intro Set</button>
      <button class="qbtn qbtn-yellow" onclick="queueLog('in_play')"    >&#x1F7E1; In Play</button>
      <button class="qbtn qbtn-blue"   onclick="queueLog('no_contact')" >&#x1F6AA; No Contact</button>
      <button class="qbtn qbtn-red"    onclick="queueLog('not_now')"    >&#x274C; Not Now</button>
      <button class="qbtn qbtn-gray"   onclick="queueLog('voicemail')"  >&#x1F4F2; Voicemail</button>
      <button class="qbtn qbtn-gray"   onclick="queueLog('dead')"       >&#x26AB; Dead</button>
    </div>
  </div>

   class="panel" id="p-route">
    <div style="display:flex;flex-direction:column;height:calc(100vh - 120px);gap:8px;overflow:hidden">

      <!-- Plan My Day header -->
      <div class="dc" style="flex-shrink:0;padding:12px">
        <div style="font-weight:800;font-size:14px;color:var(--navy);margin-bottom:10px">&#x1F4C5; Plan My Day</div>

        <!-- Row 1: Start ZIP + Time budget -->
        <div style="display:flex;gap:6px;margin-bottom:8px">
          <div style="flex:1">
            <div style="font-size:9px;color:var(--sub);font-weight:600;margin-bottom:3px">START ZIP</div>
            <input id="rzip" type="text" placeholder="e.g. 33701" maxlength="5"
              style="width:100%;padding:7px 8px;border:1px solid var(--brd);border-radius:7px;font-size:12px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none"
              oninput="rRoute()">
          </div>
          <div style="flex:1">
            <div style="font-size:9px;color:var(--sub);font-weight:600;margin-bottom:3px">TIME AVAILABLE</div>
            <select id="rtime" onchange="planMyDay()"
              style="width:100%;padding:7px 8px;border:1px solid var(--brd);border-radius:7px;font-size:12px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
              <option value="0">Manual (add stops)</option>
              <option value="2">2 hours (~4 stops)</option>
              <option value="3">3 hours (~6 stops)</option>
              <option value="4">4 hours (~8 stops)</option>
              <option value="6">Half day (~10 stops)</option>
              <option value="8">Full day (~14 stops)</option>
            </select>
          </div>
        </div>

        <!-- Row 2: Filters -->
        <div style="display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap">
          <select id="rc" onchange="rRoute()"
            style="flex:1;min-width:90px;padding:6px 7px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
            <option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option>
          </select>
          <select id="rp" onchange="rRoute()"
            style="flex:1;min-width:90px;padding:6px 7px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
            <option value="">All Priorities</option><option>CALLBACK</option><option>HOT</option><option>WARM</option>
          </select>
          <select id="rrad" onchange="rRoute()"
            style="flex:1;min-width:80px;padding:6px 7px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
            <option value="5">5 mi</option><option value="8" selected>8 mi</option>
            <option value="12">12 mi</option><option value="20">20 mi</option><option value="999">Any</option>
          </select>
        </div>

        <!-- Row 3: Action buttons -->
        <div style="display:flex;gap:6px">
          <button onclick="planMyDay()"
            style="flex:2;padding:8px;border:none;border-radius:8px;background:var(--navy);color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">
            &#x26A1; Build Optimal Route
          </button>
          <button onclick="optimizeRoute()"
            style="flex:1;padding:8px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;font-family:inherit">
            Sort
          </button>
          <button onclick="clearRoute()"
            style="flex:1;padding:8px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:10px;cursor:pointer;font-family:inherit">
            Clear
          </button>
        </div>

        <div style="font-size:9px;color:var(--sub);margin-top:6px" id="rhint">Enter start ZIP (or tap Start 📍 on any business). Set time budget. Tap Build Optimal Route.</div>
      </div>

      <!-- Two-column: list left, route+map right -->
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;flex:1;min-height:0;overflow:hidden">

        <!-- Prospect list -->
        <div style="display:flex;flex-direction:column;min-height:0">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px" id="rlist-cnt"></div>
          <div id="rlist" style="overflow-y:auto;flex:1;display:flex;flex-direction:column;gap:3px"></div>
        </div>

        <!-- Route stops + map -->
        <div style="display:flex;flex-direction:column;min-height:0;gap:8px">

          <!-- Built route -->
          <div id="day-route" style="display:none;background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:10px;flex-shrink:0;max-height:50%;overflow-y:auto">
            <div style="font-weight:700;font-size:11px;color:var(--navy);margin-bottom:2px">
              &#x1F4CD; Route — <span id="stopcnt">0</span> stops
              <span id="route-mi" style="font-size:9px;color:var(--sub);font-weight:400"></span>
            </div>
            <div id="route-time-est" style="font-size:9px;color:var(--sub);margin-bottom:7px"></div>
            <div id="day-stops"></div>
            <div style="display:flex;gap:5px;margin-top:8px">
              <button class="route-btn" onclick="openMaps()" style="flex:2">Open in Google Maps &#x2192;</button>
            </div>
          </div>

          <!-- Map -->
          <div class="map-area" id="map-area" style="flex:1;min-height:180px;border-radius:10px;overflow:hidden">
            <div class="map-empty">
              <div style="font-size:32px;margin-bottom:8px">&#x1F5FA;&#xFE0F;</div>
              <div style="font-size:11px">Enter ZIP + time budget<br>and tap Build Optimal Route</div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- SERVICE -->
  <div class="panel" id="p-service">

    <!-- Revenue forecast strip -->
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:12px" id="svc-forecast"></div>

    <!-- At-risk alert -->
    <div id="svc-atrisk" style="margin-bottom:12px"></div>

    <!-- Tabs: Calendar | Route | Reports | Referrals -->
    <div style="display:flex;gap:0;margin-bottom:12px;border:1px solid var(--brd);border-radius:8px;overflow:hidden;flex-wrap:wrap">
      <button class="svc-tab on" onclick="setSvcTab('cal')"      id="svct-cal">&#x1F4C5; Calendar</button>
      <button class="svc-tab"    onclick="setSvcTab('route')"    id="svct-route">&#x1F5FA; Route</button>
      <button class="svc-tab"    onclick="setSvcTab('reports')"  id="svct-reports">&#x1F4CB; Reports</button>
      <button class="svc-tab"    onclick="setSvcTab('tutorials')" id="svct-tutorials">&#x1F4D6; Tutorials</button>
      <button class="svc-tab"    onclick="setSvcTab('refs')"     id="svct-refs">&#x1F91D; Referrals</button>
    </div>

    <!-- Calendar view -->
    <div id="svc-cal">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Recurring clients due for service. Tap to log or reschedule.</div>
      <div id="svc-cal-list"></div>
    </div>

    <!-- Service route builder -->
    <div id="svc-route" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Build a service route for clients due this week, clustered by geography.</div>
      <div style="display:flex;gap:6px;margin-bottom:8px">
        <select id="svc-week" onchange="renderServiceRoute()"
          style="flex:1;padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="0">This week</option>
          <option value="7">Next week</option>
          <option value="14">In 2 weeks</option>
          <option value="-7">Overdue</option>
        </select>
        <button onclick="buildServiceRoute()"
          style="padding:7px 12px;border:none;border-radius:7px;background:var(--navy);color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">
          &#x26A1; Build Route
        </button>
      </div>
      <div id="svc-route-list"></div>
      <div id="svc-route-map-btn" style="display:none;margin-top:8px">
        <button onclick="openServiceMaps()" class="route-btn">Open in Google Maps &#x2192;</button>
      </div>
    </div>

    <!-- Service reports -->
    <div id="svc-reports" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Generate a printable service report for any client visit.</div>
      <select id="svc-report-client" onchange="loadReportClient()"
        style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;margin-bottom:8px">
        <option value="">Select a client...</option>
      </select>
      <div id="svc-report-preview"></div>
    </div>

    <!-- Tutorials -->
    <div id="svc-tutorials" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Step-by-step cleaning guides for each machine brand. Select brand and procedure type.</div>
      <div style="display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap">
        <select id="tut-brand" onchange="renderTutorial()"
          style="flex:1;min-width:120px;padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="">Select brand...</option>
          <option>Manitowoc</option><option>Hoshizaki</option><option>Scotsman</option>
          <option>Ice-O-Matic</option><option>Follett</option><option>Cornelius</option>
        </select>
        <select id="tut-type" onchange="renderTutorial()"
          style="flex:1;min-width:120px;padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="deep_clean">&#x1F9FC; Deep Clean</option>
          <option value="maintenance_60">&#x1F527; 60-Day Maintenance</option>
          <option value="atp_protocol">&#x1F4CA; ATP Testing Guide</option>
        </select>
      </div>
      <div id="tut-content"></div>
    </div>

    <!-- Referrals -->
    <div id="svc-refs" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Track who referred new clients. Your best customers are your best salespeople.</div>
      <div id="svc-refs-list"></div>
    </div>

  </div>

  <!-- PIPELINE -->
  <div class="panel" id="p-customers">
    <div class="fbar">
      <select id="cust-status" onchange="rCust()">
        <option value="">All Customers</option>
        <option value="customer_recurring">Recurring</option>
        <option value="customer_intro">Intro ($99 first visit)</option>
        <option value="customer_once">One-Time ($249)</option>
        <option value="quoted">Quoted - Pending</option>
        <option value="churned">Churned</option>
      </select>
      <span class="fcnt" id="cust-cnt"></span>
    </div>

    <!-- Revenue summary bar -->
    <div id="cust-summary" style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px">
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Monthly Recurring</div>
        <div style="font-size:22px;font-weight:800;color:var(--grn)" id="mrr-val">$0</div>
        <div class="fl">MRR</div>
      </div>
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Active Customers</div>
        <div style="font-size:22px;font-weight:800;color:var(--navy)" id="cust-count">0</div>
        <div class="fl">Accounts</div>
      </div>
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Annual Run Rate</div>
        <div style="font-size:22px;font-weight:800;color:var(--blu)" id="arr-val">$0</div>
        <div class="fl">ARR</div>
      </div>
    </div>

    <div id="cust-list"></div>

    <div class="tempty" id="cust-empty" style="display:none">
      <div class="ei">&#x1F91D;</div>
      <div style="font-weight:600;margin-bottom:4px">No customers yet</div>
      <div style="font-size:11px;color:var(--sub)">When you close a deal, open any prospect card and mark them as Won.</div>
    </div>
  </div>

  <!-- DATA -->
  <div<div class="panel" id="p-data">
    <div class="dc" style="margin-bottom:12px">
      <div class="dct">&#x1F4E6; Dataset &mdash; %%DATE%%</div>
      <div class="ds"><span>Total prospects</span><span class="dsv">%%TOTAL%%</span></div>
      <div class="ds"><span>Source</span><span class="dsv">FL DBPR District 3 (live)</span></div>
      <div class="ds"><span>Auto-rebuilds</span><span class="dsv">Weekly via GitHub Actions</span></div>
      <div class="ds"><span>Chronic ice offenders</span><span class="dsv">%%NCHRON%%</span></div>
      <div class="ds"><span>Callback urgent</span><span class="dsv">%%NCB%%</span></div>
      <div class="ds"><span>With phone number</span><span class="dsv">%%NPHONE%%</span></div>
    </div>

    <div class="dc" id="goals-section">
      <div class="dct">&#x1F3AF; Goals &amp; Targets</div>
      <div style="font-size:10px;color:var(--sub);margin-bottom:10px">Set your targets. The home dashboard tracks progress automatically.</div>
      <div style="display:flex;flex-direction:column;gap:8px">
        <div style="background:#f5f8fa;border-radius:8px;padding:10px">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">&#x1F91D; Client Goal</div>
          <div style="display:flex;gap:6px;align-items:center;margin-bottom:4px">
            <input id="goal-clients" type="number" min="1" max="500" placeholder="10"
              style="width:70px;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            <span style="font-size:11px;color:var(--sub)">recurring clients</span>
          </div>
          <div style="display:flex;gap:6px;align-items:center">
            <input id="goal-deadline" type="date"
              style="flex:1;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:var(--txt);outline:none">
            <span style="font-size:11px;color:var(--sub)">deadline</span>
          </div>
        </div>
        <div style="background:#f5f8fa;border-radius:8px;padding:10px">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">&#x1F4B0; Revenue Goal</div>
          <div style="display:flex;gap:6px;align-items:center">
            <input id="goal-mrr" type="number" min="100" step="100" placeholder="1490"
              style="width:80px;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            <span style="font-size:11px;color:var(--sub)">target MRR ($)</span>
          </div>
        </div>
        <div style="background:#f5f8fa;border-radius:8px;padding:10px">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">&#x1F4DE; Daily Activity</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px">
            <div>
              <div style="font-size:9px;color:var(--sub);margin-bottom:3px">Walk-ins / day</div>
              <input id="goal-walkins" type="number" min="0" max="30" placeholder="2"
                style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            </div>
            <div>
              <div style="font-size:9px;color:var(--sub);margin-bottom:3px">Calls / day</div>
              <input id="goal-calls" type="number" min="0" max="50" placeholder="5"
                style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            </div>
          </div>
        </div>
        <button onclick="saveGoals()" style="width:100%;padding:10px;background:var(--navy);color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit">Save Goals</button>
      </div>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x1F504; Rebuild With New Data</div>
      <div class="ibox">
        <b>Step 1 &mdash; Large XLS/XLSX?</b> Convert first:<br>
        <code>python convert_to_csv.py yourfile.xlsx</code><br><br>
        <b>Step 2 &mdash; Rebuild:</b><br>
        <code>python build.py file1.csv file2.csv</code><br>
        Call log and phone numbers carry over automatically.
      </div>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x1F4E7; Monday Briefing</div>
      <div style="font-size:11px;color:var(--sub);line-height:1.6">Automated weekly email via GitHub Actions. Requires <code>RESEND_API_KEY</code> and <code>BRIEFING_EMAIL</code> in GitHub Secrets.</div>
    </div>
  </div>

  <div id="mbg" onclick="closeM(event)">
  <div id="modal">
    <div class="mh" style="position:sticky;top:0;z-index:10;background:var(--surf);border-bottom:1px solid var(--brd2);padding:8px 16px;display:flex;align-items:center;justify-content:space-between">
      <div style="width:36px;height:4px;background:var(--brd);border-radius:2px"></div>
      <button onclick="closeMForce()" style="border:none;background:var(--brd2);border-radius:50%;width:28px;height:28px;font-size:14px;color:var(--sub);cursor:pointer;display:flex;align-items:center;justify-content:center;font-family:inherit;flex-shrink:0">&#x2715;</button>
    </div>
    <div class="mname" id="mn"></div>
    <div class="mloc"  id="ml"></div>
    <div class="mphsec">
      <div class="mphl">CONTACT</div>
      <div class="mphnum" id="mph"></div>
      <div class="mphacts" id="mpa"></div>
      <div id="mhrs" class="mhours"></div>
      <div id="mrat" style="font-size:9px;color:#555;margin-top:3px"></div>
    </div>
    <div id="mice" style="margin-bottom:10px"></div>
    <div style="margin-bottom:10px">
      <div class="msect">PITCH GUIDE</div>
      <div style="display:flex;gap:5px;margin-bottom:6px">
        <button id="btn-phone" onclick="setPitchMode('phone')"
          style="flex:1;padding:6px;border-radius:7px;border:1px solid var(--blu);background:#0a84ff22;color:var(--blu);font-size:10px;font-weight:700;cursor:pointer">
          &#x1F4DE; Phone Call
        </button>
        <button id="btn-walkin" onclick="setPitchMode('walkin')"
          style="flex:1;padding:6px;border-radius:7px;border:1px solid var(--brd);background:transparent;color:var(--sub);font-size:10px;font-weight:700;cursor:pointer">
          &#x1F6B6; Cold Walk-In
        </button>
      </div>
      <div class="pitch" id="mpitch"></div>
      <div class="pitch" id="mwalkin" style="display:none;border-color:#e2e8f0;color:var(--sub2)"></div>
      <div style="margin-top:8px">
        <div class="msect">OBJECTION HANDLERS</div>
        <div id="mobjections" style="display:flex;flex-direction:column;gap:5px"></div>
      </div>
    </div>
    <div style="margin-bottom:10px">
      <div class="msect">LOG THIS CONTACT</div>

      <!-- Interaction type -->
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px">
        <button id="mtype-walkin" onclick="setType('walkin')"
          style="padding:7px;border-radius:8px;border:2px solid var(--brd);background:transparent;color:var(--sub);font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;transition:.15s">
          &#x1F6B6; Walk-In
        </button>
        <button id="mtype-call" onclick="setType('call')"
          style="padding:7px;border-radius:8px;border:2px solid var(--brd);background:transparent;color:var(--sub);font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;transition:.15s">
          &#x1F4DE; Phone Call
        </button>
      </div>

      <!-- Outcome buttons -->
      <div class="ogrid">
        <button class="obtn obtn-green"  data-o="signed"     onclick="selO('signed')">&#x2705; Signed</button>
        <button class="obtn obtn-blue"   data-o="intro_set"  onclick="selO('intro_set')">&#x1F4C5; Intro Set</button>
        <button class="obtn obtn-yellow" data-o="in_play"    onclick="selO('in_play')">&#x1F7E1; In Play</button>
        <button class="obtn obtn-gray"   data-o="no_contact" onclick="selO('no_contact')">&#x1F6AA; No Contact</button>
        <button class="obtn obtn-orange" data-o="voicemail"  onclick="selO('voicemail')">&#x1F4F2; Voicemail</button>
        <button class="obtn obtn-red"    data-o="not_now"    onclick="selO('not_now')">&#x274C; Not Now</button>
        <button class="obtn obtn-dark" data-o="dead" onclick="selO('dead')" style="grid-column:1/-1">&#x26AB; Dead &mdash; Wrong fit / hard no / corporate</button>
        <button class="obtn obtn-teal" data-o="service_done" onclick="selO('service_done')" style="grid-column:1/-1">&#x1F9FC; Service Done</button>
      </div>

      <!-- Reason picker - shown only for not_now and in_play -->
      <div id="reason-wrap" style="display:none;margin-top:8px">
        <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">Reason</div>
        <div style="display:flex;flex-wrap:wrap;gap:4px" id="reason-grid"></div>
      </div>

      class="ntxt" id="mnotes" rows="2" placeholder="Notes&#x2026;"></textarea>
      <div style="display:flex;gap:6px;align-items:center;margin-top:6px;padding:6px 8px;background:#f5f8fa;border-radius:7px">
        <span style="font-size:9px;color:var(--sub);white-space:nowrap;font-weight:600">&#x1F4C5; Follow-up date:</span>
        <input type="date" id="mfollowup" style="flex:1;padding:4px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:var(--txt);outline:none">
      </div>
      <button class="btn blog" style="width:100%;margin-top:6px;padding:8px" onclick="saveL()">Save Log Entry</button>
    </div>
    <!-- CLOSE DEAL -->
    <div style="margin-bottom:10px">
      <div class="msect">CLOSE DEAL</div>
      <!-- Intro offer - most prominent -->
      <button onclick="markWon('customer_intro')"
        style="width:100%;padding:10px;border:2px solid var(--ora);border-radius:9px;background:#fff7f5;color:var(--ora);font-weight:800;font-size:12px;cursor:pointer;font-family:inherit;margin-bottom:6px">
        &#x1F525; Intro Offer  -  $99 first visit<br><span style="font-size:10px;font-weight:400">Try us once, no commitment. Recurring starts after.</span>
      </button>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:6px">
        <button onclick="markWon('customer_recurring')"
          style="padding:10px 6px;border:2px solid var(--grn);border-radius:9px;background:#ecfdf5;color:#059669;font-weight:700;font-size:11px;cursor:pointer;font-family:inherit">
          &#x1F4B0; Won  -  Recurring<br><span style="font-size:10px;font-weight:400" id="mwon-monthly"></span>
        </button>
        <button onclick="markWon('customer_once')"
          style="padding:10px 6px;border:2px solid var(--blu);border-radius:9px;background:#eff6ff;color:var(--blu);font-weight:700;font-size:11px;cursor:pointer;font-family:inherit">
          &#x1F9FC; Won  -  One-Time<br><span style="font-size:10px;font-weight:400" id="mwon-onetime"></span>
        </button>
      </div>
      <button onclick="markWon('churned')"
        style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:8px;background:transparent;color:var(--sub);font-size:10px;cursor:pointer;font-family:inherit">
        Mark as Lost / Churned
      </button>
    </div>
    <!-- CONTACTS -->
    <div style="margin-bottom:10px">
      <div class="msect">CONTACTS &amp; INTEL</div>

      <!-- Current vendor field -->
      <div style="margin-bottom:8px;padding:8px;background:#fef9ee;border:1px solid #fde68a;border-radius:7px">
        <div style="font-size:9px;font-weight:700;color:#92400e;margin-bottom:4px">&#x1F575; CURRENT ICE VENDOR</div>
        <div style="display:flex;gap:5px;align-items:center">
          <input id="mc-vendor" type="text" placeholder="e.g. Ecolab, local guy, staff cleans it..."
            style="flex:1;padding:6px;border:1px solid #fde68a;border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:var(--txt);outline:none">
          <button onclick="saveVendor()" style="padding:6px 10px;border:none;border-radius:6px;background:#92400e;color:#fff;font-size:10px;font-weight:600;cursor:pointer;font-family:inherit">Save</button>
        </div>
        <div id="mc-vendor-display" style="font-size:10px;color:#92400e;margin-top:4px;font-style:italic"></div>
      </div>

      <div id="mcontacts" style="margin-bottom:8px"></div>
      <div style="display:flex;gap:5px;margin-bottom:4px">
        <input id="mc-name" type="text" placeholder="Contact name" style="flex:2;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
        <select id="mc-role" style="flex:1;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="owner">Owner</option>
          <option value="gm">GM</option>
          <option value="manager">Manager</option>
          <option value="chef">Chef</option>
          <option value="staff">Staff</option>
        </select>
      </div>
      <input id="mc-phone" type="tel" placeholder="Direct phone (optional)" style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;margin-bottom:4px">
      <button onclick="addContact()" style="width:100%;padding:7px;border:none;border-radius:7px;background:var(--navy);color:#fff;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit">+ Save Contact</button>
    </div>
    <div style="margin-bottom:10px"><div class="msect">INSPECTION DETAILS</div><div class="mfacts" id="mfacts"></div></div>
    <div id="mhs" style="display:none;margin-bottom:10px">
      <div class="msect">CONTACT HISTORY</div><div class="mhist" id="mhist"></div>
    </div>
  </div>
</div>
<div id="toast"></div>

<script>
const P=%%DATA%%;
const PHONES=%%PHONES%%;
const ZIPS=%%ZIPS%%;

const PITCHES={
  callback:n=>`"Hi, is this <b>${n}</b>? I'm [Your Name] from Pinellas Ice Co  -  we clean and sanitize commercial ice machines locally. I'm reaching out because you recently had some health inspection issues. Ice machines are almost always part of callback inspections. We can get yours cleaned and documented before the inspector returns  -  usually takes about 2 hours. <b>Do you have time this week?</b>"`,
  overdue_urgent:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We specialize in ice machine cleaning in Pinellas and Hillsborough. I know dealing with inspection issues is stressful  -  we clean and document quickly and give you a service report you can show the inspector. <b>Can we schedule something this week?</b>"`,
  overdue:n=>`"Hi, is this <b>${n}</b>? This is [Your Name] from Pinellas Ice Co  -  commercial ice machine cleaning. Ice machines are one of the most flagged items when inspectors return. <b>Is this a good time to talk?</b>"`,
  pre_hot:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. Your next routine health inspection could be coming up very soon  -  ice machines are one of the most flagged items. We do a full clean and sanitize with a service report. <b>Want to get squared away before the inspector shows up?</b>"`,
  pre_warm:n=>`"Hi, is this <b>${n}</b>? My name is [Your Name] from Pinellas Ice Co. You have some lead time before your next inspection  -  perfect window to get a service on the books. <b>Would you be open to scheduling this month?</b>"`,
  high_risk:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We clean ice machines for restaurants in the area and based on your inspection history I wanted to reach out specifically. <b>Has anyone talked to you recently about your ice machine maintenance?</b>"`,
  routine:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We clean and sanitize commercial ice machines  -  FDA recommends service every 6 months. We're in your area. <b>Is this something that would be useful for you?</b>"`,
};
const WALKIN={
  callback:n=>`<b>Opening (to staff/gatekeeper):</b><br>&#x201C;Hey, is the manager around for one second? I&#39;m [Name] from Pinellas Ice Co. I&#39;ll be quick.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi  -  briefly: you had a callback inspection recently, and ice machines are almost always what inspectors target on the return visit. I clean and certify them, takes about 90 minutes, you get a dated service report to show the inspector. I&#39;m in the area this week. <b>Want me to take a look today?</b>&#x201D;<br><br><i>If hesitant:</i> &#x201C;No pressure  -  I can open it up and tell you what an inspector would flag. Five minutes, free.&#x201D;`,
  overdue_urgent:n=>`<b>Opening:</b><br>&#x201C;Hey, is the manager in? I&#39;ll be really quick.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi, I&#39;m [Name] from Pinellas Ice Co. Your location came up because you have an open inspection issue. Ice machines get hit hard on follow-up visits. I can clean and document yours before they return. <b>Do you have 90 minutes this week?</b>&#x201D;`,
  pre_hot:n=>`<b>Opening:</b><br>&#x201C;Hi, is the owner or manager available? Quick question about your health inspection.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co. Based on your inspection history your next one is probably coming up soon. Ice machines are one of the top-cited items  -  mold, mineral scale, slime. I service and document them so you&#39;re covered. <b>Can I show you what inspectors typically look for?</b>&#x201D;<br><br><i>Close:</i> &#x201C;I can come back at a time that works. First visit is $99 and I leave you a full compliance report.&#x201D;`,
  pre_warm:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager around for a minute?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co  -  I clean commercial ice machines. You&#39;ve got some time before your next inspection, which is actually the perfect window  -  no last-minute stress. <b>Is ice machine maintenance on your radar?</b>&#x201D;`,
  high_risk:n=>`<b>Opening:</b><br>&#x201C;Hi, is the owner or manager in? I have some information about your inspection history.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi, I&#39;m [Name] from Pinellas Ice Co. Your location has had some ice-related violations flagged. This is exactly what inspectors look for on follow-ups. <b>When was your machine last serviced?</b>&#x201D;`,
  routine:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager available?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co  -  I clean commercial ice machines. FDA recommends every 6 months. Inspectors are specifically trained to look at machine interiors now. <b>When was yours last serviced?</b>&#x201D;<br><br><i>Let their answer guide you. If they don&#39;t know  -  that&#39;s your opening.</i>`,
  overdue:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager around?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co. Your location came up because of some past inspection flags. Ice machines are one of the most targeted items when inspectors return. <b>Do you mind if I take a quick look at yours?</b>&#x201D;`,
};
const OI={
  // New outcomes
  signed:'Signed ✓', intro_set:'Intro Set', in_play:'In Play',
  no_contact:'No Contact', not_now:'Not Now', dead:'Dead',
  voicemail:'Voicemail', service_done:'Service Done ✓',
  // Legacy compat (existing log entries)
  customer_recurring:'Signed ✓', customer_once:'Signed ✓', customer_intro:'Intro Set',
  follow_up:'In Play', interested:'In Play', scheduled:'Intro Set',
  not_interested:'Not Now', no_answer:'No Contact', quoted:'In Play', churned:'Dead',
};
const OI_COLOR={
  signed:'#059669', intro_set:'#0891b2', in_play:'#d97706',
  no_contact:'#64748b', not_now:'#dc2626', dead:'#374151',
  voicemail:'#7c3aed', service_done:'#059669',
  customer_recurring:'#059669', customer_once:'#059669', customer_intro:'#0891b2',
  follow_up:'#d97706', interested:'#d97706', scheduled:'#0891b2',
  not_interested:'#dc2626', no_answer:'#64748b', quoted:'#d97706', churned:'#374151',
};
// Normalize legacy outcomes to new system for filtering
function normO(o){
  const map={customer_recurring:'signed',customer_once:'signed',customer_intro:'intro_set',
    follow_up:'in_play',interested:'in_play',scheduled:'intro_set',
    not_interested:'not_now',no_answer:'no_contact',quoted:'in_play',churned:'dead'};
  return map[o]||o;
}
const REASONS={
  already_has_service:'Already has service',
  cleans_themselves:'Cleans themselves',
  too_expensive:'Too expensive',
  landlord_handles:'Landlord handles it',
  need_partner:'Needs partner / corporate',
  just_passed:'Just passed inspection',
  no_authority:'No authority (manager)',
  franchise_corp:'Franchise / corporate contract',
  seasonal:'Seasonal / not ready',
  no_reason:'No reason given',
};
// Outcomes that show reason picker
const REASON_OUTCOMES=new Set(['not_now','in_play']);
const PO={CALLBACK:0,HOT:1,WARM:2,WATCH:3,LATER:4};
const PC={CALLBACK:'var(--cb)',HOT:'var(--hot)',WARM:'var(--warm)',WATCH:'var(--watch)',LATER:'var(--sub)'};
const ICN={V14:'food contact surfaces (V14)',V22:'non-PHF surfaces (V22)',V50:'food contact surfaces (V50)',V37:'equipment repair (V37)',V23:'utensil sanitation (V23)'};

// Objection handlers -- proven B2B field responses

// ── MACHINE TUTORIALS ─────────────────────────────────────────────────────────
const MACHINE_BRANDS = ['Manitowoc','Hoshizaki','Scotsman','Ice-O-Matic','Follett','Cornelius'];

const MACHINE_TYPES = {
  'Manitowoc':  ['Indigo NXT Series','Insight Series','QD/QY Modular','NEO Undercounter','Countertop'],
  'Hoshizaki':  ['KM Crescent Cube','DCM Cubelet','AM Modular','B Series Undercounter','Countertop'],
  'Scotsman':   ['C Series Prodigy','N Series Nugget','HID Prodigy Plus','CU Undercounter','Countertop'],
  'Ice-O-Matic':['ICEU Undercounter','GEM Nugget','B Series','ICEU Half Cube','Countertop'],
  'Follett':    ['Symphony Plus','7 Series','25 Series','Horizon Elite','Countertop Nugget'],
  'Cornelius':  ['IMI Series','Enduro Series','CR Series','UCB Undercounter','Countertop'],
};

const FILTER_TYPES = [
  'None / No filter',
  'Basic inline sediment filter',
  'Carbon block filter',
  'Scale inhibitor cartridge',
  'Everpure MH-2',
  'Everpure 4H-2',
  'Everpure MH',
  'Pentek P25',
  'Omnipure K5525',
  'Cuno CFS8112EL',
  'Manitowoc AR-10000-P',
  'Hoshizaki H9655-06',
  'Other (note below)',
];

// Chemical reference - Nu-Calgon Nickel-Safe + No-Rinse Sanitizer
const CHEM_REF = {
  cleaner: {
    name: 'Nu-Calgon Nickel-Safe Ice Machine Cleaner',
    standard: '8oz cleaner per 1 gallon warm water',
    heavy: '16oz per 1 gallon (heavy scale or first service)',
    note: 'NEVER mix with sanitizer. Rinse all cleaner completely before sanitizing.',
    color: 'Green liquid. Mixed solution will be light green/clear.',
  },
  sanitizer: {
    name: 'Nu-Calgon Ice Machine Sanitizer No-Rinse',
    ratio: '1oz per 1 gallon water = 200ppm quat solution',
    note: 'No rinsing after application. Apply and let air dry. This is the final step always.',
    color: 'Clear liquid with slight odor.',
  },
  atp: {
    meter: 'Hygiena Ensure v2 with Ultrasnap testers',
    pass_ice: '10 RLU or below — ice contact surfaces (evaporator, water trough)',
    pass_food: '30 RLU or below — general food contact surfaces (bin, scoop holder)',
    fail: 'Above threshold: re-clean and re-sanitize. Retest before leaving.',
    technique: 'Snap the Ultrasnap tube, swab 10cm² area with firm pressure using Z-pattern, insert into Ensure v2, read in 15 seconds.',
  },
  atp_sales: {
    title: '📊 Sales ATP Protocol — Where to Swab BEFORE Cleaning',
    purpose: 'Swab here to demonstrate contamination to the owner. Hand them the meter while it reads. Let the number do the selling. These surfaces reliably read 200-2000+ RLU on an uncleaned machine.',
    locations: [
      {where: 'Water trough interior', why: 'Biofilm accumulates here constantly. The wet, dark environment is ideal for bacterial growth. Almost always the highest reading in the machine.'},
      {where: 'Inside ice bin near drain', why: 'Moisture trap. Ice runoff pools here. Rarely if ever cleaned by staff. High reading guaranteed on any machine not professionally serviced.'},
      {where: 'Ice scoop handle', why: 'Hands touch this hundreds of times a day. Cross-contamination point. Visually clean but microbiologically dirty.'},
      {where: 'Underside of water curtain', why: 'Dark, permanently wet, never visible to staff. Biofilm (pink/orange slime) often starts here.'},
      {where: 'Float valve housing', why: 'Scale and organic material concentrate around the float. Hard to reach so never cleaned during routine staff wipe-downs.'},
    ],
    pitch: 'Snap the Ultrasnap. Swab the trough with a firm Z-pattern. Insert into Ensure v2. Hand the meter to the owner. Say nothing. Let them watch the number. Anything above 100 RLU is your sales pitch. Above 500 is dramatic. Above 1000 is a close.',
  },
  atp_compliance: {
    title: '\u2705 Compliance ATP Protocol — Where to Swab AFTER Cleaning',
    purpose: 'Swab here for your service report and compliance sticker. These surfaces hold sanitizer best and will read lowest post-clean. Document both pre and post readings — the delta is your proof of value.',
    locations: [
      {where: 'Center of evaporator plate', why: 'Primary ice contact surface. You just descaled and sanitized this. Should read under 5 RLU after proper service. This is your official compliance reading.'},
      {where: 'Water distribution tube exterior', why: 'Clean metal surface that holds sanitizer well. Consistent low readings post-clean.'},
      {where: 'Upper bin interior wall', why: 'Less contact than bottom, holds sanitizer well. Good secondary compliance swab.'},
    ],
    report_format: 'Pre-clean: [RLU] | Post-clean: [RLU] | Delta: [difference]. The delta is your value proof. Example: Before 847 RLU \u2192 After 6 RLU. Frame this on every report.',
    sticker: 'Apply compliance sticker near machine front or inside door. Note post-clean RLU and date. This is what an inspector sees immediately.',
  }
};

const DEEP_CLEAN = {
  'Manitowoc': {
    time: '90-120 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon water (16oz if heavy scale)', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon water', 'Spray bottles for both solutions', 'Separate 5-gallon bucket for soak'],
    steps: [
      {
        title: 'Safety & Setup',
        detail: 'Power off the machine at the switch AND unplug from wall. Never work on a powered machine. Place "Out of Service" sign on front. Put on nitrile gloves — the cleaner is acidic and will irritate skin. Set your cart with two labeled spray bottles (one CLEANER, one SANITIZER), soft brush kit, microfiber towels, and your screws bin. Mix your cleaner solution now: 8oz Nu-Calgon Nickel-Safe into 1 gallon warm water in your 5-gallon bucket. If the machine has not been serviced recently or you can see heavy white/gray scale, use 16oz per gallon. Scale looks like chalky white mineral deposits — think hard water stains on a shower. Mix sanitizer in a separate spray bottle: 1oz Nu-Calgon No-Rinse per 1 gallon water.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Open the front panel (usually 2-4 screws, or lift-and-pull). Stand back and look at the whole interior before touching anything. Tap the YouTube "Identify Components" button above and watch 60 seconds — then come back here. What you will see: (1) EVAPORATOR — fills the upper back half of the machine. Think of a car radiator — vertical metal fins or a flat plate, silver and shiny when clean. Scale makes it look chalky white or gray. Biofilm (bacteria) looks like pink, orange, or brown slime. This is the most important surface. (2) WATER CURTAIN — a clear or milky white plastic flap, about the size of a sheet of paper, hanging vertically in front of the evaporator like a shower curtain. It usually just unclips or lifts off. (3) WATER DISTRIBUTION TUBE — a horizontal white or gray plastic tube running across the top of the evaporator. Has small holes on the underside that drip water down. Think of a soaker hose. Pull it straight out. (4) WATER TROUGH — the shallow rectangular pan at the very bottom of the machine. Holds water that gets pumped up to the distribution tube. This is almost always the dirtiest part — biofilm loves it. (5) FLOAT VALVE — a small plastic ball on a wire or arm, sitting in one corner of the trough. Works exactly like the float in a toilet tank — controls the water level. Scale on the float causes it to stick. Take a photo of everything before removing anything.',
      },
      {
        title: 'Disassemble Removable Components',
        detail: 'Remove in this order, placing all parts in your 5-gallon bucket with cleaner solution to soak: (1) Water curtain — lift up and unhook from clips. It is usually just plastic clips, no tools needed. (2) Water distribution tube — pull straight out or unscrew end cap depending on model. Note which end faces which direction. (3) Float valve assembly — gentle pull, may have a small clip. (4) Water pump inlet screen if visible — small mesh screen, pull straight off. Place ALL plastic parts into your bucket of cleaner solution. Let them soak for the entire time you work on the evaporator — minimum 10 minutes.',
      },
      {
        title: 'Clean the Evaporator',
        detail: 'The evaporator is the most important surface. Using your spray bottle of cleaner solution, thoroughly wet all evaporator plates. Let the solution sit for 5 minutes — do not scrub immediately, let it work. You may see fizzing or bubbling on scale deposits — this is normal and means it is working. After 5 minutes, use your soft brush (NOT scrub pads — never abrasive on evaporator) to gently scrub top to bottom, working the solution into any scale deposits. Rinse with clean water from your spray bottle. Inspect: healthy evaporator is shiny and reflective. If you still see dull white deposits, apply cleaner again and wait 5 more minutes. Repeat until surface is clean. For very stubborn scale you can apply cleaner full strength (no dilution) directly on the deposit with a brush — let sit 10 min then rinse. Rinse thoroughly — no cleaner residue before sanitizing.',
      },
      {
        title: 'Clean the Water Trough & Interior',
        detail: 'The water trough is where biofilm loves to grow. Biofilm is the pink/orange/brown slime you may see — it is a bacterial colony and is exactly what inspectors look for. Pour some cleaner solution into the trough. Use your soft brush to scrub all surfaces including corners and the drain plug area. Use a scrub pad on stubborn deposits in the trough — it is okay to be more aggressive here than on the evaporator. Check the drain — clear any debris. Wipe down all interior walls, ceiling, and side panels with a cleaner-soaked microfiber towel. Check where the water pump sits — biofilm often hides underneath.',
      },
      {
        title: 'Clean Soaked Components',
        detail: 'Pull your soaked parts from the bucket. The cleaner should have loosened deposits significantly. Scrub each piece: Water curtain — scrub both sides, pay attention to the hinge/clip areas where biofilm accumulates. Distribution tube — use a small brush or pipe cleaner through the holes to clear any blockage. Each hole should spray freely — hold it up to light to verify. Float valve — scrub the ball/arm and housing. Check that the float moves freely. Rinse all parts thoroughly under clean water until no cleaner smell remains. Cleaner residue left on parts will interfere with sanitizer.',
      },
      {
        title: 'Clean Ice Bin & Exterior',
        detail: 'The ice bin is a food contact surface — treat it accordingly. Wipe down all bin interior surfaces with cleaner solution using microfiber towels. Pay attention to the bottom corners and drain area — biofilm pools here. Check the bin drain is clear. Clean the ice scoop and scoop holder — these are often the most contaminated surfaces because hands touch them constantly. Clean the bin door gasket if present. For exterior: wipe down all stainless surfaces. Clean the air filter — it is usually behind a front or side panel, a mesh or foam pad. Rinse it under water, let dry, or blow out with your air compressor. A clogged filter causes the machine to overheat and produce less ice.',
      },
      {
        title: 'Rinse Everything',
        detail: 'This step is critical. Any cleaner residue left in the machine will neutralize your sanitizer and leave a chemical taste in the ice. Rinse the evaporator with clean water using your spray bottle. Rinse the trough — pour clean water in and drain. Rinse all reinstalled or ready-to-reinstall components under clean water. Wipe all surfaces with a clean damp microfiber towel. If you can still smell the Nu-Calgon cleaner, rinse again. The machine should smell neutral before you sanitize.',
      },
      {
        title: 'Reassemble Components',
        detail: 'Reinstall in reverse order: (1) Float valve — snap back into place, verify it moves freely up and down. (2) Water pump screen — push back on firmly. (3) Distribution tube — slide back in with the same orientation as removed, holes facing down toward evaporator. (4) Water curtain — hook back onto clips, should hang freely in front of evaporator. Reconnect water supply if you disconnected it. Do not power on yet.',
      },
      {
        title: 'Sanitize All Surfaces',
        detail: 'Mix fresh sanitizer: 1oz Nu-Calgon No-Rinse per gallon water in your spray bottle. Spray generously on: evaporator plates (top to bottom), water trough interior, water curtain both sides, distribution tube exterior, all bin interior surfaces, bin door and gasket. DO NOT RINSE. The no-rinse formula is designed to air dry and leave a protective sanitizing film. Do not wipe dry — air dry only. This is the step that provides the antimicrobial protection between services.',
      },
      {
        title: 'ATP Testing',
        detail: 'Before powering on, run your ATP test. Using your Hygiena Ensure v2 and Ultrasnap testers: Snap the Ultrasnap tube to release the reagent. Swab a 10cm² area (about the size of your palm) on the evaporator using a firm Z-pattern — 5 strokes across, 5 strokes down. Insert swab into Ensure v2. Read in 15 seconds. PASS: 10 RLU or below on ice contact surfaces (evaporator, trough). PASS: 30 RLU or below on general surfaces (bin). FAIL: Above threshold — re-clean the area, re-sanitize, and retest. Note your RLU reading — this goes on the service report and compliance sticker.',
      },
      {
        title: 'Restart & Verify',
        detail: 'Reconnect power. Power on machine. Manitowoc will run an automatic startup cycle — do not interrupt. The first harvest cycle takes 15-30 minutes. Watch for: water flowing through distribution tube (should see even water across evaporator), ice beginning to form on evaporator plates (thin clear sheet), and clean harvest into bin. DISCARD the first full bin of ice — it may contain residual cleaning chemicals and should never be served. Verify the machine returns to normal operation with consistent ice production before leaving.',
      },
      {
        title: 'Document & Compliance Sticker',
        detail: 'Fill out your service report: date, tech name, machine brand/model, chemicals used, RLU reading, condition notes, next service date (60 days). Apply your compliance sticker in a visible location — note the RLU score and date. This sticker is what an inspector sees immediately and signals a professionally maintained machine. Leave a copy of the service report with the manager. Brief the manager on what you found, what you did, and when you will return.',
      },
    ]
  },
  'Hoshizaki': {
    time: '90-120 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon — CRITICAL: Hoshizaki evaporators are stainless steel. Any commercial ice machine cleaner works. Nickel-Safe is correct.', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon', 'Extra rinse water — Hoshizaki requires more thorough rinsing than other brands'],
    steps: [
      {
        title: 'Safety & Setup — Hoshizaki Specific',
        detail: 'Power off and unplug. IMPORTANT Hoshizaki note: their evaporators are stainless steel — more durable than other brands and compatible with any commercial ice machine cleaner. Nickel-Safe cleaner (what you have) is fine and will not damage the evaporator. However, avoid abrasive scrubbers on the evaporator surface — use soft brush only to prevent scratching. The damage from wrong chemicals is irreversible and expensive. Set up cart, mix cleaner (8oz per gallon warm water), mix sanitizer (1oz per gallon). Gloves on.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above and watch 60 seconds to orient yourself, then come back. Hoshizaki looks noticeably different from Manitowoc inside — do not assume the layout is the same. Open the front panel. What you will see: (1) EVAPORATOR — Hoshizaki makes crescent-shaped cubes. The evaporator looks like a curved metal plate or a grid of individual cup-shaped molds arranged in rows, taking up the upper section. Unlike Manitowoc which is a flat plate, this has a 3D curved shape. It is stainless steel — should look uniformly silver. Dark spots or rough texture means scale. (2) SPRAY BAR — THIS IS THE KEY DIFFERENCE FROM OTHER BRANDS. A horizontal tube, about the diameter of a pen, running across the top with tiny holes pointing down at the evaporator. When the machine runs you can see it spraying water. Pull it out of its clips — it just slides out. Hold it up to a light and look through each hole. Every hole should show a clear circle of light. (3) FLOAT SWITCH — a small white or gray cylinder about the size of a AA battery, hanging on a wire in the water reservoir. Push it up and down — it should slide freely. (4) WATER TROUGH — shallower than Manitowoc, at the base. In Florida heat this is where pink/orange biofilm (Serratia marcescens bacteria) grows most aggressively. (5) CURTAIN — clear plastic panel clipped to the front of the evaporator. Take a photo before touching anything.',
      },
      {
        title: 'Disassemble Components',
        detail: 'Remove carefully — Hoshizaki components are generally less rugged than Manitowoc and can crack if forced: (1) Curtain assembly — unclip from front of evaporator, set aside. (2) Spray bar — pull straight out from mounting clips. Hold up to light and look through the holes — each hole should be clear. Note any that are blocked. (3) Float switch — gently pull off the mounting post. The float should slide freely on the wire. (4) Water pump screen — pull off the pump inlet. Soak all plastic components in cleaner solution (8oz per gallon) in your 5-gallon bucket for minimum 10 minutes.',
      },
      {
        title: 'Clean Spray Bar Holes — Critical Step',
        detail: 'The spray bar holes are the #1 maintenance issue on Hoshizaki machines. Blocked holes cause uneven water distribution which creates irregular or incomplete crescent cubes — a telltale sign to inspectors of poor maintenance. While components are soaking, use a toothpick or the smallest brush in your kit to clear each hole on the spray bar. Push through from the inside out. Hold up to light to confirm clear. There are typically 15-25 holes on a standard spray bar. If a hole is completely blocked and toothpick cannot clear it, a thin wire or unfolded paper clip works. Rinse the bar thoroughly after clearing.',
      },
      {
        title: 'Clean Evaporator — Gentle Technique Required',
        detail: 'Apply cleaner solution to evaporator using spray bottle. Let sit 5 minutes. Use ONLY the softest brush in your kit — NEVER use scrub pads, steel wool, or abrasive materials on Hoshizaki evaporator. Even though it is stainless steel, scratches create micro-grooves that harbor bacteria. Keep technique gentle and use circular motions with light pressure. Use soft circular motions with light pressure. Rinse with clean water. Inspect: should be uniformly silver with no dull spots. Repeat cleaner application if needed. For stubborn deposits apply cleaner with brush and let sit 10 minutes before gentle scrubbing. The key word on Hoshizaki is GENTLE throughout.',
      },
      {
        title: 'Clean Float Switch & Trough',
        detail: 'Float switch: wipe the float and wire with cleaner-soaked cloth. Scale on the float causes it to stick, leading to overflow or underfill. Verify the float slides freely up and down the wire after cleaning. Water trough: scrub with soft brush and cleaner solution, paying attention to corners where biofilm accumulates. Check the drain. Wipe down all interior surfaces. Interior walls often show pink biofilm on Hoshizaki units in Florida heat — this is Serratia marcescens, a common waterborne bacteria. Scrub thoroughly with cleaner-soaked scrub pad.',
      },
      {
        title: 'Reassemble & Rinse Thoroughly',
        detail: 'Rinse all soaked components under clean water until no cleaner odor. Reinstall: (1) Float switch — back on mounting post. (2) Water pump screen — back on inlet. (3) Spray bar — back in clips, holes facing evaporator. (4) Curtain assembly — clip back onto evaporator front. THOROUGH RINSE CRITICAL on Hoshizaki: spray clean water on evaporator, flush trough with clean water, wipe all surfaces. Hoshizaki machines are more sensitive to cleaner residue than other brands — any residue left will affect ice taste and sanitizer effectiveness.',
      },
      {
        title: 'Sanitize, ATP Test & Restart',
        detail: 'Spray sanitizer (1oz per gallon) on all surfaces: evaporator, spray bar exterior, trough, curtain, bin interior. Air dry — do not wipe. ATP test on evaporator surface and trough. Pass: 10 RLU or below on ice contact surfaces. Fail: re-clean and retest. Power on — Hoshizaki will run a startup cycle. First harvest cycle on a crescent machine takes 20-35 minutes. Watch that the spray bar is spraying evenly across all evaporator molds. Discard first bin of ice. Verify crescent cubes are uniform in size and shape — irregular cubes mean water distribution issue, recheck spray bar.',
      },
    ]
  },
  'Scotsman': {
    time: '75-105 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon', 'Scotsman has an automated clean cycle — chemicals go into reservoir, not spray bottle'],
    steps: [
      {
        title: 'Safety & Scotsman Controls',
        detail: 'Power off and unplug. Scotsman Prodigy machines have a digital control panel — before unplugging, note any error codes displayed. Write them down. Common codes: E1 (harvest time too long — scale on evaporator), E2 (freeze time too long — refrigeration issue), E8 (water level — float or inlet issue). These codes tell you what to focus on during cleaning. Unplug machine. Gloves on. Remove all ice from bin.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above and watch 60 seconds first. Scotsman Prodigy is the most visually distinct machine — it looks different from both Manitowoc and Hoshizaki. Open the front panel. What you will see: (1) EVAPORATOR PLATE — Scotsman is unique because ice forms on a large VERTICAL flat plate on the back wall, not a horizontal hanging plate like Manitowoc. The plate is roughly the size of a laptop screen. When clean, it is smooth and reflective like a mirror. White crusty patches are scale. Pink or orange film is biofilm. (2) ICE THICKNESS SENSOR — THIS IS SCOTSMAN\u2019S CRITICAL COMPONENT. A small probe about the diameter of a pencil, sticking out horizontally near the middle of the evaporator. It is usually black or dark gray. The tip points toward the evaporator plate with a gap of about 3/8 inch — roughly the thickness of a pencil. You can check this gap by holding a pencil horizontally between the probe tip and the plate. Scale on this probe changes the gap and causes harvest problems. Wipe it clean at every single visit. (3) CONTROL PANEL — Scotsman Prodigy has a digital display on the front. Always check for error codes before unplugging. E1 = harvest too long (scale). E2 = freeze too long (refrigeration issue — not your problem to fix). Write down any codes. (4) DRAIN PAN — at the bottom, under the evaporator. Scotsman drain lines block with grease more than other brands. Pour a cup of water in and time it — should drain in under 30 seconds. (5) WATER CURTAIN — flat plastic panel in front of the evaporator, just unclips.',
      },
      {
        title: 'Disassemble Components',
        detail: 'Remove: (1) Water curtain — unclip from mounting brackets. (2) Ice thickness sensor — note its position carefully before removing. It is a small probe on a wire. Gently pull straight out. (3) Water distribution components if accessible. Soak plastic parts in cleaner solution. While soaking, inspect the drain pan and drain line — pour a cup of water in the drain pan and watch it drain. Should drain within 30 seconds. If slow, the drain line is blocked. Clear blockage with a thin brush or flexible cleaning brush from your kit.',
      },
      {
        title: 'Clean Ice Thickness Sensor — Critical',
        detail: 'The ice thickness sensor is Scotsman-specific and critical. Scale buildup on the sensor tip causes premature or incomplete harvest cycles — either ice falls off too early (thin, watery ice) or stays too long (frozen solid, bridged). Using a soft cloth dampened with cleaner solution, wipe the sensor tip. Do not bend the probe. Do not use abrasive materials. The tip should be clean metal with no white deposits. While cleaning, verify the mounting position — the gap between probe tip and evaporator surface should be approximately 3/8 inch. You can gauge this with a pencil held sideways — roughly the thickness of a pencil. If the gap is wrong (sensor was bent or moved) ice quality will be affected.',
      },
      {
        title: 'Clean Evaporator Plate',
        detail: 'Apply cleaner solution (8oz per gallon) to evaporator plate from top to bottom using spray bottle. Let sit 5 minutes. Vertical evaporator means cleaner will run down — that is fine and actually helps it penetrate scale. Scrub with soft brush top to bottom, paying attention to the area directly across from the thickness sensor where ice always contacts the plate. Rinse with spray bottle of clean water. Repeat if scale remains. Scotsman evaporators can tolerate slightly more scrubbing pressure than Hoshizaki but still avoid abrasive pads directly on the plate surface.',
      },
      {
        title: 'Clean Trough, Drain & Interior',
        detail: 'Trough: scrub with cleaner and brush. Check float — should move freely. Interior walls: wipe all surfaces with cleaner-soaked microfiber. Drain system: this is especially important on Scotsman. Pour cleaner solution down the drain line. Use your flexible brush to scrub the drain pan interior. Verify drain flows freely. Scotsman units in Florida kitchens get grease in the drain line which traps biofilm. If you find significant grease buildup note this in your service report.',
      },
      {
        title: 'Reassemble, Rinse & Sanitize',
        detail: 'Reinstall all components. IMPORTANT: reinstall ice thickness sensor in exact position — gap should be 3/8 inch from evaporator. Rinse all surfaces with clean water. Spray sanitizer on evaporator, curtain, trough, bin interior. Air dry. ATP test on evaporator and trough. Pass: 10 RLU or below. Power on. Scotsman Prodigy will run a startup diagnostic — allow it to complete without interruption. First harvest typically 25-35 minutes. Discard first bin. Watch that ice forms uniformly across the evaporator plate — thin patches indicate a distribution or refrigeration issue.',
      },
    ]
  },
  'Ice-O-Matic': {
    time: '75-90 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {
        title: 'Safety & Setup',
        detail: 'Power off and unplug. Ice-O-Matic units are common in fast food and cafe environments and often have heavier grease contamination than restaurant machines. This affects the exterior and air filter particularly. Gloves on. Mix cleaner 8oz per gallon warm water. Mix sanitizer 1oz per gallon. Remove all ice.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above first. Ice-O-Matic looks similar to Manitowoc internally but has one key difference that matters most. Open the front panel. What you will see: (1) EVAPORATOR — horizontal plate at the top, similar position to Manitowoc. Nickel-plated. Should be shiny silver when clean. (2) WATER DISTRIBUTION PAN — THIS IS IOM\u2019S DEFINING FEATURE. Unlike Manitowoc tube, IOM uses a flat rectangular plastic pan sitting directly above the evaporator with tiny holes on the bottom face. It looks like a plastic tray with holes punched in it, about the size of a small cutting board. Remove it by lifting straight up. Hold it up to a window or flashlight — every single hole should show a clear circle of light through it. Holes blocked with white mineral deposits look dark or cloudy. These holes are smaller than Manitowoc and block faster. Clearing them is the most important thing you do on an IOM machine. (3) BIN BAFFLE — a plastic divider panel inside the ice bin, usually white, that guides ice as it falls. Flip it over and look at the underside — this dark hidden surface grows biofilm that the owner never sees. (4) WATER TROUGH — base reservoir, scrub thoroughly. (5) WATER CURTAIN — clear plastic panel in front of evaporator, just unclips.',
      },
      {
        title: 'Disassemble Components',
        detail: 'Remove: (1) Water curtain — unclip. (2) Bin baffle — usually just lifts out. Check underside for pink/orange biofilm. (3) Distribution pan — lift out. (4) Water pump screen. Place all in cleaner soak. IMMEDIATELY inspect distribution pan holes while fresh — hold up to light and look through each hole. They should all be clear circles. Blocked holes appear dark or have white mineral plugs. Use a toothpick to clear each blocked hole now before soaking. Post-soak clearing is harder.',
      },
      {
        title: 'Clear Distribution Pan Holes — Most Important IOM Step',
        detail: 'This is the defining maintenance step on Ice-O-Matic units. Blocked distribution holes cause uneven water across the evaporator which creates scale that forms unevenly — leading to an irregular freeze pattern and poor ice quality. After soaking for 10 minutes in cleaner solution, remove pan and use toothpick on every hole. There are typically 20-40 holes on a standard unit. Every hole must be clear. After clearing, hold the pan up to a light source and look through from the bottom — every hole should show a pinpoint of light. If any remain blocked, use a thin wire. Rinse pan completely under clean water.',
      },
      {
        title: 'Clean Evaporator & Interior',
        detail: 'Apply cleaner solution to evaporator. Let sit 5 minutes. Scrub with soft brush. Scale on IOM evaporators often forms in strips corresponding to blocked distribution holes — you can actually diagnose which holes were blocked by where the scale is heaviest. Rinse thoroughly. Wipe all interior walls. Trough: scrub with brush and cleaner, clean drain. Interior inspection: IOM units in fast food get grease from cooking on interior surfaces — use your scrub pad on interior walls if grease is present. Air filter: remove (usually clips off front panel), rinse or blow out with air compressor.',
      },
      {
        title: 'Reassemble, Sanitize & Test',
        detail: 'Reinstall all components. Rinse all surfaces with clean water. Spray sanitizer on all surfaces, air dry. ATP test on evaporator and trough. Pass: 10 RLU. Power on. First harvest 15-25 minutes. Watch that water flows through ALL distribution holes evenly — you should see an even sheet of water across the evaporator. Uneven water = holes still blocked. Discard first bin of ice.',
      },
    ]
  },
  'Follett': {
    time: '60-90 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon — Follett uses stainless steel evaporator, Nickel-Safe is appropriate', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {
        title: 'Safety & Follett-Specific Setup',
        detail: 'Power off and unplug. Follett nugget ice machines work fundamentally differently from cube ice machines — they use a rotating AUGER inside an evaporator CYLINDER to continuously extrude nugget ice rather than a batch freeze/harvest cycle. This changes the cleaning process significantly. Gloves on. Mix chemicals. Remove all ice from storage bin. Listen to the machine before unplugging — unusual grinding from the auger indicates scale on cylinder walls or auger wear. Note this in your service record.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above first — Follett works completely differently from all other brands and a video helps enormously. Remove the top panel (4 screws). What you will see: (1) EVAPORATOR CYLINDER — a vertical stainless steel tube, roughly the diameter of a large coffee can (6-8 inches). This is the entire ice-making mechanism. Ice does not form on a plate — it forms as a thin layer on the INSIDE WALL of this cylinder and gets scraped off continuously by the auger. (2) AUGER — a metal helical screw (think of a giant corkscrew or drill bit) inside the cylinder. You cannot see it until you remove it — it fills the entire inside of the cylinder. The auger turns slowly and continuously, scraping ice off the cylinder wall and pushing it upward out the top. You will hear it humming when the machine runs. BEFORE UNPLUGGING: listen. Smooth hum = healthy. Grinding or clicking = scale buildup on cylinder, will need attention. (3) AUGER MOTOR — sits on top of the cylinder, usually a black box with a shaft going down. This is what drives the auger. (4) ICE CHUTE — a rectangular plastic channel leading from the cylinder top to the storage bin. Ice travels through here continuously. This gets biofilm because it is always wet and dark. (5) STORAGE BIN — the hopper below where Chewblet nugget ice collects. Soft, moist nuggets = machine running well. Hard or irregular nuggets = auger or cylinder issue.',
      },
      {
        title: 'Access Evaporator Cylinder',
        detail: 'The auger must be removed to properly clean the cylinder interior on most Follett models. Disconnect the auger drive coupling at the top (typically a hex bolt — your impact drill is useful here). Lift the auger straight up and out of the cylinder. It is heavy — 10-20 lbs depending on model. Lay auger flat on a clean surface. Inspect auger: fins should be intact and smooth. Pitting or cracking means replacement is needed — note in your report. Scale on auger fins appears as rough white deposits.',
      },
      {
        title: 'Clean Evaporator Cylinder Interior',
        detail: 'Pour cleaner solution (8oz per gallon) into the cylinder — fill approximately 1/3. Use your long brush to scrub the interior cylinder walls with circular strokes. Scale will be on the lower portion where ice forms. The solution will fizz on scale deposits — this is the acid working. Let solution sit 10 minutes. Scrub again. Drain by tilting machine slightly or using a cup to scoop out solution. Rinse with clean water poured into cylinder. Repeat rinse until no cleaner smell. Inspect cylinder wall with a flashlight — should be smooth and shiny throughout.',
      },
      {
        title: 'Clean Auger',
        detail: 'While cylinder is draining/rinsing, clean the auger. Spray or brush cleaner solution onto all auger fins. Let sit 5 minutes. Scrub with soft brush — get between the fins where scale accumulates. Rinse thoroughly. Inspect the tip of the auger (the pointed bottom end) — this is where it contacts the cylinder wall and scale wears it down. Note unusual wear. Verify auger fins are not bent or cracked. A bent fin will score the cylinder interior and create a pathway for bacterial growth.',
      },
      {
        title: 'Clean Bin, Chute & Dispenser',
        detail: 'Bin interior: scrub with cleaner solution. Pay attention to waterline area where biofilm concentrates. Drain: check for blockage, flush with clean water. Ice chute: use your smaller brush to scrub the chute interior — ice travels through here continuously and biofilm forms. On Symphony units: clean dispenser nozzle and actuator arm with cleaner-soaked cloth. The dispenser is a high-touch surface and should be sanitized thoroughly.',
      },
      {
        title: 'Reassemble, Sanitize & Test',
        detail: 'Reinstall auger into cylinder — lower straight down, engaging the drive coupling. Reconnect drive coupling bolt (snug, not overtightened — your drill on low torque or by hand). Rinse all surfaces with clean water. Spray sanitizer into cylinder, on auger, chute, bin, and dispenser. Air dry. ATP test: swab the ice chute and cylinder wall. Pass: 10 RLU. Power on. Follett machines take 15-30 minutes to reach full ice production — the cylinder must re-chill before nuggets form. Discard first production cycle. Verify nuggets are soft and consistent — hard or irregular nuggets indicate cylinder or auger issue.',
      },
    ]
  },
  'Cornelius': {
    time: '60-90 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {
        title: 'Safety & Setup',
        detail: 'Power off and unplug. Cornelius units are compact countertop or undercounter machines common in convenience stores and small cafes. They have simpler internal layouts than modular units but the cleaning process is equally important. Gloves on. Mix chemicals. Remove ice. Cornelius machines in convenience stores often have heavier use and more contamination than restaurant units — the ice is frequently touched by customers or used in fountain drinks where the machine runs constantly.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above first. Cornelius is the simplest machine internally — good news for a first-timer. Open or remove the front panel (usually 2-4 screws or a snap-off panel, varies by model). What you will see: (1) FREEZE PLATE — the evaporator on a Cornelius is simpler than modular brands. It is a flat metal plate, usually positioned vertically or at an angle near the back. Smaller than a Manitowoc evaporator — roughly the size of a hardcover book. Should be smooth and reflective when clean. White crusty patches are scale. (2) AIR FILTER — FIND THIS FIRST. Usually a foam or mesh pad behind a removable grille on the side or rear of the unit. Pull it out and hold it to light. If you cannot see through it, it is clogged. A blocked filter is the #1 reason Cornelius machines underperform. Many owners have never cleaned it. (3) WATER TROUGH — small shallow reservoir at the base. (4) DRAIN PAN — under or beside the trough. In convenience store environments this often has a sour smell from biofilm in the drain. Smell it before cleaning — if odorous, scrub extra thoroughly. (5) ICE BIN — the storage compartment below. Simpler layout with fewer places for biofilm to hide than modular machines — but check corners and the drain area.',
      },
      {
        title: 'Disassemble Accessible Components',
        detail: 'Remove what is accessible: water curtain if present (some models), any removable troughs or pans. Cornelius units have fewer removable components than larger machines — most cleaning is done in-place. Soak any removable parts in cleaner solution. IMMEDIATELY address the air filter: remove it (slide out from rear or side), hold it up to light — if you cannot see through it easily it is heavily clogged. Rinse under water. If solid with dust/grease use your air compressor to blow from inside out. Let dry. A blocked air filter causes the condenser to overheat, reducing ice production and stressing the compressor.',
      },
      {
        title: 'Clean Freeze Plate',
        detail: 'Apply cleaner solution to freeze plate using spray bottle. Let sit 5 minutes. Scrub with soft brush using horizontal strokes. Cornelius freeze plates are typically nickel-plated or stainless — compatible with Nickel-Safe. Scale on these smaller plates tends to be concentrated where water distributes. Rinse thoroughly with clean water. Inspect — should be uniformly shiny. Repeat if dull areas remain. For heavy scale: apply cleaner solution with brush, cover with a damp cloth to hold moisture, let sit 15 minutes, scrub and rinse.',
      },
      {
        title: 'Clean Trough, Drain Pan & Interior',
        detail: 'Water trough: scrub all surfaces with cleaner and brush. The trough drain is a common odor source on Cornelius units — pour cleaner solution into drain and let sit 5 minutes. Flush with clean water. Drain pan: this is where Cornelius units develop the characteristic convenience store ice machine odor. Scrub pan with scrub pad and cleaner. If there is significant slime or pink biofilm this is Serratia marcescens — be thorough. Interior walls: wipe all accessible surfaces with cleaner-soaked microfiber. Exterior: wipe stainless surfaces. Clean around water inlet valve area.',
      },
      {
        title: 'Reassemble, Sanitize & Test',
        detail: 'Reinstall all components. Make sure air filter is dry before reinstalling. Rinse all interior surfaces with clean water. Spray sanitizer on freeze plate, trough interior, drain pan, and bin interior. Air dry. ATP test: swab freeze plate and trough. Pass: 10 RLU on ice contact surfaces. Power on. Cornelius machines restart quickly — typically 10-20 minutes to first ice. Discard first production. Verify ice is clear (not cloudy) — cloudy ice indicates mineral contamination in water, possible filter issue. Note in report.',
      },
    ]
  },
};

const MAINTENANCE_60 = {
  'Manitowoc': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon (half strength for maintenance)', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon', 'Spray bottles for both'],
    steps: [
      {title:'Visual Inspection & Machine Check', detail:'Power stays ON for 60-day maintenance — you are inspecting and spot-cleaning, not doing a full chemical cycle. Observe the machine operating: listen for unusual sounds (grinding = scale on evaporator, gurgling = water level issue, clicking = harvest problem). Check ice in bin — should be clear, consistent size cubes. Cloudy or malformed ice indicates a problem. Open front panel. Check evaporator for new scale deposits since last service. Light white haze is normal mineral accumulation — heavier deposits need attention. Check for any pink/orange biofilm starting in corners or on curtain.'},
      {title:'Spot Clean Evaporator if Needed', detail:'If minor scale or light biofilm is visible: mix half-strength cleaner (4oz per gallon). Spray on affected areas only. Let sit 3 minutes. Wipe with soft brush or microfiber. Rinse with damp cloth. If scale covers more than 25% of evaporator surface or biofilm is present beyond light discoloration — elevate to full deep clean, do not proceed with 60-day maintenance only.'},
      {title:'Clean Water Trough & Float', detail:'The trough accumulates mineral deposits even between full cleanings. Wipe trough interior with cleaner-soaked microfiber. Check float valve — should move freely up and down. Light scale on float: wipe with cleaner cloth. Float stuck or heavily scaled: flag for deep clean. Check drain flows freely — pour small cup of water in trough and verify it drains.'},
      {title:'Inspect & Clear Distribution Tube', detail:'Check water distribution tube holes visually — look for any starting to cloud over. Use toothpick to clear any partially blocked holes. Verify water flows evenly across evaporator during machine operation. Uneven water means a distribution issue — note for next deep clean.'},
      {title:'Air Filter & Condenser', detail:'Remove air filter. If dusty but not clogged, blow out with air compressor (always blow from inside out to avoid pushing debris deeper). If heavily clogged, rinse with water and dry before reinstalling. Check condenser coils (usually visible behind or under air filter location) — dust on coils reduces efficiency. Use your condenser brush to gently clean coil fins if buildup is visible. Be gentle — coil fins bend easily.'},
      {title:'Bin Wipe & Sanitize', detail:'Wipe bin interior with sanitizer solution (1oz per gallon). Do not rinse. Check scoop and scoop holder — wipe both with sanitizer. Apply sanitizer spray to evaporator, trough, and curtain. Air dry. No rinsing.'},
      {title:'ATP Spot Test & Document', detail:'Run quick ATP on water trough surface. Pass: 10 RLU. If fail: do a more thorough spot clean and retest. If still failing: elevate to full deep clean today. Note ATP reading, date, any observations (new sounds, ice quality changes, scale level). Update next service date — 60 days from today. If the machine is approaching 6 months since last deep clean, schedule deep clean for next visit.'},
    ]
  },
  'Hoshizaki': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon (half strength)', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Visual & Ice Quality Check', detail:'Machine stays on. Observe ice: Hoshizaki crescent cubes should be uniform C-shapes. Irregular shapes (partial crescents, flat pieces, stuck-together cubes) indicate spray bar blockage or water distribution issue. Check ice is clear — cloudy crescent cubes indicate mineral contamination. Open panel. Inspect spray bar holes visually.'},
      {title:'Spray Bar Inspection & Clearing — Priority Step', detail:'The spray bar is the most important 60-day maintenance item on Hoshizaki. Remove spray bar (pull from clips). Hold to light — look through each hole. Any that appear cloudy or dark: use toothpick to clear. Rinse bar under water. Reinstall with holes facing evaporator. This single step prevents most common Hoshizaki ice quality issues.'},
      {title:'Float Switch Check', detail:'Check float switch — gently push float down and release. Should spring back to top position freely. If it sticks or moves slowly, mineral deposits are building. Wipe float and wire with half-strength cleaner solution. Do not use abrasive. Note: a sticking float will cause water overflow or ice production cutoff — this is a common Hoshizaki service call.'},
      {title:'Evaporator Visual Check', detail:'Look at evaporator surface. Hoshizaki stainless evaporator should be uniformly silver. Any brown, orange, or dark spots: biofilm forming. Light surface biofilm can be wiped with sanitizer solution. More than light spots: elevate to deep clean. Never use abrasive materials on the Hoshizaki evaporator — stainless scratches harbor bacteria.'},
      {title:'Bin & Sanitize', detail:'Wipe bin interior with sanitizer. Check bin drain clear. Apply sanitizer to spray bar exterior, trough, curtain. Air dry.'},
      {title:'ATP & Document', detail:'Swab trough. Pass 10 RLU. Document ATP, spray bar condition (clear/partially blocked/cleared today), evaporator condition, float switch condition. Set next service 60 days.'},
    ]
  },
  'Scotsman': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Prodigy Diagnostic Check', detail:'Before opening machine, press INFO button on Scotsman Prodigy control panel. Note any error codes displayed. E1 (harvest too long) typically means scale building on evaporator — flag for deep clean if frequent. E2 (freeze too long) is refrigeration-related — not a cleaning issue, contact a refrigeration tech. Clear minor codes after noting them.'},
      {title:'Ice Thickness Sensor — 60-Day Priority', detail:'The ice thickness sensor must be checked at every visit. Open panel. Locate the small black probe near the evaporator. Wipe probe tip with half-strength cleaner solution on a soft cloth. Check gap — should be approximately 3/8 inch from evaporator surface (thickness of a pencil). Scale buildup changes this gap. If you see white deposits on the probe tip, clean thoroughly. This sensor controls the entire harvest cycle — a dirty sensor causes premature harvest (thin ice) or delayed harvest (bridged ice).'},
      {title:'Water Curtain & Drain', detail:'Check water curtain for scale or biofilm. Wipe with cleaner if needed, rinse, sanitize. PRIORITY: check Scotsman drain system. Pour 1 cup of water in drain pan — time how long it takes to drain. Should drain in under 30 seconds. 30-60 seconds: partial blockage, flush drain line. Over 60 seconds: full blockage, clear drain line with flexible brush. Drain blockage is the #1 service call issue on Scotsman units.'},
      {title:'Air Filter & Exterior', detail:'Remove air filter. Rinse or blow out with compressor. Scotsman recommends monthly filter inspection in kitchen environments — if you see grease coating the filter, increase cleaning frequency and note in report. Wipe exterior stainless.'},
      {title:'Sanitize & Document', detail:'Spray sanitizer on curtain, trough, bin interior. Air dry. ATP on trough. Pass 10 RLU. Document: error codes noted, sensor condition, drain drain time, filter condition. Set next service 60 days.'},
    ]
  },
  'Ice-O-Matic': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Visual & Ice Check', detail:'Machine on. Check ice quality — IOM cubes should be consistent size and clear. Observe water flow if possible. Open panel and inspect distribution pan holes visually.'},
      {title:'Distribution Pan Holes — Priority Check', detail:'Remove distribution pan. Hold to light. Check every hole — even partial blockage affects ice quality. Clear any starting-to-block holes with toothpick now. It is much easier to clear early-stage blockage than full blockage. Rinse and reinstall.'},
      {title:'Bin Baffle & Trough', detail:'Check underside of bin baffle for biofilm. Pink or orange slime: wipe with half-strength cleaner, rinse, note. Trough: wipe with cleaner-soaked cloth. Check drain flows.'},
      {title:'Air Filter', detail:'Remove, clean with air compressor or rinse. IOM units in fast food environments get grease on filter — check monthly in these locations.'},
      {title:'Sanitize & Document', detail:'Spray sanitizer on evaporator, distribution pan (exterior), trough, bin. Air dry. ATP on trough. Pass 10 RLU. Document distribution hole condition, filter condition. Set 60-day return.'},
    ]
  },
  'Follett': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Auger Sound Check — Listen First', detail:'Before unplugging: power on machine and listen to auger operation. Should be smooth hum. Grinding, clicking, or irregular sounds indicate scale on cylinder walls or auger wear. Note sound quality in report — this is your early warning system. Heavy grinding: consider elevating to deep clean today.'},
      {title:'Ice Quality Check', detail:'Follett nugget ice should be soft, moist, and chewable — the signature Follett product. Hard, dry, or irregularly sized nuggets indicate auger or cylinder issue. Inspect bin for ice quality.'},
      {title:'Ice Chute & Dispenser', detail:'The ice chute is a high-frequency cleaning area — clean at every 60-day visit. Use small brush to scrub chute interior with half-strength cleaner. Rinse. On Symphony dispenser units: clean dispenser nozzle with cleaner cloth. Sanitize chute and nozzle.'},
      {title:'Bin & Drain', detail:'Wipe bin interior with sanitizer. Check drain flows. Follett bins with slow drains develop biofilm faster.'},
      {title:'Cylinder Top Access', detail:'Remove top panel and visually inspect top of auger and cylinder opening. Any visible scale or discoloration at the top: flag for deep clean. Do not attempt partial auger cleaning at 60-day service — either full deep clean or visual inspection only.'},
      {title:'Sanitize & Document', detail:'Sanitize chute, dispenser, bin. ATP on ice chute. Pass 10 RLU. Document auger sound quality, ice quality, chute condition. Set 60-day return.'},
    ]
  },
  'Cornelius': {
    time: '25-35 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Visual & Ice Quality', detail:'Cornelius 60-day service is simpler than larger machines. Check ice is clear — Cornelius countertop units in convenience stores sometimes produce cloudy ice due to high-use water with heavy mineral content. Cloudy ice = filter issue or very hard water. Note for report.'},
      {title:'Air Filter — Most Important Cornelius Step', detail:'The air filter is the #1 maintenance item on Cornelius units. Remove (slide out from side or rear). If you cannot see through it when held up to light: heavily clogged. Rinse under water and blow out with air compressor. If grease-coated (common in convenience store environments near food prep): may need replacement. A blocked air filter will visibly reduce ice production within weeks.'},
      {title:'Freeze Plate Spot Check', detail:'Open front panel. Look at freeze plate for new scale or biofilm since last deep clean. Light haze: normal. More than light haze: wipe with half-strength cleaner, rinse. Significant buildup: elevate to deep clean.'},
      {title:'Drain Pan Odor Check', detail:'Smell the drain pan area. Any sour or musty odor indicates biofilm in drain pan or drain line. Pour cup of water in drain — verify it drains freely. Wipe drain pan with cleaner if odor present. Rinse. This is the most common Cornelius complaint — recurring odor means deep clean schedule needs adjustment.'},
      {title:'Sanitize & Document', detail:'Wipe bin with sanitizer. Spray freeze plate and trough with sanitizer. Air dry. ATP on freeze plate. Pass 10 RLU. Document: filter condition, ice clarity, drain odor (yes/no), freeze plate condition. Set 60-day return.'},
    ]
  },
};


const VIDEO_LINKS={
  'Manitowoc':{
    label:'Manitowoc Official + WebstaurantStore Deep Clean Video',
    deep_url:'https://www.webstaurantstore.com/video-7936/manitowoc-detailed-cleaning-sanitation-indigo-nxt.html',
    guide_url:'https://www.partstown.com/cm/resource-center/guides/gd1/how-to-clean-a-manitowoc-ice-machine',
    official_url:'https://www.manitowocice.com/Videos',
    youtube_url:'https://www.youtube.com/results?search_query=Manitowoc+ice+machine+cleaning+tutorial+evaporator+components',
    yt_component:'https://www.youtube.com/results?search_query=Manitowoc+ice+machine+inside+evaporator+water+curtain+distribution+tube+identify',
  },
  'Hoshizaki':{
    label:'Hoshizaki Official Training + Parts Town Guide',
    deep_url:'https://www.hoshizakiamerica.com/support/training/',
    guide_url:'https://www.partstown.com/cm/resource-center/guides/gd2/how-to-clean-a-hoshizaki-ice-machine',
    official_url:'https://www.hoshizakiamerica.com/support/training/',
    youtube_url:'https://www.youtube.com/results?search_query=Hoshizaki+KM+ice+machine+cleaning+tutorial+spray+bar+components',
    yt_component:'https://www.youtube.com/results?search_query=Hoshizaki+ice+machine+open+inside+spray+bar+float+switch+identify+parts',
  },
  'Scotsman':{
    label:'Scotsman Official Cleaning Guide + Easy Ice Reference',
    deep_url:'https://scotsmanhomeice.com/how-to-clean-your-scotsman-nugget-ice-machine/',
    guide_url:'https://www.easyice.com/scotsman-ice-machine-cleaning/',
    official_url:'https://scotsmanhomeice.com/blog-ice-machine-cleaning/',
    youtube_url:'https://www.youtube.com/results?search_query=Scotsman+Prodigy+ice+machine+cleaning+tutorial+thickness+sensor',
    yt_component:'https://www.youtube.com/results?search_query=Scotsman+ice+machine+open+inside+evaporator+thickness+probe+identify+parts',
  },
  'Ice-O-Matic':{
    label:'Easy Ice IOM Guide + Parts Town Reference',
    deep_url:'https://www.easyice.com/ice-o-matic-ice-machine-cleaning/',
    guide_url:'https://www.partstown.com/cm/resource-center',
    official_url:'https://www.iceomatic.com/support',
    youtube_url:'https://www.youtube.com/results?search_query=Ice-O-Matic+ice+machine+cleaning+tutorial+distribution+pan+components',
    yt_component:'https://www.youtube.com/results?search_query=Ice-O-Matic+ice+machine+open+inside+distribution+pan+evaporator+identify',
  },
  'Follett':{
    label:'Follett Official Service Video Library',
    deep_url:'https://www.follettice.com/tech-support/service-video-library',
    guide_url:'https://www.partstown.com/cm/resource-center/guides/gd2/how-to-clean-a-follett-ice-machine',
    official_url:'https://www.follettice.com/tech-support/service-video-library',
    youtube_url:'https://www.youtube.com/results?search_query=Follett+ice+machine+cleaning+tutorial+auger+cylinder+nugget',
    yt_component:'https://www.youtube.com/results?search_query=Follett+Symphony+ice+machine+open+inside+auger+evaporator+cylinder+identify',
  },
  'Cornelius':{
    label:'Parts Town Commercial Ice Machine Resources',
    deep_url:'https://www.partstown.com/cm/resource-center',
    guide_url:'https://www.partstown.com/cm/resource-center',
    official_url:'https://www.cornelius.com/support',
    youtube_url:'https://www.youtube.com/results?search_query=Cornelius+commercial+ice+machine+cleaning+tutorial+components',
    yt_component:'https://www.youtube.com/results?search_query=commercial+countertop+ice+machine+open+inside+freeze+plate+identify+components',
  },
};

const OBJECTIONS=[
  {
    q:"We already have someone who does it",
    a:"Perfect  -  can I ask who and how often they come out? The reason I ask is most places we start working with had someone too, but it was either annual or on-call. The problem is Florida inspectors now specifically open the machine during inspections  -  they want to see dated service records, not just a clean exterior. If your current vendor is leaving you a dated compliance report after every visit you are set. If not, that is the gap we fill. I can show you what ours looks like  -  takes two minutes."
  },
  {
    q:"Not interested",
    a:"Fair enough. Real quick before I go  -  when was the last time someone actually opened up the machine? Not wiped it down, but opened the evaporator cover and documented what they found? Most managers I talk to have never seen inside their own machine. I am not trying to sell you anything right now  -  I can open it up in five minutes and tell you exactly what an inspector would flag. If it is clean you never hear from me again."
  },
  {
    q:"Call me back / Not a good time",
    a:"Completely understand. I am in this area Tuesdays and Thursdays  -  which works better? And honestly a two-minute conversation now saves us both a phone tag. The short version: your inspection history flagged ice machine violations and I specialize in exactly that. If it is not a fit I will tell you right now and be out of your hair."
  },
  {
    q:"We clean it ourselves / Staff handles it",
    a:"That is great  -  and I am sure the bin and scoop are spotless. The issue is the evaporator plates and water distribution system inside the unit. Those need chemical descaling and sanitizer that food handler training does not cover. More importantly  -  inspectors know the difference. They will open it and if there is scale or biofilm on the plates that is a high priority violation regardless of how clean everything else is. That is what we document and prevent."
  },
  {
    q:"How much does it cost?",
    a:"Monthly maintenance is $149 for the first machine, $89 for a second. That covers the full clean, sanitizer treatment, and a dated compliance report every visit. One failed callback costs more in lost revenue than six months of service. And if you want to try us first, we do a $99 intro visit  -  full service, full report, no commitment. If you like it, we set up monthly. If not, no problem."
  },
  {
    q:"We just had an inspection and passed",
    a:"That is actually the best possible time to start  -  you have the most runway before the next one. Here is the thing about Florida inspectors: they rotate, and new inspectors are specifically trained to look at ice machine interiors because it is one of the most commonly missed items. Passing once does not mean the next inspector sees it the same way. What we do is get you to a standard that passes regardless of who walks in the door  -  and gives you the paperwork to prove it."
  },
];

function dL(d){return d<0?Math.abs(d)+'d overdue':d===0?'TODAY':'+'+d+'d';}
function dC(d,p){return p==='CALLBACK'||d<0?'u':d<=21?'h':d<=45?'w':'';}
function stars(r){return r>0?'\u2605'.repeat(Math.round(r))+' '+r+'/5':'';}
function enc(s){return encodeURIComponent(s);}
function hav(la1,lo1,la2,lo2){
  const R=3958.8,a=Math.sin((la2-la1)*Math.PI/360)**2+
    Math.cos(la1*Math.PI/180)*Math.cos(la2*Math.PI/180)*Math.sin((lo2-lo1)*Math.PI/360)**2;
  return R*2*Math.atan2(Math.sqrt(a),Math.sqrt(1-a));
}

let log={},tab='today',selOut=null,selType=null,selReasonVal=null,cur=null,Q='',route=[],routeSet=new Set(),mapPros=[],routeAnchor=null;
let queueList=[],queueIdx=0;

function setType(t){
  selType=t;
  document.getElementById('mtype-walkin').className='mtype-btn'+(t==='walkin'?' mtype-on':'');
  document.getElementById('mtype-call').className='mtype-btn'+(t==='call'?' mtype-on':'');
  // Update pitch section visibility
  if(t==='walkin'){setPitchMode('walkin');}else if(t==='call'){setPitchMode('phone');}
}

function lLoad(){try{log=JSON.parse(localStorage.getItem('pic_v4')||'{}')||{};}catch(e){log={};}}
function lSave(){try{localStorage.setItem('pic_v4',JSON.stringify(log));}catch(e){}}
function phLoad(){
  let saved={};try{saved=JSON.parse(localStorage.getItem('pic_phones')||'{}')||{};}catch(e){}
  P.forEach(p=>{const ph=saved[p.id]||PHONES[p.id];if(ph&&ph.phone){p.phone=ph.phone;p.rating=ph.rating||0;p.hours=ph.hours||'';}});
}
function phSave(id,phone,hours,rating){
  let s={};try{s=JSON.parse(localStorage.getItem('pic_phones')||'{}')||{};}catch(e){}
  s[id]={phone,hours,rating};try{localStorage.setItem('pic_phones',JSON.stringify(s));}catch(e){}
  const p=P.find(x=>x.id===id);if(p){p.phone=phone;p.hours=hours;p.rating=rating;}
}
function getLC(id){const e=log[id]||[];return e.length?e[e.length-1]:null;}
function isC(id){
  const entries=log[id]||[];
  if(!entries.length)return false;
  // Do not count as contacted if ALL entries are skips
  return entries.some(e=>e.notes!=='Skipped');
}

function fp(opts){
  return P.filter(p=>{
    if(Q&&!p.name.toLowerCase().includes(Q)&&!p.city.toLowerCase().includes(Q))return false;
    if(opts.county&&p.county!==opts.county)return false;
    if(opts.pri&&p.priority!==opts.pri)return false;
    if(opts.st){const lc=getLC(p.id);const l=lc?normO(lc.outcome):'not_contacted';if(opts.st==='not_contacted'){if(isC(p.id))return false;}else if(l!==opts.st)return false;}
    if(opts.ice){if(opts.ice==='chronic'&&!p.chronic)return false;if(opts.ice==='confirmed'&&!p.confirmed)return false;if(opts.ice==='high_ice'&&'medium'!=='high')return false;if(opts.ice==='phone'&&!p.phone)return false;}
    if(opts.hc&&isC(p.id))return false;
    return true;
  }).sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||b.score-a.score);
}

let _leafletLoaded=false;
function loadLeaflet(cb){
  if(typeof L!=='undefined'){cb();return;}
  if(_leafletLoaded){setTimeout(()=>loadLeaflet(cb),100);return;}
  _leafletLoaded=true;
  const lnk=document.createElement('link');
  lnk.rel='stylesheet';
  lnk.href='https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css';
  document.head.appendChild(lnk);
  const scr=document.createElement('script');
  scr.src='https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js';
  scr.onload=cb;
  scr.onerror=()=>{window._leafletFailed=true;cb();};
  document.head.appendChild(scr);
}
function sw(t){
  tab=t;
  document.querySelectorAll('.tab').forEach((el,i)=>el.classList.toggle('on',['today','all','route','customers','service','data'][i]===t));
  document.querySelectorAll('.panel').forEach(el=>el.classList.remove('on'));
  const panel=document.getElementById('p-'+t);
  if(panel)panel.classList.add('on');
  if(t==='today'){renderBriefing();}
  else if(t==='all'){populateCityFilter();rA();}
  else if(t==='route'){rRoute();}
  else if(t==='customers'){rCust();}
  // Load leaflet on route
  if(t==='route'&&typeof L==='undefined')loadLeaflet();
}


function setF(p){sw('today');rT();}
function onS(){Q=document.getElementById('si').value.toLowerCase().trim();if(tab==='today')rT();else if(tab==='all')rA();}

function cardHTML(p){
  const last=getLC(p.id);
  // String concat only - no nested backtick templates (Safari compatibility)
  const _ph=p.phone||PHONES[String(p.id)]||'';  // p.phone is set by phLoad from localStorage
  const phH=_ph
    ?('<div class="phrow"><span class="phnum">'+_ph+'</span><a href="tel:'+_ph.replace(/\s/g,'')+'" class="abtn call-a" onclick="event.stopPropagation()">Call</a></div>')
    :('<div class="phrow"><span class="phnum none">No phone on file</span><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone number')+'" target="_blank" class="abtn find-a" onclick="event.stopPropagation()">Find</a></div>'+'<div id="ph-save-'+p.id+'" style="margin-top:4px;display:flex;gap:4px;align-items:center">'  +'<input id="ph-inp-'+p.id+'" type="tel" placeholder="Paste number here..." onclick="event.stopPropagation()" '    +'style="flex:1;padding:5px 7px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'  +'<button onclick="event.stopPropagation();saveFoundPhone('+p.id+')" '    +'style="padding:5px 9px;border:none;border-radius:6px;background:var(--navy);color:#fff;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">Save</button>'+'</div>');
  const iceH=p.chronic
    ?('<div class="icebadge chronic">&#x1F9CA; CHRONIC  -  '+p.ice_count+'x ice violations'+(p.ice_fresh?' &bull; <b>recent</b>':'')+'</div>')
    :p.confirmed?'<div class="icebadge confirmed">&#x2713; Ice violation on record'+(p.ice_fresh?' (within 6mo)':p.ice_recent?' (within 1yr)':'')+'</div>':'';
  const codesH=(p.codes||[]).length?('<div style="font-size:9px;color:#2a4860;margin-bottom:4px">Codes: '+(p.codes||[]).join(', ')+'</div>'):'';
  const insH=''?('<div class="insight">'+''+'</div>'):'';
  const cbH=(p.n_callbacks>0||p.disp_risk>=4)
    ?('<div style="font-size:9px;font-weight:600;padding:3px 8px;border-radius:5px;margin-bottom:5px;background:#fef2f2;color:#dc2626;border:1px solid #fecaca">'
      +(p.n_callbacks>0?'&#x1F6A8; '+p.n_callbacks+'x callback inspection'+(p.n_callbacks>1?'s':''):'&#x26A0;&#xFE0F; Admin complaint  -  callback due')
      +'</div>')
    :'';
  const hrH=p.hours?('<div style="font-size:9px;color:#1a3850;margin-bottom:4px">&#x1F550; '+p.hours+'</div>'):'';
  const ratH=p.rating>0?('<div style="font-size:9px;color:#c08020;margin-bottom:3px">'+stars(p.rating)+'</div>'):'';
  // Best call window based on business type
  const hr=new Date().getHours();
  const callWin=p.is_bar
    ?{good:[9,10,11],avoid:[16,17,18,19,20,21,22],tip:'Best: 9-11am before they open'}
    :{good:[9,10,14,15,16],avoid:[11,12,13,17,18,19,20],tip:'Best: 9-10am or 2-4pm. Avoid lunch/dinner rush'};
  const inGood=callWin.good.includes(hr);
  const inAvoid=callWin.avoid.includes(hr);
  const callH=inGood
    ?'<div style="font-size:8px;font-weight:700;color:#059669;margin-bottom:3px">&#x1F7E2; Good time to call now</div>'
    :inAvoid
      ?'<div style="font-size:8px;font-weight:700;color:#dc2626;margin-bottom:3px">&#x1F534; Rush hour — try later</div>'
      :'';
  const lastH=last
    ?('<div class="lastc hc" style="color:'+(OI_COLOR[last.outcome]||'#64748b')+'">'+( OI[last.outcome]||last.outcome)+' &middot; '+last.date+(last.notes?'  -  '+last.notes.slice(0,25):'')+' </div>')
    :'<div class="lastc">Not yet contacted</div>';
  const trendH=p.trending?'<div class="mi"><div class="ml">Trend</div><div class="mv bad">&#x2197; Worse</div></div>':'';
  const rtBtn=p.lat?('<button class="btn brt" onclick="event.stopPropagation();addToRoute('+p.id+')">+Route</button>'):'';
  const tc={PLATINUM:'#7c3aed',GOLD:'#d97706',SILVER:'#64748b',BRONZE:'#92400e'};
  const tbg={PLATINUM:'#ede9fe',GOLD:'#fef3c7',SILVER:'#f1f5f9',BRONZE:'#fef3c7'};
  const tierH=p.tier&&p.tier!=='COLD'?(' <span style="font-size:8px;font-weight:700;padding:2px 6px;border-radius:20px;background:'+(tbg[p.tier]||'#f1f5f9')+';color:'+(tc[p.tier]||'#64748b')+'">'+p.tier+'</span>'):'';
  const revenueH=p.monthly?('<div style="font-size:10px;font-weight:700;color:#059669;margin-bottom:5px">$'+p.monthly+'/mo &bull; '+(p.machines>1?p.machines+' machines':'1 machine')+'</div>'):'';
  const emergH=p.is_emergency?'<span class="emerg-badge">&#x1F6A8; EMERGENCY</span>':'';
  const confCol=p.confidence>=75?'#059669':p.confidence>=50?'#d97706':'#9ca3af';
  const confH='<span style="font-size:8px;font-weight:600;color:'+confCol+'" title="Prediction confidence">'+p.confidence+'% conf</span>';
  const franchH=p.biz_type==='franchise'?'<span style="font-size:8px;padding:1px 5px;border-radius:10px;background:#f0f9ff;color:#0ea5e9;border:1px solid #bae6fd">Franchise</span>':'';
  const custStatusH=(p.status&&p.status!=='prospect')
    ?('<div style="font-size:9px;font-weight:700;padding:2px 8px;border-radius:20px;display:inline-block;margin-bottom:4px;background:'+(p.status==='customer_recurring'?'#ecfdf5':p.status==='customer_once'?'#eff6ff':'#fff7f5')+';color:'+(p.status==='customer_recurring'?'#059669':p.status==='customer_once'?'var(--blu)':'var(--ora)')+'">'+({'customer_recurring':'Recurring Customer','customer_once':'One-Time Customer','quoted':'Quote Sent','churned':'Churned'}[p.status]||p.status)+'</div>')
    :'';
  return '<div class="card '+p.priority+(isC(p.id)?' done':'')+'" data-id="'+p.id+'" onclick="openM('+p.id+')">'
    +'<div class="ctop"><div class="cname">'+p.name+tierH+emergH+'</div><div style="display:flex;flex-direction:column;align-items:flex-end;gap:2px"><span class="pbadge '+p.priority+'">'+p.priority+'</span>'+confH+'</div></div>'
    +'<div class="cloc">'+p.city+', '+p.county+' '+franchH+'</div>'
    +custStatusH+revenueH+phH+ratH+callH+hrH+iceH+cbH+codesH+insH
    +'<div class="cmeta">'
      +'<div class="mi"><div class="ml">Days Until</div><div class="mv '+dC(p.days_until,p.priority)+'">'+dL(p.days_until)+'</div></div>'
      +'<div class="mi"><div class="ml">H/T Viol</div><div class="mv '+(p.high_viol>=4?'u':p.high_viol>=2?'h':'')+'">'+p.high_viol+'/'+p.total_viol+'</div></div>'
      +'<div class="mi"><div class="ml">Pred.Next</div><div class="mv" style="font-size:10px">'+p.pred_next+'</div></div>'
      +trendH
    +'</div>'
    +lastH
    +'<div class="cacts">'
      +'<button class="btn blog" onclick="event.stopPropagation();openM('+p.id+')">Log Call</button>'
      +(last&&last.notes==='Skipped'
        ?'<button class="btn bskip" onclick="event.stopPropagation();unskip('+p.id+')" style="background:#ecfdf5;color:#059669;border-color:#6ee7b7">Unskip</button>'
        :'<button class="btn bskip" onclick="event.stopPropagation();skip('+p.id+')">Skip</button>')
      +rtBtn
    +'</div>'
  +'</div>';
}

function rT(){
  const county=document.getElementById('tc').value;
  const hc=document.getElementById('thc').checked;

  function filt(p){
    if(county && p.county!==county) return false;
    if(hc && isC(p.id)) return false;
    return true;
  }

  // Act Now: emergency closures + bad inspection last 30 days + callback pending
  const actNow = P.filter(p=>filt(p) && (
    p.is_emergency ||
    (p.priority==='CALLBACK' && p.days_since<=30) ||
    (p.disp_risk>=4 && p.days_since<=45)
  )).sort((a,b)=>(b.is_emergency-a.is_emergency)||(b.score-a.score));

  // Upcoming: inspection within 60 days + ice history
  const upcoming = P.filter(p=>filt(p) &&
    !actNow.find(x=>x.id===p.id) &&
    p.days_until>=0 && p.days_until<=60 &&
    (p.confirmed||p.chronic||p.ice_rel>0)
  ).sort((a,b)=>a.days_until-b.days_until);

  // High Value: PLATINUM or GOLD by revenue
  const highVal = P.filter(p=>filt(p) &&
    !actNow.find(x=>x.id===p.id) &&
    !upcoming.find(x=>x.id===p.id) &&
    (p.tier==='PLATINUM'||p.tier==='GOLD')
  ).sort((a,b)=>(b.monthly||0)-(a.monthly||0));

  const total = actNow.length + upcoming.length + highVal.length;
  document.getElementById('tcnt').textContent = total + ' prospects';

  function renderSection(ids, gridId, emptyId){
    const g = document.getElementById(gridId);
    const e = document.getElementById(emptyId);
    if(!ids.length){g.innerHTML='';e.style.display='block';return;}
    e.style.display='none';
    g.innerHTML = ids.slice(0,40).map(cardHTML).join('');
    g.querySelectorAll('.card').forEach(card=>{
      let tx=0,ty=0;
      card.addEventListener('touchstart',ev=>{tx=ev.touches[0].clientX;ty=ev.touches[0].clientY;},{passive:true});
      card.addEventListener('touchend',ev=>{
        const dx=ev.changedTouches[0].clientX-tx;
        const dy=Math.abs(ev.changedTouches[0].clientY-ty);
        if(Math.abs(dx)>70&&dy<40){
          const id=parseInt(card.getAttribute('data-id'));
          if(!id)return;
          if(dx<0)skip(id); else openM(id);
        }
      },{passive:true});
    });
  }

  renderSection(actNow,   'tgrid-actnow',  'empty-actnow');
  renderSection(upcoming, 'tgrid-upcoming','empty-upcoming');
  renderSection(highVal,  'tgrid-highval', 'empty-highval');

  // Follow-ups due today or overdue
  const todayStr=new Date().toISOString().slice(0,10);
  const followups=P.filter(p=>{
    const entries=log[p.id]||[];
    const withFollowup=entries.filter(e=>e.followup&&e.followup<=todayStr);
    return withFollowup.length>0;
  }).sort((a,b)=>{
    const fa=(log[a.id]||[]).filter(e=>e.followup).slice(-1)[0]?.followup||'';
    const fb=(log[b.id]||[]).filter(e=>e.followup).slice(-1)[0]?.followup||'';
    return fa.localeCompare(fb);
  });
  renderSection(followups,'tgrid-followups','empty-followups');
}

// ── PRESET DEFINITIONS ───────────────────────────────────────────────────────
const PRESETS = {
  all:       {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:''},
  actnow:    {county:'',pri:'CALLBACK',ice:'confirmed',tier:'',btype:'',st:'',conf:''},
  callback:  {county:'',pri:'CALLBACK',ice:'',tier:'',btype:'',st:'',conf:''},
  phone:     {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:'',hasPhone:true},
  chronic:   {county:'',pri:'',ice:'chronic',tier:'',btype:'',st:'',conf:''},
  gold:      {county:'',pri:'',ice:'',tier:'GOLD',btype:'',st:'',conf:''},
  franchise: {county:'',pri:'',ice:'',tier:'',btype:'franchise',st:'',conf:''},
  bar:       {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:'',isBar:true},
  notyet:    {county:'',pri:'',ice:'',tier:'',btype:'',st:'not_contacted',conf:''},
  freshice:  {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:'',freshIce:true},
  multicall: {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:'',multiCall:true},
};



function clearFilters(){
  setPreset('all');
}

function rA(){
  const county  = document.getElementById('ac')?.value||'';
  const city    = document.getElementById('ac-city')?.value||'';
  const pri     = document.getElementById('ap')?.value||'';
  const st      = document.getElementById('as_')?.value||'';
  const conf    = parseInt(document.getElementById('aconf')?.value||'0');

  let list=P.filter(p=>{
    if(Q&&!p.name.toLowerCase().includes(Q)&&!p.city.toLowerCase().includes(Q))return false;
    if(county&&p.county!==county)return false;
    if(city&&p.city!==city)return false;
    if(pri&&p.priority!==pri)return false;
    if(st){
      const lc=getLC(p.id);
      const norm=lc?normO(lc.outcome):'not_contacted';
      if(st==='not_contacted'){if(isC(p.id))return false;}
      else if(norm!==st)return false;
    }
    // Apply preset exclusions
    if(presetFilter&&!presetFilter(p))return false;
    return true;
  });

  // Sort: callbacks first, then by score
  list.sort((a,b)=>{
    const pa=PO[a.priority]??99, pb=PO[b.priority]??99;
    if(pa!==pb)return pa-pb;
    return (b.score||0)-(a.score||0);
  });

  document.getElementById('acnt').textContent=list.length+' prospects';
  const g=document.getElementById('agrid');
  const empty=document.getElementById('a-empty');
  if(list.length===0){g.innerHTML='';empty.style.display='flex';return;}
  empty.style.display='none';
  g.innerHTML=list.map(p=>cardHTML(p)).join('');
  g.querySelectorAll('.card').forEach(el=>{
    const id=parseInt(el.dataset.id);
    el.onclick=()=>openM(id);
  });
}

let presetFilter=null;
function setPreset(k){
  document.querySelectorAll('.preset-btn').forEach(b=>b.classList.remove('on'));
  document.getElementById('pre-'+k)?.classList.add('on');
  // Reset city/county/status filters
  const cityEl=document.getElementById('ac-city');
  if(cityEl)cityEl.value='';

  presetFilter=null;
  if(k==='actnow'){presetFilter=p=>{
    const e=getLC(p.id);
    return p.priority==='CALLBACK'||p.priority==='HOT'||(e&&e.outcome==='no_contact'&&e.followup);
  };}
  else if(k==='callback') presetFilter=p=>p.priority==='CALLBACK';
  else if(k==='phone')    presetFilter=p=>!!p.phone;
  else if(k==='chronic')  presetFilter=p=>p.chronic;
  else if(k==='notyet')   presetFilter=p=>!isC(p.id);
  else if(k==='inplay')   presetFilter=p=>{const lc=getLC(p.id);return lc&&normO(lc.outcome)==='in_play';};
  else if(k==='freshice') presetFilter=p=>p.confirmed&&!p.chronic;
  rA();
}

function clearFilters(){
  ['ac','ac-city','ap','as_'].forEach(id=>{
    const el=document.getElementById(id);if(el)el.value='';
  });
  setPreset('all');
}

function populateCityFilter(){
  const sel=document.getElementById('ac-city');
  if(!sel)return;
  const county=document.getElementById('ac')?.value||'';
  const cities=[...new Set(P
    .filter(p=>!county||p.county===county)
    .map(p=>p.city)
    .filter(Boolean)
  )].sort();
  sel.innerHTML='<option value="">All Cities</option>'+cities.map(c=>`<option>${c}</option>`).join('');
}

// ── CALL QUEUE MODE ──────────────────────────────────────────────────────
let queueProspects=[];

function enterQueueMode(){
  // Use whatever is currently filtered in All tab
  const county = document.getElementById('ac')?.value||'';
  const city   = document.getElementById('ac-city')?.value||'';
  const pri    = document.getElementById('ap')?.value||'';

  queueProspects=P.filter(p=>{
    if(county&&p.county!==county)return false;
    if(city&&p.city!==city)return false;
    if(pri&&p.priority!==pri)return false;
    if(presetFilter&&!presetFilter(p))return false;
    // Default: only not-yet-fully-resolved
    const lc=getLC(p.id);
    const norm=lc?normO(lc.outcome):'not_contacted';
    return !['signed','dead'].includes(norm);
  }).sort((a,b)=>{
    const pa=PO[a.priority]??99,pb=PO[b.priority]??99;
    return pa!==pb?pa-pb:(b.score||0)-(a.score||0);
  });

  if(!queueProspects.length){toast('No prospects match current filters');return;}
  queueIdx=0;
  document.getElementById('queue-bg').classList.add('on');
  renderQueueCard();
}

function exitQueueMode(){
  document.getElementById('queue-bg').classList.remove('on');
  rA();renderBriefing();
}

function renderQueueCard(){
  if(queueIdx>=queueProspects.length){
    document.getElementById('queue-card-wrap').innerHTML=
      '<div style="text-align:center;padding:40px 20px">'
      +'<div style="font-size:40px;margin-bottom:12px">✓</div>'
      +'<div style="font-size:16px;font-weight:700;margin-bottom:6px">Queue complete</div>'
      +'<div style="font-size:12px;color:#888">'+queueProspects.length+' prospects worked through</div>'
      +'<button onclick="exitQueueMode()" style="margin-top:16px;padding:10px 20px;background:var(--navy);color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit">Done</button>'
      +'</div>';
    document.getElementById('queue-progress').textContent='Complete!';
    document.getElementById('queue-actions').style.display='none';
    return;
  }
  const p=queueProspects[queueIdx];
  const lc=getLC(p.id);
  const norm=lc?normO(lc.outcome):'Not contacted';
  document.getElementById('queue-progress').textContent=(queueIdx+1)+' of '+queueProspects.length;
  document.getElementById('queue-card-wrap').innerHTML=
    '<div style="background:#fff;border-radius:12px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.1)">'
    +'<div style="font-size:14px;font-weight:800;color:var(--navy);margin-bottom:4px">'+p.name+'</div>'
    +'<div style="font-size:11px;color:var(--sub);margin-bottom:8px">'+p.address+', '+p.city+'</div>'
    +(p.phone?'<a href="tel:'+p.phone+'" style="display:inline-block;background:var(--grn);color:#fff;padding:7px 14px;border-radius:8px;font-size:12px;font-weight:700;text-decoration:none;margin-bottom:8px">&#x1F4DE; '+p.phone+'</a>':'<div style="font-size:11px;color:#999;margin-bottom:8px">No phone on file</div>')
    +'<div style="display:flex;gap:6px;flex-wrap:wrap">'
    +'<span style="font-size:9px;padding:2px 7px;border-radius:10px;font-weight:700;background:var(--surf);border:1px solid var(--brd);color:var(--sub)">'+p.priority+'</span>'
    +(p.city?'<span style="font-size:9px;padding:2px 7px;border-radius:10px;background:#f0f9ff;border:1px solid #bae6fd;color:#0369a1">'+p.city+'</span>':'')
    +'<span style="font-size:9px;padding:2px 7px;border-radius:10px;background:#f8fafc;border:1px solid var(--brd);color:var(--sub)">Last: '+(OI[lc?.outcome]||'Not contacted')+'</span>'
    +'</div>'
    +(lc?.notes?'<div style="font-size:10px;color:var(--sub);margin-top:8px;padding:6px 8px;background:#f5f8fa;border-radius:6px">'+lc.notes+'</div>':'')
    +'<div style="margin-top:10px"><textarea id="q-notes" placeholder="Quick note..." rows="2" style="width:100%;padding:7px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;resize:none;outline:none"></textarea></div>'
    +'</div>';
}

function queueLog(outcome){
  const p=queueProspects[queueIdx];
  if(!p)return;
  const notes=document.getElementById('q-notes')?.value||'';
  if(!log[p.id])log[p.id]=[];
  log[p.id].push({
    outcome,type:'call',reason:null,date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),
    notes,followup:''
  });
  lSave();
  queueIdx++;
  renderQueueCard();
}

function queueNext(){
  queueIdx++;
  renderQueueCard();
}



function rRoute(){
  const county=document.getElementById('rc').value;
  const pri=document.getElementById('rp').value;
  const zip=document.getElementById('rzip').value.trim();
  const rad=parseFloat(document.getElementById('rrad').value)||999;
  // Show all except customers and hard nos
  const rExclude=new Set();
  P.forEach(p=>{
    if(p.status==='customer_recurring'||p.status==='customer_once')rExclude.add(p.id);
    const entries=log[p.id]||[];
    if(entries.length&&entries[entries.length-1].outcome==='not_interested')rExclude.add(p.id);
  });
  let data=P.filter(p=>p.lat&&p.lon&&!rExclude.has(p.id));
  if(county)data=data.filter(p=>p.county===county);
  if(pri)data=data.filter(p=>p.priority===pri);
  if(zip.length===5&&ZIPS[zip]){
    const [clat,clon]=ZIPS[zip];
    data=data.map(p=>({...p,_d:hav(clat,clon,p.lat,p.lon)}))
      .filter(p=>p._d<=rad)
      .sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||a._d-b._d);
    document.getElementById('rhint').textContent=data.length+' prospects within '+rad+'mi of '+zip;
  }else{
    data.sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||b.score-a.score);
    document.getElementById('rhint').textContent=data.length+' geocoded prospects  -  enter ZIP to filter by area';
  }
  mapPros=data.slice(0,100);
  renderRList();renderMap();
}

function renderRList(){
  const el=document.getElementById('rlist');
  const cnt=document.getElementById('rlist-cnt');
  if(!mapPros.length){
    el.innerHTML='<div style="font-size:11px;color:var(--sub);padding:8px">No geocoded prospects match. Try removing ZIP filter or expanding radius.</div>';
    if(cnt)cnt.textContent='';
    return;
  }
  if(cnt)cnt.textContent=mapPros.length+' prospects nearby';
  el.innerHTML=mapPros.map(p=>{
    const inR=routeSet.has(p.id);
    const col=PC[p.priority]||'var(--sub)';
    const distTxt=p._d?p._d.toFixed(1)+'mi':'';
    const chronicTxt=p.chronic?'<div style="font-size:8px;color:#059669;font-weight:700">CHRONIC ICE</div>':'';
    const isAnchor=routeAnchor&&routeAnchor.id===p.id;
    return '<div class="rcard'+(inR?' sel':'')+'" onclick="addToRoute('+p.id+')" style="border-left:3px solid '+col+(inR?';background:#fff7f5':'')+(isAnchor?';border-left:4px solid #7c3aed':'')+';">'
      +'<div class="rdot" style="background:'+col+'"></div>'
      +'<div style="flex:1;min-width:0">'
        +'<div class="rname" style="color:var(--navy)">'+p.name+(inR?' <span style="color:var(--ora)">&#x2713;</span>':'')+(isAnchor?' <span style="font-size:8px;color:#7c3aed">&#x1F4CD; START</span>':'')+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+p.city+' &bull; '+p.priority+' &bull; '+dL(p.days_until)+'</div>'
        +chronicTxt
      +'</div>'
      +'<div style="display:flex;flex-direction:column;align-items:flex-end;gap:3px">'
        +(distTxt?'<div class="rdist">'+distTxt+'</div>':'')
        +'<button onclick="event.stopPropagation();setRouteAnchor('+p.id+')" style="font-size:8px;padding:2px 5px;border:1px solid #7c3aed;border-radius:4px;background:#f5f3ff;color:#7c3aed;cursor:pointer;font-family:inherit">Start &#x1F4CD;</button>'
        +'<button onclick="event.stopPropagation();openM('+p.id+')" style="font-size:8px;padding:2px 5px;border:1px solid var(--brd);border-radius:4px;background:var(--surf);color:var(--sub);cursor:pointer;font-family:inherit">Details</button>'
      +'</div>'

      +'</div>';
  }).join('');
}

function renderMap(){
  const area=document.getElementById('map-area');
  if(typeof L==='undefined'||window._leafletFailed){
    // Leaflet not available (sandbox or offline) - show static list
    if(!mapPros.length){area.innerHTML='<div class="map-empty"><div style="font-size:32px">&#x1F5FA;&#xFE0F;</div><div>No prospects to map</div></div>';return;}
    area.innerHTML='<div style="padding:12px;overflow-y:auto;height:100%"><div style="font-size:10px;color:var(--sub);margin-bottom:8px">Map unavailable offline. '+(mapPros.length)+' prospects:</div>'+mapPros.slice(0,20).map(p=>'<div style="padding:5px 0;border-bottom:1px solid var(--brd2);font-size:11px;color:var(--navy)">'+p.name+'</div>').join('')+'</div>';
    return;
  }
  if(!mapPros.length){
    area.innerHTML='<div class="map-empty"><div style="font-size:36px;margin-bottom:10px">&#x1F5FA;&#xFE0F;</div><div>No prospects to map</div></div>';
    if(window._lmap){window._lmap.remove();window._lmap=null;}
    return;
  }
  // Init Leaflet map once
  if(!window._lmap){
    area.innerHTML='<div id="leaflet-map" style="width:100%;height:100%;border-radius:10px"></div>';
    window._lmap=L.map('leaflet-map',{zoomControl:true,attributionControl:false});
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
      maxZoom:18,opacity:0.85
    }).addTo(window._lmap);
    window._lmarkers=L.layerGroup().addTo(window._lmap);
  }
  window._lmarkers.clearLayers();
  const cols={CALLBACK:'#ff3b30',HOT:'#ff9f0a',WARM:'#ffd60a',WATCH:'#5e9eff',LATER:'#445066'};
  const bounds=[];
  mapPros.forEach(p=>{
    const inR=routeSet.has(p.id);
    const col=inR?'#00e5ff':cols[p.priority]||'#445066';
    const r=inR?12:p.chronic?10:8;
    const marker=L.circleMarker([p.lat,p.lon],{
      radius:r,color:col,fillColor:col,fillOpacity:0.85,weight:inR?3:1.5,
      opacity:1
    }).addTo(window._lmarkers);
    marker.bindPopup(
      '<b>'+p.name+'</b><br>'+p.city+' &bull; '+p.priority+'<br>'+dL(p.days_until)+
      (p.seats?' &bull; '+p.seats+' seats':'')+
      (p.monthly?' &bull; $'+p.monthly+'/mo':'')+
      '<br><button onclick="addToRoute('+p.id+')" '+
      'style="margin-top:4px;padding:3px 8px;border:none;border-radius:5px;background:'+col+';color:'+(p.priority==='WARM'?'#000':'#fff')+';font-size:10px;font-weight:700;cursor:pointer">'+
      (routeSet.has(p.id)?'Remove from Route':'+Route')+'</button> '+
      '<button onclick="openM('+p.id+')" '+
      'style="margin-top:3px;padding:3px 8px;border:1px solid #333;border-radius:5px;background:#111;color:#fff;font-size:10px;cursor:pointer">Details</button>',
      {maxWidth:220}
    );
    bounds.push([p.lat,p.lon]);
  });
  if(bounds.length)window._lmap.fitBounds(bounds,{padding:[20,20],maxZoom:13});
}

function addToRoute(id){
  const p=P.find(x=>x.id===id);if(!p)return;
  if(routeSet.has(id)){
    routeSet.delete(id);route=route.filter(r=>r.id!==id);
    toast(p.name.slice(0,20)+' removed from route');
  } else {
    if(route.length>=8){toast('Max 8 stops. Open in Maps first.');return;}
    routeSet.add(id);route.push(p);
    toast(p.name.slice(0,20)+' added to route ('+route.length+' stops)');
  }
  renderRList();renderMap();renderDayRoute();
}
function renderDayRoute(){
  const dr=document.getElementById('day-route');
  document.getElementById('stopcnt').textContent=route.length;
  if(!route.length){dr.style.display='none';return;}
  dr.style.display='block';

  const zip=(document.getElementById('rzip')||{}).value||'';
  let curLat=null,curLon=null;
  if(zip.length===5&&ZIPS[zip]){[curLat,curLon]=ZIPS[zip];}

  let totalMi=0,totalMin=0,tmpLat=curLat,tmpLon=curLon;
  route.forEach(p=>{
    if(tmpLat&&p.lat){const d=hav(tmpLat,tmpLon,p.lat,p.lon);totalMi+=d;totalMin+=d*DRIVE_MIN_PER_MILE;}
    totalMin+=MIN_PER_STOP[p.priority]||15;
    if(p.lat){tmpLat=p.lat;tmpLon=p.lon;}
  });

  const mi=document.getElementById('route-mi');
  const te=document.getElementById('route-time-est');
  if(mi&&totalMi>0)mi.textContent=' (~'+totalMi.toFixed(1)+'mi)';
  if(te&&totalMin>0){const h=Math.floor(totalMin/60),m=Math.round(totalMin%60);te.textContent='Est. time: '+h+'h '+m+'m (drive + visit)';}

  let runMin=0,rLat=curLat,rLon=curLon;
  // If anchor is set, show it as stop 0
  const anchorRow=routeAnchor
    ?('<div class="day-stop" style="background:#f5f3ff;border:1px solid #ddd6fe;border-radius:8px;margin-bottom:4px">'
      +'<div class="stopnum" style="background:#7c3aed">&#x1F4CD;</div>'
      +'<div style="flex:1;min-width:0">'
        +'<div style="font-weight:700;font-size:11px;color:#7c3aed">START: '+routeAnchor.name+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+routeAnchor.address+', '+routeAnchor.city+'</div>'
      +'</div>'
      +'<button onclick="clearAnchor()" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">Clear</button>'
      +'</div>')
    :'';
  document.getElementById('day-stops').innerHTML=anchorRow+route.map((p,i)=>{
    let driveMin=0;
    if(rLat&&p.lat){driveMin=Math.round(hav(rLat,rLon,p.lat,p.lon)*DRIVE_MIN_PER_MILE);}
    const stopMin=MIN_PER_STOP[p.priority]||15;
    runMin+=driveMin+stopMin;
    if(p.lat){rLat=p.lat;rLon=p.lon;}
    const ph=p.phone?('<a href="tel:'+p.phone.replace(/\s/g,'')+'" style="font-size:9px;color:var(--blu);text-decoration:none">'+p.phone+'</a>'):'';
    const driveH=driveMin>0?('<span style="font-size:8px;color:var(--sub)">+'+driveMin+'m drive</span>'):'<span style="font-size:8px;color:var(--sub)">Start</span>';
    return '<div class="day-stop">'
      +'<div class="stopnum">'+(i+1)+'</div>'
      +'<div style="flex:1;min-width:0">'
        +'<div style="font-weight:700;font-size:11px;color:var(--navy);overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+p.name+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+p.address+', '+p.city+'</div>'
        +'<div style="display:flex;gap:5px;align-items:center;margin-top:2px;flex-wrap:wrap">'
          +ph+'<span style="font-size:9px;font-weight:600;color:#059669">$'+p.monthly+'/mo</span>'
          +'<span class="pbadge '+p.priority+'" style="font-size:7px">'+p.priority+'</span>'
          +driveH
        +'</div>'
      +'</div>'
      +'<div style="display:flex;flex-direction:column;gap:3px;flex-shrink:0">'
        +'<button onclick="openM('+p.id+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--ora);border-radius:5px;background:#fff7f5;color:var(--ora);cursor:pointer;font-family:inherit">Details</button>'
        +'<button onclick="removeStop('+p.id+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">Remove</button>'
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

const MIN_PER_STOP={CALLBACK:25,HOT:20,WARM:15,WATCH:12,LATER:10};
const DRIVE_MIN_PER_MILE=3;
const PW={CALLBACK:100,HOT:80,WARM:50,WATCH:20,LATER:5};

function planMyDay(){
  const zip=(document.getElementById('rzip')||{}).value.trim();
  const hours=parseFloat((document.getElementById('rtime')||{}).value)||0;
  const county=(document.getElementById('rc')||{}).value||'';
  const pri=(document.getElementById('rp')||{}).value||'';
  const rad=parseFloat((document.getElementById('rrad')||{}).value)||8;

  if(!hours){rRoute();return;}

  // Determine start coords — anchor takes priority over home ZIP
  let slat,slon,startLabel;
  if(routeAnchor&&routeAnchor.lat){
    slat=routeAnchor.lat;slon=routeAnchor.lon;
    startLabel='From: '+routeAnchor.name.slice(0,22);
  } else if(zip.length===5&&ZIPS[zip]){
    [slat,slon]=ZIPS[zip];
    startLabel='From ZIP '+zip;
  } else {
    toast('Enter a start ZIP or tap Start on a business first');return;
  }

  const budgetMin=hours*60;

  // CRITICAL: filter radius from ACTUAL start point, not home ZIP
  // Route excludes: current customers, hard nos, already on route
  // Route INCLUDES: voicemails, no answers, follow-ups, uncontacted
  const hardExclude=new Set();
  P.forEach(p=>{
    if(p.status==='customer_recurring'||p.status==='customer_once')hardExclude.add(p.id);
    const entries=log[p.id]||[];
    const lastOutcome=entries.length?entries[entries.length-1].outcome:'';
    if(lastOutcome==='not_interested')hardExclude.add(p.id);
    if(routeSet.has(p.id))hardExclude.add(p.id);
  });
  let candidates=P.filter(p=>p.lat&&p.lon&&!hardExclude.has(p.id));
  if(county)candidates=candidates.filter(p=>p.county===county);
  if(pri)   candidates=candidates.filter(p=>p.priority===pri);
  candidates=candidates
    .map(p=>({...p,_d:hav(slat,slon,p.lat,p.lon)}))
    .filter(p=>p._d<=rad);

  if(!candidates.length){
    toast('No uncontacted prospects within '+rad+'mi. Try increasing radius.');
    return;
  }

  route=[];routeSet=new Set();
  let usedMin=0,curLat=slat,curLon=slon;
  const remaining=[...candidates];

  while(remaining.length&&usedMin<budgetMin){
    let bestIdx=-1,bestVal=-Infinity;
    remaining.forEach((p,i)=>{
      const dMin=hav(curLat,curLon,p.lat,p.lon)*DRIVE_MIN_PER_MILE;
      const sMin=MIN_PER_STOP[p.priority]||15;
      if(usedMin+dMin+sMin>budgetMin)return;
      // Boost uncontacted and voicemail/no-answer (good walk-in candidates)
      const entries=log[p.id]||[];
      const lastOutcome=entries.length?entries[entries.length-1].outcome:'';
      const contactBoost=lastOutcome===''?5
        :lastOutcome==='no_answer'||lastOutcome==='voicemail'?8
        :lastOutcome==='follow_up'||lastOutcome==='interested'?12
        :-2; // recently logged other outcome
      const val=(PW[p.priority]||0)+p.score+contactBoost-(dMin*2);
      if(val>bestVal){bestVal=val;bestIdx=i;}
    });
    if(bestIdx===-1)break;
    const chosen=remaining.splice(bestIdx,1)[0];
    const dMin=hav(curLat,curLon,chosen.lat,chosen.lon)*DRIVE_MIN_PER_MILE;
    usedMin+=dMin+(MIN_PER_STOP[chosen.priority]||15);
    route.push(chosen);routeSet.add(chosen.id);
    curLat=chosen.lat;curLon=chosen.lon;
  }

  mapPros=candidates.slice(0,100);
  renderRList();renderMap();renderDayRoute();
  const h=Math.floor(usedMin/60),m=Math.round(usedMin%60);
  const hint=document.getElementById('rhint');
  if(hint)hint.textContent=startLabel+' • '+route.length+' stops • est. '+h+'h '+m+'m of '+hours+'h';
  toast('Built '+route.length+' stops near '+startLabel);
}


function optimizeRoute(){
  if(route.length<2){toast('Add at least 2 stops to optimize');return;}
  if(route.length===2){toast('Route sorted');return;}
  const stops=[...route];
  const ordered=[stops.shift()];
  while(stops.length){
    const last=ordered[ordered.length-1];
    if(!last.lat){ordered.push(stops.shift());continue;}
    let best=0,bestD=Infinity;
    stops.forEach((s,i)=>{if(!s.lat)return;const d=hav(last.lat,last.lon,s.lat,s.lon);if(d<bestD){bestD=d;best=i;}});
    ordered.push(stops.splice(best,1)[0]);
  }
  route=ordered;routeSet=new Set(route.map(r=>r.id));
  renderRList();renderMap();renderDayRoute();
  toast('Route sorted by distance');
}

function openMaps(){
  if(!route.length)return;
  const url='https://www.google.com/maps/dir/'+route.filter(p=>p.address).map(p=>enc(p.address+', '+p.city+', FL '+p.zip)).join('/');
  window.open(url,'_blank');
}
function clearRoute(){route=[];routeSet=new Set();routeAnchor=null;renderRList();renderMap();renderDayRoute();}
function clearAnchor(){
  routeAnchor=null;
  renderDayRoute();renderRList();
  toast('Start anchor cleared');
}
function setRouteAnchor(id){
  const p=P.find(x=>x.id===id);if(!p||!p.lat){toast('This business has no map coordinates');return;}
  routeAnchor=p;
  // Update the ZIP field to show anchor is set
  const rzip=document.getElementById('rzip');
  if(rzip)rzip.placeholder='Starting from: '+p.name.slice(0,20)+'...';
  renderRList();
  toast('Route will start from: '+p.name.slice(0,25));
}
function removeStop(id){routeSet.delete(id);route=route.filter(r=>r.id!==id);renderRList();renderMap();renderDayRoute();}

// MODAL
function openM(id){
  const p=P.find(x=>x.id===id);if(!p)return;
  cur=p;selOut=null;
  document.querySelectorAll('.obtn').forEach(b=>b.classList.remove('on'));
  selType=null;selReasonVal=null;
  document.getElementById('mtype-walkin').className='mtype-btn';
  document.getElementById('mtype-call').className='mtype-btn';
  const rw2=document.getElementById('reason-wrap');if(rw2)rw2.style.display='none';
  document.getElementById('mnotes').value='';
  document.getElementById('mn').textContent=p.name;
  document.getElementById('ml').textContent=p.address+', '+p.city+', FL '+p.zip+' \u00b7 #'+p.id;
  const pe=document.getElementById('mph'),ae=document.getElementById('mpa');
  if(p.phone){
    pe.textContent=p.phone;pe.className='mphnum';
    const cl=p.phone.replace(/\\s/g,'');
    ae.innerHTML='<a href="tel:'+cl+'" class="mcall">Call Now</a><a href="sms:'+cl+'" class="msms">Text</a><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL')+'" target="_blank" class="mgoog">Google</a>';
  }else{
    pe.textContent='No phone on file';pe.className='mphnum none';
    ae.innerHTML='<a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone number')+'" target="_blank" class="mgoog">Find Phone</a><a href="https://maps.google.com/search?q='+enc(p.name+' '+p.address+' '+p.city+' FL')+'" target="_blank" class="mgoog">Maps</a>';
  }
  document.getElementById('mhrs').textContent=p.hours?'Hours: '+p.hours:'';
  document.getElementById('mrat').textContent=p.rating>0?stars(p.rating):'';
  const iceEl=document.getElementById('mice');
  if(p.confirmed||p.chronic){
    const codes=(p.codes||[]).map(c=>ICN[c]||c).join(', ');
    const parts=['<div class="micebox">'];
    parts.push('<div style="font-size:8px;color:#013a18;letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px">&#x1F9CA; ICE MACHINE HISTORY</div>');
    if(p.chronic) parts.push('<div style="font-size:13px;font-weight:700;color:var(--grn);margin-bottom:3px">CHRONIC  -  flagged in '+p.ice_count+' inspections</div>');
    if(!p.chronic&&p.confirmed) parts.push('<div style="font-size:12px;color:#55c8ff;margin-bottom:3px">Confirmed ice machine violation on record</div>');
    if(codes) parts.push('<div style="font-size:9px;color:#2a5a38;margin-bottom:3px">Violation codes: '+codes+'</div>');
    if('') parts.push('<div style="font-size:10px;color:#3a6a48;font-style:italic">'+''+'</div>');
    parts.push('</div>');
    iceEl.innerHTML=parts.join('');
  }else iceEl.innerHTML='';
  document.getElementById('mpitch').innerHTML=(PITCHES[p.pitch_type]||PITCHES.routine)(p.name);
  document.getElementById('mwalkin').innerHTML=(WALKIN[p.pitch_type]||WALKIN.routine)(p.name);
  // Objection handlers
  const objEl=document.getElementById('mobjections');
  if(objEl){
    objEl.innerHTML=OBJECTIONS.map((o,i)=>
      '<details style="background:#f5f8fa;border:1px solid var(--brd2);border-radius:7px;padding:0">'
      +'<summary style="padding:7px 10px;font-size:10px;font-weight:600;color:var(--navy);cursor:pointer;list-style:none;display:flex;align-items:center;gap:6px">'
      +'<span style="color:var(--cb)">&#x2753;</span>'+o.q+'</summary>'
      +'<div style="padding:4px 10px 9px;font-size:10px;color:var(--sub2);line-height:1.6;border-top:1px solid var(--brd2)">'+o.a+'</div>'
      +'</details>'
    ).join('');
  }
  // Build intelligence signals display
  const iceFreshness=p.ice_fresh?'Within 6mo':p.ice_recent?'Within 1yr':p.days_since_ice<999?(Math.floor(p.days_since_ice/30)+'mo ago'):'None on record';
  const iceFreshnessCol=p.ice_fresh?'r':p.ice_recent?'o':'';
  const escalationTxt=p.escalation>1.5?'Escalating fast':p.escalation>0.8?'Getting worse':p.escalation>0.3?'Slight trend up':'Stable';
  const escalationCol=p.escalation>1.5?'r':p.escalation>0.8?'o':'g';
  document.getElementById('mfacts').innerHTML=[
    // Timing
    ['Predicted Next',   p.pred_next,''],
    ['Days Until',       dL(p.days_until),''],
    ['Last Inspected',   p.last_insp,''],
    ['Days Since Insp.', p.days_since+'d',''],
    ['Inspections on File',p.n_insp,''],
    // Ice intelligence
    ['Ice Violations',   p.ice_count>0?p.ice_count+'x flagged':'None on record',p.ice_count>0?'r':''],
    ['Last Ice Violation',iceFreshness,iceFreshnessCol],
    ['Violation Breadth',p.code_diversity>0?p.code_diversity+' violation types':'None',p.code_diversity>=3?'r':p.code_diversity>=2?'o':''],
    // Inspection behavior
    ['Callback Inspections',p.n_callbacks>0?p.n_callbacks+'x (inspector returning)':'None',p.n_callbacks>=2?'r':p.n_callbacks===1?'o':''],
    ['Disposition Trend', escalationTxt,escalationCol],
    ['Failed First Visit',p.avg_visit>1.2?'Yes (avg '+p.avg_visit+'x visits)':'No',''],
    ['High/Total Viol',  p.high_viol+'/'+p.total_viol,p.high_viol>=4?'r':p.high_viol>=2?'o':''],
    ['Viol. Trajectory', p.trending?'Getting Worse':'Stable',p.trending?'r':'g'],
    // Account
    ['Confidence',       (p.confidence||0)+'%',p.confidence>=75?'g':p.confidence>=50?'o':''],
    ['Est. Machines',    p.machines||1,''],
    ['Account Tier',     p.tier||'COLD',''],
    ['Monthly Recurring','$'+(p.monthly||149)+'/mo','g'],
    ['One-Time Clean',   '$'+(p.onetime||249),'b'],
    ['Intro Offer',      '$99 first visit (no commitment)','o'],
  ].map(([l,v,c])=>'<div class="fact"><div class="fl">'+l+'</div><div class="fv '+c+'">'+v+'</div></div>').join('');
  const hist=log[id]||[];
  const hs=document.getElementById('mhs');
  if(hist.length){hs.style.display='block';document.getElementById('mhist').innerHTML=[...hist].reverse().map(e=>{
    const noteHtml=e.notes?('<div style="color:var(--sub);font-size:9px;margin-top:1px">'+e.notes+'</div>'):'';
    return '<div class="hi">'+(OI[e.outcome]||e.outcome)+'<span style="color:var(--sub);font-size:9px"> &middot; '+e.date+'</span>'+noteHtml+'</div>';
  }).join('');}
  else hs.style.display='none';
  // Show tiered pricing breakdown
  const machineBreakdown=p.machines<=1?'1 machine':
    p.machines===2?'2 machines ($149 + $89)':
    p.machines+'machines ($149 + $89 + $69×'+(p.machines-2)+')';
  document.getElementById('mwon-monthly').textContent='$'+p.monthly+'/mo  -  '+machineBreakdown;
  document.getElementById('mwon-onetime').textContent='$'+p.onetime+' one-time ('+p.machines+' machine'+(p.machines>1?'s':'')+')';
  renderContacts(id);
  renderVendor(id);
  document.getElementById('mbg').classList.add('on');
}
function closeM(e){
  if(!e||e.target.id==='mbg')document.getElementById('mbg').classList.remove('on');
}
function closeMForce(){
  document.getElementById('mbg').classList.remove('on');
}
function setPitchMode(mode){
  const phoneBtn=document.getElementById('btn-phone');
  const walkinBtn=document.getElementById('btn-walkin');
  const phoneDiv=document.getElementById('mpitch');
  const walkinDiv=document.getElementById('mwalkin');
  if(mode==='walkin'){
    phoneDiv.style.display='none';walkinDiv.style.display='block';
    phoneBtn.style.background='transparent';phoneBtn.style.color='var(--sub)';phoneBtn.style.borderColor='var(--brd)';
    walkinBtn.style.background='#0a84ff22';walkinBtn.style.color='var(--blu)';walkinBtn.style.borderColor='var(--blu)';
  }else{
    phoneDiv.style.display='block';walkinDiv.style.display='none';
    phoneBtn.style.background='#0a84ff22';phoneBtn.style.color='var(--blu)';phoneBtn.style.borderColor='var(--blu)';
    walkinBtn.style.background='transparent';walkinBtn.style.color='var(--sub)';walkinBtn.style.borderColor='var(--brd)';
  }
}
function selO(o){
  selOut=o;selReasonVal=null;
  document.querySelectorAll('.obtn').forEach(b=>b.classList.toggle('on',b.dataset.o===o));
  // Show/hide reason picker
  const rw=document.getElementById('reason-wrap');
  const rg=document.getElementById('reason-grid');
  if(rw&&rg){
    if(REASON_OUTCOMES.has(o)){
      rg.innerHTML=Object.entries(REASONS).map(([k,v])=>
        `<button class="reason-chip" onclick="selReason('${k}',this)">${v}</button>`
      ).join('');
      rw.style.display='block';
    } else {
      rw.style.display='none';
      rg.innerHTML='';
    }
  }
}
function selReason(r,el){
  selReasonVal=r;
  document.querySelectorAll('.reason-chip').forEach(c=>c.classList.remove('on'));
  if(el)el.classList.add('on');
}
function saveL(){
  if(!cur){toast('No business selected');return;}
  // Default to no_answer if no outcome selected
  if(!selOut){
    selOut='no_contact';
    document.querySelectorAll('.obtn').forEach(b=>b.classList.toggle('on',b.dataset.o==='no_contact'));
  }
  const notes=document.getElementById('mnotes').value.trim();
  const followup=(document.getElementById('mfollowup')||{}).value||'';
  const e={outcome:selOut,type:selType||null,reason:selReasonVal||null,
    date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),notes,followup};
  if(!log[cur.id])log[cur.id]=[];
  log[cur.id].push(e);lSave();
  document.getElementById('mbg').classList.remove('on');
  if(document.getElementById('mfollowup'))document.getElementById('mfollowup').value='';
  const msg=followup
    ?'✓ '+(OI[selOut]||selOut)+' logged  •  Follow-up set for '+followup
    :'✓ '+(OI[selOut]||selOut)+' logged for '+cur.name.slice(0,20);
  toast(msg);
  if(tab==='today'){rT();renderBriefing();}else if(tab==='all')rA();renderBriefing();
}
function skip(id){
  if(!log[id])log[id]=[];
  log[id].push({outcome:'no_answer',date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),notes:'Skipped'});
  lSave();rT();
}
function unskip(id){
  if(!log[id])return;
  // Remove last entry if it was a skip
  const last=log[id][log[id].length-1];
  if(last&&last.notes==='Skipped')log[id].pop();
  if(!log[id].length)delete log[id];
  lSave();if(tab==='today')rT();else if(tab==='all')rA();
  toast('Removed from skipped');
}

// PIPELINE
function statBox(label,val,unit,col){
  return '<div style="background:var(--surf);border:1px solid var(--brd);border-radius:8px;padding:8px;text-align:center">'
    +'<div style="font-size:20px;font-weight:800;color:'+col+'">'+val+'</div>'
    +'<div style="font-size:8px;color:var(--sub)">'+unit+'</div>'
    +'<div style="font-size:9px;font-weight:600;color:var(--navy)">'+label+'</div>'
    +'</div>';
}
function funnelRow(label,val,total,col){
  const pct=total>0?Math.round(val/total*100):0;
  const w=Math.max(4,pct);
  return '<div style="margin-bottom:5px">'
    +'<div style="display:flex;justify-content:space-between;font-size:10px;margin-bottom:2px">'
      +'<span style="font-weight:600;color:var(--navy)">'+label+'</span>'
      +'<span style="color:var(--sub)">'+val+' ('+pct+'%)</span>'
    +'</div>'
    +'<div style="height:6px;background:var(--brd2);border-radius:3px">'
      +'<div style="height:6px;width:'+w+'%;background:'+col+';border-radius:3px;transition:.3s"></div>'
    +'</div>'
    +'</div>';
}
function outcomeBreakdown(entries){
  const counts={};
  entries.forEach(e=>{counts[e.outcome]=(counts[e.outcome]||0)+1;});
  return Object.entries(counts).sort((a,b)=>b[1]-a[1])
    .map(([k,v])=>'<div style="display:flex;justify-content:space-between;font-size:10px;padding:2px 0">'
      +'<span style="color:var(--sub)">'+( OI[k]||k)+'</span>'
      +'<span style="font-weight:600;color:var(--navy)">'+v+'</span>'
      +'</div>'
    ).join('');
}
function rPipe(){
  const now=new Date();
  const weekAgo=new Date(now-7*864e5);
  const monthAgo=new Date(now-30*864e5);

  const allEntries=[];
  Object.entries(log).forEach(([id,entries])=>{
    entries.forEach(e=>allEntries.push({...e,id:parseInt(id)}));
  });
  function parseDate(s){
    if(!s)return null;
    const d=new Date(s);
    return isNaN(d)?null:d;
  }
  function inWindow(e,since){
    const d=parseDate(e.date);
    return d&&d>=since;
  }

  const weekEntries=allEntries.filter(e=>inWindow(e,weekAgo));
  const monthEntries=allEntries.filter(e=>inWindow(e,monthAgo));

  const totalContacted=Object.keys(log).filter(id=>log[id].length>0).length;
  const totalInterested=Object.values(log).filter(entries=>entries.some(e=>e.outcome==='interested'||e.outcome==='scheduled')).length;
  const totalScheduled=Object.values(log).filter(entries=>entries.some(e=>e.outcome==='scheduled')).length;
  const totalWon=Object.values(log).filter(entries=>entries.some(e=>['customer_recurring','customer_once'].includes(e.outcome))).length;

  const vendorCounts={};
  Object.values(vendors).forEach(v=>{
    const key=(v||'').toLowerCase().trim();
    if(key)vendorCounts[key]=(vendorCounts[key]||0)+1;
  });
  const topVendors=Object.entries(vendorCounts).sort((a,b)=>b[1]-a[1]).slice(0,4);

  const c=document.getElementById('pipec');

  c.innerHTML=
    '<div style="margin-bottom:16px">'
    +'<div style="font-weight:800;font-size:13px;color:var(--navy);margin-bottom:10px">&#x1F4CA; Activity Dashboard</div>'
    +'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:10px">'
      +statBox('This Week',weekEntries.length,'calls','var(--blu)')
      +statBox('This Month',monthEntries.length,'calls','var(--navy)')
      +statBox('Contacted',totalContacted,'total','var(--sub)')
    +'</div>'
    +'<div style="background:var(--surf);border:1px solid var(--brd);border-radius:9px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Conversion Funnel</div>'
      +funnelRow('Contacted',totalContacted,totalContacted,'#64748b')
      +funnelRow('Interested',totalInterested,totalContacted,'var(--blu)')
      +funnelRow('Scheduled',totalScheduled,totalContacted,'var(--ora)')
      +funnelRow('Closed Won',totalWon,totalContacted,'#059669')
    +'</div>'
    +(weekEntries.length?
      '<div style="background:var(--surf);border:1px solid var(--brd);border-radius:9px;padding:10px;margin-bottom:10px">'
        +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">This Week by Outcome</div>'
        +outcomeBreakdown(weekEntries)
      +'</div>'
    :'')
    +(topVendors.length>=2?
      '<div style="background:#fef9ee;border:1px solid #fde68a;border-radius:9px;padding:10px;margin-bottom:10px">'
        +'<div style="font-size:9px;font-weight:700;color:#92400e;text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">&#x1F575; Competitor Intel</div>'
        +topVendors.map(([v,n])=>'<div style="display:flex;justify-content:space-between;font-size:10px;padding:3px 0;border-bottom:1px solid #fde68a"><span style="color:#78350f;font-weight:600;text-transform:capitalize">'+v+'</span><span style="color:#92400e">'+n+'x</span></div>').join('')
        +'<div style="font-size:9px;color:#92400e;margin-top:4px">Track who you are displacing. Log vendors on each business card.</div>'
      +'</div>'
    :'')
    +'</div>'
    +'<div style="font-weight:800;font-size:13px;color:var(--navy);margin-bottom:10px">&#x1F4CB; Pipeline by Status</div>';

  const grps={
    scheduled:{l:'Scheduled / Booked',c:'#32d74b',ids:[]},
    interested:{l:'Interested',c:'#0a84ff',ids:[]},
    follow_up:{l:'Follow Up Needed',c:'#ffd60a',ids:[]},
    voicemail:{l:'Voicemail Left',c:'#5e9eff',ids:[]},
    no_answer:{l:'No Answer',c:'#445066',ids:[]},
    not_interested:{l:'Not Interested',c:'#ff3b30',ids:[]},
    service_done:{l:'Service Done',c:'#059669',ids:[]},
  };
  Object.entries(log).forEach(([id,entries])=>{if(!entries.length)return;const l=entries[entries.length-1];if(grps[l.outcome])grps[l.outcome].ids.push(parseInt(id));});
  const pm={};P.forEach(p=>pm[p.id]=p);
  const hasAny=Object.values(grps).some(g=>g.ids.length>0);
  if(!hasAny){c.innerHTML+='<div class="tempty"><div class="ei">&#x1F4DE;</div><div>Start logging calls to build your pipeline</div></div>';return;}
  c.innerHTML+=Object.entries(grps).map(([k,g])=>{
    if(!g.ids.length)return'';
    const items=g.ids.map(id=>{
      const p=pm[id];if(!p)return'';
      const last=(log[id]||[]).slice(-1)[0];
      return '<div class="pitem" onclick="openM('+id+')">'
        +'<div class="pdot" style="background:'+g.c+'"></div>'
        +'<div style="flex:1">'
          +'<div class="piname">'+p.name+'</div>'
          +'<div style="font-size:9px;color:var(--sub)">'+p.city+', '+p.county+' &bull; '+dL(p.days_until)+' &bull; '+last.date+'</div>'
          +(p.phone?('<div class="piph">'+p.phone+'</div>'):'')
          +(last.notes?('<div style="font-size:9px;color:var(--sub);margin-top:1px">'+last.notes.slice(0,50)+'</div>'):'')
          +(last.followup?('<div style="font-size:9px;font-weight:600;color:var(--ora)">&#x1F4C5; Follow-up: '+last.followup+'</div>'):'')
        +'</div>'
        +'<span class="pbadge '+p.priority+'">'+p.priority+'</span>'
        +'</div>';
    }).join('');
    return '<div class="psect"><div class="pst" style="color:'+g.c+'">'+g.l+'<span class="pct">'+g.ids.length+'</span></div>'+items+'</div>';
  }).join('');
}

// ADD PHONE
function addPhone(){
  const id=parseInt(document.getElementById('ph-id').value.trim());
  const phone=document.getElementById('ph-num').value.trim();
  const hours=document.getElementById('ph-hrs').value.trim();
  if(!id||!phone){toast('Enter a License ID and phone number');return;}
  phSave(id,phone,hours,0);
  document.getElementById('ph-id').value='';
  document.getElementById('ph-num').value='';
  document.getElementById('ph-hrs').value='';
  toast('Phone saved for #'+id);
  if(tab==='today')rT();
}

// EXPORT
function exportCSV(){
  const rows=[['Name','County','City','Phone','Hours','Rating','Priority','Days Until','Pred Next','High Viol','Total Viol','Ice Profile','Codes','Trend','Last Outcome','Date','Notes']];
  P.forEach(p=>{
    const l=getLC(p.id);
    rows.push([p.name,p.county,p.city,p.phone||'',p.hours||'',p.rating||'',p.priority,
      p.days_until,p.pred_next,p.high_viol,p.total_viol,
      p.chronic?'CHRONIC':p.confirmed?'confirmed':'none',
      (p.codes||[]).join(';'),p.trending?'worse':'stable',
      l?OI[l.outcome]||l.outcome:'',l?l.date:'',l?l.notes:'']);
  });
  const csv=rows.map(r=>r.map(c=>'"'+String(c).replace(/"/g,'""')+'"').join(',')).join('\\n');
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
  a.download='pic_prospects_'+new Date().toISOString().slice(0,10)+'.csv';
  a.click();toast('CSV exported');
}
function clrLog(){
  if(!confirm('Clear all call log data? Cannot be undone.'))return;
  log={};lSave();
  if(tab==='today'){rT();renderBriefing();}
  if(tab==='pipe')rPipe();
  toast('Call log cleared');
}
function clrCustomers(){
  if(!confirm('Clear all customer data? Cannot be undone.'))return;
  customers={};custSave();
  P.forEach(p=>{if(p.status!=='prospect')p.status='prospect';});
  if(tab==='customers')rCust();
  renderBriefing();
  toast('Customer data cleared');
}
function clrAll(){
  if(!confirm('Reset ALL data? Clears log, customers, contacts, vendors, goals. Cannot be undone.'))return;
  log={};lSave();
  customers={};custSave();
  contacts={};contactsSave();
  vendors={};vendorsSave();
  P.forEach(p=>p.status='prospect');
  if(tab==='today'){rT();renderBriefing();}
  
  else if(tab==='customers')rCust();
  else if(tab==='service')rService();
  else{rT();renderBriefing();}
  toast('All data cleared. Reloading...');
  setTimeout(()=>window.location.reload(),1200);
}
function toast(msg){const t=document.getElementById('toast');t.textContent=msg;t.classList.add('on');setTimeout(()=>t.classList.remove('on'),2200);}

// ── CUSTOMER LIFECYCLE ────────────────────────────────────────────────────────
let customers = {};  // {id: {status, won_date, service_type, notes, last_service}}

function custLoad(){
  try{customers=JSON.parse(localStorage.getItem('pic_customers')||'{}')||{};}catch(e){customers={};}
  // Apply saved statuses to P records
  P.forEach(p=>{const c=customers[p.id];if(c)p.status=c.status;});
}
function custSave(){try{localStorage.setItem('pic_customers',JSON.stringify(customers));}catch(e){}}

function markWon(status){
  if(!cur)return;
  const p=cur;
  const now=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  customers[p.id]={
    status,
    won_date: now,
    service_type: status==='customer_recurring'?'recurring':status==='customer_intro'?'intro':'one_time',
    monthly: status==='customer_recurring'?p.monthly:0,
    onetime: status==='customer_once'?p.onetime:status==='customer_intro'?99:0,
    machines: p.machines,
    name: p.name,
    address: p.address,
    city: p.city,
    phone: p.phone,
    notes: '',
    last_service: '',
    next_service: '',
    hubspot_url: '',
    square_url: '',
    machine_brand: '',
    machine_model: '',
    machine_type: '',
    filter_type: '',
    filter_installed: '',
    contract_start: '',
    contract_term: 6,
    contract_renewal: '',
    service_history: [],
    atp_history: [],
  };
  p.status=status;
  custSave();
  // Log it
  if(!log[p.id])log[p.id]=[];
  log[p.id].push({outcome:status,date:now,notes:'Deal closed'});
  lSave();
  if(status==='quoted'){
    toast('Marked as Quoted — create the quote in HubSpot.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  } else if(status==='customer_intro'){
    buildAnnualSchedule(p.id);
    toast('Intro offer booked! $99 first visit. Follow up to convert to recurring.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  } else {
    if(status==='customer_recurring')buildAnnualSchedule(p.id);
    toast(status==='customer_recurring'?'Won! Added to recurring customers.':'Won! One-time service recorded.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  }
}

function rCust(){
  const filter=document.getElementById('cust-status').value;
  const allCusts=P.filter(p=>p.status&&p.status!=='prospect');
  const shown=filter?allCusts.filter(p=>p.status===filter):allCusts;

  // MRR calculation
  const recurring=allCusts.filter(p=>p.status==='customer_recurring');
  const mrr=recurring.reduce((s,p)=>s+(p.monthly||149),0);
  document.getElementById('mrr-val').textContent='$'+mrr.toLocaleString();
  document.getElementById('cust-count').textContent=recurring.length;
  document.getElementById('arr-val').textContent='$'+(mrr*12).toLocaleString();

  const el=document.getElementById('cust-list');
  const em=document.getElementById('cust-empty');
  document.getElementById('cust-cnt').textContent=shown.length+' accounts';

  if(!shown.length){el.innerHTML='';em.style.display='block';return;}
  em.style.display='none';

  const STATUS_LABELS={
    customer_recurring:'&#x1F504; Recurring',customer_once:'&#x1F9FC; One-Time',
    customer_intro:'&#x1F525; Intro ($99)',
    quoted:'&#x1F4C4; Quote Sent',churned:'&#x274C; Churned'
  };
  const STATUS_COLORS={
    customer_recurring:'#059669',customer_once:'var(--blu)',
    customer_intro:'var(--ora)',
    quoted:'#7c3aed',churned:'var(--cb)'
  };

  // Portal ID for HubSpot links
  const portalId=loadSettings().hubspot_portal||'';

  const today=new Date();

  el.innerHTML=shown.map(p=>{
    const c=customers[p.id]||{};
    const col=STATUS_COLORS[p.status]||'var(--sub)';
    const lbl=STATUS_LABELS[p.status]||p.status;
    const rev=p.status==='customer_recurring'?('$'+p.monthly+'/mo'):p.status==='customer_once'?('$'+p.onetime+' one-time'):'';

    // Service due indicator
    let serviceDueH='';
    if(c.next_service){
      const due=new Date(c.next_service);
      const daysUntil=Math.round((due-today)/864e5);
      const dueCol=daysUntil<0?'#dc2626':daysUntil<=7?'#d97706':'#059669';
      const dueTxt=daysUntil<0?Math.abs(daysUntil)+'d overdue':daysUntil===0?'Due TODAY':daysUntil+'d until service';
      serviceDueH='<div style="font-size:9px;font-weight:700;color:'+dueCol+'">&#x1F4C5; '+dueTxt+'</div>';
    }

    // HubSpot link
    const hsUrl=portalId
      ?('https://app.hubspot.com/contacts/'+portalId+'/company/create?name='+enc(p.name)+'&city='+enc(p.city))
      :'https://app.hubspot.com';

    return '<div class="pitem" style="border-left:3px solid '+col+';flex-direction:column;align-items:stretch;gap:8px">'
      // Top row
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+', '+p.county+(p.phone?' &bull; '+p.phone:'')+'</div>'
          +'<div style="font-size:9px;font-weight:700;color:'+col+';margin-top:2px">'+lbl+'</div>'
        +'</div>'
        +'<div style="text-align:right;flex-shrink:0;margin-left:10px">'
          +(rev?'<div style="font-size:13px;font-weight:800;color:var(--grn)">'+rev+'</div>':'')
          +(c.won_date?'<div style="font-size:9px;color:var(--sub)">Since '+c.won_date+'</div>':'')
        +'</div>'
      +'</div>'
      // Service tracking row
      +'<div style="background:#f5f8fa;border-radius:7px;padding:8px;display:flex;flex-direction:column;gap:5px">'
        +'<div style="display:flex;justify-content:space-between;align-items:center">'
          +'<span style="font-size:9px;color:var(--sub);font-weight:600;text-transform:uppercase">Last Service</span>'
          +'<span style="font-size:10px;font-weight:600;color:var(--navy)">'+(c.last_service||'Not recorded')+'</span>'
        +'</div>'
        +serviceDueH
        +'<div style="display:flex;gap:5px;margin-top:3px">'
          +'<button onclick="event.stopPropagation();openServiceLog('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:none;border-radius:6px;background:var(--grn);color:#fff;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Log Visit</button>'
          +'<button onclick="event.stopPropagation();setNextService('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--sub);cursor:pointer;font-family:inherit">Set Next Due</button>'
        +'</div>'
      +'</div>'
      // Action buttons
      +'<div style="display:flex;gap:5px">'
        +'<button onclick="event.stopPropagation();openM('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--navy);cursor:pointer;font-family:inherit">Details</button>'
        +(c.hubspot_url
          ?'<a href="'+c.hubspot_url+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #ff7a59;border-radius:6px;background:#fff7f5;color:#ff7a59;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">HubSpot &#x2197;</a>'
          :'<a href="'+hsUrl+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #ff7a59;border-radius:6px;background:#fff7f5;color:#ff7a59;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">+ HubSpot</a>')
        +(c.square_url
          ?'<a href="'+c.square_url+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #00c058;border-radius:6px;background:#f0fff4;color:#00c058;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">Square &#x2197;</a>'
          :'')
      +'</div>'
      // URL link fields
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x1F517; Link Records</div>'
        +'<div style="display:flex;gap:5px;align-items:center">'
          +'<span style="font-size:9px;color:#ff7a59;font-weight:600;width:52px;flex-shrink:0">HubSpot</span>'
          +'<input type="url" placeholder="Paste HubSpot company URL..." value="'+(c.hubspot_url||'')+'"'
            +' onchange="saveHsUrl('+p.id+',this.value)"'
            +' onclick="event.stopPropagation()"'
            +' style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
        +'</div>'
        +'<div style="display:flex;gap:5px;align-items:center">'
          +'<span style="font-size:9px;color:#00c058;font-weight:600;width:52px;flex-shrink:0">Square</span>'
          +'<input type="url" placeholder="Paste Square payment link URL..." value="'+(c.square_url||'')+'"'
            +' onchange="saveSqUrl('+p.id+',this.value)"'
            +' onclick="event.stopPropagation()"'
            +' style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
        +'</div>'
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

function saveFoundPhone(id){
  const inp=document.getElementById('ph-inp-'+id);
  if(!inp)return;
  const v=inp.value.trim();
  if(!v){toast('Enter a phone number first');return;}
  const digits=v.replace(/[^\d+\-().\s]/g,'').trim();
  // Save to localStorage and update P array
  phSave(id, digits, '', 0);
  // Also update p.phone directly so cardHTML reads it correctly on re-render
  const pEntry=P.find(x=>x.id===id);
  if(pEntry)pEntry.phone=digits;
  // Update card display without full re-render
  const card=document.querySelector('[data-id="'+id+'"]');
  if(card){
    const row=card.querySelector('.phrow');
    if(row)row.innerHTML='<a href="tel:'+digits.replace(/\s/g,'')+'" class="phnum" onclick="event.stopPropagation()">'+digits+'</a>';
    const saveRow=document.getElementById('ph-save-'+id);
    if(saveRow)saveRow.style.display='none';
  }
  toast('\u2713 Phone saved: '+digits);
}
function saveHsUrl(id,v){saveCustomerUrl(id,'hubspot_url',v);}
function saveMachineBrand(id,v){saveMachineField(id,'machine_brand',v);}
function saveFilterType(id,v){saveMachineField(id,'filter_type',v);}
function saveContractStart(id,v){saveContractField(id,'contract_start',v);}
function saveContractTerm(id,v){saveContractField(id,'contract_term',v);}
function saveMachineField(id,field,value){
  if(!customers[id])customers[id]={};
  customers[id][field]=value;
  custSave();
}
function saveContractField(id,field,value){
  if(!customers[id])customers[id]={};
  customers[id][field]=value;
  // Auto-calculate renewal date
  if(field==='contract_start'||field==='contract_term'){
    const start=new Date(customers[id].contract_start||value);
    const term=parseInt(customers[id].contract_term||6);
    if(!isNaN(start)){
      const renewal=new Date(start);
      renewal.setMonth(renewal.getMonth()+term);
      customers[id].contract_renewal=renewal.toISOString().slice(0,10);
    }
  }
  custSave();
  rCust();
}
function saveSqUrl(id,v){saveCustomerUrl(id,'square_url',v);}
function saveCustomerUrl(id, field, value){
  if(!customers[id])customers[id]={};
  customers[id][field]=value.trim();
  custSave();
  // Update button if URL was set
  const label=field==='hubspot_url'?'HubSpot':'Square';
  toast(label+' link saved for this client');
}

function logService(id){
  const today=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  if(!customers[id])customers[id]={};
  customers[id].last_service=today;
  // Default next service in 30 days for recurring
  const p=P.find(x=>x.id===id);
  if(p&&p.status==='customer_recurring'){
    const next=new Date();next.setDate(next.getDate()+60);
    customers[id].next_service=next.toISOString().slice(0,10);
  }
  custSave();rCust();
  toast('Service visit logged for '+today);
}

function setNextService(id){
  const d=prompt('Set next service date (YYYY-MM-DD):',new Date(Date.now()+30*864e5).toISOString().slice(0,10));
  if(!d||!/^\d{4}-\d{2}-\d{2}$/.test(d)){toast('Invalid date format');return;}
  if(!customers[id])customers[id]={};
  customers[id].next_service=d;
  custSave();rCust();
  toast('Next service set for '+d);
}

// ── GOALS & MORNING BRIEFING ─────────────────────────────────────────────────
let goals = {
  clients:  10,
  deadline: '2025-07-01',
  mrr:      1490,
  calls:    20,
  walkins:  5,
  quotes:   3,
  closes:   1,
};

function goalsLoad(){
  try{
    const saved=JSON.parse(localStorage.getItem('pic_goals')||'null');
    if(saved)goals={...goals,...saved};
  }catch(e){}
}
function goalsSave(){
  try{localStorage.setItem('pic_goals',JSON.stringify(goals));}catch(e){}
}
function saveGoals(){
  const g=id=>parseFloat(document.getElementById(id)&&document.getElementById(id).value)||0;
  const gs=s=>document.getElementById(s)&&document.getElementById(s).value||'';
  goals.clients  = g('goal-clients')  ||10;
  goals.deadline = gs('goal-deadline')||'2025-07-01';
  goals.mrr      = g('goal-mrr')      ||1490;
  goals.calls    = g('goal-calls')    ||20;
  goals.walkins  = g('goal-walkins')  ||5;
  goals.quotes   = g('goal-quotes')   ||3;
  goals.closes   = g('goal-closes')   ||1;
  goalsSave();
  renderBriefing();
  toast('Goals saved!');
}
function initGoals(){
  goalsLoad();
  const set=(id,v)=>{const el=document.getElementById(id);if(el)el.value=v;};
  set('goal-clients',  goals.clients);
  set('goal-deadline', goals.deadline);
  set('goal-mrr',      goals.mrr);
  set('goal-calls',    goals.calls);
  set('goal-walkins',  goals.walkins);
  set('goal-quotes',   goals.quotes);
  set('goal-closes',   goals.closes);
}

function renderBriefing(){
  const now=new Date();
  const todayStr=now.toISOString().slice(0,10);
  const weekAgo=new Date(now-7*864e5);
  const dayName=now.toLocaleDateString('en-US',{weekday:'long',month:'long',day:'numeric'});

  // ── All log entries ───────────────────────────────────────────────────
  const allEntries=[];
  Object.entries(log).forEach(([id,entries])=>{
    entries.forEach(e=>allEntries.push({...e,pid:parseInt(id)}));
  });

  function parseD(s){if(!s)return null;const d=new Date(s);return isNaN(d)?null:d;}
  const weekEntries=allEntries.filter(e=>{ const d=parseD(e.date);return d&&d>=weekAgo;});

  // ── KPI ROW ───────────────────────────────────────────────────────────
  const recurring=P.filter(p=>{
    const lc=getLC(p.id);return lc&&['signed','customer_recurring','customer_once'].includes(lc.outcome);
  });
  const mrr=recurring.reduce((s,p)=>s+(p.monthly||149),0);
  const clientCount=recurring.length;

  // Pipeline = intro_set + in_play prospects
  const pipeCount=Object.keys(log).filter(id=>{
    const lc=getLC(parseInt(id));
    if(!lc)return false;
    const n=normO(lc.outcome);
    return n==='intro_set'||n==='in_play';
  }).length;

  const weekTotal=weekEntries.filter(e=>e.outcome!=='service_done').length;

  const kpiEl=id=>document.getElementById(id);
  if(kpiEl('kpi-mrr'))kpiEl('kpi-mrr').textContent='$'+mrr.toLocaleString();
  if(kpiEl('kpi-clients'))kpiEl('kpi-clients').textContent=clientCount;
  if(kpiEl('kpi-pipe'))kpiEl('kpi-pipe').textContent=pipeCount;
  if(kpiEl('kpi-week'))kpiEl('kpi-week').textContent=weekTotal;

  // ── FUNNEL (this week) ────────────────────────────────────────────────
  const wWalkins =weekEntries.filter(e=>e.type==='walkin').length;
  const wCalls   =weekEntries.filter(e=>e.type==='call').length;
  const wContacts=weekEntries.filter(e=>e.outcome!=='service_done').length;
  const wIntros  =weekEntries.filter(e=>normO(e.outcome)==='intro_set').length;
  const wSigned  =weekEntries.filter(e=>normO(e.outcome)==='signed').length;
  const wNotNow  =weekEntries.filter(e=>normO(e.outcome)==='not_now').length;

  const fEl=document.getElementById('funnel-stages');
  const fLabel=document.getElementById('funnel-week-label');
  if(fLabel)fLabel.textContent=wCalls+'c / '+wWalkins+'w this week';
  if(fEl){
    const maxV=Math.max(wContacts,1);
    const fStage=(label,val,color,sub)=>{
      const pct=Math.round(val/maxV*100);
      return '<div class="funnel-stage">'
        +'<div class="funnel-label">'+label+'</div>'
        +'<div class="funnel-bar-wrap"><div class="funnel-bar" style="width:'+pct+'%;background:'+color+'"></div></div>'
        +'<div class="funnel-val">'+val+(sub?'<span style="font-size:9px;color:#aaa"> '+sub+'</span>':'')+'</div>'
        +'</div>';
    };
    fEl.innerHTML=
      fStage('Contacts',wContacts,'#94a3b8')
      +fStage('Intro Set',wIntros,'#0891b2',wContacts?Math.round(wIntros/wContacts*100)+'%':'')
      +fStage('Signed',wSigned,'#059669',wContacts?Math.round(wSigned/wContacts*100)+'%':'')
      +fStage('Not Now',wNotNow,'#dc2626',wContacts?Math.round(wNotNow/wContacts*100)+'%':'');
  }

  // ── GOAL PACING ───────────────────────────────────────────────────────
  const gEl=document.getElementById('goal-pacing-content');
  if(gEl){
    const goalClients=goals.clients||10;
    const deadline=goals.deadline?new Date(goals.deadline):null;
    const daysLeft=deadline?Math.ceil((deadline-now)/864e5):null;
    const weeksLeft=daysLeft?Math.max(1,daysLeft/7):null;
    const needed=Math.max(0,goalClients-clientCount);
    const perWeek=weeksLeft?Math.ceil(needed/weeksLeft):null;
    // Current pace (closes last 4 weeks)
    const month4wk=new Date(now-28*864e5);
    const closedLast4=allEntries.filter(e=>{const d=parseD(e.date);return d&&d>=month4wk&&normO(e.outcome)==='signed';}).length;
    const closePace=Math.round(closedLast4/4*10)/10; // closes per week

    let paceHTML='';
    if(deadline){
      const onTrack=closePace>=perWeek;
      paceHTML+='<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:8px">'
        +'<div style="text-align:center;background:#f5f8fa;border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:var(--navy)">'+clientCount+'/'+goalClients+'</div>'
          +'<div style="font-size:8px;color:var(--sub)">clients</div></div>'
        +'<div style="text-align:center;background:#f5f8fa;border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:var(--navy)">'+(daysLeft!==null?daysLeft:'-')+'</div>'
          +'<div style="font-size:8px;color:var(--sub)">days left</div></div>'
        +'<div style="text-align:center;background:'+(onTrack?'#ecfdf5':'#fef2f2')+';border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:'+(onTrack?'#059669':'#dc2626')+'">'+perWeek+'/wk</div>'
          +'<div style="font-size:8px;color:var(--sub)">needed</div></div>'
        +'</div>';
      paceHTML+='<div style="font-size:10px;color:var(--sub)">Current pace: '
        +(closePace>0?closePace+' closes/week':'No closes logged yet')
        +' &nbsp;&#x2022;&nbsp; '
        +(onTrack?'<span class="pace-on-track">On track ✓</span>':'<span class="pace-behind">Behind pace</span>')+'</div>';
    } else {
      paceHTML='<div style="font-size:11px;color:var(--sub)">Set a goal and deadline in <b>Data</b> to see pacing.</div>';
    }
    gEl.innerHTML=paceHTML;
  }

  // ── LOSS REASON BREAKDOWN ─────────────────────────────────────────────
  const lbEl=document.getElementById('loss-breakdown');
  const leEl=document.getElementById('loss-empty');
  if(lbEl){
    // Count not_now entries with reasons (last 90 days)
    const d90=new Date(now-90*864e5);
    const reasonCounts={};
    allEntries.forEach(e=>{
      const d=parseD(e.date);
      if(!d||d<d90)return;
      if(normO(e.outcome)!=='not_now')return;
      const r=e.reason||'no_reason';
      reasonCounts[r]=(reasonCounts[r]||0)+1;
    });
    const sorted=Object.entries(reasonCounts).sort((a,b)=>b[1]-a[1]);
    if(!sorted.length){
      if(lbEl)lbEl.innerHTML='';
      if(leEl)leEl.style.display='block';
    } else {
      if(leEl)leEl.style.display='none';
      const maxR=sorted[0][1];
      lbEl.innerHTML=sorted.slice(0,5).map(([r,n])=>{
        const pct=Math.round(n/maxR*100);
        const label=REASONS[r]||r;
        return '<div style="display:flex;align-items:center;gap:8px">'
          +'<div style="font-size:10px;color:var(--txt);min-width:130px">'+label+'</div>'
          +'<div style="flex:1;height:5px;background:var(--brd2);border-radius:3px">'
            +'<div style="width:'+pct+'%;height:5px;background:#dc2626;border-radius:3px"></div></div>'
          +'<div style="font-size:10px;color:var(--sub);min-width:20px;text-align:right">'+n+'</div>'
          +'</div>';
      }).join('');
    }
  }

  // ── ACTIVE NURTURE (In Play due today or overdue) ─────────────────────
  const today2=now.toISOString().slice(0,10);
  const nurturePros=P.filter(p=>{
    const lc=getLC(p.id);
    if(!lc)return false;
    if(normO(lc.outcome)!=='in_play')return false;
    if(!lc.followup)return true; // In play with no date - show anyway
    return lc.followup<=today2;
  }).sort((a,b)=>{
    const fa=getLC(a.id)?.followup||'9999';
    const fb=getLC(b.id)?.followup||'9999';
    return fa.localeCompare(fb);
  }).slice(0,6);

  const nGrid=document.getElementById('nurture-grid');
  const nEmpty=document.getElementById('nurture-empty');
  if(nGrid){
    if(!nurturePros.length){nGrid.innerHTML='';if(nEmpty)nEmpty.style.display='block';}
    else{if(nEmpty)nEmpty.style.display='none';nGrid.innerHTML=nurturePros.map(p=>cardHTML(p)).join('');}
    nGrid.querySelectorAll('.card').forEach(el=>{
      el.onclick=()=>openM(parseInt(el.dataset.id));
    });
  }

  // ── BEST COLD TARGETS ─────────────────────────────────────────────────
  const coldTargets=P.filter(p=>{
    if(isC(p.id))return false;
    return p.priority==='CALLBACK'||p.priority==='HOT';
  }).sort((a,b)=>(PO[a.priority]??99)-(PO[b.priority]??99)||(b.score||0)-(a.score||0)).slice(0,5);

  const aGrid=document.getElementById('tgrid-actnow');
  const aEmpty=document.getElementById('empty-actnow');
  if(aGrid){
    if(!coldTargets.length){aGrid.innerHTML='';if(aEmpty)aEmpty.style.display='block';}
    else{if(aEmpty)aEmpty.style.display='none';aGrid.innerHTML=coldTargets.map(p=>cardHTML(p)).join('');}
    aGrid.querySelectorAll('.card').forEach(el=>{
      el.onclick=()=>openM(parseInt(el.dataset.id));
    });
  }
}


function swCustomers(){sw('customers');}

function buildAnnualSchedule(id){
  if(!customers[id])return;
  const c=customers[id];
  const start=new Date(c.contract_start||new Date().toISOString().slice(0,10));
  const term=parseInt(c.contract_term||6);
  const schedule=[];

  // Build visit schedule: every 60 days, 6-month visits are deep cleans
  let visit=new Date(start);
  const endDate=new Date(start);
  endDate.setMonth(endDate.getMonth()+(term===6?6:12));

  let visitNum=0;
  while(visit<=endDate){
    visitNum++;
    const isDeep=visitNum===1||(visitNum*60)%180<60;
    schedule.push({
      date: visit.toISOString().slice(0,10),
      date_display: visit.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),
      type: isDeep?'deep_clean':'maintenance_60',
      label: isDeep?'Deep Clean':'60-Day Maintenance',
      status: 'scheduled',
    });
    visit=new Date(visit);
    visit.setDate(visit.getDate()+60);
  }

  c.annual_schedule=schedule;
  c.contract_start=c.contract_start||new Date().toISOString().slice(0,10);
  const renewal=new Date(start);
  renewal.setMonth(renewal.getMonth()+term);
  c.contract_renewal=renewal.toISOString().slice(0,10);
  custSave();
}

function exportSchedulePDF(id){
  const p=P.find(x=>x.id===id);
  const c=customers[id]||{};
  if(!p||!c.annual_schedule)return;

  const win=window.open('','_blank');
  const rows=c.annual_schedule.map((s,i)=>
    '<tr style="background:'+(i%2?'#f9f9f9':'#fff')+'">'
    +'<td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">'+s.date_display+'</td>'
    +'<td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">'+(s.type==='deep_clean'?'&#x1F9FC; Full Deep Clean':'&#x1F527; 60-Day Maintenance')+'</td>'
    +'<td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:'+(s.status==='completed'?'#059669':'#94a3b8')+'">'+(s.status==='completed'?'&#x2713; Completed':'Scheduled')+'</td>'
    +'</tr>'
  ).join('');

  win.document.write('<!DOCTYPE html><html><head><title>Service Schedule - '+p.name+'</title>'
    +'<style>body{font-family:system-ui,sans-serif;padding:24px;max-width:600px;margin:0 auto;color:#1e293b}'
    +'table{width:100%;border-collapse:collapse}th{background:#1e3a5f;color:#fff;padding:10px 12px;text-align:left;font-size:12px}'
    +'@media print{button{display:none}}</style></head><body>'
    +'<div style="display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #1e3a5f;padding-bottom:16px;margin-bottom:20px">'
      +'<div><div style="font-size:20px;font-weight:800;color:#1e3a5f">Pinellas Ice Co</div>'
      +'<div style="font-size:11px;color:#64748b">Commercial Ice Machine Cleaning</div></div>'
      +'<div style="text-align:right"><div style="font-size:13px;font-weight:700">SERVICE SCHEDULE</div>'
      +'<div style="font-size:11px;color:#64748b">Generated '+new Date().toLocaleDateString()+'</div></div>'
    +'</div>'
    +'<div style="background:#f1f5f9;border-radius:8px;padding:12px;margin-bottom:16px">'
      +'<div style="font-size:14px;font-weight:700">'+p.name+'</div>'
      +'<div style="font-size:11px;color:#64748b">'+p.address+', '+p.city+', FL</div>'
      +(c.machine_brand?'<div style="font-size:11px;color:#64748b">Machine: '+c.machine_brand+'</div>':'')
      +'<div style="font-size:11px;color:#64748b">Contract: '+c.contract_start+' &rarr; '+c.contract_renewal+'</div>'
    +'</div>'
    +'<table><thead><tr><th>Date</th><th>Service Type</th><th>Status</th></tr></thead>'
    +'<tbody>'+rows+'</tbody></table>'
    +'<div style="margin-top:20px;font-size:10px;color:#94a3b8;border-top:1px solid #e2e8f0;padding-top:12px">'
    +'Pinellas Ice Co &bull; (727) 855-6873 &bull; pinellasiceco.com &bull; '
    +'Service complies with FDA Food Code 3-502.12</div>'
    +'<br><div style="display:flex;gap:8px;margin-top:8px">'+'<button onclick="window.print()" style="padding:10px 20px;background:#1e3a5f;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px">Print / Save PDF</button>'+'<button onclick="window.close()" style="padding:10px 20px;background:#fff;color:#1e3a5f;border:2px solid #1e3a5f;border-radius:6px;cursor:pointer;font-size:12px">&#x2190; Close</button>'+'</div>'
    +'</body></html>');
  win.document.close();
  setTimeout(()=>win.print(),500);
}

function geoClusterSchedule(){
  // Find clients whose next_service dates are within 5 days of each other
  // and are geographically close — suggest scheduling on the same day
  const recurring=P.filter(p=>p.status==='customer_recurring'&&p.lat&&p.lon);
  if(recurring.length<2)return null;

  const clusters=[];
  const today=new Date();

  recurring.forEach(p=>{
    const c=customers[p.id]||{};
    if(!c.next_service)return;
    const due=new Date(c.next_service);
    const daysUntil=Math.round((due-today)/864e5);

    // Find other clients within 5 days and 8 miles
    const nearby=recurring.filter(q=>{
      if(q.id===p.id)return false;
      const cq=customers[q.id]||{};
      if(!cq.next_service)return false;
      const dueQ=new Date(cq.next_service);
      const daysDiff=Math.abs(Math.round((dueQ-due)/864e5));
      const miles=hav(p.lat,p.lon,q.lat,q.lon);
      return daysDiff<=5&&miles<=8;
    });

    if(nearby.length>0){
      // Check if cluster already recorded
      const ids=[p.id,...nearby.map(n=>n.id)].sort().join(',');
      if(!clusters.find(cl=>cl.ids===ids)){
        clusters.push({ids,anchor:p,nearby,daysUntil});
      }
    }
  });

  return clusters.length?clusters:null;
}

function exportDirectoryData(){
  // Export customer data for public directory
  const recurring=P.filter(p=>p.status==='customer_recurring'||p.status==='customer_once');
  const dirData={
    generated: new Date().toISOString(),
    certified: recurring.map(p=>{
      const c=customers[p.id]||{};
      return {
        id: p.id,
        name: p.name,
        address: p.address,
        city: p.city,
        zip: p.zip,
        phone: p.phone||'',
        lat: p.lat,
        lon: p.lon,
        status: p.status,
        service_count: (c.service_history||[]).length,
        last_service: c.last_service||'',
        last_service_iso: c.last_service_iso||'',
        contract_start: c.contract_start||'',
        contract_renewal: c.contract_renewal||'',
        machine_brand: c.machine_brand||'',
        filter_type: c.filter_type||'',
        certified: p.status==='customer_recurring',
        won_date: c.won_date||'',
      };
    }),
    stats:{
      total_certified: recurring.length,
      counties_served: [...new Set(recurring.map(p=>p.county))].length,
      generated: new Date().toLocaleDateString('en-US',{month:'long',year:'numeric'}),
    }
  };

  const blob=new Blob([JSON.stringify(dirData,null,2)],{type:'application/json'});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');
  a.href=url;a.download='customers.json';
  document.body.appendChild(a);a.click();
  document.body.removeChild(a);URL.revokeObjectURL(url);
  toast('customers.json downloaded — upload to GitHub /data/ folder');
}
function setSvcTab(t){
  svcTab=t;
  document.querySelectorAll('.svc-tab').forEach(b=>b.classList.remove('on'));
  const btn=document.getElementById('svct-'+t);
  if(btn)btn.classList.add('on');
  ['cal','route','reports','tutorials','refs'].forEach(s=>{
    const el=document.getElementById('svc-'+s);
    if(el)el.style.display=s===t?'block':'none';
  });
  if(t==='cal')renderServiceCal();
  else if(t==='route')renderServiceRoute();
  else if(t==='reports')renderReports();
  else if(t==='tutorials')renderTutorial();
  else if(t==='refs')renderReferrals();
}

function rService(){
  renderForecast();
  renderAtRisk();
  setSvcTab(svcTab);
}

// ── FORECAST STRIP ────────────────────────────────────────────────────────────
function renderForecast(){
  const el=document.getElementById('svc-forecast');
  if(!el)return;

  const recurring=P.filter(p=>p.status==='customer_recurring');
  const mrr=recurring.reduce((s,p)=>s+(customers[p.id]?.monthly||p.monthly||149),0);
  const arr=mrr*12;

  // Churn assumption: 5% monthly = industry avg for small service biz
  const churnRate=0.05;
  const newPerMonth=mrr/149; // rough: each new client adds ~$149
  const mrr90=Math.round(mrr*Math.pow(1-churnRate,3)+(newPerMonth*149*3));

  // Days to goal
  const goalMrr=goals.mrr||1490;
  const goalDate=new Date(goals.deadline||'2025-07-01');
  const daysLeft=Math.max(0,Math.round((goalDate-new Date())/864e5));

  el.innerHTML=
    statBox('MRR','$'+mrr.toLocaleString(),'monthly recurring','#059669')
    +statBox('ARR','$'+arr.toLocaleString(),'annual run rate','var(--blu)')
    +statBox('90d Forecast','$'+mrr90.toLocaleString(),'at 5% churn','#7c3aed');
}

// ── AT-RISK ALERT ─────────────────────────────────────────────────────────────
function renderAtRisk(){
  const el=document.getElementById('svc-atrisk');
  if(!el)return;

  const today=new Date();
  const atRisk=P.filter(p=>{
    if(p.status!=='customer_recurring')return false;
    const c=customers[p.id]||{};
    if(!c.last_service&&!c.next_service)return true; // never logged
    if(c.next_service){
      const due=new Date(c.next_service);
      const overdueDays=Math.round((today-due)/864e5);
      return overdueDays>5; // 5+ days overdue
    }
    if(c.last_service){
      const last=new Date(c.last_service);
      const daysSince=Math.round((today-last)/864e5);
      return daysSince>35; // 35+ days since last service
    }
    return false;
  });

  if(!atRisk.length){
    el.innerHTML='<div style="background:#f0fdf4;border:1px solid #6ee7b7;border-radius:8px;padding:10px;font-size:10px;color:#059669;font-weight:600">&#x2713; All recurring clients are on schedule</div>';
    return;
  }

  el.innerHTML='<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:10px">'
    +'<div style="font-size:10px;font-weight:700;color:#dc2626;margin-bottom:6px">&#x26A0; '+atRisk.length+' client'+(atRisk.length>1?'s':'')+' at churn risk</div>'
    +atRisk.slice(0,3).map(p=>{
      const c=customers[p.id]||{};
      const msg=!c.last_service&&!c.next_service?'Never serviced since won'
        :c.next_service?Math.round((new Date()-new Date(c.next_service))/864e5)+'d overdue'
        :'35+d since last service';
      return '<div style="display:flex;justify-content:space-between;align-items:center;padding:4px 0;border-top:1px solid #fca5a5;font-size:10px">'
        +'<div><div style="font-weight:600;color:var(--navy)">'+p.name+'</div>'
        +'<div style="color:#dc2626;font-size:9px">'+msg+'</div></div>'
        +'<button onclick="swCustomers()" style="font-size:9px;padding:3px 7px;border:1px solid #dc2626;border-radius:5px;background:#fef2f2;color:#dc2626;cursor:pointer;font-family:inherit">Fix</button>'
        +'</div>';
    }).join('')
    +'</div>';
}

// ── SERVICE CALENDAR ──────────────────────────────────────────────────────────
function renderServiceCal(){
  const el=document.getElementById('svc-cal-list');
  if(!el)return;

  const today=new Date();
  const recurring=P.filter(p=>p.status==='customer_recurring');

  if(!recurring.length){
    el.innerHTML='<div class="tempty"><div class="ei">&#x1F9FC;</div><div>No recurring clients yet. Close your first deal to start tracking service.</div></div>';
    return;
  }

  // Geo-cluster suggestions
  const clusters=geoClusterSchedule();
  if(clusters&&clusters.length){
    el.innerHTML='<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:10px;font-weight:700;color:var(--blu);margin-bottom:4px">&#x1F4CD; Route Optimization Suggestion</div>'
      +clusters.slice(0,2).map(cl=>'<div style="font-size:10px;color:var(--sub);margin-bottom:3px">'
        +cl.anchor.name.slice(0,20)+' + '+cl.nearby.length+' nearby — schedule same day (within 8mi, ±5d)'
        +'</div>').join('')
      +'</div>';
  } else {
    el.innerHTML='';
  }

  // Sort by urgency: overdue first, then by next_service date
  const withDates=recurring.map(p=>{
    const c=customers[p.id]||{};
    let daysUntil=null;
    if(c.next_service){
      daysUntil=Math.round((new Date(c.next_service)-today)/864e5);
    } else if(c.last_service){
      const lastD=new Date(c.last_service);
      daysUntil=30-Math.round((today-lastD)/864e5);
    }
    return {...p,_daysUntil:daysUntil,_cust:c};
  }).sort((a,b)=>{
    if(a._daysUntil===null)return 1;
    if(b._daysUntil===null)return -1;
    return a._daysUntil-b._daysUntil;
  });

  el.innerHTML=withDates.map(p=>{
    const c=p._cust;
    const du=p._daysUntil;
    const cls=du===null?'no-date':du<0?'overdue':du<=7?'due-soon':'on-track';
    const duLabel=du===null?'No date set'
      :du<0?Math.abs(du)+'d OVERDUE'
      :du===0?'Due TODAY'
      :du===1?'Due TOMORROW'
      :'Due in '+du+'d';
    const duCol=du===null?'var(--sub)':du<0?'#dc2626':du<=7?'#d97706':'#059669';

    return '<div class="svc-card '+cls+'">'
      // Header row
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:13px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+' &bull; '+p.machines+' machine'+(p.machines>1?'s':'')+'</div>'
          +(p.phone?'<div style="font-size:10px;color:var(--blu)">'+p.phone+'</div>':'')
        +'</div>'
        +'<div style="text-align:right;flex-shrink:0;margin-left:10px">'
          +'<div style="font-size:12px;font-weight:800;color:'+duCol+'">'+duLabel+'</div>'
          +'<div style="font-size:10px;font-weight:700;color:#059669">$'+(c.monthly||p.monthly||149)+'/mo</div>'
        +'</div>'
      +'</div>'
      // Service history
      +'<div style="display:flex;justify-content:space-between;font-size:9px;color:var(--sub);background:#f5f8fa;border-radius:6px;padding:6px 8px">'
        +'<span>Last: '+(c.last_service||'Never recorded')+'</span>'
        +'<span>Next: '+(c.next_service||'Not set')+'</span>'
      +'</div>'
      // Actions
      +'<div style="display:flex;gap:5px">'
        +'<button onclick="openServiceLog('+p.id+')" style="flex:2;padding:7px;border:none;border-radius:7px;background:#059669;color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Log Service Visit</button>'
        +'<button onclick="reschedule('+p.id+')" style="flex:1;padding:7px;border:1px solid var(--brd);border-radius:7px;background:var(--surf);color:var(--sub);font-size:10px;cursor:pointer;font-family:inherit">Reschedule</button>'
        +(c.annual_schedule?'<button onclick="exportSchedulePDF('+p.id+')" style="flex:1;padding:7px;border:1px solid var(--blu);border-radius:7px;background:#eff6ff;color:var(--blu);font-size:10px;cursor:pointer;font-family:inherit">&#x1F4C5; Schedule</button>':'')
        +(p.phone?'<a href="tel:'+p.phone.replace(/\s/g,'')+'" style="flex:1;padding:7px;border:1px solid var(--blu);border-radius:7px;background:#eff6ff;color:var(--blu);font-size:10px;text-align:center;text-decoration:none;display:flex;align-items:center;justify-content:center;font-family:inherit">Call</a>':'')
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

function logServiceFromCal(id,opts){
  opts=opts||{};
  const today=new Date();
  const todayStr=today.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  const timeStr=today.toLocaleTimeString('en-US',{hour:'2-digit',minute:'2-digit'});
  if(!customers[id])customers[id]={};
  const c=customers[id];

  // Determine if this is a 6-month deep clean
  const lastDeep=c.service_history
    ?c.service_history.filter(s=>s.type==='deep_clean').slice(-1)[0]
    :null;
  const daysSinceDeep=lastDeep
    ?Math.round((today-new Date(lastDeep.date))/864e5)
    :999;
  const isDeepClean=opts.type==='deep_clean'||(daysSinceDeep>=170);

  // Build service record
  const svcRecord={
    date: today.toISOString().slice(0,10),
    date_display: todayStr,
    time: timeStr,
    type: isDeepClean?'deep_clean':'maintenance_60',
    atp: opts.atp||'',
    filter_replaced: opts.filter_replaced||false,
    filter_type: opts.filter_type||c.filter_type||'',
    notes: opts.notes||'',
    tech: 'Pinellas Ice Co',
  };

  if(!c.service_history)c.service_history=[];
  c.service_history.push(svcRecord);

  c.last_service=todayStr;
  c.last_service_iso=today.toISOString().slice(0,10);

  // Next service: 60 days for maintenance, but auto-schedule deep clean at 6 months
  const next=new Date(today);
  next.setDate(next.getDate()+60);
  c.next_service=next.toISOString().slice(0,10);

  // Flag 6-month deep clean if approaching
  const nextDeepDays=lastDeep
    ?180-Math.round((today-new Date(lastDeep.date))/864e5)
    :180-daysSinceDeep;
  c.next_deep_clean_in=Math.max(0,nextDeepDays);

  custSave();
  renderAtRisk();
  renderServiceCal();
  renderForecast();
  renderBriefing();
  const label=isDeepClean?'Deep clean':'60-day maintenance';
  toast(label+' logged ✓ Next: '+next.toLocaleDateString('en-US',{month:'short',day:'numeric'}));
}

function openServiceLog(id){
  const p=P.find(x=>x.id===id);
  const c=customers[id]||{};
  if(!p)return;

  const brand=c.machine_brand||'';
  const lastSvc=c.service_history&&c.service_history.length
    ?c.service_history[c.service_history.length-1]
    :null;

  // Build modal
  const bg=document.createElement('div');
  bg.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:200;display:flex;align-items:flex-end;justify-content:center';
  bg.id='svc-log-bg';

  bg.innerHTML='<div style="background:var(--surf);border-radius:16px 16px 0 0;padding:16px;width:100%;max-width:480px;max-height:85vh;overflow-y:auto">'
    +'<div style="font-weight:800;font-size:14px;color:var(--navy);margin-bottom:12px">Log Service Visit</div>'
    +'<div style="font-weight:600;font-size:12px;color:var(--sub);margin-bottom:10px">'+p.name+' &bull; '+p.city+'</div>'

    // Service type
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Service Type</div>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px">'
      +'<button onclick="setSvcMaint()" id="svctype-maint" class="svctype-btn on" style="padding:8px;border:2px solid var(--navy);border-radius:8px;background:#f0f4ff;color:var(--navy);font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">&#x1F527; 60-Day Maintenance</button>'
      +'<button onclick="setSvcDeep()" id="svctype-deep" class="svctype-btn" style="padding:8px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">&#x1F9FC; Deep Clean</button>'
    +'</div>'

    // ATP reading
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">ATP Reading (RLU)</div>'
    +'<div style="display:flex;gap:6px;align-items:center;margin-bottom:10px">'
      +'<input id="svc-atp" type="number" placeholder="e.g. 45" min="0" max="9999"'
        +' style="flex:1;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:13px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
      +'<div id="atp-indicator" style="font-size:9px;font-weight:700;padding:4px 8px;border-radius:6px;background:#f0fdf4;color:#059669">Enter RLU</div>'
    +'</div>'

    // Filter replaced
    +'<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;padding:8px;background:#f5f8fa;border-radius:7px">'
      +'<input type="checkbox" id="svc-filter-replaced" onchange="toggleFilterType()">'
      +'<label for="svc-filter-replaced" style="font-size:11px;font-weight:600;color:var(--navy);cursor:pointer">Filter replaced this visit</label>'
    +'</div>'
    +'<div id="svc-filter-type-row" style="display:none;margin-bottom:10px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Filter Type Installed</div>'
      +'<select id="svc-filter-type" style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
        +FILTER_TYPES.map(f=>'<option'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
      +'</select>'
    +'</div>'

    // Notes
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Service Notes</div>'
    +'<textarea id="svc-notes" rows="3" placeholder="Machine condition, issues found, recommendations..."'
      +' style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;resize:none;margin-bottom:10px"></textarea>'

    // Previous service info
    +(lastSvc?'<div style="background:#f5f8fa;border-radius:7px;padding:8px;margin-bottom:10px;font-size:10px;color:var(--sub)">'
      +'<span style="font-weight:600;color:var(--navy)">Previous: </span>'+lastSvc.date_display
      +' &bull; '+(lastSvc.type==='deep_clean'?'Deep Clean':'60-Day Maintenance')
      +(lastSvc.atp?' &bull; ATP: '+lastSvc.atp:'')
      +'</div>':'')

    // Buttons
    +'<div style="display:flex;gap:6px">'
      +'<button onclick="submitServiceLog('+id+')" style="flex:2;padding:10px;border:none;border-radius:8px;background:#059669;color:#fff;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Save Service Visit</button>'
      +'<button onclick="closeSvcLog()" style="flex:1;padding:10px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:11px;cursor:pointer;font-family:inherit">Cancel</button>'
    +'</div>'
    +'</div>';

  document.body.appendChild(bg);

  // ATP live feedback
  const atpInput=document.getElementById('svc-atp');
  if(atpInput){
    atpInput.addEventListener('input',function(){
      const v=parseInt(this.value);
      const ind=document.getElementById('atp-indicator');
      if(!ind)return;
      if(isNaN(v)){ind.textContent='Enter RLU';ind.style.background='#f5f8fa';ind.style.color='var(--sub)';}
      else if(v<=10){ind.textContent='Excellent';ind.style.background='#f0fdf4';ind.style.color='#059669';}
      else if(v<=30){ind.textContent='Good';ind.style.background='#eff6ff';ind.style.color='var(--blu)';}
      else if(v<=100){ind.textContent='Acceptable';ind.style.background='#fef9ee';ind.style.color='#d97706';}
      else{ind.textContent='Retest/Reservice';ind.style.background='#fef2f2';ind.style.color='#dc2626';}
    });
  }
}

let _svcType='maintenance_60';
function closeSvcLog(){const el=document.getElementById('svc-log-bg');if(el)el.remove();}
function toggleStep(el){const next=el.nextElementSibling;if(next)next.style.display=next.style.display==='none'?'block':'none';}

function getDiagram(brand){
  const diagrams={
    'Manitowoc': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Manitowoc Cube Ice Machine</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <!-- Machine outline -->
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Evaporator plates -->
        <rect x="60" y="25" width="200" height="55" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="45" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">EVAPORATOR PLATES</text>
        <text x="160" y="60" text-anchor="middle" font-size="8" fill="#3b82f6">Where ice forms — clean &amp; sanitize thoroughly</text>
        <text x="160" y="73" text-anchor="middle" font-size="7" fill="#64748b">Scale = white/gray crust | Biofilm = pink/orange slime</text>
        <!-- Distribution tube -->
        <rect x="55" y="85" width="210" height="14" rx="3" fill="#fef9c3" stroke="#ca8a04" stroke-width="1.5"/>
        <text x="160" y="96" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">WATER DISTRIBUTION TUBE — clear holes with toothpick</text>
        <!-- Water curtain -->
        <rect x="55" y="103" width="210" height="10" rx="2" fill="#f0fdf4" stroke="#16a34a" stroke-width="1.5"/>
        <text x="160" y="112" text-anchor="middle" font-size="8" fill="#15803d">WATER CURTAIN — remove &amp; soak in cleaner</text>
        <!-- Water trough -->
        <rect x="55" y="117" width="210" height="30" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="130" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH</text>
        <text x="160" y="142" text-anchor="middle" font-size="8" fill="#be185d">Highest ATP — biofilm source — scrub thoroughly</text>
        <!-- Float valve -->
        <circle cx="75" cy="132" r="8" fill="#fef3c7" stroke="#d97706" stroke-width="1.5"/>
        <text x="75" y="160" text-anchor="middle" font-size="7" fill="#92400e">FLOAT</text>
        <!-- Ice bin -->
        <rect x="55" y="153" width="210" height="47" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="172" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN</text>
        <text x="160" y="185" text-anchor="middle" font-size="8" fill="#64748b">Sanitize all interior surfaces — discard first batch</text>
        <!-- Arrow labels -->
        <text x="30" y="55" font-size="7" fill="#64748b" transform="rotate(-90,30,55)">TOP</text>
        <text x="30" y="180" font-size="7" fill="#64748b" transform="rotate(-90,30,180)">BOTTOM</text>
      </svg>
    </div>`,
    'Hoshizaki': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Hoshizaki Crescent Cube Machine</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Evaporator - stainless -->
        <rect x="60" y="25" width="200" height="50" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="43" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">EVAPORATOR (STAINLESS STEEL)</text>
        <text x="160" y="56" text-anchor="middle" font-size="8" fill="#3b82f6">Crescent cube molds — uniform silver when clean</text>
        <text x="160" y="69" text-anchor="middle" font-size="7" fill="#16a34a">Any commercial cleaner OK — soft brush only</text>
        <!-- Spray bar - highlighted as critical -->
        <rect x="55" y="80" width="210" height="16" rx="3" fill="#fef08a" stroke="#ca8a04" stroke-width="2"/>
        <text x="160" y="92" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">&#x26A0; SPRAY BAR — MOST CRITICAL COMPONENT</text>
        <!-- Spray bar dots -->
        <circle cx="80" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="100" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="120" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="140" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="160" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="180" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="200" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="220" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="240" cy="88" r="2" fill="#ca8a04"/>
        <text x="160" y="106" text-anchor="middle" font-size="7" fill="#92400e">Clear EVERY hole with toothpick — blocked holes = irregular ice</text>
        <!-- Float switch -->
        <rect x="55" y="112" width="80" height="20" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="95" y="122" text-anchor="middle" font-size="8" font-weight="bold" fill="#9d174d">FLOAT SWITCH</text>
        <text x="95" y="130" text-anchor="middle" font-size="6" fill="#be185d">Must slide freely</text>
        <!-- Water trough -->
        <rect x="55" y="136" width="210" height="25" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="148" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH — highest ATP reading</text>
        <text x="160" y="158" text-anchor="middle" font-size="7" fill="#be185d">Scrub biofilm (pink/orange) — common in FL heat</text>
        <!-- Bin -->
        <rect x="55" y="165" width="210" height="40" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="182" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN — Sanitize all interior surfaces</text>
        <text x="160" y="196" text-anchor="middle" font-size="7" fill="#64748b">Discard first 2 full bins after service</text>
      </svg>
    </div>`,
    'Scotsman': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Scotsman Prodigy Vertical Evaporator</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Vertical evaporator plate -->
        <rect x="60" y="20" width="70" height="140" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="95" y="55" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">VERTICAL</text>
        <text x="95" y="68" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">EVAPORATOR</text>
        <text x="95" y="82" text-anchor="middle" font-size="7" fill="#3b82f6">Scotsman unique</text>
        <text x="95" y="94" text-anchor="middle" font-size="7" fill="#3b82f6">vertical design</text>
        <text x="95" y="108" text-anchor="middle" font-size="7" fill="#3b82f6">Ice forms on</text>
        <text x="95" y="120" text-anchor="middle" font-size="7" fill="#3b82f6">flat vertical</text>
        <text x="95" y="132" text-anchor="middle" font-size="7" fill="#3b82f6">plate surface</text>
        <!-- Critical: thickness sensor -->
        <circle cx="148" cy="90" r="10" fill="#fee2e2" stroke="#dc2626" stroke-width="2"/>
        <text x="148" y="94" text-anchor="middle" font-size="7" font-weight="bold" fill="#dc2626">SENS</text>
        <line x1="130" y1="90" x2="158" y2="90" stroke="#dc2626" stroke-width="1.5"/>
        <text x="210" y="75" text-anchor="middle" font-size="9" font-weight="bold" fill="#dc2626">&#x26A0; ICE THICKNESS</text>
        <text x="210" y="87" text-anchor="middle" font-size="9" font-weight="bold" fill="#dc2626">SENSOR</text>
        <text x="210" y="99" text-anchor="middle" font-size="7" fill="#dc2626">Gap = 3/8 inch</text>
        <text x="210" y="111" text-anchor="middle" font-size="7" fill="#dc2626">(pencil thickness)</text>
        <text x="210" y="123" text-anchor="middle" font-size="7" fill="#dc2626">Clean at EVERY visit</text>
        <!-- Control panel -->
        <rect x="145" y="25" width="140" height="40" rx="4" fill="#f0fdf4" stroke="#16a34a" stroke-width="1.5"/>
        <text x="215" y="40" text-anchor="middle" font-size="9" font-weight="bold" fill="#15803d">PRODIGY CONTROL</text>
        <text x="215" y="54" text-anchor="middle" font-size="7" fill="#15803d">Check error codes before cleaning</text>
        <!-- Drain pan critical -->
        <rect x="55" y="165" width="210" height="30" rx="3" fill="#fef9c3" stroke="#ca8a04" stroke-width="2"/>
        <text x="160" y="178" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">DRAIN PAN — check at every visit</text>
        <text x="160" y="190" text-anchor="middle" font-size="7" fill="#92400e">Pour 1 cup water — should drain in under 30 seconds</text>
        <!-- Water trough -->
        <rect x="55" y="162" width="210" height="0" rx="3" fill="none"/>
      </svg>
    </div>`,
    'Ice-O-Matic': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Ice-O-Matic (similar to Manitowoc)</div>
      <svg viewBox="0 0 320 210" style="width:100%;max-height:190px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="190" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <rect x="55" y="22" width="210" height="45" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="40" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">EVAPORATOR PLATE</text>
        <text x="160" y="53" text-anchor="middle" font-size="8" fill="#3b82f6">Nickel-plated — scale forms in strips matching blocked holes</text>
        <!-- Distribution pan critical -->
        <rect x="55" y="72" width="210" height="20" rx="3" fill="#fef08a" stroke="#ca8a04" stroke-width="2"/>
        <text x="160" y="84" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">&#x26A0; DISTRIBUTION PAN — MOST CRITICAL (IOM)</text>
        <!-- Hole dots -->
        <circle cx="80" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="95" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="110" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="125" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="140" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="155" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="170" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="185" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="200" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="215" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="230" cy="82" r="1.5" fill="#ca8a04"/>
        <text x="160" y="101" text-anchor="middle" font-size="7" fill="#92400e">Clear EVERY hole with toothpick — IOM unique critical step</text>
        <rect x="55" y="105" width="210" height="25" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="117" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH — scrub biofilm from corners</text>
        <text x="160" y="127" text-anchor="middle" font-size="7" fill="#be185d">Check drain flows freely</text>
        <!-- Bin baffle -->
        <rect x="90" y="134" width="140" height="15" rx="2" fill="#f0fdf4" stroke="#16a34a" stroke-width="1.5"/>
        <text x="160" y="145" text-anchor="middle" font-size="8" font-weight="bold" fill="#15803d">BIN BAFFLE — check underside for biofilm</text>
        <!-- Ice bin -->
        <rect x="55" y="153" width="210" height="40" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="170" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN</text>
        <text x="160" y="184" text-anchor="middle" font-size="7" fill="#64748b">Sanitize all interior — discard first batch</text>
      </svg>
    </div>`,
    'Follett': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Follett Nugget Ice (Auger System)</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Evaporator cylinder - vertical -->
        <rect x="120" y="20" width="80" height="120" rx="8" fill="#dbeafe" stroke="#3b82f6" stroke-width="2"/>
        <text x="160" y="40" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">EVAPORATOR</text>
        <text x="160" y="52" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">CYLINDER</text>
        <text x="160" y="65" text-anchor="middle" font-size="7" fill="#3b82f6">Stainless steel</text>
        <text x="160" y="77" text-anchor="middle" font-size="7" fill="#3b82f6">Ice forms on</text>
        <text x="160" y="89" text-anchor="middle" font-size="7" fill="#3b82f6">inside wall</text>
        <!-- Auger inside cylinder -->
        <line x1="160" y1="25" x2="160" y2="135" stroke="#7c3aed" stroke-width="3" stroke-dasharray="4,2"/>
        <text x="160" y="108" text-anchor="middle" font-size="8" font-weight="bold" fill="#7c3aed">AUGER</text>
        <text x="160" y="120" text-anchor="middle" font-size="7" fill="#7c3aed">Pushes ice up</text>
        <!-- Motor on top -->
        <rect x="130" y="8" width="60" height="14" rx="3" fill="#f5f3ff" stroke="#7c3aed" stroke-width="1.5"/>
        <text x="160" y="19" text-anchor="middle" font-size="7" font-weight="bold" fill="#6d28d9">AUGER MOTOR</text>
        <!-- Warning about auger -->
        <rect x="210" y="45" width="85" height="45" rx="4" fill="#fef2f2" stroke="#dc2626" stroke-width="1.5"/>
        <text x="252" y="60" text-anchor="middle" font-size="8" font-weight="bold" fill="#dc2626">&#x26A0; CAUTION</text>
        <text x="252" y="72" text-anchor="middle" font-size="7" fill="#dc2626">Auger is SHARP</text>
        <text x="252" y="83" text-anchor="middle" font-size="7" fill="#dc2626">Handle carefully</text>
        <!-- Ice chute -->
        <rect x="55" y="45" width="60" height="40" rx="4" fill="#fef9c3" stroke="#ca8a04" stroke-width="1.5"/>
        <text x="85" y="62" text-anchor="middle" font-size="8" font-weight="bold" fill="#92400e">ICE CHUTE</text>
        <text x="85" y="74" text-anchor="middle" font-size="7" fill="#92400e">Clean at</text>
        <text x="85" y="83" text-anchor="middle" font-size="7" fill="#92400e">every visit</text>
        <!-- Bin -->
        <rect x="55" y="148" width="210" height="52" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="168" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">STORAGE BIN</text>
        <text x="160" y="180" text-anchor="middle" font-size="7" fill="#64748b">Listen for grinding before cleaning — indicates scale on cylinder</text>
        <text x="160" y="192" text-anchor="middle" font-size="7" fill="#64748b">Soft/moist nuggets = healthy | Hard nuggets = issue</text>
        <!-- Arrow showing auger direction -->
        <text x="205" y="100" font-size="8" fill="#7c3aed">&#x2191; ice flow</text>
      </svg>
    </div>`,
    'Cornelius': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Cornelius Compact Machine</div>
      <svg viewBox="0 0 320 190" style="width:100%;max-height:170px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="170" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Compact layout - front accessible -->
        <rect x="55" y="22" width="210" height="40" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="38" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">FREEZE PLATE</text>
        <text x="160" y="52" text-anchor="middle" font-size="7" fill="#3b82f6">Front-accessible — scale = white/gray | clean with soft brush</text>
        <!-- Air filter critical -->
        <rect x="55" y="67" width="90" height="30" rx="3" fill="#fef08a" stroke="#ca8a04" stroke-width="2"/>
        <text x="100" y="80" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">&#x26A0; AIR FILTER</text>
        <text x="100" y="91" text-anchor="middle" font-size="7" fill="#92400e">#1 Maintenance Item</text>
        <!-- Location of filter -->
        <text x="100" y="109" text-anchor="middle" font-size="7" fill="#64748b">Side or rear mount</text>
        <text x="100" y="119" text-anchor="middle" font-size="7" fill="#64748b">Rinse + blow out</text>
        <!-- Drain pan odor -->
        <rect x="155" y="67" width="110" height="30" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="210" y="80" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">DRAIN PAN</text>
        <text x="210" y="91" text-anchor="middle" font-size="7" fill="#be185d">Odor source — scrub biofilm</text>
        <!-- Water trough -->
        <rect x="55" y="103" width="210" height="22" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="114" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH — check mineral deposits</text>
        <text x="160" y="122" text-anchor="middle" font-size="7" fill="#be185d">Cloudy ice = mineral contamination = filter issue</text>
        <!-- Bin -->
        <rect x="55" y="130" width="210" height="42" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="148" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN — Sanitize thoroughly</text>
        <text x="160" y="161" text-anchor="middle" font-size="7" fill="#64748b">Simplest internal layout — faster service than modular units</text>
      </svg>
    </div>`,
  };
  return diagrams[brand]||'';
}
function setSvcMaint(){setSvcType('maintenance_60');}
function setSvcDeep(){setSvcType('deep_clean');}
function setSvcType(t){
  _svcType=t;
  const idMap={'maintenance_60':'svctype-maint','deep_clean':'svctype-deep'};
  document.querySelectorAll('.svctype-btn').forEach(b=>{
    const isOn=b.id===idMap[t];
    b.style.border=isOn?'2px solid var(--navy)':'1px solid var(--brd)';
    b.style.background=isOn?'#f0f4ff':'var(--surf)';
    b.style.color=isOn?'var(--navy)':'var(--sub)';
  });
}
function toggleFilterType(){
  const cb=document.getElementById('svc-filter-replaced');
  const row=document.getElementById('svc-filter-type-row');
  if(row)row.style.display=cb&&cb.checked?'block':'none';
}
function submitServiceLog(id){
  const atp=(document.getElementById('svc-atp')||{}).value||'';
  const notes=(document.getElementById('svc-notes')||{}).value||'';
  const filterReplaced=!!(document.getElementById('svc-filter-replaced')||{}).checked;
  const filterType=(document.getElementById('svc-filter-type')||{}).value||'';
  logServiceFromCal(id,{
    type:_svcType,
    atp,notes,
    filter_replaced:filterReplaced,
    filter_type:filterType,
  });
  if(filterReplaced&&filterType){
    if(!customers[id])customers[id]={};
    customers[id].filter_type=filterType;
    customers[id].filter_installed=new Date().toISOString().slice(0,10);
    custSave();
  }
  document.getElementById('svc-log-bg').remove();
}

function reschedule(id){
  const c=customers[id]||{};
  const current=c.next_service||new Date(Date.now()+60*864e5).toISOString().slice(0,10);
  const d=prompt('Set next service date (YYYY-MM-DD):',current);
  if(!d||!/^\d{4}-\d{2}-\d{2}$/.test(d)){toast('Invalid date');return;}
  if(!customers[id])customers[id]={};
  customers[id].next_service=d;
  // Rebuild annual schedule from new date
  if(customers[id].annual_schedule&&customers[id].annual_schedule.length){
    const today=new Date().toISOString().slice(0,10);
    // Update the next upcoming scheduled visit
    const nextIdx=customers[id].annual_schedule.findIndex(s=>s.status!=='completed'&&s.date>=today);
    if(nextIdx>=0){
      customers[id].annual_schedule[nextIdx].date=d;
      customers[id].annual_schedule[nextIdx].date_display=new Date(d).toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
      // Realign subsequent visits every 60 days from new date
      for(let i=nextIdx+1;i<customers[id].annual_schedule.length;i++){
        const prev=new Date(customers[id].annual_schedule[i-1].date);
        const next=new Date(prev);next.setDate(next.getDate()+60);
        customers[id].annual_schedule[i].date=next.toISOString().slice(0,10);
        customers[id].annual_schedule[i].date_display=next.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
      }
    }
  }
  custSave();
  renderServiceCal();
  toast('Rescheduled ✓ Schedule realigned from '+d);
}

// ── SERVICE ROUTE BUILDER ─────────────────────────────────────────────────────
function renderTutorial(){
  const brand=(document.getElementById('tut-brand')||{}).value||'';
  const type=(document.getElementById('tut-type')||{}).value||'deep_clean';
  // ATP protocol is brand-independent - render it separately
  if(type==='atp_protocol'){renderATPProtocol();return;}
  const el=document.getElementById('tut-content');
  if(!el)return;
  if(!brand){el.innerHTML='<div class="tempty"><div>Select a machine brand above</div></div>';return;}

  const data=type==='deep_clean'?DEEP_CLEAN[brand]:MAINTENANCE_60[brand];
  if(!data){el.innerHTML='<div class="tempty"><div>Tutorial not found</div></div>';return;}

  let html='<div style="background:#f0f4ff;border-radius:8px;padding:10px;margin-bottom:10px">'
    +'<div style="font-weight:800;font-size:12px;color:var(--navy)">'+brand+' — '+(type==='deep_clean'?'Full Deep Clean':'60-Day Maintenance')+'</div>'
    +'<div style="font-size:10px;color:var(--sub);margin-top:2px">⏱ Est. time: '+data.time+'</div>';

  if(data.chemicals){
    html+='<div style="font-size:9px;color:var(--sub);margin-top:4px"><b>Chemicals:</b> '+data.chemicals.join(' • ')+'</div>';
  }
  html+='</div>';

  // Video & resource links
  const vids=VIDEO_LINKS[brand]||{};
  if(vids.deep_url||vids.guide_url){
    html+='<div style="background:#fff7f5;border:1px solid #fed7aa;border-radius:8px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:10px;font-weight:700;color:#ea580c;margin-bottom:6px">&#x1F3A5; Reference Resources</div>'
      // YouTube component ID search - most useful for beginners
      +(vids.yt_component?'<a href="'+vids.yt_component+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:7px 8px;background:#ff0000;border-radius:6px;text-decoration:none;margin-bottom:5px">'
        +'<span style="font-size:14px;color:#fff">&#x25B6;</span>'
        +'<div><div style="font-size:10px;font-weight:700;color:#fff">&#x1F4CD; YouTube: Identify Components</div>'
        +'<div style="font-size:9px;color:rgba(255,255,255,.8)">Search videos showing inside of '+brand+' — identify each part visually</div></div>'
      +'</a>':'')
      +(vids.youtube_url?'<a href="'+vids.youtube_url+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:7px 8px;background:#cc0000;border-radius:6px;text-decoration:none;margin-bottom:5px">'
        +'<span style="font-size:14px;color:#fff">&#x1F9FC;</span>'
        +'<div><div style="font-size:10px;font-weight:700;color:#fff">YouTube: Full Cleaning Tutorial</div>'
        +'<div style="font-size:9px;color:rgba(255,255,255,.8)">Watch a complete '+brand+' deep clean walkthrough</div></div>'
      +'</a>':'')
      +(vids.deep_url?'<a href="'+vids.deep_url+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:6px 8px;background:#fff;border:1px solid #fed7aa;border-radius:6px;text-decoration:none;margin-bottom:5px">'
        +'<span style="font-size:14px">&#x25B6;</span>'
        +'<div><div style="font-size:10px;font-weight:600;color:#ea580c">Official Cleaning Reference</div>'
        +'<div style="font-size:9px;color:var(--sub)">Step-by-step guide with photos</div></div>'
      +'</a>':'')
      +(vids.official_url&&vids.official_url!==vids.deep_url?'<a href="'+vids.official_url+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:6px 8px;background:#fff;border:1px solid #fed7aa;border-radius:6px;text-decoration:none">'
        +'<span style="font-size:14px">&#x1F3ED;</span>'
        +'<div><div style="font-size:10px;font-weight:600;color:#ea580c">Official: '+brand+' Support</div>'
        +'<div style="font-size:9px;color:var(--sub)">Manufacturer training & documentation</div></div>'
      +'</a>':'')
    +'</div>';
  }

  // Component diagram (deep clean only)
  if(type==='deep_clean'){html+=getDiagram(brand);}

  data.steps.forEach(function(s,i){
    html+='<div style="margin-bottom:8px;border:1px solid var(--brd);border-radius:8px;overflow:hidden">'
      +'<div onclick="toggleStep(this)" style="display:flex;align-items:center;gap:8px;padding:10px;background:var(--surf);cursor:pointer">'
        +'<div style="background:var(--navy);color:#fff;border-radius:50%;width:22px;height:22px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;flex-shrink:0">'+(i+1)+'</div>'
        +'<div style="font-weight:700;font-size:11px;color:var(--navy);flex:1">'+s.title+'</div>'
        +'<div style="font-size:10px;color:var(--sub)">▼</div>'
      +'</div>'
      +'<div style="display:none;padding:10px;font-size:11px;color:var(--sub);line-height:1.5;border-top:1px solid var(--brd2)">'+s.detail+'</div>'
    +'</div>';
  });

  el.innerHTML=html;
}


function renderATPProtocol(){
  const el=document.getElementById('tut-content');
  if(!el)return;
  const s=CHEM_REF.atp_sales;
  const c=CHEM_REF.atp_compliance;
  const a=CHEM_REF.atp;

  el.innerHTML=
    // Header
    '<div style="background:#f0f4ff;border-radius:8px;padding:10px;margin-bottom:12px">'
      +'<div style="font-weight:800;font-size:12px;color:var(--navy)">&#x1F4CA; ATP Testing Guide</div>'
      +'<div style="font-size:10px;color:var(--sub);margin-top:3px">Hygiena Ensure v2 • Ultrasnap testers • All brands</div>'
    +'</div>'

    // Meter technique
    +'<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:12px;margin-bottom:10px">'
      +'<div style="font-size:11px;font-weight:700;color:var(--navy);margin-bottom:6px">&#x1F52C; How to Use the Hygiena Ensure v2</div>'
      +'<div style="font-size:11px;color:var(--sub);line-height:1.6">'
        +'<b>1.</b> Snap the Ultrasnap tube to release reagent liquid into the swab chamber.<br>'
        +'<b>2.</b> Swab a 10cm² area (palm-sized) using a firm Z-pattern — 5 strokes across, 5 strokes down.<br>'
        +'<b>3.</b> Insert swab into Ensure v2 and press READ.<br>'
        +'<b>4.</b> Result appears in 15 seconds.<br><br>'
        +'<b>Pass thresholds:</b><br>'
        +'✅ 10 RLU or below — ice contact surfaces (evaporator, trough, distribution tube)<br>'
        +'✅ 30 RLU or below — general food contact surfaces (bin, scoop, exterior)<br>'
        +'❌ Above threshold: re-clean, re-sanitize, retest. Do not leave until passing.'
      +'</div>'
    +'</div>'

    // Sales protocol
    +'<div style="background:#fef9ee;border:1px solid #fde68a;border-radius:8px;padding:12px;margin-bottom:10px">'
      +'<div style="font-size:11px;font-weight:800;color:#92400e;margin-bottom:8px">&#x1F4B0; '+s.title+'</div>'
      +'<div style="font-size:10px;color:#78350f;margin-bottom:10px;line-height:1.5">'+s.purpose+'</div>'
      +s.locations.map((l,i)=>
        '<div style="padding:8px;background:#fffbeb;border-radius:6px;margin-bottom:6px">'
          +'<div style="font-size:11px;font-weight:700;color:#92400e">'+(i+1)+'. '+l.where+'</div>'
          +'<div style="font-size:10px;color:#78350f;margin-top:2px;line-height:1.4">'+l.why+'</div>'
        +'</div>'
      ).join('')
      +'<div style="background:#d97706;border-radius:7px;padding:10px;margin-top:8px">'
        +'<div style="font-size:10px;font-weight:700;color:#fff;margin-bottom:4px">The Pitch Move</div>'
        +'<div style="font-size:10px;color:rgba(255,255,255,.9);line-height:1.5">'+s.pitch+'</div>'
      +'</div>'
    +'</div>'

    // Compliance protocol
    +'<div style="background:#f0fdf4;border:1px solid #6ee7b7;border-radius:8px;padding:12px;margin-bottom:10px">'
      +'<div style="font-size:11px;font-weight:800;color:#065f46;margin-bottom:8px">'+c.title+'</div>'
      +'<div style="font-size:10px;color:#064e3b;margin-bottom:10px;line-height:1.5">'+c.purpose+'</div>'
      +c.locations.map((l,i)=>
        '<div style="padding:8px;background:#ecfdf5;border-radius:6px;margin-bottom:6px">'
          +'<div style="font-size:11px;font-weight:700;color:#065f46">'+(i+1)+'. '+l.where+'</div>'
          +'<div style="font-size:10px;color:#064e3b;margin-top:2px;line-height:1.4">'+l.why+'</div>'
        +'</div>'
      ).join('')
      +'<div style="background:#059669;border-radius:7px;padding:10px;margin-top:8px">'
        +'<div style="font-size:10px;font-weight:700;color:#fff;margin-bottom:4px">Report Format</div>'
        +'<div style="font-size:10px;color:rgba(255,255,255,.9);line-height:1.5">'+c.report_format+'</div>'
      +'</div>'
    +'</div>'

    // Compliance sticker
    +'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:10px">'
      +'<div style="font-size:11px;font-weight:700;color:var(--navy);margin-bottom-4px">&#x1F3F7; Compliance Sticker Placement</div>'
      +'<div style="font-size:10px;color:var(--sub);margin-top:4px;line-height:1.5">'+c.sticker+'</div>'
    +'</div>';
}

function renderServiceRoute(){
  const el=document.getElementById('svc-route-list');
  if(!el)return;
  el.innerHTML='<div style="font-size:10px;color:var(--sub)">Tap Build Route to optimize service stops for selected week.</div>';
  document.getElementById('svc-route-map-btn').style.display='none';
}

function buildServiceRoute(){
  const weekOffset=parseInt(document.getElementById('svc-week').value)||0;
  const today=new Date();
  const windowStart=new Date(today.getTime()+weekOffset*864e5);
  const windowEnd=new Date(windowStart.getTime()+7*864e5);

  // Get clients due in window
  const due=P.filter(p=>{
    if(p.status!=='customer_recurring')return false;
    const c=customers[p.id]||{};
    if(!c.next_service&&!c.last_service)return weekOffset<0; // overdue filter
    if(c.next_service){
      const d=new Date(c.next_service);
      if(weekOffset===-7)return d<today; // overdue
      return d>=windowStart&&d<=windowEnd;
    }
    if(c.last_service){
      const last=new Date(c.last_service);
      const daysSince=Math.round((today-last)/864e5);
      return daysSince>=25&&daysSince<=35; // due around now
    }
    return false;
  }).filter(p=>p.lat&&p.lon);

  if(!due.length){
    const el=document.getElementById('svc-route-list');
    if(el)el.innerHTML='<div class="tempty"><div>No clients due in this window</div></div>';
    return;
  }

  // Nearest-neighbor from first client (sorted by next_service date)
  due.sort((a,b)=>{
    const ca=customers[a.id]||{},cb=customers[b.id]||{};
    return (ca.next_service||'9999').localeCompare(cb.next_service||'9999');
  });
  serviceRoute=[due[0]];
  const remaining=[...due.slice(1)];
  while(remaining.length){
    const last=serviceRoute[serviceRoute.length-1];
    let best=0,bestD=Infinity;
    remaining.forEach((p,i)=>{
      const d=hav(last.lat,last.lon,p.lat,p.lon);
      if(d<bestD){bestD=d;best=i;}
    });
    serviceRoute.push(remaining.splice(best,1)[0]);
  }

  // Calculate total miles
  let totalMi=0;
  for(let i=1;i<serviceRoute.length;i++){
    totalMi+=hav(serviceRoute[i-1].lat,serviceRoute[i-1].lon,serviceRoute[i].lat,serviceRoute[i].lon);
  }
  const estTime=Math.round(serviceRoute.length*45+totalMi*3); // 45min/stop + drive

  const el=document.getElementById('svc-route-list');
  el.innerHTML='<div style="font-size:10px;color:var(--sub);margin-bottom:8px">'
    +serviceRoute.length+' stops &bull; ~'+totalMi.toFixed(1)+'mi &bull; est. '+Math.floor(estTime/60)+'h '+Math.round(estTime%60)+'m (45min/stop)</div>'
    +serviceRoute.map((p,i)=>{
      const c=customers[p.id]||{};
      return '<div class="day-stop" style="margin-bottom:6px">'
        +'<div class="stopnum">'+(i+1)+'</div>'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:9px;color:var(--sub)">'+p.address+', '+p.city+'</div>'
          +'<div style="font-size:9px;color:#d97706">Next due: '+(c.next_service||'not set')+'</div>'
          +(p.phone?'<a href="tel:'+p.phone.replace(/\s/g,'')+'" style="font-size:9px;color:var(--blu)">'+p.phone+'</a>':'')
        +'</div>'
        +'<button onclick="logServiceFromCal('+p.id+')" style="font-size:9px;padding:4px 7px;border:none;border-radius:6px;background:#059669;color:#fff;cursor:pointer;font-family:inherit;flex-shrink:0">Done</button>'
        +'</div>';
    }).join('');

  document.getElementById('svc-route-map-btn').style.display='block';
  toast('Service route built: '+serviceRoute.length+' stops');
}

function openServiceMaps(){
  if(!serviceRoute.length)return;
  const url='https://www.google.com/maps/dir/'+serviceRoute
    .filter(p=>p.address)
    .map(p=>enc(p.address+', '+p.city+', FL '+p.zip))
    .join('/');
  window.open(url,'_blank');
}

// ── SERVICE REPORTS ───────────────────────────────────────────────────────────
function renderReports(){
  const sel=document.getElementById('svc-report-client');
  if(!sel)return;
  const recurring=P.filter(p=>p.status==='customer_recurring'||p.status==='customer_once');
  sel.innerHTML='<option value="">Select a client...</option>'
    +recurring.map(p=>'<option value="'+p.id+'">'+p.name+' — '+p.city+'</option>').join('');
  document.getElementById('svc-report-preview').innerHTML='';
}

function loadReportClient(){
  const id=parseInt(document.getElementById('svc-report-client').value);
  if(!id){document.getElementById('svc-report-preview').innerHTML='';return;}
  const p=P.find(x=>x.id===id);
  if(!p)return;
  const c=customers[id]||{};
  const today=new Date().toLocaleDateString('en-US',{weekday:'long',year:'numeric',month:'long',day:'numeric'});

  document.getElementById('svc-report-preview').innerHTML=
    '<div style="border:2px solid var(--brd);border-radius:10px;padding:16px;margin-top:8px" id="report-content">'
    // Header
    +'<div style="display:flex;justify-content:space-between;align-items:flex-start;border-bottom:2px solid var(--navy);padding-bottom:10px;margin-bottom:12px">'
      +'<div>'
        +'<div style="font-size:18px;font-weight:800;color:var(--navy)">Pinellas Ice Co</div>'
        +'<div style="font-size:10px;color:var(--sub)">Commercial Ice Machine Cleaning &amp; Sanitizing</div>'
        +'<div style="font-size:10px;color:var(--sub)">pinellasiceco.com</div>'
      +'</div>'
      +'<div style="text-align:right">'
        +'<div style="font-size:14px;font-weight:800;color:var(--navy)">SERVICE REPORT</div>'
        +'<div style="font-size:10px;color:var(--sub)">'+today+'</div>'
      +'</div>'
    +'</div>'
    // Client
    +'<div style="background:#f5f8fa;border-radius:7px;padding:10px;margin-bottom:12px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:4px">Client</div>'
      +'<div style="font-weight:700;font-size:14px;color:var(--navy)">'+p.name+'</div>'
      +'<div style="font-size:11px;color:var(--sub)">'+p.address+', '+p.city+', FL '+p.zip+'</div>'
      +(p.phone?'<div style="font-size:11px;color:var(--sub)">'+p.phone+'</div>':'')
    +'</div>'
    // Service details
    +'<div style="margin-bottom:12px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:8px">Service Performed</div>'
      +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px">'
        +reportItem('Machines Serviced',p.machines+' unit'+(p.machines>1?'s':''))
        +reportItem('Ice Bin Cleaned','&#x2713; Complete')
        +reportItem('Evaporator Plates','&#x2713; Descaled &amp; sanitized')
        +reportItem('Water Distribution','&#x2713; Flushed &amp; sanitized')
        +reportItem('Air Filter','&#x2713; Cleaned')
        +reportItem('Sanitizer Treatment','&#x2713; Applied')
        +reportItem('Compliance Standard','FDA Food Code 3-502.12')
        +reportItem('Next Service Due',new Date(Date.now()+30*864e5).toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}))
      +'</div>'
    +'</div>'
    // Condition notes
    +'<div style="margin-bottom:12px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Condition Notes</div>'
      +'<textarea id="report-notes-'+id+'" style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;color:var(--txt);background:var(--surf);outline:none;resize:none" rows="3" placeholder="Machine condition, issues found, recommendations...">'+(customers[id]&&customers[id].report_notes||'')+'</textarea>'
    +'</div>'
    // Signatures
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:10px">'
      +'<div style="border-top:1px solid var(--navy);padding-top:4px">'
        +'<div style="font-size:9px;color:var(--sub)">Technician Signature</div>'
        +'<div style="font-size:10px;color:var(--navy);margin-top:4px">Pinellas Ice Co</div>'
      +'</div>'
      +'<div style="border-top:1px solid var(--navy);padding-top:4px">'
        +'<div style="font-size:9px;color:var(--sub)">Client Signature / Date</div>'
        +'<div style="font-size:10px;color:var(--sub);margin-top:4px">&nbsp;</div>'
      +'</div>'
    +'</div>'
    // Footer
    +'<div style="font-size:9px;color:var(--sub);text-align:center;border-top:1px solid var(--brd2);padding-top:8px">'
      +'This report certifies that all ice machines listed above were cleaned, sanitized, and inspected to FDA and FL DBPR standards. Keep for your compliance records.'
    +'</div>'
    +'</div>'
    +'<div style="display:flex;gap:6px;margin-top:8px">'
      +'<button onclick="printReport()" style="flex:1;padding:9px;border:none;border-radius:8px;background:var(--navy);color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">&#x1F5A8; Print / Save PDF</button>'
      +'<button onclick="saveReportAndLog('+id+')" style="flex:1;padding:9px;border:1px solid #059669;border-radius:8px;background:#ecfdf5;color:#059669;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Log This Visit</button>'
    +'</div>';
}

function reportItem(label,val){
  return '<div style="padding:5px 7px;background:#f5f8fa;border-radius:5px">'
    +'<div style="font-size:8px;color:var(--sub);font-weight:600">'+label+'</div>'
    +'<div style="font-size:10px;font-weight:600;color:var(--navy)">'+val+'</div>'
    +'</div>';
}

function saveReportAndLog(id){
  // Save notes from report textarea
  const notesEl=document.getElementById('report-notes-'+id);
  const notes=notesEl?notesEl.value.trim():'';
  if(notes){
    if(!customers[id])customers[id]={};
    customers[id].report_notes=notes;
    custSave();
  }
  // Open service log modal with notes pre-filled
  openServiceLog(id);
  // After modal opens, fill in the notes
  setTimeout(()=>{
    const svcNotes=document.getElementById('svc-notes');
    if(svcNotes&&notes)svcNotes.value=notes;
  },100);
}

function printReport(){
  const content=document.getElementById('report-content');
  if(!content)return;
  const win=window.open('','_blank');
  win.document.write('<html><head><title>Service Report - Pinellas Ice Co</title>'
    +'<style>body{font-family:system-ui,sans-serif;padding:20px;max-width:600px;margin:0 auto;color:#1e293b}'
    +'*{box-sizing:border-box}textarea{border:1px solid #cbd5e1;border-radius:4px;padding:6px;width:100%;font-family:inherit}'
    +'@media print{button{display:none}}</style></head>'
    +'<body>'+content.outerHTML+'<br><button onclick="window.print()">Print / Save PDF</button></body></html>');
  win.document.close();
  setTimeout(()=>win.print(),400);
}

// ── REFERRAL TRACKING ─────────────────────────────────────────────────────────
function renderReferrals(){
  const el=document.getElementById('svc-refs-list');
  if(!el)return;

  const recurring=P.filter(p=>p.status==='customer_recurring'||p.status==='customer_once');
  if(!recurring.length){
    el.innerHTML='<div class="tempty"><div class="ei">&#x1F91D;</div><div>No clients yet. Close your first deal to start tracking referrals.</div></div>';
    return;
  }

  // Count referrals per client
  const refCounts={};
  recurring.forEach(p=>{
    const c=customers[p.id]||{};
    if(c.referred_by&&c.referred_by!==p.id){
      refCounts[c.referred_by]=(refCounts[c.referred_by]||0)+1;
    }
  });

  el.innerHTML=recurring.map(p=>{
    const c=customers[p.id]||{};
    const refCount=refCounts[p.id]||0;

    return '<div class="svc-card">'
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+' &bull; Won: '+(c.won_date||'?')+'</div>'
        +'</div>'
        +(refCount?'<div style="background:#ecfdf5;border:1px solid #6ee7b7;border-radius:20px;padding:3px 10px;font-size:11px;font-weight:700;color:#059669">'+refCount+' referral'+(refCount>1?'s':'')+'</div>':'')
      +'</div>'
      // Referred by
      +'<div style="font-size:9px;color:var(--sub);margin-bottom:3px">Referred by:</div>'
      +'<div style="display:flex;gap:6px;align-items:center">'
        +'<select onchange="saveReferredBy('+p.id+',this.value)" style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
          +'<option value="">None / Self-generated</option>'
          +recurring.filter(r=>r.id!==p.id).map(r=>'<option value="'+r.id+'"'+(c.referred_by===r.id?' selected':'')+'>'+r.name.slice(0,30)+'</option>').join('')
        +'</select>'
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

function saveReferredBy(bizId,refId){
  if(!customers[bizId])customers[bizId]={};
  customers[bizId].referred_by=refId?parseInt(refId):null;
  custSave();
  renderReferrals();
  toast('Referral source saved');
}

// ── SETTINGS ─────────────────────────────────────────────────────────────────
function loadSettings(){
  try{return JSON.parse(localStorage.getItem('pic_settings')||'{}')||{};}catch(e){return {};}
}
function saveSettings(){
  const portal=(document.getElementById('hs-portal')||{}).value||'';
  const homeZip=(document.getElementById('home-zip')||{}).value||'';
  try{localStorage.setItem('pic_settings',JSON.stringify({hubspot_portal:portal,home_zip:homeZip}));}catch(e){}
}
function initSettings(){
  const s=loadSettings();
  const hsel=document.getElementById('hs-portal');
  if(hsel&&s.hubspot_portal)hsel.value=s.hubspot_portal;
  const hzip=document.getElementById('home-zip');
  if(hzip&&s.home_zip)hzip.value=s.home_zip;
  // Auto-fill route start ZIP from saved home ZIP
  const rzip=document.getElementById('rzip');
  if(rzip&&s.home_zip){
    if(!rzip.value)rzip.value=s.home_zip;
    rzip.placeholder=s.home_zip;
  }
}

// ── QUOTE BUILDER ─────────────────────────────────────────────────────────────
let vendors={};  // {bizId: vendorName}

function contactsLoad(){
  try{contacts=JSON.parse(localStorage.getItem('pic_contacts')||'{}')||{};}catch(e){contacts={};}
  try{vendors=JSON.parse(localStorage.getItem('pic_vendors')||'{}')||{};}catch(e){vendors={};}
}
function vendorsSave(){
  try{localStorage.setItem('pic_vendors',JSON.stringify(vendors));}catch(e){}
}
function saveVendor(){
  if(!cur)return;
  const v=(document.getElementById('mc-vendor')||{}).value.trim();
  if(!v){toast('Enter a vendor name');return;}
  vendors[cur.id]=v;
  vendorsSave();
  renderVendor(cur.id);
  toast('Vendor saved');
}
function renderVendor(id){
  const inp=document.getElementById('mc-vendor');
  const disp=document.getElementById('mc-vendor-display');
  const v=vendors[id]||'';
  if(inp)inp.value=v;
  if(disp)disp.textContent=v?'Currently: '+v:'Not recorded yet';
}
function contactsSave(){
  try{localStorage.setItem('pic_contacts',JSON.stringify(contacts));}catch(e){}
}

function renderContacts(id){
  const el=document.getElementById('mcontacts');
  if(!el)return;
  const list=contacts[id]||[];
  if(!list.length){el.innerHTML='<div style="font-size:10px;color:var(--sub);margin-bottom:6px">No contacts saved yet</div>';return;}
  const ROLE_LABELS={owner:'Owner',gm:'GM',manager:'Manager',chef:'Chef/Kitchen',staff:'Staff'};
  el.innerHTML=list.map((c,i)=>
    '<div style="display:flex;align-items:center;gap:8px;padding:6px 0;border-bottom:1px solid var(--brd2)">'
    +'<div style="flex:1">'
      +'<div style="font-size:11px;font-weight:600;color:var(--navy)">'+c.name+'</div>'
      +'<div style="font-size:9px;color:var(--sub)">'+( ROLE_LABELS[c.role]||c.role)+(c.phone?' &bull; '+c.phone:'')+'</div>'
    +'</div>'
    +(c.phone?'<a href="tel:'+c.phone.replace(/\s/g,'')+'" style="font-size:9px;padding:3px 7px;border:1px solid var(--blu);border-radius:5px;color:var(--blu);text-decoration:none" onclick="event.stopPropagation()">Call</a>':'')
    +'<button onclick="deleteContact('+id+','+i+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer">✕</button>'
    +'</div>'
  ).join('');
}

function addContact(){
  if(!cur)return;
  const name=document.getElementById('mc-name').value.trim();
  const role=document.getElementById('mc-role').value;
  const phone=document.getElementById('mc-phone').value.trim();
  if(!name){toast('Enter a contact name');return;}
  if(!contacts[cur.id])contacts[cur.id]=[];
  contacts[cur.id].push({name,role,phone});
  contactsSave();
  document.getElementById('mc-name').value='';
  document.getElementById('mc-phone').value='';
  renderContacts(cur.id);
  toast('Contact saved');
}

function deleteContact(bizId,idx){
  if(!contacts[bizId])return;
  contacts[bizId].splice(idx,1);
  if(!contacts[bizId].length)delete contacts[bizId];
  contactsSave();
  renderContacts(bizId);
}

// ── INIT ─────────────────────────────────────────────────────────────────────
// INIT
function init(){
  lLoad();phLoad();custLoad();contactsLoad();initSettings();initGoals();setPreset('all');rT();renderBriefing();
  const si=document.getElementById('si');if(si)si.blur();
  // Show FAB on mobile
  const fab=document.getElementById('svc-fab');
  if(fab&&window.innerWidth<=480)fab.style.display='flex';
  window.addEventListener('resize',()=>{
    if(fab)fab.style.display=window.innerWidth<=480?'flex':'none';
  });
}
if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',init);}else{setTimeout(init,50);}
// ── OFFLINE DETECTION ────────────────────────────────────────────────────────
function updateOnlineStatus(){
  const banner=document.getElementById('offline-banner');
  if(!banner)return;
  if(!navigator.onLine){
    banner.classList.add('active');
    banner.style.display='flex';
    document.body.style.paddingTop='32px';
  } else {
    banner.classList.remove('active');
    banner.style.display='none';
    document.body.style.paddingTop='';
  }
}
window.addEventListener('online', updateOnlineStatus);
window.addEventListener('offline', updateOnlineStatus);
updateOnlineStatus();

// ── SERVICE WORKER ─────────────────────────────────────────────────────────
if('serviceWorker' in navigator){
  window.addEventListener('load',()=>{
    navigator.serviceWorker.register('sw.js')
      .then(reg=>{
        // Check for updates
        reg.addEventListener('updatefound',()=>{
          const newWorker=reg.installing;
          newWorker.addEventListener('statechange',()=>{
            if(newWorker.state==='installed'&&navigator.serviceWorker.controller){
              toast('New version available — reload to update');
            }
          });
        });
      })
      .catch(()=>{});
  });
}
</script>
</body>
</html>
"""

# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────
SW_JS = """const CACHE_NAME='pic-v3';
const ASSETS=['./','./ index.html'];
self.addEventListener('install',e=>{
  e.waitUntil(caches.open(CACHE_NAME).then(c=>c.addAll(ASSETS).catch(()=>{})));
  self.skipWaiting();
});
self.addEventListener('activate',e=>{
  e.waitUntil(caches.keys().then(ks=>Promise.all(ks.filter(k=>k!==CACHE_NAME).map(k=>caches.delete(k)))));
  self.clients.claim();
});
self.addEventListener('fetch',e=>{
  const url=new URL(e.request.url);
  if(url.origin===self.location.origin){
    e.respondWith(caches.match(e.request).then(cached=>{
      const fresh=fetch(e.request).then(res=>{
        if(res&&res.status===200){const copy=res.clone();caches.open(CACHE_NAME).then(c=>c.put(e.request,copy));}
        return res;
      }).catch(()=>cached);
      return cached||fresh;
    }));
  } else {
    e.respondWith(fetch(e.request).catch(()=>caches.match(e.request)));
  }
});
"""

def main():
    folder = Path(__file__).parent

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if Path(arg).is_dir():
            # Directory passed (e.g. "data/") -- find all CSV and XLSX files
            data_dir = Path(arg)
            csv_paths = (
                sorted(data_dir.glob('*.csv')) +
                sorted(data_dir.glob('*.xlsx')) +
                sorted(data_dir.glob('*.xls'))
            )
            # Exclude the license extract from inspection data
            csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                         and 'hrfood' not in p.name.lower()]
            if not csv_paths:
                print(f"\nNo inspection data files found in {data_dir}")
                sys.exit(1)
            print(f"Found {len(csv_paths)} file(s) in {data_dir}:")
            for p in csv_paths:
                print(f"  {p.name}")
            print()
        else:
            csv_paths = sys.argv[1:]
    else:
        # Auto-find in current folder
        csv_paths = (
            sorted(folder.glob('*.csv')) +
            sorted(folder.glob('*.xlsx'))
        )
        csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                     and 'hrfood' not in p.name.lower()]
        if not csv_paths:
            print("\nNo data files found. Run: python build.py data/")
            sys.exit(1)
        print(f"Auto-found {len(csv_paths)} file(s):")
        for p in csv_paths:
            print(f"  {p.name}")
        print()

    records = run(list(map(str, csv_paths)))
    print(f"\nGenerating HTML...")
    html = build_html(records)
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(html, encoding='utf-8')
    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"  Written: {OUTPUT_FILE.name} ({size_kb}KB)")
    # Write sw.js for PWA offline support
    sw_path = OUTPUT_FILE.parent / 'sw.js'
    sw_path.write_text(SW_JS, encoding='utf-8')
    print(f"  Written: sw.js")
    print(f"\n{'='*55}")
    print(f"  Done! Open prospecting_tool.html in Chrome.")
    print(f"  Your call log carries over automatically.")
    print(f"{'='*55}\n")

def main():
    folder = Path(__file__).parent

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if Path(arg).is_dir():
            # Directory passed (e.g. "data/") -- find all CSV and XLSX files
            data_dir = Path(arg)
            csv_paths = (
                sorted(data_dir.glob('*.csv')) +
                sorted(data_dir.glob('*.xlsx')) +
                sorted(data_dir.glob('*.xls'))
            )
            # Exclude the license extract from inspection data
            csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                         and 'hrfood' not in p.name.lower()]
            if not csv_paths:
                print(f"\nNo inspection data files found in {data_dir}")
                sys.exit(1)
            print(f"Found {len(csv_paths)} file(s) in {data_dir}:")
            for p in csv_paths:
                print(f"  {p.name}")
            print()
        else:
            csv_paths = sys.argv[1:]
    else:
        # Auto-find in current folder
        csv_paths = (
            sorted(folder.glob('*.csv')) +
            sorted(folder.glob('*.xlsx'))
        )
        csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                     and 'hrfood' not in p.name.lower()]
        if not csv_paths:
            print("\nNo data files found. Run: python build.py data/")
            sys.exit(1)
        print(f"Auto-found {len(csv_paths)} file(s):")
        for p in csv_paths:
            print(f"  {p.name}")
        print()

    records = run(list(map(str, csv_paths)))
    print(f"\nGenerating HTML...")
    html = build_html(records)
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(html, encoding='utf-8')
    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"  Written: {OUTPUT_FILE.name} ({size_kb}KB)")
    # Write sw.js for PWA offline support
    sw_path = OUTPUT_FILE.parent / 'sw.js'
    sw_path.write_text(SW_JS, encoding='utf-8')
    print(f"  Written: sw.js")
    print(f"\n{'='*55}")
    print(f"  Done! Open prospecting_tool.html in Chrome.")
    print(f"  Your call log carries over automatically.")
    print(f"{'='*55}\n")

if __name__ == '__main__':
    main()