#!/usr/bin/env python3
"""
Pinellas Ice Co \u2014 Prospect Tool Builder
Complete, self-contained. Drop in any folder with your CSV files and run.
"""

import sys, os, json, re, warnings, csv
from pathlib import Path
from datetime import date, timedelta
from math import radians, cos, sin, sqrt, atan2
from collections import Counter

warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
TARGET_COUNTIES  = ['Pinellas', 'Hillsborough', 'Pasco',
                    'Citrus', 'Hernando', 'Polk', 'Sumter']
MIN_SCORE        = 5   # Very low - include everything, filter in browser
TODAY            = date.today()
OUTPUT_FILE      = Path(__file__).parent / 'prospecting_tool.html'

# ── CHAIN CLASSIFICATION ──────────────────────────────────────────────────────
# Corporate chains: national vendor contracts, skip entirely
CORPORATE_CHAINS = [
    'mcdonald','burger king','wendy','chick-fil-a','starbucks','chipotle',
    'ihop','waffle house','olive garden','red lobster','outback steakhouse',
    'buffalo wild wings','panera bread','cracker barrel','texas roadhouse',
    'golden corral','panda express','jack in the box','raising cane',
    'shake shack','wingstop','habit burger',
]

# Franchise brands: independently owned, CAN be prospects
FRANCHISE_BRANDS = [
    'subway','taco bell','dunkin','pizza hut','domino','papa john',
    'little caesar','sonic','dairy queen','five guys','popeyes','kfc',
    'chilis','applebee','denny','hooters','longhorn','buffalo wild wing',
    'jersey mike','jimmy john','firehouse subs','mcalister',
    'tijuana flats','miller ale house',
]

def classify_business(name):
    """Returns ('corporate','franchise','independent') + is_chain bool."""
    n = name.lower()
    for c in CORPORATE_CHAINS:
        if c in n: return 'corporate', True
    for f in FRANCHISE_BRANDS:
        if f in n: return 'franchise', True
    return 'independent', False

# ── MACHINE ESTIMATION ────────────────────────────────────────────────────────
def est_machines(seats, is_full_bar, rank_code):
    """Estimate ice machine count from seat count + license type."""
    if rank_code in ('NOST','CNOSEAT','MFDV'): return 1
    if seats <= 0:    base = 1
    elif seats < 40:  base = 1
    elif seats < 80:  base = 1
    elif seats < 150: base = 2
    elif seats < 250: base = 2
    elif seats < 400: base = 3
    else:             base = 4
    bonus = 1 if is_full_bar else 0
    return min(base + bonus, 6)

def est_monthly(machines):
    """Estimated monthly recurring revenue at $149/mo first machine, $99 each add'l."""
    if machines <= 1: return 149
    return 149 + (machines - 1) * 99

def est_onetime(machines):
    """One-time deep clean price."""
    return 299 + (machines - 1) * 150

def account_tier(seats, rank_code, machines, chronic, confirmed):
    """Account quality tier."""
    if rank_code in ('NOST','CNOSEAT') and not confirmed and not chronic: return 'COLD'
    if seats > 0 and seats < 15 and not confirmed and not chronic: return 'COLD'
    if machines >= 3 and (chronic or confirmed): return 'PLATINUM'
    if machines >= 2 and (chronic or confirmed): return 'GOLD'
    if machines >= 2:                            return 'SILVER'
    if confirmed and chronic:                    return 'GOLD'
    if confirmed:                                return 'SILVER'
    if chronic:                                  return 'SILVER'
    if seats >= 60:                              return 'BRONZE'
    return 'COLD'

def confidence_score(n_inspections, n_years_data, has_ice_history, days_since):
    """
    Confidence in the scoring/prediction (0-100).
    Higher with more inspection history, recent data, ice violation history.
    """
    base = 30
    # More inspections = more confident
    if n_inspections >= 6:   base += 25
    elif n_inspections >= 3: base += 15
    elif n_inspections >= 2: base += 8
    # More years of data
    if n_years_data >= 4:    base += 20
    elif n_years_data >= 2:  base += 10
    # Ice history is a strong signal
    if has_ice_history:      base += 15
    # Recent inspection = more reliable
    if days_since <= 30:     base += 10
    elif days_since <= 90:   base += 5
    return min(100, base)

def seat_score_bonus(machines, rank_code):
    if rank_code in ('NOST','CNOSEAT'): return 0
    if machines >= 3: return 15
    if machines >= 2: return 10
    return 0


# ──────────────────────────────────────────────────────────────────────────────
# STEP 0: CHECK DEPENDENCIES
# ──────────────────────────────────────────────────────────────────────────────
def check_deps():
    missing = []
    for pkg, install in [('pandas','pandas'), ('sklearn','scikit-learn'), 
                          ('numpy','numpy'), ('openpyxl','openpyxl')]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(install)
    if missing:
        print("\n" + "="*55)
        print("  Missing packages. Run this once in terminal:")
        print(f"  pip install {' '.join(missing)}")
        print("="*55 + "\n")
        sys.exit(1)

check_deps()

import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN NORMALIZATION
# ──────────────────────────────────────────────────────────────────────────────
COLUMN_MAP = {
    'Inspection Date': ['InspectionDate','INSPECTION DATE','inspection_date','Date','Insp Date'],
    'License ID': ['LicenseID','LICENSE ID','license_id','License Number','LicenseNumber','Lic ID','License_ID'],
    'Business (DBA-Does Business As) Name': [
        'Business Name','DBA Name','DBA','Business','BUSINESS NAME',
        'DBA-Does Business As','Restaurant Name','RestaurantName','Establishment Name'],
    'County Name': ['County','COUNTY','county_name','CountyName'],
    'Location Address': ['Address','LOCATION ADDRESS','Street Address','StreetAddress','Addr'],
    'Location City': ['City','CITY','city_name'],
    'Location Zip Code': ['Zip','ZIP','Zip Code','ZipCode','Postal Code','PostalCode','zip_code'],
    'Inspection Type': ['InspectionType','INSPECTION TYPE','Type','Insp Type'],
    'Inspection Disposition': ['Disposition','DISPOSITION','InspectionDisposition','Result'],
    'Number of High Priority Violations': [
        'High Priority Violations','HighPriorityViolations','High Violations','HP Violations','HV'],
    'Number of Total Violations': ['Total Violations','TotalViolations','Total','TV'],
    'Number of Intermediate Violations': ['Intermediate Violations','IntermediateViolations','IV'],
    'Number of Basic Violations': ['Basic Violations','BasicViolations','BV'],
    'Visit Number': ['VisitNumber','Visit','VISIT NUMBER','Visit_Number'],
}

DISP_RISK = {
    'Inspection Completed - No Further Action': 0, 'Warning Issued': 1,
    'Call Back - Complied': 2, 'Call Back - Extension given, pending': 3,
    'Admin. Complaint Callback Complied': 3, 'Call Back - Admin. complaint recommended': 4,
    'Administrative complaint recommended': 4, 'Administrative determination recommended': 4,
    'Emergency Order Callback Complied': 3, 'Emergency Order Callback Time Extension': 4,
    'Emergency Order Callback Not Complied': 5, 'Emergency order recommended': 5,
}

ICE_CODES = {14: 100, 22: 80, 50: 70, 23: 60, 37: 50, 51: 30, 36: 20, 55: 15}

ZIP_COORDS = {
    '33511':(27.924,-82.312),'33547':(27.861,-82.212),'33570':(27.714,-82.398),
    '33572':(27.752,-82.378),'33573':(27.710,-82.354),'33578':(27.855,-82.324),
    '33579':(27.793,-82.278),'33584':(27.992,-82.269),'33592':(28.096,-82.280),
    '33594':(27.910,-82.232),'33596':(27.875,-82.258),'33598':(27.660,-82.349),
    '33602':(27.949,-82.459),'33603':(27.980,-82.468),'33604':(28.001,-82.468),
    '33605':(27.956,-82.427),'33606':(27.931,-82.474),'33607':(27.966,-82.489),
    '33609':(27.945,-82.495),'33610':(27.978,-82.406),'33611':(27.900,-82.508),
    '33612':(28.022,-82.457),'33613':(28.055,-82.448),'33614':(27.997,-82.509),
    '33615':(27.989,-82.563),'33616':(27.894,-82.532),'33617':(28.012,-82.407),
    '33618':(28.053,-82.499),'33619':(27.950,-82.389),'33620':(28.062,-82.415),
    '33621':(27.860,-82.520),'33624':(28.069,-82.535),'33625':(28.080,-82.556),
    '33626':(28.082,-82.585),'33629':(27.924,-82.522),'33634':(27.993,-82.567),
    '33635':(28.010,-82.609),'33637':(28.056,-82.380),'33647':(28.119,-82.378),
    '33701':(27.770,-82.634),'33702':(27.811,-82.647),'33703':(27.817,-82.633),
    '33704':(27.798,-82.640),'33705':(27.750,-82.638),'33706':(27.747,-82.747),
    '33707':(27.761,-82.717),'33708':(27.810,-82.800),'33709':(27.806,-82.754),
    '33710':(27.785,-82.734),'33711':(27.745,-82.682),'33712':(27.726,-82.663),
    '33713':(27.779,-82.669),'33714':(27.803,-82.685),'33715':(27.703,-82.717),
    '33716':(27.855,-82.652),'33755':(27.966,-82.800),'33756':(27.953,-82.801),
    '33759':(27.974,-82.750),'33760':(27.900,-82.687),'33761':(28.014,-82.740),
    '33762':(27.870,-82.676),'33763':(28.005,-82.748),'33764':(27.934,-82.740),
    '33765':(27.984,-82.777),'33767':(27.978,-82.828),'33770':(27.910,-82.789),
    '33771':(27.912,-82.762),'33772':(27.878,-82.795),'33773':(27.880,-82.720),
    '33774':(27.869,-82.833),'33776':(27.839,-82.815),'33777':(27.854,-82.754),
    '33778':(27.873,-82.740),'33781':(27.841,-82.717),'33782':(27.863,-82.697),
    '33785':(27.740,-82.781),'33786':(27.924,-82.826),'34677':(28.053,-82.669),
    '34681':(28.011,-82.688),'34683':(28.080,-82.765),'34684':(28.102,-82.769),
    '34685':(28.117,-82.731),'34688':(28.128,-82.704),'34689':(28.146,-82.756),
    '34695':(27.986,-82.695),'34698':(28.019,-82.775),'34652':(28.240,-82.727),
    '34653':(28.262,-82.724),'34654':(28.279,-82.666),'34655':(28.218,-82.673),
    '34667':(28.338,-82.652),'34668':(28.280,-82.697),'34669':(28.226,-82.706),
    '34690':(28.243,-82.675),'34691':(28.195,-82.754),'34638':(28.190,-82.712),
    '34637':(28.214,-82.666),'34639':(28.234,-82.628),
    '34429':(28.896,-82.580),'34442':(28.767,-82.472),'34448':(28.956,-82.582),
    '34450':(28.836,-82.330),'34452':(28.814,-82.335),'34465':(28.947,-82.469),
    '34601':(28.552,-82.388),'34604':(28.538,-82.466),'34606':(28.468,-82.549),
    '34608':(28.506,-82.540),'34609':(28.441,-82.480),'34610':(28.369,-82.547),
    '34613':(28.527,-82.517),'33801':(28.041,-81.975),'33803':(28.020,-81.970),
    '33810':(28.059,-82.036),'33811':(27.945,-82.028),'33813':(27.969,-82.001),
    '33823':(28.071,-81.897),'33830':(27.899,-81.840),'33880':(28.023,-81.731),
    '33897':(28.291,-81.643),'32162':(28.950,-81.966),'34785':(28.839,-82.051),
}

HIGH_ICE = ['bar','pub','tavern','lounge','grill','grille','sports','club',
            'seafood','sushi','hibachi','oyster','marina','brewery','taproom',
            'resort','hotel','inn','cantina','steakhouse']
LOW_ICE  = ['bakery','donut','coffee','smoothie','juice','dessert','cupcake']

# Median days until callback inspection per disposition type.
# Derived from 25k+ FL DBPR inspection interval records — data-driven, not guesses.
CALLBACK_DAYS = {
    'Emergency order recommended':               1,
    'Emergency Order Callback Not Complied':     1,
    'Administrative complaint recommended':      4,
    'Administrative determination recommended':  4,
    'Call Back - Admin. complaint recommended':  22,
    'Admin. Complaint Callback Complied':        45,
    'Warning Issued':                            12,
    'Call Back - Extension given, pending':      50,
    'Emergency Order Callback Time Extension':   62,
    'Emergency Order Callback Complied':         64,
    'Call Back - Complied':                      118,
    'Inspection Completed - No Further Action':  130,
}
STATIC_PHONES = {
    "9784166": "+1 813-993-3924",
    "2183411": "+1 813-752-2236",
    "9681679": "+1 813-657-5600",
    "9468965": "+1 813-374-5363",
    "9473519": "+1 727-314-4004",
    "9405890": "+1 813-909-6354",
    "9428645": "+1 813-265-2111",
    "9435298": "+1 813-295-8450",
    "9317331": "+1 727-242-8400",
    "9394343": "+1 813-344-3311",
    "9283918": "+1 813-815-7662",
    "9296161": "+1 727-381-0088",
    "4115014": "+1 727-844-3125",
    "3754544": "+1 727-924-2000",
    "3774587": "+1 727-586-5797",
    "3243192": "+1 727-443-4900",
    "3425752": "+1 813-654-4262",
    "3523975": "+1 813-825-1373",
    "9256830": "+1 727-270-6655",
    "9155471": "+1 813-994-9666",
    "4421753": "+1 813-961-4092",
    "7089467": "+1 727-726-4608",
    "2976346": "+1 813-374-2739",
    "6462895": "+1 727-934-4047",
    "2238952": "+1 727-327-8090",
    "5654752": "+1 727-863-0965",
    "5787421": "+1 727-614-0574",
    "7432888": "+1 727-954-7369"
}

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def normalize_columns(df):
    col_lower = {c.lower().strip(): c for c in df.columns}
    renames = {}
    for standard, variants in COLUMN_MAP.items():
        if standard in df.columns:
            continue
        for v in variants:
            if v.lower().strip() in col_lower:
                renames[col_lower[v.lower().strip()]] = standard
                break
    # Violation column variants: V14, Viol14, violation_14 → Violation 14
    viol_pat = re.compile(r'^[Vv](?:iol(?:ation)?[\s_-]?)?0?(\d+)$')
    for col in list(df.columns):
        m = viol_pat.match(col.strip())
        if m:
            new = f'Violation {int(m.group(1)):02d}'
            if new != col and col not in renames:
                renames[col] = new
    if renames:
        df = df.rename(columns=renames)
    return df

def load_csvs(paths):
    frames, ref_cols = [], None
    for p in paths:
        path = Path(p)
        if not path.exists():
            print(f"  WARNING: {path.name} not found, skipping")
            continue
        try:
            ext = path.suffix.lower()
            if ext in ('.xlsx', '.xlsm', '.xls'):
                # Stream XLSX row-by-row to handle large statewide files
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                headers = [str(c).strip() if c is not None else '' for c in rows[0]]
                data = [['' if v is None else str(v) for v in row] for row in rows[1:]]
                df = pd.DataFrame(data, columns=headers)
                print(f"  {path.name}: {len(df):,} rows (xlsx)")
            else:
                df = pd.read_csv(path, low_memory=False, encoding='utf-8',
                                 encoding_errors='replace')
                print(f"  {path.name}: {len(df):,} rows")

            df.columns = df.columns.str.strip()
            df = normalize_columns(df)
            if ref_cols is None:
                ref_cols = list(df.columns)
            frames.append(df)

        except Exception as e:
            try:
                if ref_cols and path.suffix.lower() == '.csv':
                    df = pd.read_csv(path, header=None, names=ref_cols,
                                     low_memory=False, encoding_errors='replace')
                    df = normalize_columns(df)
                    frames.append(df)
                    print(f"  {path.name}: {len(df):,} rows (headerless)")
                else:
                    print(f"  SKIP {path.name}: {e}")
            except Exception as e2:
                print(f"  SKIP {path.name}: {e2}")

    if not frames:
        raise RuntimeError("No valid data files loaded.")
    return pd.concat(frames, ignore_index=True)

def load_license_extract(data_dir=None):
    """
    Load FL DBPR active license extract.
    Returns dict keyed by License Number with phone, seats, rank, license_type.
    Tries data/ folder first, then current folder.
    """
    import pandas as pd
    candidates = []
    if data_dir:
        candidates.append(Path(data_dir) / 'hrfood3_licenses.csv')
    candidates += [
        Path(__file__).parent / 'data' / 'hrfood3_licenses.csv',
        Path(__file__).parent / 'hrfood3_licenses.csv',
    ]
    for path in candidates:
        if path.exists():
            try:
                df = pd.read_csv(path, low_memory=False, encoding_errors='replace')
                df.columns = df.columns.str.strip()
                # Normalize license number -- strip prefix letters for join
                # DBPR license numbers in extract: "SEA6213532"
                # In inspection data: numeric "6213532" or full "SEA6213532"
                result = {}
                for _, r in df.iterrows():
                    lic_raw = str(r.get('License Number', '')).strip()
                    # Store under both full and numeric forms
                    phone = str(r.get('Primary Phone Number', '') or
                                r.get('Secondary Phone Number', '') or '').strip()
                    phone = re.sub(r'[^0-9]', '', phone)
                    if len(phone) == 10:
                        phone = '+1 ' + phone[:3] + '-' + phone[3:6] + '-' + phone[6:]
                    elif len(phone) == 11 and phone.startswith('1'):
                        phone = '+1 ' + phone[1:4] + '-' + phone[4:7] + '-' + phone[7:]
                    else:
                        phone = ''
                    seats_raw = r.get('Number of Seats or Rental Units', 0)
                    try:
                        seats = int(float(str(seats_raw))) if seats_raw and str(seats_raw).strip() not in ('', 'nan') else 0
                    except:
                        seats = 0
                    entry = {
                        'phone':        phone,
                        'seats':        seats,
                        'rank':         str(r.get('Rank Code', '') or '').strip(),
                        'license_type': str(r.get('License Type Code', '') or '').strip(),
                        'risk_level':   str(r.get('Base Risk Level', '') or '').strip(),
                    }
                    result[lic_raw] = entry
                    # Also index by numeric portion for flexible joins
                    numeric = re.sub(r'[^0-9]', '', lic_raw)
                    if numeric and numeric not in result:
                        result[numeric] = entry
                n_phones = sum(1 for v in result.values() if v.get('phone'))
                print(f"  License extract: {len(df):,} records, {n_phones:,} with phones")
                return result
            except Exception as e:
                print(f"  WARNING: license extract load failed -- {e}")
    print("  License extract not found -- phones/seats from extract unavailable")
    return {}

def match_license(license_id, extract):
    """Flexible license ID matching -- try multiple formats."""
    lid = str(license_id).strip()
    if lid in extract:
        return extract[lid]
    numeric = re.sub(r'[^0-9]', '', lid)
    if numeric in extract:
        return extract[numeric]
    # Try common prefixes
    for prefix in ('SEA', 'NOS', 'MFD'):
        key = prefix + numeric
        if key in extract:
            return extract[key]
    return None

def ice_usage_label(name):
    n = name.lower()
    if any(w in n for w in HIGH_ICE): return 'high'
    if any(w in n for w in LOW_ICE):  return 'low'
    return 'medium'

def ice_insight_text(codes):
    if 'V14' in codes: return 'Food contact surfaces (V14) \u2014 inspectors directly cited food contact equipment. Ice machines are the primary target of this violation code.'
    if 'V22' in codes: return 'Non-PHF food contact surfaces (V22) \u2014 ice machines fall directly into this category.'
    if 'V50' in codes: return 'Food contact surfaces not clean (V50) \u2014 same root issue as V14/V22, cited at basic level.'
    if 'V37' in codes: return 'Equipment not in good repair (V37) \u2014 ice machine may be broken or malfunctioning.'
    if 'V23' in codes: return 'Utensil/container sanitation (V23) \u2014 includes ice scoops and bin components.'
    return ''

def _callback_fields(disposition: str, last_insp_date) -> dict:
    """Compute estimated callback inspection date from disposition type."""
    cb_days = CALLBACK_DAYS.get(disposition, 30)
    est_callback = last_insp_date + timedelta(days=cb_days)
    days_to_callback = (TODAY - est_callback).days * -1
    urgency = ('overdue'   if days_to_callback < 0  else
               'imminent'  if days_to_callback <= 7  else
               'soon'      if days_to_callback <= 21 else 'upcoming')
    return {
        'est_callback_date': est_callback.isoformat(),
        'days_to_callback':  days_to_callback,
        'callback_urgency':  urgency,
    }

def load_emergency_closures(data_dir=None):
    """Load recent DBPR emergency closure reports. Returns set of license IDs."""
    import openpyxl
    closed_ids = set()
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        if not d.exists(): continue
        for f in sorted(d.glob('EOS_Weekly_Extract_*.xlsx'), reverse=True)[:4]:
            try:
                wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows: continue
                headers = [str(h).strip() if h else '' for h in rows[0]]
                lic_col = next((i for i,h in enumerate(headers)
                                if 'license' in h.lower()), None)
                if lic_col is None: continue
                for row in rows[1:]:
                    val = row[lic_col]
                    if val:
                        num = re.sub(r'[^0-9]', '', str(val))
                        if num: closed_ids.add(num)
            except Exception:
                pass
    if closed_ids:
        print(f"  Emergency closures: {len(closed_ids)} businesses flagged")
    return closed_ids

def nominatim_geocode(address, city, state='FL', cache={}):
    """Geocode a single address using Nominatim (free, no key required).
    Uses in-memory cache to avoid duplicate requests."""
    import time, urllib.request, urllib.parse
    key = f"{address},{city},{state}"
    if key in cache:
        return cache[key]
    try:
        query = urllib.parse.urlencode({'q': f"{address}, {city}, {state}",
                                        'format': 'json', 'limit': 1})
        url = f"https://nominatim.openstreetmap.org/search?{query}"
        req = urllib.request.Request(url,
            headers={'User-Agent': 'PinellasIceCo/1.0 ice machine inspection tool'})
        with urllib.request.urlopen(req, timeout=5) as r:
            results = json.loads(r.read())
        if results:
            lat = float(results[0]['lat'])
            lon = float(results[0]['lon'])
            cache[key] = (lat, lon)
            time.sleep(1.1)  # Nominatim rate limit: 1 req/sec
            return (lat, lon)
    except Exception:
        pass
    cache[key] = None
    return None

def load_geo_cache(data_dir=None):
    """Load persistent geocoding cache from disk."""
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        p = d / 'geocache.json'
        if p.exists():
            try:
                data = json.loads(p.read_text())
                print(f"  Geocache: {len(data)} cached addresses")
                return data
            except Exception:
                pass
    return {}

def save_geo_cache(cache, data_dir=None):
    """Save geocoding cache to disk for next run."""
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        try:
            d.mkdir(exist_ok=True)
            (d / 'geocache.json').write_text(json.dumps(cache))
            return
        except Exception:
            pass

# ──────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ──────────────────────────────────────────────────────────────────────────────
def run(csv_paths):
    print(f"\n{'='*55}")
    print(f"  Pinellas Ice Co \u2014 Building Prospect Tool")
    print(f"  {TODAY}")
    print(f"{'='*55}\n")

    # LOAD
    print("Loading CSV files...")
    df = load_csvs(csv_paths)
    print(f"  Total rows loaded: {len(df):,}\n")

    # LOAD LICENSE EXTRACT (phones + seats + rank)
    print("Loading license extract...")
    data_dir = Path(csv_paths[0]).parent if csv_paths else None
    license_data = load_license_extract(data_dir)

    # LOAD EMERGENCY CLOSURES
    print("Loading emergency closures...")
    emergency_ids = load_emergency_closures(data_dir)

    # LOAD GEO CACHE
    geo_cache = load_geo_cache(data_dir)
    print()

    # COERCE
    df['inspection_date'] = pd.to_datetime(df.get('Inspection Date'), errors='coerce')
    df = df.dropna(subset=['inspection_date'])
    df['county']    = df.get('County Name',          pd.Series('', index=df.index)).astype(str).str.strip()
    df['biz_name']  = df.get('Business (DBA-Does Business As) Name',
                              pd.Series('', index=df.index)).astype(str).str.strip()
    df['address']   = df.get('Location Address',     pd.Series('', index=df.index)).astype(str).str.strip()
    df['city']      = df.get('Location City',        pd.Series('', index=df.index)).astype(str).str.strip()
    df['zip']       = df.get('Location Zip Code',    pd.Series('', index=df.index)).astype(str).str.strip().str[:5]
    df['disp_raw']  = df.get('Inspection Disposition',pd.Series('',index=df.index)).astype(str).str.strip()
    df['insp_type'] = df.get('Inspection Type',      pd.Series('', index=df.index)).astype(str).str.strip()
    df['license_id']= df.get('License ID',           pd.Series(range(len(df)), index=df.index))

    for col, src_name in [('hv','Number of High Priority Violations'),
                           ('tv','Number of Total Violations'),
                           ('iv','Number of Intermediate Violations'),
                           ('bv','Number of Basic Violations'),
                           ('vn','Visit Number')]:
        df[col] = pd.to_numeric(df.get(src_name, pd.Series(0, index=df.index)),
                                errors='coerce').fillna(0)

    # Violation codes
    for code in ICE_CODES:
        col = f'Violation {code:02d}'
        df[col] = pd.to_numeric(df.get(col, pd.Series(0, index=df.index)),
                                errors='coerce').fillna(0)

    df['ice_rel']    = sum(df[f'Violation {c:02d}'].clip(upper=1) * s for c,s in ICE_CODES.items())
    df['direct_ice'] = (df['Violation 14'].clip(upper=1) +
                        df['Violation 22'].clip(upper=1) +
                        df['Violation 50'].clip(upper=1)) > 0
    df['dr'] = df['disp_raw'].map(DISP_RISK).fillna(2)

    # Filter
    df_t = df[df['county'].isin(TARGET_COUNTIES)].copy()
    print(f"Rows in target counties: {len(df_t):,}")

    # Routine inspections
    routine = df_t[df_t['insp_type'] == 'Routine - Food'].copy()
    routine = routine.sort_values(['license_id', 'inspection_date'])
    routine['next_date'] = routine.groupby('license_id')['inspection_date'].shift(-1)
    routine['interval']  = (routine['next_date'] - routine['inspection_date']).dt.days
    rv = routine[(routine['interval'] > 0) & (routine['interval'] < 730)].copy()
    MED = float(rv['interval'].median()) if len(rv) else 125.0
    print(f"Training on {len(rv):,} inspection intervals (median={MED:.0f}d)")

    # Train model
    le = LabelEncoder()
    rv['county_enc'] = le.fit_transform(rv['county'].fillna('Unknown'))
    rv['prev']  = rv.groupby('license_id')['interval'].transform(
        lambda x: x.expanding().mean().shift(1)).fillna(MED)
    rv['month'] = rv['inspection_date'].dt.month
    FEATS = ['hv','iv','bv','tv','dr','vn','month','county_enc','prev']
    rf = RandomForestRegressor(n_estimators=120, max_depth=8, random_state=42, n_jobs=-1)
    rf.fit(rv[FEATS].fillna(0), rv['interval'])
    print("Prediction model trained.\n")

    # Per-establishment history
    ice_hist = routine.groupby('license_id').agg(
        avg_interval  =('interval','mean'),
        n_insp        =('interval','count'),
        avg_hv        =('hv','mean'),
        avg_ice       =('ice_rel','mean'),
        direct_count  =('direct_ice','sum'),
    ).reset_index()
    ice_hist['chronic']   = ice_hist['direct_count'] >= 2
    ice_hist['confirmed'] = ice_hist['direct_count'] >= 1

    # Trend slope
    def trend_slope(g):
        if len(g) < 3: return 0.0
        x = np.arange(len(g), dtype=float)
        return float(np.polyfit(x, g['hv'].values.astype(float), 1)[0])
    trends = routine.groupby('license_id').apply(trend_slope).to_dict()

    # Fired ice codes per establishment
    def get_fired(g):
        codes = []
        for code in ICE_CODES:
            col = f'Violation {code:02d}'
            if col in g.columns and (pd.to_numeric(g[col], errors='coerce').fillna(0) > 0).any():
                codes.append(f'V{code}')
        return codes[:4]
    fired = routine.groupby('license_id').apply(get_fired).to_dict()

    # Predict next inspection date
    latest = routine.sort_values('inspection_date').groupby('license_id').last().reset_index()
    latest = latest.merge(ice_hist, on='license_id', how='left')
    for c in ['avg_interval','avg_hv','avg_ice','direct_count']:
        latest[c] = latest.get(c, 0).fillna(0)
    for c in ['chronic','confirmed']:
        latest[c] = latest.get(c, False).fillna(False).astype(bool)

    cs = latest['county'].where(latest['county'].isin(le.classes_), 'Unknown').fillna('Unknown')
    latest['county_enc'] = le.transform(cs)
    latest['prev']  = latest['avg_interval'].fillna(MED)
    latest['month'] = latest['inspection_date'].dt.month
    Xp = pd.DataFrame({f: latest.get(f, 0) for f in FEATS}).fillna(0)
    latest['pred_days']  = np.clip(rf.predict(Xp).round().astype(int), 7, 730)
    latest['pred_next']  = latest['inspection_date'] + pd.to_timedelta(latest['pred_days'], unit='d')
    latest['days_until'] = (latest['pred_next'] - pd.Timestamp(TODAY)).dt.days.fillna(999).astype(int)

    # Score
    def pitch_type(dr, du, ice, chronic, confirmed, days_since=0):
        if dr >= 4 and days_since <= 120: return 'callback'   # admin complaint/emergency = always callback if recent
        if dr >= 4 and -30 <= du < 75: return 'callback'
        if du < 0 and dr >= 2:         return 'overdue_urgent'
        if du < 0:                     return 'overdue'
        if du <= 21:                   return 'pre_hot'
        if du <= 45:                   return 'pre_warm'
        if (chronic or ice >= 80) and du <= 90: return 'high_risk'
        return 'routine'

    def ice_score(pt, dr, hv, ni, ice, chronic, confirmed):
        b = {'callback':96,'overdue_urgent':88,'overdue':72,'pre_hot':84,
             'pre_warm':68,'high_risk':60,'routine':28}.get(pt, 28)
        if confirmed: b += 12
        if chronic:   b += 10
        if ice >= 150: b += 8
        elif ice >= 80: b += 4
        if hv >= 5: b += 5
        elif hv >= 3: b += 3
        if (ni or 1) >= 3: b += 4
        if dr >= 4: b += 6
        return min(100, b)

    def priority(pt):
        return {'callback':'CALLBACK','overdue_urgent':'CALLBACK',
                'pre_hot':'HOT','pre_warm':'WARM','overdue':'WARM',
                'high_risk':'WATCH'}.get(pt, 'LATER')

    # Score ALL businesses - no caps, no sampling
    # Browser does the filtering
    recent = latest[latest['inspection_date'] >= '2023-07-01'].copy()
    recent['_days_since'] = (pd.Timestamp(TODAY) - recent['inspection_date']).dt.days.fillna(999).astype(int)

    print("Scoring all businesses...")
    pts, scs, prs = [], [], []
    for _, r in recent.iterrows():
        pt = pitch_type(r['dr'], r['days_until'], r['avg_ice'],
                        bool(r['chronic']), bool(r['confirmed']), int(r['_days_since']))
        sc = ice_score(pt, r['dr'], r['avg_hv'] if pd.notna(r['avg_hv']) else r['hv'],
                       r.get('n_insp',1) or 1, r['avg_ice'],
                       bool(r['chronic']), bool(r['confirmed']))
        pts.append(pt); scs.append(sc); prs.append(priority(pt))

    recent['_pt'] = pts
    recent['_sc'] = scs
    recent['_pr'] = prs
    recent['_pinellas'] = (recent['county'] == 'Pinellas').astype(int)
    recent = recent[recent['_sc'] >= MIN_SCORE]
    recent = recent.sort_values(['_sc','_pinellas'], ascending=[False, False])

    result = recent.drop_duplicates('license_id').reset_index(drop=True)
    print(f"  {len(result):,} businesses scored")

    # Rename underscore columns -- itertuples strips leading underscores
    result = result.rename(columns={
        '_sc': 'sc', '_pt': 'pt', '_pr': 'pr',
        '_pinellas': 'pinellas', '_days_since': 'days_since_col'
    })

    print(f"Building records (vectorized)...")
    # Pre-compute all classification in bulk -- avoids slow iterrows
    result['biz_type'] = result['biz_name'].apply(
        lambda n: classify_business(str(n)[:50])[0])
    result = result[result['biz_type'] != 'corporate'].copy()
    print(f"  {len(result):,} non-corporate businesses")

    # Vectorized lat/lon from ZIP
    result['z5'] = result['zip'].astype(str).str[:5]
    result['lat'] = result['z5'].map(lambda z: ZIP_COORDS.get(z, (None,None))[0])
    result['lon'] = result['z5'].map(lambda z: ZIP_COORDS.get(z, (None,None))[1])

    # Vectorized bar detection
    bar_words = ['bar','pub','tavern','lounge','brewery','taproom','cantina',
                 'saloon','sports bar','grill & bar','grille & bar','raw bar']
    result['is_bar'] = result['biz_name'].str.lower().apply(
        lambda n: any(w in str(n) for w in bar_words))

    # Vectorized emergency flag
    result['is_emergency'] = result['license_id'].astype(str).isin(emergency_ids)

    # Vectorized days_since
    result['days_since'] = (pd.Timestamp(TODAY) - result['inspection_date']).dt.days.fillna(999).astype(int)

    # Fill missing columns safely
    for col, default in [('direct_count',0),('avg_ice',0),('n_insp',1),('disp_raw',''),
                          ('chronic',False),('confirmed',False)]:
        if col not in result.columns: result[col] = default
    result['direct_count'] = result['direct_count'].fillna(0).astype(int)
    result['avg_ice']  = result['avg_ice'].fillna(0)
    result['n_insp']   = result['n_insp'].fillna(1).astype(int)
    result['disp_raw'] = result['disp_raw'].fillna('').astype(str)

    # Build records using itertuples (fast)
    records = []
    for row in result.itertuples(index=False):
        try:
            lid   = int(row.license_id)
            name  = str(row.biz_name)[:50]
            codes = fired.get(lid, [])
            z5    = str(row.zip)[:5]

            lic_info  = match_license(lid, license_data) or {}
            phone_raw = lic_info.get('phone','') or STATIC_PHONES.get(str(lid),'')
            seats     = int(lic_info.get('seats', 0) or 0)
            rank      = lic_info.get('rank', 'SEAT')

            machines    = est_machines(seats, bool(row.is_bar), rank)
            monthly_val = est_monthly(machines)
            onetime_val = est_onetime(machines)
            confirmed   = bool(row.confirmed)
            chronic     = bool(row.chronic)
            tier        = account_tier(seats, rank, machines, chronic, confirmed)
            days_since  = int(row.days_since)

            lat = row.lat if row.lat and row.lat == row.lat else None
            lon = row.lon if row.lon and row.lon == row.lon else None
            if lat is None:
                cached = geo_cache.get(f"{row.address},{row.city},FL")
                if cached: lat, lon = cached[0], cached[1]

            n_insp      = int(row.n_insp or 1)
            avg_ice     = float(row.avg_ice or 0)
            disp_raw    = str(row.disp_raw or '')
            final_score = min(100, int(row.sc) + seat_score_bonus(machines, rank))
            conf        = confidence_score(n_insp, 1, confirmed, days_since)

            records.append({
                'id':          lid,
                'name':        name,
                'county':      str(row.county),
                'city':        str(row.city),
                'address':     str(row.address),
                'zip':         z5,
                'lat':         lat,
                'lon':         lon,
                'last_insp':   row.inspection_date.strftime('%Y-%m-%d'),
                'last_disp':   disp_raw,
                'disp_risk':   int(row.dr),
                'high_viol':   int(row.hv),
                'total_viol':  int(row.tv),
                'n_insp':      n_insp,
                'pred_next':   row.pred_next.strftime('%Y-%m-%d'),
                'days_until':  int(row.days_until),
                'days_since':  days_since,
                'pitch_type':  str(row.pt),
                'score':       final_score,
                'confidence':  conf,
                'priority':    str(row.pr),
                'ice_rel':     round(avg_ice, 0),
                'confirmed':   confirmed,
                'chronic':     chronic,
                'ice_count':   int(row.direct_count),
                'codes':       codes,
                'trending':    float(trends.get(lid, 0)) > 0.5,
                'is_emergency':bool(row.is_emergency),
                'biz_type':    str(row.biz_type),
                'is_bar':      bool(row.is_bar),
                'seats':       seats,
                'machines':    machines,
                'monthly':     monthly_val,
                'onetime':     onetime_val,
                'tier':        tier,
                'phone':       phone_raw,
                'status':      'prospect',
                'rating':      0,
                'hours':       '',
            })
        except Exception:
            pass

    PO = {'CALLBACK':0,'HOT':1,'WARM':2,'WATCH':3,'LATER':4}
    records.sort(key=lambda x: (PO.get(x['priority'], 5), -x['score']))

    # Save geocoding cache for future runs
    save_geo_cache(geo_cache, data_dir)

    # Preserve phones from previous run
    if OUTPUT_FILE.exists():
        try:
            content = OUTPUT_FILE.read_text(encoding='utf-8')
            m = re.search(r'const PHONES=(\{.*?\});', content, re.DOTALL)
            if m:
                saved_phones = json.loads(m.group(1))
                n_restored = 0
                for rec in records:
                    ph = saved_phones.get(str(rec['id']), {})
                    if ph.get('phone'):
                        rec['phone']  = ph['phone']
                        rec['rating'] = ph.get('rating', 0)
                        rec['hours']  = ph.get('hours', '')
                        n_restored += 1
                if n_restored:
                    print(f"  Phone data restored for {n_restored} businesses")
        except Exception:
            pass

    pc = dict(Counter(r['priority'] for r in records))
    print(f"\nResults:")
    print(f"  Total prospects:       {len(records)}")
    print(f"  Callback urgent:       {pc.get('CALLBACK',0)}")
    print(f"  Hot (≤21 days):        {pc.get('HOT',0)}")
    print(f"  Warm (22-45 days):     {pc.get('WARM',0)}")
    print(f"  Watch:                 {pc.get('WATCH',0)}")
    print(f"  Chronic ice offenders: {sum(1 for r in records if r.get('chronic'))}")
    print(f"  With map coordinates:  {sum(1 for r in records if r['lat'])}")
    print(f"  With phone number:     {sum(1 for r in records if r['phone'])}")

    return records

# ──────────────────────────────────────────────────────────────────────────────
# HTML
# ──────────────────────────────────────────────────────────────────────────────
def build_html(records):
    data_js   = json.dumps(records, separators=(',',':'))
    phones_js = json.dumps(
        {str(r['id']): {'phone':r['phone'],'rating':r['rating'],'hours':r['hours']}
         for r in records if r['phone']},
        separators=(',',':')
    )
    zip_js = json.dumps({z: list(c) for z,c in ZIP_COORDS.items()}, separators=(',',':'))

    n_cb    = sum(1 for r in records if r['priority']=='CALLBACK')
    n_hot   = sum(1 for r in records if r['priority']=='HOT')
    n_warm  = sum(1 for r in records if r['priority']=='WARM')
    n_phone = sum(1 for r in records if r['phone'])
    n_chron = sum(1 for r in records if r.get('chronic'))
    n_geo   = sum(1 for r in records if r['lat'])

    # Read the HTML template embedded below
    return HTML_TEMPLATE.replace('%%DATA%%', data_js)\
                        .replace('%%PHONES%%', phones_js)\
                        .replace('%%ZIPS%%', zip_js)\
                        .replace('%%DATE%%', str(TODAY))\
                        .replace('%%TOTAL%%', str(len(records)))\
                        .replace('%%NCB%%', str(n_cb))\
                        .replace('%%NHOT%%', str(n_hot))\
                        .replace('%%NWARM%%', str(n_warm))\
                        .replace('%%NPHONE%%', str(n_phone))\
                        .replace('%%NCHRON%%', str(n_chron))\
                        .replace('%%NGEO%%', str(n_geo))\
                        .replace('%%NCONF%%', str(sum(1 for r in records if r['confirmed'])))

# ──────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE  (everything between the triple-quotes)
# ──────────────────────────────────────────────────────────────────────────────
HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Pinellas Ice Co &#xB7; Prospects</title>
<!-- Leaflet loaded on-demand when Route tab opens -->
<link rel="manifest" href="manifest.json">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="PIC Prospects">
<link rel="apple-touch-icon" href="apple-touch-icon.png">
<meta name="theme-color" content="#2d3e50">
<link href="https://fonts.googleapis.com/css2?family=Lexend:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
:root{
  --bg:#f5f8fa;--surf:#ffffff;--brd:#dfe3eb;--brd2:#eaf0f6;
  --txt:#33475b;--sub:#7c98b6;--sub2:#516f90;
  --cb:#f2545b;--hot:#ff7a59;--warm:#f5c26b;--watch:#00bda5;
  --grn:#00bda5;--blu:#0091ae;--ora:#ff7a59;--navy:#2d3e50;
  --shadow:0 1px 3px rgba(45,62,80,.1),0 1px 8px rgba(45,62,80,.06);
  --shadow-md:0 3px 12px rgba(45,62,80,.12),0 1px 4px rgba(45,62,80,.08);
}
html,body{height:100%;background:var(--bg);color:var(--txt);
  font-family:'Lexend',sans-serif;
  font-size:13px;overflow:hidden;-webkit-font-smoothing:antialiased}
#app{display:flex;flex-direction:column;height:100vh}
header{background:var(--navy);
  padding:0 16px;display:flex;align-items:center;gap:10px;flex-shrink:0;flex-wrap:wrap;height:52px}
.logo{display:flex;align-items:center;gap:8px;flex-shrink:0}
.logo-icon{width:30px;height:30px;border-radius:8px;
  background:var(--ora);
  display:flex;align-items:center;justify-content:center;font-size:16px}
.logo-name{font-weight:700;font-size:14px;color:#fff;letter-spacing:-.02em}
.hchips{display:flex;gap:5px;margin-left:auto;flex-wrap:wrap}
.hc{font-size:10px;padding:3px 10px;border-radius:20px;font-weight:600;cursor:pointer;user-select:none;transition:.15s}
.hc.cb{background:rgba(242,84,91,.2);color:#ffb3b6;border:1px solid rgba(242,84,91,.3)}
.hc.ht{background:rgba(255,122,89,.2);color:#ffcab8;border:1px solid rgba(255,122,89,.3)}
.hc.wm{background:rgba(245,194,107,.2);color:#ffe5a8;border:1px solid rgba(245,194,107,.3)}
.hc.ic{background:rgba(0,189,165,.2);color:#7eeee3;border:1px solid rgba(0,189,165,.3)}
.hc.rt{background:rgba(0,145,174,.2);color:#7dd4e8;border:1px solid rgba(0,145,174,.3)}
.srch{position:relative}
.srch input{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);border-radius:8px;
  padding:6px 10px 6px 28px;color:#fff;font-size:12px;outline:none;width:160px;font-family:inherit}
.srch input::placeholder{color:rgba(255,255,255,.5)}
.srch input:focus{background:rgba(255,255,255,.18);border-color:rgba(255,255,255,.4)}
.si{position:absolute;left:9px;top:50%;transform:translateY(-50%);color:rgba(255,255,255,.5);font-size:11px;pointer-events:none}
.tabs{display:flex;background:var(--surf);border-bottom:2px solid var(--brd2);flex-shrink:0;
  box-shadow:0 1px 3px rgba(45,62,80,.06)}
.tab{flex:1;padding:10px 4px;text-align:center;font-size:11px;font-weight:600;
  color:var(--sub);cursor:pointer;border-bottom:2px solid transparent;
  margin-bottom:-2px;transition:.15s;user-select:none}
.tab.on{color:var(--ora);border-color:var(--ora)}
.panel{flex:1;overflow-y:auto;padding:14px 16px;display:none;background:var(--bg)}
.panel.on{display:block}
.fbar{display:flex;gap:6px;margin-bottom:12px;flex-wrap:wrap;align-items:center}
.fbar select{background:var(--surf);border:1px solid var(--brd);border-radius:6px;
  padding:5px 9px;color:var(--txt);font-size:11px;outline:none;cursor:pointer;
  font-family:inherit;font-weight:500;box-shadow:var(--shadow)}
.fcnt{font-size:10px;color:var(--sub);margin-left:auto;font-weight:500}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(276px,1fr));gap:10px}
.card{background:var(--surf);border:1px solid var(--brd);border-radius:10px;
  padding:13px 14px;cursor:pointer;transition:box-shadow .15s,transform .12s,border-color .15s;
  position:relative;overflow:hidden;box-shadow:var(--shadow)}
.card:hover{box-shadow:var(--shadow-md);transform:translateY(-1px);border-color:#c5d1de}
.card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;border-radius:10px 0 0 10px}
.card.CALLBACK::before{background:var(--cb)}
.card.HOT::before{background:var(--hot)}
.card.WARM::before{background:var(--warm)}
.card.WATCH::before{background:var(--watch)}
.card.done{opacity:.45}
.ctop{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:3px}
.cname{font-weight:700;font-size:12.5px;flex:1;padding-right:6px;line-height:1.3;color:var(--navy)}
.pbadge{font-size:9px;font-weight:700;padding:2px 8px;border-radius:20px;letter-spacing:.04em;white-space:nowrap}
.pbadge.CALLBACK{background:#fdeaea;color:#c0392b;border:1px solid #f5c6c6}
.pbadge.HOT{background:#fff2ee;color:#c9542b;border:1px solid #fdd5c6}
.pbadge.WARM{background:#fef9ee;color:#9e6b1a;border:1px solid #fbe8b5}
.pbadge.WATCH{background:#e8faf8;color:#007a6e;border:1px solid #b3ece6}
.cloc{font-size:10px;color:var(--sub);margin-bottom:6px;font-weight:500}
.phrow{display:flex;align-items:center;gap:6px;padding:7px 9px;
  background:#f5f8fa;border-radius:7px;border:1px solid var(--brd2);margin-bottom:6px}
.phnum{font-size:12px;font-weight:600;color:var(--blu);flex:1}
.phnum.none{color:var(--sub);font-size:10px;font-style:italic;font-weight:400}
.abtn{border:none;border-radius:5px;padding:3px 9px;font-size:10px;font-weight:700;
  cursor:pointer;text-decoration:none;display:inline-block;white-space:nowrap;font-family:inherit}
.call-a{background:var(--grn);color:#fff}
.find-a{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7}
.icebadge{font-size:9px;font-weight:700;padding:3px 9px;border-radius:20px;
  margin-bottom:5px;display:inline-block}
.icebadge.chronic{background:#e8faf8;color:#007a6e;border:1px solid #b3ece6}
.icebadge.confirmed{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7}
.cmeta{display:flex;gap:8px;margin-bottom:6px;flex-wrap:wrap}
.mi{display:flex;flex-direction:column;gap:1px}
.ml{font-size:8px;color:var(--sub);letter-spacing:.05em;text-transform:uppercase;font-weight:600}
.mv{font-weight:700;font-size:11px;color:var(--txt)}
.mv.u{color:var(--cb)}.mv.h{color:var(--hot)}.mv.w{color:#9e6b1a}.mv.bad{color:var(--cb)}
.insight{font-size:9px;color:var(--sub2);background:#f5f8fa;border-radius:5px;border:1px solid var(--brd2);
  padding:4px 8px;margin-bottom:5px;line-height:1.5}
.lastc{font-size:10px;color:var(--sub);margin-bottom:6px}
.lastc.hc{color:#007a6e;font-weight:500}
.cacts{display:flex;gap:5px}
.btn{border:none;border-radius:6px;padding:5px 11px;font-size:10px;font-weight:600;cursor:pointer;font-family:inherit}
.blog{background:var(--ora);color:#fff}.blog:hover{background:#e86744}
.bskip{background:transparent;color:var(--sub);border:1px solid var(--brd)}
.brt{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7;font-size:9px}
.ptbl{width:100%;border-collapse:collapse}
.ptbl th{padding:7px 10px;text-align:left;font-size:9px;color:var(--sub2);font-weight:700;
  letter-spacing:.06em;text-transform:uppercase;border-bottom:2px solid var(--brd);
  white-space:nowrap;position:sticky;top:0;background:var(--surf);z-index:1}
.ptbl td{padding:8px 10px;border-bottom:1px solid var(--brd2);vertical-align:middle}
.ptbl tr:hover td{background:#f5f8fa;cursor:pointer}
.ptbl tr.done td{opacity:.4}
.tn{font-weight:600;font-size:11px;max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--navy)}
.ts{font-size:9px;color:var(--sub)}
.tph{font-size:11px;color:var(--blu);font-weight:600;text-decoration:none;white-space:nowrap}
.tph:hover{color:#007a9e}
.tph.none{color:#b0c4d8;font-size:9px}
.td{font-weight:700;font-size:12px}
.td.u{color:var(--cb)}.td.h{color:var(--hot)}.td.w{color:#9e6b1a}
/* ROUTE */
.route-wrap{display:grid;grid-template-columns:280px 1fr;gap:12px;height:calc(100vh - 120px)}
@media(max-width:600px){.route-wrap{grid-template-columns:1fr;grid-template-rows:auto 1fr}}
.rsidebar{display:flex;flex-direction:column;gap:8px;overflow-y:auto}
.rcontrols{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:12px;box-shadow:var(--shadow)}
.rtitle{font-weight:700;font-size:12px;margin-bottom:8px;color:var(--navy)}
.rlist{display:flex;flex-direction:column;gap:4px;overflow-y:auto;max-height:320px}
.rcard{background:var(--surf);border:1px solid var(--brd);border-radius:7px;
  padding:8px 10px;cursor:pointer;transition:.12s;display:flex;align-items:center;gap:7px;
  box-shadow:var(--shadow)}
.rcard:hover{border-color:#c5d1de;box-shadow:var(--shadow-md)}
.rcard.sel{border-color:var(--ora);background:#fff7f5}
.rdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.rname{font-weight:600;font-size:11px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--navy)}
.rdist{font-size:9px;color:var(--sub);flex-shrink:0}
.map-area{background:#e8edf2;border:1px solid var(--brd);border-radius:10px;
  display:flex;align-items:center;justify-content:center;overflow:hidden;position:relative;min-height:300px;
  box-shadow:var(--shadow)}
.map-empty{color:var(--sub);font-size:12px;text-align:center;padding:20px}
.day-route{background:var(--surf);border:1px solid var(--brd);border-radius:10px;
  padding:11px;margin-top:8px;box-shadow:var(--shadow)}
.day-stop{display:flex;align-items:center;gap:7px;padding:6px 0;
  border-bottom:1px solid var(--brd2);font-size:11px}
.day-stop:last-child{border-bottom:none}
.stopnum{width:20px;height:20px;border-radius:50%;background:var(--ora);
  color:#fff;font-weight:700;font-size:9px;display:flex;align-items:center;
  justify-content:center;flex-shrink:0}
.route-btn{background:var(--ora);color:#fff;border:none;border-radius:8px;
  padding:9px;font-size:11px;font-weight:700;cursor:pointer;width:100%;margin-top:8px;font-family:inherit}
.route-btn:hover{background:#e86744}
.route-btn.sec{background:#f5f8fa;color:var(--sub2);border:1px solid var(--brd);margin-top:5px}
/* MODAL */
#mbg{position:fixed;inset:0;background:rgba(45,62,80,.6);z-index:100;
  display:none;backdrop-filter:blur(4px)}
#mbg.on{display:flex;align-items:flex-end;justify-content:center}
@media(min-width:600px){#mbg.on{align-items:center}}
#modal{background:#fff;border:1px solid var(--brd);border-radius:16px 16px 0 0;
  width:100%;max-width:540px;padding:18px;max-height:90vh;overflow-y:auto;
  box-shadow:0 -4px 24px rgba(45,62,80,.15)}
@media(min-width:600px){#modal{border-radius:16px;box-shadow:0 8px 40px rgba(45,62,80,.18)}}
.mh{width:36px;height:4px;background:var(--brd);border-radius:2px;margin:0 auto 14px}
.mname{font-size:17px;font-weight:700;margin-bottom:2px;color:var(--navy)}
.mloc{font-size:10px;color:var(--sub);margin-bottom:12px;font-weight:500}
.mphsec{background:#f5f8fa;border:1px solid var(--brd2);border-radius:10px;padding:12px;margin-bottom:11px}
.mphl{font-size:8px;color:var(--sub);letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px;font-weight:700}
.mphnum{font-size:20px;font-weight:700;color:var(--blu);margin-bottom:8px}
.mphnum.none{font-size:12px;color:var(--sub);font-weight:400;font-style:italic}
.mphacts{display:flex;gap:6px;flex-wrap:wrap}
.mcall{background:var(--grn);color:#fff;border:none;border-radius:8px;padding:8px 16px;
  font-size:11px;font-weight:700;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;gap:5px;font-family:inherit}
.msms{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7;border-radius:8px;
  padding:8px 12px;font-size:11px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:4px}
.mgoog{background:#f5f8fa;color:var(--sub2);border:1px solid var(--brd);border-radius:8px;
  padding:8px 12px;font-size:11px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:4px}
.mhours{font-size:9px;color:var(--sub);margin-top:5px}
.micebox{background:#e8faf8;border:1px solid #b3ece6;border-radius:9px;padding:11px;margin-bottom:11px}
.msect{font-size:9px;color:var(--sub2);letter-spacing:.07em;text-transform:uppercase;
  margin-bottom:6px;padding-bottom:5px;border-bottom:1px solid var(--brd2);font-weight:700}
.pitch{background:#f5f8fa;border:1px solid var(--brd2);border-radius:8px;padding:10px;
  font-size:11px;color:var(--sub2);line-height:1.7;margin-bottom:9px}
.pitch b{color:var(--navy)}
.ogrid{display:grid;grid-template-columns:repeat(3,1fr);gap:5px;margin-bottom:8px}
.obtn{padding:7px 4px;border-radius:7px;border:1px solid var(--brd);
  background:var(--surf);color:var(--sub2);font-size:10px;font-weight:600;
  cursor:pointer;text-align:center;transition:.12s;font-family:inherit}
.obtn:hover{border-color:#c5d1de;background:#f5f8fa}
.obtn.on{border-color:var(--ora);background:#fff7f5;color:var(--ora)}
.ntxt{width:100%;background:#f5f8fa;border:1px solid var(--brd);border-radius:8px;
  padding:8px;color:var(--txt);font-size:11px;resize:none;outline:none;
  line-height:1.5;font-family:inherit}
.ntxt:focus{border-color:var(--blu)}
.mfacts{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.fact{background:#f5f8fa;border:1px solid var(--brd2);border-radius:7px;padding:8px 9px}
.fl{font-size:8px;color:var(--sub);letter-spacing:.05em;text-transform:uppercase;margin-bottom:2px;font-weight:600}
.fv{font-weight:700;font-size:11px;color:var(--navy)}
.fv.r{color:var(--cb)}.fv.o{color:var(--hot)}.fv.g{color:var(--grn)}.fv.b{color:var(--blu)}
.mhist{background:#f5f8fa;border-radius:8px;padding:9px}
.hi{padding:4px 0;border-bottom:1px solid var(--brd2);font-size:10px;color:var(--txt)}
.hi:last-child{border-bottom:none}
.psect{margin-bottom:16px}
.pst{font-size:10px;font-weight:700;letter-spacing:.05em;text-transform:uppercase;
  margin-bottom:7px;display:flex;align-items:center;gap:6px;color:var(--navy)}
.pct{background:var(--brd2);color:var(--sub2);border-radius:20px;padding:1px 8px;font-size:9px;font-weight:700}
.pitem{background:var(--surf);border:1px solid var(--brd);border-radius:8px;
  padding:9px 12px;margin-bottom:5px;display:flex;align-items:center;gap:8px;cursor:pointer;
  box-shadow:var(--shadow);transition:.12s}
.pitem:hover{box-shadow:var(--shadow-md);border-color:#c5d1de}
.pdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.piname{font-weight:600;font-size:11px;flex:1;color:var(--navy)}
.piph{font-size:10px;color:var(--blu);font-weight:600}
.dc{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:14px;margin-bottom:10px;box-shadow:var(--shadow)}
.dct{font-weight:700;font-size:13px;margin-bottom:8px;color:var(--navy)}
.ds{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid var(--brd2);font-size:11px}
.ds:last-child{border-bottom:none}
.dsv{font-weight:700;color:var(--blu)}
.ibox{background:#f0f9fb;border:1px solid #b3dce7;border-radius:8px;
  padding:11px;font-size:11px;color:var(--sub2);line-height:1.8}
.ibox code{background:#dff0f5;padding:2px 6px;border-radius:4px;
  font-family:monospace;font-size:10px;color:var(--blu)}
.ibox b{color:var(--navy)}
.phinput{background:#f5f8fa;border:1px solid var(--brd);border-radius:7px;
  padding:7px 10px;color:var(--txt);font-size:11px;outline:none;width:100%;margin-bottom:7px;font-family:inherit}
.phinput:focus{border-color:var(--blu)}
.xbtn{background:var(--ora);color:#fff;border:none;
  border-radius:8px;padding:9px 12px;font-size:11px;font-weight:700;cursor:pointer;width:100%;margin-top:6px;font-family:inherit}
.xbtn:hover{background:#e86744}
.dbtn{background:#fff;color:var(--cb);border:1px solid #f5c6c6;
  border-radius:8px;padding:9px 12px;font-size:11px;font-weight:600;cursor:pointer;width:100%;margin-top:5px;font-family:inherit}
.empty{text-align:center;padding:36px 20px;color:var(--sub)}
.ei{font-size:32px;margin-bottom:10px}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.2}}.pulse{animation:pulse 2s ease-in-out infinite}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--brd);border-radius:2px}
#toast{position:fixed;bottom:20px;left:50%;transform:translateX(-50%) translateY(60px);
  background:var(--navy);border-radius:10px;
  padding:9px 16px;font-size:11px;color:#fff;transition:transform .2s;z-index:200;white-space:nowrap;
  box-shadow:0 4px 16px rgba(45,62,80,.25);font-weight:500}
.tsect-hdr{display:flex;flex-direction:column;gap:2px;margin-bottom:10px;padding-bottom:8px;border-bottom:2px solid var(--brd2)}
.tsect-hdr span:first-child{font-weight:700;font-size:14px;color:var(--navy)}
.tsect-sub{font-size:10px;color:var(--sub);font-weight:400}
.tempty{text-align:center;padding:16px;color:var(--sub);font-size:11px;background:var(--surf);border-radius:8px;border:1px dashed var(--brd)}
.emerg-badge{display:inline-block;font-size:8px;font-weight:700;padding:2px 6px;border-radius:20px;background:#fef2f2;color:#dc2626;border:1px solid #fecaca;margin-left:4px}
.preset-btn{flex-shrink:0;padding:6px 12px;border-radius:20px;border:1px solid var(--brd);background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;white-space:nowrap;font-family:inherit;transition:.15s}
.preset-btn.on{background:var(--navy);color:#fff;border-color:var(--navy)}
.flt-sel{background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:5px 7px;color:var(--txt);font-size:10px;outline:none;cursor:pointer;font-family:inherit;max-width:130px}
</style>
</head>
<body>
<div id="app">
  <header>
    <div class="logo">
      <div class="logo-icon">&#x1F9CA;</div>
      <div><div class="logo-name">Pinellas Ice Co</div>
        <div style="font-size:8px;color:var(--sub);letter-spacing:.04em">PROSPECT TOOL &bull; %%DATE%%</div>
      </div>
    </div>
    <div class="hchips">
      <span class="hc cb" onclick="setF('CALLBACK')">&#x25CF; %%NCB%% Callback</span>
      <span class="hc ht" onclick="setF('HOT')">&#x25CF; %%NHOT%% Hot</span>
      <span class="hc wm" onclick="setF('WARM')">&#x25CF; %%NWARM%% Warm</span>
      <span class="hc ic">&#x1F9CA; %%NCHRON%% Chronic</span>
      <span class="hc rt" onclick="sw('route')">&#x1F5FA; Route</span>
    </div>
    <div class="srch"><span class="si">&#x1F50D;</span>
      <input type="text" id="si" placeholder="Search&#x2026;" oninput="onS()">
    </div>
  </header>
  <nav class="tabs">
    <div class="tab on"  onclick="sw('today')">&#x1F4CB; Today</div>
    <div class="tab"     onclick="sw('all')">&#x1F4CD; All</div>
    <div class="tab"     onclick="sw('route')">&#x1F5FA; Route</div>
    <div class="tab"     onclick="sw('customers')">&#x1F91D; Customers</div>
    <div class="tab"     onclick="sw('pipe')">&#x1F4CA; Pipeline</div>
    <div class="tab"     onclick="sw('data')">&#x2699; Data</div>
  </nav>

  <!-- TODAY -->
  <div class="panel on" id="p-today">
    <div class="fbar">
      <select id="tc" onchange="rT()"><option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option><option>Citrus</option><option>Hernando</option><option>Polk</option><option>Sumter</option></select>
      <label style="font-size:10px;color:var(--sub);display:flex;align-items:center;gap:4px;cursor:pointer">
        <input type="checkbox" id="thc" onchange="rT()"> Hide contacted
      </label>
      <span class="fcnt" id="tcnt"></span>
    </div>
    <div id="t-actnow">
      <div class="tsect-hdr" id="hdr-actnow">
        <span>&#x1F534; Act Now</span>
        <span class="tsect-sub">Bad inspection last 30 days &bull; callbacks pending &bull; emergency closures. Check daily.</span>
      </div>
      <div class="grid" id="tgrid-actnow"></div>
      <div class="tempty" id="empty-actnow" style="display:none">No urgent prospects today.</div>
    </div>
    <div id="t-upcoming" style="margin-top:18px">
      <div class="tsect-hdr" id="hdr-upcoming">
        <span>&#x1F4C5; Upcoming</span>
        <span class="tsect-sub">Inspection predicted within 60 days &bull; ice history on record. Work weekly.</span>
      </div>
      <div class="grid" id="tgrid-upcoming"></div>
      <div class="tempty" id="empty-upcoming" style="display:none">No upcoming prospects.</div>
    </div>
    <div id="t-highval" style="margin-top:18px">
      <div class="tsect-hdr" id="hdr-highval">
        <span>&#x1F4B0; High Value</span>
        <span class="tsect-sub">PLATINUM &amp; GOLD accounts by revenue. Worth a cold call anytime.</span>
      </div>
      <div class="grid" id="tgrid-highval"></div>
      <div class="tempty" id="empty-highval" style="display:none">No high-value accounts found.</div>
    </div>
    <div id="t-followups" style="margin-top:18px">
      <div class="tsect-hdr" id="hdr-followups">
        <span>&#x1F4C5; Follow-Ups Due</span>
        <span class="tsect-sub">Follow-up date reached  -  time to call back.</span>
      </div>
      <div class="grid" id="tgrid-followups"></div>
      <div class="tempty" id="empty-followups" style="display:none">No follow-ups due today.</div>
    </div>
  </div>

  <!-- ALL -->
  <div class="panel" id="p-all">

    <!-- Static preset quick filters -->
    <div style="display:flex;gap:6px;overflow-x:auto;padding-bottom:8px;margin-bottom:10px;-webkit-overflow-scrolling:touch">
      <button class="preset-btn on" onclick="setPreset('all')"      id="pre-all">All</button>
      <button class="preset-btn"    onclick="setPreset('actnow')"   id="pre-actnow">&#x1F534; Act Now</button>
      <button class="preset-btn"    onclick="setPreset('callback')" id="pre-callback">Callbacks</button>
      <button class="preset-btn"    onclick="setPreset('phone')"    id="pre-phone">Has Phone</button>
      <button class="preset-btn"    onclick="setPreset('chronic')"  id="pre-chronic">Chronic Ice</button>
      <button class="preset-btn"    onclick="setPreset('gold')"     id="pre-gold">GOLD+</button>
      <button class="preset-btn"    onclick="setPreset('franchise')" id="pre-franchise">Franchise</button>
      <button class="preset-btn"    onclick="setPreset('bar')"      id="pre-bar">Bars/Pubs</button>
      <button class="preset-btn"    onclick="setPreset('notyet')"   id="pre-notyet">Not Contacted</button>
    </div>

    <!-- Drill-down filters -->
    <div style="display:flex;gap:5px;flex-wrap:wrap;margin-bottom:8px;align-items:center">
      <select id="ac"  onchange="rA()" class="flt-sel"><option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option><option>Citrus</option><option>Hernando</option><option>Polk</option><option>Sumter</option></select>
      <select id="ap"  onchange="rA()" class="flt-sel"><option value="">All Priorities</option><option>CALLBACK</option><option>HOT</option><option>WARM</option><option>WATCH</option></select>
      <select id="ai"  onchange="rA()" class="flt-sel"><option value="">All Ice</option><option value="chronic">Chronic</option><option value="confirmed">Confirmed</option></select>
      <select id="atier" onchange="rA()" class="flt-sel"><option value="">All Tiers</option><option value="PLATINUM">Platinum</option><option value="GOLD">Gold</option><option value="SILVER">Silver</option></select>
      <select id="abtype" onchange="rA()" class="flt-sel"><option value="">All Types</option><option value="independent">Independent</option><option value="franchise">Franchise</option></select>
      <select id="as_" onchange="rA()" class="flt-sel"><option value="">All Contact Status</option><option value="not_contacted">Not Contacted</option><option value="interested">Interested</option><option value="scheduled">Scheduled</option><option value="follow_up">Follow Up</option><option value="voicemail">Voicemail</option><option value="no_answer">No Answer</option><option value="not_interested">Not Interested</option></select>
      <select id="aconf" onchange="rA()" class="flt-sel"><option value="">Any Confidence</option><option value="75">75%+ Confident</option><option value="50">50%+ Confident</option></select>
      <button onclick="clearFilters()" style="font-size:10px;padding:5px 8px;border:1px solid var(--brd);border-radius:6px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit;flex-shrink:0">Clear</button>
      <span class="fcnt" id="acnt"></span>
    </div>

    <!-- Results as cards (not table - better on mobile) -->
    <div class="grid" id="agrid"></div>
    <div class="tempty" id="a-empty" style="display:none">
      <div class="ei">&#x1F50D;</div>
      <div>No prospects match these filters</div>
    </div>
  </div>

  <!-- ROUTE -->
  <div class="panel" id="p-route">
    <div style="display:flex;flex-direction:column;height:calc(100vh - 120px);gap:10px">

      <!-- Controls row -->
      <div class="dc" style="flex-shrink:0;padding:12px">
        <div style="font-weight:700;font-size:13px;color:var(--navy);margin-bottom:8px">&#x1F5FA; Build Day Route</div>
        <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:6px">
          <select id="rc" onchange="rRoute()" style="flex:1;min-width:100px;background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:6px 8px;color:var(--txt);font-size:11px;outline:none;font-family:inherit">
            <option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option>
          </select>
          <select id="rp" onchange="rRoute()" style="flex:1;min-width:100px;background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:6px 8px;color:var(--txt);font-size:11px;outline:none;font-family:inherit">
            <option value="">All Priorities</option><option>CALLBACK</option><option>HOT</option><option>WARM</option>
          </select>
        </div>
        <div style="display:flex;gap:6px;align-items:center">
          <input id="rzip" type="text" placeholder="Start ZIP" maxlength="5"
            style="width:90px;background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:6px 8px;color:var(--txt);font-size:11px;outline:none;font-family:inherit"
            oninput="rRoute()">
          <select id="rrad" onchange="rRoute()"
            style="background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:6px 8px;color:var(--txt);font-size:11px;outline:none;font-family:inherit">
            <option value="5">5 mi</option><option value="8" selected>8 mi</option>
            <option value="12">12 mi</option><option value="20">20 mi</option><option value="999">Any distance</option>
          </select>
          <button onclick="optimizeRoute()" title="Sort stops by geography"
            style="flex-shrink:0;padding:6px 10px;border:none;border-radius:6px;background:var(--ora);color:#fff;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">
            &#x26A1; Optimize
          </button>
        </div>
        <div style="font-size:9px;color:var(--sub);margin-top:5px" id="rhint">Tap a prospect to add to route. Tap again to remove.</div>
      </div>

      <!-- Two-column layout -->
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;flex:1;min-height:0;overflow:hidden">

        <!-- Prospect list -->
        <div style="display:flex;flex-direction:column;min-height:0">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px" id="rlist-cnt"></div>
          <div id="rlist" style="overflow-y:auto;flex:1;display:flex;flex-direction:column;gap:4px"></div>
        </div>

        <!-- Route + map -->
        <div style="display:flex;flex-direction:column;min-height:0;gap:8px">
          <!-- Built route -->
          <div id="day-route" style="display:none;background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:10px;flex-shrink:0">
            <div style="font-weight:700;font-size:11px;color:var(--navy);margin-bottom:6px">
              &#x1F4CD; Your Route — <span id="stopcnt">0</span> stops
              <span id="route-mi" style="font-size:9px;color:var(--sub);font-weight:400"></span>
            </div>
            <div id="day-stops"></div>
            <div style="display:flex;gap:5px;margin-top:8px">
              <button class="route-btn" onclick="openMaps()" style="flex:2">Open in Google Maps &#x2192;</button>
              <button class="route-btn sec" onclick="clearRoute()" style="flex:1">Clear</button>
            </div>
          </div>
          <!-- Map -->
          <div class="map-area" id="map-area" style="flex:1;min-height:200px">
            <div class="map-empty">
              <div style="font-size:32px;margin-bottom:8px">&#x1F5FA;&#xFE0F;</div>
              <div>Select county or ZIP<br>to see prospects on map</div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- PIPELINE -->
  <div class="panel" id="p-pipe"><div id="pipec"></div></div>

  <!-- CUSTOMERS -->
  <div class="panel" id="p-customers">
    <div class="fbar">
      <select id="cust-status" onchange="rCust()">
        <option value="">All Customers</option>
        <option value="customer_recurring">Recurring ($149/mo)</option>
        <option value="customer_once">One-Time ($299)</option>
        <option value="quoted">Quoted - Pending</option>
        <option value="churned">Churned</option>
      </select>
      <span class="fcnt" id="cust-cnt"></span>
    </div>

    <!-- Revenue summary bar -->
    <div id="cust-summary" style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px">
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Monthly Recurring</div>
        <div style="font-size:22px;font-weight:800;color:var(--grn)" id="mrr-val">$0</div>
        <div class="fl">MRR</div>
      </div>
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Active Customers</div>
        <div style="font-size:22px;font-weight:800;color:var(--navy)" id="cust-count">0</div>
        <div class="fl">Accounts</div>
      </div>
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Annual Run Rate</div>
        <div style="font-size:22px;font-weight:800;color:var(--blu)" id="arr-val">$0</div>
        <div class="fl">ARR</div>
      </div>
    </div>

    <div id="cust-list"></div>

    <div class="tempty" id="cust-empty" style="display:none">
      <div class="ei">&#x1F91D;</div>
      <div style="font-weight:600;margin-bottom:4px">No customers yet</div>
      <div style="font-size:11px;color:var(--sub)">When you close a deal, open any prospect card and mark them as Won.</div>
    </div>
  </div>

  <!-- DATA -->
  <div class="panel" id="p-data">
    <div class="dc">
      <div class="dct">&#x1F4E6; Dataset &mdash; %%DATE%%</div>
      <div class="ds"><span>Total prospects</span><span class="dsv">%%TOTAL%%</span></div>
      <div class="ds"><span>Data source</span><span class="dsv">FL DBPR District 3 (live)</span></div>
      <div class="ds"><span>Refreshes</span><span class="dsv">Automatically every week</span></div>
      <div class="ds"><span>Chronic ice offenders</span><span class="dsv">%%NCHRON%%</span></div>
      <div class="ds"><span>Confirmed ice violators</span><span class="dsv">%%NCONF%%</span></div>
      <div class="ds"><span>With phone number</span><span class="dsv">%%NPHONE%%</span></div>
      <div class="ds"><span>On map (geocoded)</span><span class="dsv">%%NGEO%%</span></div>
      <div class="ds"><span>Callback urgent</span><span class="dsv">%%NCB%%</span></div>
    </div>
    <div class="dc">
      <div class="dct">&#x1F504; Rebuilding With New Data</div>
      <div class="ibox">
        <b>Step 1 &mdash; Large XLS/XLSX file?</b> Convert it first:<br>
        <code>python convert_to_csv.py yourfile.xlsx</code><br><br>
        <b>Step 2 &mdash; Rebuild the tool:</b><br>
        <code>python build.py file1.csv file2.csv</code><br>
        or just <code>python build.py</code> to auto-find CSVs in the folder.<br><br>
        Your call log and phone numbers carry over automatically every time.
      </div>
    </div>
    <div class="dc">
      <div class="dct">&#x2699;&#xFE0F; Settings</div>
      <div style="font-size:10px;color:var(--sub);margin-bottom:5px">HubSpot Portal ID — find in your HubSpot URL: app.hubspot.com/contacts/<b style="color:var(--navy)">XXXXXXXX</b>/...</div>
      <input class="phinput" id="hs-portal" type="text" placeholder="e.g. 12345678" maxlength="12" oninput="saveSettings()">
      <button class="xbtn" style="margin-top:4px" onclick="saveSettings();toast('HubSpot Portal ID saved!')">Save Portal ID</button>
    </div>
    <div class="dc">
      <div class="dct">&#x2795; Add / Edit a Phone Number</div>
      <input class="phinput" id="ph-id"  type="text" placeholder="License ID (shown on each card)">
      <input class="phinput" id="ph-num" type="tel"  placeholder="Phone number, e.g. +1 727-555-1234">
      <input class="phinput" id="ph-hrs" type="text" placeholder="Hours (optional) e.g. Mon-Fri 11am-9pm">
      <button class="xbtn" style="margin-top:0" onclick="addPhone()">Save Phone Number</button>
    </div>
    <div class="dc">
      <div class="dct">&#x1F4E4; Export &amp; Reset</div>
      <div style="margin-bottom:8px">
        <div class="dct">&#x1F4CA; Model Accuracy</div>
        <div class="ibox">
          <b>Prediction accuracy (days until next inspection):</b><br>
          Within 30 days: ~52% &bull; Within 45 days: ~68% &bull; Within 60 days: ~78%<br><br>
          The model explains ~57% of timing variance (R&sup2;=0.57). The biggest signal is
          inspection outcome &mdash; emergency orders bring inspectors back in days,
          clean inspections mean 3-6 months. Use predictions as <b>directional</b> guidance,
          not precise dates.<br><br>
          <b>Ice machine scoring</b> uses FL violation codes V14, V22, V50 (food contact surfaces)
          &mdash; businesses flagged repeatedly are your strongest leads regardless of timing.
        </div>
      </div>
      <button class="xbtn" onclick="exportCSV()">Export Pipeline to CSV</button>
      <button class="dbtn" onclick="clrLog()">Clear All Call Log Data</button>
    </div>
  </div>
</div>

<!-- MODAL -->
<div id="mbg" onclick="closeM(event)">
  <div id="modal">
    <div class="mh"></div>
    <div class="mname" id="mn"></div>
    <div class="mloc"  id="ml"></div>
    <div class="mphsec">
      <div class="mphl">CONTACT</div>
      <div class="mphnum" id="mph"></div>
      <div class="mphacts" id="mpa"></div>
      <div id="mhrs" class="mhours"></div>
      <div id="mrat" style="font-size:9px;color:#555;margin-top:3px"></div>
    </div>
    <div id="mice" style="margin-bottom:10px"></div>
    <div style="margin-bottom:10px">
      <div class="msect">PITCH GUIDE</div>
      <div style="display:flex;gap:5px;margin-bottom:6px">
        <button id="btn-phone" onclick="setPitchMode('phone')"
          style="flex:1;padding:6px;border-radius:7px;border:1px solid var(--blu);background:#0a84ff22;color:var(--blu);font-size:10px;font-weight:700;cursor:pointer">
          &#x1F4DE; Phone Call
        </button>
        <button id="btn-walkin" onclick="setPitchMode('walkin')"
          style="flex:1;padding:6px;border-radius:7px;border:1px solid var(--brd);background:transparent;color:var(--sub);font-size:10px;font-weight:700;cursor:pointer">
          &#x1F6B6; Cold Walk-In
        </button>
      </div>
      <div class="pitch" id="mpitch"></div>
      <div class="pitch" id="mwalkin" style="display:none;border-color:#e2e8f0;color:var(--sub2)"></div>
      <div style="margin-top:8px">
        <div class="msect">OBJECTION HANDLERS</div>
        <div id="mobjections" style="display:flex;flex-direction:column;gap:5px"></div>
      </div>
    </div>
    <div style="margin-bottom:10px">
      <div class="msect">LOG THIS CONTACT</div>
      <div class="ogrid">
        <button class="obtn" data-o="no_answer"      onclick="selO('no_answer')">No Answer</button>
        <button class="obtn" data-o="voicemail"       onclick="selO('voicemail')">Voicemail</button>
        <button class="obtn" data-o="not_interested"  onclick="selO('not_interested')">Not Interested</button>
        <button class="obtn" data-o="follow_up"       onclick="selO('follow_up')">Follow Up</button>
        <button class="obtn" data-o="interested"      onclick="selO('interested')">Interested</button>
        <button class="obtn" data-o="scheduled"       onclick="selO('scheduled')">Scheduled!</button>
      </div>
      <textarea class="ntxt" id="mnotes" rows="2" placeholder="Notes&#x2026;"></textarea>
      <div style="display:flex;gap:6px;align-items:center;margin-top:6px">
        <label style="font-size:9px;color:var(--sub);white-space:nowrap">Follow-up:</label>
        <input type="date" id="mfollowup" style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
      </div>
      <button class="btn blog" style="width:100%;margin-top:6px;padding:8px" onclick="saveL()">Save Log Entry</button>
    </div>
    <div style="margin-bottom:10px"><div class="msect">INSPECTION DETAILS</div><div class="mfacts" id="mfacts"></div></div>
    <div id="mhs" style="display:none;margin-bottom:10px">
      <div class="msect">CONTACT HISTORY</div><div class="mhist" id="mhist"></div>
    </div>

    <!-- CLOSE DEAL -->
    <div style="margin-bottom:10px">
      <div class="msect">CLOSE DEAL</div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:6px">
        <button onclick="markWon('customer_recurring')"
          style="padding:10px 6px;border:2px solid var(--grn);border-radius:9px;background:#ecfdf5;color:#059669;font-weight:700;font-size:11px;cursor:pointer;font-family:inherit">
          &#x1F4B0; Won  -  Recurring<br><span style="font-size:10px;font-weight:400" id="mwon-monthly"></span>
        </button>
        <button onclick="markWon('customer_once')"
          style="padding:10px 6px;border:2px solid var(--blu);border-radius:9px;background:#eff6ff;color:var(--blu);font-weight:700;font-size:11px;cursor:pointer;font-family:inherit">
          &#x1F9FC; Won  -  One-Time<br><span style="font-size:10px;font-weight:400" id="mwon-onetime"></span>
        </button>
      </div>
      <button onclick="markWon('quoted')"
        style="width:100%;padding:8px;border:1px solid var(--ora);border-radius:8px;background:#fff7f5;color:var(--ora);font-weight:600;font-size:11px;cursor:pointer;font-family:inherit">
        &#x1F4C4; Generate Quote &amp; Mark Quoted
      </button>
      <button onclick="markWon('churned')"
        style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:8px;background:transparent;color:var(--sub);font-size:10px;cursor:pointer;margin-top:4px;font-family:inherit">
        Mark as Lost / Churned
      </button>
    </div>

    <!-- CONTACTS -->
    <div style="margin-bottom:10px">
      <div class="msect">CONTACTS</div>
      <div id="mcontacts" style="margin-bottom:8px"></div>
      <div style="display:flex;gap:5px;margin-bottom:4px">
        <input id="mc-name" type="text" placeholder="Name" style="flex:2;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
        <select id="mc-role" style="flex:1;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="owner">Owner</option>
          <option value="gm">GM</option>
          <option value="manager">Manager</option>
          <option value="chef">Chef</option>
          <option value="staff">Staff</option>
        </select>
      </div>
      <input id="mc-phone" type="tel" placeholder="Direct phone (optional)" style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;margin-bottom:4px">
      <button onclick="addContact()" style="width:100%;padding:7px;border:none;border-radius:7px;background:var(--navy);color:#fff;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit">+ Save Contact</button>
    </div>
  </div>
</div>
<div id="toast"></div>

<script>
const P=%%DATA%%;
const PHONES=%%PHONES%%;
const ZIPS=%%ZIPS%%;

const PITCHES={
  callback:n=>`"Hi, is this <b>${n}</b>? I'm [Your Name] from Pinellas Ice Co  -  we clean and sanitize commercial ice machines locally. I'm reaching out because you recently had some health inspection issues. Ice machines are almost always part of callback inspections. We can get yours cleaned and documented before the inspector returns  -  usually takes about 2 hours. <b>Do you have time this week?</b>"`,
  overdue_urgent:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We specialize in ice machine cleaning in Pinellas and Hillsborough. I know dealing with inspection issues is stressful  -  we clean and document quickly and give you a service report you can show the inspector. <b>Can we schedule something this week?</b>"`,
  overdue:n=>`"Hi, is this <b>${n}</b>? This is [Your Name] from Pinellas Ice Co  -  commercial ice machine cleaning. Ice machines are one of the most flagged items when inspectors return. <b>Is this a good time to talk?</b>"`,
  pre_hot:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. Your next routine health inspection could be coming up very soon  -  ice machines are one of the most flagged items. We do a full clean and sanitize with a service report. <b>Want to get squared away before the inspector shows up?</b>"`,
  pre_warm:n=>`"Hi, is this <b>${n}</b>? My name is [Your Name] from Pinellas Ice Co. You have some lead time before your next inspection  -  perfect window to get a service on the books. <b>Would you be open to scheduling this month?</b>"`,
  high_risk:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We clean ice machines for restaurants in the area and based on your inspection history I wanted to reach out specifically. <b>Has anyone talked to you recently about your ice machine maintenance?</b>"`,
  routine:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We clean and sanitize commercial ice machines  -  FDA recommends service every 6 months. We're in your area. <b>Is this something that would be useful for you?</b>"`,
};
const WALKIN={
  callback:n=>`<b>Opening (to staff/gatekeeper):</b><br>&#x201C;Hey, is the manager around for one second? I&#39;m [Name] from Pinellas Ice Co. I&#39;ll be quick.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi  -  briefly: you had a callback inspection recently, and ice machines are almost always what inspectors target on the return visit. I clean and certify them, takes about 90 minutes, you get a dated service report to show the inspector. I&#39;m in the area this week. <b>Want me to take a look today?</b>&#x201D;<br><br><i>If hesitant:</i> &#x201C;No pressure  -  I can open it up and tell you what an inspector would flag. Five minutes, free.&#x201D;`,
  overdue_urgent:n=>`<b>Opening:</b><br>&#x201C;Hey, is the manager in? I&#39;ll be really quick.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi, I&#39;m [Name] from Pinellas Ice Co. Your location came up because you have an open inspection issue. Ice machines get hit hard on follow-up visits. I can clean and document yours before they return. <b>Do you have 90 minutes this week?</b>&#x201D;`,
  pre_hot:n=>`<b>Opening:</b><br>&#x201C;Hi, is the owner or manager available? Quick question about your health inspection.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co. Based on your inspection history your next one is probably coming up soon. Ice machines are one of the top-cited items  -  mold, mineral scale, slime. I service and document them so you&#39;re covered. <b>Can I show you what inspectors typically look for?</b>&#x201D;<br><br><i>Close:</i> &#x201C;I can come back at a time that works. Service is $149 and I leave you a dated report.&#x201D;`,
  pre_warm:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager around for a minute?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co  -  I clean commercial ice machines. You&#39;ve got some time before your next inspection, which is actually the perfect window  -  no last-minute stress. <b>Is ice machine maintenance on your radar?</b>&#x201D;`,
  high_risk:n=>`<b>Opening:</b><br>&#x201C;Hi, is the owner or manager in? I have some information about your inspection history.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi, I&#39;m [Name] from Pinellas Ice Co. Your location has had some ice-related violations flagged. This is exactly what inspectors look for on follow-ups. <b>When was your machine last serviced?</b>&#x201D;`,
  routine:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager available?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co  -  I clean commercial ice machines. FDA recommends every 6 months. Inspectors are specifically trained to look at machine interiors now. <b>When was yours last serviced?</b>&#x201D;<br><br><i>Let their answer guide you. If they don&#39;t know  -  that&#39;s your opening.</i>`,
  overdue:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager around?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co. Your location came up because of some past inspection flags. Ice machines are one of the most targeted items when inspectors return. <b>Do you mind if I take a quick look at yours?</b>&#x201D;`,
};
const OI={no_answer:'No Answer',voicemail:'Voicemail',not_interested:'Not Interested',follow_up:'Follow Up',interested:'Interested',scheduled:'Scheduled!',service_done:'Service Done ✓'};
const PO={CALLBACK:0,HOT:1,WARM:2,WATCH:3,LATER:4};
const PC={CALLBACK:'var(--cb)',HOT:'var(--hot)',WARM:'var(--warm)',WATCH:'var(--watch)',LATER:'var(--sub)'};
const ICN={V14:'food contact surfaces (V14)',V22:'non-PHF surfaces (V22)',V50:'food contact surfaces (V50)',V37:'equipment repair (V37)',V23:'utensil sanitation (V23)'};

// Objection handlers -- proven B2B field responses
const OBJECTIONS=[
  {q:"We already have someone who does it",a:"Great  -  how often are they coming out? Many restaurants we work with had someone, but it was once a year or on-call. We do documented monthly maintenance so you have a dated service report every visit, which is what inspectors want to see. Happy to do a free first look so you can compare."},
  {q:"Not interested / We're good",a:"Totally understand. Quick question  -  when was the last time someone actually opened up the machine and documented it? Most managers are surprised what is inside even when everything looks fine. I can do a free 5-minute look right now, no obligation."},
  {q:"Call me back / Not a good time",a:"No problem. What is a better time  -  Tuesday morning or Thursday afternoon? I am in this area twice a week. And can I grab your name so I ask for you directly?"},
  {q:"We do it ourselves / Our staff cleans it",a:"Staff cleaning is great for the ice bin and scoop  -  the question is the evaporator plates, water distribution system, and air filter inside. Those need chemical treatment inspectors specifically check. I can show you exactly what they look at  -  takes two minutes."},
  {q:"How much does it cost?",a:"$149/month for one machine  -  full cleaning, sanitizing, and a dated compliance report every visit. One callback inspection closing you for a day costs far more. If you just want a one-time deep clean to see how we work, that is $299 with a full report included."},
  {q:"We just had an inspection / We passed",a:"That is actually the best time  -  you have the most runway before the next one. Inspectors flag ice machines on re-visits because buildup comes back fast in Florida heat. Getting on a maintenance schedule now means you are protected before the next cycle."},
];

function dL(d){return d<0?Math.abs(d)+'d overdue':d===0?'TODAY':'+'+d+'d';}
function dC(d,p){return p==='CALLBACK'||d<0?'u':d<=21?'h':d<=45?'w':'';}
function stars(r){return r>0?'\u2605'.repeat(Math.round(r))+' '+r+'/5':'';}
function enc(s){return encodeURIComponent(s);}
function hav(la1,lo1,la2,lo2){
  const R=3958.8,a=Math.sin((la2-la1)*Math.PI/360)**2+
    Math.cos(la1*Math.PI/180)*Math.cos(la2*Math.PI/180)*Math.sin((lo2-lo1)*Math.PI/360)**2;
  return R*2*Math.atan2(Math.sqrt(a),Math.sqrt(1-a));
}

let log={},tab='today',selOut=null,cur=null,Q='',route=[],routeSet=new Set(),mapPros=[];

function lLoad(){try{log=JSON.parse(localStorage.getItem('pic_v4')||'{}')||{};}catch(e){log={};}}
function lSave(){try{localStorage.setItem('pic_v4',JSON.stringify(log));}catch(e){}}
function phLoad(){
  let saved={};try{saved=JSON.parse(localStorage.getItem('pic_phones')||'{}')||{};}catch(e){}
  P.forEach(p=>{const ph=saved[p.id]||PHONES[p.id];if(ph&&ph.phone){p.phone=ph.phone;p.rating=ph.rating||0;p.hours=ph.hours||'';}});
}
function phSave(id,phone,hours,rating){
  let s={};try{s=JSON.parse(localStorage.getItem('pic_phones')||'{}')||{};}catch(e){}
  s[id]={phone,hours,rating};try{localStorage.setItem('pic_phones',JSON.stringify(s));}catch(e){}
  const p=P.find(x=>x.id===id);if(p){p.phone=phone;p.hours=hours;p.rating=rating;}
}
function getLC(id){const e=log[id]||[];return e.length?e[e.length-1]:null;}
function isC(id){return(log[id]||[]).length>0;}

function fp(opts){
  return P.filter(p=>{
    if(Q&&!p.name.toLowerCase().includes(Q)&&!p.city.toLowerCase().includes(Q))return false;
    if(opts.county&&p.county!==opts.county)return false;
    if(opts.pri&&p.priority!==opts.pri)return false;
    if(opts.st){const l=getLC(p.id)?.outcome;if(opts.st==='not_contacted'){if(isC(p.id))return false;}else if(l!==opts.st)return false;}
    if(opts.ice){if(opts.ice==='chronic'&&!p.chronic)return false;if(opts.ice==='confirmed'&&!p.confirmed)return false;if(opts.ice==='high_ice'&&'medium'!=='high')return false;if(opts.ice==='phone'&&!p.phone)return false;}
    if(opts.hc&&isC(p.id))return false;
    return true;
  }).sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||b.score-a.score);
}

let _leafletLoaded=false;
function loadLeaflet(cb){
  if(typeof L!=='undefined'){cb();return;}
  if(_leafletLoaded){setTimeout(()=>loadLeaflet(cb),100);return;}
  _leafletLoaded=true;
  const lnk=document.createElement('link');
  lnk.rel='stylesheet';
  lnk.href='https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css';
  document.head.appendChild(lnk);
  const scr=document.createElement('script');
  scr.src='https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js';
  scr.onload=cb;
  scr.onerror=()=>{window._leafletFailed=true;cb();};
  document.head.appendChild(scr);
}
function sw(t){
  tab=t;
  document.querySelectorAll('.tab').forEach((el,i)=>el.classList.toggle('on',['today','all','route','customers','pipe','data'][i]===t));
  document.querySelectorAll('.panel').forEach(el=>el.classList.remove('on'));
  document.getElementById('p-'+t).classList.add('on');
  if(t==='today')rT();
  else if(t==='all')rA();
  else if(t==='route')loadLeaflet(rRoute);
  else if(t==='customers')rCust();
  else if(t==='pipe')rPipe();
}
function setF(p){sw('today');rT();}
function onS(){Q=document.getElementById('si').value.toLowerCase().trim();if(tab==='today')rT();else if(tab==='all')rA();}

function cardHTML(p){
  const last=getLC(p.id);
  // String concat only - no nested backtick templates (Safari compatibility)
  const phH=p.phone
    ?('<div class="phrow"><span class="phnum">'+p.phone+'</span><a href="tel:'+p.phone.replace(/\s/g,'')+'\" class="abtn call-a" onclick="event.stopPropagation()">Call</a></div>')
    :('<div class="phrow"><span class="phnum none">No phone on file</span><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone')+'" target="_blank" class="abtn find-a" onclick="event.stopPropagation()">Find</a></div>');
  const iceH=p.chronic
    ?('<div class="icebadge chronic">&#x1F9CA; CHRONIC  -  '+p.ice_count+'x ice violations</div>')
    :p.confirmed?'<div class="icebadge confirmed">&#x2713; Ice machine violation on record</div>':'';
  const codesH=(p.codes||[]).length?('<div style="font-size:9px;color:#2a4860;margin-bottom:4px">Codes: '+(p.codes||[]).join(', ')+'</div>'):'';
  const insH=''?('<div class="insight">'+''+'</div>'):'';
  const cbH=(p.priority==='CALLBACK'&&p.disp_risk>=4)
    ?('<div style="font-size:9px;font-weight:600;padding:3px 8px;border-radius:5px;margin-bottom:5px;background:#fef2f2;color:#dc2626;border:1px solid #fecaca">&#x26A0;&#xFE0F; Admin complaint  -  callback inspection due</div>')
    :'';
  const hrH=p.hours?('<div style="font-size:9px;color:#1a3850;margin-bottom:4px">&#x1F550; '+p.hours+'</div>'):'';
  const ratH=p.rating>0?('<div style="font-size:9px;color:#c08020;margin-bottom:3px">'+stars(p.rating)+'</div>'):'';
  const lastH=last
    ?('<div class="lastc hc">Last: '+(OI[last.outcome]||last.outcome)+' &middot; '+last.date+(last.notes?'  -  '+last.notes.slice(0,25):'')+' </div>')
    :'<div class="lastc">Not yet contacted</div>';
  const trendH=p.trending?'<div class="mi"><div class="ml">Trend</div><div class="mv bad">&#x2197; Worse</div></div>':'';
  const rtBtn=p.lat?('<button class="btn brt" onclick="event.stopPropagation();addToRoute('+p.id+')">+Route</button>'):'';
  const tc={PLATINUM:'#7c3aed',GOLD:'#d97706',SILVER:'#64748b',BRONZE:'#92400e'};
  const tbg={PLATINUM:'#ede9fe',GOLD:'#fef3c7',SILVER:'#f1f5f9',BRONZE:'#fef3c7'};
  const tierH=p.tier&&p.tier!=='COLD'?(' <span style="font-size:8px;font-weight:700;padding:2px 6px;border-radius:20px;background:'+(tbg[p.tier]||'#f1f5f9')+';color:'+(tc[p.tier]||'#64748b')+'">'+p.tier+'</span>'):'';
  const revenueH=p.monthly?('<div style="font-size:10px;font-weight:700;color:#059669;margin-bottom:5px">$'+p.monthly+'/mo &bull; '+(p.machines>1?p.machines+' machines':'1 machine')+'</div>'):'';
  const emergH=p.is_emergency?'<span class="emerg-badge">&#x1F6A8; EMERGENCY</span>':'';
  const confCol=p.confidence>=75?'#059669':p.confidence>=50?'#d97706':'#9ca3af';
  const confH='<span style="font-size:8px;font-weight:600;color:'+confCol+'" title="Prediction confidence">'+p.confidence+'% conf</span>';
  const franchH=p.biz_type==='franchise'?'<span style="font-size:8px;padding:1px 5px;border-radius:10px;background:#f0f9ff;color:#0ea5e9;border:1px solid #bae6fd">Franchise</span>':'';
  const custStatusH=(p.status&&p.status!=='prospect')
    ?('<div style="font-size:9px;font-weight:700;padding:2px 8px;border-radius:20px;display:inline-block;margin-bottom:4px;background:'+(p.status==='customer_recurring'?'#ecfdf5':p.status==='customer_once'?'#eff6ff':'#fff7f5')+';color:'+(p.status==='customer_recurring'?'#059669':p.status==='customer_once'?'var(--blu)':'var(--ora)')+'">'+({'customer_recurring':'Recurring Customer','customer_once':'One-Time Customer','quoted':'Quote Sent','churned':'Churned'}[p.status]||p.status)+'</div>')
    :'';
  return '<div class="card '+p.priority+(isC(p.id)?' done':'')+'" data-id="'+p.id+'" onclick="openM('+p.id+')">'
    +'<div class="ctop"><div class="cname">'+p.name+tierH+emergH+'</div><div style="display:flex;flex-direction:column;align-items:flex-end;gap:2px"><span class="pbadge '+p.priority+'">'+p.priority+'</span>'+confH+'</div></div>'
    +'<div class="cloc">'+p.city+', '+p.county+' '+franchH+'</div>'
    +custStatusH+revenueH+phH+ratH+hrH+iceH+cbH+codesH+insH
    +'<div class="cmeta">'
      +'<div class="mi"><div class="ml">Days Until</div><div class="mv '+dC(p.days_until,p.priority)+'">'+dL(p.days_until)+'</div></div>'
      +'<div class="mi"><div class="ml">H/T Viol</div><div class="mv '+(p.high_viol>=4?'u':p.high_viol>=2?'h':'')+'">'+p.high_viol+'/'+p.total_viol+'</div></div>'
      +'<div class="mi"><div class="ml">Pred.Next</div><div class="mv" style="font-size:10px">'+p.pred_next+'</div></div>'
      +trendH
    +'</div>'
    +lastH
    +'<div class="cacts">'
      +'<button class="btn blog" onclick="event.stopPropagation();openM('+p.id+')">Log Call</button>'
      +(last&&last.notes==='Skipped'
        ?'<button class="btn bskip" onclick="event.stopPropagation();unskip('+p.id+')" style="background:#ecfdf5;color:#059669;border-color:#6ee7b7">Unskip</button>'
        :'<button class="btn bskip" onclick="event.stopPropagation();skip('+p.id+')">Skip</button>')
      +rtBtn
    +'</div>'
  +'</div>';
}

function rT(){
  const county=document.getElementById('tc').value;
  const hc=document.getElementById('thc').checked;

  function filt(p){
    if(county && p.county!==county) return false;
    if(hc && isC(p.id)) return false;
    return true;
  }

  // Act Now: emergency closures + bad inspection last 30 days + callback pending
  const actNow = P.filter(p=>filt(p) && (
    p.is_emergency ||
    (p.priority==='CALLBACK' && p.days_since<=30) ||
    (p.disp_risk>=4 && p.days_since<=45)
  )).sort((a,b)=>(b.is_emergency-a.is_emergency)||(b.score-a.score));

  // Upcoming: inspection within 60 days + ice history
  const upcoming = P.filter(p=>filt(p) &&
    !actNow.find(x=>x.id===p.id) &&
    p.days_until>=0 && p.days_until<=60 &&
    (p.confirmed||p.chronic||p.ice_rel>0)
  ).sort((a,b)=>a.days_until-b.days_until);

  // High Value: PLATINUM or GOLD by revenue
  const highVal = P.filter(p=>filt(p) &&
    !actNow.find(x=>x.id===p.id) &&
    !upcoming.find(x=>x.id===p.id) &&
    (p.tier==='PLATINUM'||p.tier==='GOLD')
  ).sort((a,b)=>(b.monthly||0)-(a.monthly||0));

  const total = actNow.length + upcoming.length + highVal.length;
  document.getElementById('tcnt').textContent = total + ' prospects';

  function renderSection(ids, gridId, emptyId){
    const g = document.getElementById(gridId);
    const e = document.getElementById(emptyId);
    if(!ids.length){g.innerHTML='';e.style.display='block';return;}
    e.style.display='none';
    g.innerHTML = ids.slice(0,40).map(cardHTML).join('');
    g.querySelectorAll('.card').forEach(card=>{
      let tx=0,ty=0;
      card.addEventListener('touchstart',ev=>{tx=ev.touches[0].clientX;ty=ev.touches[0].clientY;},{passive:true});
      card.addEventListener('touchend',ev=>{
        const dx=ev.changedTouches[0].clientX-tx;
        const dy=Math.abs(ev.changedTouches[0].clientY-ty);
        if(Math.abs(dx)>70&&dy<40){
          const id=parseInt(card.getAttribute('data-id'));
          if(!id)return;
          if(dx<0)skip(id); else openM(id);
        }
      },{passive:true});
    });
  }

  renderSection(actNow,   'tgrid-actnow',  'empty-actnow');
  renderSection(upcoming, 'tgrid-upcoming','empty-upcoming');
  renderSection(highVal,  'tgrid-highval', 'empty-highval');

  // Follow-ups due today or overdue
  const todayStr=new Date().toISOString().slice(0,10);
  const followups=P.filter(p=>{
    const entries=log[p.id]||[];
    const withFollowup=entries.filter(e=>e.followup&&e.followup<=todayStr);
    return withFollowup.length>0;
  }).sort((a,b)=>{
    const fa=(log[a.id]||[]).filter(e=>e.followup).slice(-1)[0]?.followup||'';
    const fb=(log[b.id]||[]).filter(e=>e.followup).slice(-1)[0]?.followup||'';
    return fa.localeCompare(fb);
  });
  renderSection(followups,'tgrid-followups','empty-followups');
}

// ── PRESET DEFINITIONS ───────────────────────────────────────────────────────
const PRESETS = {
  all:       {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:''},
  actnow:    {county:'',pri:'CALLBACK',ice:'confirmed',tier:'',btype:'',st:'',conf:''},
  callback:  {county:'',pri:'CALLBACK',ice:'',tier:'',btype:'',st:'',conf:''},
  phone:     {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:'',hasPhone:true},
  chronic:   {county:'',pri:'',ice:'chronic',tier:'',btype:'',st:'',conf:''},
  gold:      {county:'',pri:'',ice:'',tier:'GOLD',btype:'',st:'',conf:''},
  franchise: {county:'',pri:'',ice:'',tier:'',btype:'franchise',st:'',conf:''},
  bar:       {county:'',pri:'',ice:'',tier:'',btype:'',st:'',conf:'',isBar:true},
  notyet:    {county:'',pri:'',ice:'',tier:'',btype:'',st:'not_contacted',conf:''},
};

let _preset='all', _extraFilt={};

function setPreset(key){
  _preset=key;
  _extraFilt={};
  const cfg=PRESETS[key]||PRESETS.all;
  // Set dropdowns
  const set=(id,v)=>{const el=document.getElementById(id);if(el)el.value=v;};
  set('ac',cfg.county||''); set('ap',cfg.pri||''); set('ai',cfg.ice||'');
  set('atier',cfg.tier||''); set('abtype',cfg.btype||'');
  set('as_',cfg.st||''); set('aconf',cfg.conf||'');
  if(cfg.hasPhone) _extraFilt.hasPhone=true;
  if(cfg.isBar)    _extraFilt.isBar=true;
  // Highlight active preset button
  document.querySelectorAll('.preset-btn').forEach(b=>b.classList.remove('on'));
  const btn=document.getElementById('pre-'+key);
  if(btn)btn.classList.add('on');
  rA();
}

function clearFilters(){
  setPreset('all');
}

function rA(){
  const county  = document.getElementById('ac')?.value||'';
  const pri     = document.getElementById('ap')?.value||'';
  const ice     = document.getElementById('ai')?.value||'';
  const tier    = document.getElementById('atier')?.value||'';
  const btype   = document.getElementById('abtype')?.value||'';
  const st      = document.getElementById('as_')?.value||'';
  const conf    = parseInt(document.getElementById('aconf')?.value)||0;

  const data = P.filter(p=>{
    if(Q && !p.name.toLowerCase().includes(Q) && !p.city.toLowerCase().includes(Q)) return false;
    if(county && p.county!==county) return false;
    if(pri    && p.priority!==pri)  return false;
    if(ice==='chronic'   && !p.chronic)   return false;
    if(ice==='confirmed' && !p.confirmed) return false;
    if(tier   && p.tier!==tier)     return false;
    if(btype  && p.biz_type!==btype) return false;
    if(conf   && (p.confidence||0)<conf) return false;
    if(st){
      const l=getLC(p.id);
      const outcome=l?l.outcome:'';
      if(st==='not_contacted' && isC(p.id)) return false;
      else if(st!=='not_contacted' && outcome!==st) return false;
    }
    if(_extraFilt.hasPhone && !p.phone) return false;
    if(_extraFilt.isBar   && !p.is_bar) return false;
    return true;
  });

  const cnt = document.getElementById('acnt');
  if(cnt) cnt.textContent = data.length.toLocaleString()+' businesses';

  const grid  = document.getElementById('agrid');
  const empty = document.getElementById('a-empty');
  if(!data.length){
    if(grid)  grid.innerHTML='';
    if(empty) empty.style.display='block';
    return;
  }
  if(empty) empty.style.display='none';
  if(grid)  grid.innerHTML = data.slice(0,80).map(cardHTML).join('');
}

// ROUTE
function rRoute(){
  const county=document.getElementById('rc').value;
  const pri=document.getElementById('rp').value;
  const zip=document.getElementById('rzip').value.trim();
  const rad=parseFloat(document.getElementById('rrad').value)||999;
  let data=P.filter(p=>p.lat&&p.lon);
  if(county)data=data.filter(p=>p.county===county);
  if(pri)data=data.filter(p=>p.priority===pri);
  if(zip.length===5&&ZIPS[zip]){
    const [clat,clon]=ZIPS[zip];
    data=data.map(p=>({...p,_d:hav(clat,clon,p.lat,p.lon)}))
      .filter(p=>p._d<=rad)
      .sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||a._d-b._d);
    document.getElementById('rhint').textContent=`${data.length} prospects within ${rad}mi of ${zip}`;
  }else{
    data.sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||b.score-a.score);
    document.getElementById('rhint').textContent=`${data.length} geocoded prospects  -  enter ZIP to filter by area`;
  }
  mapPros=data.slice(0,100);
  renderRList();renderMap();
}

function renderRList(){
  const el=document.getElementById('rlist');
  const cnt=document.getElementById('rlist-cnt');
  if(!mapPros.length){
    el.innerHTML='<div style="font-size:11px;color:var(--sub);padding:8px">No geocoded prospects match. Try removing ZIP filter or expanding radius.</div>';
    if(cnt)cnt.textContent='';
    return;
  }
  if(cnt)cnt.textContent=mapPros.length+' prospects nearby';
  el.innerHTML=mapPros.map(p=>{
    const inR=routeSet.has(p.id);
    const col=PC[p.priority]||'var(--sub)';
    const distTxt=p._d?p._d.toFixed(1)+'mi':'';
    const chronicTxt=p.chronic?'<div style="font-size:8px;color:#059669;font-weight:700">CHRONIC ICE</div>':'';
    return '<div class="rcard'+(inR?' sel':'')+'" onclick="addToRoute('+p.id+')" style="border-left:3px solid '+col+(inR?';background:#fff7f5':'')+';">'
      +'<div class="rdot" style="background:'+col+'"></div>'
      +'<div style="flex:1;min-width:0">'
        +'<div class="rname" style="color:var(--navy)">'+p.name+(inR?' <span style="color:var(--ora)">&#x2713;</span>':'')+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+p.city+' &bull; '+p.priority+' &bull; '+dL(p.days_until)+'</div>'
        +chronicTxt
      +'</div>'
      +'<div style="display:flex;flex-direction:column;align-items:flex-end;gap:3px">'
        +(distTxt?'<div class="rdist">'+distTxt+'</div>':'')
        +'<button onclick="event.stopPropagation();openM('+p.id+')" style="font-size:8px;padding:2px 5px;border:1px solid var(--brd);border-radius:4px;background:var(--surf);color:var(--sub);cursor:pointer">Details</button>'
      +'</div>'
      +'</div>';
  }).join('');
}

function renderMap(){
  const area=document.getElementById('map-area');
  if(typeof L==='undefined'||window._leafletFailed){
    // Leaflet not available (sandbox or offline) - show static list
    if(!mapPros.length){area.innerHTML='<div class="map-empty"><div style="font-size:32px">&#x1F5FA;&#xFE0F;</div><div>No prospects to map</div></div>';return;}
    area.innerHTML='<div style="padding:12px;overflow-y:auto;height:100%"><div style="font-size:10px;color:var(--sub);margin-bottom:8px">Map unavailable offline. '+(mapPros.length)+' prospects:</div>'+mapPros.slice(0,20).map(p=>'<div style="padding:5px 0;border-bottom:1px solid var(--brd2);font-size:11px;color:var(--navy)">'+p.name+'</div>').join('')+'</div>';
    return;
  }
  if(!mapPros.length){
    area.innerHTML='<div class="map-empty"><div style="font-size:36px;margin-bottom:10px">&#x1F5FA;&#xFE0F;</div><div>No prospects to map</div></div>';
    if(window._lmap){window._lmap.remove();window._lmap=null;}
    return;
  }
  // Init Leaflet map once
  if(!window._lmap){
    area.innerHTML='<div id="leaflet-map" style="width:100%;height:100%;border-radius:10px"></div>';
    window._lmap=L.map('leaflet-map',{zoomControl:true,attributionControl:false});
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
      maxZoom:18,opacity:0.85
    }).addTo(window._lmap);
    window._lmarkers=L.layerGroup().addTo(window._lmap);
  }
  window._lmarkers.clearLayers();
  const cols={CALLBACK:'#ff3b30',HOT:'#ff9f0a',WARM:'#ffd60a',WATCH:'#5e9eff',LATER:'#445066'};
  const bounds=[];
  mapPros.forEach(p=>{
    const inR=routeSet.has(p.id);
    const col=inR?'#00e5ff':cols[p.priority]||'#445066';
    const r=inR?12:p.chronic?10:8;
    const marker=L.circleMarker([p.lat,p.lon],{
      radius:r,color:col,fillColor:col,fillOpacity:0.85,weight:inR?3:1.5,
      opacity:1
    }).addTo(window._lmarkers);
    marker.bindPopup(
      '<b>'+p.name+'</b><br>'+p.city+' &bull; '+p.priority+'<br>'+dL(p.days_until)+
      (p.seats?' &bull; '+p.seats+' seats':'')+
      (p.monthly?' &bull; $'+p.monthly+'/mo':'')+
      '<br><button onclick="addToRoute('+p.id+')" '+
      'style="margin-top:4px;padding:3px 8px;border:none;border-radius:5px;background:'+col+';color:'+(p.priority==='WARM'?'#000':'#fff')+';font-size:10px;font-weight:700;cursor:pointer">'+
      (routeSet.has(p.id)?'Remove from Route':'+Route')+'</button> '+
      '<button onclick="openM('+p.id+')" '+
      'style="margin-top:3px;padding:3px 8px;border:1px solid #333;border-radius:5px;background:#111;color:#fff;font-size:10px;cursor:pointer">Details</button>',
      {maxWidth:220}
    );
    bounds.push([p.lat,p.lon]);
  });
  if(bounds.length)window._lmap.fitBounds(bounds,{padding:[20,20],maxZoom:13});
}

function addToRoute(id){
  const p=P.find(x=>x.id===id);if(!p)return;
  if(routeSet.has(id)){routeSet.delete(id);route=route.filter(r=>r.id!==id);}
  else{if(route.length>=8){toast('Max 8 stops. Open in Maps first, then add more.');return;}routeSet.add(id);route.push(p);}
  renderRList();renderMap();renderDayRoute();
}
function renderDayRoute(){
  const dr=document.getElementById('day-route');
  document.getElementById('stopcnt').textContent=route.length;
  if(!route.length){dr.style.display='none';return;}
  dr.style.display='block';

  // Estimate total route distance
  let totalMi=0;
  for(let i=1;i<route.length;i++){
    if(route[i-1].lat&&route[i].lat)
      totalMi+=hav(route[i-1].lat,route[i-1].lon,route[i].lat,route[i].lon);
  }
  const mi=document.getElementById('route-mi');
  if(mi&&totalMi>0)mi.textContent=' (~'+totalMi.toFixed(1)+'mi)';

  document.getElementById('day-stops').innerHTML=route.map((p,i)=>{
    const ph=p.phone?('<a href="tel:'+p.phone.replace(/\s/g,'')+'\" style="font-size:10px;color:var(--blu);text-decoration:none">'+p.phone+'</a>'):'';
    const revenue='$'+p.monthly+'/mo';
    return '<div class="day-stop">'
      +'<div class="stopnum">'+(i+1)+'</div>'
      +'<div style="flex:1;min-width:0">'
        +'<div style="font-weight:700;font-size:11px;color:var(--navy);overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+p.name+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+p.address+', '+p.city+'</div>'
        +'<div style="display:flex;gap:6px;align-items:center;margin-top:2px">'
          +ph
          +'<span style="font-size:9px;font-weight:600;color:#059669">'+revenue+'</span>'
          +'<span class="pbadge '+p.priority+'" style="font-size:7px">'+p.priority+'</span>'
        +'</div>'
      +'</div>'
      +'<div style="display:flex;flex-direction:column;gap:3px;flex-shrink:0">'
        +'<button onclick="openM('+p.id+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--ora);border-radius:5px;background:#fff7f5;color:var(--ora);cursor:pointer;font-family:inherit">Details</button>'
        +'<button onclick="removeStop('+p.id+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">Remove</button>'
      +'</div>'
      +'</div>';
  }).join('');
}

function optimizeRoute(){
  if(route.length<3){toast('Need at least 3 stops to optimize');return;}
  // Nearest-neighbor from first stop
  const stops=[...route];
  const ordered=[stops.shift()];
  while(stops.length){
    const last=ordered[ordered.length-1];
    if(!last.lat){ordered.push(stops.shift());continue;}
    let best=0,bestD=Infinity;
    stops.forEach((s,i)=>{
      if(!s.lat){return;}
      const d=hav(last.lat,last.lon,s.lat,s.lon);
      if(d<bestD){bestD=d;best=i;}
    });
    ordered.push(stops.splice(best,1)[0]);
  }
  route=ordered;
  routeSet=new Set(route.map(r=>r.id));
  renderRList();renderMap();renderDayRoute();
  toast('Route optimized by distance!');
}
function openMaps(){
  if(!route.length)return;
  const url='https://www.google.com/maps/dir/'+route.filter(p=>p.address).map(p=>enc(p.address+', '+p.city+', FL '+p.zip)).join('/');
  window.open(url,'_blank');
}
function clearRoute(){route=[];routeSet=new Set();renderRList();renderMap();renderDayRoute();}
function removeStop(id){routeSet.delete(id);route=route.filter(r=>r.id!==id);renderRList();renderMap();renderDayRoute();}

// MODAL
function openM(id){
  const p=P.find(x=>x.id===id);if(!p)return;
  cur=p;selOut=null;
  document.querySelectorAll('.obtn').forEach(b=>b.classList.remove('on'));
  document.getElementById('mnotes').value='';
  document.getElementById('mn').textContent=p.name;
  document.getElementById('ml').textContent=p.address+', '+p.city+', FL '+p.zip+' \u00b7 #'+p.id;
  const pe=document.getElementById('mph'),ae=document.getElementById('mpa');
  if(p.phone){
    pe.textContent=p.phone;pe.className='mphnum';
    const cl=p.phone.replace(/\\s/g,'');
    ae.innerHTML=`<a href="tel:${cl}" class="mcall">Call Now</a><a href="sms:${cl}" class="msms">Text</a><a href="https://www.google.com/search?q=${enc(p.name+' '+p.city+' FL')}" target="_blank" class="mgoog">Google</a>`;
  }else{
    pe.textContent='No phone on file';pe.className='mphnum none';
    ae.innerHTML=`<a href="https://www.google.com/search?q=${enc(p.name+' '+p.city+' FL phone number')}" target="_blank" class="mgoog">Find Phone</a><a href="https://maps.google.com/search?q=${enc(p.name+' '+p.address+' '+p.city+' FL')}" target="_blank" class="mgoog">Maps</a>`;
  }
  document.getElementById('mhrs').textContent=p.hours?'Hours: '+p.hours:'';
  document.getElementById('mrat').textContent=p.rating>0?stars(p.rating):'';
  const iceEl=document.getElementById('mice');
  if(p.confirmed||p.chronic){
    const codes=(p.codes||[]).map(c=>ICN[c]||c).join(', ');
    const parts=['<div class="micebox">'];
    parts.push('<div style="font-size:8px;color:#013a18;letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px">&#x1F9CA; ICE MACHINE HISTORY</div>');
    if(p.chronic) parts.push('<div style="font-size:13px;font-weight:700;color:var(--grn);margin-bottom:3px">CHRONIC  -  flagged in '+p.ice_count+' inspections</div>');
    if(!p.chronic&&p.confirmed) parts.push('<div style="font-size:12px;color:#55c8ff;margin-bottom:3px">Confirmed ice machine violation on record</div>');
    if(codes) parts.push('<div style="font-size:9px;color:#2a5a38;margin-bottom:3px">Violation codes: '+codes+'</div>');
    if('') parts.push('<div style="font-size:10px;color:#3a6a48;font-style:italic">'+''+'</div>');
    parts.push('</div>');
    iceEl.innerHTML=parts.join('');
  }else iceEl.innerHTML='';
  document.getElementById('mpitch').innerHTML=(PITCHES[p.pitch_type]||PITCHES.routine)(p.name);
  document.getElementById('mwalkin').innerHTML=(WALKIN[p.pitch_type]||WALKIN.routine)(p.name);
  // Objection handlers
  const objEl=document.getElementById('mobjections');
  if(objEl){
    objEl.innerHTML=OBJECTIONS.map((o,i)=>
      '<details style="background:#f5f8fa;border:1px solid var(--brd2);border-radius:7px;padding:0">'
      +'<summary style="padding:7px 10px;font-size:10px;font-weight:600;color:var(--navy);cursor:pointer;list-style:none;display:flex;align-items:center;gap:6px">'
      +'<span style="color:var(--cb)">&#x2753;</span>'+o.q+'</summary>'
      +'<div style="padding:4px 10px 9px;font-size:10px;color:var(--sub2);line-height:1.6;border-top:1px solid var(--brd2)">'+o.a+'</div>'
      +'</details>'
    ).join('');
  }
  document.getElementById('mfacts').innerHTML=[
    ['Predicted Next',   p.pred_next,''],
    ['Days Until',       dL(p.days_until),''],
    ['High Violations',  p.high_viol,p.high_viol>=4?'r':p.high_viol>=2?'o':''],
    ['Total Violations', p.total_viol,''],
    ['Last Inspected',   p.last_insp,''],
    ['Days Since Insp.', p.days_since+'d',''],
    ['Inspections',      p.n_insp,''],
    ['Viol. Trend',      p.trending?'Getting Worse':'Stable',p.trending?'r':'g'],
    ['Ice Violations',   p.ice_count>0?p.ice_count+'x flagged':'None',''],
    ['Confidence',       (p.confidence||0)+'%',p.confidence>=75?'g':p.confidence>=50?'o':''],
    ['Seats',            p.seats>0?p.seats:'Unknown',''],
    ['Est. Machines',    p.machines||1,''],
    ['Monthly Recurring','$'+(p.monthly||149)+'/mo','g'],
    ['One-Time Clean',   '$'+(p.onetime||299),'b'],
    ['Account Tier',     p.tier||'COLD',''],
    ['Biz Type',         (p.biz_type||'independent').charAt(0).toUpperCase()+(p.biz_type||'independent').slice(1),''],
  ].map(([l,v,c])=>'<div class="fact"><div class="fl">'+l+'</div><div class="fv '+c+'">'+v+'</div></div>').join('');
  const hist=log[id]||[];
  const hs=document.getElementById('mhs');
  if(hist.length){hs.style.display='block';document.getElementById('mhist').innerHTML=[...hist].reverse().map(e=>{
    const noteHtml=e.notes?('<div style="color:var(--sub);font-size:9px;margin-top:1px">'+e.notes+'</div>'):'';
    return '<div class="hi">'+(OI[e.outcome]||e.outcome)+'<span style="color:var(--sub);font-size:9px"> &middot; '+e.date+'</span>'+noteHtml+'</div>';
  }).join('');}
  else hs.style.display='none';
  document.getElementById('mwon-monthly').textContent='$'+p.monthly+'/mo recurring';
  document.getElementById('mwon-onetime').textContent='$'+p.onetime+' one-time';
  renderContacts(id);
  document.getElementById('mbg').classList.add('on');
}
function setPitchMode(mode){
  const phoneBtn=document.getElementById('btn-phone');
  const walkinBtn=document.getElementById('btn-walkin');
  const phoneDiv=document.getElementById('mpitch');
  const walkinDiv=document.getElementById('mwalkin');
  if(mode==='walkin'){
    phoneDiv.style.display='none';walkinDiv.style.display='block';
    phoneBtn.style.background='transparent';phoneBtn.style.color='var(--sub)';phoneBtn.style.borderColor='var(--brd)';
    walkinBtn.style.background='#0a84ff22';walkinBtn.style.color='var(--blu)';walkinBtn.style.borderColor='var(--blu)';
  }else{
    phoneDiv.style.display='block';walkinDiv.style.display='none';
    phoneBtn.style.background='#0a84ff22';phoneBtn.style.color='var(--blu)';phoneBtn.style.borderColor='var(--blu)';
    walkinBtn.style.background='transparent';walkinBtn.style.color='var(--sub)';walkinBtn.style.borderColor='var(--brd)';
  }
}
function selO(o){selOut=o;document.querySelectorAll('.obtn').forEach(b=>b.classList.toggle('on',b.dataset.o===o));}
function saveL(){
  if(!cur||!selOut){toast('Pick an outcome first');return;}
  const notes=document.getElementById('mnotes').value.trim();
  const followup=(document.getElementById('mfollowup')||{}).value||'';
  const e={outcome:selOut,date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),notes,followup};
  if(!log[cur.id])log[cur.id]=[];
  log[cur.id].push(e);lSave();
  document.getElementById('mbg').classList.remove('on');
  if(document.getElementById('mfollowup'))document.getElementById('mfollowup').value='';
  const msg=followup?OI[selOut]+' logged. Follow-up: '+followup:OI[selOut]+' logged';
  toast(msg);
  if(tab==='today')rT();else if(tab==='all')rA();else if(tab==='pipe')rPipe();
}
function skip(id){
  if(!log[id])log[id]=[];
  log[id].push({outcome:'no_answer',date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),notes:'Skipped'});
  lSave();rT();
}
function unskip(id){
  if(!log[id])return;
  // Remove last entry if it was a skip
  const last=log[id][log[id].length-1];
  if(last&&last.notes==='Skipped')log[id].pop();
  if(!log[id].length)delete log[id];
  lSave();if(tab==='today')rT();else if(tab==='all')rA();
  toast('Removed from skipped');
}

// PIPELINE
function rPipe(){
  const grps={
    scheduled:{l:'Scheduled / Booked',c:'#32d74b',ids:[]},
    interested:{l:'Interested',c:'#0a84ff',ids:[]},
    follow_up:{l:'Follow Up Needed',c:'#ffd60a',ids:[]},
    voicemail:{l:'Voicemail Left',c:'#5e9eff',ids:[]},
    no_answer:{l:'No Answer',c:'#445066',ids:[]},
    not_interested:{l:'Not Interested',c:'#ff3b30',ids:[]},
    service_done:{l:'Service Done',c:'#059669',ids:[]},
  };
  Object.entries(log).forEach(([id,entries])=>{if(!entries.length)return;const l=entries[entries.length-1];if(grps[l.outcome])grps[l.outcome].ids.push(parseInt(id));});
  const pm={};P.forEach(p=>pm[p.id]=p);
  const c=document.getElementById('pipec');
  const hasAny=Object.values(grps).some(g=>g.ids.length>0);
  if(!hasAny){c.innerHTML='<div class="empty"><div class="ei">&#x1F4DE;</div><div>Start logging calls to build your pipeline</div></div>';return;}
  c.innerHTML=Object.entries(grps).map(([k,g])=>{
    if(!g.ids.length)return'';
    const items=g.ids.map(id=>{
      const p=pm[id];if(!p)return'';
      const last=(log[id]||[]).slice(-1)[0];
      return '<div class="pitem" onclick="openM('+id+')">'
        +'<div class="pdot" style="background:'+g.c+'"></div>'
        +'<div style="flex:1">'
          +'<div class="piname">'+p.name+'</div>'
          +'<div style="font-size:9px;color:var(--sub)">'+p.city+', '+p.county+' &bull; '+dL(p.days_until)+' &bull; '+last.date+'</div>'
          +(p.phone?('<div class="piph">'+p.phone+'</div>'):'')
          +(last.notes?('<div style="font-size:9px;color:var(--sub);margin-top:1px">'+last.notes.slice(0,50)+'</div>'):'')
        +'</div>'
        +'<span class="pbadge '+p.priority+'">'+p.priority+'</span>'
      +'</div>';
    }).join('');
    return `<div class="psect"><div class="pst" style="color:${g.c}">${g.l}<span class="pct">${g.ids.length}</span></div>${items}</div>`;
  }).join('');
}

// ADD PHONE
function addPhone(){
  const id=parseInt(document.getElementById('ph-id').value.trim());
  const phone=document.getElementById('ph-num').value.trim();
  const hours=document.getElementById('ph-hrs').value.trim();
  if(!id||!phone){toast('Enter a License ID and phone number');return;}
  phSave(id,phone,hours,0);
  document.getElementById('ph-id').value='';
  document.getElementById('ph-num').value='';
  document.getElementById('ph-hrs').value='';
  toast('Phone saved for #'+id);
  if(tab==='today')rT();
}

// EXPORT
function exportCSV(){
  const rows=[['Name','County','City','Phone','Hours','Rating','Priority','Days Until','Pred Next','High Viol','Total Viol','Ice Profile','Codes','Trend','Last Outcome','Date','Notes']];
  P.forEach(p=>{
    const l=getLC(p.id);
    rows.push([p.name,p.county,p.city,p.phone||'',p.hours||'',p.rating||'',p.priority,
      p.days_until,p.pred_next,p.high_viol,p.total_viol,
      p.chronic?'CHRONIC':p.confirmed?'confirmed':'none',
      (p.codes||[]).join(';'),p.trending?'worse':'stable',
      l?OI[l.outcome]||l.outcome:'',l?l.date:'',l?l.notes:'']);
  });
  const csv=rows.map(r=>r.map(c=>'"'+String(c).replace(/"/g,'""')+'"').join(',')).join('\\n');
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
  a.download='pic_prospects_'+new Date().toISOString().slice(0,10)+'.csv';
  a.click();toast('CSV exported');
}
function clrLog(){if(!confirm('Clear all call log data? This cannot be undone.'))return;log={};lSave();if(tab==='today')rT();if(tab==='pipe')rPipe();toast('Cleared');}
function toast(msg){const t=document.getElementById('toast');t.textContent=msg;t.classList.add('on');setTimeout(()=>t.classList.remove('on'),2200);}

// ── CUSTOMER LIFECYCLE ────────────────────────────────────────────────────────
let customers = {};  // {id: {status, won_date, service_type, notes, last_service}}

function custLoad(){
  try{customers=JSON.parse(localStorage.getItem('pic_customers')||'{}')||{};}catch(e){customers={};}
  // Apply saved statuses to P records
  P.forEach(p=>{const c=customers[p.id];if(c)p.status=c.status;});
}
function custSave(){try{localStorage.setItem('pic_customers',JSON.stringify(customers));}catch(e){}}

function markWon(status){
  if(!cur)return;
  const p=cur;
  const now=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  customers[p.id]={
    status,
    won_date: now,
    service_type: status==='customer_recurring'?'recurring':'one_time',
    monthly: status==='customer_recurring'?p.monthly:0,
    onetime: status==='customer_once'?p.onetime:0,
    machines: p.machines,
    name: p.name,
    address: p.address,
    city: p.city,
    phone: p.phone,
    notes: '',
    last_service: '',
    next_service: '',
  };
  p.status=status;
  custSave();
  // Log it
  if(!log[p.id])log[p.id]=[];
  log[p.id].push({outcome:status,date:now,notes:'Deal closed'});
  lSave();
  if(status==='quoted'){
    showQuote(p);
    toast('Marked as Quoted!');
  } else {
    toast(status==='customer_recurring'?'Won! Added to recurring customers.':'Won! One-time service recorded.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  }
}

function rCust(){
  const filter=document.getElementById('cust-status').value;
  const allCusts=P.filter(p=>p.status&&p.status!=='prospect');
  const shown=filter?allCusts.filter(p=>p.status===filter):allCusts;

  // MRR calculation
  const recurring=allCusts.filter(p=>p.status==='customer_recurring');
  const mrr=recurring.reduce((s,p)=>s+(p.monthly||149),0);
  document.getElementById('mrr-val').textContent='$'+mrr.toLocaleString();
  document.getElementById('cust-count').textContent=recurring.length;
  document.getElementById('arr-val').textContent='$'+(mrr*12).toLocaleString();

  const el=document.getElementById('cust-list');
  const em=document.getElementById('cust-empty');
  document.getElementById('cust-cnt').textContent=shown.length+' accounts';

  if(!shown.length){el.innerHTML='';em.style.display='block';return;}
  em.style.display='none';

  const STATUS_LABELS={
    customer_recurring:'&#x1F504; Recurring',customer_once:'&#x1F9FC; One-Time',
    quoted:'&#x1F4C4; Quote Sent',churned:'&#x274C; Churned'
  };
  const STATUS_COLORS={
    customer_recurring:'#059669',customer_once:'var(--blu)',
    quoted:'var(--ora)',churned:'var(--cb)'
  };

  // Portal ID for HubSpot links
  const portalId=loadSettings().hubspot_portal||'';

  const today=new Date();

  el.innerHTML=shown.map(p=>{
    const c=customers[p.id]||{};
    const col=STATUS_COLORS[p.status]||'var(--sub)';
    const lbl=STATUS_LABELS[p.status]||p.status;
    const rev=p.status==='customer_recurring'?('$'+p.monthly+'/mo'):p.status==='customer_once'?('$'+p.onetime+' one-time'):'';

    // Service due indicator
    let serviceDueH='';
    if(c.next_service){
      const due=new Date(c.next_service);
      const daysUntil=Math.round((due-today)/864e5);
      const dueCol=daysUntil<0?'#dc2626':daysUntil<=7?'#d97706':'#059669';
      const dueTxt=daysUntil<0?Math.abs(daysUntil)+'d overdue':daysUntil===0?'Due TODAY':daysUntil+'d until service';
      serviceDueH='<div style="font-size:9px;font-weight:700;color:'+dueCol+'">&#x1F4C5; '+dueTxt+'</div>';
    }

    // HubSpot link
    const hsUrl=portalId
      ?('https://app.hubspot.com/contacts/'+portalId+'/company/create?name='+enc(p.name)+'&city='+enc(p.city))
      :'https://app.hubspot.com';

    return '<div class="pitem" style="border-left:3px solid '+col+';flex-direction:column;align-items:stretch;gap:8px">'
      // Top row
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+', '+p.county+(p.phone?' &bull; '+p.phone:'')+'</div>'
          +'<div style="font-size:9px;font-weight:700;color:'+col+';margin-top:2px">'+lbl+'</div>'
        +'</div>'
        +'<div style="text-align:right;flex-shrink:0;margin-left:10px">'
          +(rev?'<div style="font-size:13px;font-weight:800;color:var(--grn)">'+rev+'</div>':'')
          +(c.won_date?'<div style="font-size:9px;color:var(--sub)">Since '+c.won_date+'</div>':'')
        +'</div>'
      +'</div>'
      // Service tracking row
      +'<div style="background:#f5f8fa;border-radius:7px;padding:8px;display:flex;flex-direction:column;gap:5px">'
        +'<div style="display:flex;justify-content:space-between;align-items:center">'
          +'<span style="font-size:9px;color:var(--sub);font-weight:600;text-transform:uppercase">Last Service</span>'
          +'<span style="font-size:10px;font-weight:600;color:var(--navy)">'+(c.last_service||'Not recorded')+'</span>'
        +'</div>'
        +serviceDueH
        +'<div style="display:flex;gap:5px;margin-top:3px">'
          +'<button onclick="event.stopPropagation();logService('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:none;border-radius:6px;background:var(--grn);color:#fff;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Log Service Visit</button>'
          +'<button onclick="event.stopPropagation();setNextService('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--sub);cursor:pointer;font-family:inherit">Set Next Due</button>'
        +'</div>'
      +'</div>'
      // Action buttons
      +'<div style="display:flex;gap:5px">'
        +'<button onclick="event.stopPropagation();openM('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--navy);cursor:pointer;font-family:inherit">Details</button>'
        +'<button onclick="event.stopPropagation();showQuote(P.find(x=>x.id==='+p.id+'))" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--ora);border-radius:6px;background:#fff7f5;color:var(--ora);cursor:pointer;font-family:inherit">Quote</button>'
        +'<a href="'+hsUrl+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #ff7a59;border-radius:6px;background:#fff7f5;color:#ff7a59;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">HubSpot &#x2197;</a>'
      +'</div>'
      +'</div>';
  }).join('');
}

function logService(id){
  const today=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  if(!customers[id])customers[id]={};
  customers[id].last_service=today;
  // Default next service in 30 days for recurring
  const p=P.find(x=>x.id===id);
  if(p&&p.status==='customer_recurring'){
    const next=new Date();next.setDate(next.getDate()+30);
    customers[id].next_service=next.toISOString().slice(0,10);
  }
  custSave();rCust();
  toast('Service visit logged for '+today);
}

function setNextService(id){
  const d=prompt('Set next service date (YYYY-MM-DD):',new Date(Date.now()+30*864e5).toISOString().slice(0,10));
  if(!d||!/^\d{4}-\d{2}-\d{2}$/.test(d)){toast('Invalid date format');return;}
  if(!customers[id])customers[id]={};
  customers[id].next_service=d;
  custSave();rCust();
  toast('Next service set for '+d);
}

// ── SETTINGS ─────────────────────────────────────────────────────────────────
function loadSettings(){
  try{return JSON.parse(localStorage.getItem('pic_settings')||'{}')||{};}catch(e){return {};}
}
function saveSettings(){
  const portal=(document.getElementById('hs-portal')||{}).value||'';
  try{localStorage.setItem('pic_settings',JSON.stringify({hubspot_portal:portal}));}catch(e){}
}
function initSettings(){
  const s=loadSettings();
  const el=document.getElementById('hs-portal');
  if(el&&s.hubspot_portal)el.value=s.hubspot_portal;
}

// ── QUOTE BUILDER ─────────────────────────────────────────────────────────────
function showQuote(p){
  if(!p)return;
  const isRecurring=p.status==='customer_recurring';
  const d=new Date();
  const dateStr=d.toLocaleDateString('en-US',{month:'long',day:'numeric',year:'numeric'});
  const validUntil=new Date(d.getTime()+30*864e5).toLocaleDateString('en-US',{month:'long',day:'numeric',year:'numeric'});

  document.getElementById('quote-content').innerHTML=
    '<div style="border:2px solid var(--brd);border-radius:10px;padding:16px;margin-bottom:10px">'
    // Header
    +'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:16px;padding-bottom:12px;border-bottom:1px solid var(--brd2)">'
    +'<div><div style="font-size:18px;font-weight:800;color:var(--navy)">Pinellas Ice Co</div>'
    +'<div style="font-size:10px;color:var(--sub)">Commercial Ice Machine Cleaning &amp; Sanitizing</div>'
    +'<div style="font-size:10px;color:var(--sub)">pinellasiceco.com</div></div>'
    +'<div style="text-align:right"><div style="font-size:11px;font-weight:700;color:var(--sub)">QUOTE</div>'
    +'<div style="font-size:10px;color:var(--sub)">Date: '+dateStr+'</div>'
    +'<div style="font-size:10px;color:var(--sub)">Valid: '+validUntil+'</div></div>'
    +'</div>'
    // Client
    +'<div style="margin-bottom:14px;padding:10px;background:#f5f8fa;border-radius:7px">'
    +'<div style="font-size:9px;color:var(--sub);font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Prepared For</div>'
    +'<div style="font-weight:700;font-size:14px;color:var(--navy)">'+p.name+'</div>'
    +'<div style="font-size:11px;color:var(--sub)">'+p.address+', '+p.city+', FL '+p.zip+'</div>'
    +(p.phone?'<div style="font-size:11px;color:var(--sub)">'+p.phone+'</div>':'')
    +'</div>'
    // Services
    +'<div style="margin-bottom:14px">'
    +'<div style="font-size:9px;color:var(--sub);font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Services</div>'
    // Option 1 - Recurring (highlighted if won)
    +'<div style="padding:10px;border:2px solid '+(isRecurring?'var(--grn)':'var(--brd)')+';border-radius:8px;margin-bottom:6px;background:'+(isRecurring?'#ecfdf5':'#fff')+'">'
    +'<div style="display:flex;justify-content:space-between;align-items:center">'
    +'<div><div style="font-weight:700;font-size:12px;color:var(--navy)">Monthly Maintenance Plan'+(isRecurring?' ✓ SELECTED':' (Recommended)')+'</div>'
    +'<div style="font-size:10px;color:var(--sub)">Full clean, sanitize &amp; documentation every visit</div>'
    +'<div style="font-size:10px;color:var(--sub)">'+p.machines+' machine'+(p.machines>1?'s':'')+' covered &bull; Priority scheduling &bull; Compliance report included</div>'
    +'</div>'
    +'<div style="text-align:right"><div style="font-size:20px;font-weight:800;color:'+(isRecurring?'#059669':'var(--navy)')+'">$'+p.monthly+'</div>'
    +'<div style="font-size:9px;color:var(--sub)">/month</div></div></div>'
    +'<div style="font-size:10px;color:var(--sub);margin-top:4px">First service: $'+(Math.round(p.monthly*0.7))+' (30% off based on machine condition assessment)</div>'
    +'</div>'
    // Option 2 - One-time
    +'<div style="padding:10px;border:2px solid '+(!isRecurring&&p.status==='customer_once'?'var(--blu)':'var(--brd)')+';border-radius:8px;background:'+((!isRecurring&&p.status==='customer_once')?'#eff6ff':'#fff')+'">'
    +'<div style="display:flex;justify-content:space-between;align-items:center">'
    +'<div><div style="font-weight:700;font-size:12px;color:var(--navy)">One-Time Deep Clean'+(!isRecurring&&p.status==='customer_once'?' ✓ SELECTED':'')+'</div>'
    +'<div style="font-size:10px;color:var(--sub)">Complete cleaning, sanitizing &amp; inspection report</div>'
    +'<div style="font-size:10px;color:var(--sub)">'+p.machines+' machine'+(p.machines>1?'s':'')+' &bull; Same-day availability</div></div>'
    +'<div style="text-align:right"><div style="font-size:20px;font-weight:800;color:var(--navy)">$'+p.onetime+'</div>'
    +'<div style="font-size:9px;color:var(--sub)">one-time</div></div></div>'
    +'</div></div>'
    // Why
    +'<div style="padding:10px;background:#fef9ee;border-radius:7px;border:1px solid #fde68a;margin-bottom:10px">'
    +'<div style="font-size:9px;font-weight:700;color:#92400e;margin-bottom:4px">WHY THIS MATTERS FOR YOUR BUSINESS</div>'
    +'<div style="font-size:10px;color:#92400e">Florida DBPR inspectors are trained to target ice machine interiors. Mold, slime, and mineral scale are the most common critical violations cited. A documented service record shows inspectors you take compliance seriously and can prevent costly callbacks and closures.</div>'
    +'</div>'
    // Footer
    +'<div style="font-size:9px;color:var(--sub);text-align:center;border-top:1px solid var(--brd2);padding-top:8px">'
    +'Pinellas Ice Co &bull; Licensed &amp; Insured &bull; Service Report Provided After Every Visit'
    +'</div>'
    +'</div>';

  document.getElementById('quote-bg').style.display='flex';
}

function printQuote(){
  const content=document.getElementById('quote-content').innerHTML;
  const win=window.open('','_blank');
  win.document.write('<html><head><title>Quote - Pinellas Ice Co</title>'
    +'<style>body{font-family:Lexend,sans-serif;padding:20px;max-width:600px;margin:0 auto;color:#33475b}'
    +'*{box-sizing:border-box}@media print{button{display:none}}</style></head>'
    +'<body>'+content+'<br><button onclick="window.print()">Print / Save PDF</button></body></html>');
  win.document.close();
  setTimeout(()=>win.print(),500);
}

function copyStripeLink(){
  // Generates a Stripe payment link URL
  // Replace with your actual Stripe payment links
  const p=cur;
  if(!p){toast('Open a prospect first');return;}
  const isRecurring=p.status==='customer_recurring';
  // These are placeholder Stripe links - update with real ones in Settings
  const link=isRecurring
    ?'https://buy.stripe.com/your_recurring_link?amount='+p.monthly+'&currency=usd'
    :'https://buy.stripe.com/your_onetime_link?amount='+p.onetime+'&currency=usd';
  navigator.clipboard.writeText(link).then(()=>toast('Payment link copied!')).catch(()=>{
    prompt('Copy this payment link:',link);
  });
}

// ── CONTACTS ─────────────────────────────────────────────────────────────────
let contacts = {};  // {bizId: [{name, role, phone}]}

function contactsLoad(){
  try{contacts=JSON.parse(localStorage.getItem('pic_contacts')||'{}')||{};}catch(e){contacts={};}
}
function contactsSave(){
  try{localStorage.setItem('pic_contacts',JSON.stringify(contacts));}catch(e){}
}

function renderContacts(id){
  const el=document.getElementById('mcontacts');
  if(!el)return;
  const list=contacts[id]||[];
  if(!list.length){el.innerHTML='<div style="font-size:10px;color:var(--sub);margin-bottom:6px">No contacts saved yet</div>';return;}
  const ROLE_LABELS={owner:'Owner',gm:'GM',manager:'Manager',chef:'Chef/Kitchen',staff:'Staff'};
  el.innerHTML=list.map((c,i)=>
    '<div style="display:flex;align-items:center;gap:8px;padding:6px 0;border-bottom:1px solid var(--brd2)">'
    +'<div style="flex:1">'
      +'<div style="font-size:11px;font-weight:600;color:var(--navy)">'+c.name+'</div>'
      +'<div style="font-size:9px;color:var(--sub)">'+( ROLE_LABELS[c.role]||c.role)+(c.phone?' &bull; '+c.phone:'')+'</div>'
    +'</div>'
    +(c.phone?'<a href="tel:'+c.phone.replace(/\s/g,'')+'" style="font-size:9px;padding:3px 7px;border:1px solid var(--blu);border-radius:5px;color:var(--blu);text-decoration:none" onclick="event.stopPropagation()">Call</a>':'')
    +'<button onclick="deleteContact('+id+','+i+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer">✕</button>'
    +'</div>'
  ).join('');
}

function addContact(){
  if(!cur)return;
  const name=document.getElementById('mc-name').value.trim();
  const role=document.getElementById('mc-role').value;
  const phone=document.getElementById('mc-phone').value.trim();
  if(!name){toast('Enter a contact name');return;}
  if(!contacts[cur.id])contacts[cur.id]=[];
  contacts[cur.id].push({name,role,phone});
  contactsSave();
  document.getElementById('mc-name').value='';
  document.getElementById('mc-phone').value='';
  renderContacts(cur.id);
  toast('Contact saved');
}

function deleteContact(bizId,idx){
  if(!contacts[bizId])return;
  contacts[bizId].splice(idx,1);
  if(!contacts[bizId].length)delete contacts[bizId];
  contactsSave();
  renderContacts(bizId);
}

// ── INIT ─────────────────────────────────────────────────────────────────────
// INIT
function init(){lLoad();phLoad();custLoad();contactsLoad();initSettings();setPreset('all');rT();}
if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',init);}else{setTimeout(init,50);}
if('serviceWorker' in navigator){
  window.addEventListener('load',()=>{
    navigator.serviceWorker.register('sw.js').catch(()=>{});
  });
}
</script>
</body>
</html>
"""

# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────
def main():
    folder = Path(__file__).parent

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if Path(arg).is_dir():
            # Directory passed (e.g. "data/") -- find all CSV and XLSX files
            data_dir = Path(arg)
            csv_paths = (
                sorted(data_dir.glob('*.csv')) +
                sorted(data_dir.glob('*.xlsx')) +
                sorted(data_dir.glob('*.xls'))
            )
            # Exclude the license extract from inspection data
            csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                         and 'hrfood' not in p.name.lower()]
            if not csv_paths:
                print(f"\nNo inspection data files found in {data_dir}")
                sys.exit(1)
            print(f"Found {len(csv_paths)} file(s) in {data_dir}:")
            for p in csv_paths:
                print(f"  {p.name}")
            print()
        else:
            csv_paths = sys.argv[1:]
    else:
        # Auto-find in current folder
        csv_paths = (
            sorted(folder.glob('*.csv')) +
            sorted(folder.glob('*.xlsx'))
        )
        csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                     and 'hrfood' not in p.name.lower()]
        if not csv_paths:
            print("\nNo data files found. Run: python build.py data/")
            sys.exit(1)
        print(f"Auto-found {len(csv_paths)} file(s):")
        for p in csv_paths:
            print(f"  {p.name}")
        print()

    records = run(list(map(str, csv_paths)))
    print(f"\nGenerating HTML...")
    html = build_html(records)
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(html, encoding='utf-8')
    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"  Written: {OUTPUT_FILE.name} ({size_kb}KB)")
    print(f"\n{'='*55}")
    print(f"  Done! Open prospecting_tool.html in Chrome.")
    print(f"  Your call log carries over automatically.")
    print(f"{'='*55}\n")

if __name__ == '__main__':
    main()
