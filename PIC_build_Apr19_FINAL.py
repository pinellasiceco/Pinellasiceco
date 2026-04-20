#!/usr/bin/env python3
"""
Pinellas Ice Co \u2014 Prospect Tool Builder
Complete, self-contained. Drop in any folder with your CSV files and run.
"""

import sys, os, json, re, warnings, csv
from pathlib import Path
from datetime import date, timedelta
from math import radians, cos, sin, sqrt, atan2
from collections import Counter

warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
TARGET_COUNTIES  = ['Pinellas', 'Hillsborough', 'Pasco',
                    'Citrus', 'Hernando', 'Polk', 'Sumter']
MIN_SCORE        = 5   # Very low - include everything, filter in browser
TODAY            = date.today()
OUTPUT_FILE      = Path(__file__).parent / 'prospecting_tool.html'

# ── CHAIN CLASSIFICATION ──────────────────────────────────────────────────────
# Corporate chains: national vendor contracts, skip entirely
CORPORATE_CHAINS = [
    'mcdonald','burger king','wendy','chick-fil-a','starbucks','chipotle',
    'ihop','waffle house','olive garden','red lobster','outback steakhouse',
    'buffalo wild wings','panera bread','cracker barrel','texas roadhouse',
    'golden corral','panda express','jack in the box','raising cane',
    'shake shack','wingstop','habit burger',
]

# Franchise brands: independently owned, CAN be prospects
FRANCHISE_BRANDS = [
    'subway','taco bell','dunkin','pizza hut','domino','papa john',
    'little caesar','sonic','dairy queen','five guys','popeyes','kfc',
    'chilis','applebee','denny','hooters','longhorn','buffalo wild wing',
    'jersey mike','jimmy john','firehouse subs','mcalister',
    'tijuana flats','miller ale house',
]

def classify_business(name):
    """Returns ('corporate','franchise','independent') + is_chain bool."""
    n = name.lower()
    for c in CORPORATE_CHAINS:
        if c in n: return 'corporate', True
    for f in FRANCHISE_BRANDS:
        if f in n: return 'franchise', True
    return 'independent', False

# ── MACHINE ESTIMATION ────────────────────────────────────────────────────────
def est_machines(seats, is_full_bar, rank_code):
    """Estimate ice machine count from seat count + license type."""
    if rank_code in ('NOST','CNOSEAT','MFDV'): return 1
    if seats <= 0:    base = 1
    elif seats < 40:  base = 1
    elif seats < 80:  base = 1
    elif seats < 150: base = 2
    elif seats < 250: base = 2
    elif seats < 400: base = 3
    else:             base = 4
    bonus = 1 if is_full_bar else 0
    return min(base + bonus, 6)

def est_monthly(machines):
    """Tiered: $149 first, $89 second (+$238), $69 each additional."""
    if machines <= 1: return 149
    if machines == 2: return 238
    return 238 + (machines - 2) * 69

def est_onetime(machines):
    """One-time deep clean price."""
    if machines <= 1: return 249
    if machines == 2: return 378
    return 378 + (machines - 2) * 99

def account_tier(seats, rank_code, machines, chronic, confirmed):
    """Account quality tier."""
    if rank_code in ('NOST','CNOSEAT') and not confirmed and not chronic: return 'COLD'
    if seats > 0 and seats < 15 and not confirmed and not chronic: return 'COLD'
    if machines >= 3 and (chronic or confirmed): return 'PLATINUM'
    if machines >= 2 and (chronic or confirmed): return 'GOLD'
    if machines >= 2:                            return 'SILVER'
    if confirmed and chronic:                    return 'GOLD'
    if confirmed:                                return 'SILVER'
    if chronic:                                  return 'SILVER'
    if seats >= 60:                              return 'BRONZE'
    return 'COLD'

def confidence_score(n_inspections, n_years_data, has_ice_history, days_since):
    """
    Confidence in the scoring/prediction (0-100).
    Higher with more inspection history, recent data, ice violation history.
    """
    base = 30
    # More inspections = more confident
    if n_inspections >= 6:   base += 25
    elif n_inspections >= 3: base += 15
    elif n_inspections >= 2: base += 8
    # More years of data
    if n_years_data >= 4:    base += 20
    elif n_years_data >= 2:  base += 10
    # Ice history is a strong signal
    if has_ice_history:      base += 15
    # Recent inspection = more reliable
    if days_since <= 30:     base += 10
    elif days_since <= 90:   base += 5
    return min(100, base)

def seat_score_bonus(machines, rank_code):
    if rank_code in ('NOST','CNOSEAT'): return 0
    if machines >= 3: return 15
    if machines >= 2: return 10
    return 0


# ──────────────────────────────────────────────────────────────────────────────
# STEP 0: CHECK DEPENDENCIES
# ──────────────────────────────────────────────────────────────────────────────
def check_deps():
    missing = []
    for pkg, install in [('pandas','pandas'), ('sklearn','scikit-learn'), 
                          ('numpy','numpy'), ('openpyxl','openpyxl')]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(install)
    if missing:
        print("\n" + "="*55)
        print("  Missing packages. Run this once in terminal:")
        print(f"  pip install {' '.join(missing)}")
        print("="*55 + "\n")
        sys.exit(1)

check_deps()

import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN NORMALIZATION
# ──────────────────────────────────────────────────────────────────────────────
COLUMN_MAP = {
    'Inspection Date': ['InspectionDate','INSPECTION DATE','inspection_date','Date','Insp Date'],
    'License ID': ['LicenseID','LICENSE ID','license_id','License Number','LicenseNumber','Lic ID','License_ID'],
    'Business (DBA-Does Business As) Name': [
        'Business Name','DBA Name','DBA','Business','BUSINESS NAME',
        'DBA-Does Business As','Restaurant Name','RestaurantName','Establishment Name'],
    'County Name': ['County','COUNTY','county_name','CountyName'],
    'Location Address': ['Address','LOCATION ADDRESS','Street Address','StreetAddress','Addr'],
    'Location City': ['City','CITY','city_name'],
    'Location Zip Code': ['Zip','ZIP','Zip Code','ZipCode','Postal Code','PostalCode','zip_code'],
    'Inspection Type': ['InspectionType','INSPECTION TYPE','Type','Insp Type'],
    'Inspection Disposition': ['Disposition','DISPOSITION','InspectionDisposition','Result'],
    'Number of High Priority Violations': [
        'High Priority Violations','HighPriorityViolations','High Violations','HP Violations','HV'],
    'Number of Total Violations': ['Total Violations','TotalViolations','Total','TV'],
    'Number of Intermediate Violations': ['Intermediate Violations','IntermediateViolations','IV'],
    'Number of Basic Violations': ['Basic Violations','BasicViolations','BV'],
    'Visit Number': ['VisitNumber','Visit','VISIT NUMBER','Visit_Number'],
}

DISP_RISK = {
    'Inspection Completed - No Further Action': 0, 'Warning Issued': 1,
    'Call Back - Complied': 2, 'Call Back - Extension given, pending': 3,
    'Admin. Complaint Callback Complied': 3, 'Call Back - Admin. complaint recommended': 4,
    'Administrative complaint recommended': 4, 'Administrative determination recommended': 4,
    'Emergency Order Callback Complied': 3, 'Emergency Order Callback Time Extension': 4,
    'Emergency Order Callback Not Complied': 5, 'Emergency order recommended': 5,
}

ICE_CODES = {14: 100, 22: 80, 50: 70, 23: 60, 37: 50, 51: 30, 36: 20, 55: 15}

ZIP_COORDS = {
    '33511':(27.924,-82.312),'33547':(27.861,-82.212),'33570':(27.714,-82.398),
    '33572':(27.752,-82.378),'33573':(27.710,-82.354),'33578':(27.855,-82.324),
    '33579':(27.793,-82.278),'33584':(27.992,-82.269),'33592':(28.096,-82.280),
    '33594':(27.910,-82.232),'33596':(27.875,-82.258),'33598':(27.660,-82.349),
    '33602':(27.949,-82.459),'33603':(27.980,-82.468),'33604':(28.001,-82.468),
    '33605':(27.956,-82.427),'33606':(27.931,-82.474),'33607':(27.966,-82.489),
    '33609':(27.945,-82.495),'33610':(27.978,-82.406),'33611':(27.900,-82.508),
    '33612':(28.022,-82.457),'33613':(28.055,-82.448),'33614':(27.997,-82.509),
    '33615':(27.989,-82.563),'33616':(27.894,-82.532),'33617':(28.012,-82.407),
    '33618':(28.053,-82.499),'33619':(27.950,-82.389),'33620':(28.062,-82.415),
    '33621':(27.860,-82.520),'33624':(28.069,-82.535),'33625':(28.080,-82.556),
    '33626':(28.082,-82.585),'33629':(27.924,-82.522),'33634':(27.993,-82.567),
    '33635':(28.010,-82.609),'33637':(28.056,-82.380),'33647':(28.119,-82.378),
    '33701':(27.770,-82.634),'33702':(27.811,-82.647),'33703':(27.817,-82.633),
    '33704':(27.798,-82.640),'33705':(27.750,-82.638),'33706':(27.747,-82.747),
    '33707':(27.761,-82.717),'33708':(27.810,-82.800),'33709':(27.806,-82.754),
    '33710':(27.785,-82.734),'33711':(27.745,-82.682),'33712':(27.726,-82.663),
    '33713':(27.779,-82.669),'33714':(27.803,-82.685),'33715':(27.703,-82.717),
    '33716':(27.855,-82.652),'33755':(27.966,-82.800),'33756':(27.953,-82.801),
    '33759':(27.974,-82.750),'33760':(27.900,-82.687),'33761':(28.014,-82.740),
    '33762':(27.870,-82.676),'33763':(28.005,-82.748),'33764':(27.934,-82.740),
    '33765':(27.984,-82.777),'33767':(27.978,-82.828),'33770':(27.910,-82.789),
    '33771':(27.912,-82.762),'33772':(27.878,-82.795),'33773':(27.880,-82.720),
    '33774':(27.869,-82.833),'33776':(27.839,-82.815),'33777':(27.854,-82.754),
    '33778':(27.873,-82.740),'33781':(27.841,-82.717),'33782':(27.863,-82.697),
    '33785':(27.740,-82.781),'33786':(27.924,-82.826),'34677':(28.053,-82.669),
    '34681':(28.011,-82.688),'34683':(28.080,-82.765),'34684':(28.102,-82.769),
    '34685':(28.117,-82.731),'34688':(28.128,-82.704),'34689':(28.146,-82.756),
    '34695':(27.986,-82.695),'34698':(28.019,-82.775),'34652':(28.240,-82.727),
    '34653':(28.262,-82.724),'34654':(28.279,-82.666),'34655':(28.218,-82.673),
    '34667':(28.338,-82.652),'34668':(28.280,-82.697),'34669':(28.226,-82.706),
    '34690':(28.243,-82.675),'34691':(28.195,-82.754),'34638':(28.190,-82.712),
    '34637':(28.214,-82.666),'34639':(28.234,-82.628),
    '34429':(28.896,-82.580),'34442':(28.767,-82.472),'34448':(28.956,-82.582),
    '34450':(28.836,-82.330),'34452':(28.814,-82.335),'34465':(28.947,-82.469),
    '34601':(28.552,-82.388),'34604':(28.538,-82.466),'34606':(28.468,-82.549),
    '34608':(28.506,-82.540),'34609':(28.441,-82.480),'34610':(28.369,-82.547),
    '34613':(28.527,-82.517),'33801':(28.041,-81.975),'33803':(28.020,-81.970),
    '33810':(28.059,-82.036),'33811':(27.945,-82.028),'33813':(27.969,-82.001),
    '33823':(28.071,-81.897),'33830':(27.899,-81.840),'33880':(28.023,-81.731),
    '33897':(28.291,-81.643),'32162':(28.950,-81.966),'34785':(28.839,-82.051),
}

HIGH_ICE = ['bar','pub','tavern','lounge','grill','grille','sports','club',
            'seafood','sushi','hibachi','oyster','marina','brewery','taproom',
            'resort','hotel','inn','cantina','steakhouse']
LOW_ICE  = ['bakery','donut','coffee','smoothie','juice','dessert','cupcake']

# Median days until callback inspection per disposition type.
# Derived from 25k+ FL DBPR inspection interval records — data-driven, not guesses.
CALLBACK_DAYS = {
    'Emergency order recommended':               1,
    'Emergency Order Callback Not Complied':     1,
    'Administrative complaint recommended':      4,
    'Administrative determination recommended':  4,
    'Call Back - Admin. complaint recommended':  22,
    'Admin. Complaint Callback Complied':        45,
    'Warning Issued':                            12,
    'Call Back - Extension given, pending':      50,
    'Emergency Order Callback Time Extension':   62,
    'Emergency Order Callback Complied':         64,
    'Call Back - Complied':                      118,
    'Inspection Completed - No Further Action':  130,
}
STATIC_PHONES = {
    "9784166": "+1 813-993-3924",
    "2183411": "+1 813-752-2236",
    "9681679": "+1 813-657-5600",
    "9468965": "+1 813-374-5363",
    "9473519": "+1 727-314-4004",
    "9405890": "+1 813-909-6354",
    "9428645": "+1 813-265-2111",
    "9435298": "+1 813-295-8450",
    "9317331": "+1 727-242-8400",
    "9394343": "+1 813-344-3311",
    "9283918": "+1 813-815-7662",
    "9296161": "+1 727-381-0088",
    "4115014": "+1 727-844-3125",
    "3754544": "+1 727-924-2000",
    "3774587": "+1 727-586-5797",
    "3243192": "+1 727-443-4900",
    "3425752": "+1 813-654-4262",
    "3523975": "+1 813-825-1373",
    "9256830": "+1 727-270-6655",
    "9155471": "+1 813-994-9666",
    "4421753": "+1 813-961-4092",
    "7089467": "+1 727-726-4608",
    "2976346": "+1 813-374-2739",
    "6462895": "+1 727-934-4047",
    "2238952": "+1 727-327-8090",
    "5654752": "+1 727-863-0965",
    "5787421": "+1 727-614-0574",
    "7432888": "+1 727-954-7369"
}

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def normalize_columns(df):
    col_lower = {c.lower().strip(): c for c in df.columns}
    renames = {}
    for standard, variants in COLUMN_MAP.items():
        if standard in df.columns:
            continue
        for v in variants:
            if v.lower().strip() in col_lower:
                renames[col_lower[v.lower().strip()]] = standard
                break
    # Violation column variants: V14, Viol14, violation_14 → Violation 14
    viol_pat = re.compile(r'^[Vv](?:iol(?:ation)?[\s_-]?)?0?(\d+)$')
    for col in list(df.columns):
        m = viol_pat.match(col.strip())
        if m:
            new = f'Violation {int(m.group(1)):02d}'
            if new != col and col not in renames:
                renames[col] = new
    if renames:
        df = df.rename(columns=renames)
    return df

def load_csvs(paths):
    frames, ref_cols = [], None
    for p in paths:
        path = Path(p)
        if not path.exists():
            print(f"  WARNING: {path.name} not found, skipping")
            continue
        try:
            ext = path.suffix.lower()
            if ext in ('.xlsx', '.xlsm', '.xls'):
                # Stream XLSX row-by-row to handle large statewide files
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                headers = [str(c).strip() if c is not None else '' for c in rows[0]]
                data = [['' if v is None else str(v) for v in row] for row in rows[1:]]
                df = pd.DataFrame(data, columns=headers)
                print(f"  {path.name}: {len(df):,} rows (xlsx)")
            else:
                df = pd.read_csv(path, low_memory=False, encoding='utf-8',
                                 encoding_errors='replace')
                print(f"  {path.name}: {len(df):,} rows")

            df.columns = df.columns.str.strip()
            df = normalize_columns(df)
            if ref_cols is None:
                ref_cols = list(df.columns)
            frames.append(df)

        except Exception as e:
            try:
                if ref_cols and path.suffix.lower() == '.csv':
                    df = pd.read_csv(path, header=None, names=ref_cols,
                                     low_memory=False, encoding_errors='replace')
                    df = normalize_columns(df)
                    frames.append(df)
                    print(f"  {path.name}: {len(df):,} rows (headerless)")
                else:
                    print(f"  SKIP {path.name}: {e}")
            except Exception as e2:
                print(f"  SKIP {path.name}: {e2}")

    if not frames:
        raise RuntimeError("No valid data files loaded.")
    return pd.concat(frames, ignore_index=True)

def load_license_extract(data_dir=None):
    """
    Load FL DBPR active license extract.
    Returns dict keyed by License Number with phone, seats, rank, license_type.
    Tries data/ folder first, then current folder.
    """
    import pandas as pd
    candidates = []
    if data_dir:
        candidates.append(Path(data_dir) / 'hrfood3_licenses.csv')
    candidates += [
        Path(__file__).parent / 'data' / 'hrfood3_licenses.csv',
        Path(__file__).parent / 'hrfood3_licenses.csv',
    ]
    for path in candidates:
        if path.exists():
            try:
                df = pd.read_csv(path, low_memory=False, encoding_errors='replace')
                df.columns = df.columns.str.strip()
                # Normalize license number -- strip prefix letters for join
                # DBPR license numbers in extract: "SEA6213532"
                # In inspection data: numeric "6213532" or full "SEA6213532"
                result = {}
                for _, r in df.iterrows():
                    lic_raw = str(r.get('License Number', '')).strip()
                    # Store under both full and numeric forms
                    phone = str(r.get('Primary Phone Number', '') or
                                r.get('Secondary Phone Number', '') or '').strip()
                    phone = re.sub(r'[^0-9]', '', phone)
                    if len(phone) == 10:
                        phone = '+1 ' + phone[:3] + '-' + phone[3:6] + '-' + phone[6:]
                    elif len(phone) == 11 and phone.startswith('1'):
                        phone = '+1 ' + phone[1:4] + '-' + phone[4:7] + '-' + phone[7:]
                    else:
                        phone = ''
                    seats_raw = r.get('Number of Seats or Rental Units', 0)
                    try:
                        seats = int(float(str(seats_raw))) if seats_raw and str(seats_raw).strip() not in ('', 'nan') else 0
                    except:
                        seats = 0
                    entry = {
                        'phone':        phone,
                        'seats':        seats,
                        'rank':         str(r.get('Rank Code', '') or '').strip(),
                        'license_type': str(r.get('License Type Code', '') or '').strip(),
                        'risk_level':   str(r.get('Base Risk Level', '') or '').strip(),
                    }
                    result[lic_raw] = entry
                    # Also index by numeric portion for flexible joins
                    numeric = re.sub(r'[^0-9]', '', lic_raw)
                    if numeric and numeric not in result:
                        result[numeric] = entry
                n_phones = sum(1 for v in result.values() if v.get('phone'))
                print(f"  License extract: {len(df):,} records, {n_phones:,} with phones")
                return result
            except Exception as e:
                print(f"  WARNING: license extract load failed -- {e}")
    print("  License extract not found -- phones/seats from extract unavailable")
    return {}

def match_license(license_id, extract):
    """Flexible license ID matching -- try multiple formats."""
    lid = str(license_id).strip()
    if lid in extract:
        return extract[lid]
    numeric = re.sub(r'[^0-9]', '', lid)
    if numeric in extract:
        return extract[numeric]
    # Try common prefixes
    for prefix in ('SEA', 'NOS', 'MFD'):
        key = prefix + numeric
        if key in extract:
            return extract[key]
    return None

def ice_usage_label(name):
    n = name.lower()
    if any(w in n for w in HIGH_ICE): return 'high'
    if any(w in n for w in LOW_ICE):  return 'low'
    return 'medium'

def ice_insight_text(codes):
    if 'V14' in codes: return 'Food contact surfaces (V14) \u2014 inspectors directly cited food contact equipment. Ice machines are the primary target of this violation code.'
    if 'V22' in codes: return 'Non-PHF food contact surfaces (V22) \u2014 ice machines fall directly into this category.'
    if 'V50' in codes: return 'Food contact surfaces not clean (V50) \u2014 same root issue as V14/V22, cited at basic level.'
    if 'V37' in codes: return 'Equipment not in good repair (V37) \u2014 ice machine may be broken or malfunctioning.'
    if 'V23' in codes: return 'Utensil/container sanitation (V23) \u2014 includes ice scoops and bin components.'
    return ''

def _callback_fields(disposition: str, last_insp_date) -> dict:
    """Compute estimated callback inspection date from disposition type."""
    cb_days = CALLBACK_DAYS.get(disposition, 30)
    est_callback = last_insp_date + timedelta(days=cb_days)
    days_to_callback = (TODAY - est_callback).days * -1
    urgency = ('overdue'   if days_to_callback < 0  else
               'imminent'  if days_to_callback <= 7  else
               'soon'      if days_to_callback <= 21 else 'upcoming')
    return {
        'est_callback_date': est_callback.isoformat(),
        'days_to_callback':  days_to_callback,
        'callback_urgency':  urgency,
    }

def load_emergency_closures(data_dir=None):
    """Load recent DBPR emergency closure reports. Returns set of license IDs."""
    import openpyxl
    closed_ids = set()
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        if not d.exists(): continue
        for f in sorted(d.glob('EOS_Weekly_Extract_*.xlsx'), reverse=True)[:4]:
            try:
                wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows: continue
                headers = [str(h).strip() if h else '' for h in rows[0]]
                lic_col = next((i for i,h in enumerate(headers)
                                if 'license' in h.lower()), None)
                if lic_col is None: continue
                for row in rows[1:]:
                    val = row[lic_col]
                    if val:
                        num = re.sub(r'[^0-9]', '', str(val))
                        if num: closed_ids.add(num)
            except Exception:
                pass
    if closed_ids:
        print(f"  Emergency closures: {len(closed_ids)} businesses flagged")
    return closed_ids

def nominatim_geocode(address, city, state='FL', cache={}):
    """Geocode a single address using Nominatim (free, no key required).
    Uses in-memory cache to avoid duplicate requests."""
    import time, urllib.request, urllib.parse
    key = f"{address},{city},{state}"
    if key in cache:
        return cache[key]
    try:
        query = urllib.parse.urlencode({'q': f"{address}, {city}, {state}",
                                        'format': 'json', 'limit': 1})
        url = f"https://nominatim.openstreetmap.org/search?{query}"
        req = urllib.request.Request(url,
            headers={'User-Agent': 'PinellasIceCo/1.0 ice machine inspection tool'})
        with urllib.request.urlopen(req, timeout=5) as r:
            results = json.loads(r.read())
        if results:
            lat = float(results[0]['lat'])
            lon = float(results[0]['lon'])
            cache[key] = (lat, lon)
            time.sleep(1.1)  # Nominatim rate limit: 1 req/sec
            return (lat, lon)
    except Exception:
        pass
    cache[key] = None
    return None

def load_osm_cache(data_dir):
    """Load OSM phone/hours cache from download_data.py output."""
    cache_path = Path(data_dir) / 'osm_phones.json' if data_dir else None
    if not cache_path or not cache_path.exists():
        # Try default data/ folder
        alt = Path(__file__).parent / 'data' / 'osm_phones.json'
        if not alt.exists():
            return {}
        cache_path = alt
    try:
        import re as _re
        raw = json.loads(cache_path.read_text(encoding='utf-8'))
        osm = {}
        for county_data in raw.values():
            for biz in county_data.get('businesses', []):
                name = biz.get('name', '')
                if name and biz.get('phone'):
                    key = _re.sub(r'[^a-z0-9]', '', name.lower())
                    if key not in osm:
                        osm[key] = biz
        print(f"  OSM cache: {len(osm):,} businesses with phones/hours")
        return osm
    except Exception as e:
        print(f"  OSM cache load failed: {e}")
        return {}

def osm_match(name, city, osm_cache):
    """Fuzzy match a business name to OSM cache entry."""
    import re as _re
    if not osm_cache:
        return None
    key = _re.sub(r'[^a-z0-9]', '', name.lower())
    if key in osm_cache:
        return osm_cache[key]
    words = set(_re.sub(r'[^a-z0-9 ]', ' ', name.lower()).split()) - {
        'llc','inc','the','and','bar','grill','restaurant','cafe','pub'}
    if len(words) < 2:
        return None
    best_score, best_match = 0.7, None
    for cached_biz in osm_cache.values():
        cw = set(_re.sub(r'[^a-z0-9 ]', ' ', cached_biz.get('name','').lower()).split())
        overlap = len(words & cw) / max(len(words), len(cw), 1)
        if overlap > best_score:
            cc = (cached_biz.get('city') or '').lower()
            if not cc or city.lower()[:5] in cc or cc[:5] in city.lower():
                best_score = overlap
                best_match = cached_biz
    return best_match

def load_geo_cache(data_dir=None):
    """Load persistent geocoding cache from disk."""
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        p = d / 'geocache.json'
        if p.exists():
            try:
                data = json.loads(p.read_text())
                print(f"  Geocache: {len(data)} cached addresses")
                return data
            except Exception:
                pass
    return {}

def save_geo_cache(cache, data_dir=None):
    """Save geocoding cache to disk for next run."""
    dirs = []
    if data_dir: dirs.append(Path(data_dir))
    dirs.append(Path(__file__).parent / 'data')
    dirs.append(Path(__file__).parent)
    for d in dirs:
        try:
            d.mkdir(exist_ok=True)
            (d / 'geocache.json').write_text(json.dumps(cache))
            return
        except Exception:
            pass

# ──────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ──────────────────────────────────────────────────────────────────────────────
def run(csv_paths):
    print(f"\n{'='*55}")
    print(f"  Pinellas Ice Co \u2014 Building Prospect Tool")
    print(f"  {TODAY}")
    print(f"{'='*55}\n")

    # LOAD
    print("Loading CSV files...")
    df = load_csvs(csv_paths)
    print(f"  Total rows loaded: {len(df):,}\n")

    # LOAD LICENSE EXTRACT (phones + seats + rank)
    print("Loading license extract...")
    data_dir = Path(csv_paths[0]).parent if csv_paths else None
    license_data = load_license_extract(data_dir)

    # LOAD EMERGENCY CLOSURES
    print("Loading emergency closures...")
    emergency_ids = load_emergency_closures(data_dir)

    # LOAD GEO CACHE
    geo_cache = load_geo_cache(data_dir)
    osm_cache = load_osm_cache(data_dir)
    print()

    # COERCE
    df['inspection_date'] = pd.to_datetime(df.get('Inspection Date'), errors='coerce')
    df = df.dropna(subset=['inspection_date'])
    df['county']    = df.get('County Name',          pd.Series('', index=df.index)).astype(str).str.strip()
    df['biz_name']  = df.get('Business (DBA-Does Business As) Name',
                              pd.Series('', index=df.index)).astype(str).str.strip()
    df['address']   = df.get('Location Address',     pd.Series('', index=df.index)).astype(str).str.strip()
    df['city']      = df.get('Location City',        pd.Series('', index=df.index)).astype(str).str.strip()
    df['zip']       = df.get('Location Zip Code',    pd.Series('', index=df.index)).astype(str).str.strip().str[:5]
    df['disp_raw']  = df.get('Inspection Disposition',pd.Series('',index=df.index)).astype(str).str.strip()
    df['insp_type'] = df.get('Inspection Type',      pd.Series('', index=df.index)).astype(str).str.strip()
    df['license_id']= df.get('License ID',           pd.Series(range(len(df)), index=df.index))

    for col, src_name in [('hv','Number of High Priority Violations'),
                           ('tv','Number of Total Violations'),
                           ('iv','Number of Intermediate Violations'),
                           ('bv','Number of Basic Violations'),
                           ('vn','Visit Number')]:
        df[col] = pd.to_numeric(df.get(src_name, pd.Series(0, index=df.index)),
                                errors='coerce').fillna(0)

    # Violation codes
    for code in ICE_CODES:
        col = f'Violation {code:02d}'
        df[col] = pd.to_numeric(df.get(col, pd.Series(0, index=df.index)),
                                errors='coerce').fillna(0)

    df['ice_rel']    = sum(df[f'Violation {c:02d}'].clip(upper=1) * s for c,s in ICE_CODES.items())
    df['direct_ice'] = (df['Violation 14'].clip(upper=1) +
                        df['Violation 22'].clip(upper=1) +
                        df['Violation 50'].clip(upper=1)) > 0
    df['dr'] = df['disp_raw'].map(DISP_RISK).fillna(2)

    # Filter
    df_t = df[df['county'].isin(TARGET_COUNTIES)].copy()
    print(f"Rows in target counties: {len(df_t):,}")

    # Routine inspections
    routine = df_t[df_t['insp_type'] == 'Routine - Food'].copy()
    routine = routine.sort_values(['license_id', 'inspection_date'])
    routine['next_date'] = routine.groupby('license_id')['inspection_date'].shift(-1)
    routine['interval']  = (routine['next_date'] - routine['inspection_date']).dt.days
    rv = routine[(routine['interval'] > 0) & (routine['interval'] < 730)].copy()
    MED = float(rv['interval'].median()) if len(rv) else 125.0
    print(f"Training on {len(rv):,} inspection intervals (median={MED:.0f}d)")

    # Train model
    le = LabelEncoder()
    rv['county_enc'] = le.fit_transform(rv['county'].fillna('Unknown'))
    rv['prev']  = rv.groupby('license_id')['interval'].transform(
        lambda x: x.expanding().mean().shift(1)).fillna(MED)
    rv['month'] = rv['inspection_date'].dt.month
    FEATS = ['hv','iv','bv','tv','dr','vn','month','county_enc','prev']
    rf = RandomForestRegressor(n_estimators=120, max_depth=8, random_state=42, n_jobs=-1)
    rf.fit(rv[FEATS].fillna(0), rv['interval'])
    print("Prediction model trained.\n")

    # Per-establishment history
    ice_hist = routine.groupby('license_id').agg(
        avg_interval  =('interval','mean'),
        n_insp        =('interval','count'),
        avg_hv        =('hv','mean'),
        avg_ice       =('ice_rel','mean'),
        direct_count  =('direct_ice','sum'),
    ).reset_index()
    ice_hist['chronic']   = ice_hist['direct_count'] >= 2
    ice_hist['confirmed'] = ice_hist['direct_count'] >= 1

    # ── NEW SIGNALS ────────────────────────────────────────────────────────────

    # 1. Callback inspection count per business
    #    Businesses with callbacks are actively being monitored — highest urgency
    callback_disps = ['Call Back - Admin. complaint recommended',
                      'Administrative complaint recommended',
                      'Admin. Complaint Callback Complied',
                      'Emergency Order Callback Not Complied',
                      'Emergency order recommended']
    routine['is_callback'] = routine['disp_raw'].isin(callback_disps)
    callback_hist = routine.groupby('license_id')['is_callback'].sum().reset_index()
    callback_hist.columns = ['license_id','n_callbacks']
    ice_hist = ice_hist.merge(callback_hist, on='license_id', how='left')
    ice_hist['n_callbacks'] = ice_hist['n_callbacks'].fillna(0).astype(int)

    # 2. Disposition escalation — did risk level increase over time?
    #    Compare avg disp_risk in first half of history vs second half
    def escalation_score(g):
        if len(g) < 4: return 0
        g = g.sort_values('inspection_date')
        mid = len(g)//2
        early = g.iloc[:mid]['dr'].mean()
        late  = g.iloc[mid:]['dr'].mean()
        return max(0, late - early)  # positive = getting worse
    esc = routine.groupby('license_id').apply(escalation_score)
    esc_df = esc.reset_index()
    esc_df.columns = ['license_id','escalation']
    ice_hist = ice_hist.merge(esc_df, on='license_id', how='left')
    ice_hist['escalation'] = ice_hist['escalation'].fillna(0)

    # 3. Ice violation recency — was the last ice violation recent?
    #    Recent = within 18 months, old = 3+ years ago
    ice_rows = routine[routine['direct_ice'] == True].copy()
    if len(ice_rows):
        last_ice = ice_rows.groupby('license_id')['inspection_date'].max().reset_index()
        last_ice.columns = ['license_id','last_ice_date']
        last_ice['days_since_ice'] = (pd.Timestamp(TODAY) - last_ice['last_ice_date']).dt.days
        ice_hist = ice_hist.merge(last_ice[['license_id','days_since_ice']], on='license_id', how='left')
    else:
        ice_hist['days_since_ice'] = 999
    ice_hist['days_since_ice'] = ice_hist['days_since_ice'].fillna(999).astype(int)
    ice_hist['ice_recent']  = ice_hist['days_since_ice'] <= 365   # within 1 year
    ice_hist['ice_fresh']   = ice_hist['days_since_ice'] <= 180   # within 6 months

    # 4. Violation code diversity — how many different violation types?
    def code_diversity(g):
        codes_seen = set()
        for col in [f'Violation {c:02d}' for c in ICE_CODES if f'Violation {c:02d}' in g.columns]:
            if (g[col]>0).any():
                codes_seen.add(col)
        return len(codes_seen)
    div = routine.groupby('license_id').apply(code_diversity)
    div_df = div.reset_index()
    div_df.columns = ['license_id','code_diversity']
    ice_hist = ice_hist.merge(div_df, on='license_id', how='left')
    ice_hist['code_diversity'] = ice_hist['code_diversity'].fillna(0).astype(int)

    # 5. Failed-first-visit rate — visit_number > 1 means they failed initial
    if 'vn' in routine.columns:
        visit_hist = routine.groupby('license_id')['vn'].agg(
            avg_visit='mean', max_visit='max').reset_index()
        ice_hist = ice_hist.merge(visit_hist, on='license_id', how='left')
        ice_hist['avg_visit'] = ice_hist['avg_visit'].fillna(1)
        ice_hist['has_callbacks_visit'] = ice_hist['avg_visit'] > 1.1
    else:
        ice_hist['avg_visit'] = 1.0
        ice_hist['has_callbacks_visit'] = False

    print(f"  New signals computed:")
    print(f"    Businesses with callbacks: {(ice_hist['n_callbacks']>0).sum():,}")
    print(f"    Businesses with escalation: {(ice_hist['escalation']>0.5).sum():,}")
    print(f"    Ice violation within 1yr: {ice_hist['ice_recent'].sum():,}")
    print(f"    Ice violation within 6mo: {ice_hist['ice_fresh'].sum():,}")
    print(f"    Multi-code violators: {(ice_hist['code_diversity']>=2).sum():,}")
    def trend_slope(g):
        if len(g) < 3: return 0.0
        x = np.arange(len(g), dtype=float)
        return float(np.polyfit(x, g['hv'].values.astype(float), 1)[0])
    trends = routine.groupby('license_id').apply(trend_slope).to_dict()

    # Fired ice codes per establishment
    def get_fired(g):
        codes = []
        for code in ICE_CODES:
            col = f'Violation {code:02d}'
            if col in g.columns and (pd.to_numeric(g[col], errors='coerce').fillna(0) > 0).any():
                codes.append(f'V{code}')
        return codes[:4]
    fired = routine.groupby('license_id').apply(get_fired).to_dict()

    # Predict next inspection date
    latest = routine.sort_values('inspection_date').groupby('license_id').last().reset_index()
    latest = latest.merge(ice_hist, on='license_id', how='left')
    for c in ['avg_interval','avg_hv','avg_ice','direct_count']:
        latest[c] = latest.get(c, 0).fillna(0)
    for c in ['chronic','confirmed']:
        latest[c] = latest.get(c, False).fillna(False).astype(bool)

    cs = latest['county'].where(latest['county'].isin(le.classes_), 'Unknown').fillna('Unknown')
    latest['county_enc'] = le.transform(cs)
    latest['prev']  = latest['avg_interval'].fillna(MED)
    latest['month'] = latest['inspection_date'].dt.month
    Xp = pd.DataFrame({f: latest.get(f, 0) for f in FEATS}).fillna(0)
    latest['pred_days']  = np.clip(rf.predict(Xp).round().astype(int), 7, 730)
    latest['pred_next']  = latest['inspection_date'] + pd.to_timedelta(latest['pred_days'], unit='d')
    latest['days_until'] = (latest['pred_next'] - pd.Timestamp(TODAY)).dt.days.fillna(999).astype(int)

    # Score
    def pitch_type(dr, du, ice, chronic, confirmed, days_since=0):
        if dr >= 4 and days_since <= 120: return 'callback'   # admin complaint/emergency = always callback if recent
        if dr >= 4 and -30 <= du < 75: return 'callback'
        if du < 0 and dr >= 2:         return 'overdue_urgent'
        if du < 0:                     return 'overdue'
        if du <= 21:                   return 'pre_hot'
        if du <= 45:                   return 'pre_warm'
        if (chronic or ice >= 80) and du <= 90: return 'high_risk'
        return 'routine'

    def ice_score(pt, dr, hv, ni, ice, chronic, confirmed,
                  n_callbacks=0, escalation=0, ice_recent=False,
                  ice_fresh=False, code_diversity=0, avg_visit=1.0):
        # Base by pitch type - recalibrated for better spread
        b = {'callback':55,'overdue_urgent':48,'overdue':38,'pre_hot':45,
             'pre_warm':35,'high_risk':30,'routine':15}.get(pt, 15)
        # Ice history - primary differentiator
        if ice_fresh:      b += 20  # ice violation within 6 months = urgent
        elif ice_recent:   b += 12  # ice violation within 1 year
        if chronic:        b += 15  # 2+ inspections with ice violations
        elif confirmed:    b += 8   # 1 ice violation on record
        # Callback signals - inspector already watching this business
        if n_callbacks >= 3: b += 18
        elif n_callbacks == 2: b += 12
        elif n_callbacks == 1: b += 7
        # Escalation - getting systematically worse
        if escalation > 1.5: b += 12
        elif escalation > 0.8: b += 7
        elif escalation > 0.3: b += 3
        # Code diversity - breadth of violation types
        if code_diversity >= 4: b += 8
        elif code_diversity >= 3: b += 5
        elif code_diversity >= 2: b += 2
        # Failed first visits
        if avg_visit > 1.5: b += 6
        elif avg_visit > 1.2: b += 3
        # Violation severity
        if hv >= 6: b += 8
        elif hv >= 4: b += 5
        elif hv >= 2: b += 2
        if (ni or 1) >= 6: b += 5
        elif (ni or 1) >= 3: b += 2
        if dr >= 5: b += 8
        elif dr >= 4: b += 4
        if ice >= 200: b += 6
        elif ice >= 100: b += 3
        return min(100, max(5, b))

    def priority(pt):
        return {'callback':'CALLBACK','overdue_urgent':'CALLBACK',
                'pre_hot':'HOT','pre_warm':'WARM','overdue':'WARM',
                'high_risk':'WATCH'}.get(pt, 'LATER')

    # Score ALL businesses - no caps, no sampling
    # Browser does the filtering
    recent = latest[latest['inspection_date'] >= '2023-07-01'].copy()
    recent['_days_since'] = (pd.Timestamp(TODAY) - recent['inspection_date']).dt.days.fillna(999).astype(int)

    print("Scoring all businesses...")
    pts, scs, prs = [], [], []
    for _, r in recent.iterrows():
        pt = pitch_type(r['dr'], r['days_until'], r['avg_ice'],
                        bool(r['chronic']), bool(r['confirmed']), int(r['_days_since']))
        sc = ice_score(
            pt, r['dr'],
            r['avg_hv'] if pd.notna(r['avg_hv']) else r['hv'],
            r.get('n_insp',1) or 1,
            r['avg_ice'],
            bool(r['chronic']),
            bool(r['confirmed']),
            n_callbacks   = int(r.get('n_callbacks', 0) or 0),
            escalation    = float(r.get('escalation', 0) or 0),
            ice_recent    = bool(r.get('ice_recent', False)),
            ice_fresh     = bool(r.get('ice_fresh', False)),
            code_diversity= int(r.get('code_diversity', 0) or 0),
            avg_visit     = float(r.get('avg_visit', 1.0) or 1.0),
        )
        pts.append(pt); scs.append(sc); prs.append(priority(pt))

    recent['_pt'] = pts
    recent['_sc'] = scs
    recent['_pr'] = prs
    recent['_pinellas'] = (recent['county'] == 'Pinellas').astype(int)
    recent = recent[recent['_sc'] >= MIN_SCORE]
    recent = recent.sort_values(['_sc','_pinellas'], ascending=[False, False])

    result = recent.drop_duplicates('license_id').reset_index(drop=True)
    print(f"  {len(result):,} businesses scored")

    # Rename underscore columns -- itertuples strips leading underscores
    result = result.rename(columns={
        '_sc': 'sc', '_pt': 'pt', '_pr': 'pr',
        '_pinellas': 'pinellas', '_days_since': 'days_since_col'
    })

    print(f"Building records (vectorized)...")
    # Pre-compute all classification in bulk -- avoids slow iterrows
    result['biz_type'] = result['biz_name'].apply(
        lambda n: classify_business(str(n)[:50])[0])
    result = result[result['biz_type'] != 'corporate'].copy()
    print(f"  {len(result):,} non-corporate businesses")

    # Vectorized lat/lon from ZIP
    result['z5'] = result['zip'].astype(str).str[:5]
    result['lat'] = result['z5'].map(lambda z: ZIP_COORDS.get(z, (None,None))[0])
    result['lon'] = result['z5'].map(lambda z: ZIP_COORDS.get(z, (None,None))[1])

    # Vectorized bar detection
    bar_words = ['bar','pub','tavern','lounge','brewery','taproom','cantina',
                 'saloon','sports bar','grill & bar','grille & bar','raw bar']
    result['is_bar'] = result['biz_name'].str.lower().apply(
        lambda n: any(w in str(n) for w in bar_words))

    # Vectorized emergency flag
    result['is_emergency'] = result['license_id'].astype(str).isin(emergency_ids)

    # Vectorized days_since
    result['days_since'] = (pd.Timestamp(TODAY) - result['inspection_date']).dt.days.fillna(999).astype(int)

    # Fill missing columns safely
    for col, default in [('direct_count',0),('avg_ice',0),('n_insp',1),('disp_raw',''),
                          ('chronic',False),('confirmed',False)]:
        if col not in result.columns: result[col] = default
    result['direct_count'] = result['direct_count'].fillna(0).astype(int)
    result['avg_ice']  = result['avg_ice'].fillna(0)
    result['n_insp']   = result['n_insp'].fillna(1).astype(int)
    result['disp_raw'] = result['disp_raw'].fillna('').astype(str)

    # Build records using itertuples (fast)
    records = []
    for row in result.itertuples(index=False):
        try:
            lid   = int(row.license_id)
            name  = str(row.biz_name)[:50]
            codes = fired.get(lid, [])
            z5    = str(row.zip)[:5]

            lic_info  = match_license(lid, license_data) or {}
            phone_raw = lic_info.get('phone','') or STATIC_PHONES.get(str(lid),'')
            hours_raw = ''
            rating_raw = 0
            # OSM enrichment - phone + hours for unmatched records
            if not phone_raw or not hours_raw:
                osm_hit = osm_match(name, str(row.city), osm_cache)
                if osm_hit:
                    if not phone_raw:
                        phone_raw = osm_hit.get('phone','')
                    if not hours_raw:
                        hours_raw = osm_hit.get('hours','')[:60] if osm_hit.get('hours') else ''
            seats     = int(lic_info.get('seats', 0) or 0)
            rank      = lic_info.get('rank', 'SEAT')

            machines    = est_machines(seats, bool(row.is_bar), rank)
            monthly_val = est_monthly(machines)
            onetime_val = est_onetime(machines)
            intro_val   = 99  # flat intro offer regardless of machine count
            confirmed   = bool(row.confirmed)
            chronic     = bool(row.chronic)
            tier        = account_tier(seats, rank, machines, chronic, confirmed)
            days_since  = int(row.days_since)

            lat = row.lat if row.lat and row.lat == row.lat else None
            lon = row.lon if row.lon and row.lon == row.lon else None
            if lat is None:
                cached = geo_cache.get(f"{row.address},{row.city},FL")
                if cached: lat, lon = cached[0], cached[1]

            n_insp      = int(row.n_insp or 1)
            avg_ice     = float(row.avg_ice or 0)
            disp_raw    = str(row.disp_raw or '')
            final_score = min(100, int(row.sc) + seat_score_bonus(machines, rank))
            conf        = confidence_score(n_insp, 1, confirmed, days_since)

            records.append({
                'id':          lid,
                'name':        name,
                'county':      str(row.county),
                'city':        str(row.city),
                'address':     str(row.address),
                'zip':         z5,
                'lat':         lat,
                'lon':         lon,
                'last_insp':   row.inspection_date.strftime('%Y-%m-%d'),
                'last_disp':   disp_raw,
                'disp_risk':   int(row.dr),
                'high_viol':   int(row.hv),
                'total_viol':  int(row.tv),
                'n_insp':      n_insp,
                'pred_next':   row.pred_next.strftime('%Y-%m-%d'),
                'days_until':  int(row.days_until),
                'days_since':  days_since,
                'pitch_type':  str(row.pt),
                'score':       final_score,
                'confidence':  conf,
                'priority':    str(row.pr),
                'ice_rel':     round(avg_ice, 0),
                'confirmed':   confirmed,
                'chronic':     chronic,
                'ice_count':   int(row.direct_count),
                'codes':       codes,
                'trending':    float(trends.get(lid, 0)) > 0.5,
                'is_emergency':bool(row.is_emergency),
                'biz_type':    str(row.biz_type),
                'is_bar':      bool(row.is_bar),
                'seats':       seats,
                'machines':    machines,
                'monthly':     monthly_val,
                'onetime':     onetime_val,
                'intro':       intro_val,
                'tier':        tier,
                'phone':       phone_raw,
                'status':      'prospect',
                'rating':      0,
                'hours':       hours_raw,
                # New intelligence signals
                'n_callbacks':    int(row.n_callbacks) if hasattr(row,'n_callbacks') else 0,
                'escalation':     round(float(row.escalation),2) if hasattr(row,'escalation') else 0,
                'ice_recent':     bool(row.ice_recent) if hasattr(row,'ice_recent') else False,
                'ice_fresh':      bool(row.ice_fresh) if hasattr(row,'ice_fresh') else False,
                'days_since_ice': int(row.days_since_ice) if hasattr(row,'days_since_ice') else 999,
                'code_diversity': int(row.code_diversity) if hasattr(row,'code_diversity') else 0,
                'avg_visit':      round(float(row.avg_visit),1) if hasattr(row,'avg_visit') else 1.0,
            })
        except Exception:
            pass

    PO = {'CALLBACK':0,'HOT':1,'WARM':2,'WATCH':3,'LATER':4}
    records.sort(key=lambda x: (PO.get(x['priority'], 5), -x['score']))

    # Save geocoding cache for future runs
    save_geo_cache(geo_cache, data_dir)

    # Preserve phones from previous run
    if OUTPUT_FILE.exists():
        try:
            content = OUTPUT_FILE.read_text(encoding='utf-8')
            m = re.search(r'const PHONES=(\{.*?\});', content, re.DOTALL)
            if m:
                saved_phones = json.loads(m.group(1))
                n_restored = 0
                for rec in records:
                    ph = saved_phones.get(str(rec['id']), {})
                    if ph.get('phone'):
                        rec['phone']  = ph['phone']
                        rec['rating'] = ph.get('rating', 0)
                        rec['hours']  = ph.get('hours', '')
                        n_restored += 1
                if n_restored:
                    print(f"  Phone data restored for {n_restored} businesses")
        except Exception:
            pass

    pc = dict(Counter(r['priority'] for r in records))
    print(f"\nResults:")
    print(f"  Total prospects:       {len(records)}")
    print(f"  Callback urgent:       {pc.get('CALLBACK',0)}")
    print(f"  Hot (≤21 days):        {pc.get('HOT',0)}")
    print(f"  Warm (22-45 days):     {pc.get('WARM',0)}")
    print(f"  Watch:                 {pc.get('WATCH',0)}")
    print(f"  Chronic ice offenders: {sum(1 for r in records if r.get('chronic'))}")
    print(f"  With map coordinates:  {sum(1 for r in records if r['lat'])}")
    print(f"  With phone number:     {sum(1 for r in records if r['phone'])}")

    return records

# ──────────────────────────────────────────────────────────────────────────────
# HTML
# ──────────────────────────────────────────────────────────────────────────────
def build_html(records):
    data_js   = json.dumps(records, separators=(',',':'))
    phones_js = json.dumps(
        {str(r['id']): {'phone':r['phone'],'rating':r['rating'],'hours':r['hours']}
         for r in records if r['phone']},
        separators=(',',':')
    )
    zip_js = json.dumps({z: list(c) for z,c in ZIP_COORDS.items()}, separators=(',',':'))

    n_cb    = sum(1 for r in records if r['priority']=='CALLBACK')
    n_hot   = sum(1 for r in records if r['priority']=='HOT')
    n_warm  = sum(1 for r in records if r['priority']=='WARM')
    n_phone = sum(1 for r in records if r['phone'])
    n_chron = sum(1 for r in records if r.get('chronic'))
    n_geo   = sum(1 for r in records if r['lat'])

    # Read the HTML template embedded below
    return HTML_TEMPLATE.replace('%%DATA%%', data_js)\
                        .replace('%%PHONES%%', phones_js)\
                        .replace('%%ZIPS%%', zip_js)\
                        .replace('%%DATE%%', str(TODAY))\
                        .replace('%%TOTAL%%', str(len(records)))\
                        .replace('%%NCB%%', str(n_cb))\
                        .replace('%%NHOT%%', str(n_hot))\
                        .replace('%%NWARM%%', str(n_warm))\
                        .replace('%%NPHONE%%', str(n_phone))\
                        .replace('%%NCHRON%%', str(n_chron))\
                        .replace('%%NGEO%%', str(n_geo))\
                        .replace('%%NCONF%%', str(sum(1 for r in records if r['confirmed'])))

# ──────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE  (everything between the triple-quotes)
# ──────────────────────────────────────────────────────────────────────────────

def build_html(records):
    data_js   = json.dumps(records, separators=(',',':'))
    phones_js = json.dumps(
        {str(r['id']): {'phone':r['phone'],'rating':r['rating'],'hours':r['hours']}
         for r in records if r['phone']},
        separators=(',',':')
    )
    zip_js = json.dumps({z: list(c) for z,c in ZIP_COORDS.items()}, separators=(',',':'))

    n_cb    = sum(1 for r in records if r['priority']=='CALLBACK')
    n_hot   = sum(1 for r in records if r['priority']=='HOT')
    n_warm  = sum(1 for r in records if r['priority']=='WARM')
    n_phone = sum(1 for r in records if r['phone'])
    n_chron = sum(1 for r in records if r.get('chronic'))
    n_geo   = sum(1 for r in records if r['lat'])

    # Read the HTML template embedded below
    return HTML_TEMPLATE.replace('%%DATA%%', data_js)\
                        .replace('%%PHONES%%', phones_js)\
                        .replace('%%ZIPS%%', zip_js)\
                        .replace('%%DATE%%', str(TODAY))\
                        .replace('%%TOTAL%%', str(len(records)))\
                        .replace('%%NCB%%', str(n_cb))\
                        .replace('%%NHOT%%', str(n_hot))\
                        .replace('%%NWARM%%', str(n_warm))\
                        .replace('%%NPHONE%%', str(n_phone))\
                        .replace('%%NCHRON%%', str(n_chron))\
                        .replace('%%NGEO%%', str(n_geo))\
                        .replace('%%NCONF%%', str(sum(1 for r in records if r['confirmed'])))

# ──────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE  (everything between the triple-quotes)
# ──────────────────────────────────────────────────────────────────────────────
HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Pinellas Ice Co &#xB7; Prospects</title>
<!-- Leaflet loaded on-demand when Route tab opens -->
<link rel="manifest" href="manifest.json">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="PIC Prospects">
<link rel="apple-touch-icon" href="apple-touch-icon.png">
<meta name="theme-color" content="#2d3e50">
<link href="https://fonts.googleapis.com/css2?family=Lexend:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
:root{
  --bg:#f5f8fa;--surf:#ffffff;--brd:#dfe3eb;--brd2:#eaf0f6;
  --txt:#33475b;--sub:#7c98b6;--sub2:#516f90;
  --cb:#f2545b;--hot:#ff7a59;--warm:#f5c26b;--watch:#00bda5;
  --grn:#00bda5;--blu:#0091ae;--ora:#ff7a59;--navy:#2d3e50;
  --shadow:0 1px 3px rgba(45,62,80,.1),0 1px 8px rgba(45,62,80,.06);
  --shadow-md:0 3px 12px rgba(45,62,80,.12),0 1px 4px rgba(45,62,80,.08);
}
html,body{height:100%;background:var(--bg);color:var(--txt);
  font-family:'Lexend',sans-serif;
  font-size:13px;overflow:hidden;-webkit-font-smoothing:antialiased}
#app{display:flex;flex-direction:column;height:100vh}
header{background:var(--navy);
  padding:0 16px;display:flex;align-items:center;gap:10px;flex-shrink:0;flex-wrap:wrap;height:52px}
.logo{display:flex;align-items:center;gap:8px;flex-shrink:0}
.logo-icon{width:30px;height:30px;border-radius:8px;
  background:var(--ora);
  display:flex;align-items:center;justify-content:center;font-size:16px}
.logo-name{font-weight:700;font-size:14px;color:#fff;letter-spacing:-.02em}
.hchips{display:flex;gap:5px;margin-left:auto;flex-wrap:wrap}
.hc{font-size:10px;padding:3px 10px;border-radius:20px;font-weight:600;cursor:pointer;user-select:none;transition:.15s}
.hc.cb{background:rgba(242,84,91,.2);color:#ffb3b6;border:1px solid rgba(242,84,91,.3)}
.hc.ht{background:rgba(255,122,89,.2);color:#ffcab8;border:1px solid rgba(255,122,89,.3)}
.hc.wm{background:rgba(245,194,107,.2);color:#ffe5a8;border:1px solid rgba(245,194,107,.3)}
.hc.ic{background:rgba(0,189,165,.2);color:#7eeee3;border:1px solid rgba(0,189,165,.3)}
.hc.rt{background:rgba(0,145,174,.2);color:#7dd4e8;border:1px solid rgba(0,145,174,.3)}
.srch{position:relative}
.srch input{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);border-radius:8px;padding:6px 10px 6px 28px;color:#fff;font-size:12px;outline:none;width:160px;font-family:inherit}
.srch input::placeholder{color:rgba(255,255,255,.65)}
.srch input:focus{background:rgba(255,255,255,.18);border-color:rgba(255,255,255,.4)}
.si{position:absolute;left:9px;top:50%;transform:translateY(-50%);color:rgba(255,255,255,.5);font-size:11px;pointer-events:none}
.tabs{display:flex;background:var(--surf);border-bottom:2px solid var(--brd2);flex-shrink:0;
  box-shadow:0 1px 3px rgba(45,62,80,.06)}
.tab{flex:1;padding:10px 4px;text-align:center;font-size:11px;font-weight:600;
  color:var(--sub);cursor:pointer;border-bottom:2px solid transparent;
  margin-bottom:-2px;transition:.15s;user-select:none}
.tab.on{color:var(--ora);border-color:var(--ora)}
.panel{flex:1;overflow-y:auto;padding:14px 16px;display:none;background:var(--bg)}
.panel.on{display:block}
.fbar{display:flex;gap:6px;margin-bottom:12px;flex-wrap:wrap;align-items:center}
.fbar select{background:var(--surf);border:1px solid var(--brd);border-radius:6px;
  padding:5px 9px;color:var(--txt);font-size:11px;outline:none;cursor:pointer;
  font-family:inherit;font-weight:500;box-shadow:var(--shadow)}
.fcnt{font-size:10px;color:var(--sub);margin-left:auto;font-weight:500}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(276px,1fr));gap:10px}
.card{background:var(--surf);border:1px solid var(--brd);border-radius:10px;
  padding:13px 14px;cursor:pointer;touch-action:manipulation;-webkit-tap-highlight-color:transparent;transition:box-shadow .15s,transform .12s,border-color .15s;
  position:relative;overflow:hidden;box-shadow:var(--shadow)}
.card:hover{box-shadow:var(--shadow-md);transform:translateY(-1px);border-color:#c5d1de}
.card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;border-radius:10px 0 0 10px}
.card.CALLBACK::before{background:var(--cb)}
.card.HOT::before{background:var(--hot)}
.card.WARM::before{background:var(--warm)}
.card.WATCH::before{background:var(--watch)}
.card.done{opacity:.45}
.ctop{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:3px}
.cname{font-weight:700;font-size:12.5px;flex:1;padding-right:6px;line-height:1.3;color:var(--navy)}
.pbadge{font-size:9px;font-weight:700;padding:2px 8px;border-radius:20px;letter-spacing:.04em;white-space:nowrap}
.pbadge.CALLBACK{background:#fdeaea;color:#c0392b;border:1px solid #f5c6c6}
.pbadge.HOT{background:#fff2ee;color:#c9542b;border:1px solid #fdd5c6}
.pbadge.WARM{background:#fef9ee;color:#9e6b1a;border:1px solid #fbe8b5}
.pbadge.WATCH{background:#e8faf8;color:#007a6e;border:1px solid #b3ece6}
.cloc{font-size:10px;color:var(--sub);margin-bottom:6px;font-weight:500}
.phrow{display:flex;align-items:center;gap:6px;padding:7px 9px;
  background:#f5f8fa;border-radius:7px;border:1px solid var(--brd2);margin-bottom:6px}
.phnum{font-size:12px;font-weight:600;color:var(--blu);flex:1}
.phnum.none{color:var(--sub);font-size:10px;font-style:italic;font-weight:400}
.abtn{border:none;border-radius:5px;padding:3px 9px;font-size:10px;font-weight:700;
  cursor:pointer;text-decoration:none;display:inline-block;white-space:nowrap;font-family:inherit}
.call-a{background:var(--grn);color:#fff}
.find-a{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7}
.icebadge{font-size:9px;font-weight:700;padding:3px 9px;border-radius:20px;
  margin-bottom:5px;display:inline-block}
.icebadge.chronic{background:#e8faf8;color:#007a6e;border:1px solid #b3ece6}
.icebadge.confirmed{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7}
.cmeta{display:flex;gap:8px;margin-bottom:6px;flex-wrap:wrap}
.mi{display:flex;flex-direction:column;gap:1px}
.ml{font-size:8px;color:var(--sub);letter-spacing:.05em;text-transform:uppercase;font-weight:600}
.mv{font-weight:700;font-size:11px;color:var(--txt)}
.mv.u{color:var(--cb)}.mv.h{color:var(--hot)}.mv.w{color:#9e6b1a}.mv.bad{color:var(--cb)}
.insight{font-size:9px;color:var(--sub2);background:#f5f8fa;border-radius:5px;border:1px solid var(--brd2);
  padding:4px 8px;margin-bottom:5px;line-height:1.5}
.lastc{font-size:10px;color:var(--sub);margin-bottom:6px}
.lastc.hc{color:#007a6e;font-weight:500}
.cacts{display:flex;gap:5px}
.btn{border:none;border-radius:6px;padding:5px 11px;font-size:10px;font-weight:600;cursor:pointer;font-family:inherit}
.blog{background:var(--ora);color:#fff}.blog:hover{background:#e86744}
.bskip{background:transparent;color:var(--sub);border:1px solid var(--brd)}
.brt{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7;font-size:9px}
.ptbl{width:100%;border-collapse:collapse}
.ptbl th{padding:7px 10px;text-align:left;font-size:9px;color:var(--sub2);font-weight:700;
  letter-spacing:.06em;text-transform:uppercase;border-bottom:2px solid var(--brd);
  white-space:nowrap;position:sticky;top:0;background:var(--surf);z-index:1}
.ptbl td{padding:8px 10px;border-bottom:1px solid var(--brd2);vertical-align:middle}
.ptbl tr:hover td{background:#f5f8fa;cursor:pointer}
.ptbl tr.done td{opacity:.4}
.tn{font-weight:600;font-size:11px;max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--navy)}
.ts{font-size:9px;color:var(--sub)}
.tph{font-size:11px;color:var(--blu);font-weight:600;text-decoration:none;white-space:nowrap}
.tph:hover{color:#007a9e}
.tph.none{color:#b0c4d8;font-size:9px}
.td{font-weight:700;font-size:12px}
.td.u{color:var(--cb)}.td.h{color:var(--hot)}.td.w{color:#9e6b1a}
/* ROUTE */
.route-wrap{display:grid;grid-template-columns:280px 1fr;gap:12px;height:calc(100vh - 120px)}
@media(max-width:600px){.route-wrap{grid-template-columns:1fr;grid-template-rows:auto 1fr}}
.rsidebar{display:flex;flex-direction:column;gap:8px;overflow-y:auto}
.rcontrols{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:12px;box-shadow:var(--shadow)}
.rtitle{font-weight:700;font-size:12px;margin-bottom:8px;color:var(--navy)}
.rlist{display:flex;flex-direction:column;gap:4px;overflow-y:auto;max-height:320px}
.rcard{background:var(--surf);border:1px solid var(--brd);border-radius:7px;
  padding:8px 10px;cursor:pointer;transition:.12s;display:flex;align-items:center;gap:7px;
  box-shadow:var(--shadow)}
.rcard:hover{border-color:#c5d1de;box-shadow:var(--shadow-md)}
.rcard.sel{border-color:var(--ora);background:#fff7f5}
.rdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.rname{font-weight:600;font-size:11px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--navy)}
.rdist{font-size:9px;color:var(--sub);flex-shrink:0}
.map-area{background:#e8edf2;border:1px solid var(--brd);border-radius:10px;
  display:flex;align-items:center;justify-content:center;overflow:hidden;position:relative;min-height:300px;
  box-shadow:var(--shadow)}
.map-empty{color:var(--sub);font-size:12px;text-align:center;padding:20px}
.day-route{background:var(--surf);border:1px solid var(--brd);border-radius:10px;
  padding:11px;margin-top:8px;box-shadow:var(--shadow)}
.day-stop{display:flex;align-items:center;gap:7px;padding:6px 0;
  border-bottom:1px solid var(--brd2);font-size:11px}
.day-stop:last-child{border-bottom:none}
.stopnum{width:20px;height:20px;border-radius:50%;background:var(--ora);
  color:#fff;font-weight:700;font-size:9px;display:flex;align-items:center;
  justify-content:center;flex-shrink:0}
.route-btn{background:var(--ora);color:#fff;border:none;border-radius:8px;
  padding:9px;font-size:11px;font-weight:700;cursor:pointer;width:100%;margin-top:8px;font-family:inherit}
.route-btn:hover{background:#e86744}
.route-btn.sec{background:#f5f8fa;color:var(--sub2);border:1px solid var(--brd);margin-top:5px}
/* MODAL */
#mbg{position:fixed;inset:0;background:rgba(45,62,80,.6);z-index:100;
  display:none;backdrop-filter:blur(4px)}
#mbg.on{display:flex;align-items:flex-end;justify-content:center}
@media(min-width:600px){#mbg.on{align-items:center}}
#modal{background:#fff;border:1px solid var(--brd);border-radius:16px 16px 0 0;
  width:100%;max-width:540px;padding:18px;max-height:90vh;overflow-y:auto;
  box-shadow:0 -4px 24px rgba(45,62,80,.15)}
@media(min-width:600px){#modal{border-radius:16px;box-shadow:0 8px 40px rgba(45,62,80,.18)}}
.mh{width:36px;height:4px;background:var(--brd);border-radius:2px;margin:0 auto 14px}
.mname{font-size:17px;font-weight:700;margin-bottom:2px;color:var(--navy)}
.mloc{font-size:10px;color:var(--sub);margin-bottom:12px;font-weight:500}
.mphsec{background:#f5f8fa;border:1px solid var(--brd2);border-radius:10px;padding:12px;margin-bottom:11px}
.mphl{font-size:8px;color:var(--sub);letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px;font-weight:700}
.mphnum{font-size:20px;font-weight:700;color:var(--blu);margin-bottom:8px}
.mphnum.none{font-size:12px;color:var(--sub);font-weight:400;font-style:italic}
.mphacts{display:flex;gap:6px;flex-wrap:wrap}
.mcall{background:var(--grn);color:#fff;border:none;border-radius:8px;padding:8px 16px;
  font-size:11px;font-weight:700;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;gap:5px;font-family:inherit}
.msms{background:#eaf6f9;color:var(--blu);border:1px solid #b3dce7;border-radius:8px;
  padding:8px 12px;font-size:11px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:4px}
.mgoog{background:#f5f8fa;color:var(--sub2);border:1px solid var(--brd);border-radius:8px;
  padding:8px 12px;font-size:11px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:4px}
.mhours{font-size:9px;color:var(--sub);margin-top:5px}
.micebox{background:#e8faf8;border:1px solid #b3ece6;border-radius:9px;padding:11px;margin-bottom:11px}
.msect{font-size:9px;color:var(--sub2);letter-spacing:.07em;text-transform:uppercase;
  margin-bottom:6px;padding-bottom:5px;border-bottom:1px solid var(--brd2);font-weight:700}
.pitch{background:#f5f8fa;border:1px solid var(--brd2);border-radius:8px;padding:10px;
  font-size:11px;color:var(--sub2);line-height:1.7;margin-bottom:9px}
.pitch b{color:var(--navy)}
.ogrid{display:grid;grid-template-columns:repeat(3,1fr);gap:5px;margin-bottom:8px}
.obtn{padding:7px 4px;border-radius:7px;border:1px solid var(--brd);
  background:var(--surf);color:var(--sub2);font-size:10px;font-weight:600;
  cursor:pointer;text-align:center;transition:.12s;font-family:inherit}
.obtn:hover{border-color:#c5d1de;background:#f5f8fa}
.obtn.on{border-color:var(--ora);background:#fff7f5;color:var(--ora)}
.ntxt{width:100%;background:#f5f8fa;border:1px solid var(--brd);border-radius:8px;
  padding:8px;color:var(--txt);font-size:11px;resize:none;outline:none;
  line-height:1.5;font-family:inherit}
.ntxt:focus{border-color:var(--blu)}
.mfacts{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.fact{background:#f5f8fa;border:1px solid var(--brd2);border-radius:7px;padding:8px 9px}
.fl{font-size:8px;color:var(--sub);letter-spacing:.05em;text-transform:uppercase;margin-bottom:2px;font-weight:600}
.fv{font-weight:700;font-size:11px;color:var(--navy)}
.fv.r{color:var(--cb)}.fv.o{color:var(--hot)}.fv.g{color:var(--grn)}.fv.b{color:var(--blu)}
.mhist{background:#f5f8fa;border-radius:8px;padding:9px}
.hi{padding:4px 0;border-bottom:1px solid var(--brd2);font-size:10px;color:var(--txt)}
.hi:last-child{border-bottom:none}
.psect{margin-bottom:16px}
.pst{font-size:10px;font-weight:700;letter-spacing:.05em;text-transform:uppercase;
  margin-bottom:7px;display:flex;align-items:center;gap:6px;color:var(--navy)}
.pct{background:var(--brd2);color:var(--sub2);border-radius:20px;padding:1px 8px;font-size:9px;font-weight:700}
.pitem{background:var(--surf);border:1px solid var(--brd);border-radius:8px;
  padding:9px 12px;margin-bottom:5px;display:flex;align-items:center;gap:8px;cursor:pointer;
  box-shadow:var(--shadow);transition:.12s}
.pitem:hover{box-shadow:var(--shadow-md);border-color:#c5d1de}
.pdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.piname{font-weight:600;font-size:11px;flex:1;color:var(--navy)}
.piph{font-size:10px;color:var(--blu);font-weight:600}
.dc{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:14px;margin-bottom:10px;box-shadow:var(--shadow)}
.dct{font-weight:700;font-size:13px;margin-bottom:8px;color:var(--navy)}
.ds{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid var(--brd2);font-size:11px}
.ds:last-child{border-bottom:none}
.dsv{font-weight:700;color:var(--blu)}
.ibox{background:#f0f9fb;border:1px solid #b3dce7;border-radius:8px;
  padding:11px;font-size:11px;color:var(--sub2);line-height:1.8}
.ibox code{background:#dff0f5;padding:2px 6px;border-radius:4px;
  font-family:monospace;font-size:10px;color:var(--blu)}
.ibox b{color:var(--navy)}
.phinput{background:#f5f8fa;border:1px solid var(--brd);border-radius:7px;
  padding:7px 10px;color:var(--txt);font-size:11px;outline:none;width:100%;margin-bottom:7px;font-family:inherit}
.phinput:focus{border-color:var(--blu)}
.xbtn{background:var(--ora);color:#fff;border:none;
  border-radius:8px;padding:9px 12px;font-size:11px;font-weight:700;cursor:pointer;width:100%;margin-top:6px;font-family:inherit}
.xbtn:hover{background:#e86744}
.dbtn{background:#fff;color:var(--cb);border:1px solid #f5c6c6;
  border-radius:8px;padding:9px 12px;font-size:11px;font-weight:600;cursor:pointer;width:100%;margin-top:5px;font-family:inherit}
.empty{text-align:center;padding:36px 20px;color:var(--sub)}
.ei{font-size:32px;margin-bottom:10px}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.2}}.pulse{animation:pulse 2s ease-in-out infinite}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--brd);border-radius:2px}
#toast{position:fixed;bottom:20px;left:50%;transform:translateX(-50%) translateY(60px);
  background:var(--navy);border-radius:10px;
  padding:9px 16px;font-size:11px;color:#fff;transition:transform .2s;z-index:200;white-space:nowrap;
  box-shadow:0 4px 16px rgba(45,62,80,.25);font-weight:500}
.tsect-hdr{display:flex;flex-direction:column;gap:2px;margin-bottom:10px;padding-bottom:8px;border-bottom:2px solid var(--brd2)}
.tsect-hdr span:first-child{font-weight:700;font-size:14px;color:var(--navy)}
.tsect-sub{font-size:10px;color:var(--sub);font-weight:400}
.tempty{text-align:center;padding:16px;color:var(--sub);font-size:11px;background:var(--surf);border-radius:8px;border:1px dashed var(--brd)}
.emerg-badge{display:inline-block;font-size:8px;font-weight:700;padding:2px 6px;border-radius:20px;background:#fef2f2;color:#dc2626;border:1px solid #fecaca;margin-left:4px}
.preset-btn{flex-shrink:0;padding:6px 12px;border-radius:20px;border:1px solid var(--brd);background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;white-space:nowrap;font-family:inherit;transition:.15s}
.preset-btn.on{background:var(--navy);color:#fff;border-color:var(--navy)}
.svc-tab{flex:1;padding:8px 4px;border:none;background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;font-family:inherit;border-right:1px solid var(--brd);transition:.15s}
.svc-tab:last-child{border-right:none}
.svc-tab.on{background:var(--navy);color:#fff}
.svc-card{background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:12px;margin-bottom:8px;display:flex;flex-direction:column;gap:7px}
.svc-card.overdue{border-left:4px solid #dc2626;background:#fef2f2}
.svc-card.due-soon{border-left:4px solid #d97706;background:#fef9ee}
.svc-card.on-track{border-left:4px solid #059669;background:#f0fdf4}
.svc-card.no-date{border-left:4px solid var(--sub);opacity:.7}

/* ── MOBILE RESPONSIVE ─────────────────────────────── */
@media(max-width:480px){
  /* Bottom nav for phone */
  .tabs{
    position:fixed;bottom:0;left:0;right:0;
    background:var(--surf);
    border-top:1px solid var(--brd);
    border-bottom:none;
    display:flex;overflow-x:auto;
    z-index:90;padding-bottom:env(safe-area-inset-bottom);
    box-shadow:0 -2px 8px rgba(0,0,0,.08);
  }
  .tab{
    flex:1;flex-shrink:0;
    padding:6px 4px 8px;
    font-size:8px;font-weight:700;
    display:flex;flex-direction:column;align-items:center;gap:2px;
    min-width:48px;
  }
  /* Add bottom padding so content clears bottom nav */
  .panels{padding-bottom:80px;}
  /* Header more compact */
  .hdr{padding:8px 12px 6px;}
  .hdr-title{font-size:13px;}
  /* Bigger tap targets */
  button,.btn,.cta-btn{min-height:44px;}
  .tab{min-height:52px;}
  /* Cards full width */
  .grid{grid-template-columns:1fr !important;}
  /* Briefing compact */
  #brief-stats{grid-template-columns:repeat(2,1fr) !important;}
  /* Service modal full screen on phone */
  #svc-log-bg > div{
    max-height:100vh !important;
    border-radius:0 !important;
    padding-bottom:env(safe-area-inset-bottom);
  }
  /* Offline banner */
  }

/* Offline indicator */
#offline-banner{
  display:none;
  position:fixed;top:0;left:0;right:0;
  background:#dc2626;color:#fff;
  padding:6px 12px;
  font-size:11px;font-weight:700;
  text-align:center;z-index:200;
  align-items:center;justify-content:center;gap:6px;
}
#offline-banner.active{display:flex;}

/* Quick action fab - phone only */
.fab{
  display:none !important;
}
@media(max-width:480px){
  .fab{display:flex !important;}
  #offline-banner{display:none;}
  #offline-banner.active{display:flex !important;}
}

/* Larger modal inputs on mobile */
@media(max-width:480px){
  #svc-atp{font-size:18px !important;padding:12px !important;}
  #svc-notes{font-size:14px !important;min-height:80px;}
  .svc-card{padding:14px;}
  .svc-card button{min-height:44px;font-size:12px !important;}
}
.flt-sel{background:var(--surf);border:1px solid var(--brd);border-radius:6px;padding:5px 7px;color:var(--txt);font-size:10px;outline:none;cursor:pointer;font-family:inherit;max-width:130px}

/* ── NEW OUTCOME BUTTON STYLES ─────────────────────── */
.obtn-green  { background:#ecfdf5 !important; border-color:#6ee7b7 !important; color:#059669 !important; }
.obtn-blue   { background:#ecfeff !important; border-color:#67e8f9 !important; color:#0891b2 !important; }
.obtn-yellow { background:#fffbeb !important; border-color:#fcd34d !important; color:#d97706 !important; }
.obtn-gray   { background:#f8fafc !important; border-color:#cbd5e1 !important; color:#475569 !important; }
.obtn-orange { background:#fff7ed !important; border-color:#fdba74 !important; color:#c2410c !important; }
.obtn-red    { background:#fef2f2 !important; border-color:#fca5a5 !important; color:#dc2626 !important; }
.obtn-dark   { background:#f1f5f9 !important; border-color:#94a3b8 !important; color:#374151 !important; }
.obtn-teal   { background:#f0fdfa !important; border-color:#5eead4 !important; color:#0d9488 !important; }
.obtn.on.obtn-green  { background:#059669 !important; color:#fff !important; border-color:#059669 !important; }
.obtn.on.obtn-blue   { background:#0891b2 !important; color:#fff !important; border-color:#0891b2 !important; }
.obtn.on.obtn-yellow { background:#d97706 !important; color:#fff !important; border-color:#d97706 !important; }
.obtn.on.obtn-gray   { background:#475569 !important; color:#fff !important; border-color:#475569 !important; }
.obtn.on.obtn-orange { background:#c2410c !important; color:#fff !important; border-color:#c2410c !important; }
.obtn.on.obtn-red    { background:#dc2626 !important; color:#fff !important; border-color:#dc2626 !important; }
.obtn.on.obtn-dark   { background:#374151 !important; color:#fff !important; border-color:#374151 !important; }
.obtn.on.obtn-teal   { background:#0d9488 !important; color:#fff !important; border-color:#0d9488 !important; }

/* ── TYPE BUTTON SELECTED ──────────────────────────── */
.mtype-on { background:var(--navy) !important; color:#fff !important; border-color:var(--navy) !important; }

/* ── REASON CHIP ───────────────────────────────────── */
.reason-chip { font-size:10px; padding:4px 8px; border-radius:12px; border:1px solid var(--brd2); background:var(--surf); color:var(--sub); cursor:pointer; font-family:inherit; transition:.15s; font-weight:500; }
.reason-chip.on { background:var(--navy); color:#fff; border-color:var(--navy); }

/* ── QUEUE MODE ────────────────────────────────────── */
#queue-bg { display:none; }
#queue-bg.on { display:flex !important; flex-direction:column; height:100%; }
.qbtn { padding:14px 6px; border:none; border-radius:8px; font-size:11px; font-weight:700; cursor:pointer; font-family:inherit; touch-action:manipulation; -webkit-tap-highlight-color:rgba(0,0,0,.15); user-select:none; min-height:48px; }
.qbtn-green  { background:#ecfdf5; color:#059669; }
.qbtn-yellow { background:#fffbeb; color:#d97706; }
.qbtn-blue   { background:#ecfeff; color:#0891b2; }
.qbtn-red    { background:#fef2f2; color:#dc2626; }
.qbtn-gray   { background:#f8fafc; color:#475569; }

/* ── FUNNEL STAGE ──────────────────────────────────── */
.funnel-stage { display:flex; align-items:center; gap:8px; }
.funnel-bar-wrap { flex:1; height:6px; background:var(--brd2); border-radius:3px; }
.funnel-bar { height:6px; border-radius:3px; transition:width .4s; }
.funnel-label { font-size:11px; color:var(--txt); font-weight:600; white-space:nowrap; min-width:80px; }
.funnel-val { font-size:11px; color:var(--sub); white-space:nowrap; min-width:32px; text-align:right; }

/* ── GOAL PACING ───────────────────────────────────── */
.pace-on-track  { color:#059669; font-weight:700; }
.pace-behind    { color:#dc2626; font-weight:700; }
.pace-ahead     { color:#0891b2; font-weight:700; }

</style>
</head>
<body>
<div id="app">
  <header>
    <div class="logo">
      <div class="logo-icon">&#x1F9CA;</div>
      <div><div class="logo-name">Pinellas Ice Co</div>
        <div style="font-size:8px;color:var(--sub);letter-spacing:.04em">PROSPECT TOOL &bull; %%DATE%% &bull; v5</div>
      </div>
    </div>
    <div class="hchips">
      <span class="hc cb" onclick="setF('CALLBACK')">&#x25CF; %%NCB%% Callback</span>
      <span class="hc ht" onclick="setF('HOT')">&#x25CF; %%NHOT%% Hot</span>
      <span class="hc wm" onclick="setF('WARM')">&#x25CF; %%NWARM%% Warm</span>
      <span class="hc ic">&#x1F9CA; %%NCHRON%% Chronic</span>
      <span class="hc rt" onclick="sw('route')">&#x1F5FA; Route</span>
    </div>
    <div class="srch"><span class="si">&#x1F50D;</span>
      <input type="text" id="si" placeholder="Search&#x2026;" oninput="onS()" autocomplete="off" autocorrect="off" autocapitalize="off" spellcheck="false">
    </div>
  </header>
  <nav class="tabs" id="main-nav">
    <div class="tab on"  onclick="sw('today')"     ><span class="tab-icon">&#x1F4CA;</span><span class="tab-lbl">Home</span></div>
    <div class="tab"     onclick="sw('all')"        ><span class="tab-icon">&#x1F4CD;</span><span class="tab-lbl">Prospects</span></div>
    <div class="tab"     onclick="sw('route')"      ><span class="tab-icon">&#x1F5FA;</span><span class="tab-lbl">Route</span></div>
    <div class="tab"     onclick="sw('customers')"  ><span class="tab-icon">&#x1F91D;</span><span class="tab-lbl">Clients</span></div>
    <div class="tab"     onclick="sw('service')"    ><span class="tab-icon">&#x1F9FC;</span><span class="tab-lbl">Service</span></div>
    <div class="tab"     onclick="sw('data')"       ><span class="tab-icon">&#x2699;&#xFE0F;</span><span class="tab-lbl">Settings</span></div>
  </nav>

  <!-- Offline banner -->
  <div id="offline-banner" style="display:none">&#x26A1; Offline &mdash; data saved locally, all features available</div>

  <!-- FAB: Service tab shortcut (phone only) -->
  <button id="svc-fab" onclick="sw('service')" style="display:none">&#x1F9FC;</button>

  <!-- TODAY -->
  <div class="panel on" id="p-today">

    <!-- ── KPI ROW ─────────────────────────────────── -->
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px" id="kpi-row">
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">MRR</div>
        <div style="font-size:20px;font-weight:800;color:var(--grn)" id="kpi-mrr">$0</div>
      </div>
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">Clients</div>
        <div style="font-size:20px;font-weight:800;color:var(--navy)" id="kpi-clients">0</div>
      </div>
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">Pipeline</div>
        <div style="font-size:20px;font-weight:800;color:var(--blu)" id="kpi-pipe">0</div>
      </div>
      <div class="dc" style="text-align:center;padding:10px 6px">
        <div style="font-size:8px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--sub);margin-bottom:3px">This Wk</div>
        <div style="font-size:20px;font-weight:800;color:var(--ora)" id="kpi-week">0</div>
      </div>
    </div>

    <!-- ── FUNNEL ────────────────────────────────────── -->
    <div class="dc" style="margin-bottom:12px;padding:12px 14px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
        <div style="font-weight:700;font-size:12px;color:var(--navy)">&#x1F4CA; Weekly Funnel</div>
        <div style="font-size:9px;color:var(--sub)" id="funnel-week-label"></div>
      </div>
      <div id="funnel-stages" style="display:flex;flex-direction:column;gap:6px"></div>
    </div>

    <!-- ── GOAL PACING ───────────────────────────────── -->
    <div class="dc" style="margin-bottom:12px;padding:12px 14px" id="goal-pacing-card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:700;font-size:12px;color:var(--navy)">&#x1F3AF; Goal Pacing</div>
        <button onclick="sw('data')" style="font-size:9px;padding:3px 8px;border:1px solid var(--brd);border-radius:6px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">&#x2699;&#xFE0F; Settings</button>
      </div>
      <div id="goal-pacing-content"></div>
    </div>

    <!-- ── LOSS REASON BREAKDOWN ─────────────────────── -->
    <div class="dc" style="margin-bottom:12px;padding:12px 14px" id="loss-breakdown-card">
      <div style="font-weight:700;font-size:12px;color:var(--navy);margin-bottom:8px">&#x274C; Why You're Losing</div>
      <div id="loss-breakdown" style="display:flex;flex-direction:column;gap:4px"></div>
      <div id="loss-empty" style="font-size:11px;color:var(--sub);display:none">Log some walk-ins to see loss patterns.</div>
    </div>

    <!-- ── ACTIVE NURTURE (In Play due) ─────────────── -->
    <div style="margin-bottom:18px" id="nurture-section">
      <div class="tsect-hdr">
        <span>&#x1F7E1; In Play — Follow Up Due</span>
        <span class="tsect-sub">All prospects with a follow-up date set, sorted by urgency.</span>
      </div>
      <div class="grid" id="nurture-grid"></div>
      <div class="tempty" id="nurture-empty" style="display:none">No follow-ups due. &#x2713;</div>
    </div>

    <!-- ── BEST MOVE (cold targets) ─────────────────── -->
    <div id="t-actnow">
      <div class="tsect-hdr">
        <span>&#x1F534; Best Cold Targets Today</span>
        <span class="tsect-sub">Top 5 uncontacted. High score + callback priority.</span>
      </div>
      <div class="grid" id="tgrid-actnow"></div>
      <div class="tempty" id="empty-actnow" style="display:none">No urgent prospects.</div>
    </div>

  </div>

  <div class="panel" id="p-all">

    <!-- Preset filters -->
    <div style="display:flex;gap:6px;overflow-x:auto;padding-bottom:8px;margin-bottom:10px;-webkit-overflow-scrolling:touch">
      <button class="preset-btn on" onclick="setPreset('all')"      id="pre-all">All</button>
      <button class="preset-btn"    onclick="setPreset('actnow')"   id="pre-actnow">&#x1F534; Act Now</button>
      <button class="preset-btn"    onclick="setPreset('callback')" id="pre-callback">Callbacks</button>
      <button class="preset-btn"    onclick="setPreset('phone')"    id="pre-phone">Has Phone</button>
      <button class="preset-btn"    onclick="setPreset('chronic')"  id="pre-chronic">Chronic Ice</button>
      <button class="preset-btn"    onclick="setPreset('notyet')"   id="pre-notyet">Not Contacted</button>
      <button class="preset-btn"    onclick="setPreset('inplay')"   id="pre-inplay">&#x1F7E1; In Play</button>
      <button class="preset-btn"    onclick="setPreset('freshice')" id="pre-freshice">&#x1F525; Ice Viol.</button>
      <button class="preset-btn" onclick="setPreset('followups')" id="pre-followups">&#x1F4C5; Follow-Ups</button>
    </div>

    <!-- Filters row -->
    <div style="display:flex;gap:5px;flex-wrap:wrap;margin-bottom:8px;align-items:center">
      <select id="ac"     onchange="populateCityFilter();dRa()" class="flt-sel"><option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option><option>Citrus</option><option>Hernando</option><option>Polk</option><option>Sumter</option></select>
      <select id="ac-city" onchange="dRa()" class="flt-sel"><option value="">All Cities</option></select>
      <select id="ap"     onchange="dRa()" class="flt-sel"><option value="">All Priorities</option><option>CALLBACK</option><option>HOT</option><option>WARM</option><option>WATCH</option></select>
      <select id="as_"    onchange="dRa()" class="flt-sel">
        <option value="">All Status</option>
        <option value="not_contacted">Not Contacted</option>
        <option value="in_play">In Play</option>
        <option value="intro_set">Intro Set</option>
        <option value="not_now">Not Now</option>
        <option value="signed">Signed</option>
        <option value="dead">Dead</option>
      </select>
      <button onclick="enterQueueMode()" style="font-size:10px;padding:5px 10px;border:1px solid var(--blu);border-radius:6px;background:#0a84ff22;color:var(--blu);cursor:pointer;font-family:inherit;font-weight:700;flex-shrink:0">&#x25B6; Queue</button>
      <button onclick="clearFilters()" style="font-size:10px;padding:5px 8px;border:1px solid var(--brd);border-radius:6px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit;flex-shrink:0">Clear</button>
      <span class="fcnt" id="acnt"></span>
    </div>

<div class="grid" id="agrid"></div>
    <div class="tempty" id="a-empty" style="display:none">
      <div class="ei">&#x1F50D;</div>
      <div>No prospects match these filters</div>
    </div>
  </div>

  <!-- QUEUE MODE OVERLAY -->
  <div id="queue-bg" style="display:none;position:fixed;inset:0;background:var(--bg);z-index:90;flex-direction:column">
    <div style="background:var(--navy);padding:12px 16px;display:flex;align-items:center;gap:10px;flex-shrink:0">
      <button onclick="exitQueueMode()" style="border:none;background:rgba(255,255,255,.15);color:#fff;border-radius:8px;padding:6px 12px;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">&#x2715; Exit</button>
      <div style="flex:1;text-align:center;font-size:11px;font-weight:700;color:rgba(255,255,255,.8)" id="queue-progress"></div>
      <button onclick="queueNext()" style="border:none;background:var(--ora);color:#fff;border-radius:8px;padding:6px 12px;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">Skip &#x2192;</button>
    </div>
    <div style="flex:1;overflow-y:auto;padding:14px 16px;min-height:0;-webkit-overflow-scrolling:touch">
      <div id="queue-card-wrap"></div>
    </div>
    <div style="padding:12px 16px;padding-bottom:max(12px,env(safe-area-inset-bottom));background:var(--surf);border-top:1px solid var(--brd2);display:grid;grid-template-columns:repeat(3,1fr);gap:8px;flex-shrink:0;position:relative;z-index:2" id="queue-actions">
      <button class="qbtn qbtn-green"  onclick="queueLogFull('intro_set')"  ontouchend="event.preventDefault();queueLogFull('intro_set')"  >&#x1F4C5; Intro Set</button>
      <button class="qbtn qbtn-yellow" onclick="queueLogFull('in_play')"   ontouchend="event.preventDefault();queueLogFull('in_play')"   >&#x1F7E1; In Play</button>
      <button class="qbtn qbtn-blue"   onclick="queueLogFull('no_contact')" ontouchend="event.preventDefault();queueLogFull('no_contact')" >&#x1F6AA; No Contact</button>
      <button class="qbtn qbtn-red"    onclick="queueLogFull('not_now')"   ontouchend="event.preventDefault();queueLogFull('not_now')"   >&#x274C; Not Now</button>
      <button class="qbtn qbtn-gray"   onclick="queueLogFull('voicemail')" ontouchend="event.preventDefault();queueLogFull('voicemail')" >&#x1F4F2; Voicemail</button>
      <button class="qbtn qbtn-gray"   onclick="queueLogFull('dead')"      ontouchend="event.preventDefault();queueLogFull('dead')"      >&#x26AB; Dead</button>
    </div>
  </div>

  <div class="panel" id="p-route">
    <div style="display:flex;flex-direction:column;height:calc(100vh - 120px);gap:8px;overflow:hidden">

      <!-- Plan My Day header -->
      <div class="dc" style="flex-shrink:0;padding:12px">
        <div style="font-weight:800;font-size:14px;color:var(--navy);margin-bottom:10px">&#x1F4C5; Plan My Day</div>

        <!-- Row 1: Start ZIP + Time budget -->
        <div style="display:flex;gap:6px;margin-bottom:8px">
          <div style="flex:1">
            <div style="font-size:9px;color:var(--sub);font-weight:600;margin-bottom:3px">START ZIP</div>
            <input id="rzip" type="text" placeholder="34689" value="34689" maxlength="5"
              style="width:100%;padding:7px 8px;border:1px solid var(--brd);border-radius:7px;font-size:12px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none"
              oninput="rRoute()">
          </div>
          <div style="flex:1">
            <div style="font-size:9px;color:var(--sub);font-weight:600;margin-bottom:3px">TIME AVAILABLE</div>
            <select id="rtime" onchange="planMyDay()"
              style="width:100%;padding:7px 8px;border:1px solid var(--brd);border-radius:7px;font-size:12px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
              <option value="0">Manual (add stops)</option>
              <option value="2">2 hours (~4 stops)</option>
              <option value="3">3 hours (~6 stops)</option>
              <option value="4">4 hours (~8 stops)</option>
              <option value="6">Half day (~10 stops)</option>
              <option value="8">Full day (~14 stops)</option>
            </select>
          </div>
        </div>

        <!-- Row 2: Filters -->
        <div style="display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap">
          <select id="rc" onchange="rRoute()"
            style="flex:1;min-width:90px;padding:6px 7px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
            <option value="">All Counties</option><option>Pinellas</option><option>Hillsborough</option><option>Pasco</option>
          </select>
          <select id="rp" onchange="rRoute()"
            style="flex:1;min-width:90px;padding:6px 7px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
            <option value="">All Priorities</option><option>CALLBACK</option><option>HOT</option><option>WARM</option>
          </select>
          <select id="rrad" onchange="rRoute()"
            style="flex:1;min-width:80px;padding:6px 7px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
            <option value="5">5 mi</option><option value="8" selected>8 mi</option>
            <option value="12">12 mi</option><option value="20">20 mi</option><option value="999">Any</option>
          </select>
        </div>

        <!-- Row 3: Action buttons -->
        <div style="display:flex;gap:6px">
          <button onclick="planMyDay()"
            style="flex:2;padding:8px;border:none;border-radius:8px;background:var(--navy);color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">
            &#x26A1; Build Optimal Route
          </button>
          <button onclick="optimizeRoute()"
            style="flex:1;padding:8px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:10px;font-weight:600;cursor:pointer;font-family:inherit">
            Sort
          </button>
          <button onclick="clearRoute()"
            style="flex:1;padding:8px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:10px;cursor:pointer;font-family:inherit">
            Clear
          </button>
        </div>

        <div style="font-size:9px;color:var(--sub);margin-top:6px" id="rhint">Enter start ZIP (or tap Start 📍 on any business). Set time budget. Tap Build Optimal Route.</div>
      </div>

      <!-- Two-column: list left, route+map right -->
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;flex:1;min-height:0;overflow:hidden">

        <!-- Prospect list -->
        <div style="display:flex;flex-direction:column;min-height:0">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px" id="rlist-cnt"></div>
          <div id="rlist" style="overflow-y:auto;flex:1;display:flex;flex-direction:column;gap:3px"></div>
        </div>

        <!-- Route stops + map -->
        <div style="display:flex;flex-direction:column;min-height:0;gap:8px">

          <!-- Built route -->
          <div id="day-route" style="display:none;background:var(--surf);border:1px solid var(--brd);border-radius:10px;padding:10px;flex-shrink:0;max-height:50%;overflow-y:auto">
            <div style="font-weight:700;font-size:11px;color:var(--navy);margin-bottom:2px">
              &#x1F4CD; Route — <span id="stopcnt">0</span> stops
              <span id="route-mi" style="font-size:9px;color:var(--sub);font-weight:400"></span>
            </div>
            <div id="route-time-est" style="font-size:9px;color:var(--sub);margin-bottom:7px"></div>
            <div id="day-stops"></div>
            <div style="display:flex;gap:5px;margin-top:8px">
              <button class="route-btn" onclick="openMaps()" ontouchend="event.preventDefault();openMaps()" style="touch-action:manipulation;cursor:pointer" style="flex:2">Open in Google Maps &#x2192;</button>
            </div>
          </div>

          <!-- Map -->
          <div class="map-area" id="map-area" style="flex:1;min-height:180px;border-radius:10px;overflow:hidden">
            <div class="map-empty">
              <div style="font-size:32px;margin-bottom:8px">&#x1F5FA;&#xFE0F;</div>
              <div style="font-size:11px">Enter ZIP + time budget<br>and tap Build Optimal Route</div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- SERVICE -->
  <div class="panel" id="p-service">

    <!-- Revenue forecast strip -->
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:12px" id="svc-forecast"></div>

    <!-- At-risk alert -->
    <div id="svc-atrisk" style="margin-bottom:12px"></div>

    <!-- Tabs: Calendar | Route | Reports | Referrals -->
    <div style="display:flex;gap:0;margin-bottom:12px;border:1px solid var(--brd);border-radius:8px;overflow:hidden;flex-wrap:wrap">
      <button class="svc-tab on" onclick="setSvcTab('cal')"      id="svct-cal">&#x1F4C5; Calendar</button>
      <button class="svc-tab"    onclick="setSvcTab('route')"    id="svct-route">&#x1F5FA; Route</button>
      <button class="svc-tab"    onclick="setSvcTab('reports')"  id="svct-reports">&#x1F4CB; Reports</button>
      <button class="svc-tab"    onclick="setSvcTab('tutorials')" id="svct-tutorials">&#x1F4D6; Tutorials</button>
      <button class="svc-tab"    onclick="setSvcTab('refs')"     id="svct-refs">&#x1F91D; Referrals</button>
    </div>

    <!-- Calendar view -->
    <div id="svc-cal">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Recurring clients due for service. Tap to log or reschedule.</div>
      <div id="svc-cal-list"></div>
    </div>

    <!-- Service route builder -->
    <div id="svc-route" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Build a service route for clients due this week, clustered by geography.</div>
      <div style="display:flex;gap:6px;margin-bottom:8px">
        <select id="svc-week" onchange="renderServiceRoute()"
          style="flex:1;padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="0">This week</option>
          <option value="7">Next week</option>
          <option value="14">In 2 weeks</option>
          <option value="-7">Overdue</option>
        </select>
        <button onclick="buildServiceRoute()"
          style="padding:7px 12px;border:none;border-radius:7px;background:var(--navy);color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">
          &#x26A1; Build Route
        </button>
      </div>
      <div id="svc-route-list"></div>
      <div id="svc-route-map-btn" style="display:none;margin-top:8px">
        <button onclick="openServiceMaps()" class="route-btn">Open in Google Maps &#x2192;</button>
      </div>
    </div>

    <!-- Service reports -->
    <div id="svc-reports" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Generate a printable service report for any client visit.</div>
      <select id="svc-report-client" onchange="loadReportClient()"
        style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;margin-bottom:8px">
        <option value="">Select a client...</option>
      </select>
      <div id="svc-report-preview" style="overflow-y:auto;max-height:70vh;-webkit-overflow-scrolling:touch"></div>
    </div>

    <!-- Tutorials -->
    <div id="svc-tutorials" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Step-by-step cleaning guides for each machine brand. Select brand and procedure type.</div>
      <div style="display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap">
        <select id="tut-brand" onchange="renderTutorial()"
          style="flex:1;min-width:120px;padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="">Select brand...</option>
          <option>Manitowoc</option><option>Hoshizaki</option><option>Scotsman</option>
          <option>Ice-O-Matic</option><option>Follett</option><option>Cornelius</option>
        </select>
        <select id="tut-type" onchange="renderTutorial()"
          style="flex:1;min-width:120px;padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="deep_clean">&#x1F9FC; Deep Clean</option>
          <option value="maintenance_60">&#x1F527; 60-Day Maintenance</option>
          <option value="atp_protocol">&#x1F4CA; ATP Testing Guide</option>
        </select>
      </div>
      <div id="tut-content"></div>
    </div>

    <!-- Referrals -->
    <div id="svc-refs" style="display:none">
      <div style="font-size:9px;color:var(--sub);margin-bottom:8px">Track who referred new clients. Your best customers are your best salespeople.</div>
      <div id="svc-refs-list"></div>
    </div>

  </div>

  <!-- PIPELINE -->
  <div class="panel" id="p-customers">
    <div class="fbar">
      <select id="cust-status" onchange="rCust()">
        <option value="">All Customers</option>
        <option value="customer_recurring">Recurring</option>
        <option value="customer_intro">Intro ($99 first visit)</option>
        <option value="customer_once">One-Time ($249)</option>
        <option value="quoted">Quoted - Pending</option>
        <option value="churned">Churned</option>
      </select>
      <span class="fcnt" id="cust-cnt"></span>
    </div>

    <!-- Revenue summary bar -->
    <div id="cust-summary" style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px">
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Monthly Recurring</div>
        <div style="font-size:22px;font-weight:800;color:var(--grn)" id="mrr-val">$0</div>
        <div class="fl">MRR</div>
      </div>
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Active Customers</div>
        <div style="font-size:22px;font-weight:800;color:var(--navy)" id="cust-count">0</div>
        <div class="fl">Accounts</div>
      </div>
      <div class="dc" style="text-align:center">
        <div class="fl" style="font-size:9px">Annual Run Rate</div>
        <div style="font-size:22px;font-weight:800;color:var(--blu)" id="arr-val">$0</div>
        <div class="fl">ARR</div>
      </div>
    </div>

    <div id="cust-list"></div>

    <div class="tempty" id="cust-empty" style="display:none">
      <div class="ei">&#x1F91D;</div>
      <div style="font-weight:600;margin-bottom:4px">No customers yet</div>
      <div style="font-size:11px;color:var(--sub)">When you close a deal, open any prospect card and mark them as Won.</div>
    </div>
  </div>

  <!-- DATA -->
  <div class="panel" id="p-data">
    <div class="dc" style="margin-bottom:12px">
      <div class="dct">&#x1F4E6; Dataset &mdash; %%DATE%%</div>
      <div class="ds"><span>Total prospects</span><span class="dsv">%%TOTAL%%</span></div>
      <div class="ds"><span>Source</span><span class="dsv">FL DBPR District 3 (live)</span></div>
      <div class="ds"><span>Auto-rebuilds</span><span class="dsv">Weekly via GitHub Actions</span></div>
      <div class="ds"><span>Chronic ice offenders</span><span class="dsv">%%NCHRON%%</span></div>
      <div class="ds"><span>Callback urgent</span><span class="dsv">%%NCB%%</span></div>
      <div class="ds"><span>With phone number</span><span class="dsv">%%NPHONE%%</span></div>
    </div>

    <div class="dc" id="goals-section">
      <div class="dct">&#x1F3AF; Goals &amp; Targets</div>
      <div style="font-size:10px;color:var(--sub);margin-bottom:10px">Set your targets. The home dashboard tracks progress automatically.</div>
      <div style="display:flex;flex-direction:column;gap:8px">
        <div style="background:#f5f8fa;border-radius:8px;padding:10px">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">&#x1F91D; Client Goal</div>
          <div style="display:flex;gap:6px;align-items:center;margin-bottom:4px">
            <input id="goal-clients" type="number" min="1" max="500" placeholder="10" onchange="autoMRR()"
              style="width:70px;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            <span style="font-size:11px;color:var(--sub)">recurring clients</span>
          </div>
          <div style="display:flex;gap:6px;align-items:center">
            <input id="goal-deadline" type="date"
              style="flex:1;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:var(--txt);outline:none">
            <span style="font-size:11px;color:var(--sub)">deadline</span>
          </div>
        </div>
        <div style="background:#f5f8fa;border-radius:8px;padding:10px">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">&#x1F4B0; Revenue Goal</div>
          <div style="display:flex;gap:6px;align-items:center">
            <input id="goal-mrr" type="number" min="100" step="100" placeholder="1490"
              style="width:80px;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            <span style="font-size:11px;color:var(--sub)">target MRR ($)</span>
          </div>
        </div>
        <div style="background:#f5f8fa;border-radius:8px;padding:10px">
          <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">&#x1F4DE; Daily Activity</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px">
            <div>
              <div style="font-size:9px;color:var(--sub);margin-bottom:3px">Walk-ins / day</div>
              <input id="goal-walkins" type="number" min="0" max="30" placeholder="2"
                style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            </div>
            <div>
              <div style="font-size:9px;color:var(--sub);margin-bottom:3px">Calls / day</div>
              <input id="goal-calls" type="number" min="0" max="50" placeholder="5"
                style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:13px;font-weight:700;font-family:inherit;background:#fff;color:var(--navy);outline:none;text-align:center">
            </div>
          </div>
        </div>
        <button onclick="saveGoals()" ontouchend="event.preventDefault();saveGoals()" style="width:100%;padding:10px;background:var(--navy);color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit">Save Goals</button>
      </div>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x1F504; Rebuild With New Data</div>
      <div class="ibox">
        <b>Step 1 &mdash; Large XLS/XLSX?</b> Convert first:<br>
        <code>python convert_to_csv.py yourfile.xlsx</code><br><br>
        <b>Step 2 &mdash; Rebuild:</b><br>
        <code>python build.py file1.csv file2.csv</code><br>
        Call log and phone numbers carry over automatically.
      </div>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x1F4E7; Monday Briefing</div>
      <div style="font-size:11px;color:var(--sub);line-height:1.6">Automated weekly email via GitHub Actions. Requires <code>RESEND_API_KEY</code> and <code>BRIEFING_EMAIL</code> in GitHub Secrets.</div>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x2699;&#xFE0F; App Settings</div>
      <div style="font-size:10px;color:var(--sub);margin-bottom:6px">HubSpot Portal ID &mdash; find in your HubSpot URL: app.hubspot.com/contacts/<b>XXXXXXXX</b>/</div>
      <input class="phinput" id="hs-portal" type="text" placeholder="e.g. 12345678" maxlength="12" oninput="saveSettings()">
      <div style="margin-top:10px">
        <div style="font-size:10px;color:var(--sub);margin-bottom:6px">Home Base ZIP &mdash; auto-fills your route start</div>
        <input class="phinput" id="home-zip" type="text" placeholder="34689" value="34689" maxlength="5" oninput="saveSettings()">
      </div>
      <button onclick="saveSettings();toast('\u2713 Settings saved')" ontouchend="event.preventDefault();saveSettings();toast('\u2713 Settings saved')" style="width:100%;margin-top:10px;padding:9px;background:var(--navy);color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">Save Settings</button>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x2795; Add / Edit a Phone Number</div>
      <div style="font-size:10px;color:var(--sub);margin-bottom:8px">License ID is shown on each prospect card.</div>
      <input class="phinput" id="ph-id"  type="text" placeholder="License ID (shown on card)">
      <input class="phinput" id="ph-num" type="tel"  placeholder="Phone e.g. +1 727-555-1234">
      <input class="phinput" id="ph-hrs" type="text" placeholder="Hours e.g. Mon-Fri 11am-9pm">
      <button class="xbtn" style="margin-top:4px;width:100%" onclick="addPhone()">Save Phone Number</button>
    </div>

    <div class="dc" style="margin-top:12px">
      <div class="dct">&#x1F4E4; Export &amp; Reset</div>
      <button class="xbtn" onclick="exportCSV()" style="margin-bottom:6px;width:100%">Export Pipeline to CSV</button>
      <button class="xbtn" onclick="exportDirectoryData()" style="background:#059669;color:#fff;border-color:#059669;margin-bottom:10px;width:100%">&#x1F4E5; Export Directory Data</button>
      <button class="dbtn" onclick="clrLog()" style="margin-bottom:4px;width:100%">Clear Call Log</button>
      <button class="dbtn" onclick="clrCustomers()" style="margin-bottom:4px;width:100%">Clear Customer Data</button>
      <button class="dbtn" onclick="clrAll()" style="width:100%">Clear ALL Data (full reset)</button>
    </div>

  </div>

  <div id="mbg" onclick="closeM(event)">
  <div id="modal">
    <div class="mh" style="position:sticky;top:0;z-index:10;background:var(--surf);border-bottom:1px solid var(--brd2);padding:8px 16px;display:flex;align-items:center;justify-content:space-between">
      <div style="width:36px;height:4px;background:var(--brd);border-radius:2px"></div>
      <button onclick="closeMForce()" style="border:none;background:var(--brd2);border-radius:50%;width:28px;height:28px;font-size:14px;color:var(--sub);cursor:pointer;display:flex;align-items:center;justify-content:center;font-family:inherit;flex-shrink:0">&#x2715;</button>
    </div>
    <div class="mname" id="mn"></div>
    <div class="mloc"  id="ml"></div>
    <div class="mphsec">
      <div class="mphl">CONTACT</div>
      <div class="mphnum" id="mph"></div>
      <div class="mphacts" id="mpa"></div>
      <div id="mhrs" class="mhours"></div>
      <div id="mrat" style="font-size:9px;color:#555;margin-top:3px"></div>
    </div>
    <div id="mice" style="margin-bottom:10px"></div>
    <div style="margin-bottom:10px">
      <div class="msect">PITCH GUIDE</div>
      <div style="display:flex;gap:5px;margin-bottom:6px">
        <button id="btn-phone" onclick="setPitchMode('phone')"
          style="flex:1;padding:6px;border-radius:7px;border:1px solid var(--blu);background:#0a84ff22;color:var(--blu);font-size:10px;font-weight:700;cursor:pointer">
          &#x1F4DE; Phone Call
        </button>
        <button id="btn-walkin" onclick="setPitchMode('walkin')"
          style="flex:1;padding:6px;border-radius:7px;border:1px solid var(--brd);background:transparent;color:var(--sub);font-size:10px;font-weight:700;cursor:pointer">
          &#x1F6B6; Cold Walk-In
        </button>
      </div>
      <div class="pitch" id="mpitch"></div>
      <div class="pitch" id="mwalkin" style="display:none;border-color:#e2e8f0;color:var(--sub2)"></div>
      <div style="margin-top:8px">
        <div class="msect">OBJECTION HANDLERS</div>
        <div id="mobjections" style="display:flex;flex-direction:column;gap:5px"></div>
      </div>
    </div>
    <div style="margin-bottom:10px">
      <div class="msect">LOG THIS CONTACT</div>

      <!-- Interaction type -->
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px">
        <button id="mtype-walkin" onclick="setType('walkin')"
          style="padding:7px;border-radius:8px;border:2px solid var(--brd);background:transparent;color:var(--sub);font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;transition:.15s">
          &#x1F6B6; Walk-In
        </button>
        <button id="mtype-call" onclick="setType('call')"
          style="padding:7px;border-radius:8px;border:2px solid var(--brd);background:transparent;color:var(--sub);font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;transition:.15s">
          &#x1F4DE; Phone Call
        </button>
      </div>

      <!-- Outcome buttons -->
      <div class="ogrid">
        <button class="obtn obtn-green"  data-o="signed"     onclick="selO('signed')">&#x2705; Signed</button>
        <button class="obtn obtn-blue"   data-o="intro_set"  onclick="selO('intro_set')">&#x1F4C5; Intro Set</button>
        <button class="obtn obtn-yellow" data-o="in_play"    onclick="selO('in_play')">&#x1F7E1; In Play</button>
        <button class="obtn obtn-gray"   data-o="no_contact" onclick="selO('no_contact')">&#x1F6AA; No Contact</button>
        <button class="obtn obtn-orange" data-o="voicemail"  onclick="selO('voicemail')">&#x1F4F2; Voicemail</button>
        <button class="obtn obtn-red"    data-o="not_now"    onclick="selO('not_now')">&#x274C; Not Now</button>
        <button class="obtn obtn-dark" data-o="dead" onclick="selO('dead')" style="grid-column:1/-1">&#x26AB; Dead &mdash; Wrong fit / hard no / corporate</button>
        <button class="obtn obtn-teal" data-o="service_done" onclick="selO('service_done')" style="grid-column:1/-1">&#x1F9FC; Service Done</button>
      </div>

      <!-- Reason picker - shown only for not_now and in_play -->
      <div id="reason-wrap" style="display:none;margin-top:8px">
        <div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px">Reason</div>
        <div style="display:flex;flex-wrap:wrap;gap:4px" id="reason-grid"></div>
      </div>

      class="ntxt" id="mnotes" rows="2" placeholder="Notes&#x2026;"></textarea>
      <div style="display:flex;gap:6px;align-items:center;margin-top:6px;padding:6px 8px;background:#f5f8fa;border-radius:7px">
        <span style="font-size:9px;color:var(--sub);white-space:nowrap;font-weight:600">&#x1F4C5; Follow-up date:</span>
        <input type="date" id="mfollowup" style="flex:1;padding:4px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:var(--txt);outline:none">
      </div>
      <button class="btn blog" style="width:100%;margin-top:6px;padding:8px" onclick="saveL()">Save Log Entry</button>
    </div>
    <!-- CLOSE DEAL -->
    <div style="margin-bottom:10px">
      <div class="msect">CLOSE DEAL</div>
      <!-- Intro offer - most prominent -->
      <button onclick="markWon('customer_intro')"
        style="width:100%;padding:10px;border:2px solid var(--ora);border-radius:9px;background:#fff7f5;color:var(--ora);font-weight:800;font-size:12px;cursor:pointer;font-family:inherit;margin-bottom:6px">
        &#x1F525; Intro Offer  -  $99 first visit<br><span style="font-size:10px;font-weight:400">Try us once, no commitment. Recurring starts after.</span>
      </button>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:6px">
        <button onclick="markWon('customer_recurring')"
          style="padding:10px 6px;border:2px solid var(--grn);border-radius:9px;background:#ecfdf5;color:#059669;font-weight:700;font-size:11px;cursor:pointer;font-family:inherit">
          &#x1F4B0; Won  -  Recurring<br><span style="font-size:10px;font-weight:400" id="mwon-monthly"></span>
        </button>
        <button onclick="markWon('customer_once')"
          style="padding:10px 6px;border:2px solid var(--blu);border-radius:9px;background:#eff6ff;color:var(--blu);font-weight:700;font-size:11px;cursor:pointer;font-family:inherit">
          &#x1F9FC; Won  -  One-Time<br><span style="font-size:10px;font-weight:400" id="mwon-onetime"></span>
        </button>
      </div>
      <button onclick="markWon('churned')"
        style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:8px;background:transparent;color:var(--sub);font-size:10px;cursor:pointer;font-family:inherit">
        Mark as Lost / Churned
      </button>
    </div>
    <!-- CONTACTS -->
    <div style="margin-bottom:10px">
      <div class="msect">CONTACTS &amp; INTEL</div>

      <!-- Current vendor field -->
      <div style="margin-bottom:8px;padding:8px;background:#fef9ee;border:1px solid #fde68a;border-radius:7px">
        <div style="font-size:9px;font-weight:700;color:#92400e;margin-bottom:4px">&#x1F575; CURRENT ICE VENDOR</div>
        <div style="display:flex;gap:5px;align-items:center">
          <input id="mc-vendor" type="text" placeholder="e.g. Ecolab, local guy, staff cleans it..."
            style="flex:1;padding:6px;border:1px solid #fde68a;border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:var(--txt);outline:none">
          <button onclick="saveVendor()" style="padding:6px 10px;border:none;border-radius:6px;background:#92400e;color:#fff;font-size:10px;font-weight:600;cursor:pointer;font-family:inherit">Save</button>
        </div>
        <div id="mc-vendor-display" style="font-size:10px;color:#92400e;margin-top:4px;font-style:italic"></div>
      </div>

      <div id="mcontacts" style="margin-bottom:8px"></div>
      <div style="display:flex;gap:5px;margin-bottom:4px">
        <input id="mc-name" type="text" placeholder="Contact name" style="flex:2;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
        <select id="mc-role" style="flex:1;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">
          <option value="owner">Owner</option>
          <option value="gm">GM</option>
          <option value="manager">Manager</option>
          <option value="chef">Chef</option>
          <option value="staff">Staff</option>
        </select>
      </div>
      <input id="mc-phone" type="tel" placeholder="Direct phone (optional)" style="width:100%;padding:6px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;margin-bottom:4px">
      <button onclick="addContact()" style="width:100%;padding:7px;border:none;border-radius:7px;background:var(--navy);color:#fff;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit">+ Save Contact</button>
    </div>
    <div style="margin-bottom:10px"><div class="msect">INSPECTION DETAILS</div><div class="mfacts" id="mfacts"></div></div>
    <div id="mhs" style="display:none;margin-bottom:10px">
      <div class="msect">CONTACT HISTORY</div><div class="mhist" id="mhist"></div>
    </div>
  </div>
</div>
<div id="toast"></div>

<script>
const P=%%DATA%%;
const PHONES=%%PHONES%%;
const ZIPS=%%ZIPS%%;

const PITCHES={
  callback:n=>`"Hi, is this <b>${n}</b>? I'm [Your Name] from Pinellas Ice Co  -  we clean and sanitize commercial ice machines locally. I'm reaching out because you recently had some health inspection issues. Ice machines are almost always part of callback inspections. We can get yours cleaned and documented before the inspector returns  -  usually takes about 2 hours. <b>Do you have time this week?</b>"`,
  overdue_urgent:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We specialize in ice machine cleaning in Pinellas and Hillsborough. I know dealing with inspection issues is stressful  -  we clean and document quickly and give you a service report you can show the inspector. <b>Can we schedule something this week?</b>"`,
  overdue:n=>`"Hi, is this <b>${n}</b>? This is [Your Name] from Pinellas Ice Co  -  commercial ice machine cleaning. Ice machines are one of the most flagged items when inspectors return. <b>Is this a good time to talk?</b>"`,
  pre_hot:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. Your next routine health inspection could be coming up very soon  -  ice machines are one of the most flagged items. We do a full clean and sanitize with a service report. <b>Want to get squared away before the inspector shows up?</b>"`,
  pre_warm:n=>`"Hi, is this <b>${n}</b>? My name is [Your Name] from Pinellas Ice Co. You have some lead time before your next inspection  -  perfect window to get a service on the books. <b>Would you be open to scheduling this month?</b>"`,
  high_risk:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We clean ice machines for restaurants in the area and based on your inspection history I wanted to reach out specifically. <b>Has anyone talked to you recently about your ice machine maintenance?</b>"`,
  routine:n=>`"Hi, this is [Your Name] from Pinellas Ice Co. We clean and sanitize commercial ice machines  -  FDA recommends service every 6 months. We're in your area. <b>Is this something that would be useful for you?</b>"`,
};
const WALKIN={
  callback:n=>`<b>Opening (to staff/gatekeeper):</b><br>&#x201C;Hey, is the manager around for one second? I&#39;m [Name] from Pinellas Ice Co. I&#39;ll be quick.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi  -  briefly: you had a callback inspection recently, and ice machines are almost always what inspectors target on the return visit. I clean and certify them, takes about 90 minutes, you get a dated service report to show the inspector. I&#39;m in the area this week. <b>Want me to take a look today?</b>&#x201D;<br><br><i>If hesitant:</i> &#x201C;No pressure  -  I can open it up and tell you what an inspector would flag. Five minutes, free.&#x201D;`,
  overdue_urgent:n=>`<b>Opening:</b><br>&#x201C;Hey, is the manager in? I&#39;ll be really quick.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi, I&#39;m [Name] from Pinellas Ice Co. Your location came up because you have an open inspection issue. Ice machines get hit hard on follow-up visits. I can clean and document yours before they return. <b>Do you have 90 minutes this week?</b>&#x201D;`,
  pre_hot:n=>`<b>Opening:</b><br>&#x201C;Hi, is the owner or manager available? Quick question about your health inspection.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co. Based on your inspection history your next one is probably coming up soon. Ice machines are one of the top-cited items  -  mold, mineral scale, slime. I service and document them so you&#39;re covered. <b>Can I show you what inspectors typically look for?</b>&#x201D;<br><br><i>Close:</i> &#x201C;I can come back at a time that works. First visit is $99 and I leave you a full compliance report.&#x201D;`,
  pre_warm:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager around for a minute?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co  -  I clean commercial ice machines. You&#39;ve got some time before your next inspection, which is actually the perfect window  -  no last-minute stress. <b>Is ice machine maintenance on your radar?</b>&#x201D;`,
  high_risk:n=>`<b>Opening:</b><br>&#x201C;Hi, is the owner or manager in? I have some information about your inspection history.&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hi, I&#39;m [Name] from Pinellas Ice Co. Your location has had some ice-related violations flagged. This is exactly what inspectors look for on follow-ups. <b>When was your machine last serviced?</b>&#x201D;`,
  routine:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager available?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co  -  I clean commercial ice machines. FDA recommends every 6 months. Inspectors are specifically trained to look at machine interiors now. <b>When was yours last serviced?</b>&#x201D;<br><br><i>Let their answer guide you. If they don&#39;t know  -  that&#39;s your opening.</i>`,
  overdue:n=>`<b>Opening:</b><br>&#x201C;Hi, is the manager around?&#x201D;<br><br><b>To manager:</b><br>&#x201C;Hey, I&#39;m [Name] from Pinellas Ice Co. Your location came up because of some past inspection flags. Ice machines are one of the most targeted items when inspectors return. <b>Do you mind if I take a quick look at yours?</b>&#x201D;`,
};
const OI={
  // New outcomes
  signed:'Signed ✓', intro_set:'Intro Set', in_play:'In Play',
  no_contact:'No Contact', not_now:'Not Now', dead:'Dead',
  voicemail:'Voicemail', service_done:'Service Done ✓',
  // Legacy compat (existing log entries)
  customer_recurring:'Signed ✓', customer_once:'Signed ✓', customer_intro:'Intro Set',
  follow_up:'In Play', interested:'In Play', scheduled:'Intro Set',
  not_interested:'Not Now', no_answer:'No Contact', quoted:'In Play', churned:'Dead',
};
const OI_COLOR={
  signed:'#059669', intro_set:'#0891b2', in_play:'#d97706',
  no_contact:'#64748b', not_now:'#dc2626', dead:'#374151',
  voicemail:'#7c3aed', service_done:'#059669',
  customer_recurring:'#059669', customer_once:'#059669', customer_intro:'#0891b2',
  follow_up:'#d97706', interested:'#d97706', scheduled:'#0891b2',
  not_interested:'#dc2626', no_answer:'#64748b', quoted:'#d97706', churned:'#374151',
};
// Normalize legacy outcomes to new system for filtering
function normO(o){
  const map={customer_recurring:'signed',customer_once:'signed',customer_intro:'intro_set',
    follow_up:'in_play',interested:'in_play',scheduled:'intro_set',
    not_interested:'not_now',no_answer:'no_contact',quoted:'in_play',churned:'dead'};
  return map[o]||o;
}
const REASONS={
  already_has_service:'Already has service',
  cleans_themselves:'Cleans themselves',
  too_expensive:'Too expensive',
  landlord_handles:'Landlord handles it',
  need_partner:'Needs partner / corporate',
  just_passed:'Just passed inspection',
  no_authority:'No authority (manager)',
  franchise_corp:'Franchise / corporate contract',
  seasonal:'Seasonal / not ready',
  no_reason:'No reason given',
};
// Outcomes that show reason picker
const REASON_OUTCOMES=new Set(['not_now','in_play']);
const PO={CALLBACK:0,HOT:1,WARM:2,WATCH:3,LATER:4};
const PC={CALLBACK:'var(--cb)',HOT:'var(--hot)',WARM:'var(--warm)',WATCH:'var(--watch)',LATER:'var(--sub)'};
const ICN={V14:'food contact surfaces (V14)',V22:'non-PHF surfaces (V22)',V50:'food contact surfaces (V50)',V37:'equipment repair (V37)',V23:'utensil sanitation (V23)'};

// Objection handlers -- proven B2B field responses

// ── MACHINE TUTORIALS ─────────────────────────────────────────────────────────
const MACHINE_BRANDS = ['Manitowoc','Hoshizaki','Scotsman','Ice-O-Matic','Follett','Cornelius','Kold-Draft','True','Turbo Air','Other'];
const MACHINE_TYPE_LIST = ['Cuber','Nugget/Sonic','Flaker','Shaved Ice','Undercounter','Modular','Countertop','Other'];

const MACHINE_TYPES = {
  'Manitowoc':  ['Indigo NXT Series','Insight Series','QD/QY Modular','NEO Undercounter','Countertop'],
  'Hoshizaki':  ['KM Crescent Cube','DCM Cubelet','AM Modular','B Series Undercounter','Countertop'],
  'Scotsman':   ['C Series Prodigy','N Series Nugget','HID Prodigy Plus','CU Undercounter','Countertop'],
  'Ice-O-Matic':['ICEU Undercounter','GEM Nugget','B Series','ICEU Half Cube','Countertop'],
  'Follett':    ['Symphony Plus','7 Series','25 Series','Horizon Elite','Countertop Nugget'],
  'Cornelius':  ['IMI Series','Enduro Series','CR Series','UCB Undercounter','Countertop'],
};

const FILTER_TYPES = [
  'None / No filter',
  'Basic inline sediment filter',
  'Carbon block filter',
  'Scale inhibitor cartridge',
  'Everpure MH-2',
  'Everpure 4H-2',
  'Everpure MH',
  'Pentek P25',
  'Omnipure K5525',
  'Cuno CFS8112EL',
  'Manitowoc AR-10000-P',
  'Hoshizaki H9655-06',
  'Other (note below)',
];

// Chemical reference - Nu-Calgon Nickel-Safe + No-Rinse Sanitizer
const CHEM_REF = {
  cleaner: {
    name: 'Nu-Calgon Nickel-Safe Ice Machine Cleaner',
    standard: '8oz cleaner per 1 gallon warm water',
    heavy: '16oz per 1 gallon (heavy scale or first service)',
    note: 'NEVER mix with sanitizer. Rinse all cleaner completely before sanitizing.',
    color: 'Green liquid. Mixed solution will be light green/clear.',
  },
  sanitizer: {
    name: 'Nu-Calgon Ice Machine Sanitizer No-Rinse',
    ratio: '1oz per 1 gallon water = 200ppm quat solution',
    note: 'No rinsing after application. Apply and let air dry. This is the final step always.',
    color: 'Clear liquid with slight odor.',
  },
  atp: {
    meter: 'Hygiena Ensure v2 with Ultrasnap testers',
    pass_ice: '10 RLU or below — ice contact surfaces (evaporator, water trough)',
    pass_food: '30 RLU or below — general food contact surfaces (bin, scoop holder)',
    fail: 'Above threshold: re-clean and re-sanitize. Retest before leaving.',
    technique: 'Snap the Ultrasnap tube, swab 10cm² area with firm pressure using Z-pattern, insert into Ensure v2, read in 15 seconds.',
  },
  atp_sales: {
    title: '📊 Sales ATP Protocol — Where to Swab BEFORE Cleaning',
    purpose: 'Swab here to demonstrate contamination to the owner. Hand them the meter while it reads. Let the number do the selling. These surfaces reliably read 200-2000+ RLU on an uncleaned machine.',
    locations: [
      {where: 'Water trough interior', why: 'Biofilm accumulates here constantly. The wet, dark environment is ideal for bacterial growth. Almost always the highest reading in the machine.'},
      {where: 'Inside ice bin near drain', why: 'Moisture trap. Ice runoff pools here. Rarely if ever cleaned by staff. High reading guaranteed on any machine not professionally serviced.'},
      {where: 'Ice scoop handle', why: 'Hands touch this hundreds of times a day. Cross-contamination point. Visually clean but microbiologically dirty.'},
      {where: 'Underside of water curtain', why: 'Dark, permanently wet, never visible to staff. Biofilm (pink/orange slime) often starts here.'},
      {where: 'Float valve housing', why: 'Scale and organic material concentrate around the float. Hard to reach so never cleaned during routine staff wipe-downs.'},
    ],
    pitch: 'Snap the Ultrasnap. Swab the trough with a firm Z-pattern. Insert into Ensure v2. Hand the meter to the owner. Say nothing. Let them watch the number. Anything above 100 RLU is your sales pitch. Above 500 is dramatic. Above 1000 is a close.',
  },
  atp_compliance: {
    title: '\u2705 Compliance ATP Protocol — Where to Swab AFTER Cleaning',
    purpose: 'Swab here for your service report and compliance sticker. These surfaces hold sanitizer best and will read lowest post-clean. Document both pre and post readings — the delta is your proof of value.',
    locations: [
      {where: 'Center of evaporator plate', why: 'Primary ice contact surface. You just descaled and sanitized this. Should read under 5 RLU after proper service. This is your official compliance reading.'},
      {where: 'Water distribution tube exterior', why: 'Clean metal surface that holds sanitizer well. Consistent low readings post-clean.'},
      {where: 'Upper bin interior wall', why: 'Less contact than bottom, holds sanitizer well. Good secondary compliance swab.'},
    ],
    report_format: 'Pre-clean: [RLU] | Post-clean: [RLU] | Delta: [difference]. The delta is your value proof. Example: Before 847 RLU \u2192 After 6 RLU. Frame this on every report.',
    sticker: 'Apply compliance sticker near machine front or inside door. Note post-clean RLU and date. This is what an inspector sees immediately.',
  }
};

const DEEP_CLEAN = {
  'Manitowoc': {
    time: '90-120 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon water (16oz if heavy scale)', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon water', 'Spray bottles for both solutions', 'Separate 5-gallon bucket for soak'],
    steps: [
      {
        title: 'Safety & Setup',
        detail: 'Power off the machine at the switch AND unplug from wall. Never work on a powered machine. Place "Out of Service" sign on front. Put on nitrile gloves — the cleaner is acidic and will irritate skin. Set your cart with two labeled spray bottles (one CLEANER, one SANITIZER), soft brush kit, microfiber towels, and your screws bin. Mix your cleaner solution now: 8oz Nu-Calgon Nickel-Safe into 1 gallon warm water in your 5-gallon bucket. If the machine has not been serviced recently or you can see heavy white/gray scale, use 16oz per gallon. Scale looks like chalky white mineral deposits — think hard water stains on a shower. Mix sanitizer in a separate spray bottle: 1oz Nu-Calgon No-Rinse per 1 gallon water.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Open the front panel (usually 2-4 screws, or lift-and-pull). Stand back and look at the whole interior before touching anything. Tap the YouTube "Identify Components" button above and watch 60 seconds — then come back here. What you will see: (1) EVAPORATOR — fills the upper back half of the machine. Think of a car radiator — vertical metal fins or a flat plate, silver and shiny when clean. Scale makes it look chalky white or gray. Biofilm (bacteria) looks like pink, orange, or brown slime. This is the most important surface. (2) WATER CURTAIN — a clear or milky white plastic flap, about the size of a sheet of paper, hanging vertically in front of the evaporator like a shower curtain. It usually just unclips or lifts off. (3) WATER DISTRIBUTION TUBE — a horizontal white or gray plastic tube running across the top of the evaporator. Has small holes on the underside that drip water down. Think of a soaker hose. Pull it straight out. (4) WATER TROUGH — the shallow rectangular pan at the very bottom of the machine. Holds water that gets pumped up to the distribution tube. This is almost always the dirtiest part — biofilm loves it. (5) FLOAT VALVE — a small plastic ball on a wire or arm, sitting in one corner of the trough. Works exactly like the float in a toilet tank — controls the water level. Scale on the float causes it to stick. Take a photo of everything before removing anything.',
      },
      {
        title: 'Disassemble Removable Components',
        detail: 'Remove in this order, placing all parts in your 5-gallon bucket with cleaner solution to soak: (1) Water curtain — lift up and unhook from clips. It is usually just plastic clips, no tools needed. (2) Water distribution tube — pull straight out or unscrew end cap depending on model. Note which end faces which direction. (3) Float valve assembly — gentle pull, may have a small clip. (4) Water pump inlet screen if visible — small mesh screen, pull straight off. Place ALL plastic parts into your bucket of cleaner solution. Let them soak for the entire time you work on the evaporator — minimum 10 minutes.',
      },
      {
        title: 'Clean the Evaporator',
        detail: 'The evaporator is the most important surface. Using your spray bottle of cleaner solution, thoroughly wet all evaporator plates. Let the solution sit for 5 minutes — do not scrub immediately, let it work. You may see fizzing or bubbling on scale deposits — this is normal and means it is working. After 5 minutes, use your soft brush (NOT scrub pads — never abrasive on evaporator) to gently scrub top to bottom, working the solution into any scale deposits. Rinse with clean water from your spray bottle. Inspect: healthy evaporator is shiny and reflective. If you still see dull white deposits, apply cleaner again and wait 5 more minutes. Repeat until surface is clean. For very stubborn scale you can apply cleaner full strength (no dilution) directly on the deposit with a brush — let sit 10 min then rinse. Rinse thoroughly — no cleaner residue before sanitizing.',
      },
      {
        title: 'Clean the Water Trough & Interior',
        detail: 'The water trough is where biofilm loves to grow. Biofilm is the pink/orange/brown slime you may see — it is a bacterial colony and is exactly what inspectors look for. Pour some cleaner solution into the trough. Use your soft brush to scrub all surfaces including corners and the drain plug area. Use a scrub pad on stubborn deposits in the trough — it is okay to be more aggressive here than on the evaporator. Check the drain — clear any debris. Wipe down all interior walls, ceiling, and side panels with a cleaner-soaked microfiber towel. Check where the water pump sits — biofilm often hides underneath.',
      },
      {
        title: 'Clean Soaked Components',
        detail: 'Pull your soaked parts from the bucket. The cleaner should have loosened deposits significantly. Scrub each piece: Water curtain — scrub both sides, pay attention to the hinge/clip areas where biofilm accumulates. Distribution tube — use a small brush or pipe cleaner through the holes to clear any blockage. Each hole should spray freely — hold it up to light to verify. Float valve — scrub the ball/arm and housing. Check that the float moves freely. Rinse all parts thoroughly under clean water until no cleaner smell remains. Cleaner residue left on parts will interfere with sanitizer.',
      },
      {
        title: 'Clean Ice Bin & Exterior',
        detail: 'The ice bin is a food contact surface — treat it accordingly. Wipe down all bin interior surfaces with cleaner solution using microfiber towels. Pay attention to the bottom corners and drain area — biofilm pools here. Check the bin drain is clear. Clean the ice scoop and scoop holder — these are often the most contaminated surfaces because hands touch them constantly. Clean the bin door gasket if present. For exterior: wipe down all stainless surfaces. Clean the air filter — it is usually behind a front or side panel, a mesh or foam pad. Rinse it under water, let dry, or blow out with your air compressor. A clogged filter causes the machine to overheat and produce less ice.',
      },
      {
        title: 'Rinse Everything',
        detail: 'This step is critical. Any cleaner residue left in the machine will neutralize your sanitizer and leave a chemical taste in the ice. Rinse the evaporator with clean water using your spray bottle. Rinse the trough — pour clean water in and drain. Rinse all reinstalled or ready-to-reinstall components under clean water. Wipe all surfaces with a clean damp microfiber towel. If you can still smell the Nu-Calgon cleaner, rinse again. The machine should smell neutral before you sanitize.',
      },
      {
        title: 'Reassemble Components',
        detail: 'Reinstall in reverse order: (1) Float valve — snap back into place, verify it moves freely up and down. (2) Water pump screen — push back on firmly. (3) Distribution tube — slide back in with the same orientation as removed, holes facing down toward evaporator. (4) Water curtain — hook back onto clips, should hang freely in front of evaporator. Reconnect water supply if you disconnected it. Do not power on yet.',
      },
      {
        title: 'Sanitize All Surfaces',
        detail: 'Mix fresh sanitizer: 1oz Nu-Calgon No-Rinse per gallon water in your spray bottle. Spray generously on: evaporator plates (top to bottom), water trough interior, water curtain both sides, distribution tube exterior, all bin interior surfaces, bin door and gasket. DO NOT RINSE. The no-rinse formula is designed to air dry and leave a protective sanitizing film. Do not wipe dry — air dry only. This is the step that provides the antimicrobial protection between services.',
      },
      {
        title: 'ATP Testing',
        detail: 'Before powering on, run your ATP test. Using your Hygiena Ensure v2 and Ultrasnap testers: Snap the Ultrasnap tube to release the reagent. Swab a 10cm² area (about the size of your palm) on the evaporator using a firm Z-pattern — 5 strokes across, 5 strokes down. Insert swab into Ensure v2. Read in 15 seconds. PASS: 10 RLU or below on ice contact surfaces (evaporator, trough). PASS: 30 RLU or below on general surfaces (bin). FAIL: Above threshold — re-clean the area, re-sanitize, and retest. Note your RLU reading — this goes on the service report and compliance sticker.',
      },
      {
        title: 'Restart & Verify',
        detail: 'Reconnect power. Power on machine. Manitowoc will run an automatic startup cycle — do not interrupt. The first harvest cycle takes 15-30 minutes. Watch for: water flowing through distribution tube (should see even water across evaporator), ice beginning to form on evaporator plates (thin clear sheet), and clean harvest into bin. DISCARD the first full bin of ice — it may contain residual cleaning chemicals and should never be served. Verify the machine returns to normal operation with consistent ice production before leaving.',
      },
      {
        title: 'Document & Compliance Sticker',
        detail: 'Fill out your service report: date, tech name, machine brand/model, chemicals used, RLU reading, condition notes, next service date (60 days). Apply your compliance sticker in a visible location — note the RLU score and date. This sticker is what an inspector sees immediately and signals a professionally maintained machine. Leave a copy of the service report with the manager. Brief the manager on what you found, what you did, and when you will return.',
      },
    ]
  },
  'Hoshizaki': {
    time: '90-120 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon — CRITICAL: Hoshizaki evaporators are stainless steel. Any commercial ice machine cleaner works. Nickel-Safe is correct.', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon', 'Extra rinse water — Hoshizaki requires more thorough rinsing than other brands'],
    steps: [
      {
        title: 'Safety & Setup — Hoshizaki Specific',
        detail: 'Power off and unplug. IMPORTANT Hoshizaki note: their evaporators are stainless steel — more durable than other brands and compatible with any commercial ice machine cleaner. Nickel-Safe cleaner (what you have) is fine and will not damage the evaporator. However, avoid abrasive scrubbers on the evaporator surface — use soft brush only to prevent scratching. The damage from wrong chemicals is irreversible and expensive. Set up cart, mix cleaner (8oz per gallon warm water), mix sanitizer (1oz per gallon). Gloves on.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above and watch 60 seconds to orient yourself, then come back. Hoshizaki looks noticeably different from Manitowoc inside — do not assume the layout is the same. Open the front panel. What you will see: (1) EVAPORATOR — Hoshizaki makes crescent-shaped cubes. The evaporator looks like a curved metal plate or a grid of individual cup-shaped molds arranged in rows, taking up the upper section. Unlike Manitowoc which is a flat plate, this has a 3D curved shape. It is stainless steel — should look uniformly silver. Dark spots or rough texture means scale. (2) SPRAY BAR — THIS IS THE KEY DIFFERENCE FROM OTHER BRANDS. A horizontal tube, about the diameter of a pen, running across the top with tiny holes pointing down at the evaporator. When the machine runs you can see it spraying water. Pull it out of its clips — it just slides out. Hold it up to a light and look through each hole. Every hole should show a clear circle of light. (3) FLOAT SWITCH — a small white or gray cylinder about the size of a AA battery, hanging on a wire in the water reservoir. Push it up and down — it should slide freely. (4) WATER TROUGH — shallower than Manitowoc, at the base. In Florida heat this is where pink/orange biofilm (Serratia marcescens bacteria) grows most aggressively. (5) CURTAIN — clear plastic panel clipped to the front of the evaporator. Take a photo before touching anything.',
      },
      {
        title: 'Disassemble Components',
        detail: 'Remove carefully — Hoshizaki components are generally less rugged than Manitowoc and can crack if forced: (1) Curtain assembly — unclip from front of evaporator, set aside. (2) Spray bar — pull straight out from mounting clips. Hold up to light and look through the holes — each hole should be clear. Note any that are blocked. (3) Float switch — gently pull off the mounting post. The float should slide freely on the wire. (4) Water pump screen — pull off the pump inlet. Soak all plastic components in cleaner solution (8oz per gallon) in your 5-gallon bucket for minimum 10 minutes.',
      },
      {
        title: 'Clean Spray Bar Holes — Critical Step',
        detail: 'The spray bar holes are the #1 maintenance issue on Hoshizaki machines. Blocked holes cause uneven water distribution which creates irregular or incomplete crescent cubes — a telltale sign to inspectors of poor maintenance. While components are soaking, use a toothpick or the smallest brush in your kit to clear each hole on the spray bar. Push through from the inside out. Hold up to light to confirm clear. There are typically 15-25 holes on a standard spray bar. If a hole is completely blocked and toothpick cannot clear it, a thin wire or unfolded paper clip works. Rinse the bar thoroughly after clearing.',
      },
      {
        title: 'Clean Evaporator — Gentle Technique Required',
        detail: 'Apply cleaner solution to evaporator using spray bottle. Let sit 5 minutes. Use ONLY the softest brush in your kit — NEVER use scrub pads, steel wool, or abrasive materials on Hoshizaki evaporator. Even though it is stainless steel, scratches create micro-grooves that harbor bacteria. Keep technique gentle and use circular motions with light pressure. Use soft circular motions with light pressure. Rinse with clean water. Inspect: should be uniformly silver with no dull spots. Repeat cleaner application if needed. For stubborn deposits apply cleaner with brush and let sit 10 minutes before gentle scrubbing. The key word on Hoshizaki is GENTLE throughout.',
      },
      {
        title: 'Clean Float Switch & Trough',
        detail: 'Float switch: wipe the float and wire with cleaner-soaked cloth. Scale on the float causes it to stick, leading to overflow or underfill. Verify the float slides freely up and down the wire after cleaning. Water trough: scrub with soft brush and cleaner solution, paying attention to corners where biofilm accumulates. Check the drain. Wipe down all interior surfaces. Interior walls often show pink biofilm on Hoshizaki units in Florida heat — this is Serratia marcescens, a common waterborne bacteria. Scrub thoroughly with cleaner-soaked scrub pad.',
      },
      {
        title: 'Reassemble & Rinse Thoroughly',
        detail: 'Rinse all soaked components under clean water until no cleaner odor. Reinstall: (1) Float switch — back on mounting post. (2) Water pump screen — back on inlet. (3) Spray bar — back in clips, holes facing evaporator. (4) Curtain assembly — clip back onto evaporator front. THOROUGH RINSE CRITICAL on Hoshizaki: spray clean water on evaporator, flush trough with clean water, wipe all surfaces. Hoshizaki machines are more sensitive to cleaner residue than other brands — any residue left will affect ice taste and sanitizer effectiveness.',
      },
      {
        title: 'Sanitize, ATP Test & Restart',
        detail: 'Spray sanitizer (1oz per gallon) on all surfaces: evaporator, spray bar exterior, trough, curtain, bin interior. Air dry — do not wipe. ATP test on evaporator surface and trough. Pass: 10 RLU or below on ice contact surfaces. Fail: re-clean and retest. Power on — Hoshizaki will run a startup cycle. First harvest cycle on a crescent machine takes 20-35 minutes. Watch that the spray bar is spraying evenly across all evaporator molds. Discard first bin of ice. Verify crescent cubes are uniform in size and shape — irregular cubes mean water distribution issue, recheck spray bar.',
      },
    ]
  },
  'Scotsman': {
    time: '75-105 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon', 'Scotsman has an automated clean cycle — chemicals go into reservoir, not spray bottle'],
    steps: [
      {
        title: 'Safety & Scotsman Controls',
        detail: 'Power off and unplug. Scotsman Prodigy machines have a digital control panel — before unplugging, note any error codes displayed. Write them down. Common codes: E1 (harvest time too long — scale on evaporator), E2 (freeze time too long — refrigeration issue), E8 (water level — float or inlet issue). These codes tell you what to focus on during cleaning. Unplug machine. Gloves on. Remove all ice from bin.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above and watch 60 seconds first. Scotsman Prodigy is the most visually distinct machine — it looks different from both Manitowoc and Hoshizaki. Open the front panel. What you will see: (1) EVAPORATOR PLATE — Scotsman is unique because ice forms on a large VERTICAL flat plate on the back wall, not a horizontal hanging plate like Manitowoc. The plate is roughly the size of a laptop screen. When clean, it is smooth and reflective like a mirror. White crusty patches are scale. Pink or orange film is biofilm. (2) ICE THICKNESS SENSOR — THIS IS SCOTSMAN\u2019S CRITICAL COMPONENT. A small probe about the diameter of a pencil, sticking out horizontally near the middle of the evaporator. It is usually black or dark gray. The tip points toward the evaporator plate with a gap of about 3/8 inch — roughly the thickness of a pencil. You can check this gap by holding a pencil horizontally between the probe tip and the plate. Scale on this probe changes the gap and causes harvest problems. Wipe it clean at every single visit. (3) CONTROL PANEL — Scotsman Prodigy has a digital display on the front. Always check for error codes before unplugging. E1 = harvest too long (scale). E2 = freeze too long (refrigeration issue — not your problem to fix). Write down any codes. (4) DRAIN PAN — at the bottom, under the evaporator. Scotsman drain lines block with grease more than other brands. Pour a cup of water in and time it — should drain in under 30 seconds. (5) WATER CURTAIN — flat plastic panel in front of the evaporator, just unclips.',
      },
      {
        title: 'Disassemble Components',
        detail: 'Remove: (1) Water curtain — unclip from mounting brackets. (2) Ice thickness sensor — note its position carefully before removing. It is a small probe on a wire. Gently pull straight out. (3) Water distribution components if accessible. Soak plastic parts in cleaner solution. While soaking, inspect the drain pan and drain line — pour a cup of water in the drain pan and watch it drain. Should drain within 30 seconds. If slow, the drain line is blocked. Clear blockage with a thin brush or flexible cleaning brush from your kit.',
      },
      {
        title: 'Clean Ice Thickness Sensor — Critical',
        detail: 'The ice thickness sensor is Scotsman-specific and critical. Scale buildup on the sensor tip causes premature or incomplete harvest cycles — either ice falls off too early (thin, watery ice) or stays too long (frozen solid, bridged). Using a soft cloth dampened with cleaner solution, wipe the sensor tip. Do not bend the probe. Do not use abrasive materials. The tip should be clean metal with no white deposits. While cleaning, verify the mounting position — the gap between probe tip and evaporator surface should be approximately 3/8 inch. You can gauge this with a pencil held sideways — roughly the thickness of a pencil. If the gap is wrong (sensor was bent or moved) ice quality will be affected.',
      },
      {
        title: 'Clean Evaporator Plate',
        detail: 'Apply cleaner solution (8oz per gallon) to evaporator plate from top to bottom using spray bottle. Let sit 5 minutes. Vertical evaporator means cleaner will run down — that is fine and actually helps it penetrate scale. Scrub with soft brush top to bottom, paying attention to the area directly across from the thickness sensor where ice always contacts the plate. Rinse with spray bottle of clean water. Repeat if scale remains. Scotsman evaporators can tolerate slightly more scrubbing pressure than Hoshizaki but still avoid abrasive pads directly on the plate surface.',
      },
      {
        title: 'Clean Trough, Drain & Interior',
        detail: 'Trough: scrub with cleaner and brush. Check float — should move freely. Interior walls: wipe all surfaces with cleaner-soaked microfiber. Drain system: this is especially important on Scotsman. Pour cleaner solution down the drain line. Use your flexible brush to scrub the drain pan interior. Verify drain flows freely. Scotsman units in Florida kitchens get grease in the drain line which traps biofilm. If you find significant grease buildup note this in your service report.',
      },
      {
        title: 'Reassemble, Rinse & Sanitize',
        detail: 'Reinstall all components. IMPORTANT: reinstall ice thickness sensor in exact position — gap should be 3/8 inch from evaporator. Rinse all surfaces with clean water. Spray sanitizer on evaporator, curtain, trough, bin interior. Air dry. ATP test on evaporator and trough. Pass: 10 RLU or below. Power on. Scotsman Prodigy will run a startup diagnostic — allow it to complete without interruption. First harvest typically 25-35 minutes. Discard first bin. Watch that ice forms uniformly across the evaporator plate — thin patches indicate a distribution or refrigeration issue.',
      },
    ]
  },
  'Ice-O-Matic': {
    time: '75-90 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {
        title: 'Safety & Setup',
        detail: 'Power off and unplug. Ice-O-Matic units are common in fast food and cafe environments and often have heavier grease contamination than restaurant machines. This affects the exterior and air filter particularly. Gloves on. Mix cleaner 8oz per gallon warm water. Mix sanitizer 1oz per gallon. Remove all ice.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above first. Ice-O-Matic looks similar to Manitowoc internally but has one key difference that matters most. Open the front panel. What you will see: (1) EVAPORATOR — horizontal plate at the top, similar position to Manitowoc. Nickel-plated. Should be shiny silver when clean. (2) WATER DISTRIBUTION PAN — THIS IS IOM\u2019S DEFINING FEATURE. Unlike Manitowoc tube, IOM uses a flat rectangular plastic pan sitting directly above the evaporator with tiny holes on the bottom face. It looks like a plastic tray with holes punched in it, about the size of a small cutting board. Remove it by lifting straight up. Hold it up to a window or flashlight — every single hole should show a clear circle of light through it. Holes blocked with white mineral deposits look dark or cloudy. These holes are smaller than Manitowoc and block faster. Clearing them is the most important thing you do on an IOM machine. (3) BIN BAFFLE — a plastic divider panel inside the ice bin, usually white, that guides ice as it falls. Flip it over and look at the underside — this dark hidden surface grows biofilm that the owner never sees. (4) WATER TROUGH — base reservoir, scrub thoroughly. (5) WATER CURTAIN — clear plastic panel in front of evaporator, just unclips.',
      },
      {
        title: 'Disassemble Components',
        detail: 'Remove: (1) Water curtain — unclip. (2) Bin baffle — usually just lifts out. Check underside for pink/orange biofilm. (3) Distribution pan — lift out. (4) Water pump screen. Place all in cleaner soak. IMMEDIATELY inspect distribution pan holes while fresh — hold up to light and look through each hole. They should all be clear circles. Blocked holes appear dark or have white mineral plugs. Use a toothpick to clear each blocked hole now before soaking. Post-soak clearing is harder.',
      },
      {
        title: 'Clear Distribution Pan Holes — Most Important IOM Step',
        detail: 'This is the defining maintenance step on Ice-O-Matic units. Blocked distribution holes cause uneven water across the evaporator which creates scale that forms unevenly — leading to an irregular freeze pattern and poor ice quality. After soaking for 10 minutes in cleaner solution, remove pan and use toothpick on every hole. There are typically 20-40 holes on a standard unit. Every hole must be clear. After clearing, hold the pan up to a light source and look through from the bottom — every hole should show a pinpoint of light. If any remain blocked, use a thin wire. Rinse pan completely under clean water.',
      },
      {
        title: 'Clean Evaporator & Interior',
        detail: 'Apply cleaner solution to evaporator. Let sit 5 minutes. Scrub with soft brush. Scale on IOM evaporators often forms in strips corresponding to blocked distribution holes — you can actually diagnose which holes were blocked by where the scale is heaviest. Rinse thoroughly. Wipe all interior walls. Trough: scrub with brush and cleaner, clean drain. Interior inspection: IOM units in fast food get grease from cooking on interior surfaces — use your scrub pad on interior walls if grease is present. Air filter: remove (usually clips off front panel), rinse or blow out with air compressor.',
      },
      {
        title: 'Reassemble, Sanitize & Test',
        detail: 'Reinstall all components. Rinse all surfaces with clean water. Spray sanitizer on all surfaces, air dry. ATP test on evaporator and trough. Pass: 10 RLU. Power on. First harvest 15-25 minutes. Watch that water flows through ALL distribution holes evenly — you should see an even sheet of water across the evaporator. Uneven water = holes still blocked. Discard first bin of ice.',
      },
    ]
  },
  'Follett': {
    time: '60-90 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon — Follett uses stainless steel evaporator, Nickel-Safe is appropriate', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {
        title: 'Safety & Follett-Specific Setup',
        detail: 'Power off and unplug. Follett nugget ice machines work fundamentally differently from cube ice machines — they use a rotating AUGER inside an evaporator CYLINDER to continuously extrude nugget ice rather than a batch freeze/harvest cycle. This changes the cleaning process significantly. Gloves on. Mix chemicals. Remove all ice from storage bin. Listen to the machine before unplugging — unusual grinding from the auger indicates scale on cylinder walls or auger wear. Note this in your service record.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above first — Follett works completely differently from all other brands and a video helps enormously. Remove the top panel (4 screws). What you will see: (1) EVAPORATOR CYLINDER — a vertical stainless steel tube, roughly the diameter of a large coffee can (6-8 inches). This is the entire ice-making mechanism. Ice does not form on a plate — it forms as a thin layer on the INSIDE WALL of this cylinder and gets scraped off continuously by the auger. (2) AUGER — a metal helical screw (think of a giant corkscrew or drill bit) inside the cylinder. You cannot see it until you remove it — it fills the entire inside of the cylinder. The auger turns slowly and continuously, scraping ice off the cylinder wall and pushing it upward out the top. You will hear it humming when the machine runs. BEFORE UNPLUGGING: listen. Smooth hum = healthy. Grinding or clicking = scale buildup on cylinder, will need attention. (3) AUGER MOTOR — sits on top of the cylinder, usually a black box with a shaft going down. This is what drives the auger. (4) ICE CHUTE — a rectangular plastic channel leading from the cylinder top to the storage bin. Ice travels through here continuously. This gets biofilm because it is always wet and dark. (5) STORAGE BIN — the hopper below where Chewblet nugget ice collects. Soft, moist nuggets = machine running well. Hard or irregular nuggets = auger or cylinder issue.',
      },
      {
        title: 'Access Evaporator Cylinder',
        detail: 'The auger must be removed to properly clean the cylinder interior on most Follett models. Disconnect the auger drive coupling at the top (typically a hex bolt — your impact drill is useful here). Lift the auger straight up and out of the cylinder. It is heavy — 10-20 lbs depending on model. Lay auger flat on a clean surface. Inspect auger: fins should be intact and smooth. Pitting or cracking means replacement is needed — note in your report. Scale on auger fins appears as rough white deposits.',
      },
      {
        title: 'Clean Evaporator Cylinder Interior',
        detail: 'Pour cleaner solution (8oz per gallon) into the cylinder — fill approximately 1/3. Use your long brush to scrub the interior cylinder walls with circular strokes. Scale will be on the lower portion where ice forms. The solution will fizz on scale deposits — this is the acid working. Let solution sit 10 minutes. Scrub again. Drain by tilting machine slightly or using a cup to scoop out solution. Rinse with clean water poured into cylinder. Repeat rinse until no cleaner smell. Inspect cylinder wall with a flashlight — should be smooth and shiny throughout.',
      },
      {
        title: 'Clean Auger',
        detail: 'While cylinder is draining/rinsing, clean the auger. Spray or brush cleaner solution onto all auger fins. Let sit 5 minutes. Scrub with soft brush — get between the fins where scale accumulates. Rinse thoroughly. Inspect the tip of the auger (the pointed bottom end) — this is where it contacts the cylinder wall and scale wears it down. Note unusual wear. Verify auger fins are not bent or cracked. A bent fin will score the cylinder interior and create a pathway for bacterial growth.',
      },
      {
        title: 'Clean Bin, Chute & Dispenser',
        detail: 'Bin interior: scrub with cleaner solution. Pay attention to waterline area where biofilm concentrates. Drain: check for blockage, flush with clean water. Ice chute: use your smaller brush to scrub the chute interior — ice travels through here continuously and biofilm forms. On Symphony units: clean dispenser nozzle and actuator arm with cleaner-soaked cloth. The dispenser is a high-touch surface and should be sanitized thoroughly.',
      },
      {
        title: 'Reassemble, Sanitize & Test',
        detail: 'Reinstall auger into cylinder — lower straight down, engaging the drive coupling. Reconnect drive coupling bolt (snug, not overtightened — your drill on low torque or by hand). Rinse all surfaces with clean water. Spray sanitizer into cylinder, on auger, chute, bin, and dispenser. Air dry. ATP test: swab the ice chute and cylinder wall. Pass: 10 RLU. Power on. Follett machines take 15-30 minutes to reach full ice production — the cylinder must re-chill before nuggets form. Discard first production cycle. Verify nuggets are soft and consistent — hard or irregular nuggets indicate cylinder or auger issue.',
      },
    ]
  },
  'Cornelius': {
    time: '60-90 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 8oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {
        title: 'Safety & Setup',
        detail: 'Power off and unplug. Cornelius units are compact countertop or undercounter machines common in convenience stores and small cafes. They have simpler internal layouts than modular units but the cleaning process is equally important. Gloves on. Mix chemicals. Remove ice. Cornelius machines in convenience stores often have heavier use and more contamination than restaurant units — the ice is frequently touched by customers or used in fountain drinks where the machine runs constantly.',
      },
      {
        title: 'Identify Your Components — What You Are Actually Looking At',
        detail: 'Tap "YouTube: Identify Components" above first. Cornelius is the simplest machine internally — good news for a first-timer. Open or remove the front panel (usually 2-4 screws or a snap-off panel, varies by model). What you will see: (1) FREEZE PLATE — the evaporator on a Cornelius is simpler than modular brands. It is a flat metal plate, usually positioned vertically or at an angle near the back. Smaller than a Manitowoc evaporator — roughly the size of a hardcover book. Should be smooth and reflective when clean. White crusty patches are scale. (2) AIR FILTER — FIND THIS FIRST. Usually a foam or mesh pad behind a removable grille on the side or rear of the unit. Pull it out and hold it to light. If you cannot see through it, it is clogged. A blocked filter is the #1 reason Cornelius machines underperform. Many owners have never cleaned it. (3) WATER TROUGH — small shallow reservoir at the base. (4) DRAIN PAN — under or beside the trough. In convenience store environments this often has a sour smell from biofilm in the drain. Smell it before cleaning — if odorous, scrub extra thoroughly. (5) ICE BIN — the storage compartment below. Simpler layout with fewer places for biofilm to hide than modular machines — but check corners and the drain area.',
      },
      {
        title: 'Disassemble Accessible Components',
        detail: 'Remove what is accessible: water curtain if present (some models), any removable troughs or pans. Cornelius units have fewer removable components than larger machines — most cleaning is done in-place. Soak any removable parts in cleaner solution. IMMEDIATELY address the air filter: remove it (slide out from rear or side), hold it up to light — if you cannot see through it easily it is heavily clogged. Rinse under water. If solid with dust/grease use your air compressor to blow from inside out. Let dry. A blocked air filter causes the condenser to overheat, reducing ice production and stressing the compressor.',
      },
      {
        title: 'Clean Freeze Plate',
        detail: 'Apply cleaner solution to freeze plate using spray bottle. Let sit 5 minutes. Scrub with soft brush using horizontal strokes. Cornelius freeze plates are typically nickel-plated or stainless — compatible with Nickel-Safe. Scale on these smaller plates tends to be concentrated where water distributes. Rinse thoroughly with clean water. Inspect — should be uniformly shiny. Repeat if dull areas remain. For heavy scale: apply cleaner solution with brush, cover with a damp cloth to hold moisture, let sit 15 minutes, scrub and rinse.',
      },
      {
        title: 'Clean Trough, Drain Pan & Interior',
        detail: 'Water trough: scrub all surfaces with cleaner and brush. The trough drain is a common odor source on Cornelius units — pour cleaner solution into drain and let sit 5 minutes. Flush with clean water. Drain pan: this is where Cornelius units develop the characteristic convenience store ice machine odor. Scrub pan with scrub pad and cleaner. If there is significant slime or pink biofilm this is Serratia marcescens — be thorough. Interior walls: wipe all accessible surfaces with cleaner-soaked microfiber. Exterior: wipe stainless surfaces. Clean around water inlet valve area.',
      },
      {
        title: 'Reassemble, Sanitize & Test',
        detail: 'Reinstall all components. Make sure air filter is dry before reinstalling. Rinse all interior surfaces with clean water. Spray sanitizer on freeze plate, trough interior, drain pan, and bin interior. Air dry. ATP test: swab freeze plate and trough. Pass: 10 RLU on ice contact surfaces. Power on. Cornelius machines restart quickly — typically 10-20 minutes to first ice. Discard first production. Verify ice is clear (not cloudy) — cloudy ice indicates mineral contamination in water, possible filter issue. Note in report.',
      },
    ]
  },
};

const MAINTENANCE_60 = {
  'Manitowoc': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon (half strength for maintenance)', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon', 'Spray bottles for both'],
    steps: [
      {title:'Visual Inspection & Machine Check', detail:'Power stays ON for 60-day maintenance — you are inspecting and spot-cleaning, not doing a full chemical cycle. Observe the machine operating: listen for unusual sounds (grinding = scale on evaporator, gurgling = water level issue, clicking = harvest problem). Check ice in bin — should be clear, consistent size cubes. Cloudy or malformed ice indicates a problem. Open front panel. Check evaporator for new scale deposits since last service. Light white haze is normal mineral accumulation — heavier deposits need attention. Check for any pink/orange biofilm starting in corners or on curtain.'},
      {title:'Spot Clean Evaporator if Needed', detail:'If minor scale or light biofilm is visible: mix half-strength cleaner (4oz per gallon). Spray on affected areas only. Let sit 3 minutes. Wipe with soft brush or microfiber. Rinse with damp cloth. If scale covers more than 25% of evaporator surface or biofilm is present beyond light discoloration — elevate to full deep clean, do not proceed with 60-day maintenance only.'},
      {title:'Clean Water Trough & Float', detail:'The trough accumulates mineral deposits even between full cleanings. Wipe trough interior with cleaner-soaked microfiber. Check float valve — should move freely up and down. Light scale on float: wipe with cleaner cloth. Float stuck or heavily scaled: flag for deep clean. Check drain flows freely — pour small cup of water in trough and verify it drains.'},
      {title:'Inspect & Clear Distribution Tube', detail:'Check water distribution tube holes visually — look for any starting to cloud over. Use toothpick to clear any partially blocked holes. Verify water flows evenly across evaporator during machine operation. Uneven water means a distribution issue — note for next deep clean.'},
      {title:'Air Filter & Condenser', detail:'Remove air filter. If dusty but not clogged, blow out with air compressor (always blow from inside out to avoid pushing debris deeper). If heavily clogged, rinse with water and dry before reinstalling. Check condenser coils (usually visible behind or under air filter location) — dust on coils reduces efficiency. Use your condenser brush to gently clean coil fins if buildup is visible. Be gentle — coil fins bend easily.'},
      {title:'Bin Wipe & Sanitize', detail:'Wipe bin interior with sanitizer solution (1oz per gallon). Do not rinse. Check scoop and scoop holder — wipe both with sanitizer. Apply sanitizer spray to evaporator, trough, and curtain. Air dry. No rinsing.'},
      {title:'ATP Spot Test & Document', detail:'Run quick ATP on water trough surface. Pass: 10 RLU. If fail: do a more thorough spot clean and retest. If still failing: elevate to full deep clean today. Note ATP reading, date, any observations (new sounds, ice quality changes, scale level). Update next service date — 60 days from today. If the machine is approaching 6 months since last deep clean, schedule deep clean for next visit.'},
    ]
  },
  'Hoshizaki': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon (half strength)', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Visual & Ice Quality Check', detail:'Machine stays on. Observe ice: Hoshizaki crescent cubes should be uniform C-shapes. Irregular shapes (partial crescents, flat pieces, stuck-together cubes) indicate spray bar blockage or water distribution issue. Check ice is clear — cloudy crescent cubes indicate mineral contamination. Open panel. Inspect spray bar holes visually.'},
      {title:'Spray Bar Inspection & Clearing — Priority Step', detail:'The spray bar is the most important 60-day maintenance item on Hoshizaki. Remove spray bar (pull from clips). Hold to light — look through each hole. Any that appear cloudy or dark: use toothpick to clear. Rinse bar under water. Reinstall with holes facing evaporator. This single step prevents most common Hoshizaki ice quality issues.'},
      {title:'Float Switch Check', detail:'Check float switch — gently push float down and release. Should spring back to top position freely. If it sticks or moves slowly, mineral deposits are building. Wipe float and wire with half-strength cleaner solution. Do not use abrasive. Note: a sticking float will cause water overflow or ice production cutoff — this is a common Hoshizaki service call.'},
      {title:'Evaporator Visual Check', detail:'Look at evaporator surface. Hoshizaki stainless evaporator should be uniformly silver. Any brown, orange, or dark spots: biofilm forming. Light surface biofilm can be wiped with sanitizer solution. More than light spots: elevate to deep clean. Never use abrasive materials on the Hoshizaki evaporator — stainless scratches harbor bacteria.'},
      {title:'Bin & Sanitize', detail:'Wipe bin interior with sanitizer. Check bin drain clear. Apply sanitizer to spray bar exterior, trough, curtain. Air dry.'},
      {title:'ATP & Document', detail:'Swab trough. Pass 10 RLU. Document ATP, spray bar condition (clear/partially blocked/cleared today), evaporator condition, float switch condition. Set next service 60 days.'},
    ]
  },
  'Scotsman': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Prodigy Diagnostic Check', detail:'Before opening machine, press INFO button on Scotsman Prodigy control panel. Note any error codes displayed. E1 (harvest too long) typically means scale building on evaporator — flag for deep clean if frequent. E2 (freeze too long) is refrigeration-related — not a cleaning issue, contact a refrigeration tech. Clear minor codes after noting them.'},
      {title:'Ice Thickness Sensor — 60-Day Priority', detail:'The ice thickness sensor must be checked at every visit. Open panel. Locate the small black probe near the evaporator. Wipe probe tip with half-strength cleaner solution on a soft cloth. Check gap — should be approximately 3/8 inch from evaporator surface (thickness of a pencil). Scale buildup changes this gap. If you see white deposits on the probe tip, clean thoroughly. This sensor controls the entire harvest cycle — a dirty sensor causes premature harvest (thin ice) or delayed harvest (bridged ice).'},
      {title:'Water Curtain & Drain', detail:'Check water curtain for scale or biofilm. Wipe with cleaner if needed, rinse, sanitize. PRIORITY: check Scotsman drain system. Pour 1 cup of water in drain pan — time how long it takes to drain. Should drain in under 30 seconds. 30-60 seconds: partial blockage, flush drain line. Over 60 seconds: full blockage, clear drain line with flexible brush. Drain blockage is the #1 service call issue on Scotsman units.'},
      {title:'Air Filter & Exterior', detail:'Remove air filter. Rinse or blow out with compressor. Scotsman recommends monthly filter inspection in kitchen environments — if you see grease coating the filter, increase cleaning frequency and note in report. Wipe exterior stainless.'},
      {title:'Sanitize & Document', detail:'Spray sanitizer on curtain, trough, bin interior. Air dry. ATP on trough. Pass 10 RLU. Document: error codes noted, sensor condition, drain drain time, filter condition. Set next service 60 days.'},
    ]
  },
  'Ice-O-Matic': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Visual & Ice Check', detail:'Machine on. Check ice quality — IOM cubes should be consistent size and clear. Observe water flow if possible. Open panel and inspect distribution pan holes visually.'},
      {title:'Distribution Pan Holes — Priority Check', detail:'Remove distribution pan. Hold to light. Check every hole — even partial blockage affects ice quality. Clear any starting-to-block holes with toothpick now. It is much easier to clear early-stage blockage than full blockage. Rinse and reinstall.'},
      {title:'Bin Baffle & Trough', detail:'Check underside of bin baffle for biofilm. Pink or orange slime: wipe with half-strength cleaner, rinse, note. Trough: wipe with cleaner-soaked cloth. Check drain flows.'},
      {title:'Air Filter', detail:'Remove, clean with air compressor or rinse. IOM units in fast food environments get grease on filter — check monthly in these locations.'},
      {title:'Sanitize & Document', detail:'Spray sanitizer on evaporator, distribution pan (exterior), trough, bin. Air dry. ATP on trough. Pass 10 RLU. Document distribution hole condition, filter condition. Set 60-day return.'},
    ]
  },
  'Follett': {
    time: '30-45 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Auger Sound Check — Listen First', detail:'Before unplugging: power on machine and listen to auger operation. Should be smooth hum. Grinding, clicking, or irregular sounds indicate scale on cylinder walls or auger wear. Note sound quality in report — this is your early warning system. Heavy grinding: consider elevating to deep clean today.'},
      {title:'Ice Quality Check', detail:'Follett nugget ice should be soft, moist, and chewable — the signature Follett product. Hard, dry, or irregularly sized nuggets indicate auger or cylinder issue. Inspect bin for ice quality.'},
      {title:'Ice Chute & Dispenser', detail:'The ice chute is a high-frequency cleaning area — clean at every 60-day visit. Use small brush to scrub chute interior with half-strength cleaner. Rinse. On Symphony dispenser units: clean dispenser nozzle with cleaner cloth. Sanitize chute and nozzle.'},
      {title:'Bin & Drain', detail:'Wipe bin interior with sanitizer. Check drain flows. Follett bins with slow drains develop biofilm faster.'},
      {title:'Cylinder Top Access', detail:'Remove top panel and visually inspect top of auger and cylinder opening. Any visible scale or discoloration at the top: flag for deep clean. Do not attempt partial auger cleaning at 60-day service — either full deep clean or visual inspection only.'},
      {title:'Sanitize & Document', detail:'Sanitize chute, dispenser, bin. ATP on ice chute. Pass 10 RLU. Document auger sound quality, ice quality, chute condition. Set 60-day return.'},
    ]
  },
  'Cornelius': {
    time: '25-35 min',
    chemicals: ['Nu-Calgon Nickel-Safe Cleaner: 4oz per gallon', 'Nu-Calgon No-Rinse Sanitizer: 1oz per gallon'],
    steps: [
      {title:'Visual & Ice Quality', detail:'Cornelius 60-day service is simpler than larger machines. Check ice is clear — Cornelius countertop units in convenience stores sometimes produce cloudy ice due to high-use water with heavy mineral content. Cloudy ice = filter issue or very hard water. Note for report.'},
      {title:'Air Filter — Most Important Cornelius Step', detail:'The air filter is the #1 maintenance item on Cornelius units. Remove (slide out from side or rear). If you cannot see through it when held up to light: heavily clogged. Rinse under water and blow out with air compressor. If grease-coated (common in convenience store environments near food prep): may need replacement. A blocked air filter will visibly reduce ice production within weeks.'},
      {title:'Freeze Plate Spot Check', detail:'Open front panel. Look at freeze plate for new scale or biofilm since last deep clean. Light haze: normal. More than light haze: wipe with half-strength cleaner, rinse. Significant buildup: elevate to deep clean.'},
      {title:'Drain Pan Odor Check', detail:'Smell the drain pan area. Any sour or musty odor indicates biofilm in drain pan or drain line. Pour cup of water in drain — verify it drains freely. Wipe drain pan with cleaner if odor present. Rinse. This is the most common Cornelius complaint — recurring odor means deep clean schedule needs adjustment.'},
      {title:'Sanitize & Document', detail:'Wipe bin with sanitizer. Spray freeze plate and trough with sanitizer. Air dry. ATP on freeze plate. Pass 10 RLU. Document: filter condition, ice clarity, drain odor (yes/no), freeze plate condition. Set 60-day return.'},
    ]
  },
};


const VIDEO_LINKS={
  'Manitowoc':{
    label:'Manitowoc Official + WebstaurantStore Deep Clean Video',
    deep_url:'https://www.webstaurantstore.com/video-7936/manitowoc-detailed-cleaning-sanitation-indigo-nxt.html',
    guide_url:'https://www.partstown.com/cm/resource-center/guides/gd1/how-to-clean-a-manitowoc-ice-machine',
    official_url:'https://www.manitowocice.com/Videos',
    youtube_url:'https://www.youtube.com/results?search_query=Manitowoc+ice+machine+cleaning+tutorial+evaporator+components',
    yt_component:'https://www.youtube.com/results?search_query=Manitowoc+ice+machine+inside+evaporator+water+curtain+distribution+tube+identify',
  },
  'Hoshizaki':{
    label:'Hoshizaki Official Training + Parts Town Guide',
    deep_url:'https://www.hoshizakiamerica.com/support/training/',
    guide_url:'https://www.partstown.com/cm/resource-center/guides/gd2/how-to-clean-a-hoshizaki-ice-machine',
    official_url:'https://www.hoshizakiamerica.com/support/training/',
    youtube_url:'https://www.youtube.com/results?search_query=Hoshizaki+KM+ice+machine+cleaning+tutorial+spray+bar+components',
    yt_component:'https://www.youtube.com/results?search_query=Hoshizaki+ice+machine+open+inside+spray+bar+float+switch+identify+parts',
  },
  'Scotsman':{
    label:'Scotsman Official Cleaning Guide + Easy Ice Reference',
    deep_url:'https://scotsmanhomeice.com/how-to-clean-your-scotsman-nugget-ice-machine/',
    guide_url:'https://www.easyice.com/scotsman-ice-machine-cleaning/',
    official_url:'https://scotsmanhomeice.com/blog-ice-machine-cleaning/',
    youtube_url:'https://www.youtube.com/results?search_query=Scotsman+Prodigy+ice+machine+cleaning+tutorial+thickness+sensor',
    yt_component:'https://www.youtube.com/results?search_query=Scotsman+ice+machine+open+inside+evaporator+thickness+probe+identify+parts',
  },
  'Ice-O-Matic':{
    label:'Easy Ice IOM Guide + Parts Town Reference',
    deep_url:'https://www.easyice.com/ice-o-matic-ice-machine-cleaning/',
    guide_url:'https://www.partstown.com/cm/resource-center',
    official_url:'https://www.iceomatic.com/support',
    youtube_url:'https://www.youtube.com/results?search_query=Ice-O-Matic+ice+machine+cleaning+tutorial+distribution+pan+components',
    yt_component:'https://www.youtube.com/results?search_query=Ice-O-Matic+ice+machine+open+inside+distribution+pan+evaporator+identify',
  },
  'Follett':{
    label:'Follett Official Service Video Library',
    deep_url:'https://www.follettice.com/tech-support/service-video-library',
    guide_url:'https://www.partstown.com/cm/resource-center/guides/gd2/how-to-clean-a-follett-ice-machine',
    official_url:'https://www.follettice.com/tech-support/service-video-library',
    youtube_url:'https://www.youtube.com/results?search_query=Follett+ice+machine+cleaning+tutorial+auger+cylinder+nugget',
    yt_component:'https://www.youtube.com/results?search_query=Follett+Symphony+ice+machine+open+inside+auger+evaporator+cylinder+identify',
  },
  'Cornelius':{
    label:'Parts Town Commercial Ice Machine Resources',
    deep_url:'https://www.partstown.com/cm/resource-center',
    guide_url:'https://www.partstown.com/cm/resource-center',
    official_url:'https://www.cornelius.com/support',
    youtube_url:'https://www.youtube.com/results?search_query=Cornelius+commercial+ice+machine+cleaning+tutorial+components',
    yt_component:'https://www.youtube.com/results?search_query=commercial+countertop+ice+machine+open+inside+freeze+plate+identify+components',
  },
};

const OBJECTIONS=[
  {
    q:"We already have someone who does it",
    a:"Perfect  -  can I ask who and how often they come out? The reason I ask is most places we start working with had someone too, but it was either annual or on-call. The problem is Florida inspectors now specifically open the machine during inspections  -  they want to see dated service records, not just a clean exterior. If your current vendor is leaving you a dated compliance report after every visit you are set. If not, that is the gap we fill. I can show you what ours looks like  -  takes two minutes."
  },
  {
    q:"Not interested",
    a:"Fair enough. Real quick before I go  -  when was the last time someone actually opened up the machine? Not wiped it down, but opened the evaporator cover and documented what they found? Most managers I talk to have never seen inside their own machine. I am not trying to sell you anything right now  -  I can open it up in five minutes and tell you exactly what an inspector would flag. If it is clean you never hear from me again."
  },
  {
    q:"Call me back / Not a good time",
    a:"Completely understand. I am in this area Tuesdays and Thursdays  -  which works better? And honestly a two-minute conversation now saves us both a phone tag. The short version: your inspection history flagged ice machine violations and I specialize in exactly that. If it is not a fit I will tell you right now and be out of your hair."
  },
  {
    q:"We clean it ourselves / Staff handles it",
    a:"That is great  -  and I am sure the bin and scoop are spotless. The issue is the evaporator plates and water distribution system inside the unit. Those need chemical descaling and sanitizer that food handler training does not cover. More importantly  -  inspectors know the difference. They will open it and if there is scale or biofilm on the plates that is a high priority violation regardless of how clean everything else is. That is what we document and prevent."
  },
  {
    q:"How much does it cost?",
    a:"Monthly maintenance is $149 for the first machine, $89 for a second. That covers the full clean, sanitizer treatment, and a dated compliance report every visit. One failed callback costs more in lost revenue than six months of service. And if you want to try us first, we do a $99 intro visit  -  full service, full report, no commitment. If you like it, we set up monthly. If not, no problem."
  },
  {
    q:"We just had an inspection and passed",
    a:"That is actually the best possible time to start  -  you have the most runway before the next one. Here is the thing about Florida inspectors: they rotate, and new inspectors are specifically trained to look at ice machine interiors because it is one of the most commonly missed items. Passing once does not mean the next inspector sees it the same way. What we do is get you to a standard that passes regardless of who walks in the door  -  and gives you the paperwork to prove it."
  },
];

function dL(d){return d<0?Math.abs(d)+'d overdue':d===0?'TODAY':'+'+d+'d';}
function dC(d,p){return p==='CALLBACK'||d<0?'u':d<=21?'h':d<=45?'w':'';}
function stars(r){return r>0?'\u2605'.repeat(Math.round(r))+' '+r+'/5':'';}
function enc(s){return encodeURIComponent(s);}
function hav(la1,lo1,la2,lo2){
  const R=3958.8,a=Math.sin((la2-la1)*Math.PI/360)**2+
    Math.cos(la1*Math.PI/180)*Math.cos(la2*Math.PI/180)*Math.sin((lo2-lo1)*Math.PI/360)**2;
  return R*2*Math.atan2(Math.sqrt(a),Math.sqrt(1-a));
}

let log={},tab='today',selOut=null,selType=null,selReasonVal=null,cur=null,Q='',route=[],routeSet=new Set(),mapPros=[],routeAnchor=null;
let queueList=[],queueIdx=0;

function setType(t){
  selType=t;
  document.getElementById('mtype-walkin').className='mtype-btn'+(t==='walkin'?' mtype-on':'');
  document.getElementById('mtype-call').className='mtype-btn'+(t==='call'?' mtype-on':'');
  // Update pitch section visibility
  if(t==='walkin'){setPitchMode('walkin');}else if(t==='call'){setPitchMode('phone');}
}

function lLoad(){try{log=JSON.parse(localStorage.getItem('pic_v4')||'{}')||{};}catch(e){log={};}}
let _queueAutoAdvance=false;
function lSave(){
  if(_queueAutoAdvance){
    _queueAutoAdvance=false;
    closeMForce();
    queueIdx++;
    renderQueueCard();
    toast('✓ Logged — next');
  }
  try{localStorage.setItem('pic_v4',JSON.stringify(log));}catch(e){}
}
function phLoad(){
  try{
    var saved=JSON.parse(localStorage.getItem('pic_phones')||'{}');
    Object.entries(saved).forEach(function(kv){
      var id=parseInt(kv[0]),data=kv[1];
      PHONES[kv[0]]=data.phone||'';
      var p=P.find(function(x){return x.id===id;});
      if(p){p.phone=data.phone||'';p.hours=data.hours||'';p.rating=data.rating||0;}
    });
  }catch(e){}
}
function phSave(id,phone,hours,rating){
  let s={};try{s=JSON.parse(localStorage.getItem('pic_phones')||'{}')||{};}catch(e){}
  s[id]={phone,hours,rating};try{localStorage.setItem('pic_phones',JSON.stringify(s));}catch(e){}
  const p=P.find(x=>x.id===id);if(p){p.phone=phone;p.hours=hours;p.rating=rating;}
}
function getLC(id){const e=log[id]||[];return e.length?e[e.length-1]:null;}
function isC(id){
  const entries=log[id]||[];
  if(!entries.length)return false;
  // Do not count as contacted if ALL entries are skips
  return entries.some(e=>e.notes!=='Skipped');
}

function fp(opts){
  return P.filter(p=>{
    if(Q&&!p.name.toLowerCase().includes(Q)&&!p.city.toLowerCase().includes(Q))return false;
    if(opts.county&&p.county!==opts.county)return false;
    if(opts.pri&&p.priority!==opts.pri)return false;
    if(opts.st){const lc=getLC(p.id);const l=lc?normO(lc.outcome):'not_contacted';if(opts.st==='not_contacted'){if(isC(p.id))return false;}else if(l!==opts.st)return false;}
    if(opts.ice){if(opts.ice==='chronic'&&!p.chronic)return false;if(opts.ice==='confirmed'&&!p.confirmed)return false;if(opts.ice==='high_ice'&&'medium'!=='high')return false;if(opts.ice==='phone'&&!p.phone)return false;}
    if(opts.hc&&isC(p.id))return false;
    return true;
  }).sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||b.score-a.score);
}

let _leafletLoaded=false;
function loadLeaflet(cb){
  if(typeof L!=='undefined'){cb();return;}
  if(_leafletLoaded){setTimeout(()=>loadLeaflet(cb),100);return;}
  _leafletLoaded=true;
  const lnk=document.createElement('link');
  lnk.rel='stylesheet';
  lnk.href='https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css';
  document.head.appendChild(lnk);
  const scr=document.createElement('script');
  scr.src='https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js';
  scr.onload=cb;
  scr.onerror=()=>{window._leafletFailed=true;cb();};
  document.head.appendChild(scr);
}
function sw(t){
  tab=t;
  document.querySelectorAll('.tab').forEach((el,i)=>el.classList.toggle('on',['today','all','route','customers','service','data'][i]===t));
  document.querySelectorAll('.panel').forEach(el=>el.classList.remove('on'));
  const panel=document.getElementById('p-'+t);
  if(panel)panel.classList.add('on');
  if(t==='today'){renderBriefing();}
  else if(t==='all'){const ag=document.getElementById('agrid');if(ag)delete ag._glAttached;showDebugIfNeeded();populateCityFilter();rA();}
  else if(t==='route'){rRoute();}
  else if(t==='customers'){rCust();}
  // Load leaflet on route
  if(t==='route'&&typeof L==='undefined')loadLeaflet();
}


function setF(p){sw('today');rT();}
function onS(){Q=document.getElementById('si').value.toLowerCase().trim();if(tab==='today')rT();else if(tab==='all')rA();}

function cardHTML(p){
  const last=getLC(p.id);
  // String concat only - no nested backtick templates (Safari compatibility)
  const _ph=p.phone||PHONES[String(p.id)]||'';  // p.phone is set by phLoad from localStorage
  const phH=_ph
    ?('<div class="phrow"><span class="phnum">'+_ph+'</span><a href="tel:'+_ph.replace(/\s/g,'')+'" class="abtn call-a" onclick="event.stopPropagation()">Call</a></div>')
    :('<div class="phrow"><span class="phnum none">No phone on file</span><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone number')+'" target="_blank" class="abtn find-a" onclick="event.stopPropagation()">Find</a></div>'+'<div id="ph-save-'+p.id+'" style="margin-top:4px;display:flex;gap:4px;align-items:center">'  +'<input id="ph-inp-'+p.id+'" type="tel" placeholder="Paste number here..." onclick="event.stopPropagation()" '    +'style="flex:1;padding:5px 7px;border:1px solid var(--brd);border-radius:6px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'  +'<button class="save-phone-btn" data-id="'+p.id+'" data-action="save" '    +'style="padding:5px 9px;border:none;border-radius:6px;background:var(--navy);color:#fff;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">Save</button>'+'</div>');
  const iceH=p.chronic
    ?('<div class="icebadge chronic">&#x1F9CA; CHRONIC  -  '+p.ice_count+'x ice violations'+(p.ice_fresh?' &bull; <b>recent</b>':'')+'</div>')
    :p.confirmed?'<div class="icebadge confirmed">&#x2713; Ice violation on record'+(p.ice_fresh?' (within 6mo)':p.ice_recent?' (within 1yr)':'')+'</div>':'';
  const codesH=(p.codes||[]).length?('<div style="font-size:9px;color:#2a4860;margin-bottom:4px">Codes: '+(p.codes||[]).join(', ')+'</div>'):'';
  const insH=''?('<div class="insight">'+''+'</div>'):'';
  const cbH=(p.n_callbacks>0||p.disp_risk>=4)
    ?('<div style="font-size:9px;font-weight:600;padding:3px 8px;border-radius:5px;margin-bottom:5px;background:#fef2f2;color:#dc2626;border:1px solid #fecaca">'
      +(p.n_callbacks>0?'&#x1F6A8; '+p.n_callbacks+'x callback inspection'+(p.n_callbacks>1?'s':''):'&#x26A0;&#xFE0F; Admin complaint  -  callback due')
      +'</div>')
    :'';
  const hrH=p.hours?('<div style="font-size:9px;color:#1a3850;margin-bottom:4px">&#x1F550; '+p.hours+'</div>'):'';
  const ratH=p.rating>0?('<div style="font-size:9px;color:#c08020;margin-bottom:3px">'+stars(p.rating)+'</div>'):'';
  // Best call window based on business type
  const hr=new Date().getHours();
  const callWin=p.is_bar
    ?{good:[9,10,11],avoid:[16,17,18,19,20,21,22],tip:'Best: 9-11am before they open'}
    :{good:[9,10,14,15,16],avoid:[11,12,13,17,18,19,20],tip:'Best: 9-10am or 2-4pm. Avoid lunch/dinner rush'};
  const inGood=callWin.good.includes(hr);
  const inAvoid=callWin.avoid.includes(hr);
  const callH=inGood
    ?'<div style="font-size:8px;font-weight:700;color:#059669;margin-bottom:3px">&#x1F7E2; Good time to call now</div>'
    :inAvoid
      ?'<div style="font-size:8px;font-weight:700;color:#dc2626;margin-bottom:3px">&#x1F534; Rush hour — try later</div>'
      :'';
  const lastH=last
    ?('<div class="lastc hc" style="color:'+(OI_COLOR[last.outcome]||'#64748b')+'">'+( OI[last.outcome]||last.outcome)+' &middot; '+last.date+(last.notes?'  -  '+last.notes.slice(0,25):'')+' </div>')
    :'<div class="lastc">Not yet contacted</div>';
  const trendH=p.trending?'<div class="mi"><div class="ml">Trend</div><div class="mv bad">&#x2197; Worse</div></div>':'';
  const rtBtn=p.lat?('<button class="btn brt" class="route-btn" data-id="'+p.id+'" data-action="route">+Route</button>'):'';
  const tc={PLATINUM:'#7c3aed',GOLD:'#d97706',SILVER:'#64748b',BRONZE:'#92400e'};
  const tbg={PLATINUM:'#ede9fe',GOLD:'#fef3c7',SILVER:'#f1f5f9',BRONZE:'#fef3c7'};
  const tierH=p.tier&&p.tier!=='COLD'?(' <span style="font-size:8px;font-weight:700;padding:2px 6px;border-radius:20px;background:'+(tbg[p.tier]||'#f1f5f9')+';color:'+(tc[p.tier]||'#64748b')+'">'+p.tier+'</span>'):'';
  const revenueH=p.monthly?('<div style="font-size:10px;font-weight:700;color:#059669;margin-bottom:5px">$'+p.monthly+'/mo &bull; '+(p.machines>1?p.machines+' machines':'1 machine')+'</div>'):'';
  const emergH=p.is_emergency?'<span class="emerg-badge">&#x1F6A8; EMERGENCY</span>':'';
  const confCol=p.confidence>=75?'#059669':p.confidence>=50?'#d97706':'#9ca3af';
  const confH='<span style="font-size:8px;font-weight:600;color:'+confCol+'" title="Prediction confidence">'+p.confidence+'% conf</span>';
  const franchH=p.biz_type==='franchise'?'<span style="font-size:8px;padding:1px 5px;border-radius:10px;background:#f0f9ff;color:#0ea5e9;border:1px solid #bae6fd">Franchise</span>':'';
  const custStatusH=(p.status&&p.status!=='prospect')
    ?('<div style="font-size:9px;font-weight:700;padding:2px 8px;border-radius:20px;display:inline-block;margin-bottom:4px;background:'+(p.status==='customer_recurring'?'#ecfdf5':p.status==='customer_once'?'#eff6ff':'#fff7f5')+';color:'+(p.status==='customer_recurring'?'#059669':p.status==='customer_once'?'var(--blu)':'var(--ora)')+'">'+({'customer_recurring':'Recurring Customer','customer_once':'One-Time Customer','quoted':'Quote Sent','churned':'Churned'}[p.status]||p.status)+'</div>')
    :'';
  return '<div class="card '+p.priority+(isC(p.id)?' done':'')+'" data-id="'+p.id+'">'
    +'<div class="ctop"><div class="cname">'+p.name+tierH+emergH+'</div><div style="display:flex;flex-direction:column;align-items:flex-end;gap:2px"><span class="pbadge '+p.priority+'">'+p.priority+'</span>'+confH+'</div></div>'
    +'<div class="cloc">'+p.city+', '+p.county+' '+franchH+'</div>'
    +custStatusH+revenueH+phH+ratH+callH+hrH+iceH+cbH+codesH+insH
    +'<div class="cmeta">'
      +'<div class="mi"><div class="ml">Days Until</div><div class="mv '+dC(p.days_until,p.priority)+'">'+dL(p.days_until)+'</div></div>'
      +'<div class="mi"><div class="ml">H/T Viol</div><div class="mv '+(p.high_viol>=4?'u':p.high_viol>=2?'h':'')+'">'+p.high_viol+'/'+p.total_viol+'</div></div>'
      +'<div class="mi"><div class="ml">Pred.Next</div><div class="mv" style="font-size:10px">'+p.pred_next+'</div></div>'
      +trendH
    +'</div>'
    +lastH
    +'<div class="cacts">'
      +'<button class="btn blog" class="log-call-btn" data-id="'+p.id+'" data-action="showCard" onclick="event.stopPropagation();showCard('+p.id+')" ontouchend="event.stopPropagation();event.preventDefault();showCard('+p.id+')" style="touch-action:manipulation">Log Call</button>'
      +'<button class="btn" data-id="'+p.id+'" data-action="showCard" onclick="event.stopPropagation();showCard('+p.id+')" ontouchend="event.stopPropagation();event.preventDefault();showCard('+p.id+')" style="background:#0f1f38;color:#fff;border-color:#0f1f38;touch-action:manipulation">Details</button>'
      +(last&&last.notes==='Skipped'
        ?'<button class="btn bskip" class="skip-btn" data-id="'+p.id+'" data-action="unskip" style="background:#ecfdf5;color:#059669;border-color:#6ee7b7">Unskip</button>'
        :'<button class="btn bskip" class="skip-btn" data-id="'+p.id+'" data-action="skip">Skip</button>')
      +rtBtn
    +'</div>'
  +'</div>';
}

function rT(){renderBriefing();}


function clearFilters(){
  setPreset('all');
}

let _raTimer=null;
function dRa(){clearTimeout(_raTimer);_raTimer=setTimeout(rA,80);}

function attachGridListeners(grid){
  if(!grid)return;
  // Only attach once per element - subsequent rA() calls reuse same listener
  // (delegation handles dynamically added children automatically)
  if(grid._glAttached)return;
  grid._glAttached=true;

  let _sx=0,_sy=0;
  grid.addEventListener('touchstart',function(e){
    _sx=e.touches[0].clientX;
    _sy=e.touches[0].clientY;
  },{passive:true});

  grid.addEventListener('touchend',function(e){
    const dx=Math.abs(e.changedTouches[0].clientX-_sx);
    const dy=Math.abs(e.changedTouches[0].clientY-_sy);
    if(dx>14||dy>14)return; // scroll, ignore

    // Check for action button first
    const btn=e.target.closest('[data-action]');
    if(btn){
      e.preventDefault();
      const id=parseInt(btn.dataset.id||btn.closest('[data-id]')?.dataset.id||'0');
      if(!id)return;
      const act=btn.dataset.action;
      if(act==='showCard')    showCard(id);
      else if(act==='svclog') openServiceLog(id);
      else if(act==='snext')  setNextService(id);
      else if(act==='skip')   skip(id);
      else if(act==='route')  addToRoute(id);
      else if(act==='start')  setRouteAnchor(id);
      else if(act==='save')   saveFoundPhone(id);
      else if(act==='unskip') unskip(id);
      return;
    }
    // Tap on card body (not a button)
    const card=e.target.closest('[data-id]');
    if(card&&!e.target.closest('a')&&!e.target.closest('input')){
      e.preventDefault();
      const id=parseInt(card.dataset.id||'0');
      if(id)openM(id);
    }
  },false);
}


function debugTestBtn(){
  // Called from static button - proves if openM works at all
  const p=P[0];
  if(!p){toast('No prospects loaded');return;}
  toast('Opening: '+p.name);
  openM(p.id);
}

// Show debug panel when on Prospects tab
function showDebugIfNeeded(){
  const el=document.getElementById('debug-test');
  if(el)el.style.display='block';
}


function rA(){
  const county  = document.getElementById('ac')?.value||'';
  const city    = document.getElementById('ac-city')?.value||'';
  const pri     = document.getElementById('ap')?.value||'';
  const st      = document.getElementById('as_')?.value||'';
  const conf    = parseInt(document.getElementById('aconf')?.value||'0');

  let list=P.filter(p=>{
    if(Q&&!p.name.toLowerCase().includes(Q)&&!p.city.toLowerCase().includes(Q))return false;
    if(county&&p.county!==county)return false;
    if(city&&p.city!==city)return false;
    if(pri&&p.priority!==pri)return false;
    if(st){
      const lc=getLC(p.id);
      const norm=lc?normO(lc.outcome):'not_contacted';
      if(st==='not_contacted'){if(isC(p.id))return false;}
      else if(norm!==st)return false;
    }
    // Apply preset exclusions
    if(presetFilter&&!presetFilter(p))return false;
    return true;
  });

  // Sort: follow-ups by date, others by callback/score
  if(presetFilter&&document.getElementById('pre-followups')?.classList.contains('on')){
    list.sort((a,b)=>{
      const fa=getLC(a.id)?.followup||'9999';
      const fb=getLC(b.id)?.followup||'9999';
      return fa.localeCompare(fb);
    });
  } else {
    list.sort((a,b)=>{
      const pa=PO[a.priority]??99, pb=PO[b.priority]??99;
      if(pa!==pb)return pa-pb;
      return (b.score||0)-(a.score||0);
    });
  }

  document.getElementById('acnt').textContent=list.length+' prospects';
  const g=document.getElementById('agrid');
  const empty=document.getElementById('a-empty');
  if(list.length===0){g.innerHTML='';empty.style.display='flex';return;}
  empty.style.display='none';
  requestAnimationFrame(()=>{
    const visible=list.slice(0,50);
    g.innerHTML=visible.map(p=>cardHTML(p)).join('');
    if(list.length>50){g.innerHTML+='<div style="padding:12px;text-align:center;font-size:11px;color:var(--sub)">Showing 50 of '+list.length+' — use filters to narrow results</div>';}
    attachGridListeners(g);
  });
}

let presetFilter=null;
function setPreset(k){
  document.querySelectorAll('.preset-btn').forEach(b=>b.classList.remove('on'));
  document.getElementById('pre-'+k)?.classList.add('on');
  // Reset city/county/status filters
  const cityEl=document.getElementById('ac-city');
  if(cityEl)cityEl.value='';

  presetFilter=null;
  if(k==='actnow'){presetFilter=p=>{
    const e=getLC(p.id);
    return p.priority==='CALLBACK'||p.priority==='HOT'||(e&&e.outcome==='no_contact'&&e.followup);
  };}
  else if(k==='callback') presetFilter=p=>p.priority==='CALLBACK';
  else if(k==='phone')    presetFilter=p=>!!p.phone;
  else if(k==='chronic')  presetFilter=p=>p.chronic;
  else if(k==='notyet')   presetFilter=p=>!isC(p.id);
  else if(k==='inplay')   presetFilter=p=>{const lc=getLC(p.id);return lc&&normO(lc.outcome)==='in_play';};
  else if(k==='freshice') presetFilter=p=>p.confirmed&&!p.chronic;
  else if(k==='followups') presetFilter=p=>{
    const lc=getLC(p.id);
    return lc&&lc.followup&&!isC(p.id);
  };
  rA();
}

function clearFilters(){
  ['ac','ac-city','ap','as_'].forEach(id=>{
    const el=document.getElementById(id);if(el)el.value='';
  });
  setPreset('all');
}

function populateCityFilter(){
  const sel=document.getElementById('ac-city');
  if(!sel)return;
  const county=document.getElementById('ac')?.value||'';
  const cities=[...new Set(P
    .filter(p=>!county||p.county===county)
    .map(p=>p.city)
    .filter(Boolean)
  )].sort();
  sel.innerHTML='<option value="">All Cities</option>'+cities.map(c=>`<option>${c}</option>`).join('');
}

// ── CALL QUEUE MODE ──────────────────────────────────────────────────────
let queueProspects=[];

function enterQueueMode(){
  // Use whatever is currently filtered in All tab
  const county = document.getElementById('ac')?.value||'';
  const city   = document.getElementById('ac-city')?.value||'';
  const pri    = document.getElementById('ap')?.value||'';

  queueProspects=P.filter(p=>{
    if(county&&p.county!==county)return false;
    if(city&&p.city!==city)return false;
    if(pri&&p.priority!==pri)return false;
    if(presetFilter&&!presetFilter(p))return false;
    // Default: only not-yet-fully-resolved
    const lc=getLC(p.id);
    const norm=lc?normO(lc.outcome):'not_contacted';
    return !['signed','dead'].includes(norm);
  }).sort((a,b)=>{
    const pa=PO[a.priority]??99,pb=PO[b.priority]??99;
    return pa!==pb?pa-pb:(b.score||0)-(a.score||0);
  });

  if(!queueProspects.length){toast('No prospects match current filters');return;}
  queueIdx=0;
  document.getElementById('queue-bg').classList.add('on');
  renderQueueCard();
}

function exitQueueMode(){
  document.getElementById('queue-bg').classList.remove('on');
  rA();renderBriefing();
}

function renderQueueCard(){
  const wrap=document.getElementById('queue-card-wrap');
  const acts=document.getElementById('queue-actions');
  if(queueIdx>=queueProspects.length){
    wrap.innerHTML=
      '<div style="text-align:center;padding:40px 20px">'
      +'<div style="font-size:40px;margin-bottom:12px">✓</div>'
      +'<div style="font-size:16px;font-weight:700;margin-bottom:6px">Queue complete</div>'
      +'<div style="font-size:12px;color:#888">'+queueProspects.length+' prospects worked</div>'
      +'<button onclick="exitQueueMode()" ontouchend="event.preventDefault();exitQueueMode()"'
       +' style="margin-top:16px;padding:10px 24px;background:var(--navy);color:#fff;border:none;'
       +'border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">Done</button>'
      +'</div>';
    document.getElementById('queue-progress').textContent='Complete!';
    if(acts)acts.style.visibility='hidden';
    return;
  }
  if(acts)acts.style.visibility='visible';

  const p=queueProspects[queueIdx];
  const lc=getLC(p.id);
  const entries=log[p.id]||[];
  document.getElementById('queue-progress').textContent=(queueIdx+1)+' / '+queueProspects.length;

  // Phone
  const ph=p.phone||'';
  const phoneH=ph
    ?'<a href="tel:'+ph.replace(/\D/g,'')+'" style="display:inline-flex;align-items:center;gap:8px;background:#059669;color:#fff;padding:10px 16px;border-radius:8px;font-size:14px;font-weight:700;text-decoration:none;touch-action:manipulation;margin-bottom:10px">📞 '+ph+'</a>'
    :'<div style="margin-bottom:8px"><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone')+'" target="_blank" style="font-size:11px;color:var(--blu);border:1px solid var(--blu);padding:4px 10px;border-radius:5px;text-decoration:none">Find Phone ↗</a></div>';

  // Chips
  const pBgMap={CALLBACK:'#fef2f2',HOT:'#fff7ed',WARM:'#fefce8',COOL:'#f8fafc'};
  const pClrMap={CALLBACK:'#dc2626',HOT:'#f97316',WARM:'#ca8a04',COOL:'#64748b'};
  const pc=pClrMap[p.priority]||'#64748b';
  const lColor=OI_COLOR[normO(lc?.outcome)]||'#64748b';
  const chipsH=[
    '<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;background:'+(pBgMap[p.priority]||'#f8fafc')+';color:'+pc+';border:1px solid '+pc+'40">'+p.priority+'</span>',
    p.chronic?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;background:#ecfdf5;color:#059669;border:1px solid #6ee7b740">CHRONIC ICE</span>':'',
    p.days_since>0?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;color:#dc2626;background:#fef2f2">'+p.days_since+'d overdue</span>':'',
    lc?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;color:'+lColor+';background:'+lColor+'18">'+(OI[normO(lc.outcome)]||lc.outcome)+'</span>':'',
  ].filter(Boolean).join('');

  // Inspection intel
  const intel=[
    p.n_callbacks>0?'🚨 '+p.n_callbacks+'x callback inspection'+(p.n_callbacks>1?'s':''):'',
    p.ice_count>0?'❄️ '+p.ice_count+'x ice violation'+(p.ice_count>1?'s':'')+' on record':'',
    p.high_viol>0?'⚠️ '+p.high_viol+' critical violation'+(p.high_viol>1?'s':'')+' (high priority)':'',
    p.codes&&p.codes.length?'Codes: '+p.codes.slice(0,5).join(', '):'',
    p.pred_next?'Next inspection: '+p.pred_next:'',
    p.confidence?'Confidence: '+p.confidence+'%':'',
  ].filter(Boolean);
  const intelH=intel.length
    ?'<div style="background:#f8fafc;border-radius:8px;padding:8px 10px;margin-bottom:10px;border-left:3px solid #e2e8f0">'
      +intel.map(r=>'<div style="font-size:10px;color:#475569;line-height:1.9">'+r+'</div>').join('')
     +'</div>':'';

  // Pitch + walk-in scripts
  const pitchFn=PITCHES[p.pitch_type]||PITCHES.routine;
  const walkinFn=WALKIN[p.pitch_type]||WALKIN.routine;
  const pitchH='<div style="margin-bottom:10px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Pitch Opener</div>'
    +'<div style="background:#fff7ed;border-left:3px solid var(--gold);border-radius:0 7px 7px 0;padding:8px 10px;font-size:11px;color:#475569;line-height:1.6">'+pitchFn(p.name)+'</div>'
  +'</div>';
  const walkinH='<div style="margin-bottom:10px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Walk-In Script</div>'
    +'<div style="background:#f0fdf4;border-left:3px solid #059669;border-radius:0 7px 7px 0;padding:8px 10px;font-size:11px;color:#475569;line-height:1.6">'+walkinFn(p.name)+'</div>'
  +'</div>';

  // Last note
  const lastNoteH=lc&&lc.notes
    ?'<div style="font-size:10px;color:#64748b;font-style:italic;padding:6px 8px;background:#f1f5f9;border-radius:6px;border-left:3px solid #e2e8f0;margin-bottom:10px">'+'\u201c'+lc.notes+'\u201d \u2014 '+lc.date+'</div>':'';

  // History
  const histH=entries.length
    ?'<div style="margin-bottom:10px">'
      +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">History</div>'
      +[...entries].reverse().slice(0,4).map(e=>{
        const ec=OI_COLOR[normO(e.outcome)]||'#64748b';
        return '<div style="display:flex;gap:6px;align-items:baseline;font-size:10px;padding:3px 0;border-bottom:1px solid #f1f5f9">'
          +'<span style="color:'+ec+';font-weight:700;flex-shrink:0">'+(OI[normO(e.outcome)]||e.outcome)+'</span>'
          +'<span style="color:#94a3b8;flex-shrink:0">&bull; '+e.date+'</span>'
          +(e.notes?'<span style="color:#64748b;font-style:italic"> \u2014 '+e.notes.slice(0,40)+'</span>':'')
         +'</div>';
      }).join('')
     +'</div>':'';

  // Objections
  const objH='<div style="margin-bottom:10px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Objection Handlers</div>'
    +OBJECTIONS.slice(0,4).map(o=>
      '<details style="background:#f8fafc;border:1px solid var(--brd);border-radius:6px;margin-bottom:4px">'
      +'<summary style="padding:6px 8px;font-size:10px;font-weight:600;color:var(--navy);cursor:pointer;list-style:none">❓ '+o.q+'</summary>'
      +'<div style="padding:4px 8px 8px;font-size:10px;color:#475569;line-height:1.6;border-top:1px solid var(--brd)">'+o.a+'</div>'
      +'</details>'
    ).join('')
  +'</div>';

  wrap.innerHTML=
    '<div style="background:#fff;border-radius:12px;padding:16px;box-shadow:0 2px 12px rgba(0,0,0,.1)">'
    +'<div style="margin-bottom:8px">'
      +'<div style="font-size:16px;font-weight:800;color:var(--navy);margin-bottom:2px">'+p.name+'</div>'
      +'<div style="font-size:10px;color:#94a3b8">'+p.address+', '+p.city+' · FL#'+p.id+'</div>'
    +'</div>'
    +'<div style="display:flex;gap:5px;flex-wrap:wrap;margin-bottom:10px">'+chipsH+'</div>'
    +phoneH+intelH+pitchH+walkinH+lastNoteH+histH+objH
    +'<textarea id="q-notes" placeholder="Quick note (optional)..." rows="2" '
      +'style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:12px;font-family:inherit;resize:none;outline:none;color:#1a1a2e;background:#fff;box-sizing:border-box"></textarea>'
    +'</div>';
}


function queueLogFull(outcome){
  var p=queueProspects[queueIdx];
  if(!p){toast('No prospect');return;}
  // For not_now and dead - open showCard so user can pick reason
  if(outcome==='not_now'||outcome==='dead'){
    _queueAutoAdvance=true;
    showCard(p.id);
    return;
  }
  // All other outcomes - log directly and advance
  if(!log[p.id])log[p.id]=[];
  log[p.id].push({
    outcome:outcome, type:'call', reason:null,
    date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),
    notes:'', followup:''
  });
  lSave();
  toast('\u2713 '+(OI[outcome]||outcome)+' — '+p.name.slice(0,20));
  queueIdx++;
  renderQueueCard();
}


function queueLog(outcome){
  const p=queueProspects[queueIdx];
  if(!p){toast('No prospect loaded');return;}

  // For not_now and dead — open full card so user can pick reason
  if(outcome==='not_now'||outcome==='dead'){
    openM(p.id);
    // Pre-select the outcome in the modal
    setTimeout(()=>{
      const btn=document.querySelector('.obtn[data-o="'+outcome+'"]');
      if(btn)btn.click();
    },200);
    return;
  }

  const notes=document.getElementById('q-notes')?.value||'';
  if(!log[p.id])log[p.id]=[];
  log[p.id].push({
    outcome,
    type:'call',
    reason:null,
    date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),
    notes,
    followup:''
  });
  lSave();
  toast('\u2713 '+(OI[outcome]||outcome)+' \u2014 '+p.name.slice(0,22));
  queueIdx++;
  renderQueueCard();
}


function queueNext(){
  queueIdx++;
  renderQueueCard();
}



function rRoute(){
  const county=document.getElementById('rc').value;
  const pri=document.getElementById('rp').value;
  const zip=document.getElementById('rzip').value.trim();
  const rad=parseFloat(document.getElementById('rrad').value)||999;
  // Show all except customers and hard nos
  const rExclude=new Set();
  P.forEach(p=>{
    if(p.status==='customer_recurring'||p.status==='customer_once')rExclude.add(p.id);
    const entries=log[p.id]||[];
    if(entries.length&&entries[entries.length-1].outcome==='not_interested')rExclude.add(p.id);
  });
  let data=P.filter(p=>p.lat&&p.lon&&!rExclude.has(p.id));
  if(county)data=data.filter(p=>p.county===county);
  if(pri)data=data.filter(p=>p.priority===pri);
  if(zip.length===5&&ZIPS[zip]){
    const [clat,clon]=ZIPS[zip];
    data=data.map(p=>({...p,_d:hav(clat,clon,p.lat,p.lon)}))
      .filter(p=>p._d<=rad)
      .sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||a._d-b._d);
    document.getElementById('rhint').textContent=data.length+' prospects within '+rad+'mi of '+zip;
  }else{
    data.sort((a,b)=>(PO[a.priority]??5)-(PO[b.priority]??5)||b.score-a.score);
    document.getElementById('rhint').textContent=data.length+' geocoded prospects  -  enter ZIP to filter by area';
  }
  mapPros=data.slice(0,100);
  renderRList();renderMap();
}

function renderRList(){
  const el=document.getElementById('rlist');
  const cnt=document.getElementById('rlist-cnt');
  if(!mapPros.length){
    el.innerHTML='<div style="font-size:11px;color:var(--sub);padding:8px">No geocoded prospects match. Try removing ZIP filter or expanding radius.</div>';
    if(cnt)cnt.textContent='';
    return;
  }
  if(cnt)cnt.textContent=mapPros.length+' prospects nearby';
  el.innerHTML=mapPros.map(p=>{
    const inR=routeSet.has(p.id);
    const col=PC[p.priority]||'var(--sub)';
    const distTxt=p._d?p._d.toFixed(1)+'mi':'';
    const chronicTxt=p.chronic?'<div style="font-size:8px;color:#059669;font-weight:700">CHRONIC ICE</div>':'';
    const isAnchor=routeAnchor&&routeAnchor.id===p.id;
    return '<div class="rcard'+(inR?' sel':'')+'" data-action="route" style="border-left:3px solid '+col+(inR?';background:#fff7f5':'')+(isAnchor?';border-left:4px solid #7c3aed':'')+';">'
      +'<div class="rdot" style="background:'+col+'"></div>'
      +'<div style="flex:1;min-width:0">'
        +'<div class="rname" style="color:var(--navy)">'+p.name+(inR?' <span style="color:var(--ora)">&#x2713;</span>':'')+(isAnchor?' <span style="font-size:8px;color:#7c3aed">&#x1F4CD; START</span>':'')+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+p.city+' &bull; '+p.priority+' &bull; '+dL(p.days_until)+'</div>'
        +chronicTxt
      +'</div>'
      +'<div style="display:flex;flex-direction:column;align-items:flex-end;gap:3px">'
        +(distTxt?'<div class="rdist">'+distTxt+'</div>':'')
        +'<button data-action="start" style="font-size:8px;padding:2px 5px;border:1px solid #7c3aed;border-radius:4px;background:#f5f3ff;color:#7c3aed;cursor:pointer;font-family:inherit">Start &#x1F4CD;</button>'
        +'<button data-action="openM" style="font-size:8px;padding:4px 8px;min-height:36px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--sub);cursor:pointer;font-family:inherit;touch-action:manipulation">Details</button>'
      +'</div>'

      +'</div>';
  }).join('');
  attachGridListeners(el);
}

function renderMap(){
  const area=document.getElementById('map-area');
  if(typeof L==='undefined'||window._leafletFailed){
    // Leaflet not available (sandbox or offline) - show static list
    if(!mapPros.length){area.innerHTML='<div class="map-empty"><div style="font-size:32px">&#x1F5FA;&#xFE0F;</div><div>No prospects to map</div></div>';return;}
    area.innerHTML='<div style="padding:12px;overflow-y:auto;height:100%"><div style="font-size:10px;color:var(--sub);margin-bottom:8px">Map unavailable offline. '+(mapPros.length)+' prospects:</div>'+mapPros.slice(0,20).map(p=>'<div style="padding:5px 0;border-bottom:1px solid var(--brd2);font-size:11px;color:var(--navy)">'+p.name+'</div>').join('')+'</div>';
    return;
  }
  if(!mapPros.length){
    area.innerHTML='<div class="map-empty"><div style="font-size:36px;margin-bottom:10px">&#x1F5FA;&#xFE0F;</div><div>No prospects to map</div></div>';
    if(window._lmap){window._lmap.remove();window._lmap=null;}
    return;
  }
  // Init Leaflet map once
  if(!window._lmap){
    area.innerHTML='<div id="leaflet-map" style="width:100%;height:100%;border-radius:10px"></div>';
    window._lmap=L.map('leaflet-map',{zoomControl:true,attributionControl:false});
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
      maxZoom:18,opacity:0.85
    }).addTo(window._lmap);
    window._lmarkers=L.layerGroup().addTo(window._lmap);
  }
  window._lmarkers.clearLayers();
  const cols={CALLBACK:'#ff3b30',HOT:'#ff9f0a',WARM:'#ffd60a',WATCH:'#5e9eff',LATER:'#445066'};
  const bounds=[];
  mapPros.forEach(p=>{
    const inR=routeSet.has(p.id);
    const col=inR?'#00e5ff':cols[p.priority]||'#445066';
    const r=inR?12:p.chronic?10:8;
    const marker=L.circleMarker([p.lat,p.lon],{
      radius:r,color:col,fillColor:col,fillOpacity:0.85,weight:inR?3:1.5,
      opacity:1
    }).addTo(window._lmarkers);
    marker.bindPopup(
      '<b>'+p.name+'</b><br>'+p.city+' &bull; '+p.priority+'<br>'+dL(p.days_until)+
      (p.seats?' &bull; '+p.seats+' seats':'')+
      (p.monthly?' &bull; $'+p.monthly+'/mo':'')+
      '<br><button onclick="addToRoute('+p.id+')" '+
      'style="margin-top:4px;padding:3px 8px;border:none;border-radius:5px;background:'+col+';color:'+(p.priority==='WARM'?'#000':'#fff')+';font-size:10px;font-weight:700;cursor:pointer">'+
      (routeSet.has(p.id)?'Remove from Route':'+Route')+'</button> '+
      '<button onclick="openM('+p.id+')" '+
      'style="margin-top:3px;padding:3px 8px;border:1px solid #333;border-radius:5px;background:#111;color:#fff;font-size:10px;cursor:pointer">Details</button>',
      {maxWidth:220}
    );
    bounds.push([p.lat,p.lon]);
  });
  if(bounds.length)window._lmap.fitBounds(bounds,{padding:[20,20],maxZoom:13});
}

function addToRoute(id){
  const p=P.find(x=>x.id===id);if(!p)return;
  if(routeSet.has(id)){
    routeSet.delete(id);route=route.filter(r=>r.id!==id);
    toast(p.name.slice(0,20)+' removed from route');
  } else {
    if(route.length>=8){toast('Max 8 stops. Open in Maps first.');return;}
    routeSet.add(id);route.push(p);
    toast(p.name.slice(0,20)+' added to route ('+route.length+' stops)');
  }
  renderRList();renderMap();renderDayRoute();
}
function renderDayRoute(){
  const dr=document.getElementById('day-route');
  document.getElementById('stopcnt').textContent=route.length;
  if(!route.length){dr.style.display='none';return;}
  dr.style.display='block';

  const zip=(document.getElementById('rzip')||{}).value||'';
  let curLat=null,curLon=null;
  if(zip.length===5&&ZIPS[zip]){[curLat,curLon]=ZIPS[zip];}

  let totalMi=0,totalMin=0,tmpLat=curLat,tmpLon=curLon;
  route.forEach(p=>{
    if(tmpLat&&p.lat){const d=hav(tmpLat,tmpLon,p.lat,p.lon);totalMi+=d;totalMin+=d*DRIVE_MIN_PER_MILE;}
    totalMin+=MIN_PER_STOP[p.priority]||15;
    if(p.lat){tmpLat=p.lat;tmpLon=p.lon;}
  });

  const mi=document.getElementById('route-mi');
  const te=document.getElementById('route-time-est');
  if(mi&&totalMi>0)mi.textContent=' (~'+totalMi.toFixed(1)+'mi)';
  if(te&&totalMin>0){const h=Math.floor(totalMin/60),m=Math.round(totalMin%60);te.textContent='Est. time: '+h+'h '+m+'m (drive + visit)';}

  let runMin=0,rLat=curLat,rLon=curLon;
  // If anchor is set, show it as stop 0
  const anchorRow=routeAnchor
    ?('<div class="day-stop" style="background:#f5f3ff;border:1px solid #ddd6fe;border-radius:8px;margin-bottom:4px">'
      +'<div class="stopnum" style="background:#7c3aed">&#x1F4CD;</div>'
      +'<div style="flex:1;min-width:0">'
        +'<div style="font-weight:700;font-size:11px;color:#7c3aed">START: '+routeAnchor.name+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+routeAnchor.address+', '+routeAnchor.city+'</div>'
      +'</div>'
      +'<button onclick="clearAnchor()" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">Clear</button>'
      +'</div>')
    :'';
  document.getElementById('day-stops').innerHTML=anchorRow+route.map((p,i)=>{
    let driveMin=0;
    if(rLat&&p.lat){driveMin=Math.round(hav(rLat,rLon,p.lat,p.lon)*DRIVE_MIN_PER_MILE);}
    const stopMin=MIN_PER_STOP[p.priority]||15;
    runMin+=driveMin+stopMin;
    if(p.lat){rLat=p.lat;rLon=p.lon;}
    const ph=p.phone?('<a href="tel:'+p.phone.replace(/\s/g,'')+'" style="font-size:9px;color:var(--blu);text-decoration:none">'+p.phone+'</a>'):'';
    const driveH=driveMin>0?('<span style="font-size:8px;color:var(--sub)">+'+driveMin+'m drive</span>'):'<span style="font-size:8px;color:var(--sub)">Start</span>';
    return '<div class="day-stop">'
      +'<div class="stopnum">'+(i+1)+'</div>'
      +'<div style="flex:1;min-width:0">'
        +'<div style="font-weight:700;font-size:11px;color:var(--navy);overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+p.name+'</div>'
        +'<div style="font-size:9px;color:var(--sub)">'+p.address+', '+p.city+'</div>'
        +'<div style="display:flex;gap:5px;align-items:center;margin-top:2px;flex-wrap:wrap">'
          +ph+'<span style="font-size:9px;font-weight:600;color:#059669">$'+p.monthly+'/mo</span>'
          +'<span class="pbadge '+p.priority+'" style="font-size:7px">'+p.priority+'</span>'
          +driveH
        +'</div>'
      +'</div>'
      +'<div style="display:flex;flex-direction:column;gap:3px;flex-shrink:0">'
        +'<button onclick="openM('+p.id+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--ora);border-radius:5px;background:#fff7f5;color:var(--ora);cursor:pointer;font-family:inherit">Details</button>'
        +'<button onclick="removeStop('+p.id+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer;font-family:inherit">Remove</button>'
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Type</div>'
            +'<select onchange="saveMachineType(\'+p.id+\',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_TYPE_LIST.map(function(t){return '<option value="'+t+'"'+(t===(c.machine_type||'')?'selected':'')+'>'+t+'</option>'}).join('')
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

const MIN_PER_STOP={CALLBACK:25,HOT:20,WARM:15,WATCH:12,LATER:10};
const DRIVE_MIN_PER_MILE=3;
const PW={CALLBACK:100,HOT:80,WARM:50,WATCH:20,LATER:5};

function planMyDay(){
  const zip=(document.getElementById('rzip')||{}).value.trim();
  const hours=parseFloat((document.getElementById('rtime')||{}).value)||0;
  const county=(document.getElementById('rc')||{}).value||'';
  const pri=(document.getElementById('rp')||{}).value||'';
  const rad=parseFloat((document.getElementById('rrad')||{}).value)||8;

  if(!hours){rRoute();return;}

  // Determine start coords — anchor takes priority over home ZIP
  let slat,slon,startLabel;
  if(routeAnchor&&routeAnchor.lat){
    slat=routeAnchor.lat;slon=routeAnchor.lon;
    startLabel='From: '+routeAnchor.name.slice(0,22);
  } else if(zip.length===5&&ZIPS[zip]){
    [slat,slon]=ZIPS[zip];
    startLabel='From ZIP '+zip;
  } else {
    toast('Enter a start ZIP or tap Start on a business first');return;
  }

  const budgetMin=hours*60;

  // CRITICAL: filter radius from ACTUAL start point, not home ZIP
  // Route excludes: current customers, hard nos, already on route
  // Route INCLUDES: voicemails, no answers, follow-ups, uncontacted
  const hardExclude=new Set();
  P.forEach(p=>{
    if(p.status==='customer_recurring'||p.status==='customer_once')hardExclude.add(p.id);
    const entries=log[p.id]||[];
    const lastOutcome=entries.length?entries[entries.length-1].outcome:'';
    if(lastOutcome==='not_interested')hardExclude.add(p.id);
    if(routeSet.has(p.id))hardExclude.add(p.id);
  });
  let candidates=P.filter(p=>p.lat&&p.lon&&!hardExclude.has(p.id));
  if(county)candidates=candidates.filter(p=>p.county===county);
  if(pri)   candidates=candidates.filter(p=>p.priority===pri);
  candidates=candidates
    .map(p=>({...p,_d:hav(slat,slon,p.lat,p.lon)}))
    .filter(p=>p._d<=rad);

  if(!candidates.length){
    toast('No uncontacted prospects within '+rad+'mi. Try increasing radius.');
    return;
  }

  route=[];routeSet=new Set();
  let usedMin=0,curLat=slat,curLon=slon;
  const remaining=[...candidates];

  while(remaining.length&&usedMin<budgetMin){
    let bestIdx=-1,bestVal=-Infinity;
    remaining.forEach((p,i)=>{
      const dMin=hav(curLat,curLon,p.lat,p.lon)*DRIVE_MIN_PER_MILE;
      const sMin=MIN_PER_STOP[p.priority]||15;
      if(usedMin+dMin+sMin>budgetMin)return;
      // Boost uncontacted and voicemail/no-answer (good walk-in candidates)
      const entries=log[p.id]||[];
      const lastOutcome=entries.length?entries[entries.length-1].outcome:'';
      const contactBoost=lastOutcome===''?5
        :lastOutcome==='no_answer'||lastOutcome==='voicemail'?8
        :lastOutcome==='follow_up'||lastOutcome==='interested'?12
        :-2; // recently logged other outcome
      const val=(PW[p.priority]||0)+p.score+contactBoost-(dMin*2);
      if(val>bestVal){bestVal=val;bestIdx=i;}
    });
    if(bestIdx===-1)break;
    const chosen=remaining.splice(bestIdx,1)[0];
    const dMin=hav(curLat,curLon,chosen.lat,chosen.lon)*DRIVE_MIN_PER_MILE;
    usedMin+=dMin+(MIN_PER_STOP[chosen.priority]||15);
    route.push(chosen);routeSet.add(chosen.id);
    curLat=chosen.lat;curLon=chosen.lon;
  }

  mapPros=candidates.slice(0,100);
  renderRList();renderMap();renderDayRoute();
  const h=Math.floor(usedMin/60),m=Math.round(usedMin%60);
  const hint=document.getElementById('rhint');
  if(hint)hint.textContent=startLabel+' • '+route.length+' stops • est. '+h+'h '+m+'m of '+hours+'h';
  toast('Built '+route.length+' stops near '+startLabel);
}


function optimizeRoute(){
  if(route.length<2){toast('Add at least 2 stops to optimize');return;}
  if(route.length===2){toast('Route sorted');return;}
  const stops=[...route];
  const ordered=[stops.shift()];
  while(stops.length){
    const last=ordered[ordered.length-1];
    if(!last.lat){ordered.push(stops.shift());continue;}
    let best=0,bestD=Infinity;
    stops.forEach((s,i)=>{if(!s.lat)return;const d=hav(last.lat,last.lon,s.lat,s.lon);if(d<bestD){bestD=d;best=i;}});
    ordered.push(stops.splice(best,1)[0]);
  }
  route=ordered;routeSet=new Set(route.map(r=>r.id));
  renderRList();renderMap();renderDayRoute();
  toast('Route sorted by distance');
}

function openMaps(){
  if(!route.length){toast('Add stops to your route first');return;}
  // Build Google Maps directions URL with all route stops
  const stops=route.filter(p=>p.address);
  if(!stops.length)return;
  const url='https://www.google.com/maps/dir/'+stops.map(p=>enc(p.address+', '+p.city+', FL '+p.zip)).join('/');
  // Use anchor click — most reliable across iOS Safari, PWA, and Chrome
  const a=document.createElement('a');
  a.href=url; a.target='_blank'; a.rel='noopener noreferrer';
  document.body.appendChild(a);
  a.click();
  setTimeout(()=>document.body.removeChild(a),200);
  toast('Opening Google Maps...');
}
function clearRoute(){route=[];routeSet=new Set();routeAnchor=null;renderRList();renderMap();renderDayRoute();}
function clearAnchor(){
  routeAnchor=null;
  renderDayRoute();renderRList();
  toast('Start anchor cleared');
}
function setRouteAnchor(id){
  const p=P.find(x=>x.id===id);if(!p||!p.lat){toast('This business has no map coordinates');return;}
  routeAnchor=p;
  // Update the ZIP field to show anchor is set
  const rzip=document.getElementById('rzip');
  if(rzip)rzip.placeholder='Starting from: '+p.name.slice(0,20)+'...';
  renderRList();
  toast('Route will start from: '+p.name.slice(0,25));
}
function removeStop(id){routeSet.delete(id);route=route.filter(r=>r.id!==id);renderRList();renderMap();renderDayRoute();}

// MODAL
// showCard: opens the full openM modal sheet via iOS-compatible overlay
// Works around iOS Safari inline onclick issues on injected divs
function showCard(id){
  id=parseInt(id)||id;
  var p=P.find(function(x){return x.id===id||x.id==id;});
  if(!p){toast('Not found: '+id);return;}

  var existing=document.getElementById('sc-bg');
  if(existing)existing.remove();

  var lc=getLC(p.id);
  var entries=log[p.id]||[];
  var ph=p.phone||PHONES[String(p.id)]||'';
  var c=customers[p.id]||{};

  // Phone
  var phoneH=ph
    ?'<a href="tel:'+ph.replace(/\D/g,'')+'" style="display:flex;align-items:center;gap:8px;background:#059669;color:#fff;padding:10px 16px;border-radius:8px;font-size:14px;font-weight:700;text-decoration:none;touch-action:manipulation;margin-bottom:10px">📞 '+ph+'</a><div style="display:flex;gap:6px;margin-bottom:10px"><a href="sms:'+ph.replace(/\D/g,'')+'" style="flex:1;text-align:center;padding:6px;border:1px solid #e2e8f0;border-radius:7px;font-size:11px;color:#475569;text-decoration:none">💬 Text</a><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL')+'" target="_blank" style="flex:1;text-align:center;padding:6px;border:1px solid #e2e8f0;border-radius:7px;font-size:11px;color:#475569;text-decoration:none">🔍 Google</a></div>'
    :'<div style="margin-bottom:10px;display:flex;gap:6px;align-items:center"><span style="font-size:11px;color:#94a3b8">No phone on file</span><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone')+'" target="_blank" style="font-size:11px;color:var(--blu);border:1px solid var(--blu);padding:3px 8px;border-radius:5px;text-decoration:none">Find ↗</a></div>';

  // Phone save (if no phone on file)
  var phoneSaveH='';
  if(!ph){
    phoneSaveH='<div style="display:flex;gap:6px;margin-bottom:10px">'
      +'<input id="sc-phone-input" type="tel" placeholder="Paste number here…" '
      +'style="flex:1;padding:8px;border:1px solid #e2e8f0;border-radius:7px;font-size:12px;font-family:inherit;outline:none">'
      +'<button id="sc-phone-save" style="padding:8px 12px;background:#0f1f38;color:#fff;border:none;border-radius:7px;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">Save</button>'
      +'</div>';
  }

  // Chips
  var pBg={CALLBACK:'#fef2f2',HOT:'#fff7ed',WARM:'#fefce8',COOL:'#f8fafc'};
  var pCl={CALLBACK:'#dc2626',HOT:'#f97316',WARM:'#ca8a04',COOL:'#64748b'};
  var pc=pCl[p.priority]||'#64748b';
  var lColor=lc?OI_COLOR[normO(lc.outcome)]||'#64748b':'#64748b';
  var chips=[
    '<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;background:'+(pBg[p.priority]||'#f8fafc')+';color:'+pc+'">'+p.priority+'</span>',
    p.tier&&p.tier!==p.priority?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;background:#f1f5f9;color:#475569">'+p.tier+'</span>':'',
    p.chronic?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;background:#ecfdf5;color:#059669">CHRONIC ICE</span>':'',
    p.days_since>0?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;color:#dc2626;background:#fef2f2">'+p.days_since+'d overdue</span>':'',
    lc?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;font-weight:700;color:'+lColor+';background:'+lColor+'18">'+(OI[normO(lc.outcome)]||lc.outcome)+'</span>':'',
    p.hours?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;background:#f8fafc;color:#64748b">'+p.hours+'</span>':'',
    p.rating>0?'<span style="font-size:9px;padding:3px 8px;border-radius:5px;background:#fffbeb;color:#d97706">'+stars(p.rating)+'</span>':'',
  ].filter(Boolean).join(' ');

  // Ice history
  var iceH='';
  if(p.confirmed||p.chronic){
    var codes=(p.codes||[]).map(function(c){return ICN[c]||c;}).join(', ');
    iceH='<div style="background:#ecfdf5;border:1px solid #6ee7b7;border-radius:8px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:8px;color:#013a18;letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px">🧊 ICE MACHINE HISTORY</div>'
      +(p.chronic?'<div style="font-size:12px;font-weight:700;color:#059669;margin-bottom:2px">CHRONIC — flagged in '+p.ice_count+' inspections</div>':'')
      +(!p.chronic&&p.confirmed?'<div style="font-size:11px;color:#059669;margin-bottom:2px">Confirmed ice machine violation on record</div>':'')
      +(codes?'<div style="font-size:9px;color:#2a5a38">Codes: '+codes+'</div>':'')
      +'</div>';
  }

  // Pitch + walk-in
  var pitchFn=PITCHES[p.pitch_type]||PITCHES.routine;
  var walkinFn=WALKIN[p.pitch_type]||WALKIN.routine;
  var pitchH='<div style="margin-bottom:10px"><div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Pitch Opener</div>'
    +'<div style="background:#fff7ed;border-left:3px solid #c9973a;border-radius:0 8px 8px 0;padding:9px;font-size:11px;color:#475569;line-height:1.7">'+pitchFn(p.name)+'</div></div>';
  var walkinH='<div style="margin-bottom:10px"><div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Walk-In Script</div>'
    +'<div style="background:#f0fdf4;border-left:3px solid #059669;border-radius:0 8px 8px 0;padding:9px;font-size:11px;color:#475569;line-height:1.7">'+walkinFn(p.name)+'</div></div>';

  // Objections
  var objH='<div style="margin-bottom:10px"><div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Objection Handlers</div>'
    +OBJECTIONS.map(function(o){
      return '<details style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;margin-bottom:3px">'
        +'<summary style="padding:6px 9px;font-size:10px;font-weight:600;color:#0f1f38;cursor:pointer;list-style:none">❓ '+o.q+'</summary>'
        +'<div style="padding:5px 9px 8px;font-size:10px;color:#475569;line-height:1.6;border-top:1px solid #e2e8f0">'+o.a+'</div>'
        +'</details>';
    }).join('')+'</div>';

  // Intel grid
  var iceFresh=p.ice_fresh?'Within 6mo':p.ice_recent?'Within 1yr':p.days_since_ice<999?(Math.floor(p.days_since_ice/30)+'mo ago'):'None on record';
  var esc=p.escalation>1.5?'Escalating fast':p.escalation>0.8?'Getting worse':p.escalation>0.3?'Slight trend':'Stable';
  var escCol=p.escalation>1.5?'#dc2626':p.escalation>0.8?'#d97706':'#059669';
  var factRows=[
    ['Predicted Next',p.pred_next||'—','#1e293b'],
    ['Days Until',dL(p.days_until),'#1e293b'],
    ['Last Inspected',p.last_insp||'—','#1e293b'],
    ['Inspections on File',p.n_insp,'#1e293b'],
    ['Ice Violations',p.ice_count>0?p.ice_count+'x flagged':'None',p.ice_count>0?'#dc2626':'#059669'],
    ['Last Ice Violation',iceFresh,p.ice_fresh?'#dc2626':p.ice_recent?'#d97706':'#1e293b'],
    ['Callback Inspections',p.n_callbacks>0?p.n_callbacks+'x (inspector returning)':'None',p.n_callbacks>=2?'#dc2626':p.n_callbacks===1?'#d97706':'#1e293b'],
    ['Disposition Trend',esc,escCol],
    ['High/Total Viol.',p.high_viol+'/'+p.total_viol,p.high_viol>=4?'#dc2626':p.high_viol>=2?'#d97706':'#1e293b'],
    ['Confidence',(p.confidence||0)+'%',p.confidence>=75?'#059669':p.confidence>=50?'#d97706':'#1e293b'],
    ['Est. Machines',p.machines||1,'#1e293b'],
    ['Account Tier',p.tier||'COLD','#1e293b'],
    ['Monthly Recurring','$'+(p.monthly||149)+'/mo','#059669'],
    ['One-Time Clean','$'+(p.onetime||249),'#2563eb'],
    ['Intro Offer','$99 first visit (no commitment)','#ea580c'],
  ];
  var intelH='<div style="margin-bottom:10px"><div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Intelligence</div>'
    +'<div style="background:#f8fafc;border-radius:8px;overflow:hidden;border:1px solid #e2e8f0">'
    +factRows.map(function(r,i){
      return '<div style="display:flex;justify-content:space-between;padding:5px 10px;background:'+(i%2===0?'#f8fafc':'#fff')+'">'
        +'<div style="font-size:10px;color:#64748b">'+r[0]+'</div>'
        +'<div style="font-size:10px;font-weight:600;color:'+r[2]+'">'+r[1]+'</div>'
        +'</div>';
    }).join('')+'</div></div>';

  // History
  var histH='';
  if(entries.length){
    histH='<div style="margin-bottom:10px"><div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Call History</div>'
      +[...entries].reverse().slice(0,6).map(function(e){
        var ec=OI_COLOR[normO(e.outcome)]||'#64748b';
        return '<div style="padding:5px 0;border-bottom:1px solid #f1f5f9">'
          +'<div style="display:flex;gap:6px;align-items:baseline">'
          +'<span style="font-size:10px;color:'+ec+';font-weight:700">'+(OI[normO(e.outcome)]||e.outcome)+'</span>'
          +'<span style="font-size:9px;color:#94a3b8">'+e.date+'</span>'
          +(e.reason?'<span style="font-size:9px;color:#94a3b8">· '+e.reason+'</span>':'')
          +'</div>'
          +(e.notes?'<div style="font-size:9px;color:#64748b;font-style:italic">'+e.notes+'</div>':'')
          +(e.followup?'<div style="font-size:9px;color:#0891b2">📅 Follow-up: '+e.followup+'</div>':'')
          +'</div>';
      }).join('')+'</div>';
  }

  // Service history (past logged visits)
  var svcHistH='';
  var svcHistory=(customers[p.id]||{}).service_history||[];
  if(svcHistory.length){
    svcHistH='<div style="margin-bottom:10px"><div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Service History</div>'
      +[...svcHistory].reverse().slice(0,4).map(function(sv){
        var atpStr=sv.atp?(' · ATP: '+sv.atp+' RLU'):'';  
        return '<div style="padding:5px 0;border-bottom:1px solid #f1f5f9">'
          +'<div style="font-size:10px;font-weight:700;color:#0f1f38">'+sv.date_display+(sv.type==='deep_clean'?' · 🧼 Deep Clean':' · 🔧 60-Day Maint.')+atpStr+'</div>'
          +(sv.notes?'<div style="font-size:9px;color:#475569;font-style:italic">'+sv.notes+'</div>':'')  
          +'</div>';
      }).join('')
    +'</div>';
  }

  // Outcome log section
  var logH='<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:12px;margin-bottom:10px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Log Interaction</div>'
    // Type toggle
    +'<div style="display:flex;gap:6px;margin-bottom:8px">'
    +'<button id="sc-type-call" data-sctype="call" style="flex:1;padding:7px;border:1px solid #e2e8f0;border-radius:7px;background:#0f1f38;color:#fff;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">📞 Call</button>'
    +'<button id="sc-type-walkin" data-sctype="walkin" style="flex:1;padding:7px;border:1px solid #e2e8f0;border-radius:7px;background:#f8fafc;color:#475569;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">🚶 Walk-In</button>'
    +'</div>'
    // Outcome buttons
    +'<div id="sc-obtn-row" style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px;margin-bottom:8px">'
    +'<button data-scout="signed" style="padding:8px 4px;border:2px solid #6ee7b7;border-radius:7px;background:#ecfdf5;color:#059669;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">✅ Signed</button>'
    +'<button data-scout="intro_set" style="padding:8px 4px;border:2px solid #93c5fd;border-radius:7px;background:#eff6ff;color:#2563eb;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">📅 Intro Set</button>'
    +'<button data-scout="in_play" style="padding:8px 4px;border:2px solid #fcd34d;border-radius:7px;background:#fffbeb;color:#d97706;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">🟡 In Play</button>'
    +'<button data-scout="no_contact" style="padding:8px 4px;border:2px solid #c4b5fd;border-radius:7px;background:#f5f3ff;color:#7c3aed;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">🚪 No Contact</button>'
    +'<button data-scout="voicemail" style="padding:8px 4px;border:2px solid #fdba74;border-radius:7px;background:#fff7ed;color:#ea580c;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">📲 Voicemail</button>'
    +'<button data-scout="not_now" style="padding:8px 4px;border:2px solid #fca5a5;border-radius:7px;background:#fef2f2;color:#dc2626;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">❌ Not Now</button>'
    +'<button data-scout="dead" style="padding:8px 4px;border:2px solid #94a3b8;border-radius:7px;background:#f1f5f9;color:#475569;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation;grid-column:1/3">⚫ Dead — wrong fit/hard no</button>'
    +'<button data-scout="service_done" style="padding:8px 4px;border:2px solid #86efac;border-radius:7px;background:#f0fdf4;color:#059669;font-size:10px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">🧼 Service Done</button>'
    +'</div>'
    // Reason picker
    +'<div id="sc-reason-wrap" style="display:none;margin-bottom:8px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;margin-bottom:4px">Reason</div>'
    +'<div id="sc-reason-grid" style="display:flex;flex-wrap:wrap;gap:4px"></div>'
    +'</div>'
    // Notes + follow-up
    +'<textarea id="sc-notes" placeholder="Notes…" rows="2" style="width:100%;padding:8px;border:1px solid #e2e8f0;border-radius:7px;font-size:12px;font-family:inherit;resize:none;outline:none;box-sizing:border-box;margin-bottom:6px"></textarea>'
    +'<div style="margin-bottom:8px">'
    +'<div style="font-size:9px;color:#94a3b8;font-weight:600;margin-bottom:5px">📅 Follow-up date (tap to set)</div>'
    +'<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:4px;margin-bottom:5px">'
    +'<button data-fupdays="1" style="padding:9px 4px;border:1px solid #e2e8f0;border-radius:7px;background:#f8fafc;color:#475569;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">+1d</button>'
    +'<button data-fupdays="3" style="padding:9px 4px;border:1px solid #e2e8f0;border-radius:7px;background:#f8fafc;color:#475569;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">+3d</button>'
    +'<button data-fupdays="7" style="padding:9px 4px;border:1px solid #e2e8f0;border-radius:7px;background:#f8fafc;color:#475569;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">+7d</button>'
    +'<button data-fupdays="14" style="padding:9px 4px;border:1px solid #e2e8f0;border-radius:7px;background:#f8fafc;color:#475569;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">+14d</button>'
    +'<button data-fupdays="30" style="padding:9px 4px;border:1px solid #e2e8f0;border-radius:7px;background:#f8fafc;color:#475569;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">+30d</button>'
    +'</div>'
    +'<div id="sc-fup-display" style="font-size:11px;color:#059669;font-weight:700;min-height:18px"></div>'
    +'<input id="sc-followup" type="hidden" value="">'  // value set by button taps
    +'</div>'

    +    +'<button id="sc-save-btn" style="width:100%;padding:11px;background:#059669;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">Save Log Entry</button>'
    +'</div>';

  // CLOSE DEAL section  
  var closeH='<div style="margin-bottom:12px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Close Deal</div>'
    +'<button id="sc-won-intro" style="width:100%;padding:10px;border:2px solid #f97316;border-radius:9px;background:#fff7ed;color:#ea580c;font-weight:800;font-size:12px;cursor:pointer;font-family:inherit;touch-action:manipulation;margin-bottom:6px">'
    +'🔥 Intro Offer — $99 first visit<br><span style="font-size:10px;font-weight:400">Try us once, no commitment. Recurring after.</span></button>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:6px">'
    +'<button id="sc-won-rec" style="padding:10px 6px;border:2px solid #059669;border-radius:9px;background:#ecfdf5;color:#059669;font-weight:700;font-size:11px;cursor:pointer;font-family:inherit;touch-action:manipulation">'
    +'💰 Won — Recurring<br><span id="sc-won-monthly" style="font-size:10px;font-weight:400">$'+( p.monthly||149)+'/mo</span></button>'
    +'<button id="sc-won-once" style="padding:10px 6px;border:2px solid #2563eb;border-radius:9px;background:#eff6ff;color:#2563eb;font-weight:700;font-size:11px;cursor:pointer;font-family:inherit;touch-action:manipulation">'
    +'🧼 Won — One-Time<br><span style="font-size:10px;font-weight:400">$'+(p.onetime||249)+'</span></button>'
    +'</div>'
    +'<button id="sc-lost-btn" style="width:100%;padding:7px;border:1px solid #e2e8f0;border-radius:8px;background:transparent;color:#94a3b8;font-size:10px;cursor:pointer;font-family:inherit;touch-action:manipulation">'
    +'Mark as Lost / Churned</button>'
    +'</div>';

  // Contacts & Intel section
  var contactsH='<div style="margin-bottom:12px">'
    +'<div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Contacts & Intel</div>'
    // Current vendor
    +'<div style="margin-bottom:8px;padding:8px;background:#fef9ee;border:1px solid #fde68a;border-radius:7px">'
    +'<div style="font-size:9px;font-weight:700;color:#92400e;margin-bottom:4px">🕵️ Current Ice Vendor</div>'
    +'<div style="display:flex;gap:5px">'
    +'<input id="sc-vendor" type="text" value="'+(c.vendor_name||'')+'" placeholder="e.g. Ecolab, staff cleans it…" '
    +'style="flex:1;padding:6px;border:1px solid #fde68a;border-radius:6px;font-size:11px;font-family:inherit;background:#fff;color:#1e293b;outline:none">'
    +'<button id="sc-save-vendor" style="padding:6px 10px;border:none;border-radius:6px;background:#92400e;color:#fff;font-size:10px;font-weight:600;cursor:pointer;font-family:inherit;touch-action:manipulation">Save</button>'
    +'</div></div>'
    // Contacts
    +'<div id="sc-contacts-wrap"><div style="font-size:10px;color:#94a3b8">Loading contacts…</div></div>'
    +'<button id="sc-add-contact" style="width:100%;padding:7px;border:1px dashed #e2e8f0;border-radius:7px;background:transparent;color:#64748b;font-size:10px;cursor:pointer;font-family:inherit;touch-action:manipulation;margin-top:4px">+ Add Contact</button>'
    +'</div>';

  // Build overlay
  var bg=document.createElement('div');
  bg.id='sc-bg';
  bg.style.cssText='position:fixed;inset:0;z-index:500;background:rgba(15,31,56,.75);display:flex;align-items:flex-end;justify-content:center';

  bg.innerHTML=
    '<div id="sc-sheet" style="background:#fff;width:100%;max-width:640px;max-height:93vh;border-radius:20px 20px 0 0;overflow-y:auto;-webkit-overflow-scrolling:touch;overscroll-behavior:contain">'
    +'<div style="padding:10px 0 0;text-align:center"><div style="width:36px;height:4px;background:#e2e8f0;border-radius:2px;margin:0 auto"></div></div>'
    // Sticky header
    +'<div style="position:sticky;top:0;background:#fff;border-bottom:1px solid #e2e8f0;padding:12px 16px;display:flex;justify-content:space-between;align-items:flex-start;z-index:10">'
      +'<div style="flex:1;min-width:0">'
        +'<div style="font-size:17px;font-weight:800;color:#0f1f38;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">'+p.name+'</div>'
        +'<div style="font-size:10px;color:#94a3b8">'+p.address+', '+p.city+', FL '+p.zip+' · #'+p.id+'</div>'
      +'</div>'
      +'<div style="display:flex;gap:6px;align-items:center;flex-shrink:0;margin-left:10px">'
    +'<button id="sc-route-btn" style="border:none;background:#f0fdf4;border-radius:8px;padding:6px 10px;font-size:11px;font-weight:700;color:#059669;cursor:pointer;touch-action:manipulation;font-family:inherit;-webkit-tap-highlight-color:transparent">📍 +Route</button>'
    +'<button id="sc-close" style="border:none;background:#f1f5f9;border-radius:50%;width:34px;height:34px;font-size:17px;cursor:pointer;display:flex;align-items:center;justify-content:center;touch-action:manipulation;color:#475569;font-family:inherit">✕</button>'
    +'</div>'
    +'</div>'
    +'<div style="padding:16px">'
      +'<div style="display:flex;gap:5px;flex-wrap:wrap;margin-bottom:12px">'+chips+'</div>'
      +phoneH+phoneSaveH+iceH+intelH+pitchH+walkinH+objH+histH+svcHistH+logH+closeH+contactsH
    +'</div></div>';

  document.body.appendChild(bg);

  // ── Wire all interactive elements ──────────────────────────────────────────
  var selOutcome=null, selType='call', selReason=null;

  // Phone save
  if(!ph){
    var phoneSaveBtn=document.getElementById('sc-phone-save');
    if(phoneSaveBtn){
      function doPhoneSave(e2){
        if(e2&&e2.type==='touchend')e2.preventDefault();
        var val=(document.getElementById('sc-phone-input')||{}).value||'';
        if(!val){toast('Enter a phone number');return;}
        phSave(p.id,val,p.hours||'',p.rating||0);
        p.phone=val;
        toast('Phone saved');
        bg.remove();
        showCard(p.id); // reopen with phone
      }
      phoneSaveBtn.addEventListener('touchend',doPhoneSave,false);
      phoneSaveBtn.addEventListener('click',doPhoneSave,false);
    }
  }

  // Route button
  var routeBtn=document.getElementById('sc-route-btn');
  if(routeBtn){
    function doRoute(e){if(e&&e.type==='touchend')e.preventDefault();addToRoute(p.id);toast('Added to route');}
    routeBtn.addEventListener('touchend',doRoute,false);
    routeBtn.addEventListener('click',doRoute,false);
  }
  // Close
  function doClose(e){if(e&&e.type==='touchend')e.preventDefault();bg.remove();}
  var closeBtn=document.getElementById('sc-close');
  closeBtn.addEventListener('touchend',doClose,false);
  closeBtn.addEventListener('click',doClose,false);
  bg.addEventListener('touchend',function(e){if(e.target===bg){e.preventDefault();bg.remove();}},false);

  // Follow-up day buttons — wire BEFORE type toggle
  var fupContainer=document.getElementById('sc-sheet');
  if(fupContainer){
    fupContainer.querySelectorAll('[data-fupdays]').forEach(function(btn){
      function pickFup(e){
        if(e.type==='touchend')e.preventDefault();
        var days=parseInt(btn.dataset.fupdays);
        var d=new Date(Date.now()+days*864e5);
        var iso=d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0');
        var fupInput=document.getElementById('sc-followup');
        if(fupInput)fupInput.value=iso;
        var disp=document.getElementById('sc-fup-display');
        if(disp)disp.textContent=d.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
        // Update button styles
        fupContainer.querySelectorAll('[data-fupdays]').forEach(function(b){
          b.style.background=b===btn?'#0f1f38':'#f8fafc';
          b.style.color=b===btn?'#fff':'#475569';
        });
      }
      btn.addEventListener('touchend',pickFup,false);
      btn.addEventListener('click',pickFup,false);
    });
  }

  // Type toggle
  var typeBtns=document.querySelectorAll('#sc-sheet [data-sctype]');
  function pickType(e){
    if(e.type==='touchend')e.preventDefault();
    selType=e.currentTarget.dataset.sctype;
    typeBtns.forEach(function(b){
      b.style.background=b===e.currentTarget?'#0f1f38':'#f8fafc';
      b.style.color=b===e.currentTarget?'#fff':'#475569';
    });
  }
  typeBtns.forEach(function(b){b.addEventListener('touchend',pickType,false);b.addEventListener('click',pickType,false);});

  // Outcome buttons
  var obtnRow=document.getElementById('sc-obtn-row');
  var reasonWrap=document.getElementById('sc-reason-wrap');
  var reasonGrid=document.getElementById('sc-reason-grid');
  function pickOutcome(btn){
    selOutcome=btn.dataset.scout;
    selReason=null;
    obtnRow.querySelectorAll('[data-scout]').forEach(function(b){
      b.style.opacity=b===btn?'1':'0.45';
      b.style.fontWeight=b===btn?'800':'700';
    });
    // Reason picker
    var reasonMap={
      not_now:['Already has service','Too expensive','Cleans themselves','Not the decision maker','Call back later','No reason given'],
      in_play:['Interested, needs partner','Needs corporate approval','Seasonal','Franchise contract','Just passed inspection'],
      dead:['Wrong fit','Hard no','Out of business','Corporate contract','Repeated refusals']
    };
    // Special: 'Signed' is a shortcut to Won-Recurring (creates customer record)
    if(selOutcome==='signed'){
      toast('Use the Close Deal buttons below (Won Recurring/One-Time/Intro) to record a signed deal');
      selOutcome=null;
      obtnRow.querySelectorAll('[data-scout]').forEach(function(b){b.style.opacity='1';});
      return;
    }
    if(reasonMap[selOutcome]){
      reasonGrid.innerHTML=reasonMap[selOutcome].map(function(r){
        return '<button data-scr="'+r+'" style="padding:4px 8px;border:1px solid #e2e8f0;border-radius:5px;background:#fff;font-size:10px;cursor:pointer;font-family:inherit;touch-action:manipulation">'+r+'</button>';
      }).join('');
      reasonWrap.style.display='block';
      // Wire reason chips
      reasonGrid.querySelectorAll('[data-scr]').forEach(function(rb){
        function pickR(e){
          if(e.type==='touchend')e.preventDefault();
          selReason=rb.dataset.scr;
          reasonGrid.querySelectorAll('[data-scr]').forEach(function(b){
            b.style.background=b===rb?'#0f1f38':'#fff';
            b.style.color=b===rb?'#fff':'#1e293b';
          });
        }
        rb.addEventListener('touchend',pickR,false);
        rb.addEventListener('click',pickR,false);
      });
    } else {
      reasonWrap.style.display='none';
    }
  }
  obtnRow.querySelectorAll('[data-scout]').forEach(function(btn){
    btn.addEventListener('touchend',function(e){e.preventDefault();pickOutcome(btn);},false);
    btn.addEventListener('click',function(){pickOutcome(btn);},false);
  });

  // Save log
  function doSaveLog(e){
    if(e&&e.type==='touchend')e.preventDefault();
    if(!selOutcome){toast('Pick an outcome first');return;}
    var notes=(document.getElementById('sc-notes')||{}).value||'';
    var followup=(document.getElementById('sc-followup')||{}).value||'';
    // Require follow-up date for anything that's not a terminal outcome
    var needsFollowup=['not_now','in_play'].indexOf(selOutcome)>=0;
    if(needsFollowup&&!followup){
      toast('Follow-up date required for this outcome');
      var fup=document.getElementById('sc-followup');
      if(fup){fup.style.border='2px solid #dc2626';fup.focus();}
      return;
    }
    if(!log[p.id])log[p.id]=[];
    log[p.id].push({outcome:selOutcome,type:selType,reason:selReason,
      date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),
      notes:notes,followup:followup});
    lSave();
    var msg=followup?'✓ '+(OI[selOutcome]||selOutcome)+' — follow-up '+followup:'✓ '+(OI[selOutcome]||selOutcome)+' logged';
    toast(msg);
    bg.remove();
    if(_queueAutoAdvance){_queueAutoAdvance=false;queueIdx++;renderQueueCard();}
    else if(tab==='today')renderBriefing();
    else if(tab==='all')rA();
    else renderKPIs();
  }
  var saveBtn=document.getElementById('sc-save-btn');
  saveBtn.addEventListener('touchend',doSaveLog,false);
  saveBtn.addEventListener('click',doSaveLog,false);

  // markWon helper
  function doMarkWon(status){
    var now=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
    customers[p.id]={
      status:status,won_date:now,
      service_type:status==='customer_recurring'?'recurring':status==='customer_intro'?'intro':'one_time',
      monthly:status==='customer_recurring'?p.monthly:0,
      onetime:status==='customer_once'?p.onetime:status==='customer_intro'?99:0,
      machines:p.machines,name:p.name,address:p.address,city:p.city,phone:ph,
      notes:'',last_service:'',next_service:'',hubspot_url:'',square_url:'',
      machine_brand:'',machine_model:'',machine_type:'',filter_type:'',
      filter_installed:'',contract_start:'',contract_term:6,contract_renewal:'',
      service_history:[],atp_history:[],vendor_name:'',
    };
    custSave();
    // CRITICAL: set p.status so rCust() filter finds this prospect
    p.status=status;
    if(!log[p.id])log[p.id]=[];
    log[p.id].push({outcome:status,date:now,notes:'Deal closed'});
    lSave();
    if(status==='customer_recurring')buildAnnualSchedule(p.id);
    if(status==='customer_intro')buildAnnualSchedule(p.id);
    bg.remove();
    var msg=status==='customer_intro'?'🔥 Intro booked! $99 first visit':
            status==='customer_recurring'?'🎉 Won! Added to recurring clients.':
            status==='churned'?'Marked lost/churned':'🧼 Won! One-time service recorded.';
    toast(msg);
    sw('customers');
  }

  // Close deal buttons
  function wireWon(btnId,status){
    var btn=document.getElementById(btnId);
    if(!btn)return;
    btn.addEventListener('touchend',function(e){e.preventDefault();doMarkWon(status);},false);
    btn.addEventListener('click',function(){doMarkWon(status);},false);
  }
  wireWon('sc-won-intro','customer_intro');
  wireWon('sc-won-rec','customer_recurring');
  wireWon('sc-won-once','customer_once');
  wireWon('sc-lost-btn','churned');

  // Vendor save
  var vendorSaveBtn=document.getElementById('sc-save-vendor');
  function doSaveVendor(e){
    if(e&&e.type==='touchend')e.preventDefault();
    var val=(document.getElementById('sc-vendor')||{}).value||'';
    if(!customers[p.id])customers[p.id]={name:p.name,address:p.address,city:p.city,
      phone:ph,machines:p.machines,status:'prospect',notes:'',vendor_name:''};
    customers[p.id].vendor_name=val;
    custSave();
    toast('✓ Vendor intel saved');
    // Visual confirmation on button
    if(vendorSaveBtn){
      var orig=vendorSaveBtn.textContent;
      vendorSaveBtn.textContent='✓ Saved';
      vendorSaveBtn.style.background='#059669';
      setTimeout(function(){vendorSaveBtn.textContent=orig;vendorSaveBtn.style.background='#92400e';},1500);
    }
  }
  vendorSaveBtn.addEventListener('touchend',doSaveVendor,false);
  vendorSaveBtn.addEventListener('click',doSaveVendor,false);

  // Contacts
  var cWrap=document.getElementById('sc-contacts-wrap');
  var contacts_data=contacts[p.id]||[];
  function renderSCContacts(){
    if(!contacts_data.length){cWrap.innerHTML='<div style="font-size:10px;color:#94a3b8">No contacts added yet</div>';return;}
    cWrap.innerHTML=contacts_data.map(function(ct,i){
      return '<div style="display:flex;justify-content:space-between;align-items:center;padding:5px 0;border-bottom:1px solid #f1f5f9">'
        +'<div><div style="font-size:11px;font-weight:600;color:#1e293b">'+ct.name+'</div>'
        +'<div style="font-size:9px;color:#94a3b8">'+ct.role+(ct.phone?' · '+ct.phone:'')+'</div>'
        +(ct.notes?'<div style="font-size:9px;color:#64748b;font-style:italic">'+ct.notes+'</div>':'')
        +'</div>'
        +'<button data-ci="'+i+'" style="border:none;background:transparent;color:#dc2626;font-size:12px;cursor:pointer;touch-action:manipulation">✕</button>'
        +'</div>';
    }).join('');
    // Wire delete
    cWrap.querySelectorAll('[data-ci]').forEach(function(btn){
      function delContact(e){
        if(e.type==='touchend')e.preventDefault();
        contacts_data.splice(parseInt(btn.dataset.ci),1);
        contacts[p.id]=contacts_data;
        contactsSave();
        renderSCContacts();
      }
      btn.addEventListener('touchend',delContact,false);
      btn.addEventListener('click',delContact,false);
    });
  }
  renderSCContacts();

  // Add contact
  var addContactBtn=document.getElementById('sc-add-contact');
  function doAddContact(e){
    if(e&&e.type==='touchend')e.preventDefault();
    // Show inline form instead of prompt (prompt blocked in PWA)
    var existing=document.getElementById('sc-contact-form');
    if(existing){existing.remove();return;}
    var form=document.createElement('div');
    form.id='sc-contact-form';
    form.style.cssText='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px;margin-top:6px';
    form.innerHTML=
      '<div style="font-size:9px;font-weight:700;color:#94a3b8;margin-bottom:6px">NEW CONTACT</div>'
      +'<input id="sc-ct-name" type="text" placeholder="Name" style="width:100%;padding:7px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;font-family:inherit;outline:none;box-sizing:border-box;margin-bottom:5px">'
      +'<input id="sc-ct-role" type="text" placeholder="Role (Owner, GM, Manager…)" style="width:100%;padding:7px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;font-family:inherit;outline:none;box-sizing:border-box;margin-bottom:5px">'
      +'<input id="sc-ct-phone" type="tel" placeholder="Phone (optional)" style="width:100%;padding:7px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;font-family:inherit;outline:none;box-sizing:border-box;margin-bottom:5px">'
      +'<input id="sc-ct-notes" type="text" placeholder="Notes (optional)" style="width:100%;padding:7px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;font-family:inherit;outline:none;box-sizing:border-box;margin-bottom:8px">'
      +'<div style="display:flex;gap:6px">'
      +'<button id="sc-ct-save" style="flex:1;padding:8px;background:#059669;color:#fff;border:none;border-radius:7px;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">Save Contact</button>'
      +'<button id="sc-ct-cancel" style="padding:8px 12px;background:#f1f5f9;color:#475569;border:none;border-radius:7px;font-size:12px;cursor:pointer;font-family:inherit;touch-action:manipulation">Cancel</button>'
      +'</div>';
    addContactBtn.insertAdjacentElement('afterend', form);
    // Wire save
    function saveNewContact(e2){
      if(e2&&e2.type==='touchend')e2.preventDefault();
      var name=(document.getElementById('sc-ct-name')||{}).value||'';
      if(!name){toast('Name required');return;}
      var role=(document.getElementById('sc-ct-role')||{}).value||'';
      var phone=(document.getElementById('sc-ct-phone')||{}).value||'';
      var notes=(document.getElementById('sc-ct-notes')||{}).value||'';
      if(!contacts[p.id])contacts[p.id]=[];
      contacts[p.id].push({name:name,role:role,phone:phone,notes:notes});
      contacts_data=contacts[p.id];
      contactsSave();
      form.remove();
      renderSCContacts();
      toast('Contact saved');
    }
    var saveCt=document.getElementById('sc-ct-save');
    saveCt.addEventListener('touchend',saveNewContact,false);
    saveCt.addEventListener('click',saveNewContact,false);
    var cancelCt=document.getElementById('sc-ct-cancel');
    function cancelForm(e2){if(e2&&e2.type==='touchend')e2.preventDefault();form.remove();}
    cancelCt.addEventListener('touchend',cancelForm,false);
    cancelCt.addEventListener('click',cancelForm,false);
  }
  addContactBtn.addEventListener('touchend',doAddContact,false);
  addContactBtn.addEventListener('click',doAddContact,false);
}

function openM(id){
  id=parseInt(id)||id;const p=P.find(x=>x.id===id||x.id==id);if(!p){toast('Not found: '+id+' type:'+typeof id);return;}
  try{
  _modalOpenedAt=Date.now();
  cur=p;selOut=null;
  document.querySelectorAll('.obtn').forEach(b=>b.classList.remove('on'));
  selType=null;selReasonVal=null;
  document.getElementById('mtype-walkin').className='mtype-btn';
  document.getElementById('mtype-call').className='mtype-btn';
  const rw2=document.getElementById('reason-wrap');if(rw2)rw2.style.display='none';
  document.getElementById('mnotes').value='';
  document.getElementById('mn').textContent=p.name;
  document.getElementById('ml').textContent=p.address+', '+p.city+', FL '+p.zip+' \u00b7 #'+p.id;
  const pe=document.getElementById('mph'),ae=document.getElementById('mpa');
  if(p.phone){
    pe.textContent=p.phone;pe.className='mphnum';
    const cl=p.phone.replace(/\\s/g,'');
    ae.innerHTML='<a href="tel:'+cl+'" class="mcall">Call Now</a><a href="sms:'+cl+'" class="msms">Text</a><a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL')+'" target="_blank" class="mgoog">Google</a>';
  }else{
    pe.textContent='No phone on file';pe.className='mphnum none';
    ae.innerHTML='<a href="https://www.google.com/search?q='+enc(p.name+' '+p.city+' FL phone number')+'" target="_blank" class="mgoog">Find Phone</a><a href="https://maps.google.com/search?q='+enc(p.name+' '+p.address+' '+p.city+' FL')+'" target="_blank" class="mgoog">Maps</a>';
  }
  document.getElementById('mhrs').textContent=p.hours?'Hours: '+p.hours:'';
  document.getElementById('mrat').textContent=p.rating>0?stars(p.rating):'';
  const iceEl=document.getElementById('mice');
  if(p.confirmed||p.chronic){
    const codes=(p.codes||[]).map(c=>ICN[c]||c).join(', ');
    const parts=['<div class="micebox">'];
    parts.push('<div style="font-size:8px;color:#013a18;letter-spacing:.08em;text-transform:uppercase;margin-bottom:4px">&#x1F9CA; ICE MACHINE HISTORY</div>');
    if(p.chronic) parts.push('<div style="font-size:13px;font-weight:700;color:var(--grn);margin-bottom:3px">CHRONIC  -  flagged in '+p.ice_count+' inspections</div>');
    if(!p.chronic&&p.confirmed) parts.push('<div style="font-size:12px;color:#55c8ff;margin-bottom:3px">Confirmed ice machine violation on record</div>');
    if(codes) parts.push('<div style="font-size:9px;color:#2a5a38;margin-bottom:3px">Violation codes: '+codes+'</div>');
    if('') parts.push('<div style="font-size:10px;color:#3a6a48;font-style:italic">'+''+'</div>');
    parts.push('</div>');
    iceEl.innerHTML=parts.join('');
  }else iceEl.innerHTML='';
  document.getElementById('mpitch').innerHTML=(PITCHES[p.pitch_type]||PITCHES.routine)(p.name);
  document.getElementById('mwalkin').innerHTML=(WALKIN[p.pitch_type]||WALKIN.routine)(p.name);
  // Objection handlers
  const objEl=document.getElementById('mobjections');
  if(objEl){
    objEl.innerHTML=OBJECTIONS.map((o,i)=>
      '<details style="background:#f5f8fa;border:1px solid var(--brd2);border-radius:7px;padding:0">'
      +'<summary style="padding:7px 10px;font-size:10px;font-weight:600;color:var(--navy);cursor:pointer;list-style:none;display:flex;align-items:center;gap:6px">'
      +'<span style="color:var(--cb)">&#x2753;</span>'+o.q+'</summary>'
      +'<div style="padding:4px 10px 9px;font-size:10px;color:var(--sub2);line-height:1.6;border-top:1px solid var(--brd2)">'+o.a+'</div>'
      +'</details>'
    ).join('');
  }
  // Build intelligence signals display
  const iceFreshness=p.ice_fresh?'Within 6mo':p.ice_recent?'Within 1yr':p.days_since_ice<999?(Math.floor(p.days_since_ice/30)+'mo ago'):'None on record';
  const iceFreshnessCol=p.ice_fresh?'r':p.ice_recent?'o':'';
  const escalationTxt=p.escalation>1.5?'Escalating fast':p.escalation>0.8?'Getting worse':p.escalation>0.3?'Slight trend up':'Stable';
  const escalationCol=p.escalation>1.5?'r':p.escalation>0.8?'o':'g';
  document.getElementById('mfacts').innerHTML=[
    // Timing
    ['Predicted Next',   p.pred_next,''],
    ['Days Until',       dL(p.days_until),''],
    ['Last Inspected',   p.last_insp,''],
    ['Days Since Insp.', p.days_since+'d',''],
    ['Inspections on File',p.n_insp,''],
    // Ice intelligence
    ['Ice Violations',   p.ice_count>0?p.ice_count+'x flagged':'None on record',p.ice_count>0?'r':''],
    ['Last Ice Violation',iceFreshness,iceFreshnessCol],
    ['Violation Breadth',p.code_diversity>0?p.code_diversity+' violation types':'None',p.code_diversity>=3?'r':p.code_diversity>=2?'o':''],
    // Inspection behavior
    ['Callback Inspections',p.n_callbacks>0?p.n_callbacks+'x (inspector returning)':'None',p.n_callbacks>=2?'r':p.n_callbacks===1?'o':''],
    ['Disposition Trend', escalationTxt,escalationCol],
    ['Failed First Visit',p.avg_visit>1.2?'Yes (avg '+p.avg_visit+'x visits)':'No',''],
    ['High/Total Viol',  p.high_viol+'/'+p.total_viol,p.high_viol>=4?'r':p.high_viol>=2?'o':''],
    ['Viol. Trajectory', p.trending?'Getting Worse':'Stable',p.trending?'r':'g'],
    // Account
    ['Confidence',       (p.confidence||0)+'%',p.confidence>=75?'g':p.confidence>=50?'o':''],
    ['Est. Machines',    p.machines||1,''],
    ['Account Tier',     p.tier||'COLD',''],
    ['Monthly Recurring','$'+(p.monthly||149)+'/mo','g'],
    ['One-Time Clean',   '$'+(p.onetime||249),'b'],
    ['Intro Offer',      '$99 first visit (no commitment)','o'],
  ].map(([l,v,c])=>'<div class="fact"><div class="fl">'+l+'</div><div class="fv '+c+'">'+v+'</div></div>').join('');
  const hist=log[id]||[];
  const hs=document.getElementById('mhs');
  if(hist.length){hs.style.display='block';document.getElementById('mhist').innerHTML=[...hist].reverse().map(e=>{
    const noteHtml=e.notes?('<div style="color:var(--sub);font-size:9px;margin-top:1px">'+e.notes+'</div>'):'';
    return '<div class="hi">'+(OI[e.outcome]||e.outcome)+'<span style="color:var(--sub);font-size:9px"> &middot; '+e.date+'</span>'+noteHtml+'</div>';
  }).join('');}
  else hs.style.display='none';
  // Show tiered pricing breakdown
  const machineBreakdown=p.machines<=1?'1 machine':
    p.machines===2?'2 machines ($149 + $89)':
    p.machines+'machines ($149 + $89 + $69×'+(p.machines-2)+')';
  document.getElementById('mwon-monthly').textContent='$'+p.monthly+'/mo  -  '+machineBreakdown;
  document.getElementById('mwon-onetime').textContent='$'+p.onetime+' one-time ('+p.machines+' machine'+(p.machines>1?'s':'')+')';
  renderContacts(id);
  renderVendor(id);
  document.getElementById('mbg').classList.add('on');
  }catch(err){toast('Open error: '+err.message);console.error(err);}
}
function closeM(e){
  // Ignore ghost clicks fired within 500ms of openM (iOS synthesized click bug)
  if(Date.now()-_modalOpenedAt<500)return;
  // Only close if the actual backdrop was clicked (not a child element)
  if(e&&e.target&&e.target.id==='mbg')document.getElementById('mbg').classList.remove('on');
}
function closeMForce(){
  document.getElementById('mbg').classList.remove('on');
}
function setPitchMode(mode){
  const phoneBtn=document.getElementById('btn-phone');
  const walkinBtn=document.getElementById('btn-walkin');
  const phoneDiv=document.getElementById('mpitch');
  const walkinDiv=document.getElementById('mwalkin');
  if(mode==='walkin'){
    phoneDiv.style.display='none';walkinDiv.style.display='block';
    phoneBtn.style.background='transparent';phoneBtn.style.color='var(--sub)';phoneBtn.style.borderColor='var(--brd)';
    walkinBtn.style.background='#0a84ff22';walkinBtn.style.color='var(--blu)';walkinBtn.style.borderColor='var(--blu)';
  }else{
    phoneDiv.style.display='block';walkinDiv.style.display='none';
    phoneBtn.style.background='#0a84ff22';phoneBtn.style.color='var(--blu)';phoneBtn.style.borderColor='var(--blu)';
    walkinBtn.style.background='transparent';walkinBtn.style.color='var(--sub)';walkinBtn.style.borderColor='var(--brd)';
  }
}
function selO(o){
  selOut=o;selReasonVal=null;
  document.querySelectorAll('.obtn').forEach(b=>b.classList.toggle('on',b.dataset.o===o));
  // Show/hide reason picker
  const rw=document.getElementById('reason-wrap');
  const rg=document.getElementById('reason-grid');
  if(rw&&rg){
    if(REASON_OUTCOMES.has(o)){
      rg.innerHTML=Object.entries(REASONS).map(([k,v])=>
        `<button class="reason-chip" onclick="selReason('${k}',this)">${v}</button>`
      ).join('');
      rw.style.display='block';
    } else {
      rw.style.display='none';
      rg.innerHTML='';
    }
  }
}
function selReason(r,el){
  selReasonVal=r;
  document.querySelectorAll('.reason-chip').forEach(c=>c.classList.remove('on'));
  if(el)el.classList.add('on');
}
function saveL(){
  if(!cur){toast('No business selected');return;}
  // Default to no_answer if no outcome selected
  if(!selOut){
    selOut='no_contact';
    document.querySelectorAll('.obtn').forEach(b=>b.classList.toggle('on',b.dataset.o==='no_contact'));
  }
  const notes=document.getElementById('mnotes').value.trim();
  const followup=(document.getElementById('mfollowup')||{}).value||'';
  const e={outcome:selOut,type:selType||null,reason:selReasonVal||null,
    date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),notes,followup};
  if(!log[cur.id])log[cur.id]=[];
  log[cur.id].push(e);lSave();
  document.getElementById('mbg').classList.remove('on');
  if(document.getElementById('mfollowup'))document.getElementById('mfollowup').value='';
  const msg=followup
    ?'✓ '+(OI[selOut]||selOut)+' logged  •  Follow-up set for '+followup
    :'✓ '+(OI[selOut]||selOut)+' logged for '+cur.name.slice(0,20);
  toast(msg);
  if(tab==='today')renderBriefing();else if(tab==='all')rA();else renderKPIs();
}
function skip(id){
  if(!log[id])log[id]=[];
  log[id].push({outcome:'no_answer',date:new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),notes:'Skipped'});
  lSave();rT();
}
function unskip(id){
  if(!log[id])return;
  // Remove last entry if it was a skip
  const last=log[id][log[id].length-1];
  if(last&&last.notes==='Skipped')log[id].pop();
  if(!log[id].length)delete log[id];
  lSave();if(tab==='today')rT();else if(tab==='all')rA();
  toast('Removed from skipped');
}

// PIPELINE
function statBox(label,val,unit,col){
  return '<div style="background:var(--surf);border:1px solid var(--brd);border-radius:8px;padding:8px;text-align:center">'
    +'<div style="font-size:20px;font-weight:800;color:'+col+'">'+val+'</div>'
    +'<div style="font-size:8px;color:var(--sub)">'+unit+'</div>'
    +'<div style="font-size:9px;font-weight:600;color:var(--navy)">'+label+'</div>'
    +'</div>';
}
function funnelRow(label,val,total,col){
  const pct=total>0?Math.round(val/total*100):0;
  const w=Math.max(4,pct);
  return '<div style="margin-bottom:5px">'
    +'<div style="display:flex;justify-content:space-between;font-size:10px;margin-bottom:2px">'
      +'<span style="font-weight:600;color:var(--navy)">'+label+'</span>'
      +'<span style="color:var(--sub)">'+val+' ('+pct+'%)</span>'
    +'</div>'
    +'<div style="height:6px;background:var(--brd2);border-radius:3px">'
      +'<div style="height:6px;width:'+w+'%;background:'+col+';border-radius:3px;transition:.3s"></div>'
    +'</div>'
    +'</div>';
}
function outcomeBreakdown(entries){
  const counts={};
  entries.forEach(e=>{counts[e.outcome]=(counts[e.outcome]||0)+1;});
  return Object.entries(counts).sort((a,b)=>b[1]-a[1])
    .map(([k,v])=>'<div style="display:flex;justify-content:space-between;font-size:10px;padding:2px 0">'
      +'<span style="color:var(--sub)">'+( OI[k]||k)+'</span>'
      +'<span style="font-weight:600;color:var(--navy)">'+v+'</span>'
      +'</div>'
    ).join('');
}
function rPipe(){
  const now=new Date();
  const weekAgo=new Date(now-7*864e5);
  const monthAgo=new Date(now-30*864e5);

  const allEntries=[];
  Object.entries(log).forEach(([id,entries])=>{
    entries.forEach(e=>allEntries.push({...e,id:parseInt(id)}));
  });
  function parseDate(s){
    if(!s)return null;
    const d=new Date(s);
    return isNaN(d)?null:d;
  }
  function inWindow(e,since){
    const d=parseDate(e.date);
    return d&&d>=since;
  }

  const weekEntries=allEntries.filter(e=>inWindow(e,weekAgo));
  const monthEntries=allEntries.filter(e=>inWindow(e,monthAgo));

  const totalContacted=Object.keys(log).filter(id=>log[id].length>0).length;
  const totalInterested=Object.values(log).filter(entries=>entries.some(e=>e.outcome==='interested'||e.outcome==='scheduled')).length;
  const totalScheduled=Object.values(log).filter(entries=>entries.some(e=>e.outcome==='scheduled')).length;
  const totalWon=Object.values(log).filter(entries=>entries.some(e=>['customer_recurring','customer_once'].includes(e.outcome))).length;

  const vendorCounts={};
  Object.values(vendors).forEach(v=>{
    const key=(v||'').toLowerCase().trim();
    if(key)vendorCounts[key]=(vendorCounts[key]||0)+1;
  });
  const topVendors=Object.entries(vendorCounts).sort((a,b)=>b[1]-a[1]).slice(0,4);

  const c=document.getElementById('pipec');

  c.innerHTML=
    '<div style="margin-bottom:16px">'
    +'<div style="font-weight:800;font-size:13px;color:var(--navy);margin-bottom:10px">&#x1F4CA; Activity Dashboard</div>'
    +'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:10px">'
      +statBox('This Week',weekEntries.length,'calls','var(--blu)')
      +statBox('This Month',monthEntries.length,'calls','var(--navy)')
      +statBox('Contacted',totalContacted,'total','var(--sub)')
    +'</div>'
    +'<div style="background:var(--surf);border:1px solid var(--brd);border-radius:9px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Conversion Funnel</div>'
      +funnelRow('Contacted',totalContacted,totalContacted,'#64748b')
      +funnelRow('Interested',totalInterested,totalContacted,'var(--blu)')
      +funnelRow('Scheduled',totalScheduled,totalContacted,'var(--ora)')
      +funnelRow('Closed Won',totalWon,totalContacted,'#059669')
    +'</div>'
    +(weekEntries.length?
      '<div style="background:var(--surf);border:1px solid var(--brd);border-radius:9px;padding:10px;margin-bottom:10px">'
        +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">This Week by Outcome</div>'
        +outcomeBreakdown(weekEntries)
      +'</div>'
    :'')
    +(topVendors.length>=2?
      '<div style="background:#fef9ee;border:1px solid #fde68a;border-radius:9px;padding:10px;margin-bottom:10px">'
        +'<div style="font-size:9px;font-weight:700;color:#92400e;text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">&#x1F575; Competitor Intel</div>'
        +topVendors.map(([v,n])=>'<div style="display:flex;justify-content:space-between;font-size:10px;padding:3px 0;border-bottom:1px solid #fde68a"><span style="color:#78350f;font-weight:600;text-transform:capitalize">'+v+'</span><span style="color:#92400e">'+n+'x</span></div>').join('')
        +'<div style="font-size:9px;color:#92400e;margin-top:4px">Track who you are displacing. Log vendors on each business card.</div>'
      +'</div>'
    :'')
    +'</div>'
    +'<div style="font-weight:800;font-size:13px;color:var(--navy);margin-bottom:10px">&#x1F4CB; Pipeline by Status</div>';

  const grps={
    scheduled:{l:'Scheduled / Booked',c:'#32d74b',ids:[]},
    interested:{l:'Interested',c:'#0a84ff',ids:[]},
    follow_up:{l:'Follow Up Needed',c:'#ffd60a',ids:[]},
    voicemail:{l:'Voicemail Left',c:'#5e9eff',ids:[]},
    no_answer:{l:'No Answer',c:'#445066',ids:[]},
    not_interested:{l:'Not Interested',c:'#ff3b30',ids:[]},
    service_done:{l:'Service Done',c:'#059669',ids:[]},
  };
  Object.entries(log).forEach(([id,entries])=>{if(!entries.length)return;const l=entries[entries.length-1];if(grps[l.outcome])grps[l.outcome].ids.push(parseInt(id));});
  const pm={};P.forEach(p=>pm[p.id]=p);
  const hasAny=Object.values(grps).some(g=>g.ids.length>0);
  if(!hasAny){c.innerHTML+='<div class="tempty"><div class="ei">&#x1F4DE;</div><div>Start logging calls to build your pipeline</div></div>';return;}
  c.innerHTML+=Object.entries(grps).map(([k,g])=>{
    if(!g.ids.length)return'';
    const items=g.ids.map(id=>{
      const p=pm[id];if(!p)return'';
      const last=(log[id]||[]).slice(-1)[0];
      return '<div class="pitem" onclick="openM('+id+')">'
        +'<div class="pdot" style="background:'+g.c+'"></div>'
        +'<div style="flex:1">'
          +'<div class="piname">'+p.name+'</div>'
          +'<div style="font-size:9px;color:var(--sub)">'+p.city+', '+p.county+' &bull; '+dL(p.days_until)+' &bull; '+last.date+'</div>'
          +(p.phone?('<div class="piph">'+p.phone+'</div>'):'')
          +(last.notes?('<div style="font-size:9px;color:var(--sub);margin-top:1px">'+last.notes.slice(0,50)+'</div>'):'')
          +(last.followup?('<div style="font-size:9px;font-weight:600;color:var(--ora)">&#x1F4C5; Follow-up: '+last.followup+'</div>'):'')
        +'</div>'
        +'<span class="pbadge '+p.priority+'">'+p.priority+'</span>'
        +'</div>';
    }).join('');
    return '<div class="psect"><div class="pst" style="color:'+g.c+'">'+g.l+'<span class="pct">'+g.ids.length+'</span></div>'+items+'</div>';
  }).join('');
}

// ADD PHONE
function addPhone(){
  const id=parseInt(document.getElementById('ph-id').value.trim());
  const phone=document.getElementById('ph-num').value.trim();
  const hours=document.getElementById('ph-hrs').value.trim();
  if(!id||!phone){toast('Enter a License ID and phone number');return;}
  phSave(id,phone,hours,0);
  document.getElementById('ph-id').value='';
  document.getElementById('ph-num').value='';
  document.getElementById('ph-hrs').value='';
  toast('✓ Phone saved for #'+id);
  // Re-render current view lightly
  if(tab==='all')rA();
  else if(tab==='today')renderKPIs();
}

// EXPORT
function exportCSV(){
  const rows=[['Name','County','City','Phone','Hours','Rating','Priority','Days Until','Pred Next','High Viol','Total Viol','Ice Profile','Codes','Trend','Last Outcome','Date','Notes']];
  P.forEach(p=>{
    const l=getLC(p.id);
    rows.push([p.name,p.county,p.city,p.phone||'',p.hours||'',p.rating||'',p.priority,
      p.days_until,p.pred_next,p.high_viol,p.total_viol,
      p.chronic?'CHRONIC':p.confirmed?'confirmed':'none',
      (p.codes||[]).join(';'),p.trending?'worse':'stable',
      l?OI[l.outcome]||l.outcome:'',l?l.date:'',l?l.notes:'']);
  });
  const csv=rows.map(r=>r.map(c=>'"'+String(c).replace(/"/g,'""')+'"').join(',')).join('\\n');
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
  a.download='pic_prospects_'+new Date().toISOString().slice(0,10)+'.csv';
  a.click();toast('CSV exported');
}
function clrLog(){
  if(!confirm('Clear all call log data? Cannot be undone.'))return;
  log={};lSave();
  if(tab==='today'){rT();renderBriefing();}
  toast('Call log cleared');
}
function clrCustomers(){
  if(!confirm('Clear all customer data? Cannot be undone.'))return;
  customers={};custSave();
  P.forEach(p=>{if(p.status!=='prospect')p.status='prospect';});
  if(tab==='customers')rCust();
  renderBriefing();
  toast('Customer data cleared');
}
function clrAll(){
  if(!confirm('Reset ALL data? Clears log, customers, contacts, vendors, goals. Cannot be undone.'))return;
  log={};lSave();
  customers={};custSave();
  contacts={};contactsSave();
  vendors={};vendorsSave();
  P.forEach(p=>p.status='prospect');
  if(tab==='today'){rT();renderBriefing();}
  
  else if(tab==='customers')rCust();
  else if(tab==='service')rService();
  else{rT();renderBriefing();}
  toast('All data cleared. Reloading...');
  setTimeout(()=>window.location.reload(),1200);
}
function toast(msg){const t=document.getElementById('toast');t.textContent=msg;t.classList.add('on');setTimeout(()=>t.classList.remove('on'),2200);}

// ── CUSTOMER LIFECYCLE ────────────────────────────────────────────────────────
let customers = {};  // {id: {status, won_date, service_type, notes, last_service}}

function custLoad(){
  try{customers=JSON.parse(localStorage.getItem('pic_customers')||'{}')||{};}catch(e){customers={};}
  // Apply saved statuses to P records
  P.forEach(p=>{const c=customers[p.id];if(c)p.status=c.status;});
}
function custSave(){try{localStorage.setItem('pic_customers',JSON.stringify(customers));}catch(e){}}

function markWon(status){
  if(!cur)return;
  const p=cur;
  const now=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  customers[p.id]={
    status,
    won_date: now,
    service_type: status==='customer_recurring'?'recurring':status==='customer_intro'?'intro':'one_time',
    monthly: status==='customer_recurring'?p.monthly:0,
    onetime: status==='customer_once'?p.onetime:status==='customer_intro'?99:0,
    machines: p.machines,
    name: p.name,
    address: p.address,
    city: p.city,
    phone: p.phone,
    notes: '',
    last_service: '',
    next_service: '',
    hubspot_url: '',
    square_url: '',
    machine_brand: '',
    machine_model: '',
    machine_type: '',
    filter_type: '',
    filter_installed: '',
    contract_start: '',
    contract_term: 6,
    contract_renewal: '',
    service_history: [],
    atp_history: [],
  };
  p.status=status;
  custSave();
  // Log it
  if(!log[p.id])log[p.id]=[];
  log[p.id].push({outcome:status,date:now,notes:'Deal closed'});
  lSave();
  if(status==='quoted'){
    toast('Marked as Quoted — create the quote in HubSpot.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  } else if(status==='customer_intro'){
    buildAnnualSchedule(p.id);
    toast('Intro offer booked! $99 first visit. Follow up to convert to recurring.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  } else {
    if(status==='customer_recurring')buildAnnualSchedule(p.id);
    toast(status==='customer_recurring'?'Won! Added to recurring customers.':'Won! One-time service recorded.');
    document.getElementById('mbg').classList.remove('on');
    sw('customers');
  }
}

function rCust(){
  const filter=document.getElementById('cust-status').value;
  const allCusts=P.filter(p=>p.status&&p.status!=='prospect');
  const shown=filter?allCusts.filter(p=>p.status===filter):allCusts;

  // MRR calculation
  const recurring=allCusts.filter(p=>p.status==='customer_recurring');
  const mrr=recurring.reduce((s,p)=>s+(p.monthly||149),0);
  document.getElementById('mrr-val').textContent='$'+mrr.toLocaleString();
  document.getElementById('cust-count').textContent=recurring.length;
  document.getElementById('arr-val').textContent='$'+(mrr*12).toLocaleString();

  const el=document.getElementById('cust-list');
  const em=document.getElementById('cust-empty');
  document.getElementById('cust-cnt').textContent=shown.length+' accounts';

  if(!shown.length){el.innerHTML='';em.style.display='block';return;}
  em.style.display='none';

  const STATUS_LABELS={
    customer_recurring:'&#x1F504; Recurring',customer_once:'&#x1F9FC; One-Time',
    customer_intro:'&#x1F525; Intro ($99)',
    quoted:'&#x1F4C4; Quote Sent',churned:'&#x274C; Churned'
  };
  const STATUS_COLORS={
    customer_recurring:'#059669',customer_once:'var(--blu)',
    customer_intro:'var(--ora)',
    quoted:'#7c3aed',churned:'var(--cb)'
  };

  // Portal ID for HubSpot links
  const portalId=loadSettings().hubspot_portal||'';

  const today=new Date();

  el.innerHTML=shown.map(p=>{
    const c=customers[p.id]||{};
    const col=STATUS_COLORS[p.status]||'var(--sub)';
    const lbl=STATUS_LABELS[p.status]||p.status;
    const rev=p.status==='customer_recurring'?('$'+p.monthly+'/mo'):p.status==='customer_once'?('$'+p.onetime+' one-time'):'';

    // Service due indicator
    let serviceDueH='';
    if(c.next_service){
      const due=new Date(c.next_service);
      const daysUntil=Math.round((due-today)/864e5);
      const dueCol=daysUntil<0?'#dc2626':daysUntil<=7?'#d97706':'#059669';
      const dueTxt=daysUntil<0?Math.abs(daysUntil)+'d overdue':daysUntil===0?'Due TODAY':daysUntil+'d until service';
      serviceDueH='<div style="font-size:9px;font-weight:700;color:'+dueCol+'">&#x1F4C5; '+dueTxt+'</div>';
    }

    // HubSpot link
    const hsUrl=portalId
      ?('https://app.hubspot.com/contacts/'+portalId+'/company/create?name='+enc(p.name)+'&city='+enc(p.city))
      :'https://app.hubspot.com';

    return '<div class="pitem" data-id="'+p.id+'" style="border-left:3px solid '+col+';flex-direction:column;align-items:stretch;gap:8px">'
      // Top row
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+', '+p.county+(p.phone?' &bull; '+p.phone:'')+'</div>'
          +'<div style="font-size:9px;font-weight:700;color:'+col+';margin-top:2px">'+lbl+'</div>'
        +'</div>'
        +'<div style="text-align:right;flex-shrink:0;margin-left:10px">'
          +(rev?'<div style="font-size:13px;font-weight:800;color:var(--grn)">'+rev+'</div>':'')
          +(c.won_date?'<div style="font-size:9px;color:var(--sub)">Since '+c.won_date+'</div>':'')
        +'</div>'
      +'</div>'
      // Service tracking row
      +'<div style="background:#f5f8fa;border-radius:7px;padding:8px;display:flex;flex-direction:column;gap:5px">'
        +'<div style="display:flex;justify-content:space-between;align-items:center">'
          +'<span style="font-size:9px;color:var(--sub);font-weight:600;text-transform:uppercase">Last Service</span>'
          +'<span style="font-size:10px;font-weight:600;color:var(--navy)">'+(c.last_service||'Not recorded')+'</span>'
        +'</div>'
        +serviceDueH
        +'<div style="display:flex;gap:5px;margin-top:3px">'
          +'<button onclick="event.stopPropagation();openServiceLog('+p.id+')" ontouchend="event.stopPropagation();event.preventDefault();openServiceLog('+p.id+')" ontouchend="event.stopPropagation();event.preventDefault();openServiceLog('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:none;border-radius:6px;background:var(--grn);color:#fff;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Log Visit</button>'
          +'<button onclick="event.stopPropagation();setNextService('+p.id+')" ontouchend="event.stopPropagation();event.preventDefault();setNextService('+p.id+')" ontouchend="event.stopPropagation();event.preventDefault();setNextService('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--sub);cursor:pointer;font-family:inherit">Set Next Due</button>'
        +'</div>'
      +'</div>'
      // Action buttons
      +'<div style="display:flex;gap:5px">'
        +'<button onclick="event.stopPropagation();openM('+p.id+')" style="flex:1;font-size:9px;padding:5px;border:1px solid var(--brd);border-radius:6px;background:var(--surf);color:var(--navy);cursor:pointer;font-family:inherit">Details</button>'
        +(c.hubspot_url
          ?'<a href="'+c.hubspot_url+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #ff7a59;border-radius:6px;background:#fff7f5;color:#ff7a59;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">HubSpot &#x2197;</a>'
          :'<a href="'+hsUrl+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #ff7a59;border-radius:6px;background:#fff7f5;color:#ff7a59;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">+ HubSpot</a>')
        +(c.square_url
          ?'<a href="'+c.square_url+'" target="_blank" onclick="event.stopPropagation()" style="flex:1;font-size:9px;padding:5px;border:1px solid #00c058;border-radius:6px;background:#f0fff4;color:#00c058;cursor:pointer;font-family:inherit;text-decoration:none;text-align:center;display:flex;align-items:center;justify-content:center">Square &#x2197;</a>'
          :'')
      +'</div>'
      // URL link fields
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x1F517; Link Records</div>'
        +'<div style="display:flex;gap:5px;align-items:center">'
          +'<span style="font-size:9px;color:#ff7a59;font-weight:600;width:52px;flex-shrink:0">HubSpot</span>'
          +'<input type="url" placeholder="Paste HubSpot company URL..." value="'+(c.hubspot_url||'')+'"'
            +' onchange="saveHsUrl('+p.id+',this.value)"'
            +' onclick="event.stopPropagation()"'
            +' style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
        +'</div>'
        +'<div style="display:flex;gap:5px;align-items:center">'
          +'<span style="font-size:9px;color:#00c058;font-weight:600;width:52px;flex-shrink:0">Square</span>'
          +'<input type="url" placeholder="Paste Square payment link URL..." value="'+(c.square_url||'')+'"'
            +' onchange="saveSqUrl('+p.id+',this.value)"'
            +' onclick="event.stopPropagation()"'
            +' style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
        +'</div>'
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Type</div>'
            +'<select onchange="saveMachineType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_TYPE_LIST.map(t=>'<option value="'+t+'"'+(t===(c.machine_type||'')?'selected':'')+'>'+t+'</option>').join('')
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

function saveFoundPhone(id){
  const inp=document.getElementById('ph-inp-'+id);
  if(!inp)return;
  const v=inp.value.trim();
  if(!v){toast('Enter a phone number first');return;}
  const digits=v.replace(/[^\d+\-().\s]/g,'').trim();
  // Save to localStorage and update P array
  phSave(id, digits, '', 0);
  // Also update p.phone directly so cardHTML reads it correctly on re-render
  const pEntry=P.find(x=>x.id===id);
  if(pEntry)pEntry.phone=digits;
  // Update card display without full re-render
  const card=document.querySelector('[data-id="'+id+'"]');
  if(card){
    const row=card.querySelector('.phrow');
    if(row)row.innerHTML='<a href="tel:'+digits.replace(/\s/g,'')+'" class="phnum" onclick="event.stopPropagation()">'+digits+'</a>';
    const saveRow=document.getElementById('ph-save-'+id);
    if(saveRow)saveRow.style.display='none';
  }
  toast('\u2713 Phone saved: '+digits);
}
function saveHsUrl(id,v){saveCustomerUrl(id,'hubspot_url',v);}
function saveMachineBrand(id,v){saveMachineField(id,'machine_brand',v);}
function saveMachineType(id,v){saveMachineField(id,'machine_type',v);}
function saveFilterType(id,v){saveMachineField(id,'filter_type',v);}
function saveContractStart(id,v){saveContractField(id,'contract_start',v);}
function saveContractTerm(id,v){saveContractField(id,'contract_term',v);}
function saveMachineField(id,field,value){
  if(!customers[id])customers[id]={};
  customers[id][field]=value;
  custSave();
}
function saveContractField(id,field,value){
  if(!customers[id])customers[id]={};
  customers[id][field]=value;
  // Auto-calculate renewal date
  if(field==='contract_start'||field==='contract_term'){
    const start=new Date(customers[id].contract_start||value);
    const term=parseInt(customers[id].contract_term||6);
    if(!isNaN(start)){
      const renewal=new Date(start);
      renewal.setMonth(renewal.getMonth()+term);
      customers[id].contract_renewal=renewal.toISOString().slice(0,10);
    }
  }
  custSave();
  rCust();
}
function saveSqUrl(id,v){saveCustomerUrl(id,'square_url',v);}
function saveCustomerUrl(id, field, value){
  if(!customers[id])customers[id]={};
  customers[id][field]=value.trim();
  custSave();
  // Update button if URL was set
  const label=field==='hubspot_url'?'HubSpot':'Square';
  toast(label+' link saved for this client');
}

function logService(id){
  const today=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  if(!customers[id])customers[id]={};
  customers[id].last_service=today;
  // Default next service in 30 days for recurring
  const p=P.find(x=>x.id===id);
  if(p&&p.status==='customer_recurring'){
    const next=new Date();next.setDate(next.getDate()+60);
    customers[id].next_service=next.toISOString().slice(0,10);
  }
  custSave();rCust();
  toast('Service visit logged for '+today);
}

function setNextService(id){
  // iOS-safe date picker overlay
  const existing=document.getElementById('sns-bg');
  if(existing)existing.remove();
  const defaultDate=new Date(Date.now()+60*864e5).toISOString().slice(0,10);
  const bg=document.createElement('div');
  bg.id='sns-bg';
  bg.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:600;display:flex;align-items:center;justify-content:center';
  bg.innerHTML='<div style="background:#fff;border-radius:16px;padding:20px;width:90%;max-width:320px">'
    +'<div style="font-size:13px;font-weight:700;color:#0f1f38;margin-bottom:12px">Set Next Service Date</div>'
    +'<input id="sns-date" type="date" value="'+defaultDate+'" style="width:100%;padding:10px;border:1px solid #e2e8f0;border-radius:8px;font-size:14px;font-family:inherit;outline:none;box-sizing:border-box;margin-bottom:12px">'
    +'<div style="display:flex;gap:8px">'
    +'<button id="sns-save" style="flex:1;padding:10px;background:#059669;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">Set Date</button>'
    +'<button id="sns-cancel" style="padding:10px 14px;background:#f1f5f9;color:#475569;border:none;border-radius:8px;font-size:13px;cursor:pointer;font-family:inherit;touch-action:manipulation">Cancel</button>'
    +'</div></div>';
  document.body.appendChild(bg);
  function doSNS(e){
    if(e&&e.type==='touchend')e.preventDefault();
    const d=(document.getElementById('sns-date')||{}).value||'';
    if(!d){toast('Pick a date');return;}
    if(!customers[id])customers[id]={};
    customers[id].next_service=d;
    custSave();bg.remove();rCust();
    toast('Next service: '+d);
  }
  const saveBtn=document.getElementById('sns-save');
  saveBtn.addEventListener('touchend',doSNS,false);
  saveBtn.addEventListener('click',doSNS,false);
  const cancelBtn=document.getElementById('sns-cancel');
  function doCancel(e){if(e&&e.type==='touchend')e.preventDefault();bg.remove();}
  cancelBtn.addEventListener('touchend',doCancel,false);
  cancelBtn.addEventListener('click',doCancel,false);
  bg.addEventListener('touchend',function(e){if(e.target===bg){e.preventDefault();bg.remove();}},false);
}

// ── GOALS & MORNING BRIEFING ─────────────────────────────────────────────────
let goals = {
  clients:  10,
  deadline: '2026-07-01',
  mrr:      1490,
  calls:    5,
  walkins:  2,
  quotes:   3,
  closes:   1,
};

function goalsLoad(){
  try{
    const saved=JSON.parse(localStorage.getItem('pic_goals')||'null');
    if(saved)goals={...goals,...saved};
  }catch(e){}
}
function goalsSave(){
  try{localStorage.setItem('pic_goals',JSON.stringify(goals));}catch(e){}
}
function autoMRR(){
  const clients=parseInt(document.getElementById('goal-clients')?.value)||0;
  const mrrEl=document.getElementById('goal-mrr');
  if(mrrEl&&clients>0)mrrEl.value=clients*149;
}

function saveGoals(){
  const g=id=>{const el=document.getElementById(id);return el?(parseFloat(el.value)||null):null;};
  const gs=id=>{const el=document.getElementById(id);return el?el.value||'':'';};
  // Read all values
  var c=g('goal-clients');        if(c!==null)goals.clients=c;
  var d=gs('goal-deadline');      if(d)goals.deadline=d;
  var m=g('goal-mrr');             if(m!==null)goals.mrr=m;
  // Auto-compute required calls/walkins from revenue goal + deadline
  if(goals.deadline&&goals.mrr){
    const avgMRR=149;
    const closeRate=0.05;      // 5% close rate on calls (conservative)
    const walkCloseRate=0.10;  // walk-ins 2x better
    const clientsNeeded=Math.ceil(goals.mrr/avgMRR);
    if(!c)goals.clients=clientsNeeded;
    const [dy,dm,dd]=goals.deadline.split('-').map(Number);
    const deadlineD=new Date(dy,dm-1,dd,12,0,0);
    const daysLeft=Math.max(1,Math.ceil((deadlineD-new Date())/864e5));
    const weeksLeft=Math.max(1,daysLeft/7);
    const currentClients=P.filter(p=>p.status==='customer_recurring').length;
    const clientsToGo=Math.max(0,clientsNeeded-currentClients);
    const callsPerDay=Math.max(1,Math.ceil((clientsToGo/closeRate/weeksLeft)/5));
    const walksPerDay=Math.max(0,Math.ceil((clientsToGo/walkCloseRate/weeksLeft)/5));
    goals.calls=callsPerDay;
    goals.walkins=walksPerDay;
    const callEl=document.getElementById('goal-calls');    if(callEl)callEl.value=goals.calls;
    const walkEl=document.getElementById('goal-walkins');  if(walkEl)walkEl.value=goals.walkins;
    const clEl=document.getElementById('goal-clients');    if(clEl)clEl.value=goals.clients;
  } else {
    var ca=g('goal-calls');    if(ca!==null)goals.calls=ca;
    var w=g('goal-walkins');   if(w!==null)goals.walkins=w;
  }
  goalsSave();
  renderBriefing();
  toast('\u2713 Goals saved \u2014 daily activity recalculated');
}
function initGoals(){
  goalsLoad();
  const set=(id,v)=>{const el=document.getElementById(id);if(el&&v!==undefined)el.value=v;};
  set('goal-clients',  goals.clients);
  set('goal-deadline', goals.deadline);
  set('goal-mrr',      goals.mrr);
  set('goal-calls',    goals.calls);
  set('goal-walkins',  goals.walkins);
}

function renderBriefing(){
  const now=new Date();
  const todayStr=now.toISOString().slice(0,10);
  const weekAgo=new Date(now-7*864e5);
  const dayName=now.toLocaleDateString('en-US',{weekday:'long',month:'long',day:'numeric'});

  // ── All log entries ───────────────────────────────────────────────────
  const allEntries=[];
  Object.entries(log).forEach(([id,entries])=>{
    entries.forEach(e=>allEntries.push({...e,pid:parseInt(id)}));
  });

  function parseD(s){if(!s)return null;const d=new Date(s);return isNaN(d)?null:d;}
  const weekEntries=allEntries.filter(e=>{ const d=parseD(e.date);return d&&d>=weekAgo;});

  // ── KPI ROW ───────────────────────────────────────────────────────────
  const recurring=P.filter(p=>{
    const lc=getLC(p.id);return lc&&['signed','customer_recurring','customer_once'].includes(lc.outcome);
  });
  const mrr=recurring.reduce((s,p)=>s+(p.monthly||149),0);
  const clientCount=recurring.length;

  // Pipeline = intro_set + in_play prospects
  const pipeCount=Object.keys(log).filter(id=>{
    const lc=getLC(parseInt(id));
    if(!lc)return false;
    const n=normO(lc.outcome);
    return n==='intro_set'||n==='in_play';
  }).length;

  const weekTotal=weekEntries.filter(e=>e.outcome!=='service_done').length;

  const kpiEl=id=>document.getElementById(id);
  if(kpiEl('kpi-mrr'))kpiEl('kpi-mrr').textContent='$'+mrr.toLocaleString();
  if(kpiEl('kpi-clients'))kpiEl('kpi-clients').textContent=clientCount;
  if(kpiEl('kpi-pipe'))kpiEl('kpi-pipe').textContent=pipeCount;
  if(kpiEl('kpi-week'))kpiEl('kpi-week').textContent=weekTotal;

  // ── FUNNEL (this week) ────────────────────────────────────────────────
  const wWalkins =weekEntries.filter(e=>e.type==='walkin').length;
  const wCalls   =weekEntries.filter(e=>e.type==='call').length;
  const wContacts=weekEntries.filter(e=>e.outcome!=='service_done').length;
  const wIntros  =weekEntries.filter(e=>normO(e.outcome)==='intro_set').length;
  const wSigned  =weekEntries.filter(e=>normO(e.outcome)==='signed').length;
  const wNotNow  =weekEntries.filter(e=>normO(e.outcome)==='not_now').length;

  const fEl=document.getElementById('funnel-stages');
  const fLabel=document.getElementById('funnel-week-label');
  if(fLabel)fLabel.textContent=wCalls+'c / '+wWalkins+'w this week';
  if(fEl){
    const maxV=Math.max(wContacts,1);
    const fStage=(label,val,color,sub)=>{
      const pct=Math.round(val/maxV*100);
      return '<div class="funnel-stage">'
        +'<div class="funnel-label">'+label+'</div>'
        +'<div class="funnel-bar-wrap"><div class="funnel-bar" style="width:'+pct+'%;background:'+color+'"></div></div>'
        +'<div class="funnel-val">'+val+(sub?'<span style="font-size:9px;color:#aaa"> '+sub+'</span>':'')+'</div>'
        +'</div>';
    };
    fEl.innerHTML=
      fStage('Contacts',wContacts,'#94a3b8')
      +fStage('Intro Set',wIntros,'#0891b2',wContacts?Math.round(wIntros/wContacts*100)+'%':'')
      +fStage('Signed',wSigned,'#059669',wContacts?Math.round(wSigned/wContacts*100)+'%':'')
      +fStage('Not Now',wNotNow,'#dc2626',wContacts?Math.round(wNotNow/wContacts*100)+'%':'');
  }

  // ── GOAL PACING ───────────────────────────────────────────────────────
  const gEl=document.getElementById('goal-pacing-content');
  if(gEl){
    const goalClients=goals.clients||10;
    const deadline=goals.deadline?new Date(goals.deadline):null;
    const daysLeft=deadline?Math.ceil((deadline-now)/864e5):null;
    const weeksLeft=daysLeft?Math.max(1,daysLeft/7):null;
    const needed=Math.max(0,goalClients-clientCount);
    const perWeek=weeksLeft?Math.ceil(needed/weeksLeft):null;
    // Current pace (closes last 4 weeks)
    const month4wk=new Date(now-28*864e5);
    const closedLast4=allEntries.filter(e=>{const d=parseD(e.date);return d&&d>=month4wk&&normO(e.outcome)==='signed';}).length;
    const closePace=Math.round(closedLast4/4*10)/10; // closes per week

    let paceHTML='';
    if(deadline){
      const onTrack=closePace>=perWeek;
      paceHTML+='<div style="display:grid;grid-template-columns:repeat(2,1fr);gap:6px;margin-bottom:8px">'
        +'<div style="text-align:center;background:#f5f8fa;border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:var(--grn)">$'+mrr.toLocaleString()+'</div>'
          +'<div style="font-size:8px;color:var(--sub)">MRR now</div></div>'
        +'<div style="text-align:center;background:#f5f8fa;border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:var(--navy)">'+clientCount+'/'+goalClients+'</div>'
          +'<div style="font-size:8px;color:var(--sub)">clients</div></div>'
        +'<div style="text-align:center;background:#f5f8fa;border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:var(--navy)">'+(daysLeft!==null?daysLeft:'-')+'</div>'
          +'<div style="font-size:8px;color:var(--sub)">days left</div></div>'
        +'<div style="text-align:center;background:'+(onTrack?'#ecfdf5':'#fef2f2')+';border-radius:7px;padding:8px">'
          +'<div style="font-size:18px;font-weight:800;color:'+(onTrack?'#059669':'#dc2626')+'">'+perWeek+'/wk</div>'
          +'<div style="font-size:8px;color:var(--sub)">needed</div></div>'
        +'</div>';
      paceHTML+='<div style="font-size:10px;color:var(--sub)">Current pace: '
        +(closePace>0?closePace+' closes/week':'No closes logged yet')
        +' &nbsp;&#x2022;&nbsp; '
        +(onTrack?'<span class="pace-on-track">On track ✓</span>':'<span class="pace-behind">Behind pace</span>')+'</div>';
    } else {
      paceHTML='<div style="font-size:11px;color:var(--sub)">Set a goal and deadline in <b>Data</b> to see pacing.</div>';
    }
    gEl.innerHTML=paceHTML;
  }

  // ── LOSS REASON BREAKDOWN ─────────────────────────────────────────────
  const lbEl=document.getElementById('loss-breakdown');
  const leEl=document.getElementById('loss-empty');
  if(lbEl){
    // Count not_now entries with reasons (last 90 days)
    const d90=new Date(now-90*864e5);
    const reasonCounts={};
    allEntries.forEach(e=>{
      const d=parseD(e.date);
      if(!d||d<d90)return;
      if(normO(e.outcome)!=='not_now')return;
      const r=e.reason||'no_reason';
      reasonCounts[r]=(reasonCounts[r]||0)+1;
    });
    const sorted=Object.entries(reasonCounts).sort((a,b)=>b[1]-a[1]);
    if(!sorted.length){
      if(lbEl)lbEl.innerHTML='';
      if(leEl)leEl.style.display='block';
    } else {
      if(leEl)leEl.style.display='none';
      const maxR=sorted[0][1];
      lbEl.innerHTML=sorted.slice(0,5).map(([r,n])=>{
        const pct=Math.round(n/maxR*100);
        const label=REASONS[r]||r;
        return '<div style="display:flex;align-items:center;gap:8px">'
          +'<div style="font-size:10px;color:var(--txt);min-width:130px">'+label+'</div>'
          +'<div style="flex:1;height:5px;background:var(--brd2);border-radius:3px">'
            +'<div style="width:'+pct+'%;height:5px;background:#dc2626;border-radius:3px"></div></div>'
          +'<div style="font-size:10px;color:var(--sub);min-width:20px;text-align:right">'+n+'</div>'
          +'</div>';
      }).join('');
    }
  }

  // ── ACTIVE NURTURE & COLD TARGETS (only when on Home tab) ──────────
  if(tab!=='today')return; // KPI-only update handled by renderKPIs
  // Defer heavy card rendering to next animation frame
  requestAnimationFrame(()=>{
  const today2=now.toISOString().slice(0,10);
  // All prospects with a follow-up date set (any outcome), not yet clients
  const today_iso=now.toISOString().slice(0,10);
  const in7=new Date(now.getTime()+7*864e5).toISOString().slice(0,10);
  const in14=new Date(now.getTime()+14*864e5).toISOString().slice(0,10);

  const allFollowups=P.filter(p=>{
    if(isC(p.id))return false;
    const lc=getLC(p.id);
    return lc&&lc.followup;
  }).sort((a,b)=>{
    const fa=getLC(a.id)?.followup||'9999';
    const fb=getLC(b.id)?.followup||'9999';
    return fa.localeCompare(fb);
  });

  // Split into groups
  const fuOverdue =allFollowups.filter(p=>(getLC(p.id)?.followup||'9999')<today_iso);
  const fuToday   =allFollowups.filter(p=>getLC(p.id)?.followup===today_iso);
  const fuSoon    =allFollowups.filter(p=>{const d=getLC(p.id)?.followup||'';return d>today_iso&&d<=in7;});
  const fuUpcoming=allFollowups.filter(p=>{const d=getLC(p.id)?.followup||'';return d>in7&&d<=in14;});

  // Legacy: nurturePros for backward compat
  const nurturePros=allFollowups.slice(0,6);

  const nGrid=document.getElementById('nurture-grid');
  const nEmpty=document.getElementById('nurture-empty');
  if(nGrid){
    if(!allFollowups.length){nGrid.innerHTML='';if(nEmpty)nEmpty.style.display='block';}
        else{
      if(nEmpty)nEmpty.style.display='none';
      function fuSection(label,color,items){
        if(!items.length)return '';
        return '<div style="font-size:9px;font-weight:700;color:'+color+';text-transform:uppercase;letter-spacing:.06em;padding:6px 0 4px">'+label+' ('+items.length+')</div>'
          +items.map(p=>cardHTML(p)).join('');
      }
      nGrid.innerHTML=
        fuSection('⚠️ Overdue','#dc2626',fuOverdue)
        +fuSection('📅 Today','#059669',fuToday)
        +fuSection('📆 This Week','#d97706',fuSoon)
        +fuSection('🔭 Upcoming','#2563eb',fuUpcoming);
      attachGridListeners(nGrid);
    }
  }

  // ── BEST COLD TARGETS ─────────────────────────────────────────────────
  const coldTargets=P.filter(p=>{
    if(isC(p.id))return false;
    return p.priority==='CALLBACK'||p.priority==='HOT';
  }).sort((a,b)=>(PO[a.priority]??99)-(PO[b.priority]??99)||(b.score||0)-(a.score||0)).slice(0,5);

  const aGrid=document.getElementById('tgrid-actnow');
  const aEmpty=document.getElementById('empty-actnow');
  if(aGrid){
    if(!coldTargets.length){aGrid.innerHTML='';if(aEmpty)aEmpty.style.display='block';}
    else{if(aEmpty)aEmpty.style.display='none';aGrid.innerHTML=coldTargets.map(p=>cardHTML(p)).join('');
    attachGridListeners(aGrid);}

  }
  }); // end requestAnimationFrame
}


function renderKPIs(){
  // Fast-path: just update the 4 KPI numbers without re-rendering cards
  const now=new Date();
  const weekAgo=new Date(now-7*864e5);
  const recurring=P.filter(p=>{const lc=getLC(p.id);return lc&&['signed','customer_recurring','customer_once'].includes(lc.outcome);});
  const mrr=recurring.reduce((s,p)=>s+(p.monthly||149),0);
  const pipeCount=Object.keys(log).filter(id=>{const lc=getLC(parseInt(id));return lc&&['in_play','intro_set','follow_up','interested','scheduled','quoted'].includes(lc.outcome);}).length;
  const weekEntries=[];
  Object.values(log).forEach(entries=>entries.forEach(e=>{const d=new Date(e.date);if(d>=weekAgo)weekEntries.push(e);}));
  const weekTotal=weekEntries.filter(e=>e.outcome!=='service_done').length;
  const el=id=>document.getElementById(id);
  if(el('kpi-mrr'))el('kpi-mrr').textContent='$'+mrr.toLocaleString();
  if(el('kpi-clients'))el('kpi-clients').textContent=recurring.length;
  if(el('kpi-pipe'))el('kpi-pipe').textContent=pipeCount;
  if(el('kpi-week'))el('kpi-week').textContent=weekTotal;
}

function swCustomers(){sw('customers');}

function buildAnnualSchedule(id){
  if(!customers[id])return;
  const c=customers[id];
  // Parse date as LOCAL (not UTC) to avoid off-by-one errors
  const startStr=c.contract_start||new Date().toISOString().slice(0,10);
  const [sy,sm,sd]=startStr.split('-').map(Number);
  const start=new Date(sy,sm-1,sd,12,0,0); // noon to avoid any DST edge
  const term=parseInt(c.contract_term||6);
  const schedule=[];

  // Build visit schedule: every 60 days, starting 60 days from contract start
  let visit=new Date(start.getFullYear(),start.getMonth(),start.getDate()+60,12,0,0);
  const endDate=new Date(start);
  endDate.setMonth(endDate.getMonth()+(term===6?6:12));

  let visitNum=0;
  while(visit<=endDate){
    visitNum++;
    const isDeep=visitNum===1||(visitNum*60)%180<60;
    // Format as local YYYY-MM-DD (avoid UTC shift)
    const iso_local=visit.getFullYear()+'-'+String(visit.getMonth()+1).padStart(2,'0')+'-'+String(visit.getDate()).padStart(2,'0');
    schedule.push({
      date: iso_local,
      date_display: visit.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'}),
      type: isDeep?'deep_clean':'maintenance_60',
      label: isDeep?'Deep Clean':'60-Day Maintenance',
      status: 'scheduled',
    });
    visit=new Date(visit);
    visit.setDate(visit.getDate()+60);
  }

  c.annual_schedule=schedule;
  c.contract_start=c.contract_start||new Date().toISOString().slice(0,10);
  const renewal=new Date(start);
  renewal.setMonth(renewal.getMonth()+term);
  c.contract_renewal=renewal.toISOString().slice(0,10);
  custSave();
}

function exportSchedulePDF(id){
  const p=P.find(x=>x.id===id);
  const c=customers[id]||{};
  if(!p||!c.annual_schedule)return;

  const win=window.open('','_blank');
  const rows=c.annual_schedule.map((s,i)=>
    '<tr style="background:'+(i%2?'#f9f9f9':'#fff')+'">'
    +'<td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">'+s.date_display+'</td>'
    +'<td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">'+(s.type==='deep_clean'?'&#x1F9FC; Full Deep Clean':'&#x1F527; 60-Day Maintenance')+'</td>'
    +'<td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:'+(s.status==='completed'?'#059669':'#94a3b8')+'">'+(s.status==='completed'?'&#x2713; Completed':'Scheduled')+'</td>'
    +'</tr>'
  ).join('');

  win.document.write('<!DOCTYPE html><html><head><title>Service Schedule - '+p.name+'</title>'
    +'<style>body{font-family:system-ui,sans-serif;padding:24px;max-width:600px;margin:0 auto;color:#1e293b}'
    +'table{width:100%;border-collapse:collapse}th{background:#1e3a5f;color:#fff;padding:10px 12px;text-align:left;font-size:12px}'
    +'@media print{button{display:none}}</style></head><body>'
    +'<div style="display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #1e3a5f;padding-bottom:16px;margin-bottom:20px">'
      +'<div><div style="font-size:20px;font-weight:800;color:#1e3a5f">Pinellas Ice Co</div>'
      +'<div style="font-size:11px;color:#64748b">Commercial Ice Machine Cleaning</div></div>'
      +'<div style="text-align:right"><div style="font-size:13px;font-weight:700">SERVICE SCHEDULE</div>'
      +'<div style="font-size:11px;color:#64748b">Generated '+new Date().toLocaleDateString()+'</div></div>'
    +'</div>'
    +'<div style="background:#f1f5f9;border-radius:8px;padding:12px;margin-bottom:16px">'
      +'<div style="font-size:14px;font-weight:700">'+p.name+'</div>'
      +'<div style="font-size:11px;color:#64748b">'+p.address+', '+p.city+', FL</div>'
      +(c.machine_brand?'<div style="font-size:11px;color:#64748b">Machine: '+c.machine_brand+'</div>':'')
      +'<div style="font-size:11px;color:#64748b">Contract: '+c.contract_start+' &rarr; '+c.contract_renewal+'</div>'
    +'</div>'
    +'<table><thead><tr><th>Date</th><th>Service Type</th><th>Status</th></tr></thead>'
    +'<tbody>'+rows+'</tbody></table>'
    +'<div style="margin-top:20px;font-size:10px;color:#94a3b8;border-top:1px solid #e2e8f0;padding-top:12px">'
    +'Pinellas Ice Co &bull; (727) 855-6873 &bull; pinellasiceco.com &bull; '
    +'Service complies with FDA Food Code 3-502.12</div>'
    +'<br><div style="display:flex;gap:8px;margin-top:8px">'+'<button onclick="window.print()" style="padding:10px 20px;background:#1e3a5f;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px">Print / Save PDF</button>'+'<button onclick="window.close()" style="padding:10px 20px;background:#fff;color:#1e3a5f;border:2px solid #1e3a5f;border-radius:6px;cursor:pointer;font-size:12px">&#x2190; Close</button>'+'</div>'
    +'</body></html>');
  win.document.close();
  setTimeout(()=>win.print(),500);
}

function geoClusterSchedule(){
  // Find clients whose next_service dates are within 5 days of each other
  // and are geographically close — suggest scheduling on the same day
  const recurring=P.filter(p=>p.status==='customer_recurring'&&p.lat&&p.lon);
  if(recurring.length<2)return null;

  const clusters=[];
  const today=new Date();

  recurring.forEach(p=>{
    const c=customers[p.id]||{};
    if(!c.next_service)return;
    const due=new Date(c.next_service);
    const daysUntil=Math.round((due-today)/864e5);

    // Find other clients within 5 days and 8 miles
    const nearby=recurring.filter(q=>{
      if(q.id===p.id)return false;
      const cq=customers[q.id]||{};
      if(!cq.next_service)return false;
      const dueQ=new Date(cq.next_service);
      const daysDiff=Math.abs(Math.round((dueQ-due)/864e5));
      const miles=hav(p.lat,p.lon,q.lat,q.lon);
      return daysDiff<=5&&miles<=8;
    });

    if(nearby.length>0){
      // Check if cluster already recorded
      const ids=[p.id,...nearby.map(n=>n.id)].sort().join(',');
      if(!clusters.find(cl=>cl.ids===ids)){
        clusters.push({ids,anchor:p,nearby,daysUntil});
      }
    }
  });

  return clusters.length?clusters:null;
}

function exportDirectoryData(){
  // Export customer data for public directory
  const recurring=P.filter(p=>p.status==='customer_recurring'||p.status==='customer_once'||p.status==='customer_intro');
  const dirData={
    generated: new Date().toISOString(),
    certified: recurring.map(p=>{
      const c=customers[p.id]||{};
      return {
        id: p.id,
        name: p.name,
        address: p.address,
        city: p.city,
        zip: p.zip,
        phone: p.phone||'',
        lat: p.lat,
        lon: p.lon,
        status: p.status,
        service_count: (c.service_history||[]).length,
        last_service: c.last_service||'',
        last_service_iso: c.last_service_iso||'',
        contract_start: c.contract_start||'',
        contract_renewal: c.contract_renewal||'',
        machine_brand: c.machine_brand||'',
        filter_type: c.filter_type||'',
        certified: p.status==='customer_recurring',
        won_date: c.won_date||'',
      };
    }),
    stats:{
      total_certified: recurring.length,
      counties_served: [...new Set(recurring.map(p=>p.county))].length,
      generated: new Date().toLocaleDateString('en-US',{month:'long',year:'numeric'}),
    }
  };

  const blob=new Blob([JSON.stringify(dirData,null,2)],{type:'application/json'});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');
  a.href=url;a.download='customers.json';
  document.body.appendChild(a);a.click();
  document.body.removeChild(a);URL.revokeObjectURL(url);
  toast('customers.json downloaded — upload to GitHub /data/ folder');
}
function setSvcTab(t){
  svcTab=t;
  document.querySelectorAll('.svc-tab').forEach(b=>b.classList.remove('on'));
  const btn=document.getElementById('svct-'+t);
  if(btn)btn.classList.add('on');
  ['cal','route','reports','tutorials','refs'].forEach(s=>{
    const el=document.getElementById('svc-'+s);
    if(el)el.style.display=s===t?'block':'none';
  });
  if(t==='cal')renderServiceCal();
  else if(t==='route')renderServiceRoute();
  else if(t==='reports')renderReports();
  else if(t==='tutorials')renderTutorial();
  else if(t==='refs')renderReferrals();
}

function rService(){
  renderForecast();
  renderAtRisk();
  setSvcTab(svcTab);
}

// ── FORECAST STRIP ────────────────────────────────────────────────────────────
function renderForecast(){
  const el=document.getElementById('svc-forecast');
  if(!el)return;

  const recurring=P.filter(p=>p.status==='customer_recurring');
  const mrr=recurring.reduce((s,p)=>s+(customers[p.id]?.monthly||p.monthly||149),0);
  const arr=mrr*12;

  // Churn assumption: 5% monthly = industry avg for small service biz
  const churnRate=0.05;
  const newPerMonth=mrr/149; // rough: each new client adds ~$149
  const mrr90=Math.round(mrr*Math.pow(1-churnRate,3)+(newPerMonth*149*3));

  // Days to goal
  const goalMrr=goals.mrr||1490;
  const goalDate=new Date(goals.deadline||'2025-07-01');
  const daysLeft=Math.max(0,Math.round((goalDate-new Date())/864e5));

  el.innerHTML=
    statBox('MRR','$'+mrr.toLocaleString(),'monthly recurring','#059669')
    +statBox('ARR','$'+arr.toLocaleString(),'annual run rate','var(--blu)')
    +statBox('90d Forecast','$'+mrr90.toLocaleString(),'at 5% churn','#7c3aed');
}

// ── AT-RISK ALERT ─────────────────────────────────────────────────────────────
function renderAtRisk(){
  const el=document.getElementById('svc-atrisk');
  if(!el)return;

  const today=new Date();
  const atRisk=P.filter(p=>{
    if(p.status!=='customer_recurring')return false;
    const c=customers[p.id]||{};
    if(!c.last_service&&!c.next_service)return true; // never logged
    if(c.next_service){
      const due=new Date(c.next_service);
      const overdueDays=Math.round((today-due)/864e5);
      return overdueDays>5; // 5+ days overdue
    }
    if(c.last_service){
      const last=new Date(c.last_service);
      const daysSince=Math.round((today-last)/864e5);
      return daysSince>35; // 35+ days since last service
    }
    return false;
  });

  if(!atRisk.length){
    el.innerHTML='<div style="background:#f0fdf4;border:1px solid #6ee7b7;border-radius:8px;padding:10px;font-size:10px;color:#059669;font-weight:600">&#x2713; All recurring clients are on schedule</div>';
    return;
  }

  el.innerHTML='<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:10px">'
    +'<div style="font-size:10px;font-weight:700;color:#dc2626;margin-bottom:6px">&#x26A0; '+atRisk.length+' client'+(atRisk.length>1?'s':'')+' at churn risk</div>'
    +atRisk.slice(0,3).map(p=>{
      const c=customers[p.id]||{};
      const msg=!c.last_service&&!c.next_service?'Never serviced since won'
        :c.next_service?Math.round((new Date()-new Date(c.next_service))/864e5)+'d overdue'
        :'35+d since last service';
      return '<div style="display:flex;justify-content:space-between;align-items:center;padding:4px 0;border-top:1px solid #fca5a5;font-size:10px">'
        +'<div><div style="font-weight:600;color:var(--navy)">'+p.name+'</div>'
        +'<div style="color:#dc2626;font-size:9px">'+msg+'</div></div>'
        +'<button onclick="swCustomers()" style="font-size:9px;padding:3px 7px;border:1px solid #dc2626;border-radius:5px;background:#fef2f2;color:#dc2626;cursor:pointer;font-family:inherit">Fix</button>'
        +'</div>';
    }).join('')
    +'</div>';
}

// ── SERVICE CALENDAR ──────────────────────────────────────────────────────────
function renderServiceCal(){
  const el=document.getElementById('svc-cal-list');
  if(!el)return;

  const today=new Date();
  const recurring=P.filter(p=>p.status==='customer_recurring');

  if(!recurring.length){
    el.innerHTML='<div class="tempty"><div class="ei">&#x1F9FC;</div><div>No recurring clients yet. Close your first deal to start tracking service.</div></div>';
    return;
  }

  // Geo-cluster suggestions
  const clusters=geoClusterSchedule();
  if(clusters&&clusters.length){
    el.innerHTML='<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:10px;font-weight:700;color:var(--blu);margin-bottom:4px">&#x1F4CD; Route Optimization Suggestion</div>'
      +clusters.slice(0,2).map(cl=>'<div style="font-size:10px;color:var(--sub);margin-bottom:3px">'
        +cl.anchor.name.slice(0,20)+' + '+cl.nearby.length+' nearby — schedule same day (within 8mi, ±5d)'
        +'</div>').join('')
      +'</div>';
  } else {
    el.innerHTML='';
  }

  // Sort by urgency: overdue first, then by next_service date
  const withDates=recurring.map(p=>{
    const c=customers[p.id]||{};
    let daysUntil=null;
    if(c.next_service){
      daysUntil=Math.round((new Date(c.next_service)-today)/864e5);
    } else if(c.last_service){
      const lastD=new Date(c.last_service);
      daysUntil=30-Math.round((today-lastD)/864e5);
    }
    return {...p,_daysUntil:daysUntil,_cust:c};
  }).sort((a,b)=>{
    if(a._daysUntil===null)return 1;
    if(b._daysUntil===null)return -1;
    return a._daysUntil-b._daysUntil;
  });

  el.innerHTML=withDates.map(p=>{
    const c=p._cust;
    const du=p._daysUntil;
    const cls=du===null?'no-date':du<0?'overdue':du<=7?'due-soon':'on-track';
    const duLabel=du===null?'No date set'
      :du<0?Math.abs(du)+'d OVERDUE'
      :du===0?'Due TODAY'
      :du===1?'Due TOMORROW'
      :'Due in '+du+'d';
    const duCol=du===null?'var(--sub)':du<0?'#dc2626':du<=7?'#d97706':'#059669';

    return '<div class="svc-card '+cls+'">'
      // Header row
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:13px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+' &bull; '+p.machines+' machine'+(p.machines>1?'s':'')+'</div>'
          +(p.phone?'<div style="font-size:10px;color:var(--blu)">'+p.phone+'</div>':'')
        +'</div>'
        +'<div style="text-align:right;flex-shrink:0;margin-left:10px">'
          +'<div style="font-size:12px;font-weight:800;color:'+duCol+'">'+duLabel+'</div>'
          +'<div style="font-size:10px;font-weight:700;color:#059669">$'+(c.monthly||p.monthly||149)+'/mo</div>'
        +'</div>'
      +'</div>'
      // Service history
      +'<div style="display:flex;justify-content:space-between;font-size:9px;color:var(--sub);background:#f5f8fa;border-radius:6px;padding:6px 8px">'
        +'<span>Last: '+(c.last_service||'Never recorded')+'</span>'
        +'<span>Next: '+(c.next_service||'Not set')+'</span>'
      +'</div>'
      // Actions
      +'<div style="display:flex;gap:5px">'
        +'<button onclick="openServiceLog('+p.id+')" style="flex:2;padding:7px;border:none;border-radius:7px;background:#059669;color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit">&#x2713; Log Service Visit</button>'
        +'<button onclick="reschedule('+p.id+')" style="flex:1;padding:7px;border:1px solid var(--brd);border-radius:7px;background:var(--surf);color:var(--sub);font-size:10px;cursor:pointer;font-family:inherit">Reschedule</button>'
        +(c.annual_schedule?'<button onclick="exportSchedulePDF('+p.id+')" style="flex:1;padding:7px;border:1px solid var(--blu);border-radius:7px;background:#eff6ff;color:var(--blu);font-size:10px;cursor:pointer;font-family:inherit">&#x1F4C5; Schedule</button>':'')
        +(p.phone?'<a href="tel:'+p.phone.replace(/\s/g,'')+'" style="flex:1;padding:7px;border:1px solid var(--blu);border-radius:7px;background:#eff6ff;color:var(--blu);font-size:10px;text-align:center;text-decoration:none;display:flex;align-items:center;justify-content:center;font-family:inherit">Call</a>':'')
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

function logServiceFromCal(id,opts){
  opts=opts||{};
  const today=new Date();
  const todayStr=today.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  const timeStr=today.toLocaleTimeString('en-US',{hour:'2-digit',minute:'2-digit'});
  if(!customers[id])customers[id]={};
  const c=customers[id];

  // Determine if this is a 6-month deep clean
  const lastDeep=c.service_history
    ?c.service_history.filter(s=>s.type==='deep_clean').slice(-1)[0]
    :null;
  const daysSinceDeep=lastDeep
    ?Math.round((today-new Date(lastDeep.date))/864e5)
    :999;
  const isDeepClean=opts.type==='deep_clean'||(daysSinceDeep>=170);

  // Build service record
  const svcRecord={
    date: today.toISOString().slice(0,10),
    date_display: todayStr,
    time: timeStr,
    type: isDeepClean?'deep_clean':'maintenance_60',
    atp_pre: opts.atp_pre||'',
    atp: opts.atp||'',
    filter_replaced: opts.filter_replaced||false,
    filter_type: opts.filter_type||c.filter_type||'',
    machine_brand: opts.machine_brand||c.machine_brand||'',
    machine_model: opts.machine_model||c.machine_model||'',
    machine_serial: opts.machine_serial||c.machine_serial||'',
    units: opts.units||1,
    notes: opts.notes||'',
    tech: 'Pinellas Ice Co',
  };

  if(!c.service_history)c.service_history=[];
  c.service_history.push(svcRecord);

  c.last_service=todayStr;
  c.last_service_iso=today.toISOString().slice(0,10);

  // Next service: 60 days for maintenance, but auto-schedule deep clean at 6 months
  const next=new Date(today);
  next.setDate(next.getDate()+60);
  c.next_service=next.toISOString().slice(0,10);

  // Flag 6-month deep clean if approaching
  const nextDeepDays=lastDeep
    ?180-Math.round((today-new Date(lastDeep.date))/864e5)
    :180-daysSinceDeep;
  c.next_deep_clean_in=Math.max(0,nextDeepDays);

  custSave();
  renderAtRisk();
  renderServiceCal();
  renderForecast();
  renderBriefing();
  const label=isDeepClean?'Deep clean':'60-day maintenance';
  toast(label+' logged ✓ Next: '+next.toLocaleDateString('en-US',{month:'short',day:'numeric'}));
}

function openServiceLog(id){
  const p=P.find(x=>x.id===id);
  const c=customers[id]||{};
  if(!p)return;

  const brand=c.machine_brand||'';
  const lastSvc=c.service_history&&c.service_history.length
    ?c.service_history[c.service_history.length-1]
    :null;

  // Build modal
  const bg=document.createElement('div');
  bg.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:200;display:flex;align-items:flex-end;justify-content:center';
  bg.id='svc-log-bg';

  bg.innerHTML='<div style="background:var(--surf);border-radius:16px 16px 0 0;padding:16px;width:100%;max-width:480px;max-height:85vh;overflow-y:auto">'
    +'<div style="font-weight:800;font-size:14px;color:var(--navy);margin-bottom:12px">Log Service Visit</div>'
    +'<div style="font-weight:600;font-size:12px;color:var(--sub);margin-bottom:10px">'+p.name+' &bull; '+p.city+'</div>'

    // Service type
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Service Type</div>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px">'
      +'<button onclick="setSvcMaint()" ontouchend="event.preventDefault();setSvcMaint()" id="svctype-maint" class="svctype-btn on" style="padding:8px;border:2px solid var(--navy);border-radius:8px;background:#f0f4ff;color:var(--navy);font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">&#x1F527; 60-Day Maintenance</button>'
      +'<button onclick="setSvcDeep()" ontouchend="event.preventDefault();setSvcDeep()" id="svctype-deep" class="svctype-btn" style="padding:8px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:10px;font-weight:700;cursor:pointer;font-family:inherit">&#x1F9FC; Deep Clean</button>'
    +'</div>'

    // Machine info + ATP pre/post
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:6px">Machine Info <span style="font-weight:400;font-style:italic">(auto-saved per client)</span></div>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:10px">'
      +'<select id="svc-machine-brand" style="padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
      +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
      +'</select>'
      +'<select id="svc-machine-type" style="padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
      +MACHINE_TYPE_LIST.map(t=>'<option value="'+t+'"'+(t===(c.machine_type||'')?'selected':'')+'>'+t+'</option>').join('')
      +'</select>'
      +'<input id="svc-machine-model" type="text" placeholder="Model #" value="'+(c.machine_model||'')+'"\'\n        +\' style="padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
      +'<input id="svc-machine-serial" type="text" placeholder="Serial #" value="'+(c.machine_serial||'')+'"\'\n        +\' style="padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
      +'<input id="svc-units" type="number" min="1" max="10" placeholder="# units" value="'+(c.machines||1)+'"\'\n        +\' style="padding:7px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
    +'</div>'
    // ATP Readings
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:6px">ATP Readings (RLU) <span style="font-weight:400;font-style:italic">— required for report</span></div>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:8px">'
      +'<div>'
        +'<div style="font-size:9px;color:#dc2626;font-weight:700;margin-bottom:4px">⚠ Before (pre-clean)</div>'
        +'<input id="svc-atp-pre" type="number" placeholder="e.g. 847" min="0" max="9999"'
          +' style="width:100%;padding:10px;border:2px solid #fca5a5;border-radius:7px;font-size:18px;font-weight:800;font-family:inherit;background:#fff;color:#dc2626;outline:none;text-align:center">'
      +'</div>'
      +'<div>'
        +'<div style="font-size:9px;color:#059669;font-weight:700;margin-bottom:4px">✓ After (post-clean)</div>'
        +'<input id="svc-atp" type="number" placeholder="e.g. 6" min="0" max="9999"'
          +' style="width:100%;padding:10px;border:2px solid #6ee7b7;border-radius:7px;font-size:18px;font-weight:800;font-family:inherit;background:#fff;color:#059669;outline:none;text-align:center">'
      +'</div>'
    +'</div>'
    +'<div id="atp-indicator" style="font-size:9px;font-weight:700;padding:4px 8px;border-radius:6px;background:#f0fdf4;color:#059669;margin-bottom:8px;display:none">✓ Pass</div>'

    // Filter replaced
    +'<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;padding:8px;background:#f5f8fa;border-radius:7px">'
      +'<input type="checkbox" id="svc-filter-replaced" onchange="toggleFilterType()">'
      +'<label for="svc-filter-replaced" style="font-size:11px;font-weight:600;color:var(--navy);cursor:pointer">Filter replaced this visit</label>'
    +'</div>'
    +'<div id="svc-filter-type-row" style="display:none;margin-bottom:10px">'
      +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Filter Type Installed</div>'
      +'<select id="svc-filter-type" style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
      +FILTER_TYPES.map(f=>'<option'+(f===(c.filter_type||'')?' selected':'')+'>'+f+'</option>').join('')
      +'</select>'
    +'</div>'

    // Notes
    +'<div style="font-size:9px;font-weight:700;color:var(--sub);text-transform:uppercase;margin-bottom:5px">Service Notes</div>'
    +'<textarea id="svc-notes" rows="3" placeholder="Machine condition, issues found, recommendations..."'
      +' style="width:100%;padding:8px;border:1px solid var(--brd);border-radius:7px;font-size:11px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none;resize:none;margin-bottom:10px"></textarea>'

    // Previous service info
    +(lastSvc?'<div style="background:#f5f8fa;border-radius:7px;padding:8px;margin-bottom:10px;font-size:10px;color:var(--sub)">'
      +'<span style="font-weight:600;color:var(--navy)">Previous: </span>'+lastSvc.date_display
      +' &bull; '+(lastSvc.type==='deep_clean'?'Deep Clean':'60-Day Maintenance')
      +(lastSvc.atp?' &bull; ATP: '+lastSvc.atp:'')
      +'</div>':'')

    // Buttons
    +'<div style="display:flex;gap:6px">'
      +'<button onclick="submitServiceLog('+id+')" ontouchend="event.preventDefault();submitServiceLog('+id+')" style="flex:2;padding:10px;border:none;border-radius:8px;background:#059669;color:#fff;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">&#x2713; Save Service Visit</button>'
      +'<button onclick="closeSvcLog()" ontouchend="event.preventDefault();closeSvcLog()" style="flex:1;padding:10px;border:1px solid var(--brd);border-radius:8px;background:var(--surf);color:var(--sub);font-size:11px;cursor:pointer;font-family:inherit">Cancel</button>'
    +'</div>'
    +'</div>';

  document.body.appendChild(bg);

  // ATP live feedback
  const atpInput=document.getElementById('svc-atp');
  if(atpInput){
    atpInput.addEventListener('input',function(){
      const v=parseInt(this.value);
      const ind=document.getElementById('atp-indicator');
      if(!ind)return;
      if(isNaN(v)){ind.textContent='Enter RLU';ind.style.background='#f5f8fa';ind.style.color='var(--sub)';}
      else if(v<=10){ind.textContent='Excellent';ind.style.background='#f0fdf4';ind.style.color='#059669';}
      else if(v<=30){ind.textContent='Good';ind.style.background='#eff6ff';ind.style.color='var(--blu)';}
      else if(v<=100){ind.textContent='Acceptable';ind.style.background='#fef9ee';ind.style.color='#d97706';}
      else{ind.textContent='Retest/Reservice';ind.style.background='#fef2f2';ind.style.color='#dc2626';}
    });
  }
}

let _svcType='maintenance_60';
function closeSvcLog(){const el=document.getElementById('svc-log-bg');if(el)el.remove();}
function toggleStep(el){const next=el.nextElementSibling;if(next)next.style.display=next.style.display==='none'?'block':'none';}

function getDiagram(brand){
  const diagrams={
    'Manitowoc': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Manitowoc Cube Ice Machine</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <!-- Machine outline -->
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Evaporator plates -->
        <rect x="60" y="25" width="200" height="55" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="45" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">EVAPORATOR PLATES</text>
        <text x="160" y="60" text-anchor="middle" font-size="8" fill="#3b82f6">Where ice forms — clean &amp; sanitize thoroughly</text>
        <text x="160" y="73" text-anchor="middle" font-size="7" fill="#64748b">Scale = white/gray crust | Biofilm = pink/orange slime</text>
        <!-- Distribution tube -->
        <rect x="55" y="85" width="210" height="14" rx="3" fill="#fef9c3" stroke="#ca8a04" stroke-width="1.5"/>
        <text x="160" y="96" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">WATER DISTRIBUTION TUBE — clear holes with toothpick</text>
        <!-- Water curtain -->
        <rect x="55" y="103" width="210" height="10" rx="2" fill="#f0fdf4" stroke="#16a34a" stroke-width="1.5"/>
        <text x="160" y="112" text-anchor="middle" font-size="8" fill="#15803d">WATER CURTAIN — remove &amp; soak in cleaner</text>
        <!-- Water trough -->
        <rect x="55" y="117" width="210" height="30" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="130" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH</text>
        <text x="160" y="142" text-anchor="middle" font-size="8" fill="#be185d">Highest ATP — biofilm source — scrub thoroughly</text>
        <!-- Float valve -->
        <circle cx="75" cy="132" r="8" fill="#fef3c7" stroke="#d97706" stroke-width="1.5"/>
        <text x="75" y="160" text-anchor="middle" font-size="7" fill="#92400e">FLOAT</text>
        <!-- Ice bin -->
        <rect x="55" y="153" width="210" height="47" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="172" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN</text>
        <text x="160" y="185" text-anchor="middle" font-size="8" fill="#64748b">Sanitize all interior surfaces — discard first batch</text>
        <!-- Arrow labels -->
        <text x="30" y="55" font-size="7" fill="#64748b" transform="rotate(-90,30,55)">TOP</text>
        <text x="30" y="180" font-size="7" fill="#64748b" transform="rotate(-90,30,180)">BOTTOM</text>
      </svg>
    </div>`,
    'Hoshizaki': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Hoshizaki Crescent Cube Machine</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Evaporator - stainless -->
        <rect x="60" y="25" width="200" height="50" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="43" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">EVAPORATOR (STAINLESS STEEL)</text>
        <text x="160" y="56" text-anchor="middle" font-size="8" fill="#3b82f6">Crescent cube molds — uniform silver when clean</text>
        <text x="160" y="69" text-anchor="middle" font-size="7" fill="#16a34a">Any commercial cleaner OK — soft brush only</text>
        <!-- Spray bar - highlighted as critical -->
        <rect x="55" y="80" width="210" height="16" rx="3" fill="#fef08a" stroke="#ca8a04" stroke-width="2"/>
        <text x="160" y="92" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">&#x26A0; SPRAY BAR — MOST CRITICAL COMPONENT</text>
        <!-- Spray bar dots -->
        <circle cx="80" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="100" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="120" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="140" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="160" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="180" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="200" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="220" cy="88" r="2" fill="#ca8a04"/>
        <circle cx="240" cy="88" r="2" fill="#ca8a04"/>
        <text x="160" y="106" text-anchor="middle" font-size="7" fill="#92400e">Clear EVERY hole with toothpick — blocked holes = irregular ice</text>
        <!-- Float switch -->
        <rect x="55" y="112" width="80" height="20" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="95" y="122" text-anchor="middle" font-size="8" font-weight="bold" fill="#9d174d">FLOAT SWITCH</text>
        <text x="95" y="130" text-anchor="middle" font-size="6" fill="#be185d">Must slide freely</text>
        <!-- Water trough -->
        <rect x="55" y="136" width="210" height="25" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="148" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH — highest ATP reading</text>
        <text x="160" y="158" text-anchor="middle" font-size="7" fill="#be185d">Scrub biofilm (pink/orange) — common in FL heat</text>
        <!-- Bin -->
        <rect x="55" y="165" width="210" height="40" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="182" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN — Sanitize all interior surfaces</text>
        <text x="160" y="196" text-anchor="middle" font-size="7" fill="#64748b">Discard first 2 full bins after service</text>
      </svg>
    </div>`,
    'Scotsman': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Scotsman Prodigy Vertical Evaporator</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Vertical evaporator plate -->
        <rect x="60" y="20" width="70" height="140" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="95" y="55" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">VERTICAL</text>
        <text x="95" y="68" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">EVAPORATOR</text>
        <text x="95" y="82" text-anchor="middle" font-size="7" fill="#3b82f6">Scotsman unique</text>
        <text x="95" y="94" text-anchor="middle" font-size="7" fill="#3b82f6">vertical design</text>
        <text x="95" y="108" text-anchor="middle" font-size="7" fill="#3b82f6">Ice forms on</text>
        <text x="95" y="120" text-anchor="middle" font-size="7" fill="#3b82f6">flat vertical</text>
        <text x="95" y="132" text-anchor="middle" font-size="7" fill="#3b82f6">plate surface</text>
        <!-- Critical: thickness sensor -->
        <circle cx="148" cy="90" r="10" fill="#fee2e2" stroke="#dc2626" stroke-width="2"/>
        <text x="148" y="94" text-anchor="middle" font-size="7" font-weight="bold" fill="#dc2626">SENS</text>
        <line x1="130" y1="90" x2="158" y2="90" stroke="#dc2626" stroke-width="1.5"/>
        <text x="210" y="75" text-anchor="middle" font-size="9" font-weight="bold" fill="#dc2626">&#x26A0; ICE THICKNESS</text>
        <text x="210" y="87" text-anchor="middle" font-size="9" font-weight="bold" fill="#dc2626">SENSOR</text>
        <text x="210" y="99" text-anchor="middle" font-size="7" fill="#dc2626">Gap = 3/8 inch</text>
        <text x="210" y="111" text-anchor="middle" font-size="7" fill="#dc2626">(pencil thickness)</text>
        <text x="210" y="123" text-anchor="middle" font-size="7" fill="#dc2626">Clean at EVERY visit</text>
        <!-- Control panel -->
        <rect x="145" y="25" width="140" height="40" rx="4" fill="#f0fdf4" stroke="#16a34a" stroke-width="1.5"/>
        <text x="215" y="40" text-anchor="middle" font-size="9" font-weight="bold" fill="#15803d">PRODIGY CONTROL</text>
        <text x="215" y="54" text-anchor="middle" font-size="7" fill="#15803d">Check error codes before cleaning</text>
        <!-- Drain pan critical -->
        <rect x="55" y="165" width="210" height="30" rx="3" fill="#fef9c3" stroke="#ca8a04" stroke-width="2"/>
        <text x="160" y="178" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">DRAIN PAN — check at every visit</text>
        <text x="160" y="190" text-anchor="middle" font-size="7" fill="#92400e">Pour 1 cup water — should drain in under 30 seconds</text>
        <!-- Water trough -->
        <rect x="55" y="162" width="210" height="0" rx="3" fill="none"/>
      </svg>
    </div>`,
    'Ice-O-Matic': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Ice-O-Matic (similar to Manitowoc)</div>
      <svg viewBox="0 0 320 210" style="width:100%;max-height:190px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="190" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <rect x="55" y="22" width="210" height="45" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="40" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">EVAPORATOR PLATE</text>
        <text x="160" y="53" text-anchor="middle" font-size="8" fill="#3b82f6">Nickel-plated — scale forms in strips matching blocked holes</text>
        <!-- Distribution pan critical -->
        <rect x="55" y="72" width="210" height="20" rx="3" fill="#fef08a" stroke="#ca8a04" stroke-width="2"/>
        <text x="160" y="84" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">&#x26A0; DISTRIBUTION PAN — MOST CRITICAL (IOM)</text>
        <!-- Hole dots -->
        <circle cx="80" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="95" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="110" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="125" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="140" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="155" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="170" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="185" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="200" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="215" cy="82" r="1.5" fill="#ca8a04"/>
        <circle cx="230" cy="82" r="1.5" fill="#ca8a04"/>
        <text x="160" y="101" text-anchor="middle" font-size="7" fill="#92400e">Clear EVERY hole with toothpick — IOM unique critical step</text>
        <rect x="55" y="105" width="210" height="25" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="117" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH — scrub biofilm from corners</text>
        <text x="160" y="127" text-anchor="middle" font-size="7" fill="#be185d">Check drain flows freely</text>
        <!-- Bin baffle -->
        <rect x="90" y="134" width="140" height="15" rx="2" fill="#f0fdf4" stroke="#16a34a" stroke-width="1.5"/>
        <text x="160" y="145" text-anchor="middle" font-size="8" font-weight="bold" fill="#15803d">BIN BAFFLE — check underside for biofilm</text>
        <!-- Ice bin -->
        <rect x="55" y="153" width="210" height="40" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="170" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN</text>
        <text x="160" y="184" text-anchor="middle" font-size="7" fill="#64748b">Sanitize all interior — discard first batch</text>
      </svg>
    </div>`,
    'Follett': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Follett Nugget Ice (Auger System)</div>
      <svg viewBox="0 0 320 220" style="width:100%;max-height:200px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="200" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Evaporator cylinder - vertical -->
        <rect x="120" y="20" width="80" height="120" rx="8" fill="#dbeafe" stroke="#3b82f6" stroke-width="2"/>
        <text x="160" y="40" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">EVAPORATOR</text>
        <text x="160" y="52" text-anchor="middle" font-size="9" font-weight="bold" fill="#1e40af">CYLINDER</text>
        <text x="160" y="65" text-anchor="middle" font-size="7" fill="#3b82f6">Stainless steel</text>
        <text x="160" y="77" text-anchor="middle" font-size="7" fill="#3b82f6">Ice forms on</text>
        <text x="160" y="89" text-anchor="middle" font-size="7" fill="#3b82f6">inside wall</text>
        <!-- Auger inside cylinder -->
        <line x1="160" y1="25" x2="160" y2="135" stroke="#7c3aed" stroke-width="3" stroke-dasharray="4,2"/>
        <text x="160" y="108" text-anchor="middle" font-size="8" font-weight="bold" fill="#7c3aed">AUGER</text>
        <text x="160" y="120" text-anchor="middle" font-size="7" fill="#7c3aed">Pushes ice up</text>
        <!-- Motor on top -->
        <rect x="130" y="8" width="60" height="14" rx="3" fill="#f5f3ff" stroke="#7c3aed" stroke-width="1.5"/>
        <text x="160" y="19" text-anchor="middle" font-size="7" font-weight="bold" fill="#6d28d9">AUGER MOTOR</text>
        <!-- Warning about auger -->
        <rect x="210" y="45" width="85" height="45" rx="4" fill="#fef2f2" stroke="#dc2626" stroke-width="1.5"/>
        <text x="252" y="60" text-anchor="middle" font-size="8" font-weight="bold" fill="#dc2626">&#x26A0; CAUTION</text>
        <text x="252" y="72" text-anchor="middle" font-size="7" fill="#dc2626">Auger is SHARP</text>
        <text x="252" y="83" text-anchor="middle" font-size="7" fill="#dc2626">Handle carefully</text>
        <!-- Ice chute -->
        <rect x="55" y="45" width="60" height="40" rx="4" fill="#fef9c3" stroke="#ca8a04" stroke-width="1.5"/>
        <text x="85" y="62" text-anchor="middle" font-size="8" font-weight="bold" fill="#92400e">ICE CHUTE</text>
        <text x="85" y="74" text-anchor="middle" font-size="7" fill="#92400e">Clean at</text>
        <text x="85" y="83" text-anchor="middle" font-size="7" fill="#92400e">every visit</text>
        <!-- Bin -->
        <rect x="55" y="148" width="210" height="52" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="168" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">STORAGE BIN</text>
        <text x="160" y="180" text-anchor="middle" font-size="7" fill="#64748b">Listen for grinding before cleaning — indicates scale on cylinder</text>
        <text x="160" y="192" text-anchor="middle" font-size="7" fill="#64748b">Soft/moist nuggets = healthy | Hard nuggets = issue</text>
        <!-- Arrow showing auger direction -->
        <text x="205" y="100" font-size="8" fill="#7c3aed">&#x2191; ice flow</text>
      </svg>
    </div>`,
    'Cornelius': `<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:10px;margin-bottom:10px">
      <div style="font-size:10px;font-weight:700;color:var(--navy);margin-bottom:8px">&#x1F4CD; Component Diagram — Cornelius Compact Machine</div>
      <svg viewBox="0 0 320 190" style="width:100%;max-height:170px" xmlns="http://www.w3.org/2000/svg">
        <rect x="20" y="10" width="280" height="170" rx="8" fill="#f8fafc" stroke="#94a3b8" stroke-width="2"/>
        <!-- Compact layout - front accessible -->
        <rect x="55" y="22" width="210" height="40" rx="4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.5"/>
        <text x="160" y="38" text-anchor="middle" font-size="10" font-weight="bold" fill="#1e40af">FREEZE PLATE</text>
        <text x="160" y="52" text-anchor="middle" font-size="7" fill="#3b82f6">Front-accessible — scale = white/gray | clean with soft brush</text>
        <!-- Air filter critical -->
        <rect x="55" y="67" width="90" height="30" rx="3" fill="#fef08a" stroke="#ca8a04" stroke-width="2"/>
        <text x="100" y="80" text-anchor="middle" font-size="9" font-weight="bold" fill="#92400e">&#x26A0; AIR FILTER</text>
        <text x="100" y="91" text-anchor="middle" font-size="7" fill="#92400e">#1 Maintenance Item</text>
        <!-- Location of filter -->
        <text x="100" y="109" text-anchor="middle" font-size="7" fill="#64748b">Side or rear mount</text>
        <text x="100" y="119" text-anchor="middle" font-size="7" fill="#64748b">Rinse + blow out</text>
        <!-- Drain pan odor -->
        <rect x="155" y="67" width="110" height="30" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="210" y="80" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">DRAIN PAN</text>
        <text x="210" y="91" text-anchor="middle" font-size="7" fill="#be185d">Odor source — scrub biofilm</text>
        <!-- Water trough -->
        <rect x="55" y="103" width="210" height="22" rx="3" fill="#fce7f3" stroke="#db2777" stroke-width="1.5"/>
        <text x="160" y="114" text-anchor="middle" font-size="9" font-weight="bold" fill="#9d174d">WATER TROUGH — check mineral deposits</text>
        <text x="160" y="122" text-anchor="middle" font-size="7" fill="#be185d">Cloudy ice = mineral contamination = filter issue</text>
        <!-- Bin -->
        <rect x="55" y="130" width="210" height="42" rx="3" fill="#f1f5f9" stroke="#64748b" stroke-width="1.5"/>
        <text x="160" y="148" text-anchor="middle" font-size="9" font-weight="bold" fill="#334155">ICE BIN — Sanitize thoroughly</text>
        <text x="160" y="161" text-anchor="middle" font-size="7" fill="#64748b">Simplest internal layout — faster service than modular units</text>
      </svg>
    </div>`,
  };
  return diagrams[brand]||'';
}
function setSvcMaint(){setSvcType('maintenance_60');}
function setSvcDeep(){setSvcType('deep_clean');}
function setSvcType(t){
  _svcType=t;
  const idMap={'maintenance_60':'svctype-maint','deep_clean':'svctype-deep'};
  document.querySelectorAll('.svctype-btn').forEach(b=>{
    const isOn=b.id===idMap[t];
    b.style.border=isOn?'2px solid var(--navy)':'1px solid var(--brd)';
    b.style.background=isOn?'#f0f4ff':'var(--surf)';
    b.style.color=isOn?'var(--navy)':'var(--sub)';
  });
}
function toggleFilterType(){
  const cb=document.getElementById('svc-filter-replaced');
  const row=document.getElementById('svc-filter-type-row');
  if(row)row.style.display=cb&&cb.checked?'block':'none';
}
function submitServiceLog(id){
  const atp=(document.getElementById('svc-atp')||{}).value||'';
  const atpPre=(document.getElementById('svc-atp-pre')||{}).value||'';
  const notes=(document.getElementById('svc-notes')||{}).value||'';
  const filterReplaced=!!(document.getElementById('svc-filter-replaced')||{}).checked;
  const filterType=(document.getElementById('svc-filter-type')||{}).value||'';
  const machineBrand=(document.getElementById('svc-machine-brand')||{}).value||'';
  const machineType=(document.getElementById('svc-machine-type')||{}).value||'';
  const machineModel=(document.getElementById('svc-machine-model')||{}).value||'';
  const machineSerial=(document.getElementById('svc-machine-serial')||{}).value||'';
  const units=parseInt((document.getElementById('svc-units')||{}).value||'1')||1;

  logServiceFromCal(id,{
    type:_svcType,
    atp,atp_pre:atpPre,notes,
    filter_replaced:filterReplaced,
    filter_type:filterType,
    machine_brand:machineBrand,
    machine_model:machineModel,
    machine_serial:machineSerial,
    units,
  });

  // Save machine info and filter to customer record for future reports
  if(!customers[id])customers[id]={};
  if(machineBrand)customers[id].machine_brand=machineBrand;
  if(machineType)customers[id].machine_type=machineType;
  if(machineModel)customers[id].machine_model=machineModel;
  if(machineSerial)customers[id].machine_serial=machineSerial;
  if(units>1)customers[id].machines=units;
  if(atpPre)customers[id].atp_pre_last=atpPre;
  if(atp)customers[id].atp_post_last=atp;
  if(filterReplaced&&filterType){
    customers[id].filter_type=filterType;
    customers[id].filter_installed=new Date().toISOString().slice(0,10);
  }
  // Set last service to today, next service 60 days out
  const today_str=new Date().toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
  const today_iso=new Date().toISOString().slice(0,10);
  const next_iso=new Date(Date.now()+60*864e5).toISOString().slice(0,10);
  customers[id].last_service=today_str;
  customers[id].next_service=next_iso;
  // Add to service history
  if(!customers[id].service_history)customers[id].service_history=[];
  customers[id].service_history.push({
    date_display:today_str,date:today_iso,
    type:svcType,label:svcType==='deep_clean'?'Deep Clean':'60-Day Maintenance',
    atp:atp,atp_pre:atpPre,notes:notes,filter_replaced:filterReplaced,status:'completed'
  });
  // Add to ATP history for reporting
  if(!customers[id].atp_history)customers[id].atp_history=[];
  if(atp||atpPre)customers[id].atp_history.push({date:today_iso,pre:atpPre,post:atp});
  // Log service_done outcome with ATP data for report generation
  if(!log[id])log[id]=[];
  log[id].push({
    outcome:'service_done',type:'service',reason:null,
    date:today_str,atp:atp,atp_pre:atpPre,notes:notes,
    filter_replaced:filterReplaced,filter_type:filterType,
  });
  lSave();
  custSave();
  document.getElementById('svc-log-bg').remove();
  toast('\u2713 Service logged \u2014 next due '+next_iso);
  // Update weekly funnel / briefing
  if(typeof renderBriefing==='function')setTimeout(renderBriefing,100);
  rCust();
}

function reschedule(id){
  const c=customers[id]||{};
  const current=c.next_service||new Date(Date.now()+60*864e5).toISOString().slice(0,10);
  const d=prompt('Set next service date (YYYY-MM-DD):',current);
  if(!d||!/^\d{4}-\d{2}-\d{2}$/.test(d)){toast('Invalid date');return;}
  if(!customers[id])customers[id]={};
  customers[id].next_service=d;
  // Rebuild annual schedule from new date
  if(customers[id].annual_schedule&&customers[id].annual_schedule.length){
    const today=new Date().toISOString().slice(0,10);
    // Update the next upcoming scheduled visit
    const nextIdx=customers[id].annual_schedule.findIndex(s=>s.status!=='completed'&&s.date>=today);
    if(nextIdx>=0){
      customers[id].annual_schedule[nextIdx].date=d;
      customers[id].annual_schedule[nextIdx].date_display=new Date(d).toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
      // Realign subsequent visits every 60 days from new date
      for(let i=nextIdx+1;i<customers[id].annual_schedule.length;i++){
        const prev=new Date(customers[id].annual_schedule[i-1].date);
        const next=new Date(prev);next.setDate(next.getDate()+60);
        customers[id].annual_schedule[i].date=next.toISOString().slice(0,10);
        customers[id].annual_schedule[i].date_display=next.toLocaleDateString('en-US',{month:'short',day:'numeric',year:'numeric'});
      }
    }
  }
  custSave();
  renderServiceCal();
  toast('Rescheduled ✓ Schedule realigned from '+d);
}

// ── SERVICE ROUTE BUILDER ─────────────────────────────────────────────────────
function renderTutorial(){
  const brand=(document.getElementById('tut-brand')||{}).value||'';
  const type=(document.getElementById('tut-type')||{}).value||'deep_clean';
  // ATP protocol is brand-independent - render it separately
  if(type==='atp_protocol'){renderATPProtocol();return;}
  const el=document.getElementById('tut-content');
  if(!el)return;
  if(!brand){el.innerHTML='<div class="tempty"><div>Select a machine brand above</div></div>';return;}

  const data=type==='deep_clean'?DEEP_CLEAN[brand]:MAINTENANCE_60[brand];
  if(!data){el.innerHTML='<div class="tempty"><div>Tutorial not found</div></div>';return;}

  let html='<div style="background:#f0f4ff;border-radius:8px;padding:10px;margin-bottom:10px">'
    +'<div style="font-weight:800;font-size:12px;color:var(--navy)">'+brand+' — '+(type==='deep_clean'?'Full Deep Clean':'60-Day Maintenance')+'</div>'
    +'<div style="font-size:10px;color:var(--sub);margin-top:2px">⏱ Est. time: '+data.time+'</div>';

  if(data.chemicals){
    html+='<div style="font-size:9px;color:var(--sub);margin-top:4px"><b>Chemicals:</b> '+data.chemicals.join(' • ')+'</div>';
  }
  html+='</div>';

  // Video & resource links
  const vids=VIDEO_LINKS[brand]||{};
  if(vids.deep_url||vids.guide_url){
    html+='<div style="background:#fff7f5;border:1px solid #fed7aa;border-radius:8px;padding:10px;margin-bottom:10px">'
      +'<div style="font-size:10px;font-weight:700;color:#ea580c;margin-bottom:6px">&#x1F3A5; Reference Resources</div>'
      // YouTube component ID search - most useful for beginners
      +(vids.yt_component?'<a href="'+vids.yt_component+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:7px 8px;background:#ff0000;border-radius:6px;text-decoration:none;margin-bottom:5px">'
        +'<span style="font-size:14px;color:#fff">&#x25B6;</span>'
        +'<div><div style="font-size:10px;font-weight:700;color:#fff">&#x1F4CD; YouTube: Identify Components</div>'
        +'<div style="font-size:9px;color:rgba(255,255,255,.8)">Search videos showing inside of '+brand+' — identify each part visually</div></div>'
      +'</a>':'')
      +(vids.youtube_url?'<a href="'+vids.youtube_url+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:7px 8px;background:#cc0000;border-radius:6px;text-decoration:none;margin-bottom:5px">'
        +'<span style="font-size:14px;color:#fff">&#x1F9FC;</span>'
        +'<div><div style="font-size:10px;font-weight:700;color:#fff">YouTube: Full Cleaning Tutorial</div>'
        +'<div style="font-size:9px;color:rgba(255,255,255,.8)">Watch a complete '+brand+' deep clean walkthrough</div></div>'
      +'</a>':'')
      +(vids.deep_url?'<a href="'+vids.deep_url+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:6px 8px;background:#fff;border:1px solid #fed7aa;border-radius:6px;text-decoration:none;margin-bottom:5px">'
        +'<span style="font-size:14px">&#x25B6;</span>'
        +'<div><div style="font-size:10px;font-weight:600;color:#ea580c">Official Cleaning Reference</div>'
        +'<div style="font-size:9px;color:var(--sub)">Step-by-step guide with photos</div></div>'
      +'</a>':'')
      +(vids.official_url&&vids.official_url!==vids.deep_url?'<a href="'+vids.official_url+'" target="_blank" style="display:flex;align-items:center;gap:6px;padding:6px 8px;background:#fff;border:1px solid #fed7aa;border-radius:6px;text-decoration:none">'
        +'<span style="font-size:14px">&#x1F3ED;</span>'
        +'<div><div style="font-size:10px;font-weight:600;color:#ea580c">Official: '+brand+' Support</div>'
        +'<div style="font-size:9px;color:var(--sub)">Manufacturer training & documentation</div></div>'
      +'</a>':'')
    +'</div>';
  }

  // Component diagram (deep clean only)
  if(type==='deep_clean'){html+=getDiagram(brand);}

  data.steps.forEach(function(s,i){
    html+='<div style="margin-bottom:8px;border:1px solid var(--brd);border-radius:8px;overflow:hidden">'
      +'<div onclick="toggleStep(this)" style="display:flex;align-items:center;gap:8px;padding:10px;background:var(--surf);cursor:pointer">'
        +'<div style="background:var(--navy);color:#fff;border-radius:50%;width:22px;height:22px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;flex-shrink:0">'+(i+1)+'</div>'
        +'<div style="font-weight:700;font-size:11px;color:var(--navy);flex:1">'+s.title+'</div>'
        +'<div style="font-size:10px;color:var(--sub)">▼</div>'
      +'</div>'
      +'<div style="display:none;padding:10px;font-size:11px;color:var(--sub);line-height:1.5;border-top:1px solid var(--brd2)">'+s.detail+'</div>'
    +'</div>';
  });

  el.innerHTML=html;
}


function renderATPProtocol(){
  const el=document.getElementById('tut-content');
  if(!el)return;
  const s=CHEM_REF.atp_sales;
  const c=CHEM_REF.atp_compliance;
  const a=CHEM_REF.atp;

  el.innerHTML=
    // Header
    '<div style="background:#f0f4ff;border-radius:8px;padding:10px;margin-bottom:12px">'
      +'<div style="font-weight:800;font-size:12px;color:var(--navy)">&#x1F4CA; ATP Testing Guide</div>'
      +'<div style="font-size:10px;color:var(--sub);margin-top:3px">Hygiena Ensure v2 • Ultrasnap testers • All brands</div>'
    +'</div>'

    // Meter technique
    +'<div style="background:#fff;border:1px solid var(--brd);border-radius:8px;padding:12px;margin-bottom:10px">'
      +'<div style="font-size:11px;font-weight:700;color:var(--navy);margin-bottom:6px">&#x1F52C; How to Use the Hygiena Ensure v2</div>'
      +'<div style="font-size:11px;color:var(--sub);line-height:1.6">'
        +'<b>1.</b> Snap the Ultrasnap tube to release reagent liquid into the swab chamber.<br>'
        +'<b>2.</b> Swab a 10cm² area (palm-sized) using a firm Z-pattern — 5 strokes across, 5 strokes down.<br>'
        +'<b>3.</b> Insert swab into Ensure v2 and press READ.<br>'
        +'<b>4.</b> Result appears in 15 seconds.<br><br>'
        +'<b>Pass thresholds:</b><br>'
        +'✅ 10 RLU or below — ice contact surfaces (evaporator, trough, distribution tube)<br>'
        +'✅ 30 RLU or below — general food contact surfaces (bin, scoop, exterior)<br>'
        +'❌ Above threshold: re-clean, re-sanitize, retest. Do not leave until passing.'
      +'</div>'
    +'</div>'

    // Sales protocol
    +'<div style="background:#fef9ee;border:1px solid #fde68a;border-radius:8px;padding:12px;margin-bottom:10px">'
      +'<div style="font-size:11px;font-weight:800;color:#92400e;margin-bottom:8px">&#x1F4B0; '+s.title+'</div>'
      +'<div style="font-size:10px;color:#78350f;margin-bottom:10px;line-height:1.5">'+s.purpose+'</div>'
      +s.locations.map((l,i)=>
        '<div style="padding:8px;background:#fffbeb;border-radius:6px;margin-bottom:6px">'
          +'<div style="font-size:11px;font-weight:700;color:#92400e">'+(i+1)+'. '+l.where+'</div>'
          +'<div style="font-size:10px;color:#78350f;margin-top:2px;line-height:1.4">'+l.why+'</div>'
        +'</div>'
      ).join('')
      +'<div style="background:#d97706;border-radius:7px;padding:10px;margin-top:8px">'
        +'<div style="font-size:10px;font-weight:700;color:#fff;margin-bottom:4px">The Pitch Move</div>'
        +'<div style="font-size:10px;color:rgba(255,255,255,.9);line-height:1.5">'+s.pitch+'</div>'
      +'</div>'
    +'</div>'

    // Compliance protocol
    +'<div style="background:#f0fdf4;border:1px solid #6ee7b7;border-radius:8px;padding:12px;margin-bottom:10px">'
      +'<div style="font-size:11px;font-weight:800;color:#065f46;margin-bottom:8px">'+c.title+'</div>'
      +'<div style="font-size:10px;color:#064e3b;margin-bottom:10px;line-height:1.5">'+c.purpose+'</div>'
      +c.locations.map((l,i)=>
        '<div style="padding:8px;background:#ecfdf5;border-radius:6px;margin-bottom:6px">'
          +'<div style="font-size:11px;font-weight:700;color:#065f46">'+(i+1)+'. '+l.where+'</div>'
          +'<div style="font-size:10px;color:#064e3b;margin-top:2px;line-height:1.4">'+l.why+'</div>'
        +'</div>'
      ).join('')
      +'<div style="background:#059669;border-radius:7px;padding:10px;margin-top:8px">'
        +'<div style="font-size:10px;font-weight:700;color:#fff;margin-bottom:4px">Report Format</div>'
        +'<div style="font-size:10px;color:rgba(255,255,255,.9);line-height:1.5">'+c.report_format+'</div>'
      +'</div>'
    +'</div>'

    // Compliance sticker
    +'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:10px">'
      +'<div style="font-size:11px;font-weight:700;color:var(--navy);margin-bottom-4px">&#x1F3F7; Compliance Sticker Placement</div>'
      +'<div style="font-size:10px;color:var(--sub);margin-top:4px;line-height:1.5">'+c.sticker+'</div>'
    +'</div>';
}

function renderServiceRoute(){
  const el=document.getElementById('svc-route-list');
  if(!el)return;
  el.innerHTML='<div style="font-size:10px;color:var(--sub)">Tap Build Route to optimize service stops for selected week.</div>';
  document.getElementById('svc-route-map-btn').style.display='none';
}

function buildServiceRoute(){
  const weekOffset=parseInt(document.getElementById('svc-week').value)||0;
  const today=new Date();
  const windowStart=new Date(today.getTime()+weekOffset*864e5);
  const windowEnd=new Date(windowStart.getTime()+7*864e5);

  // Get clients due in window
  const due=P.filter(p=>{
    if(p.status!=='customer_recurring')return false;
    const c=customers[p.id]||{};
    if(!c.next_service&&!c.last_service)return weekOffset<0; // overdue filter
    if(c.next_service){
      const d=new Date(c.next_service);
      if(weekOffset===-7)return d<today; // overdue
      return d>=windowStart&&d<=windowEnd;
    }
    if(c.last_service){
      const last=new Date(c.last_service);
      const daysSince=Math.round((today-last)/864e5);
      return daysSince>=25&&daysSince<=35; // due around now
    }
    return false;
  }).filter(p=>p.lat&&p.lon);

  if(!due.length){
    const el=document.getElementById('svc-route-list');
    if(el)el.innerHTML='<div class="tempty"><div>No clients due in this window</div></div>';
    return;
  }

  // Nearest-neighbor from first client (sorted by next_service date)
  due.sort((a,b)=>{
    const ca=customers[a.id]||{},cb=customers[b.id]||{};
    return (ca.next_service||'9999').localeCompare(cb.next_service||'9999');
  });
  serviceRoute=[due[0]];
  const remaining=[...due.slice(1)];
  while(remaining.length){
    const last=serviceRoute[serviceRoute.length-1];
    let best=0,bestD=Infinity;
    remaining.forEach((p,i)=>{
      const d=hav(last.lat,last.lon,p.lat,p.lon);
      if(d<bestD){bestD=d;best=i;}
    });
    serviceRoute.push(remaining.splice(best,1)[0]);
  }

  // Calculate total miles
  let totalMi=0;
  for(let i=1;i<serviceRoute.length;i++){
    totalMi+=hav(serviceRoute[i-1].lat,serviceRoute[i-1].lon,serviceRoute[i].lat,serviceRoute[i].lon);
  }
  const estTime=Math.round(serviceRoute.length*45+totalMi*3); // 45min/stop + drive

  const el=document.getElementById('svc-route-list');
  el.innerHTML='<div style="font-size:10px;color:var(--sub);margin-bottom:8px">'
    +serviceRoute.length+' stops &bull; ~'+totalMi.toFixed(1)+'mi &bull; est. '+Math.floor(estTime/60)+'h '+Math.round(estTime%60)+'m (45min/stop)</div>'
    +serviceRoute.map((p,i)=>{
      const c=customers[p.id]||{};
      return '<div class="day-stop" style="margin-bottom:6px">'
        +'<div class="stopnum">'+(i+1)+'</div>'
        +'<div style="flex:1;min-width:0">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:9px;color:var(--sub)">'+p.address+', '+p.city+'</div>'
          +'<div style="font-size:9px;color:#d97706">Next due: '+(c.next_service||'not set')+'</div>'
          +(p.phone?'<a href="tel:'+p.phone.replace(/\s/g,'')+'" style="font-size:9px;color:var(--blu)">'+p.phone+'</a>':'')
        +'</div>'
        +'<button onclick="logServiceFromCal('+p.id+')" style="font-size:9px;padding:4px 7px;border:none;border-radius:6px;background:#059669;color:#fff;cursor:pointer;font-family:inherit;flex-shrink:0">Done</button>'
        +'</div>';
    }).join('');

  document.getElementById('svc-route-map-btn').style.display='block';
  toast('Service route built: '+serviceRoute.length+' stops');
}

function openServiceMaps(){
  if(!serviceRoute.length)return;
  const url='https://www.google.com/maps/dir/'+serviceRoute
    .filter(p=>p.address)
    .map(p=>enc(p.address+', '+p.city+', FL '+p.zip))
    .join('/');
  window.location.href=url;
}

// ── SERVICE REPORTS ───────────────────────────────────────────────────────────
function renderReports(){
  const sel=document.getElementById('svc-report-client');
  if(!sel)return;
  const recurring=P.filter(p=>p.status==='customer_recurring'||p.status==='customer_once'||p.status==='customer_intro');
  sel.innerHTML='<option value="">Select a client...</option>'
    +recurring.map(p=>'<option value="'+p.id+'">'+p.name+' — '+p.city+'</option>').join('');
  document.getElementById('svc-report-preview').innerHTML='';
}

function loadReportClient(){
  const id=parseInt(document.getElementById('svc-report-client').value);
  if(!id){document.getElementById('svc-report-preview').innerHTML='';return;}
  const p=P.find(x=>x.id===id);
  if(!p)return;
  const c=customers[id]||{};

  // Pull ATP readings from last service log entry
  const entries=log[id]||[];
  const lastSvc=entries.filter(e=>e.outcome==='service_done').slice(-1)[0];
  // Pull ATP from multiple sources: log entry, atp_history, or direct customer fields
  const atpHistory=(c.atp_history||[]).slice(-1)[0];
  const atpPre=lastSvc?.atp_pre||atpHistory?.pre||c.atp_pre_last||'—';
  const atpPost=lastSvc?.atp||atpHistory?.post||c.atp_post_last||'—';
  const hasATP=(atpPre!=='—'||atpPost!=='—');

  const today=new Date();
  const dateStr=today.toLocaleDateString('en-US',{weekday:'long',year:'numeric',month:'long',day:'numeric'});
  // Use actual scheduled next_service if available, else today+60
  const nextSvcDate=c.next_service?new Date(c.next_service.replace(/-/g,'/')):new Date(today.getTime()+60*864e5);
  const nextSvc=nextSvcDate.toLocaleDateString('en-US',{month:'long',day:'numeric',year:'numeric'});
  const reportNum='PIC-'+today.getFullYear()+String(today.getMonth()+1).padStart(2,'0')+String(today.getDate()).padStart(2,'0')+'-'+id;

  const machineBrand=c.machine_brand||'Commercial Ice Machine';
  const machineModel=c.machine_model||'';
  const machineSerial=c.machine_serial||'';
  const filterReplaced=c.filter_replaced||false;
  const filterType=c.filter_type||'Everpure i2000(2) Insurice';

  // ATP result color
  function atpColor(val){
    const n=parseInt(val);
    if(isNaN(n))return '#64748b';
    if(n<=10)return '#059669';
    if(n<=30)return '#d97706';
    return '#dc2626';
  }
  function atpLabel(val){
    const n=parseInt(val);
    if(isNaN(n))return '';
    if(n<=10)return 'PASS';
    if(n<=30)return 'MARGINAL';
    return 'FAIL';
  }

  const reportHTML=
    // ── HEADER ──────────────────────────────────────────────────────────
    '<div style="display:flex;justify-content:space-between;align-items:flex-start;padding-bottom:14px;margin-bottom:14px;border-bottom:3px solid #0f1f38">'+
      '<div style="display:flex;align-items:center;gap:12px">'+
        '<img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAZAAAAEKCAIAAACzIJS+AAEAAElEQVR42uz9Z7gt2VUeCo8w56yqtdZOJ58+naVWt1IrNkKoEQogoWgyxrIRYIKNSbaxuf7AxmDjgH2NMRcHbIONAGMRjS2iIsqxFVvqVudw8tlxhao5R7g/5trt+H3289zr/tx4vT+6z3POXmuvqlU1aoR3vC+aGfxnQACHFVZYYYX/tYAIQP/V366i1QorrPC/Ivy/FbBWWGGFFf4XRVidghVWWOEJkmCtMqwVVljhCQNcBawVVljhCYNVwFphhRWeMFXhKmCtsMIKqwxrhRVWWGEVsFZYYYVVwFphhRVWWAWsFVZYYYVVwFphhRX+N8OKh7XCCiusMqwVVlhhhf+3seJhrbDCCqsMa4UVVlhhFbBWWGGFVcBaYYUVVlgFrBVWWGGFVcBaYYUVVgFrhRVWWGEVsFZYYYUVVgFrhRVWWAWsFVZYYYVVwFphhRVWWAWsFVZYYRWwVlhhhRVWAWuFFVZYYRWwVlhhhVXAWmGFFVZYBawVVlhhhVXAWmGFFVYBa4UVVlhhFbBWWGGFFVYBa4UVVlgFrBVWWGGFVcBaYYUVVlgFrBVWWGEVsFZYYYUVVgFrhRVWWGEVsFZYYYVVwFphhRVWWAWsFVZYYYVVwFphhRVWAWuFFVZYYRWwVlhhhVXAWmGFFVZYBawVVlhhhVXAWmGFFVYBa4UVVlhhFbBWWGGFFVYBa4UVVlgFrBVWWGGFVcBaYYUVVlgFrBVWWGEVsFZYYYUV/pdA+N/hIN3d3c0MABDx/8H7AIADIAI4+H/xr/WdD3/m8C8BAcHd/3vvjQD+P/IZENAB/r/88H/zTerx+v/AX/533/A/O3n/Awf1OOB/9Lz9bwX8T/BH7dDqbfxHNFCBmgJACGF1Ea/wvyFEBACZ/+gUUn9kA5aqMnN9whwcHNx1192f/vRn7rv//nvuuTfnbCohBCYGIHdDIiIGcERHREIwB3AMgQGpphJEzEzErCLuxswxBABULe6OgKKiYiFwTNEBak5nZu5ARCFEdyNCMwMHIkZCMwcANwNEQCek+lA0BzB3AAREQjcFRARUM1VhIg5cn5/uy/zRzQjZAQC9PlXNnYjrPxLiMi8EckBTA3BEDMxFBncDQEIiqj+G4A6AiATgSAjuZqZmKmoO7k7kTIGYicjBCRwQmchM1SwQE5Go5lIQiTkwIyIyMSISkzsQkrm5uddstZ52MzVzMHB3B3MlIEdnDkyMQGaGiEToDiImosSYYgI0tcLEzBERl1kXArirWT10JBIRd2cmIgQAImJiBzRzNzV3AFcVU0FkBARENYVlTuuqhTmFkJjQXb1m1A6AyIHd3N0PE20XKYhAhMxBVN2XiS0TAbiDuYM7ICAAIqBoISQgQABABvAQAhHp8rU1mTRwRwRAdDNzY2QKgYiYMMTYNKlpmqc97Wk33HD9k2+6KTADQL0OmXkVsP4XLQDrtWhmv/s7v/umN/3y+z/w0XNnr2TxGNumaZAMEAJ3gzgDxpiKDCmGYRiICcFn/SIiEQUpGYlEBnNiZpGBOQJBE1ORwkTm4IhMuuhzYkYMzKxuWobxZN1EinpkEjMT4cgl9+4UmKSIOzoomCEF1QVz5w6pibksmtS6Q2A2dwAqWtBdVUNIgMjIQ16Yq5uHGBEhxJZCKH3ftKPFYo4GgBaYSpGm6UoRRsta1ADMmrZBQKQIhOxFVNyRQ0ohmXvRQUXaplUtIkaEISZG6vMQQxKVyDCfL5iDaBmPWlEXcQBoU5gu5ogOwDE0gKAqABhDcMecMyfSIjGEYShN15paDDRfzFNKRXKgYGbgFkNk5qEsVIQ4MTGQgXts1koZEDDnvmlalULIBpTaBK6unrVETiqGaG5CHIsWcs+lIALHFkxDTKaKjEzs7v0wC9xiiIzAiICobqA5lz5wg5TchZkQMTVj11KkT3GkDlJ6MylDbrpxydnMHAwRiDhwAxTcFCnIsCgyA8CmXTczU0EEDBSIFv0ixARmIbCohtAAeKAoWvph3jbrpQwhsqohmkphDiVnDklNArOUAZBCaBlNzFV6jiN0FekBMkDo2nDDjdfefvsXfv3Xf93LXvaSGrb+H7ZEVgHr/3089iT5t//2zT/5k//0Qx/65MbmmVNX3XD6zDVHjx7nEPMwb8ax68ZerCiISGQOKZbcI2PTRHDMxR0QRFNgRFgMCwJoujbnHtzb8SSEQMTzxZwpNqmVknsxUx11rQFOZweSh9FojYisJjOqkREDLRYLwhAC1yjmACIaY9NP9w2AQ2NlwMjj0dhKTomZue+H2CTNWd2QEhEy0/TgwMwcqIncNgk4FHECR4RciqnFFAgBAUSsbVqwXNRyMXAI7MzBkYpoJGy6URHputQmUsOD6SIPuUkJ0fu+70YjM2FCNTcAAG8jSynIkQJJLiJaiqrq+qTri0wP+hAnMVBKEAKrZA7MhGaogCoamRZ9diRCjxz7YaFugFTTRqaQ+z42jZkMw8CUYuQQWYuEwCFQzrmIuFvbdeTQdF0u6uboaET9kFMMZeg5UM6DiBJQHnonpsCROXBQM2SMxOawGHo3iDGaewzBXVMb82KYzXMusrW5WYYFhUAcmtTkxTw27XR2EGLjgAgEiCo5JZaSh35OIbbtyIxCDIvFPDIx43w2FbXRqCNEN6MQYoOgUIqLCIXoKm03cnACA9O2G5Uii/liY2Mym/cGSOhogkB9LmIQY0gpmrmooXtqGAHniwU4tG07m877flDpy7B33733PvTgPfP55Ze85Iu+53u/+yu/4vUAIKLM/MRt/P2RCliqGkI4f/7Ct3/7d/z7f//bG5tPvuUZzz12/Kr5fDGU/mBvO/fz2XRnf+/ceG0rcuzns248NuemSVcunbvqmuuuXHr0mmufdP7C+Y31jflsPplsTKcHiNg0jZoiB1UD1/F4cjCbcQgmEmMCgL3dy8eOnXSDg4MDtZJSWsymJ8/cuHPlwnh9kxAAfH/3coxJRZu2maxtXbl84fjpM9OD/en+Lri6+2Tj+DCbUgxuGpmIQ2q6xXw6n+1vbB1TMXNbW1vrF4vFYq6S27ZbzA+OHD0xXyxyv1jf2Njd3Zusb073d1PTxLZBwH6Rt44e3du+1I7GiFxKURUiAkQpOcXmYH9/NB679oG5GW8c7O+NunEuJcW4t7c7Go+GkttmNPR9Ox7lYQ7uSAQAbZN2d/bG4/X5dL+bjMuwaEfj/d1dAO661kxyv+hGbS6CSKrAIbr72trmbDFflnWm8+l+0yWkAI6m0nXd9OBgNF7L0qfU5X4IKalhCAF06MaT2fRAZWBEUe3GG7PZlCiEEMowD6HhQAS+GIYYG1OJqcnDXMSQeNS1ZZgjBREZTyZSVE0dwFSYExBpyTn3Gxvr04Mpcurn+5G5aUdFJDaNlZ44ORIiDYv9rWOn9/d2KcTFdO/osWPz6b6pNt0YADh0s+kU0FwlBDb3rh3NDnbb8Zi5EZGtI2v7u9upmcznMwBHjjGmPCxcB6aAxMQxL2ZrG5uz+dB2HZiqZAdcDD0gMnNKrZu5ZSIueUAEdB/6eTfaPDjYPXrsqo3NY2tr6yF2KY0unH/kU5943+7OI6997at/5md+6vTpk6VICE/Q8hD/6ASsGq0+9alPfcVXfNV99z30rOd++cbmVXsHu3s7l/rh4PzDnwAYRuOjqRnfcNNNB3tXFrPFyWtuGBbTrunOnzt74tTpnd29jY21R+6//8z1N6BDKcNivhNim5rJsOh3d85dfe11DhGRdneujEaj6cH0zDXX7+xemU8Pzpy5am9/hu6APptNY2xUcx76E1ffKEM/mkyuXLgYEsfYNk26eP48Et7yjGc8cO89KXJRdi0bW0e2L50HhMBxfXNzPt2fTqcmebx+JKR2frDnYKYaQzzMfSai2rRjkTyfH5y59sZ+yGVY7G7v3HDT02cHe7nf37188Zrrb3j4oQeOnzwpiqoaOcxm07WNdRXZOnr84rmzqWtmeztHjp1a3zpx7pF7S5HUTgLBbLrbjkaXzz968vQ121d2Tp06fvHCxSc/9Rmf+/THzlxzHQDs7Bxcd/2Nn//cp06duXp/d7fp2u1LF9a2tkrRxXRGjLHpCHDoF4ic2mZtba3tuksXL2gZNjY21HE2nTLT7u7B1dfdcOn8o+O1zdn+lZOnz1y+cK5pm/Fk89jJ01cuXRTJJZcjR4/t7u6ZmQz9sVOnKLYXHn2QyZlpPNmY7m4v+n5t/WgMUcp8vpjHGFJq5v1iPFlfzIbYjAhtd/tCUd3YPBoC7+3sMMFk4ygABMLty5c2jx4rpUTG3Z2d01dfN53ODnZ3r73xuu0rVwwCmi4WUwDvRmt5GJquW8ynZ645c98990wm6ypy/NSZKxfPt13XdePF7CCkpCr9Yt61bTELMc6m/ahrh7wA8IO9g82jxxYH+9ODnfHaliO1Dc+nB4RYpBw9cWYxm25sbDz84AM3PPkpF85fVBuaJuVcci7uNlnbSCkNQ/Eyz0Nux2tNk3auXFR1pDDb39vf2+66jVNnbrnhSc9ox+OD/d33/+FbT59ofunf/Mvbb/+iIhqWLS1fBaz/v1WCd9zx8Ve84tU7O8Otz/nSzSPHH374/p3t8zuXPt926ZZnveB5L/qym5/zQubm6KnjQ3YTSaNxXvRDdpPCbbM/zyGyzg6OnTgxFIxkDWlffL0Lfc7oQ4ipWBMY+8UsxjhflPGoK6VktW4ynhUH0SZCX4qLmimAj9cms5m0bSTNKQFxq07nL+2ePta2bbe736OX/VmetFHVixYm4hApRXRzzUUAORrEPJ91bRRRdWfwFBPGBrxEpkUpQ5bxeG2tS7PFdN6XMye2ZotFMZjuTk8e3+yzGBK6gdnmOF3ZnRWKLppCwNJD2+YhT9bWNsZNycO8zwYBHfb299bWJ8PQn9zq9qb92vr4/OXZ8WNHZ7NZSJyLRUAxHYZZ240io5nv7c2JGwBIBLN+0UzGjJzYhlzMoUnJzA+ms/VxiimJmpRCIVkZmrZb9BlCZFdGYPD5MKj5xsZaG3C+GAaFpkmahz4LUcAQm6bZ3TvYXEtqGAmDi3pxpBBacF30pWuiAw3F1yYpoB1kB7fLewdHN8e9UorQBJ5Oe0YmpFyKilBIFLgM2bRMNjYJaXt77+pTG+cv71Fo83weEhXx1HZkauCEfOrE2vb2PiIl9qyhH4Ym0uakMVMxj02cz8Xcx21XyuBmbSIg6guA26hLpZQh9wsNTYxdi3l+4M59No5NQDu+2Rzs73WTdXRfDGKUSjEnnB7MjmyMAAjcUQuiArWEPiwOjCOA729fuOvTH/vkB9/16Q+9p9+bXn/Tc256xgvm84PPfvzD25cf/Nmf/Qff8A1fJyLM4YkWs/5IBCwzI6Kzj579ohe95PKVfNvzvqQZjc5e3H74/o9Odx957otf+2Vf/20nr7tlfW19Z3c2G3LJi4uXDwy5jb41CheuDL14Sn7q+Ob2tPQZh6HPw/CMm05vT8ul/f70kbYUPZiXvs9njk6Ob40evTx99PLs2PGtEYNKyerTrONxCkjzvozamPt8bKO9tNcPeTi12R3Z2Hjo8oytj8xt5Ed3C7OVnI9vjfenQz/M2xSaFHfmpY0MDrEJbdNOFzZdDEfW2hTo0UuzfhiObXTOsS9gRY6uj05vxbMX9/ZncwA8c2IdAR+9MgUOk0kzXwxFVA2uPb4+6uID53Y2uhAIiWD7YLE/L5MmsNligL7k48e31rumiE4HDRw2xuncxW0mmIxSShEAHHE6n6+Nu/NXZg5uYKO22xi3Fy7tj7qIACePrt//4AUCHHVdm3go0hcNTbzq2Nr23v7+PKcY6xBz92Dats3aqGOC2aBg1jWxi7g7H+aDdE04sdb1Q37o0s7pY+tmHkMYREUpBlwMi/1p74Cnj26I+F4vCnD6yLgN8PCF/bVJs9Y1o0T9Yri414+auDlpZ4PtzuYnNmL2ZrHoZ4uMRGdObFzY3icdYkxtDG2kRy9NHfHYxqgJeO7SwcZaB2iR42zwrCJObRP29+ccIKa4OWpioIs7Mzdfm7STUbq8fbDR4LhthwwcedHno5tjIN456LWUSRu7FHenPaJurE/G425796CYjdowbsL+olw60LW2OboRHr54kIuPu7ZNPEowny9S07Upjho8uz1D8CbGSdecu3KgWroUj210Z6/MJqOWdJh0aTHo9lzHXXtia7K+MRlyX+bb733LL37gt3+1n8+f+pyXnT59/cc/9L67P/feX//1X/jKr3y9iB0yHnwVsB7vgPW613zNW373D57//JcdP3XDI2cfvf+eD3OQr/+uv/GMF7++7/vPf/4+6Renjm485Zq1J1+1dt2Jyf4sd4lTwBDCwXwYdzEFJo77s1xKGYbh+NF1A5r1OSIORYdikXHSJQTrxbNC7SZnsUBgQCK6tdaIqhq46aSN016YadIwAu7Oc0Az9ybG2aBqrqpdE4hZzUw9xTCogSkAxkCElMUWg4zaGAl2ZwXQm8gpUC8+ZFkbNS373mwABFEYNexApYiaO2KXeJGljYSA464Rczd1d1GrpysyqmpRNLcUQxPjfJCsFmNIhIt+SIGQMEUWBXPLRTbHzcGiiCpU3gCiqgMAgY27Zp5Vi3CM4FaKhcBiutaFIasjBOJclAkd3AGYAhHlXAxcDQkxBhRVcFvrWnNYDLmNnNWYmBnMCQCYScxLzuOu6bMRQi86ipgCmQMzijqCmUFWb2Iws6aJ834A9xQbUSXwIev6uFnkDCbqWIqNurQYBMFTQCYyIERXczBQM3EghBTY3MW0axs3TzEUlcUid21aEhtMu7ZZDNI1vMiamBBxKCqqidGBBtGuIVWIKbq7maPbuI1XDvpKsljreHdewKGNwQEC2bwvxMkdm4hZVEWIGYCKSmBGhFETsiiYt4myyHjU7uzn+84e3Hfu4MGLs7N7et0Np46cODa/8sgv//j33/vx93zxK964Pl5751vfMiwe/ciH33PzLU9RNSJaBazHu3X18//659/4Td9x0y1ffOa6G7f3Fg/c87FJK6/+tr99zbO/6MKjD52/fHD9qe5Ln3X6i55xer2LQ9/v7u2ZqqpWmlVKUbRSiDwwlTJ0XbekBjmIiKoQIofg7iLLPyNxzplw+XPEwdxjjPVuzkUIAAlzFjWNgdWMmc2MEId+ICZiRsQQIgKIlDp1NjVzr21pcBM1ZiZwB2cO9eXETMz9kJlQzQJTEXO3EJYErSGXSkNrm2SV5USUc1EpMSVwsMpOAidiByi5ADiCI2IlCHEIzOwAuGQBIQBy/VgIIiKl4CG1x82QkDiYailFVZumcTM1I6KmaYZcqDL+65sR4iH1y1TMHYk0S2qaGg2JSdXqjeTmIbD/J8sF7o7gauZmjgTuTAyEpqZ1eA+ORIGDaAFAVS8lh8AAyIyIFGNQdUR0s1IkhFCJUWYWY6zDf1V1d1OtjDlECBzMrPK5AjMHElEAjDG6u7kTYhEBcF2SuiyEUI8CAcw9BFYVQiJmBAQiNyciRHQAdCPiIiJS1DxwIHREMndiWnLTzBCAiOq5ZSJzd3cEAMQmpW7UcUyX9/p3fOLcH959sL3wG687trXWvuWf/thd73vr8174qn62/+53/MYXv+gLf+/3fwPqVf7EoTo8sQNW5ekdHBw877kvunBp9pznf6lAPHv2nkc+/+7XffvfueWlX7t96ewj53de9qwT3/CSJ53caB548OGzjzxi5g7QjUabmxuiquZmioCj0Sjn3LXtbD7FOmlHyqWYW9u2poZEplpKZuJ2NBYpi/ncwcfjcR6G+t2PRqN68THzfDYXkRRDiBEQAUBFzS2GKCJEmHMREVHpulGMsbJN81AAYBj6EGIIwczn81lMqWtbVTEzd5+srUkpSFSKpCbt7+0zUx4GIuYQRqORiMSUCHE2O0ipQWIppevaXIqINKlRlcV8llJqmnY2m9VEdWNzAxxyzovFHBFTSkREzF3X5SH3iwURp6YRkSULkQjMFos5BzZzKaVtWzU1UXOPKZVh4BBSaphJVXPOo/GIKQx5AHcDJ8RDJgrO53NVqb83xGgiIjoeT5BwGHozczMzZ0Ywa9qWKGSR+WwaY0SklFIpUk/sxsbGYjEvpRBxCDEPg5qORuPdvZ2ubU0lpraIxJhECnNo2max6PvFLKWEwG2bAKAUUZGQoorWcIMITKRmDpVDMKvHy8Q5D4jYNG2IwcykFERU1dQ0Uoq7E7FoMVFmVvfAHEJcLBZN24y6kburqalWYq5Imc5mk7X1wFwjpgPknBFpbTKpJNi+nxNRyZmZCZGY+r5fzBcAjIjXXHPViZPHz+/O//Fv3/PRe7affvOJzcnRX/t7f/6hz95x+5d+zV2f+OhnP/O2f/EzP/Onv+2NpWgITxiiwxM7YNX06l/93Ju++Vu+5Uk3vejMtU+7eOniQw+89+ZnfuEL/tSPqQ7z6d4XP+Oqb/myJ83nw/vf854Q8MSJEydPn05Nm1KKkUsWB0fEJqWc64WFHLgURYBSStM0ogIAIURErExsRFDRENi9EsGXRRIRFykESISH99eS+kyETMyBpKi7xxjMwcwIoYgikZuFwETk7iKKiKUIETKRu5sDE5oZMal65SiCW10bqxuMFMjUOAR0EJWaxCwp1eCB2QHcjRDNwc2JSUTAHZEAgAO7m7sjkpm7G4AjESG6V9K4A6KZElJ9LjuAL48RKr+8/kyMoeQCiKbKIYA7MaoYBzpkdWvdssT6Pu5IiEiSBQhMrfL4ze3wJDAAEFMl9RNSDcGVZ49IbgYIZp5SAABT58AiSoT1fDKzioYY3E1Fl8mdeeWFLgn3NcV1r8mLL1npUEoJISLC0A+IGAKHmKRkBzDTGJOp1vSKA3s9y+ZI6O7MXIq4WWpSvWLNvG0bVRURImLmkkvdMUBAdw+B+2FAhBCiFHE3RGRmJAIHDqG+NgY2t3oSicjdACDnIqVcvHhxb3c3xubpz7jFEX/iNz732XOzJ1+1UUT/+Q9++7VXndncPP7B979rcw0+8YkPrG+sucMTJcniH/7hH35Cl4RE+EM/9DfuueeRG296Vmq77e1H+/nll33zj7abx+ez2XOuHX/na2/ph/yZT30mBHrebV9w4tTJFKOpuZmUUivBukYjIqoKgKaG7mbOIdSdlFKKqYoIIiChqSKBu5eS66vM3NTdalmhOZeSs6rWULJcCgEouYioA0gRM0eAWmKAO4CLiKnnXADcVGvhZuo1Z6mXbCn1X+vWj5VcpAgAmKkUUTUT64dBVEXNzOvqtZu718VtkCLLcsXc1BBJVXORUnKNrfXNAbzGK1MtInB4iuqeR8nFzLQ2sQDMlqvQtQwspZiYinBkFa074GYuovVsE6GaVd61iJQiUmqyq2aGsEzHArObMxEREVPJuRSpoUSK1DDnajVsihiAi+jQ51oimR/+GEA9H4QoRUKMZpXSi17TXlNCFFEpYm7gWEO5iNSSzdzKkAEREN0hD0NtBSKiiqqIufsyeruqErO7uZnK8hqoSWI9kJwHU5ciiChS3IGYa7dBREspDq6qKkLEIbI7SCn1eshD7+bggIhqdadIVY1roRpDk9LG5uaRo0fB9eLFK+vr68+/+fjnHtk34tHWsTyUz33wd06ceUpqR3fd+f6bnnzTc5/3bFUlemLsGz6BtyLr4+v82XMf/OCHu/Hx9Y0jqWmmuxduvPXF1z3ttmG6HxG+8vYbiPDSpcunrzr1vNuejwj7ewc5Z0Jk5toKMTdRESlIVEnANeevUUNU3C3GgITmy+tDRFVMirpDjKl2i3x5hSo4MAcHIKIQQ4gBCWII9Snatk1gJiLw5Q3vajWAMrGa1uzs8KHnZhpjCJEd3JbdVkRCQFcVZgqB69O1LsqpG9FhlDIrRaQUB0cENTWrrG+uyXXNWWrbtR6jG6hZXWo005yzWe28mLs7QCml5AJes7R6SuqvWrZR6vcSm9h0jTsQk6iYQc1NiAmJRNTNETwPg6txIA7MjCEGYnZ0VUlNElVidHdRqdlNSpFDUDN3r/8ECKqWc1mmGojMDAi5lDz0ZuKmNfQCgoPXZISIAAzASimAICI5Z0CoubOpItVszlW1xnuOoZ5GKQUBVE1Ea5Ctf1/jl6nW58cyd3NXVSkFkdy9SJFSzGq0l7puqmbDMJSSzSwEjjEE5sAsIqUMs9mi5ExL5QVDwhqmzI2Y8HAp0tyQSdWKyDAMiHDsxMlR116+tL02Sq993smHHt3fvrx928te2W1sPfzQ/UePH+/GV/38m37l8GzAKmD9Tw9YAHDHJz55+cq5I0dOEPGlC4/ODi5f98wX98PgqrffcuSG02t7B/PFfD5Zm6hqyaVtU+1hiYiqIRLVux+w5NIPvaiAe5ECCDVjdwBVVbXalTDzEEJVgCBiKWIGag643DtGQlVBJCIytTxkU/flzjDmkgGciJAQCQmBmYgJid09hADu4Ia4rIw4cN/nWrIsP3kpUgQcEAgRVQ0cSln2lbC2Uelw4Rjc3Gp+ZKpu5gaqhoCEVNMwVWViXBaaFAJDbSETE6CaStG6BU1EiBRTIg5Y2/GItTntpocN+BrU8tAPbq6iy1zOrFTkIqXUlk0IIcRQf6AUzbkAoJbSpIQAUEtXIkIKMQCAubsZI4bAtsyOl6GQiRBgWeUBAECsnAx3M8vDEIhUpJbwplqb2SGwmQOgA+Y+D0NWEUev+9NEWM+wFu0XPTiYaAiBAoe6QY3LOvKxjioxu0M9cCmChDFGYq5PNUICrL1ybpoWEN2csNaeaGYiddtJHLAbjZqm7bqWmWu8W55ItzoNcHUTdQcRGfoiubi71cGfI4CfOHXKVXa39555w9ZNJ8j62WTr+I3Pedne5UcBwomTN7zv/R+4887P1cXbVcB6PALWJz7xSQBp2pYYzEF0zqPN2Wxw1WdcO1b1Rx5+pGlSiqHkUnffH8s+Qoy1r1yfXoiY+1xFEYjIzaVIGbKq6uFN5+a1paqmiChFkEBVmJCQ1FRE81BMDamGDkKiPAzzRV+KiqiZlWVh5QBoDkWW9aaqDf2gokRcdbVqWAyBTdystso8NYmptpm8FFnW9sxmy9X/Q80AjDFyCIHZHczBHazGCdPlrycCAFUrpYgu67WlVpZ5TQBrTOEQmAIzxxSJyZb9bzZzRKrlqpqpKhEGDjHGWskhIhKJmhSRouBVSCMwsS6TiwxAiESEzOxmZpBLqeeKiACphoD6DPBllY3uUKfytfRFIoQaUjHGGEOoz4n6qtQ07lD701UPoTYA651fG/+1VdSNOsJlr20pp+GOhHXCq2790A/9QIGZmZAJl4cZOCBi1ZNgJlUjJjfPdTvajJk5BASgGrqI6kHZMnU3YgohxhibpmFCd3eHWjly5MNmKIG7iph5zrk29QAAaan9gDV/R6w/QCEMuaQYb3/qUcBAIR09c20/LFJsuzaVsvvWP3gn/K+ibvbfxxNeKOrhh88CcAgRDFXK+uapI8dP9sPi+Fp48pmNftGXUk4cP1oHy/UZVcVJzN1F5rlw4MrxSWlZW9UWci06iBkQTd2WGUQV6HNVRajteUGkIgruRDAapa5rAWB2MHcAJOi6JoaRIyGQlNz3eRhyASFC0+V9UqVDikiIwc1N1cEJSVXMnarMCyK4m5oUQwAOXHvbVQSmxlQxCTGGEFRtWSGqImKMDIi1RbtM9hitdpQAECGmYGoObiJYNVVq2wlJREJgOezFmHmI3KRUGVWlZERwRTOrc0lTdzAVAaxdZjVTRAohLFu76ObmYmbGgUMIiKDmRCSlmBox1SIFEVUE0MCdAyOCm4kIhyCiUoqoxBgqqcJUK03Ba+boNUAQEYooIomUEIL7skGOhFrLN1VHQASOwVRKKTUKh0A1CRYRUcXD3jbHVNMuR19Gzyr2YlrrwToz9eW2ONYeJdfDFK0RDRCYKKbQL3rmoFr7p6Ao9XmDiGbOgeoVYpXyxlQbWLUXpiqqnlJSMxNVMQAgRCd8LO9b31yf7u27+emj45wvLaYHV585MRqPAiNyBOAPf/gj8IRRccAncMBCRHd78IEHAeJiPs85x4Dz+ezixSvHxydGI2bCXPLJkyc2tjZrBwQclmufDgZASBiRGAlJ1YahhBBCYFGrlch/OkgNIdaCi5ge63T64RNvPG5ioOm0v/NzD33sY3fdccfd9953DsCGIYfI1117dHOjPXPV6Rfc9swbbjyzubnRNImZc859P1TpKTNjJnfAmlQEqiN/WjbmoZSSUjKwxETEiCBFEQlwKZZCIZQiKipiTPXltZnqlV6EAcwdHZYiWjWtILLaG653F6K7V8IRp2TulrUUCcxFpQraqNiszEMMWLeBiep9UlvOhGgGNQurHIVa5ohIndBziEwM5HX+VZkKbqbuzKH2jmpuxYHrDPGx4EvENfGJkd1iZcaFFP1QFQyXXxqAo4rUDLFm0N2oO5xjqFdpKiY3d2ZwH/ohpUQhVqpZHZKYqIIxMwJwDIg1m3MpJcbo5rVvUAlQtYQ0cyQC0xhTnWwguJuLFCL2w25ALbpBrWlbEYkx1QSOaKl0pqoh0uHMxDggEzs4EBBzzcfr07fkgkQhxJQICVWEmUQUHFSUCEMIecgnjq51DfZFe0+LRd7Z2Y5phJgeeeQhM3+iyM48wTMshyIFMcbUlJJVc0wtMaW2KVBHXba/v4eIo66twyZVM7MQQjxkA0pWd4GlZh2burshYH1uO0AMoZRS766q6FayVOJP1yVi3ts9uONj9//huz/zzj/86Gc+++CVy5eJIwCUvECiEMIHP3AXIKqUrg3HjqRnPPWG229/3tNvffLNN99w4sSRbjQqQxlyNqR6o5qqiztRHSFVamtMZOqVqSSiTASIDo7uKgZLWT4IMRCRmiP4ss3k4A41JaxC0SZaJ4eAAGJ12sD8mMRzbXK5iPpyFgYAGDiIi5RMxDEEc6g1FDggLDUAa0SojAE3N9BacFWuIxED1ogm9aG+TDdqw2hJY3a3Q1qDKiABQiVqHaa9VItxomX966q1LDWzygyosth0uHcSmAEx5wyHAtamTsyA4AjMbGqpSVBzIvBStJ612qoHdDG1AsxMhAhYpfWWMoMATKS65FHFGOpJrM2HKh9YP5FoWZaW6rQkzWqdctZHSC2iS8nuYKqAUH9RjcG+/G1VrBHNjQhjDPUc1k5F7an5MhwuLwYzNfeubbqGMSUHphDbbjI9OEBqHn307P7+/ubmxuF3sQpY//OSLELm4K6BuRuNL19RAN/YWO+HwVAJPRcZjyddNwIwIq4LyVSvIAdmqvWFmVddRhExNCIyMEQkX1Y1APXJCUwUA0/GnaqefeTinZ+95/0fuOt97//8vfc+MpsdxKaNTXDd67O1Ma9PaBh8Pt1zYKRxjBFwbXcfPvCRh9/7oUdTx0c3x0+68eSzbn3S8557y7Oe9ZSTp46GmESkX+QaWCvn4PDGpqp4W/OFUqRSlIm5pjbgECIxB0BAE4Bawvjh7BsAcVlw1VSuZgjgtb0NjzWvVAGBQwBAQ1gSKLRUZnoKsZZUJpVJBrXEBoDKtjVEWL4v+pImtGQYceCSMwasbAsiRgDipW4qIqoUAK8fDxCrYmrNYR/j2yORqeGS1ELurmocCNwRkLm2hDQwmYObAVXWhdZPhURDPzCTWf1kj7HA6sATiCqNbtkZrLT1mqYR1WrWH+tVOTggmRsC1Qyr9sSRqM45arCvo0YAQCcz4dpVFAUANQ0U6jdb+3FEXOvrZaR2+E864rUrtxw+mi1bEzVAhxirpG3lyqgaqgKAiNajSBwMfTRed/PcT0Xd1EvOIvJEyVD+CIidY22KSBkCwWQySW2nQLu7u6WUmFIuZioAwKH2RNTcGdwMEgeqjG1f3iDLVRNCrGxKBK3lIeN43MTUDP3s4YfPf+D9n/3Ahz738Ts+//CjF7MU0wUS7B48art5fZK+8LanftmX3f6CFz739KkT04OD3b39hx56+GN3fPJTn/rMvZ+//+LlA/PR2sZ1HNcuX9l/9Nz2O999Zwy/dfU1V99y0zUvvv2pX/KS51x//VWpGQH4YtZXyvVSchcQGcEdCWMKlegM4ESIyJXUapbrMDGGcKi/ixxYxZgYHGrJ4G7g4LakUC67GAgqtpTfVas3gIggojMwVzktrdK+RFgLUgSsrEt3Z2Q77AOGwHU+WcucmnEQYmX9VBYbEavV2QKKaNt1wzDURcWaZ7mbFGVmIqoDw2UW5lCZHIQ1vhgz1YGJFDlc/EEHQEJ77DARvPbRK8fBjAP5skFGHFBF8TDE1lCI6My1uek1wh8mMeAAZShIklLKMqAjgBt4CIEBrOpZ185dzePsMPiaqdUnjbtZ0fKY5jURlVIq31WXtDUkQu2FQ4gpmNQnWXUqwMe+uACwJKlVtonoY63ASl4BADVkpMvnHpFh0XYTuXgZsEHiJ9BqzhM+YNXrdrFYmEFKTRG/sr03ObmVAscQQgw7O/td2wBAv1jUgU698NS0lEpFcApL1njtmPjyarMYuG1CCDgMcscdd73jXZ/8w3d9+qGHdi9dvhBTAJfZ/NLBwRXJ+zfccOqVX/clX/7KlzznWc94yk1PotiAq4oiIoUIAN8C4GZnz1345Mc/+a53/eGHPnLH/ffdc/bcPsAoxDF06w8/un3v/Vf+/W9/YGPjl575zCe/8Auf9rSnXvvUm6+56qqTx49vqnrfDyWX3GdijikioKoRh5qb1Fygjg6JyA8fyw6AACqGiEjASIih1rbggMsSB+qjvo4Rkarsug9Dn1LCw3NSSkFE5qimMUYVrZVOvYcpcGVgP5Z61GBW91WrmjuAc6glTGUtcWVj1DBR6aDMVKmhtY4i4sPt3GWCE5i1qncym4KZE4GbqxsxMRGEqlBMWpdazM2dDskcgNC0rZm6w2GwdrDa8BFEcl8SNUIIIVBdWqp5jZkhk1ndteRKzqqVbAzLPlodQy53jJpGzZiIm6r4LgBA1RLFvKqzt02bS3GzkIKKiRkREjN4jXFavyyvNa8aIjCzg4FBXfaq8+jUJHdnDsu54SEFhwhMTNUIDcFS8OPHJt1kPF7riAncmOOhi9IqYD0O6ZU7AAVO7jabzVWka2I3GunBZQAIIayvT0S1Sal+38TVdMBiCLUNr6YhMBKruagxUYzctQmQti/vfuSjd7/73Z/6wIfu+ew9F65cvpIYEfJsvjP0Z5vGzpw++hWvfemXfukXf/GXvPiaq08DUi5ycb+/OO0f2dHdAUZNQDvYaPxoB0cn4cjmxqte88pXveaVKvLIo+fuu/+Bj3zkEx/96B2f++yD9z3w6aEHEVgsxpcub7/rDz+2trE1avi6a7a+8AVPf/aznnTbFzzzzFUnY6AiJWfNWSqfqz5F65ZGDQRq6u6Bo9ckyvwxUpIUfaxnVFMnCgEca+GIhE3baOXxm6NTvVeXNVGAuoVbM5r6YCckMzUzMPU6x6hkLaLKqj9cWfYQuJQioiGEx4R6mZmZ3F1UA7ObIeBjzZT6ShFBqpS1uvQCladqrkRobuR0uB4ONUhBfTBBXSHAynfgZdfM3Xxpw8Fc6/1aVD1WIS5LNjM8tMyq5d6yZQ6wDNYApqbuHBhM67iWKSydRB5j/wO6Wm1I/UdmMno1NJkvFsxExEUUARGRQwA3UaksjZrl1dOy9CtAREJHqAMHAAgczK123Eyt5BJSqOoOS5YhuAEq8GzwCzsa0rjSWji4uT5ROA1/FDIsIgYwKYVjY3WRgTkXXW8aB5vP59vbe0eObKracqvGLTBD4PrAdHAizqWkmFLktusA7MqV/Xe+6yPvfNcn3vueO8+f3+sXQiSLYerDhd4Xmxtrt3/Zc17xim/9wi987g3XXbe5tQEA09ni7ke2z+3Kgzs6s3iwkJ1ZiYFylpRCMWDwrTGvBz++tnf1Vrr2aHP65InrrrvmpS/5YgDY3du/667P3/P5B971znd/7KMfv+/eh3d2F2V+dV47vXt5585PPyQ2PXrs6M1PvvZ5z7vlVa9+0S23XL91ZN3dp9O5m9exd2WJLmmjiR9rWtc619RLKZVfibgsH6p5TG2vMDMz181KPKRBLjvBTDVGHKYbXhcJiakmdVxptMtuDvqyd+5ICAa1/FJR5lCHrcwEh3O6+oeaQakauBNTiFFFKnPdmZatdPSl3AXhYT3ldUJqqrW7fLjnSAAeAy/pZvVAAvlyIKAcAld+rFZ3IgdCVK9cjmWXm5YU1MP7Get6IADGFOufEUlNYwx1rFHLTCeEw9KRiKFGwOXe5XKhB5bkVmRiZlbTyjt2d1QF8Jpk1e2IuvhZ47eZ+2GlX/9bsqgoEoXAiBBjqOy3yhM2RCROKQ3qQy5dKi2VxXxvfjBDN1XJQ69qq4D1eNJH6/MUA6fpwd58dtCtDRwhMBXxI0eOpMQlS51elSxMSkwhBGJazPsY+ciRTXC5eGnnN37zk5/6zEPvfs8nH3rovIqo7Pf9/rDYceifctOZl7/81S99yYue9vSn3njDdcg09LI/W3zi3u0rA+72cP7ApjOZ9jLqAN3WW4iRoEv9MDREMdB8kAcvlXYnwkN2YiseG+Wr1voTEzi9xsc32hfc9uwXfMHz3vCGr57NZvfde/9HPvKJ3/+Dd3/4Q58+d37vYEbj8fqVK9vvvXz5rW9/zz/5Z2++8fozL37xs1/60tue//xnHjm6ubO9HwIDIiDXllIlWJlaamLlZFWqgaiiwuHWJHCI1SmmbnhgkboCzUS45IuTualoTZxUXYoQERCqiAss9/yYRLSoLAVkEBiw0peYqO0aUyVmZioilY1dE6ucK/tLETE1iasYi0MRqZP/olaKuDkHrntIUhTQiagUXSY+7kRYS31DD1iVfGgY8pId7lZrWOa6A8SHazfL5n0NPcR1ugqqWsQAnKnSC2Il8dZPjnXUQShFlvQ0r/ZFWItyAtTHkkQ1cy8m9SlATtXnzQGwDnaobqTXKtiYmZbk1ppSwXK5R8UNDkVmEJBVDRBdyrKg8+UEuea2h7NaCxx2ptta8vqxk6XkEeHBdN9FJmtrbduBozksDUZWAet/frSy5Y1HAcBjoK5rEjuDp7BcTNnf25usjSfjUe3vhBArBVlU2xCPHFnb3tl/61s/9Cu/+q5Pffrs/fddIWCKi6Hf3t15xGH69Kff+MUvevXLXvbil73sJRsb6yC6KPLghb0rc9hewPa8PLonuzMZtXHUMKPGQKY+X/QIEJXN1N0jQxEzp5NbLQGoo5T+/kv+yBWIARr2IyM+vRlPT2C98VNHxs+89WnPvPUZ3/wtb7h48cIdH/vE29/xvre944P33nP/dDZsbl0f0/rZi+Xnf/E9P/umt9147bHv/Z4//pVf+dK+L0vJFwcidlw2WaQIIGxuTsChLnkzIfEIqaY2bqZuZuohsoqqeq73ISgS1vXjUCeP4OvrXR0oujlgIsbFIpesNRZ0bWBeZnQq5p4qNWQ+H2jJ9loWdyGFuuvTNhGwjjChjsnm80GsWjTCuEtIWIrU2V8IJFmg4doWP2QPVRoKmBkxA+BsOgAhiK+td0QBYHk3ioiZV1lkd+UQYgh1vTm10c1LLpV2Nxm3IbLXJrobExfRfpGZGXmZ2NQe05LWwIc2kZVoAlC1MUrOzLS5MXazWiTWae9hkqv9onCMZl4NG5c9SF+qkpl5XXTnQONRW1ljokJEog4AaekDhkiUh6JmKksWBRIxIjPHFLaObLkqmB49stVMRhsbW7Edl5Jj04U4hieS9xc+0QNW5bkQRXL3frGYrB1Zn0y4S6NO67c+Go8mk0nVJKoj3hAY3Maj9s477/n7/+AXP3v3ztnzg8oC0XJ/fjq9gjw7c/roq1/9pa977ctf8pLbjxw5igjT+fDQpfnD23J2Bned73emQxMoRgoEo45NZDGACBO4uRGzukvJDshU+0pggCJQxLqOFn1JEfMwOIS9OW4vyucvLRBo0qXN8fSaDbp2E0+sxY3NrVd++Ste+eWvmM6md955zx13fPxDH/z0+z7wuXPnLhXBtY2tex648Cf+5A/9/R//C9/35//4wf68CtTUx37JtjxFov/ht99fFaVKyXUlez4vMaA5xpSoLqiYb65Prj5zbHNzvLHejMfdbDHoYSqkIgjwB2/7+HQ6MFWtGD6Yzm+44cTznvPk+aIPTh/91H3b29OUYiml70vTBFUlxpe/5PlVoic1qVaUDm6qDPSxO+45f243lwLogUPf98+69fprr7tKRUX1gx+5e3dvGBZDzsKBmzalyA5eRS9EdMhiZjFw2zUxBBETKS9/yTNTCAjwsTvuPXt2R0VESt3lXl8fveiLngkAlW/gyz1KykOJMRx23+ADH7rz4GCo+89NE/b351dddfQZT72uBiMRZSZbnmdYygjlfCikswwigBAiL+bDBz94p5rPZwMSIFBqAhGLwZOfdObG648Ng1T2rLvU0q8uJ9XBRQy8tj452N973/s/84lPPnT+/O72zs7BwTwXBbBrrj426pobrj913XUnr7vu9MkTW5PJaDZbIKGpVfoVEe3v7R85sqlm82LjyE0TY0i1sKRQV1pXTffHiexet+usbdsYYgh4/uwjB9OcunIg85pLl5xVakKODp6apCJdm973vo/9iT/5/5lPnWDotQz9TpPw6U9/ysu/9GWveOXLnvGMp544dhQA5ovFgxf2zu/pp86WB3fyUFQBx20YtQwGJZe5+SgFZpQi6NBEFMCOYi4WGYsoUFCgECgxqWibgprHSEQQUnS14+upqjOo2kFf9rN+/mxGhCOTdLzDoyM4scan1uMzb336F9z27G/9Njj3yNlPffrut739A7/7+++5995Lp0+f+NEf+9fPec5TvvjFz17MM2IlIgIzq0pK8dFHLn3Lt/3QdN+7ZoQUidjdzIpp4dhwGAMooKfYxoDjcdzcGl937dHXveZFr3nNi9ou9fOhKkMdHMy+9Vv/+sUL09GoNfO2HV24cO4v/IWv/YLbbi5F1ibrP/KjP/0Hv3/HiRPHSilArFIMtO8Xv/LLP/7qV79wf29W+YzuzkSO0HXN3/m7b/r1X3/31taaiIQQr1x+9Cd/4ru/67u/cW/em8qf/c6//dDDQ2J2U2RwxJRGDqpmhEF1UFOTDADgTLEZBjmxFd71jn9g3o1Gzd/6W2/6zd983+bGSHWIgff2d297/o3veMfPDsPSHnW5bVPFlw+jTIzxr/zgP/zgB+/e2NisLbwrl7bf+E2v/rl/+UO7ezPCxy47SCmKCFhlckEdXNYNqrqRtDYZ/cHvf/hPvvHH1sZtKWpu7iUEJI67e+Xrvvblv/Dz/8dikUNgVWdmCst2noqqaEqcS/n5f/o7v/Gb77rrrrPzRTFTN1XLgI7IRZyRQgghwtGjm5tb42/+xle+8Rtfv1hkDlT12k21aZqU4iAkarN5JjApi7bhpo2SF4gjXAWsx5GEBXU6YybgzowA1jQB3YtYTBGRU5NyLuCuRZ0dAC9c3P6u7/s780VPVCYTfP2XPOc5z3vuF9z27Oc+7zlN1wCA5nL58t6jB/CpK/zIjk7ng5qr+bhjIipaBeMgNaEBiLTs9wSCLCoKhMaI6hiboApuykwO0CUoCimwOqlJnc/kYk2iRVbJ0sXADA1gP+i0F4Pm4tw/dV4mjZ0+Ek5O8lNP+LFjW1/+yi951ate8ld+4Mo3vOEvfuyOc8zNP/qpf3v77bfi8nwQLisga1ty8BTmk3Hs2oxE4FiNAolbAHdYBI6AGCO4lmExPLR35d677/+9333/L/7C7/6jn/pLN1x/an9vyjEgYdt420lqxbR0oxDTzngExNykWERTkLbJxH0icLcmkrkyxb/wl37iqU+97qqrjs+nfdOmqlboZou+H49oPPbxGEp2IomJ/XDaBQime+CzdrQJ1ofQFFEkVR0SJwRCLymwBXFTwEgM6KVJS5WKwKFtpG37ydpoGKSJZJZiCkMuqoYURKrGPCA4B8YlzwDNPAZvGx11KKUwU9thm2C5WsTkbk4EYEVs2QivC09eKzus+laAQIS/+mtvj2zdCGgYEMlMHJCZjh9p3/XOj3zqk/c89anXzxcDApah4KGIjLt1Xdrdm33Hd/y997znk22Dgb1rHJcjYHQzQCEKtTGvanvb23d++s5rrpp867d+9aLPdQsdkcytaZuSC8eWABzIAKUYLKkwhMtZ6ipg/c9HXThF8MX8wNxDbJBaR0CXwOQObZMuXTh3sL/fdR0ipNSY2cbG2k//4zfd/dm7T506sbe384//8T947eteVXscOZfpweyeHXpkRy8c6D0X+92FxECusj5OogBIaksxqMg0ZFnKngZCBwhBBACUCMRAzZkI3CIhoA1FMdDuLLcJK4GTmBdZHbwf3NwDGDqqRfPAoaqDs0nJg17SyeUsw4lw80mWUuZ9VtXjx47+xE/8Hy9/+Z8BSJ/8xJ3nzl08fvxY3+clh3NJc6/VHCKQGoAPJjPAIGUgCogMABoaoqBFAYpo7w4hMLO9/R3v+2Nf8b3/7jd/4qrTx3IuDlXVPZgpuOdhbgYigoA5l7ZrwEW1uIlIcScAVBtibB556ML3/8A/fdPP/VCICABaqmQdMfEw9Cq55IWosaOU/el0VjUzOAR3FCl9f0DoaigiIYHKMNicuQWvkmSKyKoLRDJLi0UehtK0nbku+rlqyXlROZy59CnFtu1yziVncCjilVw6O5jFlKDS1s36xdzc1FQku5NqLtJX0v6ST1vjWoxmdUAZENxgKUWLhAQ07tqPf+Jzb33rByeTpooyFlkgsAOWMovBLl3ZfdObfvvH/973ysGMkGKKxGhmWjTGUPLwxm/8q+9598dPnNwc8qIfxKwAYd/3WsRciDClLnCMsSWkpp2sr01GXXc460C3Zd9yf3dvMhk1sfV6UGqUOil5sZgDOtITyecrPLHTKzcHMK9qk4DoIXg7GnkYiU+RsO+HnPPGxrGYInhdsIXZbP57v/f29Um7u7vz577zm177ulft7R4A+PKZg3huTz/60Dy79+axDeM2aI+qhibDYEwQmBgxMLoRMlbOT4pUR/IpUGDYn2ciUsMYYiAvCg2DiI/bGBkW2SgkB+uaSOjozojobE6izgQcYmybQZyZTmz6mWPpOde2J0dWqdQbyYrgdD6/6cnX33LLjZ/61EN7ZXjwwbMnThyrPuaqVttGABCY0aHymMwKgCGCyHbu9ziNwYUwiopZcE/rG5tMIZeCgEePrH3uzs9//1/+iV9589/jEHIpdeGQwA0MHMCBA1ZdzRACoAK4O4EvKQ4AnnPe2lz/nd9591/9kZ/92z/6p4ehNG3y5cITlDKoCQC4CXIAFJEhpjjvB3RAADMBE0dUELWDg8t31+Oq2hLgdW9JiSIzmTHDBnFsu2apgFg3oiuN1iE1yVSGPhNhalIVcliu+JiHFBBB1B2cMLgBYECipS5jnbqmWPlrCCDgdce76vEfZluKwHWF8Fd+7e17+wcnjm0NRRHBTYosQpy4o6qsTZpf+/W3fc93f92xY1u5SJ1U1DXvSTv6J//019/29vecuer4fL5vboQ4XxwYlJufcu3x48eIeD4fdnauTKf9Yj7f3euJ5rNZFl1K6NTwCuCmfuTYsVHXzPri7lIKcHK3YRhUDDFU0v9qNefx+PSIyEQA5uhEDIDICYln88WRFogQOWxubaUmLS8nJEIw08Uw70YTLnr77bdND2ZFZLnogGjmL3syPftE2O3p4qzdKXhuXy8TZHGIBFKYOTEy1u/ZAKCLleDnajpJoTecD0YUukRtSouhZEU1D4RMYKa5QCAkInWsfk2SC4CZQ5cCO857KQU2G7nhaHrKyfjk42kcndCHLJFxKvjBs9ixP/O4LASms7matik5YAhxGOaETEwAXDW8QgwOpYhzmLgpM+7uXfmqr3zF933fGxfzeWoaAJ/PF7u7e+9/34f+1b/6DZGtSqDvh2Fra/3tb/vAHR/73HOeffNs3rsrQHWoKg7J3UNI5s6xNryzSF+H7rW5WEXbh2F6ZHP0L/75rz3rGTe88Ru/fH9vWk1xRl2L6ODqXmdb7uYxNQAQY8g5FxmqYL6qNKE5OJCv+qrX/dk/+4bLl644GDN1Xefm1WyiaVpiJOT19ZEU6boGwREMTPxwpgyAqW0Wi0xE5m5Fq+JNTCkPubJMY+BDZqohLtWrKXAlOplZFc6u0XDJ39Ll+nEpGQC6lAD83IXtX/u1d01Go2FYOICYxAhbW+NLlwuhg3NK+Mgj53/5l3/3+//SNy2GpT51bKIUUfff/u33tU1TigA6I+U8nDhx7Kd/+q+/8IXPHY/HdQ12d2//YP/gyvb2Aw+c/fRnHnzrW9999NiGqaoIVgE4xNQ0JeeBIIRoUtxAh0Ue5qWIg1VFVnvClIT4RG+6IzNVHhaHsNyhdW/IR11TZfD6RT8aNUuVtaUemzIhxXXXfSZou2YoJQVqI7o7Bszia6M0av36Y9y2MQteOijbM7nv8uL+y7DX2zzbKCG7j5pQNeSISN27JqqhqyBACjEEMrcYgwyWQnVPAAeSQWKoukUEDpGh7eKQswcexGOEY+vx1quam47SiY0k1bxPzRDPTfHd98u77pk/PKUbtuCnvuJIsAxuhOzgqstl/WUnhTDnYu5VMc4Bq4woALmnq89c89znPGs+H6oegIpy4K/4itfdeuut3/ldf2+9neTcGzgzDbPysY98+jnPfkrbRPBiZggJcKk46rZMMHW5lENEQVWJAiGqS6XFu+lklP7qX/vpW5/5pKc/7dpFX6pTS9s2HBJzMBUirtvRdbIJy7VnXG5nm5RBjh295oUvfMHO9kHTNLWpXNXTmcgB6lZzFbdXNVVxVyQGWRKURGQxXxATIbopEdczZmZNm0TU1FwtcKxSFuiAyEQBDtOQ6gRRqWcAoGpV6QIQtQjEOtPw0aj9qf/r395/39mTxzaKODHuHwyvfe3tX/1VL/6mb/7BtckI0N11NO5+/k1v+ZZv+Yp21NYcENyJ8GD/4Py5C8zBDNAhxri9vfc1X/vqV73qJXs7+9ODae12pZBOnzp1443Xf8Ftz/26r4Uf/Ct/enowH4YSY9V9DEjobrPZjAi6cZNicHRObYwxRQhMiOyOK6b74xixlnRfKkOfUnAVKf24jYkF3FRsY3NDxTBgJV67WS4Fkao4IyAHxC7S3gAfPauzYmb+5BPp2CiMk4L7bJGB6NgYjnZ484nJ3lwuTv3Bnfz5C30W2J9JChTYiDgwDuqlOIONEg0OxBQJFlkdFP2xtX5sm7pOQSYSAyEghWY+4JERHh/xNcfik442R1tXt2HIiWnRlwd3/e336R/e218+yGCCxAeBdnYPEmbCEEJyX3RdV3WuapmznEkg1JiFUHdl0QxKzqUMqppLHnJPREw8nWVmesUrX3L9tb9w7vwBM7oTM6uUz9/zIDjMZ4uhSHXicadqE1NEKpW07VqqWmOARBGA1IacZzGOkGIRGXXtwUH/13/kX775l38EwefTedNEZjZVByJqEIN7US1FRFTc1LXAoXqqubvL9vbloe8BZOhLFW1tmlSl6M0sxjAMuba9Q+TaKHAHJEYgABSVpQ4PoRTh4DHGykAXcxFBorZt6t47U1AXpIAYSrWjKeUxu6Oua5fGhUstnMd8E52Jd7Z33/wr75yMxwbgrohJyvDHXv/yP/bHXnLTTf/sgQcurE2SmY5Ho8/f++i///fvfMM3vv5gb0pVMwhRRKp+gwM5QBHb2jz2jnd87Ffe/Ltf/srbN9ZHh65xsljkK1d2EBEBUwqTyaSUUoU96r2RtaytT1JK7takGCejC8MM3Opjvq7402r5+XFuvqtCn3U2W6iZQ6AQxWZVfH13d2djY62ShhlJRGOIgRFgYajWzx+9MvzMR/pPn+17AUSfFb/qaPfk4+2Tt+imo3R8hCfGlbhMRW19nDYn8LQz3e1PGu0t/PyefP5S2R1sobA3FzGfNCHGBG6yGOaqjOxg5AaE5uhFQgolS1jKWfJCHJE2Wv6CG9NNm+X6Y2MEV/XAFIDOTst7Hlh8+lz57KP7D23LWhdHpIY+y6oWYgyah1JKXUcBxCKiasxQGdvMFKv5qwpUdy+HahWTi1IlhYYAiA7QtMnUJ20zXt+UR3Y5RKQAQGa6XFMJpKXUsTugL9Xmlhr2WnJBRwd2QFUhoiHnl770he9736dMnWMQ1Y2N8bv+8KN/+8d/6Yf/6jfncoWYbanK5WZDgABOdeGOiYvpoU4LIHLVI3OHR89e3tubxhBDZCZyn1Vee2AgjkePrFdlqmVlRAGQEKp411Jvq0q4hMhVWxWX24ZYFYpFVGQ4tJBQhOBuTNy0LU7nzCyiIcS+HxAhxoS47MfVcyuiTRN//60fuuuuhzcmjXhBhtl8dtNTrvvyL//C0Wj0hj/xx37oB//R+lpdS8e2aX7+F3/vq77my+quQggsRba2Nq655uTn7nqo7Rr36pnqu7sH3/Ydf/Pmm2689Vm3POPpT7r22pOnTm3deP3pk6ePAdJ8tuj7bAaIXql2NfAR0f7+wdGjR9ygbRvnMN482jRdffAs5cGfOCYUT/iS0MERQ+QYYrsYSiAPANNZOYgFEfthGPre1yf1h0sphORgw3DQz/su4Z3n8795y5XPnJ1etRUTk5hvtWGYD+cu6iPn8Y4RjSNcvY5HOrxqI167FY6ODQl61UB4bILHxvHpp+Ms64V9fWgbHj3Q/d7UUB0mo5YRpWjbhr63ImZgKUD1rRkEFiWnwDceT7ecaq/b4kn0XGIpJaBP5/rhh4YPPip3ni+X54Cgawm31ppKfGUARmQkQIwxhkiqhTCXXAhp6UvqBocP6qXtiwlC+o/eycsKBx4TaEfEjc21t739g5/97CNd2znoUjUM6apTx2rLGdDAFSC4GXKqy851EM9Mtfo8VFmg+Vy++qte/cIvfO4P/tA/P316oqZ9v9hcH/3kP/yFZ916w2te/cKhH2KoNtTVaaZU+nt1o1GR1LYUBiRwEXfe3Dr2jj/87B++4vtqxKm+iOCOGIg553LttSd+481/I6YIDkM/hMjuSgQqBXEpAWi23AGoiRs4LPqBmerK3nIRELAyQqtEIrjUqpaRq/mbilQGSc65rgfVsrouUfSL4Wf/5W8yZgc2Lalpple2v/Irv+7o0c3FfPG1X/vKn/7pX8pZiIJqGY+bOz5+zzve8ZFXfNkL5osBAESVib/+G171lt9+J+JGZYGIZiTqunT3Pfd99nMPvpkYALouXXP11vOf97TnP//pL/yiW2+++frFItfGiJQSqoNhyUPf1yX/+WIeQzubzd20SYHQ+QmWYD3xeViE5G4hxsCcEl+c7gNYl1j6rKLra2uE6If+4yICBEzEoU1tbHD2H+6c5+vziTGiY1FXBxAcNczoo2TJJahf3OPtRbhvR7Yu4XprpyZ4/VE+sRbaAGqWi03aOGn4+iNhEEfmi1N/cLs8uC2DURpFM3Pi2IQW3cwHsUBw7bHullPN6bGOI7RNAEQVaZt4cYafPpvfde/8jofns0VpIqDjuItZ6va/p8AOFBzAIcbQMHWjzkxipGqYXOuTEKObl2IAkJpIHH0ptlk9wjiGBJXrIBYiB8aDaf8Hb3vHD//wP0ZwRDARAHOnJoVn3npTtRRr28kBHSxFO91VDZBrX99UchEEd6i6Bcv7+Qf+j2972zs++r73f2ZzfVzlALuu+9Ef+bnnPvdp1113om4d1PlJ9azvRl3XNiXnECiEAG64nJcwAJec89ATRzcBwKpuWNd7S4G86GMKdY+nHXVLuWkAWH6eKspQVfGWKsyllJQiEeGhr6qUgYjd1EzB1QyQgpoBYmxi1ZJGxBAiUpVXs7p/jkSRuW3Te9/ziY98+DOjrorDQM798WPrf/JPvkaKiNiTn3z9q1714n/9r99y8uTJ6hhWsvz6v3vf615zu4jFFEPg2XT+dV/z6j/4g/e+6ed/7ejRq0KISFxdh5oUqWXAegSz+x6Yfe6uh/7Nv33rkSNr3/iNr//L3/8nkcltKc4RmAK3zDyfzyfr6yJaMuShnx3sT2dTKYMo4VLyZBWwHpdBYf3vbD7vh1k3miC1Vy5vnz6uKQYHFykXL17e3FyPMVbdOwAw86Ff9IthNGmyWLXirELfzBxRh4JzkPGIRHyu3gYnAgPZnenZPfzko24lbzT+lFOjp5xM120FRmsiC4KjI/oNR/DGI2mWm7svDPdt26WDMh1MDbuIo8S3Xtveejod6ZAJh+LMjCbF8KF9fNs9wwfu3nn40n4zimjWMrgjI5a+mDtiYCJ0FzN0a1KqVhrqxtSAz6XkPAxV0Xip146xqtQzB8TqS4aqsrW19Zbf/uAHP/itqoYYulETU3Pu3N4DDzycgjdNKlLAoUntpUvnXvSiZ7zgC581my1SSkQ1Z4nuWoVQ5vMeEU2tiCKGGFsiFC9EAdHzIET0j37yL7/ild9dsjGhqDVN88CDF3/0b/yrn/vZvzL0JYSGiKXkGBtmWMxn1UpnKaYIXH28ANCsEAVAM50RRdVMFMHr1p2UIZttVH/WyhQHIERGZABFClWqxZe2Z5WEacysqqqKVaIfmDmUMjerdLYq+lyYQNX6fqhyYlVvfikQhgwMdbWwaiW++VfeVkTG40kRiSleunTlDd/wiqfd8qSdnf0QGdy/6qu+9Jd+6fertKFoGY/S29/2gTs/c8+Tbrp+MVvEGJmDiP6Tn/4bN95w3U/91L+6cOEcEo7GG107rhZh7qomgBQCpEghpPl88Xf/zs/dc9eD/+xn/krNEGusjYHdve1aBEwxWUiJkSMBUVnOJaLaavn5cWtfHW6fgqm5bW4dH69PiMgK1DLn2LFjbZsA0FRjCNXERdWQSLQ045Gmpsiieo+aiwDPRY81BABM0ARC4vmiMFuR7EAJAKPvLOxjZ4c7L/QbXTy51V69CVev49HRUqlNDdrgz74mPee6uL/Q+y/30+xHJ+HqrWaSvIiqQSAPCJfm/t77hvc9WM7PfHt3Pk6+vt6ow1CU0FNKUiSbhRjAxCgAWCAo9h+ps64FwM1dRGuOQMTq1RW4TuerqslSXLgK121f3r1w7kpdf1argy9vmxYARA2RAtHly5eaZvi7P/6XR6PRfD7nwEt9+CoiCoiO4JSalFJkplG3QRwQnDkgAhNmKWb2tKfd9Pd//Hu+9dt+bH19DQFyHjbW21/9lbe+/OXP3ticLGW5CBFBVYooEYUQvMqeWgFooEpvofWLK6a9anYgdwUEAAFApNjPs+lWk1pVqdov1bZ+KWaA1TseU4olCxFKkZhStRFdWjoSImIehqVVD1QlPIbqtOMQQwwxVGP6qptapcXq0DBnadp0990PveV3372+OTZ3IjCDrg2vf93L9/bns/kQA6nos575lJtvvv7BB8+3bXKHEPDChfO/8Eu/92M/9t1DnzmwubsaAPz1v/7df/wbXvPmN//W+973wfvuO3fu3KX9goiJCFOKTdMCuKiqzgnDsaNrv/Jr777tC5/7vd/zVbu7+zHGPJQqsmVmHGAyarQNW0e3kJKaHyroq6quAtbjmmS5aYzkMz842OXAaka8XBNbLBbVOs6rhiwCM3OIprnkkktpiEfjdjFbNIRMlNUWg2y13cYo7C2yKrgJuJpTyU4oiYECd4kCW2S8uD+c35dPPoygcmIj3XSivXoznlyjjqxlynkYN/HWM22l8BTRnDWSF+c7HpEPn7MPPrh4ZKegD4Tg4FkQHPr5AsCdOIuBippDiMCkpRgQOcSmAfcYKFIEMNFFG8lUQgzzeU9kRBgO/a8QGcFMC0JThZaql3XTRAdRVcSlqKHpgkNnqovFYja7cstTr/6/fvIfPuvWp29v73ZtY0UJAYmX+sWMgAageRj6fgiMDgau5goA1QgxxeQOly/v/PFvePV733/HP/nHv3Hq1LFSRBXaNvytv/VzN1y/1bWp+qGCG0JIMR0azau5Ya1YEEJqd3b3/tSfeN33fu8bdncPqn1ZbUWJFHMPIaxP1hCh5gtV7ctdwJWWmoWMSEVKFX3nEJhJDSITYF3fM6uCXETuWF/oJlU6vz4dS85UnbXN3CHGWBM2KeJuMdCb3vRbly9fOnH8qKoAoamur238tR/5uR/+0V+MKaUUmEkl7+0PMbK5Ibiqrq2tvflX3vpnvuNrTpw8spgvqlYXIE0PZjded/Vf+6vflwfZ2d179Ny5z3z6zgcePHf3XQ/eccdn73/gUeZm1I1URb1Hgbbxf/NLb/nmN34ZIYooEpha3y9CCAhYRANBX3zoFwSEyITshwqFq4D1+LTdCQDdlJbSQk4U3FRczGx6MLty+cpkMq7JCAeurSBEVMncRs1DDCSDMDExilkuspj3RRsKTERNpBRIhNQsECZGU3Und5esxtRGAnBGWKhd2h2u7A9itDkORyfxycfijUfDehwcCLNUKdvA/JnL/m8+NvvMw9N5KR0DqiHhoi/MwduIYMhsZrCkI1kVY6v8dTBputR2cTymqroZOKCjg3IgQqzeqzEGYpTagAWo/lSAlY9F6H4wvTCd7SCa2xwAARiAAQrSZDIePfWWG1//+q994xu/9sSJ4wcHs6aJIXJllrtldzrMPqA6J3NgDrF+HkQCUCLmEN2cmWIMOZe/+Te/58MfvvOzn3l4PBmJagi4fWXn4vlHmzZVDQmkgOhN07h76poyACEsDaXNwMWNjh+/6uabb5oezEMIzFSb/tVDAQGqoHtK6dC1wWtj3gER2F1FXbVynZiJS5FDly+riViodn0xEXENu7TcVLUakpa6gyHUiy+XUu93RGxSOn/h8q//xtvWJmtmYLb8dFnKufOXODRVTr5+F02sD1F0FwBsUnrk0fP/+k2/9UM/+B39kBGxaRt3H7LM5v10NgeErm2eevOTn/OsZ1Qy7c7O7rvf/aEf+sF/8PAjeykFMxAdQgxnz9534fzla6870/d9CJEQp7OpiDSIs0XfdTZM91LkrhtfubJDDLQ0JVox3R+XA3BzAAohmAMBmFpeTCdsni2G2LZ0/Q3XhhDroF2ru/LQA3I7nrj3bjbvh+hCSEMRUwGk1LSTUQxoBATgs35Qw66N4G7g4uTmzCEyqZRSpAnERKOGimJgKOrbB/3ZK7NPPpS2xvTGF2xctcFDMSLuor/j7uFvvqsv/fzYGMBdnYCDOjQtinjfZ1AhdJUSOTIBU6oGxrX/BgQxBAQoCkWU2MyNWIZhkFoHmiJRkRKcKyHbzVQG4gmFWDdv9/Zmt9/+nNe97ovEFNwWi342nZrq6auunkzGT7rxSc961jPX1tcW80XfD03biEjJNRfzx1wU6326pE06mGoIVFVJqwkzOPS5VDXLftGPuuYnf+L7X/f6vyBFkFzUCJE5uTsRgxcAZk5WpSSGXBvqyxiB1UBU5vO5iZirOchQMym0pTjqYyqpVbTZRAy8ZtcKQEjExE0KgRGAfDleXMoiu3vJMgy5SCmSAxOH4LIULI2RR+OmCgeKCNFyMti2XZVxyX1pmvRLv/CWBx+8cOrkqVIyAiMAU0TXGMnBkJmY3Aos7S3ITZGCqpoMo9HoV3/trd/1574hNclMz5+/uLa2tnVko5QyDLkajOScpwezamAeQnj961+xsz399m//ka7b8HpETEudvypPYuZIMTVajYPMESw1bWra2fSAiJhCNRNaZViPGwWrABgAqkrOeXPzSAhRjWJqkKrEGhyak0OMMQRHcvCi0nMbHHkYMmKJsQEO1Ruqn+dAnpgMVB2REMzE1J3qKNjMVHIpyOgOIGLMKOpEOGTNRZi5DeAuu3P4nc/Nv+1FWyjZVGdqv/XZUoZ+o6NF8UDsYuaA4CUX5EO/GvTEFJoYAhFABBxERZVBRaTk2KSIKjElgrJYDGpMHHM/VEsbUyVEhWVe9tjcHaHybiDnctNTbv7O7/wzZv8NCo6ZTw+mF85fbtqEhKqymPdd11bpiyVXviofA4cYK0sgNYkDAxBRNNPaEBqGUkoRkaZJi8Vw2223/s2/+We+53v/webmmqsYMEL1tbalrjoE5oCHblbV7gGBwKsAXi65XLy8f+ni5a5rxuMRMS3mC3c3h7YJzCRik8mYmGKKplVKlcALYFWL14ceuaCiauhOMbGLdF1TipaSien40bWUIsJS/E5KIQ7EYf+gv/e+czs7B9V6dn2tm80HlVKXw0Lgq04f2dvbf8tb3ptSKwbmzIRi/cH+lSEvmGjZ/LPsVhAoppa56dqtEJvqbDvq2rvufuQtv/3eN7zhVQj4C7/wW//qF97+DV//5V/wgqfdfNOZjfVxjOxmuBHcpPoOnjt3/ld//a3deLI0dqQ49MPWqRMbm+slF62G2G5u2rUjcNiYdJyIU2uO876vPuhWOYqrgPU41IOHatvmoMy0mE8NlFNTilILTJBz2d092NpaXwpCVjVeB0RITadljlh5p0Zs4JQVmgCBaX0Ux02UPUkJRNkAOTC65UEMIDAxY8lFDWIIoy6CuxKZGSOkGByXpOWW8MEr+a6LcvOxAIh3PTq98+G96J4HBNPC2LapX/R9XziwFgmECICqTlw7IKUoM43aVERdIaRE6EQYYhTVGJBTUC1NG1PTVPu86utZB+1qBg4hBNOhLsMiRiYe+mExXxxMZ8whBK5ehAho7pUw0Y1acA8xmNlo3BERZHjMcdaXTg4SiJBIRPOQRRSRiRNqrjsftVCSooGDou3tz/70t3zN+9/3iV/65T/YOrJmqr40C0SAgBzdHXGp3ilZqpCdIwCiim6sr/2H3/nw77/t46LqWtbWt1Qt595BmTlFFimlzH/1zX//xhuuGYYMaADV8otVtW3bRx/Ze81rvt/dVAQwNU3rbswITv3QA8Jv/frfuOWpNwCEymJHYlVbX9v4wPs//7KXfzc4MTEQBqZSimohooPp9PrrTnzw/T/7e7/3/g99+HPrm+siBRwBaH9/7wd/6Due99ynDosByVMKappzblLTduncue2/+P0/XR2hzUxVmxD/+b/49de+9vb1tfFstvjYhz9z950PhCZce83xa68+eeaaq5rEp08dQ6K9ven29v573/vR++57aDJuRHszj9H39i4+7/lfcuTI5nQ6b1JCIhV1MwcMgRTcii8Odvv5Toph6HuV3iHCKsN6nOpBODQlTjGmbm1988q99w/9Yv0ogRZ34BDW1ibVYhcBRUqtrkQEXNsmIZiqhhQRQSSbgyhtbnRt4CYS1zrMsUmhCWhGccR9llw0tqlLMBQJjAhggIlRALJqFnfvmTkwRw7Z8Mqs6JFg5k0TQwjuoP2MEIDCbG+/DJlS4+6MHogoJvNo7ipSFzvAXB3BjYhSjK6CbuOWEcHcmKBJXYxYjRjAAQnrNhmHIKJNk5ibpevzoWwdh3hY1rkdGoUikZRSXSzQHRmkTjQBBHQxnS+tCIERCjgwRySSXFKTulGnonXtDomrXvChU47UPlcIPAz5b/3Y97z/Ax89d26vaxtVhcBuyBwAAoVmGLK5uXtgRgyVmYpIDs4cFvP59EBibNztYHqOMLhrKQuiaA5Siut0WMxrTEUIhOFwgwuRqJSiCzUrzEFknvt0yFCDpd/i0jQamANTMHI1AQeiIFnMBJGYY29Sma5mZegXe3tpd2/v53/+PzgGQDbtiflgNr/xhmu++7veuL42cqsl2uGUCCEPJTXxzW9+x+/+3vuPHNkyFXWdTNqPfexz73znh7/yK14+DEOT4MiRbhiGhx88d+/dDzuwSAkcgMhN1cp4PJpMOimDg6aUpgcHk7XRn/tzf6oUqcPw6vrVtI2q5FLmi5yCISakiBSrUdChkdoqYD0eIasaH5lKyTkjAFASxb7v51jMgcEP9vfbrlmaEms133NETE3XJEoxSAhoxXxpx4qhXSzyZjtpo7npbGHVwylgqGo2gUCgjoQZXPqiTESETNAmRgAnMw1IaEhAqKU0gbA60amMG96ZlxBjiuTExQ2QwNXMqq+MiS7p42rmTu7MXHJG18AMKg1jQpehn8+TwoIpMTelLKp3c0yxsq6rEDETDcMQY4qJmKO5pqYNMTJTatJ83kPNbxBF1F1KySEEEKkXcUoRUcyBkMbrE6ImhHFIY+ImJQ4HvS+NSV2kEJK5InndDqk2CpWpWnkGqqYiJ0+d+D//z7/yJ97wA8whxgQQzQsSEyfmcYwpxkhMIXBMbYjjEFsjJ0Y3BaDATKwckqkbEGFotI4UWCUwekpBRGLiEAgphdABZCQiJFWNTXDnx0h5jhJjEimBYx5mSIjESBpiE5uR+ZywdTVCNMtEhoQcOCAXKVW7lYOvrbV33f3QBz5015Ej6+YWU+LAV65c+bqv/uNdCpcubhMhc7UqWzqnOThg963f8tV/8PsfZCJuR+AWQ2wy/uIv/cFXfuWX7u/vz+cHqkeI42QSrRNAImKRAgCEaF4AQGThYCq2ffnRja3Jz/yzv/fsW2+ZHszcjGOsprCiWKmtKXApJecsUhBrBxCW85NVwHpcUix3MwDKwxAY27ZTtWGYrRGhgxQhpsAUY6wWb0YMEQChbcd5p2DXsAu4dE1wpKwYYmhiXMy9S7zVkTmM23CoK4lIHAiGwZsYxNQceCmMRUzkbrnoUPnoKeVi/ZAJgwMMxZiTuY3atDZJc/FhXgYtHApaMVXk4JJjDESATnVd2YkDIbiVIbsKhsjuvYhHVod5SIgUYxwW0zxM1WZel6sRTZcOBgEpNSHEsL2zf+nStOnGiAnA+/mV2XRWN84CB3cz1RC4ZhaHnnoO4MOQAaCIxsgisn8w275yMU2nZppSM59vI2FKEQA4xJ3dncX84fPncikSQxr6bVOJISwAqylkDMHc9/enr3nNS/7iX3zDD/+1v0vYWeWAgROS6r4dyl2FJu3v729vf357OwEc7jAeehUib7hmAEYCprFIRnT31ERtu3FqYtOk6XR7Ov3UdLoJMAUgAKnkBgACCAAZIAEIQDw0qgglF0RYDP32lc9tX3n40Ii2/moFKP/JywWAAQP4vtjmv/v3f/jQg59rmjVx1ZIBuvUx/Kk3vhaZu1G7NJQmQgBmJuYQWdW+/FUv+oIXPPVd73x30x4RKcwJkX/zt972uc/e/+Iv+YJf/813XLp4OQ9DdSMKKaJ7iJU7xma55EUug/uwtTn5+m941Q/8pe981rOfvre7X3+FA1SirIq0bQMATYxNF7eObqQ0VhUHeyxirQLW40TBqvrIRNHdAbjt1taOHE2jjvOUmZh4bX1CFIZhqN4HKurgAJr7eemQETEEdV8aCBYxckRqAqJ7L15Ei3ouAgm7tpXaTQby7O4m5inioh/M680PdZKnZkwwbpk5ElvOuUgoRRfZ9mbZwUIKBKJSYkwUAd3nuaB7CCylgKkjUXWWVzPAmBLHhgiaGAi8iKlZigxVmROdAErRwGFhQzV2r10YVWtS+kt/+dtVLRCFNCZmyf0zn3lzHoa6/cvEzFgdPcBdRKj6JRASIXOojsqBR3/7b3/XfLEw9SLiTsx02223DsMQU8pD+Z7vfeMb/sSXq4I5dm3LDM9+9q3zeU9I1WTQa4in2C/67/nuN66tjUxzShHBiYK5qsrLXvrSUqRpGnf7i3/xG/d2LwP4fHbg7qVoahM45lIIQ92JCYEJQ2raroshRhlsMhnnIRPiV3/Vl95yy8m2aVRLlReqwoEqwlx9Cal6o4aQqhP18eNbUvSbvumrzz/6ghgZiQFcSq6rOSXnpmkcFAxoaRkWy1Cuv+GG0XjzB37gT8cUQoghhPl8uOGaM1edOdX3Q3VprIVpCJxzcRFTBUIm/MEf/LbbbrupG43qGnjXNcNQ9vcPXvPql99227PvvPPuOz/72XvuffD8uYsXzl/Iud++clnNmybE2Jy56qnXXnf6uc99zhe98LZbbrlp6PPe3sGyJ1gHxADuPllfqyn7bL5A7HZ2d11L2436xcLVa/9+FbAeJxYWBwZg4kAccy5mtujLyGE8mSDC/v7B5z57zzXXntlYXwOoXm9e5dVjakxlMQykji5tihzIIaYUWiQE3xrFSZcoAqN2TF0bwZHRAbyIO6IbMHMWHSWipXkmJMIUKWcFd1EjNkYat6EGIw7YJlZARzdFoyBFidhBkAg5lKEQIcWkor4kxcKSsGMiaiYSYjDE2CRzGI1HXTeWQSbj0WQyNjdVO6zSHBH7fogpff/3/9n/4tT1iz7nUrUcqjtpdWN3t3io3xADm7mo0qGj6h//hj/2X7zPwd7BbLpomuQAX/PVr/7PGoyI/aJXNQ4kpdrLW13rGxZDjPHPf9+3/tffab8YhmFAJHD4jm//U//Nx5Rppa07Mf633qGvazTf8i3fyIHBH1sl/O+jXwxq9v1/4c8dTkzBzKs3xP/PzgQg+X9h+G7q0+msPiar72zOWYQCB8ClSe3Q55e/9Iu+7Etv/68/xs723tpk/CUvfsFLX/LCOhgacnaHxWJRi2wOYdR1HBgASpYrl3eo6sESEZEjVjYsIj70wIOnTp9q21E1oO1GE5F+fjAlIsSaza8C1uPWdl9qA6CKxMTM1AYEldliIWqqdtXVVx07dsTNAUGKBmYKWPpBimDbchwhcwyEyDmLiBRmQGojTyKQVz83YdB+cHNoYnBzMGsDceJFP4AbUCLwQSQLMDG4qnmbQtslMU+AbWJElKKm1qaUZZj3PYE2TTJxJiJumVXM0AyZhsWgKnDIIiIkihxiJARMjSO1KVQ9GQAncsDBLNSIjAB1tId17Cdo5lcu76hWjTqpM6mUIuJymJhzrpbLgFC19NS8bs8yL1OtSurZ3t6ry9NLshUiBW5jrGouOzv7pgq1ke8eQ4wxIFb6m5ohgCMCMbdt5657u/tI6OYiEg71F5eMUABz39neJSJfdsErgRPMrBI1ObCIVjWrmGJVYiCiwAyIRLS3t39ongx1TFa3lWoLvNpcmy0VrQCWJSUR7+9P3Ww5K2BiZnM3MY7sZtV+CRBMLARWtRCCL70al3tA1TK+ptzMZOpmGkOoVaEfejib6u7uvpmFECr9lUOo7JOYgrsf7M+qlgYzxxhEJIYUqhEswOVL29XTNaZYTcYQ8VB+o46JMXA4fuI4M0Pl1IuhaWqa8dr6/s4OQuWprALW40PCqkpPAEhUyqDqhEYgACh5cPP19fXJZK1elq6GCCEGNTGwYTGVjhZGLaCYgSkghhDVXN2JMLIhuZsHRA7RHdtE5D6oQaXSIwZmdZMiMYYmkAKZQUAI0c3KsEAKYRCnkNSKqjJypV8zA1Fw9yYSg2bjImoOSEwOnBILidUbDSSXwFjE2M3d0CWXZlR9D4Dm84Ub5rIY+kErj9vdzZERAOo9RkxsBg7o4G5EWA2RHTzEWG8gd6vKyBQCuSNCKcUOb+aqD5VisOVDogYlVCCqZiCq7s5hqbFVCfaqS896ZqqU0yJCaoBOSO4eOChYrWEPsxUTtxoPa7SqpuyVpB8CujNUMoRDihERi0q1PA2BTQ/NSplH41i7nLlIIFIRJGIKFEhNQ4iAoCLAXA3lDRwcESCGALBs5yHV/Wcz8rrwbGpYnXIQqm195bhV1YP6Nojk7G4WYhApROS29GeuutH1/yGEpRMqYoihalFWTqkURZQQIiIye3WiZSZVzUWJWFUn6xN3rzTUat0qpeRFSSksJymIjNS0LTgEpq5LPUfJgiEocGg6ZkL01WrO4xqzAGzoexVZzA+GYZGzteqTriWiYegvXrh84v9m7z+jLbmqc394hrWqau99QueWWmrl2MpCESEJiZyDyDkYMJgMAswFY5OMbWyijX0NBixyMlECkQyIJIGQhIRyVqtz9wk7VNVac873w6zdePzvO95vV+PlPyQPj+EhQ/c5e1etNcPz/J71a/2m7Z5+QAAre1WIDBxCjNQZpUlNAyICJFEmYjRiqDgm8cmOhUCBzB16DirqhxKRXABpCoEIUJvkOHFLbY5F9BhXJBaRejyZTFoGTckiKYTYGobg5EejgJrUNQFMqCqExgwxsoExOwsrAAYiDLFQTarAIXLsVAudvhnM2c0q4t0cAhoYMfn9b+BReiDe8SEa70WSmufdE5FvtNCB82DMBN45gq+qOk1uzhnAOARVc1kWM7lOXUS7nE7zNhGJEQyZmQMTkuvzp2ccAIJP0NTMteDOds4u6/pjmKCZqcO/GJ2V7Hw+MP8E3IljBtAFCDm8GAmyZD+OPRdeskznOAYOdAYwNa8JPUfWf3JJ2cuoEIJkiVVh6t5D9AMFwZCD610JUIlEsohmk7IosiiTk8q6hMZO1ouYUkJCMGNmH1l4FrSoOAYHEP2/6UR5l4mAWdO0ZkaELtY1wCJGIvBtOCi0mup6Mjc/LyJFwRBhsGIFWDKdOJzWgO9PzbmP/nHIKwBWZS8WZRarJ8uS08z8HOZGwZq6UVWHI0MX1Sfuw9Cs0raYU85QRDIwAEMCJIwcY6AycgxUuN8LLAYigjYJISFYEVnVkmqb8mCm541ARCgCqoUyQlLNORvxjqVm3PRSq0asmqteaCUDVFa3gTkWBYg1TRvYL0sjxjYBcoGmzBQsQSgAORCmtjWkIiBGIuacEhc4O9s3/xB6pXMmQ4w5ZUQixr3WVhHxukBNCYm6lxx0mn5OSGoW2B25XRp7YEBCNfMceW+9vfUxUyT2lz+EoKqExAHUwBCIGQycto7kLZ4iAEFnSvY6SEA8Q9CPIVdde1NjaiGyFywcQ+cINCMmLyKJ2AD8kCImZ4GamuTsW1JCdL27v/Oqzu3SP2LBAACQ2dnnXoh3/x8Bb9XJ6TVISIQCGCh4fcT+R4F/wDQ9jsnU0AUmiJZNspi6oUe8eur4P7aXCoui5mR9fzCnn7xDr8AVyIQE7ImtmNWmS1WNMTCzR0aLiGQhP2E7tCESUyxKU1WjJEg9Gg8X63HDiL5MQOP7eVj33dAduyynrFmYYwxhMNMbDhvKNQKUVbF27ZqyV0iW6fTa3LCCzCGWPSYiz+30MEzNohR5IqApRYRIKGZMIEmTgopm914RBSYx4+jGUTIAQqyzZckpKwCmnIlUc2am2bn+pMlMmpMRc1VgahIxm4I7/hCsFURyeJOpSBEYEDVxjCE3DRURQ0SzwNC0SavoyQsIFIIHTZmKNm3rzy6iEoUpaByIpiog7d5572F8d2ZmTV1zYEIwcAg6GJqakeH0rDC3+oTAZgrqcF0f2xoi+qDH3AmElCQBoJN/zRQydMVRt5gy6sCoCHsDRMGcjUlEIObqoRCDqIonACI4dIYJFTTE6N5JRDT38RCAQgwBDNo2pzZ5H+a/vmaJRaGq2XUGCL4q9TOCmMBPcQADU1EvDSmwD5JMTSwTcs45xICdn8jAfddZEICDmz4NCJEgxpiSA9ohBgcvAhIxEqCYgSsDiRxeajFGh+2YTyRx2kagn7IgWZz9zUzen4IBM03GYzVj5BgZiJjYTLMIASwt7M6DwYqVq9RMWomB+/25qj8zHE6YC+jSpu8/sO4rKRaAtU3LkftVGUKZUztTxB4RAJRluWt5j5qFwJJziIGAYi+EUE3qxXHgtp4EycKIoUBQRkMOCjZXhZm+rujxruVUlUaAZS8sDtuyIFYhClmgX5FZdIWBB/WhQcqCaAyIhLGgcatFVSSxpmnMCE0BQYyWhykUlaRJFiGmEIJkI0aVbCpVQRDI6SDGoKklxJQFIBNak01DmUQNUXIejesYK8KsBohQFgUTQ4DUpk6x6UtDNQY3+mpOiYgA0JlZrlENRWRiVQWEnFpVjUX0Mbmn1XrOqIiYdV2ej/D9vw7mm34nfELbthyCQ7mwy8XwuRYwkaKpmjTtdMWGIl0ZCAAFF65lAwQQnExqM0PsgmqcLy8qpuI0Yc0ZIDs/jwgZScWIvfgGl+A52N2tkTkLEXb0C0A1MFEkzG32X7Zb+CHk1KG1XF1LTEyFqcVYmJmCSRYCMAMP6UFEgyCikQnUUYEQA8cQRDXn7KEbrYhPvUTE77mckuPGVE2kdd50CMEZYR0tzD9Zb1dzbhtB7xURfFpvKSOB+ONoBgZIpDlXvX4sChGpiqKmKKJIwCGomUgL4AOvP40S5U//wHIGnXMKVFWTD54sobPrUsrzK0pVFUAwSDlhEmYGTEAxMqpHjhJURWizWM5NwiwCaiISCwbUJMqBYmAGyAIAGphELBIhmwJQgFQLEcbAwUQcpwc4Oyjq5dyKxaKQJldlJGrqNkfQnDOH6HkYWQ2Jtc1EoEh1AssphMAcVCWbUWAA9oQCkZRTI/0i52zBQixiUSFM2radjGsDa1PyXaFfzP7GAtleyoL5CcWMCAboJN/g7M2c3SXOHAhRwXLKKuJBzWrmogfs0mhQVIkwhujDHT/yyjIi+jov5Cxtm2IMoubJtznnGLlXlSEQUAQQFclZPCFRsrRJctYs4v3OYNCrqiKlxN10jH0ZAthhXnIWMAgxgGt3a2VmD1jz4tFrT6+kbHpwlmWsqsLPCDAjN1SaEbGaZdGmTgLg2wZVcSyMx1YzU68qiioCEFg2BeTgqbFNk83AkIiQga0TjIBTCYlRxHzjCQaOkPemMneceLcGQWBWVU8qAbO6rouyyNl9QgQI3gJTCCnlEALuvb4RmaOJaLdd5bZNxEyUl4fjlWtnBzMDDL0QC+IAoP+nIOP+Cuv/+m8RmAEsS6onTRIoI89XgxCjiBCh58UjoQ93CTHnFIs+IQpIURSErSI2WYjQb7Y6ZULauKa6aYfnqkNOaqDJiJgBbK4Xs2CTUpsTYOiVgVFyhl6PI8CwbnO2sl8VBVcF9ou9OXGKBsxECJBNciIiQiiLMB43HIhMi7IYN0nMtK0z98wosCBijAFUiiJSEYSoqsrAxAxM2kyWZmYqdyxPdi9S1yqTqZmAqpRVxcw5504OCiHnrKpt2zIHv+qJKcYQmADRFaQAaKaiyoFDjD7DFlFmNnBNKecsBqCiqhJj0baJCW+88dbde8ZFWdaTZp99Vh6wcT0AgkkWqaqi1+vfdc/2e+7Zdeutm++4e/uOnYsLi0MVO3jjmg37rjr4oP2P2nTg/vutCYGXl0eEcO21t96zeVfggMxubWAkVePgxkmIEclzTNv22OMP3bBhXU4ZtaPRxhhSR63qxmQzM1Uswu13bL3nnp2bN+9cWBxt3bp7+/ZFJJidjYcefMDs/MxBG1cdcfj+q1atGI/rlBIYiObBoCqrmaWl4V1377jttnvvuGPLzbfem1IqAh911MaZmf7RRx24776r1q2dR+ThcFJWhSH52gEJmdl9MCqC09aYAhGTqhVl6dhb38+YCjP54FxU/Qox9VEgQjdw68CnPr6vej1XzPmkjH2OprZi5QqXX6jKnmG7fdvu1AzbSR2YEcmncPcfWPdRP+gAP5+wenbWZLjYpJxzg2ZtSoDoCfUeHacibTsSJTA0EY4lc2DIopBFTI1jaCZtGWcRpUQRMQJhhJQVkFQt58zEozoFIiaqyjKLAFivjEkAARuxbGgmpjIcatuC5KzCOSsgpiw5axkRQwjOrDNrk4qZ5ByqKufWTAghFqXmjMjMBTD5/IcImMNEwCM/2zbHWMRY+UehooEZmdDMja+IwFxqB1fQ4PgngBhj27SxKHyF77c6ESGzmhGzZgEE4kBk3YQIkQN36zBE5qAmLm9AwlgUPvpFk1e+6j2/uWbr6hVrtm3b/qpXPOF9f/uKpaURM62Yn73yqps+9q//dcVvbllYmIzHjUhWEA5RcjI1LgZkUPXg4IPXP+2CB7/oBY8aDPqf/8KlH/rQF9fvu1/dtGoWYsEUEFnETB2+OqnHi4P+7MKe4UUX/eVBB24YjyYOPibyARxzIJdfzs72b7zp7n/5t0su/f4VO7YvmgWz1sAAIzi5AUFVqxL327DiCY8767Wvfqqvj1fMz1z9+1s++/lLf3X5DVu27BmOmpRaNfOxWuAIBlUR1qxbcfyxB7/0JY8/64HHLC6MOATq8NuA3Ok8iFlyR4VV9W4XupUhOMk6cwiAGDkaWISoojkll4/wlEEWILhyTU1RTbJgDC7N8KOwbZMZDIfLg0G/6vezWBqnquAyMmhiVAQEc33Mn8Yb/6eNl3HsNwARBSdrV72Zqj8rOVk0Yi7KckUoAJFD96ISYVEwE6iaqEKuAyRNLXNExqRSRsYY+gWFKqyegaxSEuSURSGJ9SJRDKo2aVRMe5FFRdRErWBjimqaAQGZgk2alogChVjEGIOYiCRgllSPBHoFFZEk5RhpIjnEaKom2USrQAkIkczEQFMrTGBFGZmNqBUghF6vMISqqppmPBovNo2Br8wQCBCZQuCUsofT+VjXKcwq6oR0X8blzjSjZoABiUBEQA0ANJvroD2pwXE+qWmRiBhFsuswAQCARdy8TUQEoJKGAIOqDFURvJroVeVnP3fxq177T23dzsz1A1M9WUptozLOmkGBQw94UMQyhLnf/uaG0dL4+c97eEoNo3EEwnGvxCx5YWGrWebIOQmoEmMsepIl5+FkMjZz6xV49aEKRIRobduC6mBQfe1rP37dhf+2a+fy/Fy/KnFpaVs9WSQCl4yL5GYyGsytm5/d77bb93z5Kz95zSsvAMRer/jYv33lr9/1yaZJZUEcw3B5UXOTc5P9UgnV3Mwq4sG2rTu+dcfW7//gyr997589+5kPreuEiFmUiXLKOWczLcsyBJqOFy2n7GeQF7OmiiH4OltEmMg3oSEGFEAit/x3h1cgN5x5tKKkzIENwIFlzIEIvcIiZgAE0Lpp22RA5GLhPymh+/9LWsKUJatp04zNTC1oFihZRIlw0kxmZ2d9CgNdKIMBWm7HWJZi2GbhLgaFDExVLFtEQ7UIYup0UUxNcryfKIllBFcA5DZnQpRMilhETEnArbRIQGgA/X4vi04mkzZp3WrdJo7MZAhmIjklj/IEUwZDU5EMwAWT5DYQJwUEQwpq1ooSIKGJMgIGpuHycDhc7g/mQGuHSejUmqNZmH3xb77q9thn3/GTC60YmdmnMzRtSBAAmDRnQGcqW0oJwGJRpDZ1Aveu4QBVAMCmaTt0ihEFalPtyX0pN01qRaTXK6697ubXvekDADa/ctDU48WF5fPOO+XUU47ZsO9qjyZNKe/Yufsn/33NLy6/KTeZCNo2VVU5HA01Nym1bZKZ2fItb3n+TL9KbV03TV2nECIicYhV1TfVww7fmJMURVAD30Wqqqc9z872rvn9Ta941fsRitWrqpyb3bt2nXn6EQ85/9QN+62dnakAdGl5vPneHVdccfUlP/i9weqUB03Trly14uv/9YNXv+7vV66cn50pRuNRyunxjz7j+OMO37BhHSA1Tbt1647f/e7m7//gyl6/P+iHnOvXveHDhx263ymnHDWZtGVRuKQGwBCpbZPXf2oaQiRVnwZ2awcz6ACq6CoHb9lUO52ZmdaTJpYFKpiamM9k97oPkAhEsW1bDiyi4/F4dnZGTIh4diYuzq/kWNV1qyIx+obw/pbwPqkPvU0CQCaMgYqiQMszgwI5tEkNQM3aNrlNwRfPJhoKDiHGWJpkyXnSQskWyESViYvIyMEAsur6FeWglFErmhUwABgREGovFOO6NQMDqspCJEciRBq3EhDqNqesgAhIgDBJVkYuS1ZLhC2oKXKExKBk2K/ipGlzNk0pRAaDoigkJUALjEkNO4hdMEM0U5MqooVKDGKgwCUTTcbjqkJV8a1TEZ2EB4CgIt4KMaGBBSJwCr5p8J2bKXSe5y7ziphxeu8SQc4SY+wu+RDM1DWgYqAekEUUQ5wOs03VhU7dAj4WzMyE+PH/+NquXcsb9lnbNk2b8f3vf8tfvPwZ/2eI5/LS8jv++l+JcP/9NsTpYWQAIfbqZjQ3O/uuv3nt/49nwj2S/iu7Pst16oAQQvzEf3xraXG0z7oyiywvLT//eY/4l4++Hf8PGtRkMr7oM1/fsWs8OxggYlNPPvZvXymLMgZqm7Yswsf/9zse97iH/Z9/+4c+dNFb3/ZvMcyWRRiNx1/5yo/POuuEyaSdinAcyIxe9ZELZEEBp2vWjtXsOjbyWZVXiQgoWbu9J1JRlkgYOk4GS9aUspkZEyCDASFyjGpK3WpFYyiKyG2rIpkZTEVNcucHux8vc9/0hERMBKBEQQSaps25VTVRG6XkRo+UMnEAMzLwCF8RMAWmoNKo5EFV9GIARFVgUlOdnyl6ERFxEKTP0jCJZNOu5SIOrnFvcoocCopFUVUFThppRRSBmQy5TdK0aXamUuhCOR2rZgYFo9TJSjQjEQOisiIhFBEwLQpycasiAwKXpbecIQY2cUE5kQaGnDKRUIgiGawoq6qIcTKu/TTv0kwRi8ACIEZiOk5KxEyAZk5QqKrC/14qAjNbdmYzCaEY1KOmCiyEYMSmJkKErl2oyqBCZSxFoclmXc6YIqFp9l2bSSIUU11eXr7mmptmBrOiWjdy/HFHv/xlT11eHjpd2vf1iORS73/6pwu7w2t5iAAhMIDbB9EArrrqhplBr6kbclaqqluvzXR2djA7M0AiMA0xWCd4BVCXDrS33HJXDCHlNovNr+i/5c1/1jRpaXFY9kqX5nv2B3N46UueBSaapUmy+Z4tN9+8eWZ2YGZLw9GLn//4xz3uYbt37XHfn+tIAYCZXv7yp33m85fc8Id7Vq6cjRF++9s/TMbjEKipaz8WicibdAycBepkajY3KNFRFYbJMasmRBhCAKQicFnEpIambco5ZY+ysKRNk4g5BFBV5BAjx0jut3StW04CIEVZ5pSLEsdN23JMbQsWYtHLaZfkbBbt/gPrPhu6i2QAokDjySQENqNRnYM5SoFcO7q0uNjv9cRXhEySM4cSuUFrUxIAC4wpo0kOpEUMPvM2w16vOPuw/reuG88MShEd14mJEWF5nIhjGUlFVbNqFA1mighNK2Jd+IqajhtJaoADYqx6xSwW/aqtRZPGRqWEbGKCIWc1hRBjTgpIzMEoqKEouU+EGAhUckscc6PGuanRADhw4AKpTKntAjmIsog3vynnflXsHjYf/EVbQ2wFRAAIiKEKlFIiBKC0z4oitVK3CcG8i5y0basggKoWSXM7LguOZECQsiBAr19M2km/CCJ23qHlOetFAIkghoIIFBQAQ4iG5jrvpml37V5yYI4BrF45NxqNzcBniwDG1AW3SM57du1BnzwB+OnqFqAQcHExPf1Z71ZNZdkzILcJxchzs/M7d+5+1Sse87KXPnG4PGHGva4XFUPAwKFpmvFo5BjZ1DZr1qzu96o2tSESmqqA2yE94X1xYclRq/MrZrdt3bm8NCx6PYCMwKtWr0wpeZQiEYm60gJVlJAOOnC/a39/FwAwwT2bt+/avbxu7UqVbhDuRSggBpDfbaV/+81kVR/K0gYFDes8TpCQUE1yUrGZfshJI5sIxBj7FbdZJ3Ue9EMmSgJNjdlShFxVcZKTSA3MDz2In3AIjAVjoFhE9wAVMaioqkHBg9mZnFPbTFISRJfd3y8cvU8KLO3aekkpM+GkrQczMzFy27QxZAKjEHpV5blPexkpFIuU6rZeHlQhlv0kVtcpUJgrgZiXGuO2ZpwRg9TqEeurhwv94s4mGfYGoSCs27YHLFliLJS1aYUCTJqsZgCEDJEYXFBglpJwVSxNsiqrmmuj62RMnBoxMOYoSIYkIG3TVoGSIBhkM4Gg2I0z1FQzqEBADYGy6HDUtslmK/I5u0jaK1d2XIGZRoymMqrbP+ySxVHNkVObC4YEOKhiyqKpSW0KITLh0rAeVKHfL5JxM2kF0GLkEEma0bhRoKIsFAmkNTWORc4Z0fYk2nfV2kceXCzX2qmLzBDcUmNgRhxd/YQIvq4lJEAty7JpWyIGMCJ2agIgEjOH4FLPbmcPyFwAWFH2TPHOO+/NuQ6RDSITq2mb6vnZ1Qu7F3ft3E1IBm7EnAqdCIlYRYijA6FFBdHpzxZDNAUObJ070fzoQaRYkEv5eoOeAxoAwQDKsgyBp4Ynl4cqApoBEe+zz6qcs3fK0hkGtfMARYLOK2lFj5Zz+v1OqWyi1DRJOwWpaMoWAvdLAq0BMTA1rYqOqkCKTKZZEpclE6JKK0omCSNHApXE1VFrqhgDOiXVjJjbpoYYEIERkHgiapJjiKLqDb7dP8O6j0SjNtX5AgFCjJUCcYi+cBERZB6NJmAYmADIh0DuXO71Z1O7Wy0PepHbCQAsN0gEp26ITzp+dnVITQIzq4GOXkf7zvVu2d7cvlu2jWSSjBF6FZNZRpPARVlGxCal0TjlrGrC5BJNDRyIMURmZhQQyb0qLIzbnFswEWTJqiCqaiIhxqKgNludSdFAJ9axb0BE25TRDAjaRhMWc73Scw77g1KkCWyuV/b33J9+U8XAc73iKZvSOMUMuDxuioDIDMhNmwyYQZloeZIJY4EyUwYjrhNHBCAQgJyI4nxSSwYGWBhEAAMMVEbGqsCT1+lkklMrEGOIAcFSOyJIoA2CFjGoSFWVVa8whRAq4sni0rKIqAh0GnLrLJAiHuThXpbpza9migCScoj64heeu3LFiqbNdZJuRYwYmdu2PfWUo3OWjvOPKCIuyPRs3SpwVfWsU8PAeFxPJvXKlbOd9sk9zNz9MABmCm7iqcqy1+sn0SJWyO3y8iinvDdSDBB8NJVyAtOdOxe9aMk5r5gtyyIYdKJ3EGBCZhCR8chWFfDkTTEgB0RVKYpQtzmJZQM1yNmPQZytgpsgBkGHtTYZEGO/DGrWZJjpFawyTGAIc0UghJNXa8qmKqK+1zFC6tLYABCMQXIe15OxGZuhgd5vfr7PTqzOMkFkTAigqUmpaSNC4JhFcl03dTM/P0eIKpKz5pz6/XJ+bsUdd+4aVKuKqhQIiy2v6OMJ6+OjNs2cvG/o94q6kW5SoDYR7bOdsrE4aT/bsZzvWNDrt+VRa0tNRkQCbeqWijCcZEAsi6LNOTARgyLPzVTLLWrOOTMimRmqEaISk3X7nJxF3chGockmOUPXDhGFgGYA4BPVEArJkggbSYSVqabUVlVfIYYYEToWiqulACG1qc0y14/POzmiWRYFjCoWYzAw0QhqHtUhIkVR5Jyxw9MYmHEMkjNhP8SIhC4cVZMuXAeAmZloUrdCsSgICXNOJplCQRxEEpLXStTv91avns+yHYlnZuZuumnLVVfd+MCzTlpaGiJgR3cAQITBoByNajMtyyK16rAz1ZYDa9P0B/2/fe/r5+YGqWkdRweGbt7mEHPO41HNgVwdFmLMOTMHFSGkWBQrV83klJj7iLRj59KvL7/uGc94RF3XAKDOkFEpiziYqXLOXiwy88ygX1ZFO2zAcHZ2xWU/v1pFyqoYjyb2PzawvX6xe8/ClVfeXPVKRFKBVasGg0GvbdPMoIoxqtnS0ihnAQRR3LQeT9qvRKLUJKKCmR2A42tBA2QiRMgpeYntH7ipMhEXMTXt1MSN3UwfgJhUYZy0V1VOf2dEYA9hxKoM3A80v9KrXcnJPer3ewnvq54QDFQRmZlTTpO6DpHLslCVcePSwTAYVPV4HGMBYERUlmVV9fr9XtsmqOZnZfn4A/sHlnbeEf39BkZgYjRpzXPe3ffjlJWkaGbr5+OGFXDifr0dI7l6S3vXrna5xSzWZ+xXMWdDBAOrm2TAZpZTIuOCi1jEdtyGwJM65aysKYTYtslRCFyUCmZmgoyBKs+nx4DMROT2aeQyELYpoyUVlDbNz8/mZvHue7YNBgOU3TOzPdf1mGnO6qUWM4vC4rgLdu9oAVm6ATMAGLAqIo1HDXWEb/DzS1vLSZEQ23bvmGPqNjdRdXc2cUB3/BiAWYwFcR8sMvfA6k4GgfTIR57zgx9eVxRFapNke+OFH/7wh15/1FEHzvR7wAxAKnl5OPnvn173D3/3pZ07F0868YB/+sBrAXAwMyCM7iZMKW/fujMw5pRFMgIbAkzJCkxUFKEz0wH4+ev+5JwTYv/cc0/55jd+GUPRtG2/N/v37//cxv3XnXra0UURO4YaRCS6+ppb3vXuL45Hk332GXzog6/ZsP8+p5569CXf++3MTMmBrr9x87ve/fE3v+VFK1YM0JRDACIVXVwaveWtH7vrru3zc70Qi+XlbWeeccLM7Ew9GV98ya9uv2PHpqP3O+uBx+esRKiqWTQrA5oIMVAgNmIz63gSYMzmZsachYhCjG1WAEJDHWVElpw5kGuynIwIrYbATGQqDhTKIku7F2dmZ+bm5xAtcEht5lCqGRMhkufC3X9g3TcnFgKC+Ywy51j2AKRe3rF634Nnir4ZOEJ7fkXfOQWIiEwcwtlnn/Lt7/xi5fzczit/9Yq/fPzBBx8IAFm0aVrMGUwke52MZuBoTyJMSZoMMQaCtE9fNx5dDZt49xL89PZmsYYsWCedLaly7jrQpIEmCRdOVstM1LTNpE0UA2kJJgaYcjIqRIyYRRKTAoiaIRFw6DQy5AwSSGJokg1qwWS4PBx94P3/fs9de2Is9jtg3WGHHpxSho4AB6kR50AAuviTACgG7mB9aj6IcbCMgTF1OmznZaW2jSEQgiu5GYkCM1OnWgC0qS8aRCmwR8E45arXGxD3OPY4DMEUEVPKL37hk7/y5R9dccUN+27Yp6r4tjt2P+lpf7Np04HHHnvY3MygaZrbbr/z9tu23X33brQwHE7adkk055wAiENhYEVRJYEXvfwDIQaEkLMhGSKWRY+ZmHU8Gr/qL574+MecNRyOfJWpooAWYkDW4Wj09Kc96j8vuuTaa+9eu3btTMC771580lPeecLxh248cMO6dSurqti+fWHXruUrf/OH8VBzSqvXUpvSfAyv+YtnXXzJryeTNDOYHfQGH/3XS773o2tPPvmoAzeuK4pi956le+/dfc3vb7/l5lvXrF0Dpjt37j7kkNV/9mcXqOT//fFvXfiWT/R6/Uk9evdfP/c1r3rG0vIoxuBiEcnimivNomaEGJh8YSqqITASxhDalCS1HiqB0543hMCIQAChg964L9p3FClnZiLCsiqrXpVSnqkKijQ7V61YvabqDXp9QQzEQMT3H1j3xT80dbGrKgCZCCBzOdtmAw4OIy/KsmnaGEInrlNr6vbZz7rg3z/xhV07dqSUH/PY15374AecdebJJ598xOGHHxj7VUptXbcqhgQunHBrNRExU04JkIA5K5SBNq3jI9dVeya6eTHdvC1tXsrLjUpGAiuKgCAJIDI6PhsMYihGdTKTyXCiqjFSb1ClpE3TGEAIbBA7/QSCGaQ2p5zIlBCNSIAJtFi4d/tNNz35S9dfd/3dq+bX3Hbrb1776tf1+v3FhWViQkIELAoCAFElYmRwFQ8ymhoCciARbds2+vwYXVeAbpB2n00nvAKIjACQU5aMUzQNIKITnQwsu8qUyCX1WVOT2nHdjkaTppkAQJvyzEz/S19434tf8o4f/uiKXm9m0JuTHH99+c0//+VNAKKSIgdCQpC6WUop77fxAWBd1kzbtjlbXS9RiL+98mpVQFOAiNz3x4AphhD27F568hPPcmKBQ/imaD0wtbZN8ytmP/Of73reC99+xeXXx6IczMzH2Lvq6lt/e9XNsBdl0Y6KQAi4vLy0dv06RJyMmzMfeMK//+ub3/DGf7rnnh29alCW/Tvv3HHrbdsAWQ3RA9+1Dmx7FpYktccds+FjH3vbfvuvb9v01a/9iCivXlXs3NF+93u/eu1rnul9nIruZRaKCAZ27YZzxEQyGCgYCqkjI7uCCNx9bR3lVZmDAwuJSB2k4/wGpwx1Xh8NRexVMZMogKS2KpFQTFrE8k/HSvgnz8OiECKAhcAhhCRhMh6aQlkUKAaqw+Fw96496/dZ5/tyDgEZ67resGGff/3oe5785JfU0FtYDF/+0k++9rVf9atw/LEHP+GCB59xxjGHHbJ/UZSTSdO2KWfzMa0vvFUNCXKbA5OI49lkVQlr96Fj18blptgxtlt3NLftyjtHGTlyiEXgLGLAkT0oAMGMA6FYCJzGk8SlAqBZyhJiBCNkf4rEpE51Ui6MOC1sgz33pHtu2v2HX2ojxqEM6bZbb3/6U89/xZ8/d2lxGTtSp1PnHM8gKgKAxMQccs4+c8lJCDGGYGBFEZ3p7jnsbdMimqj4MWYGbcpFGd3yklN2zLkrFTmELLkDyYsQsSnsu3Z2zcq8Yd/eTDk3N9fznIvJpFm3dtVXv/yBiy762re+/cMbbrh1z8Jdk9EkZQuxD4glF7OD3v77rzvl1NMf9rBzTzv1eNejDXqxLJYDlv2iEUtVnwIXkusYqMkTMwyEvX4Vi37JTVk4ucpU1MAIg6hIVkAsq3I0nGzcf9/vfP3DX/zyN7/21W9cf/3d43FOSZIYUTDTGBGsnpldcfzxx516ynGPetSDY4xt26acnvGMR59wwpGf+MTnLrvs8nu37VpYGDdt5jBACgitSg6E+6xbddRRhz3xCQ+74CmPKMtiYc/iqlUrnvuch1/2i9/efdcQ0Z7xtOc3k0ayWAyImNrkCSCE4LgbQvMplvNRPcQbDH0A2iXimLpyxQCYAnpOMBESsqFvOb0ZVFEm2rV9B6xdM19W23YvlSuqbfduGS6NOVQq2aHYdv8M6z6UYnXinaKIbWJEmgwXF/csUtpudnDVG6xcOReY/Dbz0SOHsLw8Ov/8cy65+KK3/OXfXXf9Xbt3L1S9QdsMfnX5+Ps/uGzdvmtPOuHQBz3ohHPOOe244w4ry2I0bto2NU1bllEAzMQ1x1Wv7HjoiE0WAywxH7qSjlhdDRvdupRv2Q23LaAImOFk0opCFs05g2RAKIqIHuETQmqTqoIFNExt64FUFAuhqgVr770Vdt/Z3HrNZMc9klUkL40Wm3rx8EM3vPoVL3v1a14qWVLbEHPVq1TUVA01BJIMIqIqiIGJkllqMxOFwOoIJ6LJuFZVd0GjKgcy7Q4vIkxtDjEAYIxFzlnUmcusajEyEWGIgECITW6JzEw/+E9/hYyDft/r39HyiAI7XNBUX/6y57z4RU/bfM+9exb2bN+x+8477lbQubm5/ffbZ78N+65bt24wMzCztmlzluFo8sxnPfrhjzit16sCB1FBDzQLseqVqW3aNoto1St8vDa/YuVoVCNiLAIopJxjjIrJWaOIlFIue9XLX/78F73gGffcs2Xr9u07d+zatWe3u7v322+fudmZDRs2rF+/JsaiqZuUhZkJYXFx+eCD9n//+9++tLS8ffuO22+/68abbt62bUcs4oEHbpwdDNauXbNx43777ru+LMvRcFxPmiKW43H9guc/Yf/9115zzbUnnXjsmWec1DSpLGN3g3anFXpmGFgWEeJARCLKgc18qC9OIoxF7AgyiI6E9mWoR+c2dQqh2waaGRr0elVgXrV2jQNt2mxp7KEbljLkbB2h/34d1n11XKmKB9hSWYQsJQBgIAkhSeXMuMl4Upals5yISMVNoTwcjc584Gnf/8Hnb7jh5h/88Gc/+9mvf/WLKxcWFwaz65pEv/rN5h//4vaZj1585qlHn3feyQ866/gjjzqgLApVG+o4pda1gk2TujEQmbM0MUZBSll6VXFYL27aPyyMs6maUa+KGSSn1kzLyCKMxGY5q4okA2QOkmXSZFVRtWaSod6ett4md10n226CLKK4Z2HH8tLWMtIpp5/wvOe95gmPf9T69evGo4mSVr1eF/rZ0XqxbZKaqlpOCZkmk5oDOzcUlf0/FELw+BxCqOu6i3AlijESgMNknKunKjCNZpAudwdUxBADkYiWZZdONpiZ6fUqD+Z29WY3AkcUsd17FjiE9fvuu/GAA0L847hXxbLknGU8mkxn+8rE69at2WefddpBP7tu1FNCVTQwqZmIhBCgkzkhF1Gn95MzvNxy5MVEynnb1p0h8Ib99tlv/w1Vr/gfOhkAgKauJ5OmnrTsdAsAEUUgv7SIaePG/Q868MCHPezB/4/nsW1zPZkMlychME3zHUbDySMf/qBHPuIcU1taGvb6PRVJrW8hDcymMEVDooDEzDl1+zskR+gYogZmF38yEyGqWozR3dE+zo0xTsfo2LbJtRdWRBWJVWVqc4O+lFHnBiGWw/FI0R8XVL3/wLqvZA3a0QKwTblpag49KgaqWpZlWZYhxpm5uaoq1QwV9ppHc4YQeDQaI+Cxx2464YRjX/fal910062XXvqjn/3sl5dfcfXWbRPkinHNT3954w9+cuOqmS+edOzGcx9yymmnHXXAAfvvs+9qQGzqptHksxsEUFN/REIgQsopCZGA9Asyw5QlBjbN/YrGDhRw170iALZ1nRWKqmgAEiK1Y9m6WW777eiu32s9gXa5gbBz+z2mi+vWrn7mS5/5rGddcOppJ/V7vaWF4Z7di7GI4IYPQIHsgVdu02NiAK16vZwzRsopmws1RQDAiYAeF5qyeKQdIdk0vJ6IxNTzEVS7PtHfCfdLcwhI6H+auW5UwSMekCgwqaqAaVbvOr1YU9WcUtu2plaUQUSZ0X2+IspMqibmDFdrU2ImE4+jVsecNW3yzaCCeRgSIoGJOaYAOuhNoGBmWRQRm6b1Vx0Ner3SzOq6NrPxaOTEvRiDZAWw0IXiemAQ7qWwIxGoIWNqU2OtLGedyvqiM8KYiNBDa7rkNDAOvLg4BAQzLIqg6rhBFlHpWjniwI4Mo0Aq6vN4NUPpcu2JUT0ZTBVAAZiIU0quNQ2RswgzpZQ5sGQhxrIqnb01MztbVYUqpCShsonGyWTSaXoVPMvj/gPrvplhdeZn8aE7oORxEbjfH1AzbNu2nuBdd2/bb8PamUHPFHyF627YDiAFOplMPG/y0EMPfvWrX/bqV73srrvu/vnPf/mzn/36m9+6dPM914RYUFrzy8t3/PRXN/RmZ9asKM8959jzHnz6iScftXH/dQAwHI6yaNsmRCyK6Es0IkJyBrzhH/MyjJmJTLJkMxAhwjZZppAJ27pN99xa7b67vvumxXtuSvVu5HJSLy8u3N2v+NGPfPATnvCohzzk/EMOPdjUFheXmknrn4DkDGYeVydi09gHH5CDK4CI0FQ4RJ/OeifbtqmjCQJSCG4RRyZXJUF3KnXUY09zcSynioS9mynpiE6IgBzULMYAiC6tImbJncDd0ABAsnBw3I9SpJxzEaNPyT3zigOjGBk5ENffQCBCVbBOfV7EYGqpTTF2UT1ITjomj/YxNATsonVUiLgsCjNz1o0TREOgnDLHAAimKt2RHZg9FA2IusCLbmJKrCZdT2bAzF4hmhoym7qSuUsICCGoZD/znJ8HXfoOUSRTVQMy82Wfi0L8EfFez3nQOYtnR4YQchZmdJ+554+hemokNJOGQzDTWEQERP90zVQtxpByahrr9fpIkESbegjaEqg0E5jare8/sO67MwvAyipWvWr3nglA5hCauk65RaIQw6GHbkRTJ0OJWa47RVIIgUiZGbtcKRsPRzkrB1qzZu0zn/W0Zz7raRdeeNfFF1968cXfueqqa/fs2RXK1chrN9cz//rvl3zyU9/fuHHN+eed+ohHnH7yA45av24VmA6H45TFhz4hMMh0m+lK6JQAIYtKFjADVVBNxrmowmihvvMm3HJ32Hr7eHn3uF5q03DbjtsBlg85+NBXvvzVz3zW0486+mhE2LVj1/atO6pej5mnKyYw10anBObxorkzITF7AWqmsSglZ0eDCoA5Uxg6gYIThHUaUOgHjYkCQNM0iBA4eKWBxC5VdB25x9mbTvnHZqIqdQ4hmvpcptOhIqGJqRgH3msY9khX8NZHPRHDJuO6IyB3dC3IWYoyBmAzED+xVAnJIxsweNeJxiRq1mZPh8joETXY0ftUVTU3DRLF0C01/EjyVtBXop6BBgAqknOOIez11iQV16Aik1p3lHfhaWYhson5oW+AJlOxvgEzOekYECVL27T+BBZlkVJSMzSQnKnjJRPhNAUMLcaoKqlNjovxCYT4fgMUAbXjkVLOikAOvDaP6mAywKZuXNZDiBQYDEKMZVUSUeC+WvoTet//xINUOy8hhVgCwPz86ljOAXJ/MJi1mplS2959173r16+NMaSUmbnXK/3RMQBnhKcsiBhDIDJiA4CU0p7dLTOuW7fuL175Zy996YtvufXWn/7055d895I//OGmO+68hXAuzmzYuXvy6c/++HNf+dVBB649/ZTDHnjm8eecc8LqVYOVK+dFZDyqfSXHzNF5LxRitsl4NFxuJGcsKm1yu2tLe/cNsOU2XVwi1DrXO3fevbDnjqoMD3vouc993tMfcv6D161b06a8e/ceYir7HQZ3r3OlcwizQe6uSiI281BCRQQ/rIEpFgUTAULbtAZYlkVOqQOQFgWAiYiImlv+Uo5FwcyxiB1YSsS8MNAub8I8yF48otHcRssemM4ECJLEJzWOdMoqZsBKSF1+gqpK7vxVotp5lRGJKcTgWTmpzoE5tWkvFAG6paTRVOgtqq4/6qCdgByDZOlKFSMfqLkBCMxMDcixjh0LzLUdJp31DzuWDkzNgAigSBRj9EMHDDxTCwBEJadcFNF3EUgIphSC804JyVkMAOjfnesnXFecU0aSnLAoozujwSwEBoSmaUFBIHt+V1lVIQRRces1IoRQ+KjPsd0cgrPewSyD+I9dMlZVzy+nflm0BP3+gENlgLHse6gF3S8cvW+LrCiiOaWU2iJGM10cTUqrCZGY169fMzPTTylXvSo1bU6CpH5HCRIRimgMwfR/GMoQkRER2ratd9Ux8EEHHnD0yw572cuev23b9p/+9LJvf+uS3/3u2jvuXBjVTQi9G+uVN9987af/85v77rPPsccc+PjHnn3KKUcedNA+8/ODnDXl3LatKwzANMZAVYDx0vjW65ZvuKrZegc045zr8WRxYeFugMmJJxz95Nde+MhHPeL4444qq6Jp2tFoZGZe7XejjS7YSv1ANADJHgCjaDrdAzCAiIq/4Z1gzQy631Tr8RgJA8dOHJ8yMRGiKDIH0dSNpdRc/Bm6WkN8RquixISKqkqAfqUDIqp7A3y32C21/KXvxt4eI+Z6VCIiJsI2JfcjqXX5C25RMjEObIagQETdaWqWknh6BBHHwO4qnUYEqgGMRhNmZO6i5P3s9rmeqnb0QXHslyERGIJBiNRFMXeyYRXJnpGNiKGLjAU1Q8AYg3aBQMihm80ZgAddgFlqBRGTZQ9PC04T7bLRUK3bzGonX/AkQZcoZ2LcmztPZFQwIqp/ERymGyQBM2eNem22d2/erU2YUs5+cjHRcDicBFzcs0tNkZgDGqauzLz/wLovhKOIIQSzNhAwMlEYjydEGBi1NVFjkeXlkRmURXTyuoKRAjCEEImgbTMYtG3yxJcuKiqQimdnGmEwgMlkUteTEHhmMPvUpz75qU998p7di1ddedV3Lr7kF7/8zTXX3rA0HJfViu3bln60ffuPfnzdPuvWPuABRxx/3P4nnXzoSScdtmbNalUhxF27l3vD3cPfXNneem2z617T1I6WFpZ2tfXCuvWDFzz3cU9/xtPOPueBg5l+UzeLi8u6oFVZOmFF1dSUkfam+Pk4BgAdZTmVoXfrMQBxFSJMmU2SXfVjIUYVMbWiLMBAwVNLgz+47WhCzByDq8xiDMD+TpCzMDn4/g/NIIaYUvJLHglTSg4j91SIDt5AqF2mDmk3VJqCNLGL4/ZlvKSMCNOlpDIydEdMZlcCM5mBiIXA00hWy1lEJXBQVUNjBBVjQiYmYtfZd31ZB93yhEZBBO3Kky5hUMVMFdnjztBjxDwGoos1DRFMParUm1MwUMtT9iERkSeemXUqBOp2cKYqKuay3hgdfIxmwMRINNUW/PHEI0QREzVPvRVVAnI4hIlS173u9f/D3q1ol1LvfyZCm9oQOWVW1fG49mkFgQJoahqaLe+XNdyX1hz0cUNdJzV2nWRkztLNjCfjev36taEjdit2SBAF1CygphyYiV3q7XtzkewPFrPDOc0YqXtjbc+uRUeEPPBBDzzvoedORvXvr732e9/7waXf//E111y7tHxvDCt3cfr+D7Z+93sRUQ86cPUDTjns4IP3u+fehauuuWVxz7hdXjLQPUs7d+64IUY++6wzXvSi55x2+in777d/UcbJZLKwezGLllXpJw8xq7p1xkSFkDBQZ7BR5WBMnERFsk+gp+v/vXWWAbglKZhaKFizokd4SZ7GzZvH8xFi1asYCQlTaj2k3t9zT/TrnNUAAJBS1m4r3z3xVVmlnNumCYE9BqJ7Jz3YGtBU/V/66NA91dI4Ht6F/kDYxYv6XpAIPdpnbxaGl4QA6Cv5tm2LskQiJgb0PtF8GaciHDjG4KNrX26aWc5pKsKEnNtu1DUtVxGQQ8ii3t4iEgffmVpOydtSZu4qIrQiFOA/T9esg5da4kz9JEVRuDe5KMm6mHpjopQyYPcLqWRwMUrbEhEwAwAxhhidzEXMpiqgIEpTgpAPTAWkC4cmJgSXDbP/lCEgUs5tjAUgM1NjBVDBsXQu6f+Qc9x/YP1fV43612xNkwwsMEvOO7bcu3rFwfMBI9Ps7OyBHMG0bQQRiFi6ZHNGAzUlQMmCbDK1cXi+t5cCqckco2f+CmAnJmZOOSPbcHm4vKQhhpNPPum00095wxtfe+utt37n2xd/85vf+/111w+Xd4e4z9zs/nfcAbfetjMEALJ+fyCp3b7ttra9e936NS/+s2c/+9lPPeusM4sYhsPxeDJuWg6ROQagLhtK1XLKfjgTAoXoOXouKXTXBdJeoDd5YaUqBuiGwem4y0xNVaVVAIgxcHBXGgKYijeYoWkaImpSZmIKBAjE2KWfGzCjmWp2eQgwk+QMAN2isUv90w4EOGWQeoKLinif4ns077bI2Cd9Uz9wRiLJqqqxCAAoKfsR46n0IbCr+BXBTelEFIuIYDklH0WJGRITkk5xVCKiWYiZmRyW4Es/zxz0sD//P1KbYhGyZFMQ9aRS8q7Qf3JiUO3GQ6raRZ9CR8vyE2d686mXRRzYJWxE1HZx353oBP9HddTN3gADswGkti3KQlRNZFolg3Urk67IRuRYIDN1YbQqIuIzKfYEAxEza5omBmZm4kBUqalKruumSRkpWJc0cv+BdZ8IsVQVgEMIIdDycCEWYW7FiiISZBKzlNLCwuKqVSsiI3TuLQAEEDXCnCXn7Dt4JM+R4Q534Hd7JxSQTlhsCuJxx0SIXBUigoDD4QgQYgybjj76uOOOefWrX3nb7bf/8peXX/q9713z+5vuuee2piXJBpj6vf5gMHPeeUc87elveuhDzj3ggP3bJjd1MxqO3N9HhJKzqnm+gzPjXJ3vP5XjFnL2wYoFDtb1MwDepPjrjGRqwGAGKXU6b8ApuxMxp4RI0xcJRcQM1BIHFpGyV6kI+zmlklVceE1IbvmRlPxI8rAZFXGlkpmEGDNgzjKdkTMyhsAWouScswCIgc+APDiSVFSyhCK4si4w23TKHkLw6ZgLJjsti9tCEcyUOQZgEZlOycgTg/wbDIFyVmbCgFkyEjpsHgAYMEma1rABEDgGFUkpYxdU42CfhACxKMwst223yiBwqaqqhRhSm6qqEkkuQ/Nvw7M/us8QXBTSqYtd8yFZ3GqeUjIFDmQ+lDIVkbIsDTAyZRFvM8UB2x6DBOarFV8UuLCui6MERKLUZg5MCDml1LYzMytFpE5ZHMZq4neUmUz70PsPrPtmjEUEkNvUIJKItW2jCjlriIyAKhlM67rp93qSxS9AQiQk9YsoBCcQEQAiZBH0iTAhAjuzcfocdBV/iCF0Hcr0YgQnq8B4MkltS0THbDr62GM3veQlL1hYWLr9jjtuuumWu+/evHG//Tbsu8/6fdcfeMDGsirG48me3YtghkSBAxGZhzsgdm8m4bTHMB/rdK6vztraQTIBgMDz0o0QBPf+R8xDN2MRzcD3DGpdkeKNwJQwgwgEBIFZTXn69HMIXQ5V7vQBLl6bbidI90bpAeL0xzYA4i5ggbpiyrowUcSiiGbmn7P/+xCIiDB2Pmd/J8HlDkS+JovMPqDxbsv1rb4BRAICRqIYCzNtUyIjfws9lzAENAAwjSEgkoJ5ppaaBY7+m7owDQ0CByXzKREgOd95bzfKHJC6T54D7a2oYhFT8rNMQowACOqjBgIEN3X6H6Iqvq0WgK5GBoTQ+ZlTzn4Vhhg1K5qaAE2Z9zGGpmkRqd/vi3bsQ0QiYu8AEUBUXFcXAgOSrw5DYBVhDiICmGJV1ZPFerSc2xbBAOhP6H3/098SIgBQjBHAYhHAtD/ol0XQ1iNkaDAYzMzMqIrT+LykypK9N+wmNI7X8EaCUER8WZ5S9tBgB6dgQCAw0azJoXdTPgF1KZhmIcYYw3g88TiGfq934vHHn3Ti8XuvseWl4Xg0aZoGkWKnhHKXifigV0VDYJkqNmMMAD5AQUAsyui2spyyqfqQwjxpPbl6CVUkxoDIHjtYFNF8B++5hC538vkOd9GwHirjqzSf+HJg62AMSISqprLXDq1lUSIREbudzdPtPJow58xM4nNokBjCdMUJKefkKQtmsSxSyr7kYMamSWBKhIaQs3q6hIl4K+sQvlhGle4E8Uxv6kL6vMo0EWX0JFSLMZhqh6UGU9MAIecMPpkW8TKTiX1YNu3QkABbab1d9OpGVDRrRy5N6h6gnCQEBjORbprmea0qyoEBwRQ8Lp66vYGHCUpZdrZHxCCqKWdQI6bUJkBQUXKUIfmYD73cBjBjCiF4fpoP47BrS9HU/FxzN7uP3j17CQHKqlLTInIkmIgMh02Mvd5gZjxqTASM/nSEo/gnP8PywA9DaHN25fvObfds3OcQZnR3VU1pOBxxYOiCqqiTuQOoKXWJyg7h6MalgAhqYopgZVW6XKuLDzCgwG65QPIcQ805xRgdEEVdAQJqBtlGzcjzS8A87YZMNRaupTaYap+dBGJghFhUpUr2eiTEgFOLDDEBmGRP7qEYg6/YabrQxo61A2VZAID/JGVVEiFkpSKAGQpmUURy8b2qhhgDsxNORDXGgApEqJKROKvTlimE4Ceuh6G5NdpH2iKiol7/hSIQoxmYU4ERs2RCaupmWjB2yvWcsmYJZfRDMzBxDGaWc5YsSuTDNUQEwy7G1bMGAZmJOExV4PRHCSuCaKfV8kQJ8M8XzI/1oihEhdw3g1gWRXemICE6rlPAoCiLDtGSIEsnPigoIiFA6MLmRQE8hzk4P8+3hC7+2Is5dNmwTw+ZvQw0VdkbWhH3ejm56yed0wCIhEBEbsvIbaJp3e3TQI6Bu6xcpgCSRURg2gZS4FgEF4suLy/Pzs1JVgGcmR1s0yzq/vZIHA3uV7rf11tCzdmIC2vqWPZn5lYFxLIIPj7YumXrypUrBnGgHXLTEIED55R9YgUGLr2ZviFAzusXy4Am4nVLN8lC9GvfpeHgIwNAAMgiZmaWQwjYXZKARgjA0A31RcRHZq6E8r/UB6Wez2wmOWdRQ7MYoxJ22lDH5irkLNPpq+sPO3i5d6j+5zDz3vUCAqhozt3o2olvKqoisSw9o9ztSgCGKF76+fmNoGYQnDM31S6RSxDwj0t0MPCCSLJod7Jol4g3rc4QMYZopm2T/EYwAwqkat0hg6ji/2NImFOORVAxV8AiAnPwALeO/eSwOgBEzz1WFfS9WM7CYcowEAVQMAjEKeUu6oLZu3yfc5mqovmknAhd6K/Q5f10JiZXkLa5+za6kxS69nnKd/bxZ2fcCe5JMkQyNN2razMjBEJWL4W6yb0/DNQVgKpe3wXEnMVvZU809JkXosYYcxZT42CS1e2Q/kRxRNVu9rp390JM87MzQ5PDjz6sN7eCmShwB9W+vyW8rwZY6ObegBaYAQ0tx6IMsYBMANg0zczMTL/fdwVw26YQ2MGMROzaKGR2fVDnXZ/WbgBYVoXffpqlV/Xgj7HeXtOYKyfU490BXTzxP7CC5nEYiFRE9rA8Xyi5T03NDMy9bxCxOwTBIpGJxrLwrZBrlHw9NB08+zVOjiXxhgI9ppGmkVNMCJ2iunPhmAFi7Cw2KFm4Kqb7JyDCoujZ3vBhl6GqIRhOFfQ+PHZzNZj6tJ8D2VT77qrOogiqata9DiF01RwicCAw4BC6bbo3k65+NAwUMKBIJnKaXWfBdBQqEsUOVKAAQAXnlMEsxsiBXZ6KU0kX+UQ8MBKZqE7pndYpnow5mhkCApEr6QHR7dZePAKAoyxU1eWohUMgtNNSuX6EXTo/netxZ7BGVQVmAxPRwFSURSfLoK6692Gi3zruhRDVvW4Bp7yapz2xDxC9O+aqKnF6W3eNwnQA6h1oztrtEc0MdOWqVQBoKpISUVjatUWakaSc21alNeA/neRn+1OfYXXVAIJ5fF1OWVQkt6lt2iRVxBBjUca2SU6q8z7Cv+4Yok/BsorXxTSNw/PlEWInJlYwyHma8AJTrhark7PVOPiMX8V1yeRyG41FJKNOt40A2ZC7uFAfUziTwE8T11GqmqQMYFrX/t52Zg5VQN80IQeC3M1ELHWbeyCXU4ecEkzxAqbTyboLMs1tzOw7JjRExKZtmckpCMScU967ZTCApm2LGLFTM/7R2u8BxioqWXPOsfAJPXpQbTfW184a1xkARQEgloXvvHzJ6C2SZoshAPmWwxVaHvaa9z6siIRIDpbx24KYc0q+AHVNgKh2QFTAlHMIjGbEJCmLaIwB3KsE6MxnT7pKKampl6UcGAGIMKfcqnrVF0IgYi+iGQkMxMTUvDvLIr7NVDBXdHmDP73UQFSnV5z/8BADizORuwmX7P2lXD1DgHuHm95Uep0tWTiQiMUiitseAJEgJeGpjp/ZA0Gyf5Dbt22fm5+dmZ0d180Yw6SFlFIsKqIRhwgm928J78Mplpo/yhwDUYjFfFGWILXkxExlGffsWRqPJyEEFSOCEAMHJkCXjAJ0PBNXIndlAQAx55x95QcANN2CZfctw94ODEQ7fBoRI02zyKeQgDZNdQS+RGNXfndO1E7fiZBFmNknQUUR/UkVyV2K6HQUIyI+bM6p2w+pqo/eutcpS7duF1cbqfk+ETp1mSMuHQ/gU3kA6PWqlHNqkxmEYOzyCvT6UZpa2jbHGIgpZ/X2WdVUxWNlPefK1Ny0SETagQc8IhCnHxUCIgePaenW/NgtyxhApr+gEiFRzCKAwIRq5i28KywAIbtkXxQJ/mjuc52D2TQyx2F27IZHAAsxGJiKBWYAlCwmexP5cK+ZvG3aIoa96TUhhgBh7z2nIlhEEzOPnzCyLs0BkTCLqrZZuk/YTfUq3iKrgfnPH2PIvt/ImlolJg7MgTt7Uzd9++Oi06k72JXm4HQKh9/6qtHHoGZGiGoKCrlOsSgk56xQxI4XONuvRKs4Pz87v7pta/SsIOtIZ/cfWPeJDsvnlKHsVVXTq3IaL+xZmDngmJwbU4sxDgb9Xq+nqswoOWcRJhIzJkYk89ejGwP5N2cUQtumbl3f2V+6gpuZ2jb7VeawuqlImvdaYkwtZ/E4rL2ih26iAWBTYidQZ0lz9nzq2lXw900km3eMKu5c4cDRrCyjmQ2Hky7ZmKbeN4SqjP1BlbLVdRNimGbGYLcBgE40b51dhnyp55JOVz+4zhDQvNQixJwVEYsixBjrpvUuKYbg7sDJpOXAKuZBqSEG89mPGiJ39PfpAe3moS4bsqsoEab2ZYrRC66ZmT4RNU3yGhEQug2ul4RMoFYUxXSnhwagOSPuVdujg5t9S+t1HBKaYkqJyT3V/m+sW5NMm3ybjshEjbHTTHkUkLd4zBRCmVyhSl3V8z9kcMCB52YHSDQaTZxjtVdAryocuCiKlNrl5VFZFKnNgECBfNdJBAYwN9sn5tFwLGpMhIQqHWrij5MQBUBzNYJvfmha+gMSKhpYKApiMmAFLcuiKKJHm6yY723bk9o2EWGI1HEp7udh3YeaBgRQlSalpmnGgbk/O7tcS3Tch9ny8tAMev0eEQJ34YAAioSaO62DoJnZYNDjQDnluknT3fkUBdONC1BFuIsCAAWQnDEEoO4r99eG0Fau6AMhGA6HtbMPwIiYisBF2VOzpskel0KE00EKugrRANzu58wDyU7pFmbcs3vxxz/+Xd2kc84+YeMB+7RtQkAmMjBmvvX2zT/4/hXnPfjkY487rJ3+CoQEYGgEiK6Y7yYsPhELRMR7i0Qfkag6zMBSamdnegAwnjR10xSR+4N+26a779p2xW9vblM69+zj166Zzyl1ZYt2dRwwdst1Ea85XS4FiJKze4aA/LNCDI7GV1WNkb/w+e9u3bbrRS96Yq9XppSnyWMGYM7DVwL0zR9gB+Qyd/PoXqmqS8D8iOn1et7KlWVs6tS2bff1+YlpZmCzMxUyN3VLhLEowXQ0qmmaD6giCEYO0HKogxoHUlGGLo3J+9qyij/7xe/uvGPLYx71oH6/UjUVow5/qLffvPn73//F2Wefcsyxh0oWb4fdgUBMknNVxe9+//Kbb7z76U8/f8WKuZxzx6sRdc9DR1XtsKvkhOsp0AaI2B9LQAzM/qn6hxVCELGlcauQJxPJuS6r/mTSAKgr5u8/sO6jE4uJAaxtGpEMiE07NrAVs725ttc27VJulpaH8ytWMJPzgk07qlk38yAgYJFcxPDhf/7iT356zZFHHvjXb3thapq2TR5HitOd9zRcBgC8+BdT8J7LiwHXFo+Go/e9//Pbty2vXT375jc9Y2oMhl5ZfumrP7r00l8z4wue/5hTH3D08nDsY6A2JVP19aI3e66F8NGPtFJWBRPuWRhe+NaPL+4ZffqTb9l4wD51nULoYJUx8qte848//fFvTzrhiO9//yPUEdnZRLK6xZ+RKKfEFEUVAafYLAkx5CwukGzbDAgxMBH2B4PvXPzL73znl7ffvmUymaxePXPCcYe+7KVPuuqaW17wgnfOzvW+/fV/WLN6TlSdtQKAObelj4SBOmW2Wc6ZoNt8EZOaObSembPlqUQWelX53z+5/CUv/ftJPVq9au4FL3rieNwQqx893rkXRezsjYiq2rQZAKan/DRrHhDNiKgo4o5duz/8rk9dccWNTT1+9CNOe+WrnsrsszafARgz9aryy1/76Q9+eNU9d2/uV+GwQ/d//OMfdPIDNrVt8k4zpawisSwCQPai3mdNTE3dMRRzln6vvOHG2x/z5P813LH1Ix9+68v//EnEhMRt2zZ1WxbxTW/5yHe+/d0zzjjnxz/+iP/hMfoUFVPOg175u6tvveBZf1vv2aGSX/e6Z+0Z1/4E+iqTiNum5cBonUpjWhj6becTDQUgpzqoCAVm4l3Lu5FwMBiUMYzULE9QkuYWQQkZwZVf9x9Y90VH6BJwIqKmaSRLCKWqjcYTW9yNeHCItGJ+rojRbc8uoey6Oe9MVJlNRKoyXn317d/42k/OfcgZzBxiUICqKqoyeoUwmbRNk3xn3+tVZRkliyEOl8c21d2klIrAy6PJ577w4z27RgD5IQ858ZxzT9q9a7Hfr3bs3P337//c7bdsHk/Gp51y9FlnnoA4qcpYFNEPnZTSaFSbWgikCmUZioI7ILpa27RIvGrFSkuQshZFmJ/vh8DD4aRt06BfPvuZj9i5dek5z3lU1atAIZaxbVMynZ/rc2DJeTisQww0RejOzw+YSFSHw4mBhViaGrH6VLiumzdc+NH/+trPlpcXV6ycXb9+1R233fGNr39xn33mjzjqsKqEQT8Soy8N+zNlr1+aas46Hk/cy5dFyzL0qhKZRGQ0nHilSgCxV+QsIjY72wuBU8rjSSOqhx120MMecoZKPvOBJ9aTptcrmGk8rsuyKIsIaONxk8WZf8CBZwYVIU6aFsyKGABxOKwBUc0IYDgcvfBF7/3Rj68+dtPBMcA7/uai0bB+53teNpm0ZVG0bSuSi8ivv/Bf/v3jFwPiPuvmNTVf/do3N9+7/aLPnDSZNGUZy6pAIslSN636UxSDi+CauhkMqsDcpmymSLhy5fyJxx62sGf9cccdyoHf975P//7am1/2kic96OyTc5YHnXXSby6/9jGPOYOZiyJAwcyBiVxPpmarVs2fcsKhud33gQ88Pov0+6WqjidNWcRqUKU2exnumt5QhP6gIoS2STlLDGhAZuyFGE4dV6rW7/eZWUwlW5NlebgMzIbAIQaOBi3i/QfWfVJemWhKLUAgKnyhNjM7x0TD4YQzSRbuh6IsvEtXMMvickrpCL+gYrnzstr8bG8wM5ib6bl4em62f9fdW6/7w52bN+9ctWpw4gmH7rvPOmKa7fdvuXXz9X+4azxu9ttv5fHHHUZMktXHIkCoYkXE2ZmwsNR88qJLzzn3ZCQqy/iFL//8lpvv3rB+5dZtMqlbABj0yuFocvkVN2y9dycgHnLohmOOPRTMUpLBTLWwsPTrH12/Y9dSvyqOOHy/I488QLKmlCb1JBa0Z2H5yitvioE2HX3wzFx/Mmme8fSHnXTiEQceuK+ZLS4ub9u+e37F7EEH7vubK2++884t++676rRTj0kptU3T6xWTOv/hhrvuvXf7yhUzDzj5yH5VLi0ux1hwYMk6M9t/w5s+9On/vGT1yrkXv+hxz3v+YzZuXL+8NPr3j3/t7HNPve22zSKas3sMbWZQ7dy1eMWlN47reuP+a0464TDmMBlPql6xZ2H465tvWB6OelVx0gmHzQz64/GkTbJz18KKFbP77rPmit9cv23b7gMPWHvscUdMJu2++6396Mdejwjr1q5smrRjx54s+YgjDrznnu3XXHMrE5x88lErVs5OJm0RQ9u2v7niluG4PvKI/U1t1+5Rr9c7YOMq1wpUveKqq2+8+qrbN+yz6nOf/av5+f7ZD3r9D39844XL4xjDZFIjYq9XXn3NLZ/9/A/7vfi61z35z//8gsU9S5d+75fHHHd427S9Ki4ujq+6+tbFxeXBoDrmmENmZ/ptm+q63rZtWFXFunWrbrjxzns2bz9g//Wbjjl0NBzvu++aj33ktaZw+GEbF5eG3/z2L375iytPOvHww488IIT4kpc88ZGPPPWQQw5IOW3fuUeyxSKmNo0nLROuWT1/4AFrP/rBl8fABx+83/LyaOfORUQ85JB9d2xf+PUV1wWmY489rN8rR8NxjNxM8nW/v3VhYXHjAevnZmcWF0e9Xrlu3UpAcDu3qTgnx61IjEikg6i2ZhXHqurPARBxtxi5/8C6L0SjU1uGIhkHFklNPRqNJjNrYnR3Ooeq6qmJZAkh7J2gl1VpHqmAOC2nzelxKamoDgblx//j2x/4py/fcdeOts2meXY2fPFz73zYI896399d9K8f+8a2rYvI2OsVj330ae9//6vm5gaeKWDKopqz1E09M9O/9Pu/vep3Nx1z7CEqevH3Lgekts1tmxcXx4D4jW/9/A1v/Pfdu5ZF28Cwcn72jW965stf/qTAdNllV73xjf9y081b6rqNDEz0zne+4KlPe2jT1FWPf/TjK//X2z6x+a6tK1fMbjxw3Uc/8poTTzrq3e/6j7/9+y885tGnf+mL7/qXj331bX/ziSc87hzU8I1v/qzqxRjpL17x5L96+wvLIv7gx79917s/e8MN9wxHw7Iqzjh10wf/6S+OOnLjZNKkNvX71ZW/u/6LX/rB7OzMwx9+6kc+emFKuW3aVStn3vueVzHjtdfe2qHKmXu98uMf/+YHP/y1m2/ejARVUTzq4Q/4wIdetW7dqne+99P/8s/frMd13bRzc7NHHrnvx/759cefcMRHPvKl//WOjz/yEafO9nsXXXRJWcX5ucGFb3jO697wnD/84banPeOv25S+/pV3r14997gn/C9me+mLH/v+D3x1y5Zdq1bOHHHEfh/72BsPO3T/u+/e+prXfuC/f3atiK5d0+9X1e137HzwuSd/59vvHY8aA0hNmpud6fd5Yc/wqqtuKsryzru2Pezhp5dV0ZmfREMIu3cvEkkINFoer5zvr1m94uWHHTBcXgaAj33s6+99z0V13Tap7Q+KY489+GP//KbjTjji3z/wjbf/1b8/6UlnV1X/c5/9Xn/Q6/d7b3rTs/7i5Rf84fpbLnj62+rG/usL73zHOz9+2+1bDjvi0L//x8//zXs+efWVn/ns53780Y9+4XnPe+z7/valT7zg7TffvGVmtur3yiLwtq0Lb3nTM5/73Ic88SlvH47ar33hHUccuf/DHvU6RHvpi5/wLx/7xpYtu1atnD/umIM//ok3HXjgvr+78vo3vulffn/NrU2arFw1u2bVittu3f6IR57xxS/8zeLCMBYRAMDIxxrNpHGFYF1PBHumxmhNPXRk9J9ULOGflO/x/2tL2M3dEYoYmYhD4FCZiuQsIimlO2+/Yzgcm4Gq5pRdkN02Tc5ZsuYsOXcSl15ZqGQDK8v47e/84s1v+fj2HUuPffRpX/vSX3/6E29+8uMffMopx3z6U995z7s+XUX+4Ade8V9ffee55xz71a/89F/++etlWajaNPidmjavWjV71gM37dy5+K1v/qQs4k033/XLX95yyIEbDjlw37bN9aQGgLIozjjlqP/45Bsv/d4/vOIVjwOTf/+3r9977/Zduxdf8coP3njTPcdu2v9j//zqf/vXNzzwzKPWrlvlAdRVr/+FL/zwYQ85+c1vflZ/UF177Z0f/8TFZjaZtJNJm3MG1brR2ZmVl132h+3bd3zkA6847ZRDAuMnP/mtO+7YeuddW1716o/ecOM9T3vK2d/62nue87Tzf/GL6/76nRel5H+8hRh+/otrJuMkOT3soaep2sKeZVNrmrS8PDRApqCAzGHVypkf//eVF77ln3fv2vXOv3r2t7/+3guefM53LrnqPe/5PBL3y/IpTzzrS5//qy989u1HHL7h6qvv+sQnL0bAxYVxL1a/uOz3V199w1+//YUPPGNTavVDH/zy7bfeQwCj5boea9vKeFw3dVpaTO96zxee+uRz3vrmZ87PDX73u1s/9amLmen1b/zQd77z8yc89ozvf/cDZ5914j2bdxx33MYLL3xK02RVVdGl5dGxxx56wRPPMdC3ve1f3/b2f/vLNz/9wx9++cxgUMTo9OTR8vjkE448YP/1OeXPfO77D3/Eaz/+8a8vLiz5TAAAHv6IUz590ds+c9FfHbvp4N/99pZ//MfPmtlw2BL3fvCj3//q19e/7rXPeMBJh9V1+/5//MJtt20OHEbLzWh5kkWf9fTz52arpaXJOeec+JdvftGK+dk9uxa2b1uYTDIiPvupD3vly570ypc+fuOGFUsLSyHo6WdsAiRJOh6OmqYBQBNZXpj8/d9/4VlPf8jrX3NBCPSLX1z38Y9/R6R969v+/bLLrn3Mo0+/9LsfefDZJ994413HHbfxwjc+LacUAqc2pTY1TSuigKAqbdOKWd2kts5F2RcF5gIRVRRM9X5rzn1yWk3NqgD9XiFZ2rZtmrZuJlE1qSBSzgkAB4PBlB/kCGsGNFXzOCYAy1mKsgyuPNBshp/93KVtW5926qb/+I+3zs3NqKTnPO9Rqvatb/1MtHnwece98EWPI2YF/c1vbrvssuuWlkdlWcik3QuuZYAnPO6sH/7491//5q/e/vYX/ugn12y5a+cLnv2MpYWlH/7kd1lUc37YQ095whMftGv30tLS+KhNh8bip4uLw9Fo8pOfXrn5nl377rP6gx987WmnHwNgFzz1vBjCH/5wa2Daszj+85c94b3veamZ3Hzr3V/+4o8337uLwMiF3EgKIGKpTQfsv+azF719v437bDxw7fNf+N5JbUtLw9/89votm7efdNJh73j78/fdsOaoIzf++orrf/Ob62666c5Nmw5eWhoR4s6dS6rAEQaDHgIwIVAHKQZTYjdR57KMX/rqDyeT0blnH/vq1z6zV/U27L/m57+49rLLrr3xhjvf8pbnSc5btu4ykEO+ue73196xsDACsCxpPBkeecDGr37lfQcfvP8ZP/j185//nnEtm+/due+GVWaGTq1B5ICLS+0b3/C0N7/5mQBw6233fOnL/720NNy1e/Hqq29ZuXLuVa98xhlnHhMjX/LdK/r92QeeeaKqEFNKsmLF7MWX/OKHP/ptjLg0HB174Pr3vu9ll//6hh/+9+UPPuekU0/dVE8aEVmzbuU//P0rX/u6f7z7zs2XX7Hn8suv+/SnvvOv//bmIw4/4HVveIap3b15x3jSbDrusN///vZ77tnh3uaUc7+a/Y9PvOXUUzZ97/u/fP5z3zVcnmzevPOgg9cTgokAwjOf9cgPfeTLt9++7SEPPu3lr3gKgLRtS6zuqnjN655SFMUVl1/zxS//aDhq3nThM8477wG33HJnWTKZxCIgUQi0vDR661uf/453vBgAbr9jyze+8bMdO3Zt377rhhvvXLWi//I/f8oZZx43GY2/8tWf9arBCccdOhxOALAoo2v6AJSQ+v1ejIGRYiwas2a8HGJJGOt62IWU4f0H1n0iGiWiwAFAlpaW1ktGRNMkopM6zVJAwl6vf9AhB/V6Vdu0vvvfS4NFUKQuxsJUwUw0A1jObdO0O3cs5dweeOD6fq+3Z/dCR+8w3LVraW5+5jvf/fV/HfgUZDTLSws1EQyXJ2VRABoANE1rBiL4kPNPOe+8Uy7+1s8/8clvf+XrPy8HxXOe/ZDPfObHiL3FxZpCuPKaa9/97k9fd+0dS0uTGOL8XD/lVkTvunMbAB9y8MZjjz10z+5FX1QWMRYxEhMAnXP2Cb66mp+fndS1x2kNx22H0EKc1PVwOD544z7r1q1qmnZmdhaQc7ambm+86S4i3bJ15/EnPo84FEVYWp4s7Vm8++6txx13mJ/rRVECUJts+/Y90yhzNNOcZWrNZY7F7l2Lf7julsHs4LdX3XLI4U8t+z3J7fatO3ObR6PJN775k7/+m0/cevNmBZufnyliyJI9AWQ8Wj7o4P0OOHBDPWlmZ/oce62kWERCTDkjILlgyoCYzj77+LpuACAWwe+kqlf2+8XWbXs++7lL9tl31be/fdniYr161Qrv9yVrVZV79iy/+S3/ctNNd73j7S+89g93f+WrP3ntaz68sFR/+lNf/sA/vf6M04/NORdFMZ40555z4vcv/ei3v/XTL33pe1f/7parr7rtVa/60MUXv/+iT37jIx/9yk03bUkCq1fMVb0qS6fhrEejgw445sgjD2zbtGLFnNt0ql5hAKmzbcLy8rhNWSQtLo7qySQEUjMVm0waQGvq5obrb3n28955yy07XvTCR1144XPG44kZSE6A3qMhI8cQH3TWiX7vrlu3JosY6OrVq9evX/n7q2+7+Hu/PPjQfb998S8n4yyKop5DQWbWpuRUa1UJMaAb+BWq2d7OrU1bT1Kq26YxzYRGFO4/sO6bMRYiEQDFWHnOFFIsZ1YwUdMkZgSA8XhoTqo0SCmJZJ9qxyKyuEO4C79NWVRbyRIjD/qRwHbt2BkiFTGmlGZn+3WdyjJOxvWxpx3z9Kc9YveuZQUtijA726+qom1bt+U1TWqbmmcGVa98+lPO+e8fXvnRf/nWnZu3n3Xm0ZuO2pjVMIRde5Z3bN/5pgs/cu3vb3/CE895wQsfu2fX8PWv/TAShRh7VUGMO3btWVwYrl4zO5nU/V5hZilL1tw049FoQgRENJnUXjJKzillgDblBGZtm0Bz3bQ+p+ugyWbM7PvN2UHxhtc+TRQRaXampypHbzrYY41Tm047fZOZllX1hS9d+oynnb9q1byZElc7d+xu2oaQ0dBBKytWzEyG9TFHHfDsZz16z+I4t4nZev1i69adL33Z300mzSteccETnnDOV7/6kw996MsxBHAAIPF4kuvxJARu2jalNufcNKlNyd92VWnb5PKl0agmBABbWlxCzG2TZgf99773Fa96zT9+6qLvfu4LP1xaXD7rgce9650vNLA2OZAT/nD97Tt3Lm3cuP55z31cf1Bt3b5w0ed/OOj3jjjyiLMeeOJwOFa1LK6ttfXrV734z5744j974mtf/fcX/ef3t2zd8x+f+vbb3vqRlOTPX3bBk590/n/9108/9KEvGpjnSRLjZDLOKXO/MjVDUss5uX1dU25NTURUBExUNUaSLBwAcJRTS0iTyeS1r//nW27e+vCHnvKRD76aPKbDkJlzlpwcaSNmMh6NXYamkAHatk1lWb7rXS954Yve/YEPf/FjH/va7j2jo47e78I3Pq2D0yJw4MBhGnREk/G43+8Tx5xTaYrMTTNWFUA1U6T7eVj34T/s2TBZBr1+OxgEpogKhAUjANb1ZOeOHYcedjhY53YOoTQDZnFSOJmS+7OYQuCy7PWqXlHG00/f9P1LL7/l5ns/+9nvPeEJZy+Pxl/+r59e8KRzTzzx8F/96rrxMB+96eAHPXBTauWyX/1hzZoV8/MzddMyOslPQwiDmVJVzz/v+COP2rBlyx5GfP7zHmkYZmf6VX9W1O69d9vWLbtXrZq74CkPfch5p3z6099FjoE1p3T66cfG8LUt23b/7d//55ve9Ew0/dxnLj32hCMOOXh9EePMzOxU4y6xKEIsYhGJuVcWgStmJqKAHIoKyVS1qioPcGVEQjr7rOM/8P7Pj4dpv437Pv5xZ4nIDTfcfe+9uw45ZGNdN0URm6Y975yTzj/vpB/+4Oprr7v7RS9934tf9LgYw113bfvs5y59x189fzDTZ4qR44oVc496xJnf+97vJjVu2nToOWefgAA33HBXiOEP1982GjaHHrrvq1/z9A37rv7oP3+l6lWBCZD6vR6HmX6/x87tRYsFuzCdiWKIiCSiMXBkmun1aRq6MegPOPaLMqrZvZv37Nk5fO7zHnXmmcetWDHz6EedGUIcDccxdGyWdWtXzAx6dSP/+ZmLX/e6Zz7vOQ/7q7/+5NLy8rnnnH7q6ZsW9iwRYlnEe+7e8q73fPJJT37oscccHIrYJBCFVatmlpeWwfiwg/d/wxuft2GfVZ//wk+rcjAzOyiKIgQui7IsnDwFvX4ZigKUqrLs9YoYYxEMzfr9stfr9/qzC3uWJpM2xGKm3w9c9nuVqLzhwn++/PIbj960//96+/M2b921e+fSPutXrV4zW/V6gYuyLKpeURYFBxYVRGAKITAR+9F2zdV37Nm1/PKXX/Dwh5/Rq+Jxmw5atXbV8tKw6GapCGxu7TaV1LZSlrEoihhFGQBVUqobk2Qqf1qMqf8XAPwQgF2817S5acWAAgEAqllZlAcfcigitm0iFERMTXKQbk4C5pFvCghN205Gk8lkcWlptLQ0fs7zHvedi3999e9uecUr//G97/vkwu7h1q1LJxx7yF/8xQXfv/Ty3197y2Mf/8ZDDt6QW73uurvf+e4XH/+Xz26W2hDYTBcWRlu37Fy1aq5p27WrV55//vHvfs9FZ5x2wqMfcRqCFWUYjyZ33LFl9epVBx+y769/ff2b3vihv/vbi7ZtX0yp3rZj+8f/478+9IELL3jy2Rd95ruf+I/vfPvbP+uXxXXX3/XRj7zh8EPXL+xZ3LV7vLw0Sm0ysOWlOqc8GScTWR7WOQ+Xlyc5p7puU724sLjkmTpNk7dv3xMJl5eH55936iMf+cCLL7n82c9/58b916Lx9dfdc/oZmx587nFTuh60bfrwh17/7Of+zZW/vfkb3/zlt779CwSrJ9ls6YXPf/jc/NzS4lIVaWl5+PSnPfwrX7vs57+84dFPeMtBB6yrynj1HzY/7clnv+0vnxlLvvWWu84//6VEPBy1saDPf/H7T3/muaCa28Xl5QmA5dSK6NLicj3JqU0iujysU8qpTSoyGbZ7FtvsVnCVpaVxaoZN3TDzXXdtMaCf//yazXdvRcRvf/Nn55574hMef64bpEajyeGHH/CkJ5/zb//2rY/8y9c//4Uf79q12OsVK1f0PveFH51+6rEvfeljFhdGlYQ/XHfHpz99yWe/cNmG/dZoloXdy6r56U8//8wzNr37XZ+44+4tj3vMa5GKLVsWizL8/OdX//znv42RJ5Ol0WgMBqltx6PJcGnc1q2Xh8vD0eJi3aZUxGL//db++pfXf+nLP//cl3797//7NYic8yKx/fDHV3zmootn52a2bNnxlKf+5dLSuB5PLnjyw//zP9+6Z8/ScLiwuLQkIsvLzfLysttX67qejBrJy8PhhJm3bduFSD/60W/u3bxtZqboV9WJJx7xlKc8FIgkZ5f4cwgUkCkUVcmBTLUoYy6KEENV9XuDmaRKoQKTPx1VA/A73vGOP9kRliHCZz7z2VtvvWPN2kPnVszV9XhxcdfGkx5G5WA+NA89eUOvV01G49SmEH2+3lFQCNDHup7BaaYx0Gh5sv/GjY94xOknnHD4ypVzj33cWcQQSQvGE44/5G1ve+Fppx+zZs3Kxz7u7F6/KAtG1UMP2f/1b3jWU598TojsTnpEGE/qlNOZZ2w695yTAfDQQ/c79dQjn/GMhxx00D6EkEVXrRqcesrh5593ygMfePzOXQsicuihB/zTP71yZrbaZ58Vj3n0OUcdedBDHvKAVatmmcxA16xZ8dKXPP6CCx4MAHXSk08+/NxzTli/fhUz33bHljVrV539oONOecBR27bvXr1uzXkPfsADHnDkjp17Vq6aP+/ck0855egQ4/JyvTycPOCUw887/wHr1q1+zOMetGbNrKTMhOvXrXj60857/euftmr1DDivglhU16yef9KTzlm9en5m0Fu1avaAjWvOPPPYV7/6WY9/wrkLe5bV7PTTjzzrQSesW7f60Y964Nx8r9crQsD52f7jHnvGS178mBNPOOKwQ/dr23amXz360Q963/teGQkPPHDV2WedEEOpQKefsenM0zYhsYiOJ/n44w855+zjV6+cSymfeurh55x74sxgMGnScccfdNaDjl2zZoWZ3bN560GH7Pfgc048/oTDrvzd9T+/7OoYZOeuPdu37fr15Td89Ws/Y7aHPuyMlHJZlsx8/vmnrFk7H0OYGZQnnnDIO97xghe+4LHLw8loODz7QSf0+z0kWrN21QEHbOz3CzKdnSlOP/3It771uc957qPX77Nm3w1rRuORZNl01CH/9IG/qApYvXrm/PNOjTHmnE87fdOZZxwfi7i0NNq5a+Goo/Z/6MNOmZ+fuePOLYccsv/5D37AuvWrTj316Ml4UlW8bt3cUy44NyuIyvnnn3LSCUfWdXvKKUeceMLhRx154PHHHXrUkQedfNLhp5626Y47N++33+pHPuLMtetW37N5+zHHHHT2g47bf+O+BjBcnqxZs+qRj3jgscce9vtrb77ssquYcPv2PTfeeOdPL7vuW9/59exMee6DH2AGIQQP48k5Z9Hdu3aGEMqyuvK2BShmJos7rv/1z/ff/6BJ09xzx60zfXzZn79kdnZ2L4T6/6/rkz8h3+P/4x9VJYSHP+pxP7j0h4cccs5Bhx+2c+e999578zkv/jtesWEu73zvi08dVMVdd969dv36IkbnUvpX4hEArmB2S7BI7g96ZVmqymTc+A6xrMrRaERIvX4PAJq6SUmKMhRFHI8bVen3KzdMtG1GAhFNbdvvVVWvULXRaKJqTNQfVKo2Hrdm0h9U7t0dDidVVTDz8vJwbm62bVpiDoFVdDyuRWV+fnYyqeu6LcvY7/dSm1KWXq9ExKZumiYBWFkWZVWqynBp3OuXsShUZDgc93pVLCKAjUd1U7e9fq/qFQAwGddZJDD3+lVdN03dFmWsqqJt2iw2lba56lCZaGa2L6KjUU0EZVUQYdMkEZud7QPAZFKLWAhUVeVkUueUiqIoy9i2qW2lP6j2BoK1TVOUJQC4sqQoCwAbDSdghkz9XgUIk3EtojOzfQBsmlZE+v0KANs25ZwBsCwCoDHHz37m28993jtPPvGwiy/50PzKFWUR3/Wu//237/vymWcc9a1vvq9ps1NMmWh2buAOpBgLU1EfSwMsL498tBcC96pY120WIeJ+v0Kk4XCsZjP9qmmTqg0GPQB1DVDbJjD132U0HLuRsD/oEWFdN6bmj8pkXHuQRCxC2zZFUaaUHHnatq2ZlmX1/7yAVSeTpqyiG+xTmweDCpCaps0pi4h/1wDw7W//9IlP/OvnPPvBn/rU/xIxDuGVr/yH//jUDx9w0sHfveQfsujeDToiSJa6rufmZsuq+MDXr1/UmXb33V/4h7cde9wZ27ZvufqK/163Jl7xm19s2LDPFCJ4f0v4f1OH5QymwNzr94h4Mh4lkV5ZBqqcn7di1eqqqhyrAs4tMFA0pC403EPVi6Jo6nYyaogpRjaDlKVeWEZCBdmzexERi7IgwrbJ9aQlQmJaXh6rWhGjmrEhmIUQsujCwkhEYozMZACjUeO5CQA4mbSjYU1EgXkyboiwKOLy8hCRrE3ug/VhxfLSSCSXRQEGi4tD54csLQ5FNER2aExK0jRDAAshNE0ajWqnl9RNOxyOzSzGgIQpJx1JzrkjrpgtLQ2ZKBZRsg6XJw4O7yJwphg8Vd21e7GInjNK49HEo8QAcHFxmZk9N6Jt02RcAwARuvzHlyHLyyMmAgDHUUwmjVMrcsqj0STG4OgKRlhaHBpAYDazPbuXHE1hqosLy9T9wEAETdOORpMV8zN7FkaING7sy1/92ZFHHnDvlp3f/+FVk3F94gmHVb1ezmP/sE1tYc+yv7djawCRCVVrJGAiAySipmmbuiWioow55aWlkbO4mGg8aVSUAy0tLvujgtPo1rpJAODR4ogwHk26A4JweXmkorGIIYSUc0oJCNvhxBU1k3FtaqGIqR2pWYce2xtazTgZt+6EB8Tl5UkX6IwYQmiadjSerFgxNxzVAHbzLZs/+/lL16xZsXXr7muuvWcyak55wJGDwWB5OO5izBBEMhLGyKJiak3dKuec6sl4OTVNx1kzuD9I9b6y5lgXqKmmKeWiKAnZebJqmlNq2rYeT6qygi4FQBE95iCzx1KK5AxIaIFxGv+bUiYkj2CQ7GEHaAYiMuU3kWZJKbsdWpxm5VQGwI4AqeqXagyBmFLb5ixI6FDAEEN2Ela2nAAJDZywbICQTcHAQ/falBCRqbNeG2IIZApJsmeCTSt5dNYcILSqRB30PaUUQtAsXBAxZ1FUSzkzUVZRbT1VmJhULbWJiLLnbiAycxGdHaqeVSOSxTp0n4hj8GxvQrrnN3JgFbUsHKZkHhEIgQglSzNJsShCcBSnqQqD4+HVAaR+nkoWj4BUtTa3SOjRLzHwpG6e8uQH//ePr/jpz37/mtd+kBlzm9etX/nyVzz2jRc+czyauFPaTMlz3xFD4Jyzmap6Aph/d+S5zQ5syE6GUPVwB4QOMWgdJ7mjLHYKtCQhBAPQlKdJ8aaiuc5T8qoAsYMVbRp16dmxVJAzshGgKoumbZmDn4Ng02dQ1QA5kNuMRCTn7LfIZDJ56PkPuOCCMy753uXPe967CUEEV65a8eIXP/Ttf/WCSd2Ygmj2C9Ijsl3TU5VlKErRmHlQlL1eVQUKKui6nvsPrPvu2AIAVcgpI1EzWV7YuWXVQUdhA2DQ7/XCPkW35jfzpAIPwkEACsEAvKghInA/NE1RyQgAyB3IDZiD0z5V1MlHjF12nqkmz8IwI/aAaPXNnQ/bRIQDcwyqigBK6GgtgA4iPAWuBwdseX6fO12dNuOcQmcEmiqAIRozI0XHwphpLApzLLqq466c/QQOtMy5YycRlUXRiTms28D9ERwOQFMUsjOFwVOhAaYGTEIE7dJNFc35U7wXW+gYWBMzc+opePoGOj6YdS/3BhkFPCy+a9WRwBNBnJg1vS3Iz1MAqnpFynl+xfxFF73rqqtuvGfzThFZvWpu4/7rDz/ygPGoTlm8CCSMxI5ytja1nvzuOBnJ/mUJAHb59UzMhF1OvDgym7rYZwTopkKSRcBpLezxS0gdWsPDX9UMHcAdiEOHpoIuydY6DhcAqHawLbCiKJywqjkbdlWbqkjugtGQkIGnPwiaWn/Q/9Sn3nH5r6/duWtpeXk8M6gOPXT/E048vG1z26RYRJmO0ZmJCGdmZz2qtogxahg1Q8AARCmbSGNQ/OkUWH/aB1aX4ASAxIEYJbXIBRczrVjAIGaicu/me1evWV3EmLOvBbOpKWgXHqXd2+IoSCLai5DR6RPmGEvVlkNwhkkXLPU/yB4IHo/sRxAaEQA0dc0cENGPyBACqBk5Co49P2JvHQcElp3zNP0bzClxGrsCARBNc4dAcpfy9G8kQGybxulLbpns+rssokJMouY1pKeS+kGAiJq7TDMvIbqwGego9X6qdn20J3yKOqjLE/Ek544viqiqMQTruGCaU0Iiz3BX1L0JjzLNLlZTdyqIKiComAdnea1FiB0+WCSlFrGD/3HglBIxnXLqMWc+MPrUrWnzeFTjFLzp77mXJ1mEPY6MyEGdFHhvYPU0lNDaNpkqEwFg27Z+iImqw5r9fwFUTQNSk5IfQoCovmk27SjvXVyI5SR7P0yfP6gKGnpSPYPjctX9A07dAfOkMzM1QpSc1RSQiKhNbQzRQAkppczEZz3oRISOjp/adnlp7HeMGy1SlhDYH9HxaDQ7P0dEWaVOragwR2Y2zY7h92/k/gPr//o/0xAKQKL+oJfaGaY4s2LeIBiwH1IcQhGjVw/+lip69vc0LcIADDhwl19i4Dlx5Ghiz41CkCx+5f5xXzFlK3OXaZwRkJgNUFRAjZCgi5kITohHMoMuRUpEPXxBzWgK5DVVr3GKglNKqOBPt2dQT9FIEkKQrIAYmEQ9/JWVKKN4/2jTk9VPYefBd28TYcFRTT33oZuxT0OJ/QhUM8daeCfYNcKIBuhqoJRSCOioYvQMnqJgZu+aoaM6/zHsDwBjcAYp5Cwdkg86Z5Uzxryi3Itm6uTeiBwCKmXJDtpUNTBUsYU9S9Nv0GiKgQ6BJRug72TYTJwyLKrgx0MXeT8tZkWc++rs1pzF4cgwxVJ30Ydqotlx1SmnGKMHSTJ5erY50o9j4ZMKFXWg/d663kjNqfvQfSNONTSaJuM6PHoKOEUEMVQxAAWwsijUDLSb24rJ8tK469VVixi7maCpf1khdN+FH96e/JSyCoRMg7aZ1KMlf5ydu33/gXUfibCYGQBzkrrOqU392RVF2RsPlwrMw+F4zaq5+fkVzEGkFcmqxPz/ae9dYq3Ljvu+qlqPvfd53tf36mar2S3SeiR2B7JEAzYFJYgRJIgHFjOI4WlEyfPAgyDJQNA0EzuBAxgOYk2CRHJgJ0bgDBxZiSwoelC2xI5Iyu02+TWb3/O+zmvvvdaqqgxqn9tfx85Agaj4o9cPBHH6fvecs+8++9SuVavq/yecJNjRLEYmOXaz8FIAQEVgFu+d914mRWEABWEBcubjdKeGzMwueJMhF1uBqNilqFMPvRChn3xiyJlxAwgXdt47ImZGs65jNlklZjGPGUQyVXIiMjHoSVkYFKeCm/lTyaQ8rOqDFxHnzR90kqpzzluhyoz/LA3LQ0Ii711hsaMCMsNC88EmiyAmIjgt4gitLmNLK++9uTSHEI+GzfbY1ouKHpzzXBgITc0CplemY2MKMkspJl5sg0fApVjOS86ZRc3k9ZAyRcyJj+FmUmEm05VCQoJSCiJZpoaIpZjDKTlbIztvr2/uW8JicmPMxbJmZo5muShKRKUUCx6qylIIHZF5zKjJGR0zKnLOF2aYBOmdiVub1cW0B8jT3cJWay74nIrt0Zufrjl0+OCnhk9QIgohMIuKAE1LZjhqK8QYppEpERFJ44iEMUbT47/b8ptEShFyKof+sOGuv7mWMpQinjygg9fJ5uv1N1I1/fwYfckJyeV0uHn54uxzb7149jKxeu+3203fH05P11aeVxHyzlZ+cLRBFiks4IgsK7eFZs6TkZyb1LvVbqdIk+HSlMEQqohVms0btfDkAGhyocHMUHOx62KyeFFw3iFCyYVZbAjRbonm8SmT7SuEEHIuR+snNo12W5eZ9LDYG1s25FzJ5WjuAM45jzCmZAUsQHBEiqYADqIqKal4m+ZVUTQ5X1YFBRRJxfbvWdjcOiwRICIEb+0CyuJ9KKUoKGT1wZusq/eBAgpoTgVAHR39OSwKmHk9TZE3xKD2FxSzIpzsqixpSikROQCNTUNECGyr9hiDlbdLYVJUAOHpq55zntyPRHzwKuK8B5jcIMzFi9zkgWaxyXZFRV0RhiLkXM5sBquTg6SAINtK2Hrcj7ZsoKI556Mtm5lToLIULqWw996hs4r6ccUPmpkc+eDQBmREVW2vwHbulBwxC6gQkXMhZwZV8k5V3GSio1z4zl316LlbJqMDBOe8WaLFtvG2OTtwaCjFlpp5Ktr3Pag6F1xVHP2jC1gqALzf9zknjygMq5M1OgWk69s9Es5mHZEzLzxhIU/K6pwT86OfcqVJmBiOS6GpTKqqMpnEeSRAmIbbERDICuK28yMs5lzyyv0KEbhtG9uGsrL0ZCyu6oMD0Fx46l+dLhgyy2LbzHZk8seToLOIOOcdKB/v1SrKxzY6q4NM8sFApbD3DkBTKsF60ABocn41GXxxzoUQ7IAJkKJHQOEyNfEAInlb3YAN9Ts0ZV6Lkmb/ZwJFMYacs5KZ0R4L5yIK4KM/rk/ND5AtYQzBT2tRFWW5M2clJJgUxpXMEnnK9QAJzfwG0Wz4wNromjbC8Q8TUYeTG4ioot2ERKaVJ6KnoyYnoh5V54lwOiqPKgKeYCrMwaRSf8x9pj1ZBeccThs+go6icwhAMTKL1QpcjMIcY2RRM6blUlSdaRxbslwy23oZEZxzIXhR8I6YRdT0jiEELyKWHdu2EOp00/HeAaCiWcopOecsQrkIk1q9HrV2m49fPFPBWeTw6PsU2zENMbZEHSLD69Pq/nrrYSFS2zYAqXCSIuic974f+v2h7JL7lX/8uKS0XC5D8JMjklrzoJhxvCNn4QOJiEjkmG0rlFxKLjlnEUlDKrmwWvKPRy8wsqo/Hy12rWwjZgSPyFwsp1ARW4OomtHy5O6Xc7FHzpGwlFyEbWqMzCHRCqhcuJSC0y2dS+Hppj4ZWwESOuf1FYdn60so0xMhjQkVuPA4jGkYzc3Mnnt3Z7f/OFqlT1t6RKhsax/TSYfx0JecbQfATp05DJoFsTAPfT95Dlpy6uybT8fPC1GnYxYVU0y/K0sjTtY4d/saIqwq5BwoOEdWC7eai4qZ2hdRGfrB/H5UlMwwcXIbmXxarUg0va9M+7Z2RUxbD4CqOo5jKcWHYIt5OnaTmPcPIXnb7gQFFStZmomgc5aIaRpHLsWWijklO7cmzRaCN3NDYck5KyhzgaP5MyKmnFPKqJpSySkJs3NERJyZC4M5qooACNHU/GE1LEcEFs3NLtM5QBwOA5cizETYti2o/IN/+NV+lPOT2eU3vz5un8fYoTJAz5yrgN8f2WgOvvvOuwDQtpjTMJ+vSXX77HGITeza33tWHn/0rOsagGl9V7iklBVAlEspMnmX3iVYVi5lnTIKdM45R6GJsYneLh8itZAGGmKwGFVKtin545Jtqp1aaLMiGE2XvpRS6K6t4ehFKiosggTMzCLOkYVWct4FH4LnqSgkd6peSCCTYioBaikZEbz31uM+dQNMCzFAwrZtfQy2cHBuahOzX5r2wMiJqu0AkCN78cmUxTqmuCiC997cs+37YIdUSjGfxxij/VyFfXDWBjDVoURyyYWLuc5Ys2KIwdwVU8oswmxe8LYLASKiYi0DbCbPVt5CBBEWzuYjbc0WPnhmVgRA5JwBLAGZNhBEhEVyziJs7SalZItlU8+EvQ5hznn6sABs6WorLPNwZi4pZ3OlFWEVtgOerHrIWeX7rvcqxGh7kSUXLoU5q8r0+mhl/qIqaUwIkHJmEUJ03h+LZWgiSMXqmlMzifJ0YkVBzLQVAKf7hEhJOTSRHFkxrmmb66ur3/7mrlufDklvr68caDtbXl9fIlqL6+vimvPaOz/D577/XQC9uXkZYndx78H5vTef/t6vPPyT/+79s/b2pvlff+3r/9Eb97pZt73d9Ie9cz4Er4qiIlys1cCRs8zKeQcIwkKEzgU1pxbvps5CnaJYCN4uG8sdfAgm2jetoo72zdZqoKDOE4CbvgaOSNTWUMxsCxZrtAkx6mS3eUzmC1MkZYGpgWhyajF3MiLyAVVBRR05aojNd1rE3LCbECxG63HLzVZq6Imc8wDHTAaQSMRWXla90qnRzE2Ve9N7sD01+2KQI7UCj/WXILKI816PG5Em++mIRNSWhywSfKBjvcSM3MdhdI6sZ83SPSTkwiLsnBcAdGhDmkAQnIsxWrjnSaqcSmFzXSyZffC2u+eCV57s2r3tEjJOW7z6yb6YtWsUKd75qd0EMecytdER5THd9cTaEpgo8DCaIoLidOtSBnKIRCBiPrXmG1Cm3MoM7sUFjzgNJ3vwTROtOhm8B3O0D2EKSSDTjRRsgxeDC8cqFdvHaYt6szYLIYiK2a4OOYtwSSWEgIjL1RqA//4/fP+mfXDRxu2BP/7ab91/+H3CWtQVHs/O763Wa7OyqzWs72pPA4nIe++9F+Nss3m2Pr039IeHn/nhr/1fv6y3T8LFw8V69Wsf7z//6+//xJ9+bzafbzYbBSzMwzDGEBAhl2ylblEJ3mueQlLK2TuvqimnSJHIyeTHSUSQUy7MzhEzqELhUrKp03hVTaV4546u4iWlMYZohfkibCFgHMo4jiEES6a4MCLmkoUVVX30ynZjZ9NyQ0c5F1SdOl11CgFSZGpcQMylqEjTtiWXYhtyhR1hKaXk7EKwpqec2QUah8EWaJnZBy+swny53bZN42PgPDkMlynPpDwmWw1N9ouEw+FA5Gx7i9zUi5vHZKWTY0i6205FmysoOUkSImeNRXaSc7EymfoQyDkt0+K9mLPRK6VGQOSSnXNm3EqIOVuLLBcrwxGKTh6R5BwwmF2gpTMiwqU4a5dD1GPJvZQyqs1OOdv6TCnZxohtwNl2ITlHk/2a5lKkMCD4EAiRuWQkq39bQLRzZUMCzMzCKqIy7RuwcE5JNEwt+ce7HYqUknHqHRWryTnrwhNJafQ+IGHJ5oRKNgxgPal2kygKtm8Tm0hIi+WCCH711776fzwGCt317aG/uXzx4e+8+9n3mHPJB+Xbhw9/eDbrpo2FGrC+mwUszDn/4A/+wB//4z/8la/89ub66XbzxsXDs4uL73/8v//ie3/xPwlee2r/9vuHff9bX/zRH7i4d1FK6Q9901g5AwhJVXIpVrOMwdvMM4K1LCI5KrlYK5Zt0pUsxy7wIMJWAPGOp6m6MYUYQdU7B9NGmFNER8Qg09wWqOmFm6imNUyzqAMk0uBDLgUQXUCPfhrxI0eRbPcNwYzd0aTXVESE7TsurIiYSyEiy0Rs2QvmxuEcMztH09Y+i3fe5DER0Icwd4SAwhybRkVtC8JsnNumkWPjFACUwk3b2BiAydGEEK1zLaVERABqS2+aNKxVVZumEVCnWnIppRBRKjl478lhiGDdswAhxJwz0nRPAvC2urVIzaUUFu+9FaGtEd/7IHpceU6G0bZLC8JqjQVESM577+3eYwbgwkKObOMfEUMIKWURDSGCZRzT2g2tuwlEWNR7h6LoJztuZo5NO72rTJ675ByUYmtPRMBpTIgs5fTOVmEIqmS3XlXnKI9ZwdpyHag65239aJ1qwXlrbfE+qKK14ogogjCzRw8iCsAsTdd2XRu87w+Hf/zVD//2Vwc9eSv2O2zar/zy36KcFsvzJ9/+aDi8ADj8yI/8iPfe6rA1YH3Xk6z1+uRLX/rSV77yW/3+5vG3vhGa2Wff/dd///1f/s7/+Xfuf+FLzbJsYf2LX33y4sVv/cQXPvfZz75xcrICgJSK+ZgLa9PNRBgUYpzMKSGA996u6ZynlePkbDrVp8E5h1NHpWBrEoC+bVtbrVhO0c7Beo4dYWha6zIlxJxzLsUmyGLTEKKoIsAwjkTUOEfOdq8dvNJhObWYH03SrRJk5XG2Bi4kcpTGPLWni3iPPkSzSmdmEtfEqMdNdGEOTbD5ZNvGyjlbaDt2TgqgmU67qUSiQrbBimRtk8fy3FTvciGYn3NsrEWTrT9rSrIAhNl7H2OcAgHbYlFV1YfgvQdV7x0cJ+/sSySi5BBkarCwKGDfc52GoqZtQnI0dY0Rmcs3Ioio9aZxYXEOADyXEIMlzVYAskaVRdPcGdxbB5yoeofDmBDAEZH3oJpSsnkaO1eEMJXwEe2orMPA5rosLFpF014zRp8LHxvppvaa42V2tFXBxpaTiGSJ9t2IBai2XctT26jaBJElXDSdMs39/p98+OQ3f+effdifxfM30m7vlqvLf/a1j//Rr/zgv/anhmF4+fLpYX/tffNn/+y/DcfBn9cgTXl95WWMvu8ff+vxn//J/+Ab3/j9s7N3Tk7fefDo84uT1ccf/Obb/95PdZ/7022bx914HvJDd/l9p/jHPrN+42J1cb6OMXrv7TolT4QogsOYiMAHP4xj9MEK5HbdE6HazhozHUOA88G6k7z3KU2SDFbHLzkDqgg2Tez7w11WlYvYq/WHPnoK0RN555xV3h25MSVrosHj7lXJycfAmVkE0MrM1gyBzOytvcDWTAAheFabVxHbnLK+MNu5P47psrnjMIt3rkwBiG0jTFRE2KznAQGJHDkRxqnbExwhMx/6IXhvQbybdSkXBPDeEUJKGUCRyHZRp5ougircGVAToXUIWCHPEVqHqkhx5GzJVkoBtHZ8sMHkadgAEVTuzOjh+OrOu/1hbKM/jvOqCgOSPw4M2dZbzolQnQ/kvL2WpY1THxPCsb8cps7ewiqFnNnEgSh4h2KdsccwOZ0ftV0/VhHviG2yEpBVLMMNITBrzonuGkpFY/Bq+ZTIFJERpyqpIrlpQmJqFQb0wXPOOWcinG6c6ABEAbmU29vN733wna/+02df/2jo8ezh2w+xm2Uf908/+vv/zX8xd81sfv6db39w9fKj65fv/9iP/am/9/f+5/V6PTWp1ID13SbnfDj0f+sXf/Gnf+YvhbBw/nR9+s4P/fB77Wz5+MN/1H3+C5/5t/7ickZ42HEZQxdg/3KBaR1ld9g/ujibz+fbXS/K6ONyMRtT2h964Tyytk0TUVIq5GLbNrmUUsqQeXfol7Nm3pAnt1itrm766GA57y63vYe8nDWF4WTWXG2Hq9t9281CdKsWBPDFbT/vOtObH8chZ17MY2aZdU0TwmHoARXBj0XGnGdNcA6HzOvFbEwMiCpwvTm0AYOjWRNEoQ24OyR0NGQuhc9PFznlLvj1ork9pJxSyuI9ttEpuZTKOHKfiidEQBbmUkLbztsICiw8DiOArhftmNKTy92903UTY59L1xCqDiO3TWDFzNwG6ocMBFJ41sC+H1frlQMZkoTYhuB2u0MeRx9czrLoIig0TXTO9WMGhcUs7oaEiIt5ywIh+FzyMOQmxOvbja3dZm1g0d0+iwiBhOB32+1uSBdna0++afyQSyCfWAtnohADtR4E6Ho/to0LLgypjGkYxzE46kJoGicsAo4F9/1AwOvl4nTZsciQxqaNBbzkMo7jkDMCtU1EhDEVhwDkQgg5JQQJ3qtySTl6b61jSZAIx8xEmFkRwfuQFT3A7tAv510Z05jGpp3NuphyWc7j5e0eRIIPVrqSXFwIMeB2PxRR0wrdH/qmid6RD74JxKIqEIIbMkciZn1xvd1vNuvVcjELFyeL7WGInoZheLHDD3etCzG2y9VyMY6HuFg8/Se/+zt/52+cn96fL+9/8/e/9p2PvnHYfavkm//yr/5XP/OXvmwDFTVg/RFxe3OzPxz+s//0P/+bP//ftt1FiBdnZ2+88+6fmC9Pv/PRV4du9ehH/53Tt35ouZwHj2XYh+CVwZqHvHP73d57RyF4GdeL5nY3eB+WM99n3hxKM2uX0RHq9bZfzJokcHVgKGU5b22GzhGuujCmsi+KIkNm37YLZIzx8Yv9Zx6ur/flJCIoFhUPqoA8DocxvXmxeHJbzk+7w5CLQN+Pwrko3Fu3Y9aSkgteQmycU4WrfXnjvL3ZjQ40DynMmnvLeHnbF9GT1aywXN2M84isPPR5NYsqadvr979xvs+82R/mTRizNJ52IyfGWetEQbn0KbeNf+O02/RcxtS0kRzuDqWksW1C9MQimz5HonbWKLnNze5s3YHqmHQxp5vbw3rdEsLNLi/bcHEyu+71att7le0hxUA5Mw/p9HQxD3S9l9M53PZwsY5F9Nn1ENu46tzN5uCc7LfjO28/Gg/9mCUJzLsQCQrjzXZoGrp3svj4xSHnYbWYLSLuMwPLrrjFIuxGOfRD4fFiEYeBfcSnN+XtB6sskApv92XWhWUkJHzx4na+WgHRbtcT8KztThpQwlxKAnrz/vr55WE9jx8+3d9bEmdsQ7nd57ZtHEoRV0qZtx5EmzYOh2E+a7o2Fubn27Sa+Tz0N710rQ/OEYZDUeESA8QQ9rshNLTd873TWSopZUGA6MO89busIhy9A8RNn8qYQ3BOeMgMBKqwXs4F4NCPMUbbQKDgCjMK7oc8DmMb43rVtE7HrF3jrjf7Wdtt+jRrozi/OFmkze1v/i//4+brv/bWZ97q1g8ef/jhk48/3G+f7bcffPHP/Jlf+IX/4fTsvGma16Zy/T0QsEopl5dXN9dXP/XlL//qr/5q05yTW6xW99/9/vcuHn6mH/bffvx1XJ7P3vxjF2//0Pz0QTNfQ+pFMhDl/e6zn//s0ycvWFwAJu4vHr7x7OmLXoLzoj4kovNZmws3Xbe9vQGi3XY/ny9SP5yvu+0h51xmAcei21Eenfqx6M0gc1J0gWLst1cs0jbLNrihuEgsOZ+eLve5HLa73cifeXhyeX2QwojsiHC+CMROVCCy8FDEuTBvgpL2QyLUPB7my7N+2H/m3uLxk9sQwmrW7A557MeTRXNzOHjf3Oz61dzHpjtseteEZj7TzKHp9ptt0wQoTB5y4eiomYVdUuVCUsqYIUZCunr+/OLRw1LYibKMGucRpSiklN95dPKtJ9eztokOh1wQaHd92zW4Pjs59OV0vbi62WGIjtPVpsfYEGFDDp1XKYXiacdZICcFihcnzYvbPhAuQn52uVnOT7x383l7fbvvWn+7GbwLqrltQy4pjwmoefTo/OnVMA6H0DbSH9SHJuJ2e2gXK9ZyuL1drNZA7Akx5XFII4T7909fPL9ExXnnUfh2c6CmYeV52+y243o5A++Wy1bBh7S/3R7OTuabUYeeNcnp2iuF/WF03arkYTZrt5v9ctlsNtvT9UlWn0pug4b57HKXT1x5+eKqbRerLgLqzX4IVLz3w5CJIJfeN6sGcjtrLjeHe6fzF1fDyWI+DIMPPjND7A777byJ+6RNIODRxwjOl+2OPCxnTV9i03bjYbcbCyJ61CSggH2fLu6vogNQIGREunlxdf7g7ObpR8+++bXrJ4/Hb35tPesevfn25nb7rW9+kFJ/dfm475/du3f6C//9f/cnf/THutmMXh/jnO+FgGWVrKurq4+//e2f+vKX33///bY9FaG2fbRY3Hv7B/7EfD6XtL29fZ4lFYwJIunhsL2czWYg5d5nPv/4g6+ePXgLS3r+7KP1/bfT4YDNenP50WL9wHk362Yvnn775Ox8f/PC+/m4vzp7493tzfOzs4eiePPy43HYdot1Vu8UM4/NbKHD7Tjs7z96++m3vtHMFuvzN2+vn49CTWzT4TKzLs4etXG5XJ7dXn9H1W1vnsWAIXTbYWxn7elq9eLycr0+ZSXvPadxt7nFZtk1Ded91iD756fr0xeXL0/PLnabbWi6GNF7t91ulxffh2WXSimCZRxURYBVtGnn3kNKvUdMJc3aBUrZjalbrMuwT4dtSenswZsKLii8vL31bRvQDf3tyb1Hh811bOYlHx698dY//eADdPjowYMXz58reRl3TduFbp7T2M7ONG3Vt8PmSpib5RmBjGn0cY6k83m7P+TUb7r5ynVL6S8Ph342P0XZP3vyeL56QxTO3/w8lrS5ufQh9sO4mjVjf5PTAYAcIHUza7HN+fDg3sPdfjcOoyc8HPaz9fnY731sczqcXNzfvHiJTBrcMB4clNgu0tifnz/Yb69ds4oBUDRzGrO0XQdEjrwDjzrmUlhLztKFcLvbrE/vKWizvs/7l6oAFIL3Ny+/hb51PuZ0WC5Wi9X6yYubWRP62xeC7byLIQSKC85DKq7vd13wV5cfnb/xuf7qsju7R7pvu+76arvomv6w93GW+muKS023Z2cX+xLa+Tptn++Gfnnx1vVH31yePxiGrXJpZ+tI+enTJ6Aym6+a+ZqZb69ezE/OPeGw2xCWMQ2z+cNmMd8/+TopnKwu7j94iyF+9K0P9tur/f4qDU+Hce8D/I2//tf/3J/792fzxWuUXn3vBCwA2Gw2Nzc3Hz3+1n/8l//yb/z6bwQ/J7dgxtXpW00zW51crNanq9VyMe9ubl/GZhabdjabUfDD4eBCTHnoAg1j+c63H5+cP9xtt/fvnd5eb09OL8ZxLCU1IbDgzaZ/dP805Ryatt/dEtB83o39rh8T+kDkF7P5zfYGpKT+0C2Ww5Dbpg0eQmy32z2CsrBKGdKwXF6slutx3KZhAOcV9XDo54vlfrebBQTQnNJieZbL2LWdj/7yxXWczZvG7zY3s9mCyN1uNovVfL/rmVF1SKk8evTO5dXLxXwZgpc8iCiQeALnw/PnL13bStqdnj/YH/bRh3tni+vb2yLqvdvvhntnZ7ebDYuG2HoS5sQM0flcCjkvDBRc8ETkmdmHNrTx+dPHCjE2i832ZjGPHv3Jxfnz58+Xy8V4GFz0szaoKLO1iRUE4JJ9bJJIgEToEL1o6fs9pzxfnpbCb7z58ObqMsZ2zEPwnQDd3Gy8c2N/8CiqfLs7dLMZOj8/Od/f3p6sF60vu/3Irpu3JKXcbndcdLFY5zJC2XsfhyS7w/7kZL1cLC5fvLx48OCw60E5p3F9errv+83mNsbu9GTmCV5ebx0SIgnI9na/PlnfbG4/97m3nz15enHvvhRBF/f7GweyT+QCRScplfmsQ8T9fts24fJyf3J60kY3jPnk/PTqxWXompJYMYypxAa4qOTMRRbr2WF3WMyXT548/ezbb+73h9ligYqbm5dAThSa2GXm9UmXc+kPaTabi5JK6re7xXIeo1PRvu99aLVkkXxz8+L8/mefPHl6/+GjzS6XnDY3l9/+9nd2m2fKz4hgHF52XfezP/uzf+Ev/Ier1Xq1Wr1mzUzfMwELQDeb7c3N9c31zV/5K3/1b/78z6tKjAsX5qBONXDmxfJ+0619DAjSzeZjfwhNW0qOcZlLBkAkp8r9YZdzP+s6UCegCjSbLcfDjfNRkFQ4BA9AXBJR4NxLGRQwpb5rXNNd3Gx3y0WjrDntu/lq7A9N0wL6XIDL4ENg5sIjgWcui3kHRMOY0njwIQor+RBcmHWLze2lEhGhd2E2X+42t32/bbo5ofT9HpBW6weH7XPfLA+7DZI0IcwWJyxQcvEOVEtOWQGWy1OUNIzjWErbdv1h33RLFnaIPlDOPKa0Wp+X/mazvV6e3MspxdiM46gK3gcAzWlQJR+Dh0K+LTnt99vVyTmCHHY7Flmf3uOcUxpS5sVqncctZ4lt27YNStltb1VVwTVNS86z8GHo592MXEhjGdPYzeakI2CMTVQe2m7x/OV1jA4V22623+1KOjTdLHp36PeI3nTG0LUIRACL5WK/34Kbpdyvlif7wy2zCLOPLcrARVMaZ4uVlkPJRRSaGJUPPjYibhz7EBtAMsVaT54l23R5E10ae8AWoMQY2nY29D0AALk2+uDDzfY253R6ciLQ7fcbBGnblksRxZL7pgnD2JdUfGhC20EZWFwuKTSz0Hb76+cuzNs27DZX6CKhomrXNrFbDP3giHxsWPLu9pKBVstTR6AUuRQXYhm2okSkCOCctTsU54IwK3nA0HaLfnu13d+OwziOG04bxDL21wrps2+//XM/93M//uNfXCxXp6en+Bo5fH3PBSxQ1f1+d3tzW0r5pX/wv/21v/Zff+Urvw0AiCGGDtF5v2IB4ayam3YhJTsfx3HXtCeI4HyrkrkkpKbwCJKca4BIFZpmttt+3DarwuqDG4ddiAvn25IGIAZRFiQoedxQWHvfgI6zxdnm5kk3P0vDtp2thv4QwnxMvXNOJKuICjgfiKTk0i7WZdgpQN8futkp536+OAMt2/2GkFTZh3keN0ihm82GoRdhzsPZ/e/b3b7MKZFzTdumYRPjQqlVHkLEYSwAzpFae0TXrfe7K+ep7/fed0BUxr0LDQKKYtvOuBwO+6uuOwWknLJzHp0nLCKc08HFlUhGyaJkhhHWd5VT70L0jgAclyQKMbph7BEDYXCemEdmxWm0MZrsk/NUxj35TiQxc4htSdt2/iingyOIzXLMQ0lDcCgCRK6U7IMHkJxLtzjRvC8lM6tznQgTocqAYTUO29lsCYBp3CEFAOYytrOzw/ZyffZgv7ssaSTXOu+Ichqzjyvlkcu+6c4VkMe9D1FVRDwABw99v226sxC7w+7F6vTB5ualQnBEgOIQBUjKwXvfzt/Y3DxzTgFB8sGFlagQFAFKY+9dDG037l80s3vKQ4xd3+/n89Ocs8i2lCQaEL0Jc6HrVAl1WKzPDrvrnAbvWuYhxBliSHkIMfS7l+3shEtG8ggQmvlhfx1io6yAjsg7HzhvkHToN6A5lx4A2qb90pf+/M/89E+/8+673Wx+enpK9PqNEn9PBSxjHMfN7W3hcn159Uu/9Ev/09/9u7/+67+x221f+ZU70TKEO32GT35i/0+v/NPUWXl8UI6PydQgYdJDkeMv3J1SeuUtEIBfeXx3rcgr70WvHAwRetHxlX8VAAfAx3Zf+ef+hE99svDJaJgdIQLkTz8Fj4/1+NZ8fCDHf5VPnxA9vqweX1aOv0av/Pn4yq/hv+iQ7GBePVd3fz79v/w5d8f56h9Or/wTvHIa+dNHogD++GHp8Sd4/B9/+k3p0++oAAGgIHrV8s+dQH3lr747BgJwANme+OlPXF+5CPnT542PP7/7IPT4RD6eOoFPTfzhK8fw6mmU46sBALz99tv/5k/8xJd+8iff+zfec94vl6vFckmvW271PRuwrMn7cDjs97uc83a7++CDD373d3/3/fff//g7Hz9/9nyz2aAp5gEcdSMnbVnT+rBz4sipCgubEt6x6XwagmUWGxhGRNNkuFPUUlAi551jEVMXcN4srURh0tsqXACm7lCbd7V+cWtktselZBtksxENm3qdhIxFTFdTVXPOCBBjRKKciwh75+9m2Y7KOSwsNkpisoJENvxoOhaTaJy9u3PueFQyqYPqMfAgmloKEXnnrdXadLWYeRyTNegTmqYFWCO4CRIc1TZtKBDuXvYobjFdkKbNwoVNKfTuS24TiKAqoPjJE2GS0FAVYUJHjkwTFSbRfRVh+2RN+xiPZkvMAqDBB+e9jf2ZNiEz26d2J4tqOkKiMvXTT3YbamLWptwwfYKAMjntoHPeFB3wqFn6yVilAh9n74/fQXXO21yBycOb4ru12pothXXDHz2KFCaNXJOpAUKSSVRSmyZe3Lu/mM/feeedH//iF7/whS/cv38fAEII69OTJjb4ekYruFO2/J5EhPt+2B/2luQQnAAABHRJREFUOWUunEsupfR9nybjvEmAaVJ1l+PIro3LmnaBiBmFm3ehKd7aLAszTyq0eBSQmS5xNG1StGlbLoBoTzEhhklvd+qPd4CoR/+x4zcWzeJBuJTC5MgRicqdo9edXo0FspwzADYxgAUsZu8dTt9hWw+CqIiod855Z2bXjpxpG+SSJ6UwEVDFSd1p+uJNouwAR9MNNUcfi32gIKDBextGG4beFBqQ3N04jgV3a6M/6lLZWMv0xbsbOIGj/QeL2LDU9OZo3jRu0hwz0Z6jVLBpXk+KzJPihHCZ8h01+c1J7fMofvWJkDmGGI7zw2oSYFy4cKGjtM4n5YbjJwiTON4nDkMifDxXk8Wf3TNYRLhYkIM7yzCTwzZpIDfdpdCmtc2rzXzGcBKTcHYxqExyzCaVC5MooMlC2GiSiJrWc/B+sVh4H2azLjaNvW8367q2ex2Xgf+qBKy7K62UnMY05lRysQsLX7kc74bdjl8bMI22V1/h6EeAx9RsSjhemQjGuwzBHHdMLhKPVg7wybjWnS3vZJ8F+MlL3X2j1GY6AE0YC4/xwgS/TUDmbsxwMrKbpoWPR3X8fF/5temvO2Yu8EoacjzCTy937vKLyc5n8j+bToJF/buIMyVQYmoI8Mr5MeHDTx5/okz+6qNXAsExSt7lFMexu+MxfHIOp4P85IAt1nzq01E45kevKPVMp12OE0ugr3w6cBRcw1c+nbsDRvjUz+/+2OO1RHfPQPzEpfSVPGt6krzy0UxGScffpaNXz1HbyAYK4Xjw9vjVcAp38so2J++c88F7H0II3rnXNqv6VyxgffJlmJxc4P/xBf7/P6TCd1mISNWcNfAP752+28dscQtfp0/heA+b3Of+QM/Uf0HV7g9yheqrYf/VcA6fuhPXgFWpVCp/dAGL6jmoVCqvCzVgVSqV1wWtAatSqdQMq1KpVGrAqlQqNWBVKpVKDViVSqVSA1alUqkBq1KpVGrAqlQqlRqwKpVKDViVSqVSA1alUqnUgFWpVGrAqlQqlRqwKpVK5Q9A1cOqVCo1w6pUKpUasCqVSg1YlUql8i87VXG0UqnUDKtSqVRqwKpUKjVgVSqVSg1YlUqlUgNWpVKpAatSqVRqwKpUKpUasCqVSg1YlUqlUgNWpVKp1IBVqVRqwKpUKpUasCqVSqUGrEqlUgNWpVKp1IBVqVQqNWBVKpUasCqVSqUGrEqlUqkBq1Kp1IBVqVQqNWBVKpVKDViVSqUGrEqlUqkBq1KpVGrAqlQqNWBVKpVKDViVSqVSA1alUnldwBqwKpVKzbAqlUrlDxmtAatSqdQMq1KpVP6QqTWsSqVSM6xKpVKpAatSqdSAValUKjVgVSqVSg1YlUqlBqxKpVKpAatSqVRqwKpUKjVgVSqVSg1YlUql8v+JOvxcqVRqhlWpVCo1YFUqlRqwKpVKpQasSqVSqQGrUqnUgFWpVCo1YFUqlUoNWJVKpQasSqVS+ZeLakJRqVRqhlWpVCp/uOkV1oBVqVRei2gFoADwfwPvorMkgF+xwAAAAABJRU5ErkJggg==" style="height:52px;width:auto;object-fit:contain">'+
        '<div>'+
          '<div style="font-size:10px;color:#64748b;margin-top:2px">Commercial Ice Machine Sanitation &amp; Compliance</div>'+
          '<div style="font-size:10px;color:#64748b">pinellasiceco.com &nbsp;&bull;&nbsp; Pinellas County, FL</div>'+
        '</div>'+
      '</div>'+
      '<div style="text-align:right">'+
        '<div style="font-size:13px;font-weight:800;color:#0f1f38;letter-spacing:.05em">COMPLIANCE REPORT</div>'+
        '<div style="font-size:10px;color:#64748b;margin-top:2px">Report No. '+reportNum+'</div>'+
        '<div style="font-size:10px;color:#64748b">'+dateStr+'</div>'+
        '<div style="display:inline-block;margin-top:4px;padding:2px 8px;background:#ecfdf5;border:1px solid #6ee7b7;border-radius:4px;font-size:9px;font-weight:700;color:#059669;letter-spacing:.06em">FL DBPR COMPLIANT</div>'+
      '</div>'+
    '</div>'+

    // ── CLIENT + MACHINE ROW ─────────────────────────────────────────────
    '<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:14px">'+
      '<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px">'+
        '<div style="font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:5px">Establishment</div>'+
        '<div style="font-size:14px;font-weight:800;color:#0f1f38;margin-bottom:2px">'+p.name+'</div>'+
        '<div style="font-size:10px;color:#475569">'+p.address+'</div>'+
        '<div style="font-size:10px;color:#475569">'+p.city+', FL '+p.zip+'</div>'+
        (p.phone?'<div style="font-size:10px;color:#475569;margin-top:2px">'+p.phone+'</div>':'')+
        '<div style="font-size:8px;color:#94a3b8;margin-top:4px">FL License: '+id+'</div>'+
      '</div>'+
      '<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px">'+
        '<div style="font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:5px">Equipment</div>'+
        '<div style="font-size:12px;font-weight:700;color:#0f1f38;margin-bottom:2px">'+machineBrand+(machineModel?' — '+machineModel:'')+'</div>'+
        (machineSerial?'<div style="font-size:10px;color:#475569">S/N: '+machineSerial+'</div>':'')+
        '<div style="font-size:10px;color:#475569;margin-top:2px">Units serviced: '+(p.machines||1)+'</div>'+
        '<div style="font-size:10px;color:#475569">Next service due: '+nextSvc+'</div>'+
        '<div style="font-size:8px;color:#94a3b8;margin-top:4px">Service interval: 60 days</div>'+
      '</div>'+
    '</div>'+

    // ── ATP RESULTS — THE CENTERPIECE ────────────────────────────────────
    '<div style="margin-bottom:14px">'+
      '<div style="font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px">ATP Contamination Testing &mdash; Hygiena Ensure&#8482; Luminometer</div>'+
      (hasATP?
        '<div style="display:grid;grid-template-columns:1fr auto 1fr;gap:8px;align-items:center">'+
          // Pre
          '<div style="text-align:center;background:#fef2f2;border:2px solid #fca5a5;border-radius:10px;padding:12px">'+
            '<div style="font-size:9px;font-weight:700;color:#dc2626;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">Before Service</div>'+
            '<div style="font-size:36px;font-weight:900;color:'+atpColor(atpPre)+';line-height:1">'+atpPre+'</div>'+
            '<div style="font-size:10px;font-weight:700;color:#64748b;margin-top:2px">RLU</div>'+
            '<div style="font-size:9px;font-weight:700;color:'+atpColor(atpPre)+';margin-top:4px">'+atpLabel(atpPre)+'</div>'+
          '</div>'+
          // Arrow
          '<div style="text-align:center">'+
            '<div style="font-size:20px;color:#059669;font-weight:700">&rarr;</div>'+
            '<div style="font-size:8px;color:#64748b;margin-top:2px">90%+ reduction<br>required</div>'+
          '</div>'+
          // Post
          '<div style="text-align:center;background:#ecfdf5;border:2px solid #6ee7b7;border-radius:10px;padding:12px">'+
            '<div style="font-size:9px;font-weight:700;color:#059669;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px">After Service</div>'+
            '<div style="font-size:36px;font-weight:900;color:'+atpColor(atpPost)+';line-height:1">'+atpPost+'</div>'+
            '<div style="font-size:10px;font-weight:700;color:#64748b;margin-top:2px">RLU</div>'+
            '<div style="font-size:9px;font-weight:700;color:'+atpColor(atpPost)+';margin-top:4px">'+atpLabel(atpPost)+'</div>'+
          '</div>'+
        '</div>'+
        '<div style="margin-top:8px;padding:6px 10px;background:#f0fdf4;border-radius:6px;font-size:9px;color:#166534">'+
          '<b>Compliance standard:</b> FDA Food Code requires &lt;10 RLU on ice contact surfaces. Readings &gt;30 RLU represent a potential health code violation. '+
          'ATP (Adenosine Triphosphate) testing is the gold standard for food contact surface sanitation verification used by health departments nationally.'+
        '</div>'
      :
        '<div style="background:#f8fafc;border:1px dashed #e2e8f0;border-radius:8px;padding:12px;text-align:center;color:#94a3b8;font-size:11px">'+
          'Log ATP readings in the service tab to populate this section'+
        '</div>'
      )+
    '</div>'+

    // ── SERVICES PERFORMED ───────────────────────────────────────────────
    '<div style="margin-bottom:14px">'+
      '<div style="font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px">Services Performed This Visit</div>'+
      '<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'+
        svcRow('Ice bin disassembly &amp; cleaning','&#x2713;')+
        svcRow('Evaporator plate descaling','&#x2713;')+
        svcRow('Water distribution system flush','&#x2713;')+
        svcRow('Internal sanitizer application','&#x2713;')+
        svcRow('Water curtain &amp; splash guard','&#x2713;')+
        svcRow('Air filter cleaning','&#x2713;')+
        svcRow('Float valve inspection','&#x2713;')+
        svcRow('Exterior wipe-down &amp; inspection','&#x2713;')+
        (filterReplaced?svcRow('Water filter replacement — '+filterType,'&#x2713;'):'<div style="padding:5px 8px;background:#f8fafc;border-radius:5px;font-size:9px;color:#94a3b8">Water filter: Not replaced this visit</div>')+
        svcRow('Chemical used','Nu-Calgon Nickel-Safe + No-Rinse Sanitizer')+
        svcRow('Compliance standard','FDA Food Code 3-502.12 &amp; FL 64E-11')+
        svcRow('Next service due',nextSvc)+
      '</div>'+
    '</div>'+

    // ── CONDITION NOTES ──────────────────────────────────────────────────
    '<div style="margin-bottom:14px">'+
      '<div style="font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">Technician Notes &amp; Observations</div>'+
      '<textarea id="report-notes-'+id+'" style="width:100%;padding:10px;border:1px solid #e2e8f0;border-radius:8px;font-size:11px;font-family:inherit;color:#1e293b;background:#fff;outline:none;resize:none;line-height:1.6" rows="3" placeholder="Machine condition, scale level, biofilm observed, recommendations, items to monitor...">'+(c.report_notes||'')+'</textarea>'+
    '</div>'+

    // ── CERTIFICATION BLOCK ──────────────────────────────────────────────
    '<div style="background:#0f1f38;border-radius:8px;padding:12px;margin-bottom:14px">'+
      '<div style="font-size:9px;font-weight:700;color:#c9973a;text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">Certification</div>'+
      '<div style="font-size:10px;color:#e0d8cc;line-height:1.7">'+
        'I certify that the ice machine equipment listed above was cleaned, descaled, and sanitized in accordance with FDA Food Code 3-502.12, Florida Administrative Code 64E-11, and manufacturer service specifications. '+
        'All chemical products used are EPA-registered and NSF/ANSI 60 certified for food equipment contact surfaces. '+
        'ATP readings were obtained using a calibrated Hygiena Ensure&#8482; luminometer with Ultrasnap&#8482; test swabs.'+
      '</div>'+
    '</div>'+

    // ── SIGNATURES ───────────────────────────────────────────────────────
    '<div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:14px">'+
      '<div>'+
        '<div style="height:40px;border-bottom:1.5px solid #0f1f38;margin-bottom:4px"></div>'+
        '<div style="font-size:9px;font-weight:700;color:#475569">Technician Signature</div>'+
        '<div style="font-size:10px;color:#0f1f38;font-weight:600;margin-top:2px">Pinellas Ice Co</div>'+
        '<div style="font-size:9px;color:#94a3b8">'+dateStr+'</div>'+
      '</div>'+
      '<div>'+
        '<div style="height:40px;border-bottom:1.5px solid #0f1f38;margin-bottom:4px"></div>'+
        '<div style="font-size:9px;font-weight:700;color:#475569">Authorized Representative</div>'+
        '<div style="font-size:10px;color:#94a3b8;margin-top:2px">Print name &amp; sign</div>'+
        '<div style="font-size:9px;color:#94a3b8">Date</div>'+
      '</div>'+
    '</div>'+

    // ── FOOTER ───────────────────────────────────────────────────────────
    '<div style="border-top:1px solid #e2e8f0;padding-top:8px;display:flex;justify-content:space-between;align-items:center">'+
      '<div style="font-size:8px;color:#94a3b8">'+
        'Pinellas Ice Co &bull; Commercial Ice Machine Sanitation &bull; Pinellas County, FL &bull; pinellasiceco.com'+
      '</div>'+
      '<div style="font-size:8px;color:#94a3b8">Retain for health inspection compliance records</div>'+
    '</div>';

  document.getElementById('svc-report-preview').innerHTML=
    '<div style="border:1px solid #e2e8f0;border-radius:10px;padding:20px;margin-top:8px;background:#fff;color:#1e293b;font-family:system-ui,sans-serif" id="report-content">'+
    reportHTML+
    '</div>'+
    '<div style="display:flex;gap:8px;margin-top:10px">'+
      '<button onclick="printReport()" style="flex:1;padding:10px;border:none;border-radius:8px;background:#0f1f38;color:#fff;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">&#x1F5A8; Print / Save PDF</button>'+
      '<button onclick="saveReportAndLog('+id+')" style="flex:1;padding:10px;border:1px solid #059669;border-radius:8px;background:#ecfdf5;color:#059669;font-size:12px;font-weight:700;cursor:pointer;font-family:inherit;touch-action:manipulation">&#x2713; Save &amp; Log Visit</button>'+
    '</div>';
}

function svcRow(label,val){
  return '<div style="display:flex;gap:6px;align-items:flex-start;padding:5px 8px;background:#f8fafc;border-radius:5px">'+
    '<span style="font-size:9px;font-weight:700;color:#059669;flex-shrink:0">'+val+'</span>'+
    '<span style="font-size:9px;color:#475569;line-height:1.4">'+label+'</span>'+
  '</div>';
}


function reportItem(label,val){
  return '<div style="padding:5px 7px;background:#f5f8fa;border-radius:5px">'
    +'<div style="font-size:8px;color:var(--sub);font-weight:600">'+label+'</div>'
    +'<div style="font-size:10px;font-weight:600;color:var(--navy)">'+val+'</div>'
    +'</div>';
}

function saveReportAndLog(id){
  // Save notes from report textarea
  const notesEl=document.getElementById('report-notes-'+id);
  const notes=notesEl?notesEl.value.trim():'';
  if(notes){
    if(!customers[id])customers[id]={};
    customers[id].report_notes=notes;
    custSave();
  }
  // Open service log modal with notes pre-filled
  openServiceLog(id);
  // After modal opens, prefill all fields from saved customer data
  setTimeout(()=>{
    const c2=customers[id]||{};
    const lastSvc2=c2.service_history&&c2.service_history.length
      ?c2.service_history[c2.service_history.length-1]:null;
    const set=(elId,val)=>{const el=document.getElementById(elId);if(el&&val)el.value=val;};
    set('svc-machine-brand', c2.machine_brand||'');
    set('svc-machine-model', c2.machine_model||'');
    set('svc-machine-serial', c2.machine_serial||'');
    set('svc-units', c2.machines||1);
    set('svc-filter-type', c2.filter_type||'Everpure i2000(2) Insurice EV9612-22');
    // Don't pre-fill ATP - must be fresh each visit
  },120);
}

function printReport(){
  const content=document.getElementById('report-content');
  if(!content)return;
  const win=window.open('','_blank');
  win.document.write('<html><head><title>Service Report - Pinellas Ice Co</title>'
    +'<style>body{font-family:system-ui,sans-serif;padding:20px;max-width:600px;margin:0 auto;color:#1e293b}'
    +'*{box-sizing:border-box}textarea{border:1px solid #cbd5e1;border-radius:4px;padding:6px;width:100%;font-family:inherit}'
    +'@media print{button{display:none}}</style></head>'
    +'<body>'+content.outerHTML+'<br><button onclick="window.print()">Print / Save PDF</button></body></html>');
  win.document.close();
  setTimeout(()=>win.print(),400);
}

// ── REFERRAL TRACKING ─────────────────────────────────────────────────────────
function renderReferrals(){
  const el=document.getElementById('svc-refs-list');
  if(!el)return;

  const recurring=P.filter(p=>p.status==='customer_recurring'||p.status==='customer_once'||p.status==='customer_intro');
  if(!recurring.length){
    el.innerHTML='<div class="tempty"><div class="ei">&#x1F91D;</div><div>No clients yet. Close your first deal to start tracking referrals.</div></div>';
    return;
  }

  // Count referrals per client
  const refCounts={};
  recurring.forEach(p=>{
    const c=customers[p.id]||{};
    if(c.referred_by&&c.referred_by!==p.id){
      refCounts[c.referred_by]=(refCounts[c.referred_by]||0)+1;
    }
  });

  el.innerHTML=recurring.map(p=>{
    const c=customers[p.id]||{};
    const refCount=refCounts[p.id]||0;

    return '<div class="svc-card">'
      +'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
        +'<div style="flex:1">'
          +'<div style="font-weight:700;font-size:12px;color:var(--navy)">'+p.name+'</div>'
          +'<div style="font-size:10px;color:var(--sub)">'+p.city+' &bull; Won: '+(c.won_date||'?')+'</div>'
        +'</div>'
        +(refCount?'<div style="background:#ecfdf5;border:1px solid #6ee7b7;border-radius:20px;padding:3px 10px;font-size:11px;font-weight:700;color:#059669">'+refCount+' referral'+(refCount>1?'s':'')+'</div>':'')
      +'</div>'
      // Referred by
      +'<div style="font-size:9px;color:var(--sub);margin-bottom:3px">Referred by:</div>'
      +'<div style="display:flex;gap:6px;align-items:center">'
        +'<select onchange="saveReferredBy('+p.id+',this.value)" style="flex:1;padding:5px;border:1px solid var(--brd);border-radius:6px;font-size:10px;font-family:inherit;background:var(--surf);color:var(--txt);outline:none">'
          +'<option value="">None / Self-generated</option>'
          +recurring.filter(r=>r.id!==p.id).map(r=>'<option value="'+r.id+'"'+(c.referred_by===r.id?' selected':'')+'>'+r.name.slice(0,30)+'</option>').join('')
        +'</select>'
      +'</div>'

      // Machine profile + contract
      +'<div style="display:flex;flex-direction:column;gap:5px;padding:8px;background:#f5f8fa;border-radius:7px;margin-top:6px">'
        +'<div style="font-size:8px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px">&#x2699; Machine Profile</div>'
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Brand</div>'
            +'<select onchange="saveMachineBrand('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +MACHINE_BRANDS.map(b=>'<option value="'+b+'"'+(b===(c.machine_brand||'')?'selected':'')+'>'+b+'</option>').join('')
              +'<option value="">Unknown</option>'
            +'</select>'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Filter</div>'
            +'<select onchange="saveFilterType('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +FILTER_TYPES.map(f=>'<option value="'+f+'"'+(f===(c.filter_type||'')?'selected':'')+'>'+f+'</option>').join('')
            +'</select>'
          +'</div>'
        +'</div>'
        // Contract dates
        +'<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-top:3px">'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Contract Start</div>'
            +'<input type="date" value="'+(c.contract_start||'')+'" onblur="saveContractStart('+p.id+',this.value)" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
          +'</div>'
          +'<div>'
            +'<div style="font-size:8px;color:var(--sub)">Term</div>'
            +'<select onchange="saveContractTerm('+p.id+',parseInt(this.value))" onclick="event.stopPropagation()"'
              +' style="width:100%;padding:4px;border:1px solid var(--brd);border-radius:5px;font-size:10px;font-family:inherit;background:#fff;color:var(--txt);outline:none">'
              +'<option value="6"'+(c.contract_term===6?' selected':'')+'>6 months</option>'
              +'<option value="12"'+(c.contract_term===12?' selected':'')+'>12 months</option>'
            +'</select>'
          +'</div>'
        +'</div>'
        +(c.contract_renewal?'<div style="font-size:9px;color:#d97706;font-weight:600">Renews: '+c.contract_renewal+'</div>':'')
      +'</div>'

      +'</div>';
  }).join('');
}

function saveReferredBy(bizId,refId){
  if(!customers[bizId])customers[bizId]={};
  customers[bizId].referred_by=refId?parseInt(refId):null;
  custSave();
  renderReferrals();
  toast('Referral source saved');
}

// ── SETTINGS ─────────────────────────────────────────────────────────────────
function loadSettings(){
  try{return JSON.parse(localStorage.getItem('pic_settings')||'{}')||{};}catch(e){return {};}
}
function saveSettings(){
  const portal=(document.getElementById('hs-portal')||{}).value||'';
  const homeZip=(document.getElementById('home-zip')||{}).value||'';
  try{localStorage.setItem('pic_settings',JSON.stringify({hubspot_portal:portal,home_zip:homeZip}));}catch(e){}
}
function initSettings(){
  const s=loadSettings();
  const DEFAULT_HOME_ZIP='34689'; // 1324 Live Oak Pkwy, Tarpon Springs, FL
  const hsel=document.getElementById('hs-portal');
  if(hsel&&s.hubspot_portal)hsel.value=s.hubspot_portal;
  const hzip=document.getElementById('home-zip');
  if(hzip){
    if(s.home_zip)hzip.value=s.home_zip;
    else if(!hzip.value)hzip.value=DEFAULT_HOME_ZIP;
  }
  // Auto-fill route start ZIP from saved home ZIP (fallback to default)
  const rzip=document.getElementById('rzip');
  if(rzip){
    const z=s.home_zip||DEFAULT_HOME_ZIP;
    if(!rzip.value)rzip.value=z;
    rzip.placeholder=z;
  }
}

// ── QUOTE BUILDER ─────────────────────────────────────────────────────────────
let vendors={};  // {bizId: vendorName}

function contactsLoad(){
  try{contacts=JSON.parse(localStorage.getItem('pic_contacts')||'{}')||{};}catch(e){contacts={};}
  try{vendors=JSON.parse(localStorage.getItem('pic_vendors')||'{}')||{};}catch(e){vendors={};}
}
function vendorsSave(){
  try{localStorage.setItem('pic_vendors',JSON.stringify(vendors));}catch(e){}
}
function saveVendor(){
  if(!cur)return;
  const v=(document.getElementById('mc-vendor')||{}).value.trim();
  if(!v){toast('Enter a vendor name');return;}
  vendors[cur.id]=v;
  vendorsSave();
  renderVendor(cur.id);
  toast('Vendor saved');
}
function renderVendor(id){
  const inp=document.getElementById('mc-vendor');
  const disp=document.getElementById('mc-vendor-display');
  const v=vendors[id]||'';
  if(inp)inp.value=v;
  if(disp)disp.textContent=v?'Currently: '+v:'Not recorded yet';
}
function contactsSave(){
  try{localStorage.setItem('pic_contacts',JSON.stringify(contacts));}catch(e){}
}

function renderContacts(id){
  const el=document.getElementById('mcontacts');
  if(!el)return;
  const list=contacts[id]||[];
  if(!list.length){el.innerHTML='<div style="font-size:10px;color:var(--sub);margin-bottom:6px">No contacts saved yet</div>';return;}
  const ROLE_LABELS={owner:'Owner',gm:'GM',manager:'Manager',chef:'Chef/Kitchen',staff:'Staff'};
  el.innerHTML=list.map((c,i)=>
    '<div style="display:flex;align-items:center;gap:8px;padding:6px 0;border-bottom:1px solid var(--brd2)">'
    +'<div style="flex:1">'
      +'<div style="font-size:11px;font-weight:600;color:var(--navy)">'+c.name+'</div>'
      +'<div style="font-size:9px;color:var(--sub)">'+( ROLE_LABELS[c.role]||c.role)+(c.phone?' &bull; '+c.phone:'')+'</div>'
    +'</div>'
    +(c.phone?'<a href="tel:'+c.phone.replace(/\s/g,'')+'" style="font-size:9px;padding:3px 7px;border:1px solid var(--blu);border-radius:5px;color:var(--blu);text-decoration:none" onclick="event.stopPropagation()">Call</a>':'')
    +'<button onclick="deleteContact('+id+','+i+')" style="font-size:9px;padding:3px 6px;border:1px solid var(--brd);border-radius:5px;background:transparent;color:var(--sub);cursor:pointer">✕</button>'
    +'</div>'
  ).join('');
}

function addContact(){
  if(!cur)return;
  const name=document.getElementById('mc-name').value.trim();
  const role=document.getElementById('mc-role').value;
  const phone=document.getElementById('mc-phone').value.trim();
  if(!name){toast('Enter a contact name');return;}
  if(!contacts[cur.id])contacts[cur.id]=[];
  contacts[cur.id].push({name,role,phone});
  contactsSave();
  document.getElementById('mc-name').value='';
  document.getElementById('mc-phone').value='';
  renderContacts(cur.id);
  toast('Contact saved');
}

function deleteContact(bizId,idx){
  if(!contacts[bizId])return;
  contacts[bizId].splice(idx,1);
  if(!contacts[bizId].length)delete contacts[bizId];
  contactsSave();
  renderContacts(bizId);
}

// ── INIT ─────────────────────────────────────────────────────────────────────
// INIT
function init(){
  lLoad();phLoad();custLoad();contactsLoad();initSettings();initGoals();setTimeout(function(){renderBriefing();},150);
  const si=document.getElementById('si');if(si)si.blur();
  // FAB hidden - using tab navigation instead

  // Global touch handler - catches taps on dynamically injected cards
  // iOS Safari drops onclick on innerHTML-injected elements; this doesn't
  var _tsx=0,_tsy=0,_tsTime=0;
  document.addEventListener('touchstart',function(e){
    _tsx=e.touches[0].clientX;
    _tsy=e.touches[0].clientY;
    _tsTime=Date.now();
  },{passive:true});

  document.addEventListener('touchend',function(e){
    // Ignore scrolls
    var dx=Math.abs(e.changedTouches[0].clientX-_tsx);
    var dy=Math.abs(e.changedTouches[0].clientY-_tsy);
    if(dx>12||dy>12)return;
    // Ignore rapid re-fires
    if(Date.now()-_tsTime>1000)return;

    var t=e.target;

    // Action button?
    var btn=t.closest('[data-action]');
    if(btn){
      var act=btn.dataset.action;
      var card=btn.closest('[data-id]');
      var id=parseInt(btn.dataset.id||(card&&card.dataset.id)||'0');
      if(!id)return;
      e.preventDefault();
      if(act==='showCard')      showCard(id);
      else if(act==='svclog') openServiceLog(id);
      else if(act==='snext')  setNextService(id);
      else if(act==='skip')     skip(id);
      else if(act==='route')    addToRoute(id);
      else if(act==='unskip')   unskip(id);
      else if(act==='start')    setRouteAnchor(id);
      return;
    }

    // Tap on card body?
    var card2=t.closest('[data-id]');
    if(card2&&!t.closest('a')&&!t.closest('input')&&!t.closest('textarea')&&!t.closest('select')){
      var id2=parseInt(card2.dataset.id||'0');
      if(id2){e.preventDefault();showCard(id2);}
    }
  },false);
}
if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',init);}else{setTimeout(init,50);}

// ── GLOBAL TOUCH DELEGATION (iPad Safari fix) ─────────────────────────────
// iPad Safari won't fire onclick on div elements set via innerHTML.
// One delegated touchend listener handles ALL card taps across every tab.
(function(){
  let _tx=0, _ty=0;
  document.addEventListener('touchstart', function(e){
    _tx=e.touches[0].clientX;
    _ty=e.touches[0].clientY;
  }, {passive:true});

  document.addEventListener('touchend', function(e){
    const dx=Math.abs(e.changedTouches[0].clientX-_tx);
    const dy=Math.abs(e.changedTouches[0].clientY-_ty);
    if(dx>10||dy>10) return; // was a scroll, not a tap

    const t=e.target;

    // ── Card buttons (check before card to avoid double-fire) ──────────────
    const logBtn=t.closest('.log-call-btn');
    if(logBtn){const id=parseInt(logBtn.dataset.id);if(id)openM(id);e.preventDefault();return;}

    const routeBtn=t.closest('.route-btn');
    if(routeBtn){const id=parseInt(routeBtn.dataset.id);if(id)addToRoute(id);e.preventDefault();return;}

    const skipBtn=t.closest('.skip-btn');
    if(skipBtn){const id=parseInt(skipBtn.dataset.id);if(id)skip(id);e.preventDefault();return;}

    const savePhBtn=t.closest('.save-phone-btn');
    if(savePhBtn){const id=parseInt(savePhBtn.dataset.id);if(id)saveFoundPhone(id);e.preventDefault();return;}

    // ── Card main body tap → open details ─────────────────────────────────
    const card=t.closest('.card');
    if(card){
      // Don't open if tapping a link or input
      if(t.closest('a,input,textarea,button,select'))return;
      const id=parseInt(card.dataset.id);
      if(id)openM(id);
      e.preventDefault();
      return;
    }

    // ── Queue action buttons ───────────────────────────────────────────────
    const qBtn=t.closest('.qbtn');
    if(qBtn){
      const fn=qBtn.dataset.action;
      if(fn==='next') queueNext();
      else if(fn) queueLog(fn);
      e.preventDefault();
      return;
    }

    // ── Queue Details button ───────────────────────────────────────────────
    const qDetails=t.closest('.q-details-btn');
    if(qDetails){
      const id=parseInt(qDetails.dataset.id);
      if(id)openM(id);
      e.preventDefault();
      return;
    }
  }, {passive:false});
})();

// ── OFFLINE DETECTION ────────────────────────────────────────────────────────
function updateOnlineStatus(){
  const banner=document.getElementById('offline-banner');
  if(!banner)return;
  if(!navigator.onLine){
    banner.classList.add('active');
    banner.style.display='flex';
    document.body.style.paddingTop='32px';
  } else {
    banner.classList.remove('active');
    banner.style.display='none';
    document.body.style.paddingTop='';
  }
}
window.addEventListener('online', updateOnlineStatus);
window.addEventListener('offline', updateOnlineStatus);
updateOnlineStatus();

// ── SERVICE WORKER ─────────────────────────────────────────────────────────
// SW disabled during development - re-enable when stable
if('serviceWorker' in navigator){
  // Unregister any existing service worker so browser fetches fresh code
  navigator.serviceWorker.getRegistrations().then(regs=>{
    regs.forEach(r=>r.unregister());
  });
}
</script>
</body>
</html>
"""

# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────
SW_JS = """const CACHE_NAME='pic-v5';
const ASSETS=['./','./ index.html'];
self.addEventListener('install',e=>{
  e.waitUntil(caches.open(CACHE_NAME).then(c=>c.addAll(ASSETS).catch(()=>{})));
  self.skipWaiting();
});
self.addEventListener('activate',e=>{
  e.waitUntil(caches.keys().then(ks=>Promise.all(ks.filter(k=>k!==CACHE_NAME).map(k=>caches.delete(k)))));
  self.clients.claim();
});
self.addEventListener('fetch',e=>{
  const url=new URL(e.request.url);
  if(url.origin===self.location.origin){
    e.respondWith(caches.match(e.request).then(cached=>{
      const fresh=fetch(e.request).then(res=>{
        if(res&&res.status===200){const copy=res.clone();caches.open(CACHE_NAME).then(c=>c.put(e.request,copy));}
        return res;
      }).catch(()=>cached);
      return cached||fresh;
    }));
  } else {
    e.respondWith(fetch(e.request).catch(()=>caches.match(e.request)));
  }
});
"""

def main():
    folder = Path(__file__).parent

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if Path(arg).is_dir():
            # Directory passed (e.g. "data/") -- find all CSV and XLSX files
            data_dir = Path(arg)
            csv_paths = (
                sorted(data_dir.glob('*.csv')) +
                sorted(data_dir.glob('*.xlsx')) +
                sorted(data_dir.glob('*.xls'))
            )
            # Exclude the license extract from inspection data
            csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                         and 'hrfood' not in p.name.lower()]
            if not csv_paths:
                print(f"\nNo inspection data files found in {data_dir}")
                sys.exit(1)
            print(f"Found {len(csv_paths)} file(s) in {data_dir}:")
            for p in csv_paths:
                print(f"  {p.name}")
            print()
        else:
            csv_paths = sys.argv[1:]
    else:
        # Auto-find in current folder
        csv_paths = (
            sorted(folder.glob('*.csv')) +
            sorted(folder.glob('*.xlsx'))
        )
        csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                     and 'hrfood' not in p.name.lower()]
        if not csv_paths:
            print("\nNo data files found. Run: python build.py data/")
            sys.exit(1)
        print(f"Auto-found {len(csv_paths)} file(s):")
        for p in csv_paths:
            print(f"  {p.name}")
        print()

    records = run(list(map(str, csv_paths)))
    print(f"\nGenerating HTML...")
    html = build_html(records)
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(html, encoding='utf-8')
    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"  Written: {OUTPUT_FILE.name} ({size_kb}KB)")
    # Write sw.js for PWA offline support
    sw_path = OUTPUT_FILE.parent / 'sw.js'
    sw_path.write_text(SW_JS, encoding='utf-8')
    print(f"  Written: sw.js")
    print(f"\n{'='*55}")
    print(f"  Done! Open prospecting_tool.html in Chrome.")
    print(f"  Your call log carries over automatically.")
    print(f"{'='*55}\n")

def main():
    folder = Path(__file__).parent

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if Path(arg).is_dir():
            # Directory passed (e.g. "data/") -- find all CSV and XLSX files
            data_dir = Path(arg)
            csv_paths = (
                sorted(data_dir.glob('*.csv')) +
                sorted(data_dir.glob('*.xlsx')) +
                sorted(data_dir.glob('*.xls'))
            )
            # Exclude the license extract from inspection data
            csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                         and 'hrfood' not in p.name.lower()]
            if not csv_paths:
                print(f"\nNo inspection data files found in {data_dir}")
                sys.exit(1)
            print(f"Found {len(csv_paths)} file(s) in {data_dir}:")
            for p in csv_paths:
                print(f"  {p.name}")
            print()
        else:
            csv_paths = sys.argv[1:]
    else:
        # Auto-find in current folder
        csv_paths = (
            sorted(folder.glob('*.csv')) +
            sorted(folder.glob('*.xlsx'))
        )
        csv_paths = [p for p in csv_paths if 'licenses' not in p.name.lower()
                     and 'hrfood' not in p.name.lower()]
        if not csv_paths:
            print("\nNo data files found. Run: python build.py data/")
            sys.exit(1)
        print(f"Auto-found {len(csv_paths)} file(s):")
        for p in csv_paths:
            print(f"  {p.name}")
        print()

    records = run(list(map(str, csv_paths)))
    print(f"\nGenerating HTML...")
    html = build_html(records)
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(html, encoding='utf-8')
    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"  Written: {OUTPUT_FILE.name} ({size_kb}KB)")
    # Write sw.js for PWA offline support
    sw_path = OUTPUT_FILE.parent / 'sw.js'
    sw_path.write_text(SW_JS, encoding='utf-8')
    print(f"  Written: sw.js")
    print(f"\n{'='*55}")
    print(f"  Done! Open prospecting_tool.html in Chrome.")
    print(f"  Your call log carries over automatically.")
    print(f"{'='*55}\n")

if __name__ == '__main__':
    main()